"""Atualiza faturamento médico a partir de planilha (paciente, associado, valor, conferência)."""
from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path

from django.db.models import Q

from faturamento_medico.models import FaturamentoMedico, ItemServico
from faturamento_medico.procedimento_utils import eh_procedimento_transvaginal
from faturamento_medico.services.transvaginal_lancamento import (
    linha_planilha_sem_guia,
    obter_faturamento_transvaginal,
    separar_item_transvaginal,
)

CODIGO_MATERIAIS = '88888'
CODIGO_MEDICAMENTO = '99999'


@dataclass
class LinhaPlanilha:
    data: date
    paciente: str
    nome_associado: str
    procedimento: str
    modalidade: str
    valor: Decimal
    guia: str = ''
    numero_guia_lancada: str = ''
    lote_externo: str = ''
    protocolo: str = ''
    senha: str = ''


def _normalizar_texto(valor: str) -> str:
    txt = (valor or '').strip().upper()
    txt = unicodedata.normalize('NFKD', txt)
    txt = ''.join(c for c in txt if not unicodedata.combining(c))
    txt = re.sub(r'\s+', ' ', txt)
    return txt


def _parse_data(raw: str) -> date:
    raw = (raw or '').strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f'Data inválida: {raw!r}')


def _parse_valor(raw: str) -> Decimal:
    s = (raw or '').strip().replace('R$', '').strip()
    if not s:
        raise ValueError('Valor vazio')
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    elif s.count('.') > 1:
        parts = s.split('.')
        s = ''.join(parts[:-1]) + '.' + parts[-1]
    return Decimal(s)


def _tipo_linha_especial(linha: LinhaPlanilha) -> str | None:
    proc = _normalizar_texto(linha.procedimento)
    if not proc:
        return None
    if 'MEDICAMENTO' in proc:
        return 'medicamento'
    if 'MATERIAIS' in proc or ('ENSUMO' in proc and 'AMBULATORIAL' in proc):
        return 'materiais'
    return None


def _codigo_servico_linha(linha: LinhaPlanilha) -> str | None:
    tipo = _tipo_linha_especial(linha)
    if tipo == 'materiais':
        return CODIGO_MATERIAIS
    if tipo == 'medicamento':
        return CODIGO_MEDICAMENTO
    return None


def _associado_final(paciente: str, associado: str) -> str:
    assoc = (associado or '').strip()
    if not assoc or _normalizar_texto(assoc) == 'TITULAR':
        return paciente.strip()
    return assoc


def _similaridade(a: str, b: str) -> float:
    na = _normalizar_texto(a)
    nb = _normalizar_texto(b)
    if not na or not nb:
        return 0.0
    if na in nb or nb in na:
        return 0.95
    return SequenceMatcher(None, na, nb).ratio()


def _detectar_delimitador(amostra: str) -> str:
    if '\t' in amostra:
        return '\t'
    if ';' in amostra and amostra.count(';') >= amostra.count(','):
        return ';'
    return ','


def _mapear_colunas(fieldnames: list[str] | None) -> dict[str, str]:
    if not fieldnames:
        raise ValueError('Arquivo sem cabeçalho.')
    mapa: dict[str, str] = {}
    for col in fieldnames:
        chave = _normalizar_texto(col).replace(' ', '_')
        mapa[chave] = col
    aliases = {
        'data': ('DATA',),
        'paciente': ('PACIENTE',),
        'nome_associado': ('NOME_ASSOCIADO', 'ASSOCIADO', 'NOME_DO_ASSOCIADO', 'TITULAR'),
        'procedimento': ('PROCEDIMENTO', 'EXAME', 'SERVICO'),
        'modalidade': ('MODALIDADE', 'MOD', 'MODALID'),
        'valor': ('VALOR', 'VALOR_TOTAL'),
        'guia': ('GUIA_AUT', 'GUIA', 'NUMERO_DA_GUIA', 'NUMERO GUIA', 'N_DA_GUIA'),
        'numero_guia_lancada': (
            'GUIA_LANCADA',
            'NUMERO_GUIA_LANCADA',
            'N_GUIA_LANCADA',
            'GUIA_PRESTADOR',
            'GUIA_DO_PRESTADOR',
        ),
        'lote_externo': ('LOTE',),
        'protocolo': ('PROTOCOLO',),
        'senha': ('SENHA', 'SENHA_AUTORIZACAO'),
    }
    obrigatorias = ('data', 'paciente', 'procedimento', 'modalidade', 'valor')
    resultado: dict[str, str] = {}
    for destino, opcoes in aliases.items():
        for op in opcoes:
            if op in mapa:
                resultado[destino] = mapa[op]
                break
        else:
            if destino in obrigatorias:
                raise ValueError(f'Coluna obrigatória ausente: {destino}')
    return resultado


def _campo_planilha_str(val) -> str:
    if val is None:
        return ''
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    s = str(val).strip()
    if s.lower() in ('none', 'nan'):
        return ''
    return s


def _parse_data_flex(raw) -> date:
    if isinstance(raw, date):
        return raw
    if isinstance(raw, datetime):
        return raw.date()
    return _parse_data(str(raw or ''))


def _parse_valor_flex(raw) -> Decimal:
    if raw is None:
        raise ValueError('Valor vazio')
    if isinstance(raw, (int, float, Decimal)):
        return Decimal(str(raw))
    return _parse_valor(str(raw or ''))


def _linha_from_row(row: dict, colunas: dict[str, str], idx: int) -> LinhaPlanilha | None:
    paciente = (row.get(colunas['paciente']) or '').strip()
    if not paciente:
        return None
    try:
        valor = _parse_valor_flex(row.get(colunas['valor']))
    except ValueError:
        return None

    def opt(key: str) -> str:
        col = colunas.get(key)
        if not col:
            return ''
        return _campo_planilha_str(row.get(col))

    return LinhaPlanilha(
        data=_parse_data_flex(row.get(colunas['data'])),
        paciente=paciente,
        nome_associado=str(row.get(colunas.get('nome_associado', ''), '') or '').strip(),
        procedimento=str(row.get(colunas['procedimento']) or '').strip(),
        modalidade=str(row.get(colunas['modalidade']) or '').strip().upper(),
        valor=valor,
        guia=opt('guia'),
        numero_guia_lancada=opt('numero_guia_lancada'),
        lote_externo=opt('lote_externo'),
        protocolo=opt('protocolo'),
        senha=opt('senha'),
    )


def carregar_planilha_xlsx(caminho: Path) -> list[LinhaPlanilha]:
    from openpyxl import load_workbook

    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError('Planilha vazia.')
    headers = [str(c or '').strip() for c in rows[0]]
    colunas = _mapear_colunas(headers)
    linhas: list[LinhaPlanilha] = []
    for idx, raw in enumerate(rows[1:], start=2):
        row = {headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers))}
        try:
            linha = _linha_from_row(row, colunas, idx)
            if linha:
                linhas.append(linha)
        except ValueError as exc:
            raise ValueError(f'Linha {idx}: {exc}') from exc
    return linhas


def carregar_planilha(caminho: Path) -> list[LinhaPlanilha]:
    if caminho.suffix.lower() in ('.xlsx', '.xlsm', '.xltx'):
        return carregar_planilha_xlsx(caminho)
    texto = caminho.read_text(encoding='utf-8-sig')
    delim = _detectar_delimitador(texto[:4096])
    reader = csv.DictReader(texto.splitlines(), delimiter=delim)
    colunas = _mapear_colunas(reader.fieldnames)
    linhas: list[LinhaPlanilha] = []
    for idx, row in enumerate(reader, start=2):
        try:
            linha = _linha_from_row(row, colunas, idx)
            if linha:
                linhas.append(linha)
        except ValueError as exc:
            raise ValueError(f'Linha {idx}: {exc}') from exc
    return linhas


def _filtro_convenio(convenio: str, *, prefixo_faturamento: bool = True) -> Q:
    nome = (convenio or '').strip().upper()
    campo = 'faturamento__convenio' if prefixo_faturamento else 'convenio'
    if not nome:
        return Q(**{f'{campo}__icontains': 'BOMBEIRO'})
    if nome == 'FUSEX':
        return Q(**{f'{campo}__iexact': 'FUSEX'})
    return Q(**{f'{campo}__icontains': nome})


def _normalizar_guia(guia: str) -> str:
    return re.sub(r'\D', '', (guia or '').strip())


def _modalidades_equivalentes(a: str, b: str) -> bool:
    na = _normalizar_texto(a)
    nb = _normalizar_texto(b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    par = {('MR', 'RM'), ('CR', 'RX'), ('TC', 'CT')}
    return (na, nb) in par or (nb, na) in par


def _modalidade_filter_linha(linha: LinhaPlanilha) -> Q:
    mods: set[str] = set()
    mod = (linha.modalidade or '').strip().upper()
    if mod in ('MR', 'RM'):
        mods.update({'MR', 'RM'})
    elif mod in ('TC', 'CT'):
        mods.update({'TC', 'CT'})
    elif mod:
        mods.add(mod)
    proc = _normalizar_texto(linha.procedimento)
    if proc.startswith('RX') or proc.startswith('RX-') or proc.startswith('RX '):
        mods.update({'CR', 'RX'})
    if proc.startswith('RM') or proc.startswith('RM ') or proc.startswith('RM-'):
        mods.update({'MR', 'RM'})
    if not mods:
        return Q()
    filtro = Q()
    for m in mods:
        filtro |= Q(modalidade__iexact=m)
    return filtro


def _paciente_compativel(linha: LinhaPlanilha, fat: FaturamentoMedico) -> bool:
    guia_csv = _normalizar_guia(linha.guia)
    guia_fat = _normalizar_guia(fat.guia or '')
    if guia_csv and guia_fat and (guia_csv == guia_fat or guia_csv in guia_fat or guia_fat in guia_csv):
        return True

    paciente_norm = _normalizar_texto(linha.paciente)
    fat_nome = _normalizar_texto(fat.nome or '')
    fat_assoc = _normalizar_texto(fat.nome_associado or '')
    if not paciente_norm:
        return False
    for candidato in (fat_nome, fat_assoc):
        if not candidato:
            continue
        if candidato == paciente_norm:
            return True
        if _similaridade(linha.paciente, candidato) >= 0.82:
            return True
    return False


def _score_item(linha: LinhaPlanilha, item: ItemServico) -> float:
    fat = item.faturamento
    score = 0.0
    paciente_norm = _normalizar_texto(linha.paciente)
    assoc_csv = _normalizar_texto(linha.nome_associado or '')
    fat_nome = _normalizar_texto(fat.nome or '')
    fat_assoc = _normalizar_texto(fat.nome_associado or '')

    if fat_nome == paciente_norm or fat_assoc == paciente_norm:
        score += 4.0
    elif paciente_norm in fat_nome or fat_nome in paciente_norm:
        score += 2.5
    elif paciente_norm in fat_assoc or fat_assoc in paciente_norm:
        score += 2.0

    if assoc_csv and assoc_csv != 'TITULAR':
        if fat_assoc == assoc_csv:
            score += 2.5
        elif _similaridade(assoc_csv, fat_assoc) >= 0.82:
            score += 1.5

    score += _similaridade(linha.procedimento, item.servico or '') * 3.5

    if linha.modalidade and _modalidades_equivalentes(item.modalidade or '', linha.modalidade):
        score += 1.0

    guia_csv = _normalizar_guia(linha.guia)
    guia_fat = _normalizar_guia(fat.guia or '')
    if guia_csv and guia_fat:
        if guia_csv == guia_fat or guia_csv in guia_fat or guia_fat in guia_csv:
            score += 5.0
        else:
            score -= 4.0

    valor_item = item.total if item.total is not None else item.valor
    if valor_item is not None:
        diff = abs(Decimal(valor_item) - linha.valor)
        if diff <= Decimal('0.01'):
            score += 4.0
        elif diff <= Decimal('50'):
            score += 1.0
        else:
            score -= 1.5

    return score


def _buscar_faturamento_linha(
    linha: LinhaPlanilha,
    *,
    empresa_id: int,
    convenio: str,
    tolerancia_mes: bool = False,
) -> FaturamentoMedico | None:
    qs = (
        FaturamentoMedico.objects
        .filter(empresa_id=empresa_id)
        .filter(_filtro_convenio(convenio, prefixo_faturamento=False))
    )
    if tolerancia_mes:
        qs = qs.filter(
            data__year=linha.data.year,
            data__month=linha.data.month,
        )
    else:
        qs = qs.filter(data=linha.data)
    guia_norm = _normalizar_guia(linha.guia)
    if guia_norm:
        qs_guia = qs.filter(guia__icontains=guia_norm)
        if qs_guia.exists():
            qs = qs_guia
    melhor: FaturamentoMedico | None = None
    melhor_score = 0.0
    for fat in qs:
        if not _paciente_compativel(linha, fat):
            continue
        score = 4.0
        if guia_norm and _normalizar_guia(fat.guia or '') == guia_norm:
            score += 3.0
        if score > melhor_score:
            melhor_score = score
            melhor = fat
    return melhor


def _item_eh_materiais_ou_medicamento(item: ItemServico) -> bool:
    cod = (item.codigo_servico or '').strip()
    if cod in (CODIGO_MATERIAIS, CODIGO_MEDICAMENTO):
        return True
    proc = _normalizar_texto(item.servico or '')
    return 'MEDICAMENTO' in proc or 'MATERIAIS' in proc or ('ENSUMO' in proc and 'AMBULATORIAL' in proc)


def _buscar_item_materiais_medicamento(
    linha: LinhaPlanilha,
    *,
    empresa_id: int,
    convenio: str,
    ids_usados: set[int],
) -> ItemServico | None:
    codigo = _codigo_servico_linha(linha)
    if not codigo:
        return None
    fat = _buscar_faturamento_linha(linha, empresa_id=empresa_id, convenio=convenio)
    if not fat:
        return None
    melhor: ItemServico | None = None
    melhor_score = 0.0
    for item in fat.itens_servico.all():
        if item.id in ids_usados:
            continue
        if not _item_eh_materiais_ou_medicamento(item):
            continue
        item_cod = (item.codigo_servico or '').strip()
        score = 3.0
        if item_cod == codigo:
            score += 2.0
        diff = abs(_valor_item(item) - linha.valor)
        if diff <= Decimal('0.01'):
            score += 3.0
        elif diff <= Decimal('100'):
            score += 1.0
        if score > melhor_score:
            melhor_score = score
            melhor = item
    if melhor is not None and melhor_score >= 5.0:
        return melhor
    return None


def _criar_item_materiais_medicamento(
    linha: LinhaPlanilha,
    fat: FaturamentoMedico,
    *,
    dry_run: bool,
) -> ItemServico | None:
    codigo = _codigo_servico_linha(linha)
    if not codigo:
        return None
    servico = (linha.procedimento or '').strip()[:200]
    if dry_run:
        item = ItemServico(
            faturamento=fat,
            codigo_servico=codigo,
            servico=servico,
            modalidade=(linha.modalidade or '')[:20],
            valor=linha.valor,
            total=linha.valor,
            conferido=True,
            status_conferencia='CONFERIDO',
        )
        item.id = 0
        return item
    item = ItemServico.objects.create(
        faturamento=fat,
        codigo_servico=codigo,
        servico=servico,
        modalidade=(linha.modalidade or '')[:20],
        valor=linha.valor,
        total=linha.valor,
        conferido=True,
        status_conferencia='CONFERIDO',
    )
    fat.atualizar_total()
    return item


def _criar_item_faltante(
    linha: LinhaPlanilha,
    fat: FaturamentoMedico,
    *,
    dry_run: bool,
) -> ItemServico | None:
    codigo = _codigo_servico_linha(linha)
    servico = (linha.procedimento or '').strip()[:200]
    if eh_procedimento_transvaginal(servico):
        fat = obter_faturamento_transvaginal(fat)
    if dry_run:
        item = ItemServico(
            faturamento=fat,
            codigo_servico=codigo or '',
            servico=servico,
            modalidade=(linha.modalidade or '')[:20],
            valor=linha.valor,
            total=linha.valor,
            conferido=True,
            status_conferencia='CONFERIDO',
        )
        item.id = 0
        return item
    item = ItemServico.objects.create(
        faturamento=fat,
        codigo_servico=codigo or '',
        servico=servico,
        modalidade=(linha.modalidade or '')[:20],
        valor=linha.valor,
        total=linha.valor,
        conferido=True,
        status_conferencia='CONFERIDO',
    )
    fat.atualizar_total()
    return item


def _candidatos_query(
    linha: LinhaPlanilha,
    *,
    empresa_id: int,
    convenio: str,
    apenas_pendentes: bool,
    ignorar_guia: bool = False,
    tolerancia_mes: bool = False,
):
    qs = (
        ItemServico.objects
        .select_related('faturamento')
        .filter(faturamento__empresa_id=empresa_id)
        .filter(_filtro_convenio(convenio))
    )
    if tolerancia_mes:
        qs = qs.filter(
            faturamento__data__year=linha.data.year,
            faturamento__data__month=linha.data.month,
        )
    else:
        qs = qs.filter(faturamento__data=linha.data)
    guia_norm = _normalizar_guia(linha.guia)
    if guia_norm and not ignorar_guia:
        qs_guia = qs.filter(faturamento__guia__icontains=guia_norm)
        if qs_guia.exists():
            qs = qs_guia
    if apenas_pendentes:
        qs = qs.exclude(status_conferencia='CONFERIDO').exclude(conferido=True)
    qs = qs.filter(_modalidade_filter_linha(linha))
    return qs


def _buscar_item(
    linha: LinhaPlanilha,
    *,
    empresa_id: int,
    convenio: str,
    ids_usados: set[int],
    apenas_pendentes: bool = True,
    ignorar_guia: bool = False,
    tolerancia_mes: bool = False,
) -> ItemServico | None:
    melhor: ItemServico | None = None
    melhor_score = 0.0

    for item in _candidatos_query(
        linha,
        empresa_id=empresa_id,
        convenio=convenio,
        apenas_pendentes=apenas_pendentes,
        ignorar_guia=ignorar_guia,
        tolerancia_mes=tolerancia_mes,
    ):
        if item.id in ids_usados:
            continue
        if not _paciente_compativel(linha, item.faturamento):
            continue
        score = _score_item(linha, item)
        if score > melhor_score:
            melhor_score = score
            melhor = item

    if melhor is None or melhor_score < 5.0:
        return None
    return melhor


def _buscar_item_com_fallbacks(
    linha: LinhaPlanilha,
    *,
    empresa_id: int,
    convenio: str,
    ids_usados: set[int],
    apenas_pendentes: bool = True,
) -> ItemServico | None:
    for ignorar_guia in (False, True) if _normalizar_guia(linha.guia) else (False,):
        item = _buscar_item(
            linha,
            empresa_id=empresa_id,
            convenio=convenio,
            ids_usados=ids_usados,
            apenas_pendentes=apenas_pendentes,
            ignorar_guia=ignorar_guia,
        )
        if item is not None:
            return item
        item = _buscar_item(
            linha,
            empresa_id=empresa_id,
            convenio=convenio,
            ids_usados=ids_usados,
            apenas_pendentes=apenas_pendentes,
            ignorar_guia=ignorar_guia,
            tolerancia_mes=True,
        )
        if item is not None:
            return item
    return None


def _ja_conferido_no_banco(linha: LinhaPlanilha, *, empresa_id: int, convenio: str) -> bool:
    melhor_score = 0.0
    for item in _candidatos_query(linha, empresa_id=empresa_id, convenio=convenio, apenas_pendentes=False):
        if not (item.conferido or item.status_conferencia == 'CONFERIDO'):
            continue
        if not _paciente_compativel(linha, item.faturamento):
            continue
        score = _score_item(linha, item)
        if score > melhor_score:
            melhor_score = score
    return melhor_score >= 5.0


def _valor_item(item: ItemServico) -> Decimal:
    v = item.total if item.total is not None else item.valor
    return Decimal(v or 0)


def _aplicar_documentacao_faturamento(fat: FaturamentoMedico, linha: LinhaPlanilha) -> list[str]:
    """Atualiza guia, guia lançada, lote e protocolo no faturamento."""
    if eh_procedimento_transvaginal(linha.procedimento):
        linha = linha_planilha_sem_guia(linha)
    update_fields: list[str] = []
    if linha.guia and (fat.guia or '').strip() != linha.guia:
        fat.guia = linha.guia[:50]
        update_fields.append('guia')
    if linha.numero_guia_lancada and (fat.numero_guia_lancada or '').strip() != linha.numero_guia_lancada:
        fat.numero_guia_lancada = linha.numero_guia_lancada[:50]
        update_fields.append('numero_guia_lancada')
    lote = linha.lote_externo
    if lote and lote not in ('0',) and (fat.lote or '').strip() != lote:
        fat.lote = lote[:50]
        update_fields.append('lote')
    if linha.protocolo and (fat.guia_lancada or '').strip() != linha.protocolo:
        fat.guia_lancada = linha.protocolo[:50]
        update_fields.append('guia_lancada')
    if linha.senha and (fat.senha or '').strip() != linha.senha:
        fat.senha = linha.senha[:50]
        update_fields.append('senha')
    if update_fields:
        fat.save(update_fields=update_fields)
    return update_fields


def _aplicar_linha_item(
    linha: LinhaPlanilha,
    item: ItemServico,
    *,
    dry_run: bool,
    stats: dict,
    ids_usados: set[int],
) -> None:
    fat: FaturamentoMedico = item.faturamento
    if eh_procedimento_transvaginal(linha.procedimento):
        linha = linha_planilha_sem_guia(linha)
        if not dry_run:
            item = separar_item_transvaginal(item)
            fat = item.faturamento
    assoc = _associado_final(linha.paciente, linha.nome_associado)
    ja_conferido = item.status_conferencia == 'CONFERIDO' or item.conferido
    valor_atual = _valor_item(item)
    valor_ok = abs(valor_atual - linha.valor) <= Decimal('0.01')
    precisa_valor = not valor_ok or not ja_conferido

    if dry_run:
        acao = 'CORRIGIR' if ja_conferido else 'CONFERIR'
        if not precisa_valor:
            acao = 'DOC'
        stats['atualizados'] += 1
        doc = []
        if linha.guia:
            doc.append(f'guia={linha.guia}')
        if linha.numero_guia_lancada:
            doc.append(f'guia_lanc={linha.numero_guia_lancada}')
        if linha.lote_externo and linha.lote_externo != '0':
            doc.append(f'lote={linha.lote_externo}')
        if linha.protocolo:
            doc.append(f'protocolo={linha.protocolo}')
        if linha.senha:
            doc.append(f'senha={linha.senha}')
        stats['detalhes'].append(
            f"DRY-RUN {acao} item #{item.id} fat #{fat.id}: "
            f"{linha.paciente} | {linha.modalidade} | valor {valor_atual} -> {linha.valor}"
            + (f" | {', '.join(doc)}" if doc else '')
        )
        ids_usados.add(item.id)
        return

    try:
        fat.nome = linha.paciente[:200]
        fat.nome_associado = assoc[:200]
        update_fat = ['nome', 'nome_associado']
        fat.save(update_fields=update_fat)

        doc_fields = _aplicar_documentacao_faturamento(fat, linha)
        alterou_codigo = False

        if precisa_valor:
            item.valor = linha.valor
            item.total = linha.valor
            if linha.procedimento:
                item.servico = linha.procedimento[:200]
            codigo = _codigo_servico_linha(linha)
            if codigo:
                item.codigo_servico = codigo
            item.conferido = True
            item.status_conferencia = 'CONFERIDO'
            item.save()
            fat.atualizar_total()
        elif _codigo_servico_linha(linha):
            codigo = _codigo_servico_linha(linha)
            update_item: list[str] = []
            if codigo and (item.codigo_servico or '').strip() != codigo:
                item.codigo_servico = codigo
                update_item.append('codigo_servico')
            if linha.procedimento and (item.servico or '').strip() != linha.procedimento.strip():
                item.servico = linha.procedimento[:200]
                update_item.append('servico')
            if update_item:
                item.save(update_fields=update_item)
                alterou_codigo = True

        if precisa_valor or doc_fields or alterou_codigo:
            stats['atualizados'] += 1
            ids_usados.add(item.id)
            acao = 'CORRIGIDO' if ja_conferido else 'OK'
            if not precisa_valor and doc_fields:
                acao = 'DOC'
            stats['detalhes'].append(
                f"{acao} item #{item.id} fat #{fat.id}: {linha.paciente} | {linha.modalidade} | R$ {linha.valor}"
            )
        else:
            stats['ja_conferidos_banco'] += 1
            ids_usados.add(item.id)
    except Exception as exc:
        stats['erros'] += 1
        stats['detalhes'].append(f"ERRO {linha.paciente}: {exc}")


def aplicar_atualizacoes(
    linhas: list[LinhaPlanilha],
    *,
    empresa_id: int,
    convenio: str = 'CORPO DE BOMBEIRO',
    dry_run: bool = False,
) -> dict:
    stats = {
        'linhas': len(linhas),
        'atualizados': 0,
        'ignorados_conferidos': 0,
        'ja_conferidos_banco': 0,
        'nao_encontrados': 0,
        'erros': 0,
        'detalhes': [],
    }
    ids_usados: set[int] = set()

    for linha in linhas:
        if _tipo_linha_especial(linha):
            item = _buscar_item_materiais_medicamento(
                linha,
                empresa_id=empresa_id,
                convenio=convenio,
                ids_usados=ids_usados,
            )
            if item is None:
                fat = _buscar_faturamento_linha(linha, empresa_id=empresa_id, convenio=convenio)
                if fat is None:
                    stats['nao_encontrados'] += 1
                    stats['detalhes'].append(
                        f"NÃO ENCONTRADO: {linha.data:%d/%m/%Y} | {linha.paciente} | "
                        f"{linha.procedimento[:40]} | R$ {linha.valor}"
                    )
                    continue
                item = _criar_item_materiais_medicamento(linha, fat, dry_run=dry_run)
                if item is None:
                    stats['nao_encontrados'] += 1
                    continue
                if dry_run:
                    stats['atualizados'] += 1
                    stats['detalhes'].append(
                        f"DRY-RUN CRIAR item materiais/med fat #{fat.id}: "
                        f"{linha.paciente} | cod {_codigo_servico_linha(linha)} | R$ {linha.valor}"
                    )
                    continue
                doc_fields = _aplicar_documentacao_faturamento(fat, linha)
                stats['atualizados'] += 1
                stats['detalhes'].append(
                    f"CRIADO item #{item.id} fat #{fat.id}: {linha.paciente} | "
                    f"cod {_codigo_servico_linha(linha)} | R$ {linha.valor}"
                )
                if doc_fields:
                    stats['detalhes'][-1] += f" | doc: {', '.join(doc_fields)}"
                if item.id:
                    ids_usados.add(item.id)
                continue

            _aplicar_linha_item(linha, item, dry_run=dry_run, stats=stats, ids_usados=ids_usados)
            continue

        item = _buscar_item_com_fallbacks(
            linha,
            empresa_id=empresa_id,
            convenio=convenio,
            ids_usados=ids_usados,
        )
        if item is None:
            item = _buscar_item_com_fallbacks(
                linha,
                empresa_id=empresa_id,
                convenio=convenio,
                ids_usados=ids_usados,
                apenas_pendentes=False,
            )
        if item is None:
            fat = _buscar_faturamento_linha(
                linha,
                empresa_id=empresa_id,
                convenio=convenio,
            )
            if fat is None:
                fat = _buscar_faturamento_linha(
                    linha,
                    empresa_id=empresa_id,
                    convenio=convenio,
                    tolerancia_mes=True,
                )
            if fat is not None:
                item = _criar_item_faltante(linha, fat, dry_run=dry_run)
                if item is not None:
                    if dry_run:
                        stats['atualizados'] += 1
                        stats['detalhes'].append(
                            f"DRY-RUN CRIAR item fat #{fat.id}: {linha.paciente} | "
                            f"{linha.modalidade} | R$ {linha.valor}"
                        )
                        continue
                    doc_fields = _aplicar_documentacao_faturamento(fat, linha)
                    stats['atualizados'] += 1
                    stats['detalhes'].append(
                        f"CRIADO item #{item.id} fat #{fat.id}: {linha.paciente} | "
                        f"{linha.modalidade} | R$ {linha.valor}"
                    )
                    if doc_fields:
                        stats['detalhes'][-1] += f" | doc: {', '.join(doc_fields)}"
                    if item.id:
                        ids_usados.add(item.id)
                    continue
            stats['nao_encontrados'] += 1
            stats['detalhes'].append(
                f"NÃO ENCONTRADO: {linha.data:%d/%m/%Y} | {linha.paciente} | {linha.modalidade} | R$ {linha.valor}"
            )
            continue

        if item.status_conferencia == 'CONFERIDO' or item.conferido:
            if item.id in ids_usados:
                stats['ignorados_conferidos'] += 1
                continue

        _aplicar_linha_item(linha, item, dry_run=dry_run, stats=stats, ids_usados=ids_usados)

    return stats
