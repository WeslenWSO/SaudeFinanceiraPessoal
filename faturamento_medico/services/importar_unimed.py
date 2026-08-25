"""Importação de relatório UNIMED (.txt ou .pdf — Produção)."""

from __future__ import annotations

import io
import logging
import os
import re
import unicodedata
from datetime import datetime
from typing import Any

import pdfplumber
from django.utils import timezone

from faturamento_medico.models import FaturamentoMedico, ItemServico
from servicos_medicos.models import ServicosMedicos

logger = logging.getLogger(__name__)

# Cabeçalho do modelo Excel UNIMED (mesma ordem lógica do .txt Produção)
UNIMED_XLSX_HEADERS = [
    'Lote',
    'Guia',
    'Cod. Usuario',
    'Nome Usuario',
    'Plano',
    'Cod.Servico',
    'Desc.Servico',
    'Tp. Grau',
    'Data',
    'Qtde',
    'Participacao %',
    'Valor Unit',
    'Valor Total',
    'Cod.Rel',
    'Observacao',
]

# Planilha UNIMED Produção: linhas 1–4 são título/filtros; cabeçalho na linha 5
UNIMED_XLSX_LINHA_CABECALHO = 5

try:
    import pypdfium2 as pdfium
except ImportError:
    pdfium = None

try:
    import pytesseract
    from PIL import Image

    OCR_AVAILABLE = True
except ImportError:
    pytesseract = None
    Image = None
    OCR_AVAILABLE = False

_OCR_ENGINE_OK: bool | None = None


def _fold(s: str) -> str:
    if not s:
        return ''
    d = unicodedata.normalize('NFKD', str(s))
    d = ''.join(c for c in d if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', d).strip().lower()


def _money_br(raw: str) -> float:
    s = (raw or '').strip()
    if not s:
        return 0.0
    s = s.replace('R$', '').replace(' ', '')
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_data(raw: str):
    s = (raw or '').strip()
    if not s:
        return timezone.now().date()
    for fmt in ('%d/%m/%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return timezone.now().date()


def _parse_int(raw: str) -> int:
    s = (raw or '').strip().replace(',', '.')
    if not s:
        return 1
    try:
        return int(float(s))
    except ValueError:
        return 1


def _linha_ignorada(cells: list[str]) -> bool:
    blob = _fold(' '.join(cells))
    if not blob:
        return True
    if 'tipo de guia' in blob:
        return True
    if 'vl. total pago' in blob or 'valor total pago' in blob:
        return True
    if blob in ('producao', 'produção'):
        return True
    if blob.startswith('lote guia'):
        return True
    return False


def _mapear_colunas_pdf(headers: list[str]) -> dict[str, int | None]:
    folded = [_fold(h) for h in headers]
    col_map: dict[str, int | None] = {
        'lote': None,
        'guia': None,
        'cod_usuario': None,
        'nome_usuario': None,
        'cod_servico': None,
        'desc_servico': None,
        'guia_prest': None,
        'data': None,
        'qtde': None,
        'via': None,
        'valor_unit': None,
        'valor_total': None,
    }

    for i, h in enumerate(folded):
        if not h:
            continue
        if h == 'lote' or h.startswith('lote '):
            col_map['lote'] = i
        elif h == 'guia':
            col_map['guia'] = i
        elif 'guia prest' in h:
            col_map['guia_prest'] = i
        elif 'cod' in h and 'usuario' in h:
            col_map['cod_usuario'] = i
        elif 'nome' in h and 'usuario' in h:
            col_map['nome_usuario'] = i
        elif 'cod' in h and 'serv' in h:
            col_map['cod_servico'] = i
        elif 'desc' in h and 'serv' in h:
            col_map['desc_servico'] = i
        elif h == 'data':
            col_map['data'] = i
        elif h.startswith('qtde') or h == 'qt':
            col_map['qtde'] = i
        elif h == 'via':
            col_map['via'] = i
        elif 'valor unit' in h:
            col_map['valor_unit'] = i
        elif 'valor (r$)' in h or h == 'valor r$' or (h.startswith('valor') and 'r$' in h):
            col_map['valor_total'] = i

    return col_map


def _mapear_colunas_unimed_xlsx(headers: list[str]) -> dict[str, int | None]:
    """Mapeia cabeçalho da planilha UNIMED (export .txt / Excel Produção)."""
    folded = [_fold(h) for h in headers]
    col_map: dict[str, int | None] = {
        'lote': None,
        'guia': None,
        'cod_usuario': None,
        'nome_usuario': None,
        'cod_servico': None,
        'desc_servico': None,
        'data': None,
        'qtde': None,
        'valor_unit': None,
        'valor_total': None,
        'cod_rel': None,
        'observacao': None,
    }

    for i, h in enumerate(folded):
        if not h:
            continue
        if h == 'lote' or h.startswith('lote '):
            col_map['lote'] = i
        elif h == 'guia':
            col_map['guia'] = i
        elif 'cod' in h and 'usuario' in h:
            col_map['cod_usuario'] = i
        elif 'nome' in h and 'usuario' in h:
            col_map['nome_usuario'] = i
        elif ('cod' in h or 'codigo' in h) and 'serv' in h and 'usuario' not in h:
            col_map['cod_servico'] = i
        elif 'desc' in h and 'serv' in h:
            col_map['desc_servico'] = i
        elif h == 'data':
            col_map['data'] = i
        elif h.startswith('qtde') or h == 'qt' or h.startswith('qtde/via'):
            col_map['qtde'] = i
        elif 'valor unit' in h:
            col_map['valor_unit'] = i
        elif 'valor total' in h or 'valor (r$)' in h or h == 'valor r$':
            col_map['valor_total'] = i
        elif 'cod.rel' in h or h == 'cod rel' or h.startswith('cod rel'):
            col_map['cod_rel'] = i
        elif 'observ' in h:
            col_map['observacao'] = i
        # Ignorados: Plano, Tp. Grau, Participação %, Valor Ref.

    return col_map


def _valor_celula_xlsx(raw) -> str:
    if raw is None:
        return ''
    if hasattr(raw, 'strftime'):
        try:
            return raw.strftime('%d/%m/%Y')
        except Exception:
            pass
    return str(raw).strip()


def parse_unimed_xlsx(xlsx_bytes: bytes) -> tuple[dict[str, dict], set[tuple[str, str]]]:
    """Planilha Excel (.xlsx) — cabeçalho na linha 5 (relatório UNIMED Produção)."""
    from io import BytesIO

    from openpyxl import load_workbook

    grupos: dict[str, dict] = {}
    servicos_unicos: set[tuple[str, str]] = set()

    wb = load_workbook(filename=BytesIO(xlsx_bytes), data_only=True, read_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        for _ in range(UNIMED_XLSX_LINHA_CABECALHO - 1):
            next(rows_iter, None)

        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError(
                f'Planilha UNIMED inválida: esperado cabeçalho na linha {UNIMED_XLSX_LINHA_CABECALHO}.'
            )

        headers = [_valor_celula_xlsx(c) for c in header_row]
        col_map = _mapear_colunas_unimed_xlsx(headers)
        if col_map.get('lote') is None or col_map.get('guia') is None or col_map.get('cod_servico') is None:
            cols = ', '.join(h for h in headers if h) or '(vazio)'
            raise ValueError(
                f'Cabeçalho na linha {UNIMED_XLSX_LINHA_CABECALHO} deve conter Lote, Guia e Cod.Serviço. '
                f'Colunas encontradas: {cols}'
            )

        for row in rows_iter:
            cells = [_valor_celula_xlsx(c) for c in row]
            if _linha_ignorada(cells):
                continue

            lote = _cel(cells, col_map.get('lote'))
            guia = _cel(cells, col_map.get('guia'))
            cod_servico = _cel(cells, col_map.get('cod_servico'))
            if not lote or not guia or not cod_servico:
                continue

            _adicionar_linha_grupo(
                grupos,
                servicos_unicos,
                lote=lote,
                guia=guia,
                cod_usuario=_cel(cells, col_map.get('cod_usuario')),
                nome_usuario=_cel(cells, col_map.get('nome_usuario')),
                cod_servico=cod_servico,
                desc_servico=_cel(cells, col_map.get('desc_servico')),
                data=_parse_data(_cel(cells, col_map.get('data'))),
                qtde=_parse_int(_cel(cells, col_map.get('qtde'))),
                valor_unit=_money_br(_cel(cells, col_map.get('valor_unit'))),
                valor_total=_money_br(_cel(cells, col_map.get('valor_total'))),
                cod_rel=_cel(cells, col_map.get('cod_rel')),
                observacao=_cel(cells, col_map.get('observacao')),
            )
    finally:
        wb.close()

    return grupos, servicos_unicos


def _cel(cells: list[str], idx: int | None) -> str:
    if idx is None or idx >= len(cells):
        return ''
    return str(cells[idx] or '').strip()


def _adicionar_linha_grupo(
    grupos: dict[str, dict[str, Any]],
    servicos_unicos: set[tuple[str, str]],
    *,
    lote: str,
    guia: str,
    cod_usuario: str,
    nome_usuario: str,
    cod_servico: str,
    desc_servico: str,
    data,
    qtde: int,
    valor_unit: float,
    valor_total: float,
    cod_rel: str = '',
    observacao: str = '',
    porte: str = '',
    percentual: float = 0,
) -> None:
    if not lote or not guia or not cod_servico:
        return

    chave = f'{lote}_{guia}'
    if chave not in grupos:
        grupos[chave] = {
            'lote': lote,
            'guia': guia,
            'carteirinha': cod_usuario,
            'nome': nome_usuario,
            'data': data,
            'cod_rel': cod_rel,
            'servicos': [],
        }

    grupos[chave]['servicos'].append({
        'codigo': cod_servico,
        'descricao': desc_servico,
        'porte': porte,
        'qt': qtde,
        'percentual': percentual,
        'valor': valor_unit,
        'total': valor_total,
        'observacao': observacao,
    })
    servicos_unicos.add((cod_servico, desc_servico))


def _configurar_tesseract() -> bool:
    global _OCR_ENGINE_OK
    if _OCR_ENGINE_OK is not None:
        return _OCR_ENGINE_OK
    if not OCR_AVAILABLE or pytesseract is None:
        _OCR_ENGINE_OK = False
        return False

    candidatos = []
    try:
        from django.conf import settings

        cmd_env = getattr(settings, 'TESSERACT_CMD', None) or os.environ.get('TESSERACT_CMD', '')
        if cmd_env:
            candidatos.append(cmd_env.strip())
    except Exception:
        cmd_env = os.environ.get('TESSERACT_CMD', '')
        if cmd_env:
            candidatos.append(cmd_env.strip())

    candidatos.extend([
        r'C:\Program Files\Tesseract-OCR\tesseract.exe',
        r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
    ])
    for cmd in candidatos:
        if cmd and os.path.isfile(cmd):
            pytesseract.pytesseract.tesseract_cmd = cmd
            break

    try:
        pytesseract.get_tesseract_version()
        _OCR_ENGINE_OK = True
    except Exception as exc:
        logger.warning('Tesseract indisponível: %s', exc)
        _OCR_ENGINE_OK = False
    return _OCR_ENGINE_OK


def _ocr_habilitado() -> bool:
    """OCR Tesseract — no Render usa build.sh (tesseract-ocr); desligar com UNIMED_OCR_RENDER=false."""
    if os.environ.get('RENDER', '').strip().lower() in ('true', '1', 'yes'):
        flag = os.environ.get('UNIMED_OCR_RENDER', 'true').strip().lower()
        if flag in ('false', '0', 'no'):
            return False
    return OCR_AVAILABLE and pdfium is not None and _configurar_tesseract()


def _gemini_habilitado() -> bool:
    """Gemini síncrono no Render derruba o worker (502/500); só se UNIMED_GEMINI_RENDER=true."""
    if os.environ.get('RENDER', '').strip().lower() in ('true', '1', 'yes'):
        return os.environ.get('UNIMED_GEMINI_RENDER', '').strip().lower() in ('true', '1', 'yes')
    return True


def _ocr_pdf_para_texto(pdf_bytes: bytes) -> str:
    """OCR página a página (PDF imagem / scan)."""
    if not OCR_AVAILABLE or pdfium is None or not _configurar_tesseract():
        return ''
    on_render = os.environ.get('RENDER', '').strip().lower() in ('true', '1', 'yes')
    try:
        max_pages = int(os.environ.get('UNIMED_OCR_MAX_PAGES', '0') or '0')
    except ValueError:
        max_pages = 0
    if on_render and max_pages <= 0:
        max_pages = 12
    scale = 200 / 72 if on_render else 300 / 72
    try:
        with pdfium.PdfDocument(pdf_bytes) as pdf:
            partes: list[str] = []
            total = len(pdf)
            limite = min(total, max_pages) if max_pages > 0 else total
            for idx in range(limite):
                page = pdf[idx]
                bitmap = page.render(scale=scale)
                pil_image = bitmap.to_pil()
                if not isinstance(pil_image, Image.Image):
                    pil_image = Image.frombytes(pil_image.mode, pil_image.size, pil_image.tobytes())
                try:
                    partes.append(
                        pytesseract.image_to_string(pil_image, lang='por+eng', config='--psm 6')
                    )
                except Exception:
                    partes.append(
                        pytesseract.image_to_string(pil_image, lang='por', config='--psm 6')
                    )
            if max_pages > 0 and total > limite:
                partes.append(f'[OCR limitado a {limite} de {total} paginas no servidor]')
            return '\n'.join(partes)
    except Exception as exc:
        logger.warning('Falha OCR UNIMED: %s', exc)
        return ''


def _grupos_de_texto(texto: str) -> tuple[dict[str, dict], set[tuple[str, str]]]:
    grupos: dict[str, dict] = {}
    servicos_unicos: set[tuple[str, str]] = set()
    for linha in (texto or '').splitlines():
        parsed = _parse_linha_pdf_texto(linha)
        if not parsed:
            continue
        obs = f"Guia prest.: {parsed['guia_prest']}" if parsed.get('guia_prest') else ''
        _adicionar_linha_grupo(
            grupos,
            servicos_unicos,
            lote=parsed['lote'],
            guia=parsed['guia'],
            cod_usuario=parsed['cod_usuario'],
            nome_usuario=parsed['nome_usuario'],
            cod_servico=parsed['cod_servico'],
            desc_servico=parsed['desc_servico'],
            data=parsed['data'],
            qtde=parsed['qtde'],
            valor_unit=parsed['valor_unit'],
            valor_total=parsed['valor_total'],
            observacao=obs,
        )
    return grupos, servicos_unicos


def _grupos_de_linhas_gemini(linhas: list[dict[str, Any]]) -> tuple[dict[str, dict], set[tuple[str, str]]]:
    grupos: dict[str, dict] = {}
    servicos_unicos: set[tuple[str, str]] = set()
    for raw in linhas:
        lote = str(raw.get('lote') or '').strip()
        guia = str(raw.get('guia') or '').strip()
        cod_servico = str(raw.get('cod_servico') or '').strip()
        if not lote or not guia or not cod_servico:
            continue
        guia_prest = str(raw.get('guia_prest') or '').strip()
        obs = f'Guia prest.: {guia_prest}' if guia_prest else ''
        _adicionar_linha_grupo(
            grupos,
            servicos_unicos,
            lote=lote,
            guia=guia,
            cod_usuario=str(raw.get('cod_usuario') or '').strip(),
            nome_usuario=str(raw.get('nome_usuario') or '').strip(),
            cod_servico=cod_servico,
            desc_servico=str(raw.get('desc_servico') or '').strip(),
            data=_parse_data(str(raw.get('data') or '')),
            qtde=_parse_int(str(raw.get('qtde') or '1')),
            valor_unit=_money_br(str(raw.get('valor_unit') or '')),
            valor_total=_money_br(str(raw.get('valor_total') or '')),
            observacao=obs,
        )
    return grupos, servicos_unicos


def parse_unimed_txt(content: str) -> tuple[dict[str, dict], set[tuple[str, str]]]:
    """Arquivo .txt separado por ponto e vírgula (cabeçalho na 1ª linha)."""
    grupos: dict[str, dict] = {}
    servicos_unicos: set[tuple[str, str]] = set()
    lines = content.splitlines()

    for line in lines[1:]:
        if not line.strip():
            continue
        parts = line.split(';')
        if len(parts) < 13:
            continue

        lote = parts[0].strip()
        guia = parts[1].strip()
        cod_usuario = parts[2].strip()
        nome_usuario = parts[3].strip()
        cod_servico = parts[5].strip()
        desc_servico = parts[6].strip()
        data = _parse_data(parts[8].strip())
        qtde_via = parts[9].strip()
        valor_unit = _money_br(parts[11].strip())
        valor_total = _money_br(parts[12].strip())
        cod_rel = parts[13].strip() if len(parts) > 13 else ''
        observacao = parts[14].strip() if len(parts) > 14 else ''
        # Ignorados: parts[4] Plano, parts[7] Tp. Grau, parts[10] Participação %

        _adicionar_linha_grupo(
            grupos,
            servicos_unicos,
            lote=lote,
            guia=guia,
            cod_usuario=cod_usuario,
            nome_usuario=nome_usuario,
            cod_servico=cod_servico,
            desc_servico=desc_servico,
            data=data,
            qtde=_parse_int(qtde_via),
            valor_unit=valor_unit,
            valor_total=valor_total,
            cod_rel=cod_rel,
            observacao=observacao,
        )

    return grupos, servicos_unicos


def _parse_unimed_pdf_table(table: list[list[Any]], col_map: dict[str, int | None]) -> tuple[dict, set]:
    grupos: dict[str, dict] = {}
    servicos_unicos: set[tuple[str, str]] = set()

    for row in table:
        cells = [str(c or '').strip() for c in row]
        if _linha_ignorada(cells):
            continue

        lote = _cel(cells, col_map.get('lote'))
        guia = _cel(cells, col_map.get('guia'))
        if not lote.isdigit() or not guia.isdigit():
            continue

        cod_usuario = _cel(cells, col_map.get('cod_usuario'))
        nome_usuario = _cel(cells, col_map.get('nome_usuario'))
        cod_servico = _cel(cells, col_map.get('cod_servico'))
        desc_servico = _cel(cells, col_map.get('desc_servico'))
        guia_prest = _cel(cells, col_map.get('guia_prest'))
        data = _parse_data(_cel(cells, col_map.get('data')))
        qtde = _parse_int(_cel(cells, col_map.get('qtde')))
        valor_unit = _money_br(_cel(cells, col_map.get('valor_unit')))
        valor_total = _money_br(_cel(cells, col_map.get('valor_total')))

        obs_parts = []
        if guia_prest:
            obs_parts.append(f'Guia prest.: {guia_prest}')
        via = _cel(cells, col_map.get('via'))
        if via:
            obs_parts.append(f'Via: {via}')

        _adicionar_linha_grupo(
            grupos,
            servicos_unicos,
            lote=lote,
            guia=guia,
            cod_usuario=cod_usuario,
            nome_usuario=nome_usuario,
            cod_servico=cod_servico,
            desc_servico=desc_servico,
            data=data,
            qtde=qtde,
            valor_unit=valor_unit,
            valor_total=valor_total,
            observacao=' | '.join(obs_parts),
        )

    return grupos, servicos_unicos


def _parse_pdf_tabelas(pdf_bytes: bytes) -> tuple[dict[str, dict], set[tuple[str, str]]]:
    grupos: dict[str, dict] = {}
    servicos_unicos: set[tuple[str, str]] = set()
    col_map: dict[str, int | None] | None = None

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                for table in _extrair_tabelas_pdf(page):
                    if not table or len(table) < 2:
                        continue

                    header_idx = None
                    for i, row in enumerate(table[:5]):
                        folded = _fold(' '.join(str(c or '') for c in row))
                        if 'lote' in folded and 'guia' in folded and 'serv' in folded:
                            header_idx = i
                            break

                    if header_idx is None:
                        continue

                    headers = [str(c or '') for c in table[header_idx]]
                    candidate_map = _mapear_colunas_pdf(headers)
                    if candidate_map.get('lote') is None or candidate_map.get('guia') is None:
                        continue

                    if col_map is None:
                        col_map = candidate_map

                    partial_grupos, partial_servicos = _parse_unimed_pdf_table(
                        table[header_idx + 1 :],
                        col_map,
                    )
                    grupos, servicos_unicos = _merge_grupos(
                        grupos, partial_grupos, servicos_unicos, partial_servicos
                    )
    except Exception as exc:
        logger.warning('Falha ao ler tabelas PDF UNIMED: %s', exc)

    return grupos, servicos_unicos


def _merge_grupos(
    base: dict[str, dict],
    extra: dict[str, dict],
    servicos_base: set[tuple[str, str]],
    servicos_extra: set[tuple[str, str]],
) -> tuple[dict[str, dict], set[tuple[str, str]]]:
    for chave, dados in extra.items():
        if chave not in base:
            base[chave] = dados
        else:
            base[chave]['servicos'].extend(dados['servicos'])
    servicos_base.update(servicos_extra)
    return base, servicos_base


def _parse_linha_pdf_texto(line: str) -> dict[str, Any] | None:
    """Fallback: linha de texto extraída do PDF (relatório Produção)."""
    line = re.sub(r'\s+', ' ', (line or '').strip())
    if _linha_ignorada([line]):
        return None

    m_data = re.search(r'(\d{2}/\d{2}/\d{4})', line)
    if not m_data:
        return None

    left = line[: m_data.start()].strip()
    right = line[m_data.end() :].strip()
    data_str = m_data.group(1)

    lm = re.match(
        r'^(\d+)\s+(\d+)\s+(\d+)\s+(.+?)\s+[A-Z]\s+(\d+)\s+(.+?)\s+(\d+)\s+(\w+)\s*$',
        left,
    )
    if not lm:
        return None

    rm = re.match(
        r'^(\d+)\s+(\d+)\s+[\d.,]+\s+\S+\s+\S+\s+.+?\s+([\d.,]+)\s+([\d.,]+)\s*$',
        right,
    )
    if not rm:
        rm = re.match(r'^(\d+)\s+(\d+)\s+.+?\s+([\d.,]+)\s+([\d.,]+)\s*$', right)
    if not rm:
        return None

    return {
        'lote': lm.group(1),
        'guia': lm.group(2),
        'cod_usuario': lm.group(3),
        'nome_usuario': lm.group(4).strip(),
        'cod_servico': lm.group(5),
        'desc_servico': lm.group(6).strip(),
        'guia_prest': lm.group(7),
        'data': _parse_data(data_str),
        'qtde': _parse_int(rm.group(1)),
        'valor_unit': _money_br(rm.group(3)),
        'valor_total': _money_br(rm.group(4)),
    }


def _parse_unimed_pdf_texto(pdf_bytes: bytes) -> tuple[dict[str, dict], set[tuple[str, str]]]:
    grupos: dict[str, dict] = {}
    servicos_unicos: set[tuple[str, str]] = set()

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                texto = page.extract_text() or ''
                for linha in texto.splitlines():
                    parsed = _parse_linha_pdf_texto(linha)
                    if not parsed:
                        continue
                    obs = ''
                    if parsed.get('guia_prest'):
                        obs = f"Guia prest.: {parsed['guia_prest']}"
                    _adicionar_linha_grupo(
                        grupos,
                        servicos_unicos,
                        lote=parsed['lote'],
                        guia=parsed['guia'],
                        cod_usuario=parsed['cod_usuario'],
                        nome_usuario=parsed['nome_usuario'],
                        cod_servico=parsed['cod_servico'],
                        desc_servico=parsed['desc_servico'],
                        data=parsed['data'],
                        qtde=parsed['qtde'],
                        valor_unit=parsed['valor_unit'],
                        valor_total=parsed['valor_total'],
                        observacao=obs,
                    )
    except Exception as exc:
        logger.warning('Falha ao ler texto PDF UNIMED: %s', exc)

    return grupos, servicos_unicos


def _extrair_tabelas_pdf(page) -> list[list[list[Any]]]:
    configs = [
        {'vertical_strategy': 'lines', 'horizontal_strategy': 'lines', 'intersection_tolerance': 8},
        {'vertical_strategy': 'text', 'horizontal_strategy': 'text'},
        {'vertical_strategy': 'lines_strict', 'horizontal_strategy': 'lines_strict'},
        {},
    ]
    seen: list[list[list[Any]]] = []
    for cfg in configs:
        try:
            tables = page.extract_tables(cfg) or []
        except Exception:
            tables = []
        for table in tables:
            if table and len(table) >= 2:
                seen.append(table)
    return seen


def parse_unimed_pdf(pdf_bytes: bytes) -> tuple[dict[str, dict], set[tuple[str, str]], list[str]]:
    """
    Relatório UNIMED «Produção» em PDF.
    Ordem: texto/tabela nativa → OCR (Tesseract) → Google Gemini.
    Ignora: Plano, Tp. Grau, Participação %, Valor Ref.
    """
    avisos: list[str] = []

    grupos, servicos_unicos = _parse_pdf_tabelas(pdf_bytes)
    if not grupos:
        grupos, servicos_unicos = _parse_unimed_pdf_texto(pdf_bytes)
        if grupos:
            avisos.append('PDF lido pelo texto embutido (pdfplumber).')

    if not grupos and _ocr_habilitado():
        avisos.append('Texto nativo não encontrado; aplicando OCR (Tesseract)…')
        texto_ocr = _ocr_pdf_para_texto(pdf_bytes)
        if texto_ocr.strip():
            grupos, servicos_unicos = _grupos_de_texto(texto_ocr)
            if grupos:
                avisos.append(f'OCR extraiu {len(grupos)} guia(s); revise os dados importados.')
        else:
            avisos.append('OCR local não produziu texto (Tesseract indisponível ou PDF ilegível).')
    elif not grupos:
        avisos.append('Texto nativo não encontrado; OCR local indisponível neste servidor.')

    if not grupos:
        from SaudeFinanceira.gemini_config import get_gemini_api_key

        if not _gemini_habilitado():
            avisos.append(
                'PDF escaneado: Gemini desativado no Render (requisição HTTP tem limite ~30s). '
                'Se o OCR não extraiu linhas, exporte o relatório UNIMED «Produção» em .txt.'
            )
            detalhe = ' '.join(a for a in avisos if a)
            raise ValueError(
                'Não foi possível ler o PDF UNIMED «Produção» (sem texto selecionável). '
                f'{detalhe}'
            )

        if not get_gemini_api_key():
            avisos.append('GEMINI_API_KEY não configurada no servidor.')
            detalhe = ' '.join(a for a in avisos if a)
            raise ValueError(
                'Não foi possível importar o PDF UNIMED «Produção». '
                f'{detalhe} Configure GEMINI_API_KEY no Render ou envie o relatório em .txt.'
            )

        avisos.append('Enviando PDF ao Google Gemini…')
        try:
            from faturamento_medico.services.unimed_pdf_gemini import extract_unimed_linhas_gemini

            linhas, gemini_avisos = extract_unimed_linhas_gemini(pdf_bytes)
            avisos.extend(gemini_avisos)
            if linhas:
                grupos, servicos_unicos = _grupos_de_linhas_gemini(linhas)
        except Exception as exc:
            logger.exception('Falha Gemini UNIMED')
            avisos.append(f'Erro ao chamar Gemini: {exc}')

    if not grupos:
        detalhe = ' '.join(a for a in avisos if a)
        raise ValueError(
            'Não foi possível importar o PDF UNIMED «Produção». '
            f'{detalhe} Configure GEMINI_API_KEY no Render se o Gemini falhou.'
        )

    return grupos, servicos_unicos, avisos


def persistir_unimed(
    grupos: dict[str, dict],
    servicos_unicos: set[tuple[str, str]],
    empresa_id: int,
    *,
    codigo_relatorio: str = '',
) -> tuple[int, int, int]:
    """Cria serviços, faturamentos e itens. Retorna (servicos_criados, faturamentos, itens)."""
    codigo_relatorio = (codigo_relatorio or '').strip()
    servicos_criados = 0
    for cod, desc in servicos_unicos:
        if not ServicosMedicos.objects.filter(codigo=cod).exists():
            ServicosMedicos.objects.create(
                codigo=cod,
                servicos=desc,
                porte_anestesico=None,
            )
            servicos_criados += 1

    faturamentos_criados = 0
    itens_criados = 0

    for dados in grupos.values():
        faturamento = FaturamentoMedico.objects.create(
            empresa_id=empresa_id,
            lote=dados['lote'],
            guia=dados['guia'],
            carteirinha=dados['carteirinha'],
            nome=dados['nome'],
            data=dados['data'],
            convenio='UNIMED',
            codigo_relatorio=dados.get('cod_rel') or codigo_relatorio or '',
            status='pendente',
        )

        for servico in dados['servicos']:
            ItemServico.objects.create(
                faturamento=faturamento,
                codigo_servico=servico['codigo'],
                servico=servico['descricao'],
                porte=servico.get('porte') or '',
                percentual=servico.get('percentual') or 0,
                qt=servico.get('qt') or 1,
                valor=servico.get('valor') or 0,
                total=servico.get('total') or 0,
            )
            itens_criados += 1

        faturamento.atualizar_total()
        faturamentos_criados += 1

    return servicos_criados, faturamentos_criados, itens_criados
