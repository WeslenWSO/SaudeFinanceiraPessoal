"""Importa relatório RIS (.xlsx) para FaturamentoMedico."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

from django.db import transaction
from openpyxl import load_workbook

from faturamento_medico.models import FaturamentoMedico, ItemServico
from faturamento_medico.views import (
    _celula_texto,
    _eh_status_agendamento_cancelado,
    _parse_data_ris,
    _parse_valor_ris,
)

BATCH_SIZE = 300


def _mapear_colunas(headers: list[str]) -> dict[str, int | None]:
    lower = [h.lower() for h in headers]

    def idx(*nomes):
        for nome in nomes:
            nome_l = nome.lower()
            if nome_l in lower:
                return lower.index(nome_l)
        return None

    return {
        'unidade': idx('Unidade'),
        'data': idx('Data'),
        'paciente': idx('Paciente'),
        'cns': idx('Cartão Nacional de Saúde', 'Cartao Nacional de Saude'),
        'cpf': idx('CPF'),
        'lote': idx('Número do lote', 'Numero do lote'),
        'procedimento': idx('Procedimento'),
        'prioridade': idx('Prioridade'),
        'horario_inicio': idx('Horário de início', 'Horario de inicio'),
        'horario_fim': idx('Horário de fim', 'Horario de fim'),
        'modalidade': idx('Modalidade'),
        'valor': idx('Valor'),
        'agendado_via': idx('Agendado via', 'Agendado Via'),
        'status': idx('Status do Agendamento'),
        'motivo_cancelamento': idx(
            'Motivo Cancelamento/Desistência/Deleção',
            'Motivo Cancelamento/Desistencia/Delecao',
        ),
        'medico': idx('Médico', 'Medico'),
        'medico_solicitante': idx('Médico solicitante', 'Medico solicitante'),
        'tecnico': idx('Técnico', 'Tecnico'),
        'checkin_por': idx('Check-in por', 'Check-in Por'),
        'agendado_por': idx('Agendado por', 'Agendado Por'),
        'convenio': idx('Viabilidade'),
        'tag': idx('Tag'),
        'indicacao': idx('Indicação clínica', 'Indicacao clinica'),
        'descricao': idx('Descrição', 'Descricao'),
        'obs_pagamento': idx('Observações de Pagamento', 'Observacoes de Pagamento'),
    }


def _ler_grupos(ws, col: dict) -> tuple[list[dict], int, int, set[date]]:
    def get(row, chave):
        i = col.get(chave)
        if i is None or i >= len(row):
            return None
        return row[i]

    grupos_map: dict[str, dict] = {}
    linhas_ignoradas = 0
    linhas_canceladas = 0
    datas_planilha: set[date] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue

        status_raw = _celula_texto(get(row, 'status'), 50)
        paciente = _celula_texto(get(row, 'paciente'), 200)
        procedimento = _celula_texto(get(row, 'procedimento'), 200)
        if not paciente or not procedimento:
            linhas_ignoradas += 1
            continue

        data_fat = _parse_data_ris(get(row, 'data'))
        datas_planilha.add(data_fat)
        cpf = _celula_texto(get(row, 'cpf'), 50)
        cns = _celula_texto(get(row, 'cns'), 50)
        convenio = _celula_texto(get(row, 'convenio'), 100) or 'Particular'
        valor = _parse_valor_ris(get(row, 'valor'))
        if _eh_status_agendamento_cancelado(status_raw):
            linhas_canceladas += 1

        medico_solicitante = _celula_texto(get(row, 'medico_solicitante'), 200) or ''
        chave = (
            f'{paciente}|{data_fat.isoformat()}|{cpf}|{convenio}|'
            f'{status_raw or "ok"}|{medico_solicitante}'
        )
        modalidade = _celula_texto(get(row, 'modalidade'), 20)
        agendado_via = _celula_texto(get(row, 'agendado_via'), 50) or 'RIS'

        if chave not in grupos_map:
            horario_inicio = _celula_texto(get(row, 'horario_inicio'), 20) or None
            horario_fim = _celula_texto(get(row, 'horario_fim'), 20) or None
            horario = ''
            if horario_inicio or horario_fim:
                horario = f'{(horario_inicio or "")} - {(horario_fim or "")}'.strip(' -')
            prioridade = _celula_texto(get(row, 'prioridade'), 50) or None
            urgencia = 'Não'
            if prioridade and prioridade.lower() not in ('eletivo', ''):
                urgencia = 'Sim'
            obs_pag = _celula_texto(get(row, 'obs_pagamento'))
            observacao = f'Pagamento: {obs_pag}' if obs_pag else None

            grupos_map[chave] = {
                'lote': _celula_texto(get(row, 'lote'), 50) or None,
                'carteirinha': cns or None,
                'cpf': cpf or None,
                'horario': horario or None,
                'horario_inicio': horario_inicio,
                'horario_fim': horario_fim,
                'prioridade': prioridade,
                'status_agendamento': status_raw or None,
                'motivo_cancelamento': _celula_texto(get(row, 'motivo_cancelamento'), 255) or None,
                'nome': paciente,
                'nome_associado': paciente,
                'data': data_fat,
                'local': _celula_texto(get(row, 'unidade'), 200) or None,
                'medico': _celula_texto(get(row, 'medico'), 200) or None,
                'medico_solicitante': _celula_texto(get(row, 'medico_solicitante'), 200) or None,
                'tecnico': _celula_texto(get(row, 'tecnico'), 200) or None,
                'checkin_por': _celula_texto(get(row, 'checkin_por'), 200) or None,
                'agendado_por': _celula_texto(get(row, 'agendado_por'), 200) or None,
                'convenio': convenio,
                'tag': _celula_texto(get(row, 'tag'), 100) or None,
                'indicacao_clinica': _celula_texto(get(row, 'indicacao')) or None,
                'descricao': _celula_texto(get(row, 'descricao')) or None,
                'agendado_via': agendado_via,
                'urgencia': urgencia,
                'observacao': observacao,
                'servicos': [],
            }

        grupos_map[chave]['servicos'].append({
            'descricao': procedimento,
            'modalidade': modalidade,
            'com_contraste': 'contraste' in procedimento.lower(),
            'valor': valor,
            'total': valor,
        })

    return list(grupos_map.values()), linhas_ignoradas, linhas_canceladas, datas_planilha


def importar_ris_planilha(
    empresa_id: int,
    caminho: Path | str,
    *,
    substituir_periodo: bool = False,
) -> dict:
    """Importa planilha RIS (.xlsx) para a empresa informada."""
    path = Path(caminho)
    if not path.is_file():
        raise FileNotFoundError(f'Arquivo não encontrado: {path}')

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    headers = [_celula_texto(cell.value).lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = _mapear_colunas(headers)

    obrigatorias = ['data', 'paciente', 'procedimento', 'valor']
    faltando = [c for c in obrigatorias if col[c] is None]
    if faltando:
        wb.close()
        raise ValueError(
            'Arquivo fora do modelo RIS. Colunas obrigatórias não encontradas: '
            + ', '.join(faltando)
        )

    grupos, linhas_ignoradas, linhas_canceladas, datas_planilha = _ler_grupos(ws, col)
    wb.close()

    apagados = 0
    if substituir_periodo and datas_planilha:
        di = min(datas_planilha)
        df = max(datas_planilha)
        apagados, _ = FaturamentoMedico.objects.filter(
            empresa_id=empresa_id,
            data__gte=di,
            data__lte=df,
        ).delete()

    faturamentos_criados = 0
    itens_criados = 0

    with transaction.atomic():
        for offset in range(0, len(grupos), BATCH_SIZE):
            lote_grupos = grupos[offset:offset + BATCH_SIZE]
            fat_objs: list[FaturamentoMedico] = []
            for dados in lote_grupos:
                total = sum((s['total'] for s in dados['servicos']), Decimal('0'))
                fat_objs.append(FaturamentoMedico(
                    empresa_id=empresa_id,
                    lote=dados['lote'],
                    carteirinha=dados['carteirinha'],
                    cpf=dados['cpf'],
                    horario=dados['horario'],
                    horario_inicio=dados['horario_inicio'],
                    horario_fim=dados['horario_fim'],
                    prioridade=dados['prioridade'],
                    status_agendamento=dados['status_agendamento'],
                    motivo_cancelamento=dados['motivo_cancelamento'],
                    nome=dados['nome'],
                    nome_associado=dados['nome_associado'],
                    data=dados['data'],
                    local=dados['local'],
                    medico=dados['medico'],
                    medico_solicitante=dados['medico_solicitante'],
                    tecnico=dados['tecnico'],
                    checkin_por=dados['checkin_por'],
                    agendado_por=dados['agendado_por'],
                    convenio=dados['convenio'],
                    tag=dados['tag'],
                    indicacao_clinica=dados['indicacao_clinica'],
                    descricao=dados['descricao'],
                    agendado_via=dados['agendado_via'],
                    urgencia=dados['urgencia'],
                    observacao=dados['observacao'],
                    codigo_relatorio=None,
                    status='pendente',
                    total=total,
                ))

            FaturamentoMedico.objects.bulk_create(fat_objs, batch_size=BATCH_SIZE)

            item_objs: list[ItemServico] = []
            for fat, dados in zip(fat_objs, lote_grupos):
                for servico in dados['servicos']:
                    item_objs.append(ItemServico(
                        faturamento=fat,
                        codigo_servico='',
                        servico=servico['descricao'],
                        modalidade=servico['modalidade'] or None,
                        com_contraste=servico['com_contraste'],
                        porte='',
                        qt=1,
                        valor=servico['valor'],
                        total=servico['total'],
                    ))

            ItemServico.objects.bulk_create(item_objs, batch_size=BATCH_SIZE)
            faturamentos_criados += len(fat_objs)
            itens_criados += len(item_objs)

    return {
        'faturamentos_criados': faturamentos_criados,
        'itens_criados': itens_criados,
        'linhas_ignoradas': linhas_ignoradas,
        'linhas_canceladas': linhas_canceladas,
        'apagados_antes': apagados,
        'data_min': min(datas_planilha).isoformat() if datas_planilha else None,
        'data_max': max(datas_planilha).isoformat() if datas_planilha else None,
        'grupos': len(grupos),
    }
