from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils import timezone
from django.utils.formats import number_format
from django.http import HttpResponse, Http404, JsonResponse
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from decimal import Decimal, InvalidOperation
import uuid
import logging
import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime, date, timedelta
from difflib import SequenceMatcher
from urllib.parse import urlencode
from .lote_utils import faturamento_elegivel_lote, faturamento_tem_lote_interno, ids_lotes_internos
from servicos_medicos.models import Convenio
from empresa.models import Empresa
from .forms import (
    FaturamentoMedicoForm,
    FaturamentoDocumentacaoForm,
    DocumentoAnexadoForm,
    ItemServicoForm,
    ItemServicoFormSet,
    ServicoDisponivelForm,
)
from .models import (
    DocumentoAnexado,
    ExtratoPagamentoConvenio,
    FaturamentoMedico,
    ItemServico,
    Lote,
    MetaModalidadeSolicitante,
    ServicoDisponivel,
)
from .utils import processar_arquivos_com_gemini, processar_arquivos_com_ocr

logger = logging.getLogger(__name__)


def _moeda_br(valor):
    """Formata valor no padrão brasileiro com milhar: 2.100,00"""
    try:
        return number_format(valor or 0, decimal_pos=2, force_grouping=True, use_l10n=True)
    except (TypeError, ValueError):
        return '0,00'


def _lote_protocolo_faturamento_grid(faturamento, ids_internos):
    lote_val = (faturamento.lote or '').strip()
    return {
        'lote_convenio': '' if lote_val in ids_internos else lote_val,
        'protocolo': (faturamento.guia_lancada or '').strip(),
    }

# Status de agendamento fora da lista principal de faturamento
STATUS_AGENDAMENTO_CANCELADOS = (
    'Cancelado',
    'Desistência',
    'Desistencia',
    'Deletado',
    'Deleção',
    'Delecao',
)


def _q_status_agendamento_cancelados():
    q = Q()
    for status in STATUS_AGENDAMENTO_CANCELADOS:
        q |= Q(status_agendamento__iexact=status)
    return q


def _filtros_listagem_faturamento(request, use_session_fallback=False):
    """Lê filtros da listagem (GET; opcionalmente complementa com sessão)."""
    g = request.GET
    sess = request.session.get('faturamento_filters') or {} if use_session_fallback else {}

    def pick(key, default=''):
        # Parâmetro presente na URL (mesmo vazio) tem prioridade sobre a sessão
        if key in g:
            v = g.get(key)
            return str(v).strip() if v is not None else ''
        if use_session_fallback:
            sv = sess.get(key)
            if sv is not None and str(sv).strip():
                return str(sv).strip()
        return default

    if 'convenio' in g:
        convenios = [c.strip() for c in g.getlist('convenio') if c and str(c).strip()]
    elif use_session_fallback:
        convenios = [c for c in (sess.get('convenio') or []) if c and str(c).strip()]
    else:
        convenios = []

    hoje = date.today()
    data_inicio = pick('data_inicio')
    data_fim = pick('data_fim')
    if not data_inicio:
        data_inicio = (
            (sess.get('data_inicio') if use_session_fallback else None)
            or hoje.replace(day=1).strftime('%Y-%m-%d')
        )
    if not data_fim:
        proximo_mes = hoje.replace(day=28) + timedelta(days=4)
        data_fim = (
            (sess.get('data_fim') if use_session_fallback else None)
            or (proximo_mes - timedelta(days=proximo_mes.day)).strftime('%Y-%m-%d')
        )

    return {
        'nome': pick('nome'),
        'guia': pick('guia'),
        'anestesista': pick('anestesista'),
        'status': pick('status'),
        'status_conferencia': pick('status_conferencia'),
        'lote': pick('lote'),
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'convenios': convenios,
        'codigo_relatorio': pick('codigo_relatorio'),
    }


def _query_export_faturamento(filtros):
    """Monta query string do export Excel a partir dos filtros efetivos da listagem."""
    params = []
    for key in (
        'nome', 'guia', 'anestesista', 'status', 'status_conferencia', 'lote',
        'data_inicio', 'data_fim', 'codigo_relatorio',
    ):
        val = filtros.get(key)
        if val:
            params.append((key, val))
    for conv in filtros.get('convenios') or []:
        params.append(('convenio', conv))
    return urlencode(params)


_FAT_LISTAGEM_FILTER_KEYS = (
    'nome', 'guia', 'codigo_relatorio', 'anestesista',
    'data_inicio', 'data_fim', 'status', 'status_conferencia', 'lote',
)


def _tem_filtros_na_query(request):
    """True se a URL traz parâmetros de filtro (mesmo vazios), exceto paginação."""
    if request.GET.get('limpar'):
        return False
    for key in _FAT_LISTAGEM_FILTER_KEYS:
        if key in request.GET:
            return True
    return 'convenio' in request.GET


def _filtros_dict_from_session(sess):
    return {
        'nome': sess.get('nome') or '',
        'guia': sess.get('guia') or '',
        'anestesista': sess.get('anestesista') or '',
        'status': sess.get('status') or '',
        'status_conferencia': sess.get('status_conferencia') or '',
        'lote': sess.get('lote') or '',
        'data_inicio': sess.get('data_inicio') or '',
        'data_fim': sess.get('data_fim') or '',
        'codigo_relatorio': sess.get('codigo_relatorio') or '',
        'convenios': [c for c in (sess.get('convenio') or []) if c],
    }


def _query_listagem_faturamento(filtros, *, per_page=None):
    qs = _query_export_faturamento(filtros)
    if per_page:
        extra = urlencode({'per_page': per_page})
        qs = f'{qs}&{extra}' if qs else extra
    return qs


def _salvar_filtros_listagem_sessao(request, filtros, *, per_page='25'):
    request.session['faturamento_filters'] = {
        'nome': filtros.get('nome') or '',
        'guia': filtros.get('guia') or '',
        'anestesista': filtros.get('anestesista') or '',
        'status': filtros.get('status') or '',
        'status_conferencia': filtros.get('status_conferencia') or '',
        'lote': filtros.get('lote') or '',
        'data_inicio': filtros.get('data_inicio') or '',
        'data_fim': filtros.get('data_fim') or '',
        'convenio': filtros.get('convenios') or [],
        'codigo_relatorio': filtros.get('codigo_relatorio') or '',
        'per_page': str(per_page),
    }
    request.session.modified = True


def _url_ftlistar_com_filtros_sessao(request):
    sess = request.session.get('faturamento_filters') or {}
    url = reverse('faturamento_medico:ftlistar')
    if not sess:
        return url
    qs = _query_listagem_faturamento(
        _filtros_dict_from_session(sess),
        per_page=sess.get('per_page'),
    )
    return f'{url}?{qs}' if qs else url


def _redirect_ftlistar_com_filtros_sessao(request):
    return redirect(_url_ftlistar_com_filtros_sessao(request))


def _aplicar_filtros_faturamento_qs(qs, filtros):
    """Aplica os mesmos filtros da listagem ao queryset de faturamentos."""
    nome = filtros.get('nome') or ''
    guia = filtros.get('guia') or ''
    anestesista = filtros.get('anestesista') or ''
    status = filtros.get('status') or ''
    status_conferencia = filtros.get('status_conferencia') or ''
    lote = filtros.get('lote') or ''
    data_inicio = filtros.get('data_inicio') or ''
    data_fim = filtros.get('data_fim') or ''
    convenios = filtros.get('convenios') or []
    codigo_relatorio = filtros.get('codigo_relatorio') or ''

    if nome:
        qs = qs.filter(Q(nome__icontains=nome))
    if guia:
        qs = qs.filter(guia__icontains=guia)
    if codigo_relatorio:
        qs = qs.filter(codigo_relatorio__icontains=codigo_relatorio)
    if anestesista:
        qs = qs.filter(anestesista__icontains=anestesista)
    if status:
        qs = qs.filter(status=status)
    if data_inicio:
        qs = qs.filter(data__gte=data_inicio)
    if data_fim:
        qs = qs.filter(data__lte=data_fim)
    if convenios:
        q_objects = Q()
        for conv in convenios:
            if conv:
                q_objects |= _q_convenio_filtro(conv)
        qs = qs.filter(q_objects)
    if lote == '__sem__':
        qs = qs.filter(Q(lote__isnull=True) | Q(lote=''))
    elif lote:
        qs = qs.filter(lote=lote)

    qs = qs.exclude(_q_status_agendamento_cancelados())

    if status_conferencia:
        if status_conferencia == 'CONFERIDO':
            qs = qs.filter(
                Q(itens_servico__status_conferencia='CONFERIDO')
                | Q(itens_servico__conferido=True)
            ).distinct()
        else:
            qs = qs.filter(
                itens_servico__status_conferencia=status_conferencia,
                itens_servico__conferido=False,
            ).distinct()
    return qs


def _q_convenio_filtro(nome: str) -> Q:
    """Correspondência exata de convênio (FUSEX não inclui FUSEX ISENTO / FUSEX PASS)."""
    conv = (nome or '').strip()
    if not conv:
        return Q()
    return Q(convenio__iexact=conv)


def _stats_de_grid_linhas(grid_linhas):
    """Totais por convênio/anestesista a partir das linhas do grid (mesma base do RESUMO)."""
    from collections import defaultdict

    conv = defaultdict(lambda: {'total_valor': Decimal('0'), 'quantidade': 0})
    anest = defaultdict(lambda: {'total_valor': Decimal('0'), 'quantidade': 0})
    valor_total = Decimal('0')
    fat_ids = set()

    for linha in grid_linhas:
        try:
            valor = Decimal(str(linha.get('valor') or 0))
        except (InvalidOperation, ValueError, TypeError):
            valor = Decimal('0')
        valor_total += valor

        fat = linha.get('faturamento')
        if fat is not None:
            fat_ids.add(fat.id)

        nome_conv = ((getattr(fat, 'convenio', None) if fat else None) or '').strip() or 'Não informado'
        conv[nome_conv]['total_valor'] += valor
        conv[nome_conv]['quantidade'] += 1

        nome_anest = ((getattr(fat, 'anestesista', None) if fat else None) or '').strip()
        if nome_anest:
            anest[nome_anest]['total_valor'] += valor
            anest[nome_anest]['quantidade'] += 1

    stats_convenio = sorted(
        [
            {'convenio': k, 'total_valor': v['total_valor'], 'quantidade': v['quantidade']}
            for k, v in conv.items()
        ],
        key=lambda s: -float(s['total_valor'] or 0),
    )
    stats_anestesista = sorted(
        [
            {'anestesista': k, 'total_valor': v['total_valor'], 'quantidade': v['quantidade']}
            for k, v in anest.items()
        ],
        key=lambda s: -float(s['total_valor'] or 0),
    )
    for stat in stats_convenio:
        stat['total_valor_fmt'] = _moeda_br(stat.get('total_valor') or 0)
    for stat in stats_anestesista:
        stat['total_valor_fmt'] = _moeda_br(stat.get('total_valor') or 0)

    return {
        'stats_convenio': stats_convenio,
        'stats_anestesista': stats_anestesista,
        'valor_total': valor_total,
        'total_faturamentos': len(fat_ids),
    }


def _status_linha_faturamento(faturamento, item=None):
    if item is not None:
        return item.status_conferencia_badge()
    if not (faturamento.guia or '').strip():
        return 'FALTA DE GUIA', 'warning'
    if not faturamento.total:
        return 'FALTA DE VALOR NA TABELA', 'danger'
    return 'PENDENTE', 'secondary'


def _modalidade_faturamento_item(faturamento, item=None):
    if item and item.modalidade:
        return item.modalidade
    obs = faturamento.observacao or ''
    if 'Modalidade:' in obs:
        for parte in obs.splitlines():
            if parte.strip().lower().startswith('modalidade:'):
                valor = parte.split(':', 1)[-1].strip()
                if valor:
                    return valor
    return '-'


MODALIDADES_SOLICITANTE = (
    ('US', 'Ultrassonografia'),
    ('CT', 'Tomografia'),
    ('MG', 'Mamografia'),
    ('CR', 'Raio X'),
    ('MR', 'Ressonância'),
    ('EG', 'EEG'),
    ('EC', 'ECG'),
)

METAS_MODALIDADES_SOLICITANTE = (
    ('MR', 'Ressonância'),
    ('US', 'Ultrassonografia'),
    ('CR', 'Raio X'),
    ('CT', 'Tomografia'),
    ('MG', 'Mamografia'),
    ('EG', 'EEG'),
)


def _normalizar_codigo_modalidade(codigo):
    cod = (codigo or '').strip().upper()
    if cod == 'RX':
        cod = 'CR'
    return cod if cod and cod != '-' else 'OUTROS'


SOLICITANTE_NAO_INFORMADO = 'Não informado'
SOLICITANTE_SIMILARIDADE_MIN = 0.90


def _chave_nome_solicitante(texto: str) -> str:
    """Normaliza nome do solicitante para agrupar grafias repetidas."""
    bruto = (texto or '').strip()
    if not bruto:
        return '__sem__'
    t = _normalizar_texto_filtro(bruto)
    for prefixo in (
        'DR.', 'DR ', 'DRA.', 'DRA ', 'PROF.', 'PROF ', 'PROFA.', 'PROFA ',
        'MEDICO ', 'MEDICA ',
    ):
        if t.startswith(prefixo):
            t = t[len(prefixo):].strip()
    t = re.sub(r'^CRM\s*\d+\s*[-–./]?\s*', '', t, flags=re.I)
    t = re.sub(r'\s*\(?CRM\s*\d+\)?\s*', ' ', t, flags=re.I)
    t = re.sub(r'[^\w\s]', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t or '__vazio__'


def _primeiro_nome_compativel(a: str, b: str) -> bool:
    if a == b:
        return True
    if not a or not b:
        return False
    if len(a) > 4 and len(b) > 4 and a.rstrip('S') == b.rstrip('S'):
        return True
    return SequenceMatcher(None, a, b).ratio() >= 0.88


def _sobrenome_compativel(a: str, b: str) -> bool:
    if a == b:
        return True
    if SequenceMatcher(None, a, b).ratio() >= 0.84:
        return True
    min_len = min(len(a), len(b))
    if min_len >= 5 and a[:5] == b[:5]:
        return True
    return False


def _partes_meio_compativeis(partes_a: list[str], partes_b: list[str]) -> bool:
    meio_a = ' '.join(partes_a).strip()
    meio_b = ' '.join(partes_b).strip()
    if not meio_a or not meio_b:
        return True
    if meio_a == meio_b:
        return True
    return SequenceMatcher(None, meio_a, meio_b).ratio() >= 0.82


def _solicitantes_mesma_pessoa(chave_a: str, chave_b: str) -> bool:
    """True se dois nomes normalizados parecem ser o mesmo solicitante."""
    if chave_a == chave_b:
        return True
    if chave_a in ('__sem__', '__vazio__') or chave_b in ('__sem__', '__vazio__'):
        return False
    if SequenceMatcher(None, chave_a, chave_b).ratio() >= SOLICITANTE_SIMILARIDADE_MIN:
        return True
    partes_a = chave_a.split()
    partes_b = chave_b.split()
    if len(partes_a) < 2 or len(partes_b) < 2:
        return False
    if not _primeiro_nome_compativel(partes_a[0], partes_b[0]):
        return False
    if not _sobrenome_compativel(partes_a[-1], partes_b[-1]):
        return False
    return _partes_meio_compativeis(partes_a[1:-1], partes_b[1:-1])


def _chave_agrupamento_solicitante(texto: str) -> str:
    """
    Chave de agrupamento: primeiro + segundo nome (ex.: MARCO|ANTONIO).
    Une MARCO/MARCOS, PIMENTEL/MENEZES etc. quando o nome base coincide.
    """
    chave = _chave_nome_solicitante(texto)
    if chave in ('__sem__', '__vazio__'):
        return chave
    partes = chave.split()
    if len(partes) < 2:
        return chave
    primeiro = partes[0]
    if len(primeiro) >= 6 and primeiro.endswith('S'):
        primeiro = primeiro[:-1]
    return f'{primeiro}|{partes[1]}'


def _escolher_nome_exibicao_solicitante(nomes: list[str]) -> str:
    """Escolhe o nome mais completo/frequente para exibir no grupo."""
    if not nomes:
        return SOLICITANTE_NAO_INFORMADO
    contagem = defaultdict(int)
    for nome in nomes:
        limpo = (nome or '').strip()
        if limpo:
            contagem[limpo] += 1
    return max(contagem.keys(), key=lambda n: (contagem[n], len(n)))


def _construir_grupos_solicitante(raws) -> tuple[dict[str, str], dict[str, dict]]:
    """
    Agrupa nomes repetidos do solicitante.
    Retorna (raw -> nome canônico, grupos por nome canônico).
    """
    por_chave = defaultdict(list)
    for raw in raws:
        original = (raw or '').strip()
        if not original:
            por_chave['__sem__'].append('')
            continue
        por_chave[_chave_agrupamento_solicitante(original)].append(original)

    chaves = [c for c in por_chave.keys() if c not in ('__sem__', '__vazio__')]
    pai = {c: c for c in chaves}

    def _find(chave):
        while pai[chave] != chave:
            pai[chave] = pai[pai[chave]]
            chave = pai[chave]
        return chave

    def _unir(a, b):
        ra, rb = _find(a), _find(b)
        if ra != rb:
            pai[rb] = ra

    for i, ca in enumerate(chaves):
        rep_a = max(por_chave[ca], key=len)
        nome_a = _chave_nome_solicitante(rep_a)
        for cb in chaves[i + 1:]:
            rep_b = max(por_chave[cb], key=len)
            nome_b = _chave_nome_solicitante(rep_b)
            if _solicitantes_mesma_pessoa(nome_a, nome_b):
                _unir(ca, cb)

    grupos_raw = defaultdict(list)
    for chave, nomes in por_chave.items():
        if chave in ('__sem__', '__vazio__'):
            grupos_raw[chave].extend(nomes)
            continue
        grupos_raw[_find(chave)].extend(nomes)

    mapeamento: dict[str, str] = {}
    grupos: dict[str, dict] = {}

    for chave, nomes in grupos_raw.items():
        variantes = sorted({n for n in nomes if n is not None})
        if chave in ('__sem__', '__vazio__') or not variantes:
            canonico = SOLICITANTE_NAO_INFORMADO
            variantes = ['']
        else:
            canonico = _escolher_nome_exibicao_solicitante(variantes)
        if canonico in grupos:
            variantes = sorted(set(grupos[canonico]['variantes'] + variantes))
        grupos[canonico] = {
            'nome': canonico,
            'variantes': variantes,
            'qtd_variantes': len([v for v in variantes if v]),
        }
        for raw in variantes:
            mapeamento[raw] = canonico

    mapeamento[''] = SOLICITANTE_NAO_INFORMADO
    return mapeamento, grupos


def _canonico_solicitante(raw, mapeamento: dict[str, str]) -> str:
    original = (raw or '').strip()
    if not original:
        return SOLICITANTE_NAO_INFORMADO
    return mapeamento.get(original, original)


def _filtrar_por_solicitantes(qs, solicitantes_sel, grupos: dict[str, dict]):
    if not solicitantes_sel:
        return qs
    q_sol = Q()
    for selecionado in solicitantes_sel:
        if selecionado == SOLICITANTE_NAO_INFORMADO:
            q_sol |= Q(medico_solicitante__isnull=True) | Q(medico_solicitante='')
            continue
        info = grupos.get(selecionado)
        variantes = info['variantes'] if info else [selecionado]
        for variante in variantes:
            if variante:
                q_sol |= Q(medico_solicitante__iexact=variante)
    return qs.filter(q_sol) if q_sol else qs


MESES_PT = (
    '', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro',
)


def _periodo_abrange_mais_de_um_mes(data_inicio: date, data_fim: date) -> bool:
    if data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio
    return (data_inicio.year, data_inicio.month) != (data_fim.year, data_fim.month)


def _meses_no_periodo(data_inicio: date, data_fim: date) -> list[tuple[int, int]]:
    if data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio
    meses = []
    cursor = data_inicio.replace(day=1)
    fim = data_fim.replace(day=1)
    while cursor <= fim:
        meses.append((cursor.year, cursor.month))
        if cursor.month == 12:
            cursor = cursor.replace(year=cursor.year + 1, month=1)
        else:
            cursor = cursor.replace(month=cursor.month + 1)
    return meses


def _rotulo_mes_ano(ano: int, mes: int) -> str:
    nome = MESES_PT[mes] if 1 <= mes <= 12 else str(mes)
    return f'{nome}/{ano}'


def _montar_modalidades_card(dados_mes, codigos_modalidade, labels_modalidade):
    modalidades = []
    for codigo in codigos_modalidade:
        qtd = dados_mes['modalidades'].get(codigo, 0)
        if qtd:
            modalidades.append({
                'codigo': codigo,
                'label': labels_modalidade.get(codigo, codigo),
                'quantidade': qtd,
            })
    if dados_mes.get('outros'):
        modalidades.append({
            'codigo': 'OUTROS',
            'label': 'Outros',
            'quantidade': dados_mes['outros'],
        })
    return modalidades


def _carregar_metas_solicitante(empresa_id) -> dict[str, dict[str, int]]:
    metas = {}
    if not empresa_id:
        return metas
    for row in MetaModalidadeSolicitante.objects.filter(empresa_id=empresa_id):
        metas.setdefault(row.solicitante, {})[row.modalidade] = row.meta
    return metas


def _metas_solicitante_grupo(metas_map, solicitante, grupos):
    """Metas cadastradas para o solicitante (inclui grafias do grupo)."""
    nomes = {solicitante}
    info = grupos.get(solicitante) or {}
    for variante in info.get('variantes') or []:
        if variante:
            nomes.add(variante)
    for nome_grupo, dados in grupos.items():
        if nome_grupo == solicitante:
            continue
        variantes = dados.get('variantes') or []
        if solicitante in variantes or any(v in nomes for v in variantes):
            nomes.add(nome_grupo)
            nomes.update(v for v in variantes if v)
    agregado = {}
    for nome in nomes:
        for codigo, valor in (metas_map.get(nome) or {}).items():
            agregado[codigo] = max(agregado.get(codigo, 0), valor)
    return agregado


def _realizado_por_modalidade(dados_resumo) -> dict[str, int]:
    return {
        codigo: qtd
        for codigo, qtd in dados_resumo.get('modalidades', {}).items()
        if qtd
    }


def _enriquecer_modalidades_com_meta(modalidades, realizado_map, metas_solicitante):
    metas = metas_solicitante or {}
    por_codigo = {item['codigo']: item for item in modalidades}
    for codigo, label in METAS_MODALIDADES_SOLICITANTE:
        meta_val = metas.get(codigo)
        if not meta_val:
            continue
        qtd = realizado_map.get(codigo, por_codigo.get(codigo, {}).get('quantidade', 0))
        faltam = max(0, meta_val - qtd)
        atingiu = qtd >= meta_val
        dados_meta = {
            'meta': meta_val,
            'atingiu_meta': atingiu,
            'status_label': 'Meta batida' if atingiu else f'Faltam {faltam}',
            'status_css': 'success' if atingiu else 'danger',
        }
        if codigo in por_codigo:
            por_codigo[codigo].update(dados_meta)
        else:
            modalidades.append({
                'codigo': codigo,
                'label': label,
                'quantidade': qtd,
                **dados_meta,
            })
    return modalidades


def _lista_metas_from_dict(metas_modalidade, realizado_map=None):
    realizado_map = realizado_map or {}
    itens = []
    for codigo, label in METAS_MODALIDADES_SOLICITANTE:
        meta = metas_modalidade.get(codigo)
        if not meta:
            continue
        qtd = realizado_map.get(codigo, 0)
        faltam = max(0, meta - qtd)
        atingiu = qtd >= meta
        itens.append({
            'codigo': codigo,
            'label': label,
            'meta': meta,
            'quantidade': qtd,
            'faltam': faltam,
            'atingiu_meta': atingiu,
            'status_label': 'Meta batida' if atingiu else f'Faltam {faltam}',
            'status_css': 'success' if atingiu else 'danger',
        })
    return itens


def _resumo_metas_solicitante(metas_list):
    if not metas_list:
        return None
    total = len(metas_list)
    atingidas = sum(1 for meta in metas_list if meta['atingiu_meta'])
    pendentes = total - atingidas
    if atingidas == total:
        return {
            'total': total,
            'atingidas': atingidas,
            'pendentes': pendentes,
            'status': 'atingida',
            'label': 'Todas as metas batidas',
            'css': 'success',
        }
    if atingidas == 0:
        return {
            'total': total,
            'atingidas': atingidas,
            'pendentes': pendentes,
            'status': 'nao_atingida',
            'label': 'Nenhuma meta batida',
            'css': 'danger',
        }
    return {
        'total': total,
        'atingidas': atingidas,
        'pendentes': pendentes,
        'status': 'parcial',
        'label': f'{atingidas} de {total} metas batidas',
        'css': 'warning',
    }


def _metas_form_solicitante(solicitante, metas_map, grupos=None):
    metas = _metas_solicitante_grupo(metas_map, solicitante, grupos or {})
    return [
        {
            'codigo': codigo,
            'label': label,
            'valor': metas.get(codigo, ''),
        }
        for codigo, label in METAS_MODALIDADES_SOLICITANTE
    ]


def _salvar_metas_modalidade_solicitante(request):
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('faturamento_medico:listar_exames_por_solicitante')

    solicitante = (request.POST.get('solicitante') or '').strip()
    if not solicitante or solicitante == SOLICITANTE_NAO_INFORMADO:
        messages.error(request, 'Solicitante inválido para definir metas.')
        redirect_qs = (request.POST.get('redirect_qs') or '').strip()
        url = reverse('faturamento_medico:listar_exames_por_solicitante')
        return redirect(f'{url}?{redirect_qs}' if redirect_qs else url)

    for codigo, _ in METAS_MODALIDADES_SOLICITANTE:
        bruto = (request.POST.get(f'meta_{codigo}') or '').strip()
        try:
            meta = max(0, int(bruto)) if bruto else 0
        except ValueError:
            meta = 0
        if meta > 0:
            MetaModalidadeSolicitante.objects.update_or_create(
                empresa_id=empresa_id,
                solicitante=solicitante,
                modalidade=codigo,
                defaults={'meta': meta},
            )
        else:
            MetaModalidadeSolicitante.objects.filter(
                empresa_id=empresa_id,
                solicitante=solicitante,
                modalidade=codigo,
            ).delete()

    messages.success(request, f'Metas atualizadas para {solicitante}.')
    redirect_qs = (request.POST.get('redirect_qs') or '').strip()
    url = reverse('faturamento_medico:listar_exames_por_solicitante')
    return redirect(f'{url}?{redirect_qs}' if redirect_qs else url)


def _novo_resumo_solicitante(codigos_modalidade, por_mes=False):
    resumo = {
        'total': 0,
        'valor': Decimal('0'),
        'modalidades': {codigo: 0 for codigo in codigos_modalidade},
        'outros': 0,
    }
    if por_mes:
        resumo['meses'] = defaultdict(lambda: _novo_resumo_solicitante(codigos_modalidade, por_mes=False))
    return resumo


def _acumular_modalidade_resumo(resumo, modalidade, valor, codigos_modalidade):
    try:
        resumo['valor'] += Decimal(str(valor or 0))
    except (InvalidOperation, ValueError, TypeError):
        pass
    resumo['total'] += 1
    cod_mod = _normalizar_codigo_modalidade(modalidade)
    if cod_mod in resumo['modalidades']:
        resumo['modalidades'][cod_mod] += 1
    else:
        resumo['outros'] += 1


def _badge_status_agendamento(status):
    texto = (status or '').strip() or '-'
    if texto == '-':
        return texto, 'secondary'
    if _eh_status_agendamento_cancelado(status):
        return texto, 'danger'
    norm = (
        texto.lower()
        .replace('ê', 'e')
        .replace('é', 'e')
        .replace('ç', 'c')
        .replace('ã', 'a')
    )
    if 'conclu' in norm or 'realiz' in norm:
        return texto, 'success'
    if 'confirm' in norm:
        return texto, 'primary'
    if 'aguard' in norm or 'pend' in norm or 'andamento' in norm:
        return texto, 'warning'
    return texto, 'secondary'


def _filtrar_por_status_agendamento(qs, status_sel):
    if not status_sel:
        return qs.exclude(_q_status_agendamento_cancelados())
    q_status = Q()
    for status in status_sel:
        if status == 'Não informado':
            q_status |= Q(status_agendamento__isnull=True) | Q(status_agendamento='')
        else:
            q_status |= Q(status_agendamento__iexact=status)
    return qs.filter(q_status) if q_status else qs


def _eh_status_agendamento_cancelado(status):
    if not status:
        return False
    normalizado = (
        str(status)
        .strip()
        .lower()
        .replace('ê', 'e')
        .replace('é', 'e')
        .replace('ç', 'c')
        .replace('ã', 'a')
    )
    return normalizado in {'cancelado', 'desistencia', 'deletado', 'delecao'}


def _parse_hora_minutos(valor):
    """Converte 'HH:MM' / 'H:MM' em minutos desde 00:00. Retorna None se inválido."""
    if valor is None:
        return None
    texto = str(valor).strip()
    if not texto:
        return None
    # Aceita "07:30 - 07:45" pegando só o primeiro horário
    if ' - ' in texto:
        texto = texto.split(' - ', 1)[0].strip()
    texto = texto.replace('.', ':')
    partes = texto.split(':')
    if len(partes) < 2:
        return None
    try:
        hora = int(partes[0])
        minuto = int(partes[1][:2])
    except (TypeError, ValueError):
        return None
    if hora < 0 or hora > 23 or minuto < 0 or minuto > 59:
        return None
    return hora * 60 + minuto


def _parse_data_filtro(valor) -> date | None:
    """
    Converte string de filtro em date (YYYY-MM-DD ou DD/MM/YYYY).
    Rejeita ano fora de 2000-2100. Corrige 0206-07-29 -> 2026-07-29
    (ano 206 no filtro liberava todo o historico desde 2023).
    """
    if isinstance(valor, datetime):
        valor = valor.date()
    if isinstance(valor, date):
        return valor if 2000 <= valor.year <= 2100 else None
    texto = (valor or '').strip()
    if not texto:
        return None
    # 0206-MM-DD (ano 206) -> 2026-MM-DD (digitos 0,2,0,6 -> 2,0,2,6)
    m = re.match(r'^0(\d)(\d)(\d)-(\d{2})-(\d{2})$', texto)
    if m:
        texto = f'2{m.group(2)}{m.group(1)}{m.group(3)}-{m.group(4)}-{m.group(5)}'
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            d = datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
        if 2000 <= d.year <= 2100:
            return d
        return None
    return None


def _periodo_filtro_padrao(hoje: date | None = None) -> tuple[date, date]:
    hoje = hoje or date.today()
    ini = hoje.replace(day=1)
    proximo = hoje.replace(day=28) + timedelta(days=4)
    fim = proximo - timedelta(days=proximo.day)
    return ini, fim


def _periodo_ultimos_dias(hoje: date | None = None, dias: int = 15) -> tuple[date, date]:
    """Período inclusivo: últimos N dias (ex.: 15 = hoje e os 14 anteriores)."""
    hoje = hoje or date.today()
    if dias < 1:
        dias = 1
    return hoje - timedelta(days=dias - 1), hoje


def _montar_dias_regua(data_min: date, data_max: date) -> list[dict]:
    """Lista de dias para a régua visual do filtro."""
    if data_max < data_min:
        data_min, data_max = data_max, data_min
    dias = []
    d = data_min
    while d <= data_max:
        dias.append({
            'iso': d.isoformat(),
            'label': d.strftime('%d/%m'),
            'titulo': d.strftime('%d/%m/%Y'),
            'fim_semana': d.weekday() >= 5,
            'primeiro_mes': d.day == 1,
        })
        d += timedelta(days=1)
    return dias


def _intervalos_sobrepoem(ini_a, fim_a, ini_b, fim_b):
    if None in (ini_a, fim_a, ini_b, fim_b):
        return False
    if fim_a <= ini_a:
        fim_a = ini_a + 1
    if fim_b <= ini_b:
        fim_b = ini_b + 1
    return ini_a < fim_b and ini_b < fim_a


def _horario_slot(faturamento):
    ini = _parse_hora_minutos(faturamento.horario_inicio)
    fim = _parse_hora_minutos(faturamento.horario_fim)
    if ini is None and faturamento.horario:
        partes = str(faturamento.horario).split(' - ')
        if len(partes) >= 2:
            ini = _parse_hora_minutos(partes[0])
            fim = _parse_hora_minutos(partes[1])
        else:
            ini = _parse_hora_minutos(faturamento.horario)
    return ini, fim


def _duracao_slot_minutos(ini, fim):
    if ini is None or fim is None:
        return 0
    if fim <= ini:
        return 0
    return int(fim - ini)


def _fmt_minutos_hhmm(total_minutos: int) -> str:
    total_minutos = max(0, int(total_minutos or 0))
    h, m = divmod(total_minutos, 60)
    if h and m:
        return f'{h}h {m:02d}min'
    if h:
        return f'{h}h'
    return f'{m} min'


def _chave_slot_paciente_maquina(linha: dict):
    """
    Agrupa por paciente + máquina + horário (não por exame).
    Vários exames do mesmo paciente no mesmo slot contam 1 vez.
    """
    fat = linha.get('faturamento')
    if not fat:
        return None
    ini, fim = _horario_slot(fat)
    paciente = (getattr(fat, 'nome', None) or '').strip().upper()
    maquina = (linha.get('maquina_codigo') or '').strip().upper()
    data = getattr(fat, 'data', None)
    if not paciente or not maquina or not data:
        return None
    if ini is None and fim is None:
        horario_txt = (
            f"{getattr(fat, 'horario_inicio', '') or ''} - "
            f"{getattr(fat, 'horario_fim', '') or ''}"
        ).strip(' -') or (getattr(fat, 'horario', None) or '').strip()
        if not horario_txt:
            return None
        return (data, paciente, maquina, horario_txt)
    return (data, paciente, maquina, ini, fim)


def _resumo_slots_por_situacao(grid_linhas: list, situacao: str) -> dict:
    """
    1 slot por (paciente, máquina, horário) na situação informada
    (MAQUINA_VAGA ou REUTILIZADA). Exames no mesmo slot não somam de novo.
    """
    slots = {}
    for linha in grid_linhas:
        if linha.get('situacao') != situacao:
            continue
        chave = _chave_slot_paciente_maquina(linha)
        if not chave:
            continue
        if chave in slots:
            slots[chave]['qtd_exames'] += 1
            continue
        fat = linha['faturamento']
        ini, fim = _horario_slot(fat)
        minutos = _duracao_slot_minutos(ini, fim)
        if getattr(fat, 'horario_inicio', None) or getattr(fat, 'horario_fim', None):
            horario_txt = f"{fat.horario_inicio or ''} - {fat.horario_fim or ''}".strip(' -')
        else:
            horario_txt = (fat.horario or '').strip()
        slots[chave] = {
            'data': fat.data,
            'paciente': fat.nome or '-',
            'maquina': linha.get('maquina') or '-',
            'maquina_codigo': linha.get('maquina_codigo') or '',
            'horario': horario_txt or '-',
            'minutos': minutos,
            'qtd_exames': 1,
            'reutilizado_por': (linha.get('reutilizado_por') or '-') if situacao == 'REUTILIZADA' else '-',
        }

    itens = sorted(
        slots.values(),
        key=lambda s: (s['data'] or date.min, s['maquina'], s['horario'], s['paciente']),
    )
    total_min = sum(s['minutos'] for s in itens)
    por_maquina = {}
    for s in itens:
        cod = s['maquina_codigo'] or s['maquina']
        bucket = por_maquina.setdefault(
            cod,
            {'maquina': s['maquina'], 'slots': 0, 'minutos': 0},
        )
        bucket['slots'] += 1
        bucket['minutos'] += s['minutos']

    por_maquina_lista = [
        {**v, 'minutos_fmt': _fmt_minutos_hhmm(v['minutos'])}
        for v in sorted(por_maquina.values(), key=lambda x: x['maquina'])
    ]

    return {
        'slots': itens,
        'total_slots': len(itens),
        'total_minutos': total_min,
        'total_fmt': _fmt_minutos_hhmm(total_min),
        'por_maquina': por_maquina_lista,
    }


def _resumo_maquinas_paradas(grid_linhas: list) -> dict:
    """
    Calcula slots únicos (paciente+máquina+horário) e agrega a exibição
    por dia + máquina: soma tempo parado e soma exames no slot.
    """
    base = _resumo_slots_por_situacao(grid_linhas, 'MAQUINA_VAGA')
    por_dia = {}
    for s in base['slots']:
        data = s.get('data')
        cod = s.get('maquina_codigo') or s.get('maquina') or ''
        chave = (data, cod)
        if chave not in por_dia:
            por_dia[chave] = {
                'data': data,
                'maquina': s.get('maquina') or '-',
                'maquina_codigo': cod,
                'minutos': 0,
                'qtd_exames': 0,
                'slots': 0,
            }
        por_dia[chave]['minutos'] += int(s.get('minutos') or 0)
        por_dia[chave]['qtd_exames'] += int(s.get('qtd_exames') or 0)
        por_dia[chave]['slots'] += 1

    por_dia_lista = sorted(
        por_dia.values(),
        key=lambda r: (r['data'] or date.min, r['maquina']),
    )
    for r in por_dia_lista:
        r['minutos_fmt'] = _fmt_minutos_hhmm(r['minutos'])

    # Totais por máquina a partir do mesmo recorte filtrado (dia+máquina)
    por_maquina = {}
    for r in por_dia_lista:
        cod = r['maquina_codigo'] or r['maquina']
        bucket = por_maquina.setdefault(
            cod,
            {'maquina': r['maquina'], 'maquina_codigo': cod, 'slots': 0, 'minutos': 0, 'qtd_exames': 0},
        )
        bucket['slots'] += r['slots']
        bucket['minutos'] += r['minutos']
        bucket['qtd_exames'] += r['qtd_exames']
    por_maquina_lista = [
        {**v, 'minutos_fmt': _fmt_minutos_hhmm(v['minutos'])}
        for v in sorted(por_maquina.values(), key=lambda x: x['maquina'])
    ]

    return {
        **base,
        'slots': por_dia_lista,  # tabela agregada dia + máquina (já filtrada)
        'slots_detalhe': base['slots'],
        'por_maquina': por_maquina_lista,
        'total_exames': sum(r['qtd_exames'] for r in por_dia_lista),
    }


def _resumo_horarios_reutilizados(grid_linhas: list) -> dict:
    """
    Horários reutilizados (respeita grid já filtrado):
    - Conta só exames que ocuparam o lugar
    - 1 slot = paciente + máquina + horário (vários exames no mesmo slot: 1 tempo)
    - Exibição agregada: dia + máquina (sem paciente)
    """
    slots = {}
    ids_uso_vistos = set()  # evita contar o mesmo agenda em várias linhas canceladas
    for linha in grid_linhas:
        if linha.get('situacao') != 'REUTILIZADA':
            continue
        maquina_nome = linha.get('maquina') or '-'
        maquina_cod = (linha.get('maquina_codigo') or '').strip().upper()
        for uso in linha.get('reutilizacoes') or []:
            fat_uso = uso.get('faturamento')
            id_uso = None
            if fat_uso is not None:
                ini, fim = _horario_slot(fat_uso)
                data = getattr(fat_uso, 'data', None) or uso.get('data')
                paciente = (getattr(fat_uso, 'nome', None) or uso.get('paciente') or '').strip()
                if getattr(fat_uso, 'horario_inicio', None) or getattr(fat_uso, 'horario_fim', None):
                    horario_txt = f"{fat_uso.horario_inicio or ''} - {fat_uso.horario_fim or ''}".strip(' -')
                else:
                    horario_txt = (getattr(fat_uso, 'horario', None) or uso.get('horario') or '').strip()
                minutos = _duracao_slot_minutos(ini, fim)
                id_uso = getattr(fat_uso, 'id', None) or uso.get('id')
            else:
                data = uso.get('data')
                paciente = (uso.get('paciente') or '').strip()
                horario_txt = (uso.get('horario') or '').strip()
                ini = _parse_hora_minutos(horario_txt.split(' - ')[0] if ' - ' in horario_txt else horario_txt)
                fim = _parse_hora_minutos(horario_txt.split(' - ')[1]) if ' - ' in horario_txt else None
                minutos = _duracao_slot_minutos(ini, fim)
                id_uso = uso.get('id')

            if id_uso is not None:
                if id_uso in ids_uso_vistos:
                    continue
                ids_uso_vistos.add(id_uso)

            paciente_key = paciente.upper()
            if not data or not paciente_key or not maquina_cod:
                continue
            if ini is not None or fim is not None:
                chave = (data, paciente_key, maquina_cod, ini, fim)
            else:
                if not horario_txt:
                    continue
                chave = (data, paciente_key, maquina_cod, horario_txt)

            if chave in slots:
                slots[chave]['qtd_exames'] += 1
                continue

            slots[chave] = {
                'data': data,
                'maquina': uso.get('maquina') or maquina_nome,
                'maquina_codigo': maquina_cod,
                'horario': horario_txt or '-',
                'minutos': minutos,
                'qtd_exames': 1,
            }

    detalhe = sorted(
        slots.values(),
        key=lambda s: (s['data'] or date.min, s['maquina'], s['horario']),
    )

    # Agrega por dia + máquina (exibição)
    por_dia = {}
    for s in detalhe:
        data = s.get('data')
        cod = s.get('maquina_codigo') or s.get('maquina') or ''
        chave = (data, cod)
        if chave not in por_dia:
            por_dia[chave] = {
                'data': data,
                'maquina': s.get('maquina') or '-',
                'maquina_codigo': cod,
                'minutos': 0,
                'qtd_exames': 0,
                'slots': 0,
            }
        por_dia[chave]['minutos'] += int(s.get('minutos') or 0)
        por_dia[chave]['qtd_exames'] += int(s.get('qtd_exames') or 0)
        por_dia[chave]['slots'] += 1

    por_dia_lista = sorted(
        por_dia.values(),
        key=lambda r: (r['data'] or date.min, r['maquina']),
    )
    for r in por_dia_lista:
        r['minutos_fmt'] = _fmt_minutos_hhmm(r['minutos'])

    por_maquina = {}
    for r in por_dia_lista:
        cod = r['maquina_codigo'] or r['maquina']
        bucket = por_maquina.setdefault(
            cod,
            {'maquina': r['maquina'], 'maquina_codigo': cod, 'slots': 0, 'minutos': 0, 'qtd_exames': 0},
        )
        bucket['slots'] += r['slots']
        bucket['minutos'] += r['minutos']
        bucket['qtd_exames'] += r['qtd_exames']

    total_min = sum(r['minutos'] for r in por_dia_lista)
    return {
        'slots': por_dia_lista,
        'slots_detalhe': detalhe,
        'total_slots': len(detalhe),
        'total_minutos': total_min,
        'total_fmt': _fmt_minutos_hhmm(total_min),
        'total_exames': sum(r['qtd_exames'] for r in por_dia_lista),
        'por_maquina': [
            {**v, 'minutos_fmt': _fmt_minutos_hhmm(v['minutos'])}
            for v in sorted(por_maquina.values(), key=lambda x: x['maquina'])
        ],
    }


# Modalidade → (chave da máquina, nome exibido)
# CR (RIS) e RX representam a mesma máquina de Raio X
MAQUINAS_POR_MODALIDADE = {
    'US': ('US', 'Máquina de Ultrassom'),
    'EG': ('EG', 'Máquina de Eletroencefalograma'),
    'MR': ('MR', 'Máquina de Ressonância'),
    'CT': ('CT', 'Máquina de Tomografia'),
    'RX': ('RX', 'Máquina de Raio X'),
    'CR': ('RX', 'Máquina de Raio X'),
    'MG': ('MG', 'Máquina de Mamografia'),
}


def _maquina_por_modalidade(modalidade):
    """Retorna (chave_maquina, nome_maquina) a partir do código de modalidade."""
    cod = (modalidade or '').strip().upper()
    if not cod or cod == '-':
        return None, 'Modalidade não informada'
    if cod in MAQUINAS_POR_MODALIDADE:
        return MAQUINAS_POR_MODALIDADE[cod]
    return cod, f'Máquina {cod}'


def _normalizar_texto_filtro(texto: str) -> str:
    t = unicodedata.normalize('NFKD', (texto or ''))
    t = ''.join(c for c in t if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', t.upper().strip())


def _resolver_codigos_maquina(valor: str) -> set[str]:
    """
    Resolve filtro de máquina por código (MR, US…) ou nome
    ('Máquina de Ressonância', 'ressonancia', etc.).
    """
    bruto = (valor or '').strip()
    if not bruto:
        return set()
    up = bruto.upper().strip()
    norm = _normalizar_texto_filtro(bruto)
    # Remove prefixo comum "MAQUINA DE/DA"
    norm_limpo = re.sub(r'^MAQUINA\s+(DE\s+|DA\s+)?', '', norm).strip()

    codigos: set[str] = set()
    # Código direto (MR, US, RX…)
    for chave, nome in MAQUINAS_POR_MODALIDADE.values():
        if up == chave or norm == chave:
            codigos.add(chave)
    if up in MAQUINAS_POR_MODALIDADE:
        chave, _ = MAQUINAS_POR_MODALIDADE[up]
        codigos.add(chave)
    if codigos:
        return codigos

    # Por nome / trecho do nome
    for chave, nome in dict(MAQUINAS_POR_MODALIDADE.values()).items():
        nome_n = _normalizar_texto_filtro(nome)
        nome_n_limpo = re.sub(r'^MAQUINA\s+(DE\s+|DA\s+)?', '', nome_n).strip()
        if (
            norm == nome_n
            or norm_limpo == nome_n_limpo
            or (norm_limpo and norm_limpo in nome_n)
            or (nome_n_limpo and nome_n_limpo in norm)
            or (norm_limpo and nome_n_limpo.startswith(norm_limpo))
            or (norm_limpo and norm_limpo in nome_n_limpo)
        ):
            codigos.add(chave)

    # Aliases frequentes
    aliases = {
        'RESSONANCIA': 'MR',
        'RM': 'MR',
        'MRI': 'MR',
        'ULTRASSOM': 'US',
        'ULTRA SOM': 'US',
        'TOMOGRAFIA': 'CT',
        'TOMO': 'CT',
        'RAIO X': 'RX',
        'RAIOX': 'RX',
        'RX': 'RX',
        'MAMOGRAFIA': 'MG',
        'MAMO': 'MG',
        'EEG': 'EG',
        'ELETROENCEFALOGRAMA': 'EG',
    }
    if norm_limpo in aliases:
        codigos.add(aliases[norm_limpo])
    for alias, chave in aliases.items():
        if alias in norm_limpo or norm_limpo in alias:
            codigos.add(chave)

    return codigos


def _parece_filtro_maquina(texto: str) -> bool:
    """True se o texto do campo paciente parece nome/código de máquina."""
    n = _normalizar_texto_filtro(texto)
    if not n:
        return False
    if n.startswith('MAQUINA'):
        return True
    if _resolver_codigos_maquina(texto):
        # Evita confundir nome de paciente curto com código US/MR isolado
        if len(n) <= 3 and n.isalpha():
            return n in {c for c, _ in MAQUINAS_POR_MODALIDADE.values()} | set(MAQUINAS_POR_MODALIDADE)
        return 'MAQUINA' in n or any(
            p in n for p in (
                'RESSON', 'ULTRA', 'TOMO', 'RAIO', 'MAMO', 'ELETRO', 'EEG',
            )
        )
    return False


def _chaves_maquina_do_faturamento(faturamento):
    """Conjunto de chaves de máquina presentes nos itens do faturamento."""
    chaves = set()
    for item in faturamento.itens_servico.all():
        chave, _ = _maquina_por_modalidade(item.modalidade)
        if chave:
            chaves.add(chave)
    return chaves


def _analisar_vaga_maquina(faturamento, modalidade, candidatos_ativos):
    """
    Verifica se o horário cancelado foi reutilizado na MESMA máquina (modalidade).
    US/EG/MR/CT/RX(CR)/MG.
    """
    ini, fim = _horario_slot(faturamento)
    local = (faturamento.local or '').strip().upper()
    chave_maquina, nome_maquina = _maquina_por_modalidade(modalidade)

    if not chave_maquina:
        return {
            'situacao': 'SEM_MODALIDADE',
            'situacao_css': 'secondary',
            'situacao_label': 'Sem modalidade',
            'maquina': nome_maquina,
            'maquina_codigo': '-',
            'reutilizado_por': '-',
            'reutilizacoes': [],
        }

    reutilizacoes = []

    for outro in candidatos_ativos:
        if outro.id == faturamento.id:
            continue
        if outro.data != faturamento.data:
            continue
        outro_local = (outro.local or '').strip().upper()
        if local and outro_local and local != outro_local:
            continue

        # Obrigatório: mesma máquina (modalidade)
        chaves_outro = _chaves_maquina_do_faturamento(outro)
        if chave_maquina not in chaves_outro:
            continue

        ini_o, fim_o = _horario_slot(outro)
        sobrepoe = _intervalos_sobrepoem(ini, fim, ini_o, fim_o)
        if not sobrepoe:
            h1 = (faturamento.horario or '').strip()
            h2 = (outro.horario or '').strip()
            if not h1 or not h2 or h1 != h2:
                continue

        procs = []
        mods_usadas = []
        for it in outro.itens_servico.all():
            chave_it, _ = _maquina_por_modalidade(it.modalidade)
            if chave_it == chave_maquina:
                if (it.servico or '').strip():
                    procs.append((it.servico or '').strip())
                if (it.modalidade or '').strip():
                    mods_usadas.append((it.modalidade or '').strip().upper())
        if not procs:
            procs = [(outro.servico or '-')]

        horario_txt = ''
        if outro.horario_inicio or outro.horario_fim:
            horario_txt = f"{outro.horario_inicio or ''} - {outro.horario_fim or ''}".strip(' -')
        else:
            horario_txt = (outro.horario or '').strip()

        reutilizacoes.append({
            'id': outro.id,
            'faturamento': outro,
            'paciente': outro.nome or '-',
            'status': outro.status_agendamento or '-',
            'horario': horario_txt or '-',
            'data': outro.data,
            'local': outro.local or '-',
            'convenio': outro.convenio or '-',
            'agendado_via': outro.agendado_via or '-',
            'procedimentos': ', '.join(procs),
            'modalidade': ', '.join(dict.fromkeys(mods_usadas)) or (modalidade or '-'),
            'maquina': nome_maquina,
        })

    base = {
        'maquina': nome_maquina,
        'maquina_codigo': chave_maquina,
    }

    if reutilizacoes:
        nomes = ', '.join(
            f"{r['paciente']} ({r['status']})" for r in reutilizacoes[:3]
        )
        if len(reutilizacoes) > 3:
            nomes += f' +{len(reutilizacoes) - 3}'
        return {
            **base,
            'situacao': 'REUTILIZADA',
            'situacao_css': 'success',
            'situacao_label': 'Horário reutilizado',
            'reutilizado_por': nomes,
            'reutilizacoes': reutilizacoes,
        }

    return {
        **base,
        'situacao': 'MAQUINA_VAGA',
        'situacao_css': 'warning',
        'situacao_label': 'Máquina vaga',
        'reutilizado_por': '-',
        'reutilizacoes': [],
    }


def listar_faturamentos(request):
    """Lista todos os faturamentos médicos com filtros"""
    if request.GET.get('limpar'):
        request.session.pop('faturamento_filters', None)
        request.session.modified = True
        return redirect('faturamento_medico:ftlistar')

    tem_filtros_url = _tem_filtros_na_query(request)
    if not tem_filtros_url:
        sess = request.session.get('faturamento_filters') or {}
        if sess:
            return redirect(_url_ftlistar_com_filtros_sessao(request))

    empresa_id = request.session.get('empresa_id')
    if empresa_id:
        faturamentos = FaturamentoMedico.objects.filter(empresa_id=empresa_id).order_by('-data')
    else:
        faturamentos = FaturamentoMedico.objects.all().order_by('-data')

    filtros = _filtros_listagem_faturamento(request, use_session_fallback=not tem_filtros_url)
    faturamentos = _aplicar_filtros_faturamento_qs(faturamentos, filtros)
    nome = filtros['nome']
    guia = filtros['guia']
    anestesista = filtros['anestesista']
    status = filtros['status']
    status_conferencia = filtros['status_conferencia']
    lote = filtros['lote']
    data_inicio = filtros['data_inicio']
    data_fim = filtros['data_fim']
    convenios = filtros['convenios']
    codigo_relatorio = filtros['codigo_relatorio']

    def _label_convenio_curto(nome, max_len=36):
        n = (nome or '').strip() or 'Não informado'
        if len(n) <= max_len:
            return n
        return n[: max_len - 1].rstrip() + '…'

    # Buscar convênios disponíveis para a empresa
    convenios_disponiveis = []
    if empresa_id:
        from servicos_medicos.models import Convenio
        convenios_disponiveis = list(Convenio.objects.filter(empresa_id=empresa_id).order_by('nome'))
        if not convenios_disponiveis:
            # Convênios padrão se nenhum for encontrado para a empresa
            convenios_padrao = [
                {'nome': 'CBSAUDE'},
                {'nome': 'PM'},
                {'nome': 'UNIMED'},
                {'nome': 'BRADESCO'},
                {'nome': 'GEAP'},
                {'nome': 'SAUDE CAIXA'},
                {'nome': 'POSTAL SAUDE'},
                {'nome': 'FUSEX'},
                {'nome': 'LIFE EMPRESARIAL'},
                {'nome': 'CASSI'},
                {'nome': 'GCARD'},
                {'nome': 'PERSONAL NET'},
            ]
            convenios_disponiveis = convenios_padrao

    # Buscar lotes disponíveis para a empresa
    lotes_disponiveis = []
    lotes_filtro = []
    if empresa_id:
        lotes_disponiveis = [
            lote for lote in Lote.objects.filter(empresa_id=empresa_id, baixado=False)
            .prefetch_related('linhas_extrato_pagamento')
            .order_by('-id')
            if lote.aberto_para_adicionar()
        ]
        lotes_filtro = list(Lote.objects.filter(empresa_id=empresa_id).order_by('-id'))
        # Inclui valores de lote vindos da importação RIS (texto livre)
        ids_lote = {str(l.id) for l in lotes_filtro}
        lotes_extras = (
            FaturamentoMedico.objects
            .filter(empresa_id=empresa_id)
            .exclude(Q(lote__isnull=True) | Q(lote=''))
            .values_list('lote', flat=True)
            .distinct()
        )
        for valor in lotes_extras:
            if valor and str(valor) not in ids_lote:
                lotes_filtro.append(type('LoteExtra', (), {
                    'id': valor,
                    'convenio': 'RIS/Outro',
                    'total_lote': None,
                })())

    # Grid pós-filtragem (modelo RIS): uma linha por procedimento
    faturamentos = faturamentos.prefetch_related('itens_servico')
    grid_linhas = []
    ids_lotes_int = ids_lotes_internos(empresa_id) if empresa_id else set()

    # Cache de preços da tabela por empresa/convênio (código e descrição)
    precos_por_codigo = set()
    precos_por_descricao = set()
    if empresa_id:
        from servicos_medicos.models import TabelaPreco
        tabelas = (
            TabelaPreco.objects
            .filter(empresa_id=empresa_id)
            .select_related('codigo_servico', 'convenio')
        )
        for t in tabelas:
            conv = (t.convenio.nome or '').strip().upper()
            cod = (t.codigo_servico.codigo or '').strip().upper()
            desc = (t.codigo_servico.servicos or '').strip().upper()
            if conv and cod:
                precos_por_codigo.add((conv, cod))
            if conv and desc:
                precos_por_descricao.add((conv, desc))

    def _tem_preco_tabela(faturamento, item):
        if not item:
            return False
        valor = item.total if item.total is not None else item.valor
        if valor is None or valor == 0:
            return False
        conv = (faturamento.convenio or '').strip().upper()
        if not conv:
            return False
        cod = (item.codigo_servico or '').strip().upper()
        if cod and (conv, cod) in precos_por_codigo:
            return True
        desc = (item.servico or '').strip().upper()
        if desc and (conv, desc) in precos_por_descricao:
            return True
        # Sem código/descrição na tabela do convênio
        if not precos_por_codigo and not precos_por_descricao:
            # Sem tabela cadastrada: considera falta de preço
            return False
        return False

    def _modalidade_item(faturamento, item=None):
        if item and item.modalidade:
            return item.modalidade
        obs = faturamento.observacao or ''
        if 'Modalidade:' in obs:
            for parte in obs.splitlines():
                if parte.strip().lower().startswith('modalidade:'):
                    valor = parte.split(':', 1)[-1].strip()
                    if valor:
                        return valor
        return '-'

    for faturamento in faturamentos:
        itens = list(faturamento.itens_servico.all())
        if not itens:
            status_label, status_css = _status_linha_faturamento(faturamento)
            if status_conferencia and status_label != status_conferencia:
                continue
            grid_linhas.append({
                'faturamento': faturamento,
                'item': None,
                'paciente': faturamento.nome or '-',
                'nome_associado': faturamento.nome_associado or faturamento.nome or '-',
                'procedimento': faturamento.servico or '-',
                'modalidade': _modalidade_item(faturamento),
                'com_contraste': 'contraste' in (faturamento.servico or '').lower(),
                'valor': faturamento.total or 0,
                'valor_fmt': _moeda_br(faturamento.total or 0),
                'conferido': False,
                'status_conferencia': status_label,
                'status_conferencia_css': status_css,
                'mostrar_selecao_lote': True,
                'tem_lote_interno': faturamento_tem_lote_interno(
                    faturamento, ids_internos=ids_lotes_int
                ),
                'elegivel_lote': faturamento_elegivel_lote(
                    faturamento, ids_internos=ids_lotes_int
                ),
                **_lote_protocolo_faturamento_grid(faturamento, ids_lotes_int),
            })
            continue
        itens_filtrados = []
        for item in itens:
            status_label, status_css = _status_linha_faturamento(faturamento, item)
            if status_conferencia and status_label != status_conferencia:
                continue
            itens_filtrados.append((item, status_label, status_css))
        for idx, (item, status_label, status_css) in enumerate(itens_filtrados):
            valor_item = item.total if item.total is not None else (item.valor or 0)
            grid_linhas.append({
                'faturamento': faturamento,
                'item': item,
                'paciente': faturamento.nome or '-',
                'nome_associado': faturamento.nome_associado or faturamento.nome or '-',
                'procedimento': item.servico or '-',
                'modalidade': _modalidade_item(faturamento, item),
                'com_contraste': item.com_contraste,
                'valor': valor_item,
                'valor_fmt': _moeda_br(valor_item),
                'conferido': item.conferido,
                'status_conferencia': status_label,
                'status_conferencia_css': status_css,
                'mostrar_selecao_lote': idx == 0,
                'tem_lote_interno': faturamento_tem_lote_interno(
                    faturamento, ids_internos=ids_lotes_int
                ),
                'elegivel_lote': faturamento_elegivel_lote(
                    faturamento, ids_internos=ids_lotes_int
                ),
                **_lote_protocolo_faturamento_grid(faturamento, ids_lotes_int),
            })

    # Resumo por modalidade (conforme filtros / grid de procedimentos)
    MODALIDADES_RESUMO = [
        ('US', 'QUANTIDADE DE ULTRASSONOGRAFIA'),
        ('CT', 'QUANTIDADE DE TOMOGRAFIA'),
        ('MG', 'QUANTIDADE DE MAMOGRAFIA'),
        ('CR', 'QUANTIDADE DE RAIO X'),
        ('MR', 'QUANTIDADE DE RESSONÂNCIA'),
        ('EG', 'QUANTIDADE DE ELETROENCEFALOGRAMA'),
        ('EC', 'QUANTIDADE DE ELETROCARDIOGRAMA'),
    ]
    contagem_modalidade = {codigo: 0 for codigo, _ in MODALIDADES_RESUMO}
    resumo_quantidade_total = 0
    for linha in grid_linhas:
        resumo_quantidade_total += 1
        codigo_mod = (linha.get('modalidade') or '').strip().upper()
        # CR e RX contam juntos como Raio X
        if codigo_mod == 'RX':
            codigo_mod = 'CR'
        if codigo_mod in contagem_modalidade:
            contagem_modalidade[codigo_mod] += 1

    resumo_modalidades = [
        {
            'codigo': codigo,
            'label': label,
            'quantidade': contagem_modalidade[codigo],
        }
        for codigo, label in MODALIDADES_RESUMO
    ]

    totais_grid = _stats_de_grid_linhas(grid_linhas)
    stats_convenio = totais_grid['stats_convenio']
    stats_anestesista = totais_grid['stats_anestesista']
    valor_total = totais_grid['valor_total']
    resumo_valor_total = valor_total
    total_faturamentos = totais_grid['total_faturamentos']

    grafico_convenio_labels = [_label_convenio_curto(s.get('convenio')) for s in stats_convenio]
    grafico_convenio_keys = [(s.get('convenio') or '').strip() or 'Não informado' for s in stats_convenio]
    grafico_convenio_valores = [float(s.get('total_valor') or 0) for s in stats_convenio]
    grafico_convenio_qtde = [int(s.get('quantidade') or 0) for s in stats_convenio]

    MODALIDADES_LABEL_CURTO = {
        'US': 'Ultrassonografia',
        'CT': 'Tomografia',
        'MG': 'Mamografia',
        'CR': 'Raio X',
        'MR': 'Ressonância',
        'EG': 'EEG',
        'EC': 'ECG',
    }
    grafico_modalidade_labels = [
        MODALIDADES_LABEL_CURTO.get(m['codigo'], m['codigo']) for m in resumo_modalidades
    ]
    grafico_modalidade_valores = [m['quantidade'] for m in resumo_modalidades]

    # Gráfico por procedimento (top 8 + Outros)
    contagem_procedimento = defaultdict(int)
    for linha in grid_linhas:
        nome_proc = (linha.get('procedimento') or '').strip() or 'Não informado'
        contagem_procedimento[nome_proc] += 1
    procedimentos_ordenados = sorted(
        contagem_procedimento.items(), key=lambda x: (-x[1], x[0].lower())
    )
    TOP_PROCEDIMENTOS = 8

    def _label_proc_curto(nome, max_len=42):
        n = (nome or '').strip()
        if len(n) <= max_len:
            return n
        return n[: max_len - 1].rstrip() + '…'

    if len(procedimentos_ordenados) <= TOP_PROCEDIMENTOS:
        grafico_procedimento_labels = [_label_proc_curto(n) for n, _ in procedimentos_ordenados]
        grafico_procedimento_valores = [q for _, q in procedimentos_ordenados]
    else:
        top = procedimentos_ordenados[:TOP_PROCEDIMENTOS]
        resto = sum(q for _, q in procedimentos_ordenados[TOP_PROCEDIMENTOS:])
        grafico_procedimento_labels = [_label_proc_curto(n) for n, _ in top] + ['Outros']
        grafico_procedimento_valores = [q for _, q in top] + [resto]

    # Detalhe por convênio (modalidade + procedimento) para drill-down nos gráficos
    MOD_CODIGO_LABEL = dict(MODALIDADES_LABEL_CURTO)
    por_convenio_mod = defaultdict(lambda: defaultdict(int))
    por_convenio_proc = defaultdict(lambda: defaultdict(int))
    for linha in grid_linhas:
        fat = linha.get('faturamento')
        conv = ((getattr(fat, 'convenio', None) if fat else None) or '').strip() or 'Não informado'
        codigo_mod = (linha.get('modalidade') or '').strip().upper()
        if codigo_mod == 'RX':
            codigo_mod = 'CR'
        label_mod = MOD_CODIGO_LABEL.get(codigo_mod, codigo_mod if codigo_mod and codigo_mod != '-' else 'Outros')
        por_convenio_mod[conv][label_mod] += 1
        nome_proc = (linha.get('procedimento') or '').strip() or 'Não informado'
        por_convenio_proc[conv][nome_proc] += 1

    detalhe_convenio = {}
    for conv in set(list(por_convenio_mod.keys()) + list(por_convenio_proc.keys())):
        mods = sorted(por_convenio_mod[conv].items(), key=lambda x: (-x[1], x[0]))
        procs = sorted(por_convenio_proc[conv].items(), key=lambda x: (-x[1], x[0].lower()))
        if len(procs) > TOP_PROCEDIMENTOS:
            top_p = procs[:TOP_PROCEDIMENTOS]
            resto_p = sum(q for _, q in procs[TOP_PROCEDIMENTOS:])
            proc_labels = [_label_proc_curto(n) for n, _ in top_p] + ['Outros']
            proc_vals = [q for _, q in top_p] + [resto_p]
        else:
            proc_labels = [_label_proc_curto(n) for n, _ in procs]
            proc_vals = [q for _, q in procs]
        detalhe_convenio[conv] = {
            'modalidade': {
                'labels': [n for n, _ in mods],
                'valores': [q for _, q in mods],
            },
            'procedimento': {
                'labels': proc_labels,
                'valores': proc_vals,
            },
        }

    # Paginação da tabela (resumo/gráficos usam a lista completa)
    PER_PAGE_OPCOES = (25, 100, 250, 500, 1000)
    try:
        per_page = int(request.GET.get('per_page') or 25)
    except (TypeError, ValueError):
        per_page = 25
    if per_page not in PER_PAGE_OPCOES:
        per_page = 25
    grid_total = len(grid_linhas)
    paginator = Paginator(grid_linhas, per_page)
    page_obj = paginator.get_page(request.GET.get('page') or 1)
    grid_linhas = list(page_obj.object_list)

    # Armazenar filtros na sessão (preserva ao voltar de edição / gerar lote / menu)
    _salvar_filtros_listagem_sessao(request, filtros, per_page=per_page)

    context = {
        'faturamentos': faturamentos,
        'grid_linhas': grid_linhas,
        'grid_total': grid_total,
        'page_obj': page_obj,
        'per_page': per_page,
        'per_page_opcoes': PER_PAGE_OPCOES,
        'total_faturamentos': total_faturamentos,
        'valor_total': valor_total,
        'valor_total_fmt': _moeda_br(valor_total),
        'resumo_modalidades': resumo_modalidades,
        'resumo_quantidade_total': resumo_quantidade_total,
        'resumo_valor_total': resumo_valor_total,
        'resumo_valor_total_fmt': _moeda_br(resumo_valor_total),
        'grafico_modalidade_labels': json.dumps(grafico_modalidade_labels, ensure_ascii=False),
        'grafico_modalidade_valores': json.dumps(grafico_modalidade_valores),
        'grafico_procedimento_labels': json.dumps(grafico_procedimento_labels, ensure_ascii=False),
        'grafico_procedimento_valores': json.dumps(grafico_procedimento_valores),
        'grafico_convenio_labels': json.dumps(grafico_convenio_labels, ensure_ascii=False),
        'grafico_convenio_keys': json.dumps(grafico_convenio_keys, ensure_ascii=False),
        'grafico_convenio_valores': json.dumps(grafico_convenio_valores),
        'grafico_convenio_qtde': json.dumps(grafico_convenio_qtde),
        'detalhe_convenio': json.dumps(detalhe_convenio, ensure_ascii=False),
        'status_conferencia_choices': ItemServico.STATUS_CONFERENCIA_CHOICES,
        'status_faturamento_choices': FaturamentoMedico.FATURAMENTO_STATUS_CHOICES,
        'stats_convenio': stats_convenio,
        'stats_anestesista': stats_anestesista,
        'convenios_disponiveis': convenios_disponiveis,
        'lotes_disponiveis': lotes_disponiveis,
        'lotes_filtro': lotes_filtro,
        'filtros': {
            'nome': nome,
            'guia': guia,
            'anestesista': anestesista,
            'status': status,
            'status_conferencia': status_conferencia,
            'lote': lote,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'convenio': convenios,
            'codigo_relatorio': codigo_relatorio,
            'per_page': str(per_page),
        },
        'export_query_string': _query_export_faturamento(filtros),
        'listagem_query_string': _query_listagem_faturamento(filtros, per_page=per_page),
    }

    return render(request, 'faturamento_medico/listar.html', context)


def listar_cancelados(request):
    """Lista procedimentos com status Cancelado, Desistência ou Deletado."""
    empresa_id = request.session.get('empresa_id')
    if empresa_id:
        faturamentos = FaturamentoMedico.objects.filter(empresa_id=empresa_id).order_by('-data', 'nome')
    else:
        faturamentos = FaturamentoMedico.objects.all().order_by('-data', 'nome')

    faturamentos = faturamentos.filter(_q_status_agendamento_cancelados())

    nome = (request.GET.get('nome') or '').strip()
    convenios = request.GET.getlist('convenio')
    status_agendamento = request.GET.get('status_agendamento')
    situacao_vaga = (request.GET.get('situacao_vaga') or '').strip()
    maquina_raw = (request.GET.get('maquina') or '').strip()

    # Campo "Paciente" preenchido com nome de máquina (erro comum) → usa como filtro de máquina
    if nome and not maquina_raw and _parece_filtro_maquina(nome):
        maquina_raw = nome
        nome = ''

    codigos_maquina = _resolver_codigos_maquina(maquina_raw)
    # Código canônico para o <select> (primeiro do conjunto)
    maquina = next(iter(sorted(codigos_maquina)), '') if codigos_maquina else ''

    hoje = date.today()
    dias_regua_qtd = 15
    regua_ini, regua_fim = _periodo_ultimos_dias(hoje, dias_regua_qtd)
    di_padrao, df_padrao = regua_ini, regua_fim
    di = _parse_data_filtro(request.GET.get('data_inicio'))
    df = _parse_data_filtro(request.GET.get('data_fim'))
    if di is None:
        di = di_padrao
    if df is None:
        df = df_padrao
    if di > df:
        di, df = df, di
    if di < regua_ini:
        di = regua_ini
    if df > regua_fim:
        df = regua_fim
    if di > df:
        di, df = df, di
    data_inicio = di.isoformat()
    data_fim = df.isoformat()

    if nome:
        faturamentos = faturamentos.filter(Q(nome__icontains=nome))
    # Sempre aplica período (date objects — evita ano 0206 / string inválida)
    faturamentos = faturamentos.filter(data__gte=di, data__lte=df)
    if convenios:
        q_objects = Q()
        for conv in convenios:
            if conv:
                q_objects |= _q_convenio_filtro(conv)
        faturamentos = faturamentos.filter(q_objects)
    if status_agendamento:
        faturamentos = faturamentos.filter(status_agendamento__iexact=status_agendamento)

    faturamentos = list(faturamentos.prefetch_related('itens_servico'))

    # Candidatos ativos no mesmo período (outros status) para checar reuso do horário
    datas = {f.data for f in faturamentos if f.data}
    ativos_qs = FaturamentoMedico.objects.none()
    if datas:
        ativos_qs = (
            FaturamentoMedico.objects
            .filter(data__in=datas)
            .exclude(_q_status_agendamento_cancelados())
            .prefetch_related('itens_servico')
        )
        if empresa_id:
            ativos_qs = ativos_qs.filter(empresa_id=empresa_id)
    ativos_por_data = {}
    for ativo in ativos_qs:
        ativos_por_data.setdefault(ativo.data, []).append(ativo)

    def _passa_filtro_maquina(analise: dict) -> bool:
        if not codigos_maquina:
            return True
        return (analise.get('maquina_codigo') or '').strip().upper() in codigos_maquina

    grid_linhas = []
    for faturamento in faturamentos:
        candidatos = ativos_por_data.get(faturamento.data, [])
        itens = list(faturamento.itens_servico.all())
        if not itens:
            analise = _analisar_vaga_maquina(faturamento, '-', candidatos)
            if situacao_vaga and analise['situacao'] != situacao_vaga:
                continue
            if not _passa_filtro_maquina(analise):
                continue
            grid_linhas.append({
                'faturamento': faturamento,
                'procedimento': faturamento.servico or '-',
                'modalidade': '-',
                'valor': faturamento.total or 0,
                'valor_fmt': _moeda_br(faturamento.total or 0),
                **analise,
            })
            continue
        for item in itens:
            modalidade = item.modalidade or '-'
            analise = _analisar_vaga_maquina(faturamento, modalidade, candidatos)
            if situacao_vaga and analise['situacao'] != situacao_vaga:
                continue
            if not _passa_filtro_maquina(analise):
                continue
            valor_item = item.total if item.total is not None else (item.valor or 0)
            grid_linhas.append({
                'faturamento': faturamento,
                'procedimento': item.servico or '-',
                'modalidade': modalidade,
                'valor': valor_item,
                'valor_fmt': _moeda_br(valor_item),
                **analise,
            })

    # Resumos só com linhas já filtradas (período, convênio, status, máquina, vaga…)
    resumo_parada = _resumo_maquinas_paradas(grid_linhas)
    resumo_reutilizado = _resumo_horarios_reutilizados(grid_linhas)

    # Marca na grade: vaga = 1º exame do slot cancelado; reutilizado = exames que ocuparam o lugar
    vistos_slot = set()
    vistos_uso = set()
    for linha in grid_linhas:
        situacao = linha.get('situacao')
        linha['slot_unico'] = False
        linha['minutos_slot'] = 0
        linha['minutos_slot_fmt'] = ''
        if situacao == 'MAQUINA_VAGA':
            chave = _chave_slot_paciente_maquina(linha)
            if not chave or chave in vistos_slot:
                continue
            vistos_slot.add(chave)
            fat = linha['faturamento']
            ini, fim = _horario_slot(fat)
            linha['slot_unico'] = True
            linha['minutos_slot'] = _duracao_slot_minutos(ini, fim)
            linha['minutos_slot_fmt'] = _fmt_minutos_hhmm(linha['minutos_slot'])
            linha['minutos_parada'] = linha['minutos_slot']
            linha['minutos_parada_fmt'] = linha['minutos_slot_fmt']
            continue
        if situacao != 'REUTILIZADA':
            continue
        # Só marca tempo com base nos exames que usaram o horário (não nos cancelados)
        minutos_novos = 0
        for uso in linha.get('reutilizacoes') or []:
            id_uso = uso.get('id')
            if id_uso is None or id_uso in vistos_uso:
                continue
            vistos_uso.add(id_uso)
            fat_uso = uso.get('faturamento')
            if fat_uso is not None:
                ini, fim = _horario_slot(fat_uso)
            else:
                ht = (uso.get('horario') or '').strip()
                ini = _parse_hora_minutos(ht.split(' - ')[0] if ' - ' in ht else ht)
                fim = _parse_hora_minutos(ht.split(' - ')[1]) if ' - ' in ht else None
            minutos_novos += _duracao_slot_minutos(ini, fim)
        if minutos_novos > 0:
            linha['slot_unico'] = True
            linha['minutos_slot'] = minutos_novos
            linha['minutos_slot_fmt'] = _fmt_minutos_hhmm(minutos_novos)

    convenios_disponiveis = []
    if empresa_id:
        from servicos_medicos.models import Convenio
        convenios_disponiveis = list(Convenio.objects.filter(empresa_id=empresa_id).order_by('nome'))

    # Opções de máquina: mapa fixo + códigos presentes no resultado filtrado
    maquinas_opcoes = []
    vistos = set()
    for chave, nome_maq in MAQUINAS_POR_MODALIDADE.values():
        if chave not in vistos:
            vistos.add(chave)
            maquinas_opcoes.append({'codigo': chave, 'nome': nome_maq})
    for linha in grid_linhas:
        cod = (linha.get('maquina_codigo') or '').strip().upper()
        if not cod or cod in vistos:
            continue
        vistos.add(cod)
        maquinas_opcoes.append({
            'codigo': cod,
            'nome': linha.get('maquina') or f'Máquina {cod}',
        })
    maquinas_opcoes.sort(key=lambda m: m['nome'])

    # Querystring limpa para badges (evita datas duplicadas/corrompidas na URL)
    params_filtro = {
        'data_inicio': data_inicio,
        'data_fim': data_fim,
    }
    if nome:
        params_filtro['nome'] = nome
    if status_agendamento:
        params_filtro['status_agendamento'] = status_agendamento
    if situacao_vaga:
        params_filtro['situacao_vaga'] = situacao_vaga
    convenios_sel = [c for c in convenios if c]
    if convenios_sel:
        params_filtro['convenio'] = convenios_sel
    qs_sem_maquina = urlencode(params_filtro, doseq=True)

    partes_filtro = [
        f'{di.strftime("%d/%m/%Y")} → {df.strftime("%d/%m/%Y")}',
    ]
    if nome:
        partes_filtro.append(f'paciente: {nome}')
    if status_agendamento:
        partes_filtro.append(f'status: {status_agendamento}')
    if maquina:
        nome_m = next((m['nome'] for m in maquinas_opcoes if m['codigo'] == maquina), maquina)
        partes_filtro.append(f'máquina: {nome_m}')
    if situacao_vaga == 'MAQUINA_VAGA':
        partes_filtro.append('só máquina vaga')
    elif situacao_vaga == 'REUTILIZADA':
        partes_filtro.append('só horário reutilizado')
    if convenios_sel:
        if len(convenios_sel) <= 2:
            partes_filtro.append('convênio: ' + ', '.join(convenios_sel))
        else:
            partes_filtro.append(f'{len(convenios_sel)} convênios')
    filtros_rotulo = ' · '.join(partes_filtro)

    context = {
        'grid_linhas': grid_linhas,
        'total_procedimentos': len(grid_linhas),
        'total_maquina_vaga': resumo_parada['total_slots'],
        'total_reutilizadas': resumo_reutilizado['total_slots'],
        'resumo_parada': resumo_parada,
        'resumo_reutilizado': resumo_reutilizado,
        'valor_total': sum((linha['valor'] or 0) for linha in grid_linhas),
        'valor_total_fmt': _moeda_br(sum((linha['valor'] or 0) for linha in grid_linhas)),
        'convenios_disponiveis': convenios_disponiveis,
        'status_opcoes': ['Cancelado', 'Desistência', 'Deletado'],
        'maquinas_opcoes': maquinas_opcoes,
        'qs_sem_maquina': qs_sem_maquina,
        'filtros_rotulo': filtros_rotulo,
        'filtros': {
            'nome': nome,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'convenio': convenios,
            'status_agendamento': status_agendamento or '',
            'situacao_vaga': situacao_vaga,
            'maquina': maquina,
        },
        'dias_regua': _montar_dias_regua(regua_ini, regua_fim),
        'regua_min': regua_ini.isoformat(),
        'regua_max': regua_fim.isoformat(),
        'periodo_inicio_fmt': di.strftime('%d/%m/%Y'),
        'periodo_fim_fmt': df.strftime('%d/%m/%Y'),
    }
    return render(request, 'faturamento_medico/listar_cancelados.html', context)


def listar_exames_por_solicitante(request):
    """Relatório de exames agrupados por médico solicitante."""
    if request.method == 'POST':
        return _salvar_metas_modalidade_solicitante(request)

    empresa_id = request.session.get('empresa_id')
    if empresa_id:
        qs_base = FaturamentoMedico.objects.filter(empresa_id=empresa_id)
    else:
        qs_base = FaturamentoMedico.objects.all()

    hoje = date.today()
    di_padrao, df_padrao = _periodo_filtro_padrao(hoje)
    di = _parse_data_filtro(request.GET.get('data_inicio')) or di_padrao
    df = _parse_data_filtro(request.GET.get('data_fim')) or df_padrao
    if di > df:
        di, df = df, di

    solicitantes_sel = [s.strip() for s in request.GET.getlist('solicitante') if s and str(s).strip()]
    status_agendamento_sel = [
        s.strip() for s in request.GET.getlist('status_agendamento') if s and str(s).strip()
    ]
    qs_periodo = qs_base.filter(data__gte=di, data__lte=df)

    status_disponiveis = sorted({
        (status or '').strip() or 'Não informado'
        for status in qs_periodo.values_list('status_agendamento', flat=True).distinct()
    }, key=str.lower)

    mapa_solicitante, grupos_solicitante = _construir_grupos_solicitante(
        qs_periodo.values_list('medico_solicitante', flat=True).distinct()
    )
    solicitantes_disponiveis = [
        info for _, info in sorted(grupos_solicitante.items(), key=lambda x: x[0].lower())
    ]

    qs = _filtrar_por_status_agendamento(qs_periodo, status_agendamento_sel)
    qs = _filtrar_por_solicitantes(qs, solicitantes_sel, grupos_solicitante)

    qs = qs.order_by('-data', 'nome').prefetch_related('itens_servico')
    codigos_modalidade = [codigo for codigo, _ in MODALIDADES_SOLICITANTE]
    labels_modalidade = dict(MODALIDADES_SOLICITANTE)
    periodo_multimes = _periodo_abrange_mais_de_um_mes(di, df)
    metas_map = _carregar_metas_solicitante(empresa_id)

    grid_linhas = []
    cards_map = defaultdict(lambda: _novo_resumo_solicitante(codigos_modalidade, periodo_multimes))

    for faturamento in qs:
        solicitante = _canonico_solicitante(faturamento.medico_solicitante, mapa_solicitante)
        itens = list(faturamento.itens_servico.all())

        def _registrar_linha(procedimento, modalidade, valor, item=None):
            status_label, status_css = _status_linha_faturamento(faturamento, item)
            status_ag_label, status_ag_css = _badge_status_agendamento(faturamento.status_agendamento)
            grid_linhas.append({
                'data': faturamento.data,
                'data_fmt': faturamento.data.strftime('%d/%m/%Y') if faturamento.data else '-',
                'paciente': faturamento.nome or '-',
                'procedimento': procedimento,
                'modalidade': modalidade or '-',
                'status': status_label,
                'status_css': status_css,
                'status_agendamento': status_ag_label,
                'status_agendamento_css': status_ag_css,
                'valor': valor,
                'valor_fmt': _moeda_br(valor),
                'solicitante': solicitante,
                'convenio': faturamento.convenio or '-',
            })
            card = cards_map[solicitante]
            _acumular_modalidade_resumo(card, modalidade, valor, codigos_modalidade)
            if periodo_multimes and faturamento.data:
                chave_mes = (faturamento.data.year, faturamento.data.month)
                _acumular_modalidade_resumo(card['meses'][chave_mes], modalidade, valor, codigos_modalidade)

        if not itens:
            _registrar_linha(
                faturamento.servico or '-',
                _modalidade_faturamento_item(faturamento),
                faturamento.total or 0,
            )
            continue
        for item in itens:
            valor_item = item.total if item.total is not None else (item.valor or 0)
            _registrar_linha(
                item.servico or '-',
                _modalidade_faturamento_item(faturamento, item),
                valor_item,
                item,
            )

    cards_resumo = []
    meses_periodo = _meses_no_periodo(di, df) if periodo_multimes else [(di.year, di.month)]
    for nome, dados in sorted(cards_map.items(), key=lambda x: (-x[1]['total'], x[0].lower())):
        card = {
            'nome': nome,
            'total': dados['total'],
            'valor_fmt': _moeda_br(dados['valor']),
            'variantes': grupos_solicitante.get(nome, {}).get('variantes', []),
            'qtd_variantes': grupos_solicitante.get(nome, {}).get('qtd_variantes', 0),
            'periodo_multimes': periodo_multimes,
            'metas_form': _metas_form_solicitante(nome, metas_map, grupos_solicitante),
            'metas_ref': _metas_solicitante_grupo(metas_map, nome, grupos_solicitante),
        }
        if periodo_multimes:
            meses_card = []
            metas_ref = card['metas_ref']
            for ano, mes in meses_periodo:
                dados_mes = dados['meses'].get((ano, mes))
                if not dados_mes and not metas_ref:
                    continue
                if not dados_mes:
                    dados_mes = {
                        'total': 0,
                        'valor': Decimal('0'),
                        'modalidades': {codigo: 0 for codigo in codigos_modalidade},
                        'outros': 0,
                    }
                realizado_mes = _realizado_por_modalidade(dados_mes)
                modalidades_mes = _montar_modalidades_card(
                    dados_mes, codigos_modalidade, labels_modalidade,
                )
                _enriquecer_modalidades_com_meta(modalidades_mes, realizado_mes, metas_ref)
                metas_list_mes = _lista_metas_from_dict(metas_ref, realizado_mes)
                meses_card.append({
                    'label': _rotulo_mes_ano(ano, mes),
                    'ano': ano,
                    'mes': mes,
                    'total': dados_mes['total'],
                    'valor_fmt': _moeda_br(dados_mes['valor']),
                    'modalidades': modalidades_mes,
                    'metas': metas_list_mes,
                    'resumo_metas': _resumo_metas_solicitante(metas_list_mes),
                })
            card['meses'] = meses_card
            card['modalidades'] = []
            card['metas'] = []
            card['resumo_metas'] = None
        else:
            ano, mes = di.year, di.month
            metas_ref = card['metas_ref']
            realizado_map = _realizado_por_modalidade(dados)
            modalidades = _montar_modalidades_card(
                dados, codigos_modalidade, labels_modalidade,
            )
            _enriquecer_modalidades_com_meta(modalidades, realizado_map, metas_ref)
            metas_list = _lista_metas_from_dict(metas_ref, realizado_map)
            card['modalidades'] = modalidades
            card['meses'] = []
            card['metas'] = metas_list
            card['resumo_metas'] = _resumo_metas_solicitante(metas_list)
            card['mes_label'] = _rotulo_mes_ano(ano, mes)
        cards_resumo.append(card)

    cards_resumo.sort(key=lambda c: (-c['total'], c['nome'].lower()))

    totais_solicitante = {card['nome']: card['total'] for card in cards_resumo}
    grid_linhas.sort(key=lambda linha: (
        -totais_solicitante.get(linha['solicitante'], 0),
        -(linha['data'].toordinal() if linha['data'] else 0),
        linha['paciente'].lower(),
    ))

    valor_total = sum((linha.get('valor') or 0) for linha in grid_linhas)
    context = {
        'grid_linhas': grid_linhas,
        'cards_resumo': cards_resumo,
        'total_exames': len(grid_linhas),
        'valor_total_fmt': _moeda_br(valor_total),
        'solicitantes_disponiveis': solicitantes_disponiveis,
        'status_disponiveis': status_disponiveis,
        'filtros': {
            'data_inicio': di.isoformat(),
            'data_fim': df.isoformat(),
            'solicitante': solicitantes_sel,
            'status_agendamento': status_agendamento_sel,
        },
        'periodo_fmt': f'{di.strftime("%d/%m/%Y")} → {df.strftime("%d/%m/%Y")}',
        'periodo_multimes': periodo_multimes,
        'metas_modalidades_opcoes': METAS_MODALIDADES_SOLICITANTE,
        'redirect_qs': request.GET.urlencode(),
    }
    return render(request, 'faturamento_medico/listar_exames_por_solicitante.html', context)


FUZZY_NOME_MIN_RATIO = 0.90


def _normalizar_nome_servico(texto):
    """Uppercase, sem acentos e espaços extras — para comparar nomes de procedimento."""
    if not texto:
        return ''
    t = unicodedata.normalize('NFKD', str(texto))
    t = ''.join(c for c in t if not unicodedata.combining(c))
    t = re.sub(r'\s+', ' ', t.upper().strip())
    return t


def _nome_base_procedimento(texto):
    """Remove complemento entre parênteses (descrição longa no faturamento)."""
    base = _normalizar_nome_servico(texto)
    if '(' in base:
        base = base.split('(', 1)[0].strip()
    return base


def _similaridade_nome_procedimento(a, b):
    """
    Similaridade 0–1 entre dois nomes de procedimento.
    Compara texto completo e nome base (antes de parênteses).
    """
    candidatos_a = {_normalizar_nome_servico(a), _nome_base_procedimento(a)}
    candidatos_b = {_normalizar_nome_servico(b), _nome_base_procedimento(b)}
    candidatos_a.discard('')
    candidatos_b.discard('')
    if not candidatos_a or not candidatos_b:
        return 0.0
    melhor = 0.0
    for na in candidatos_a:
        for nb in candidatos_b:
            if na == nb:
                return 1.0
            melhor = max(melhor, SequenceMatcher(None, na, nb).ratio())
    return melhor


def _pool_precos_convenio(empresa_id, convenio_nome):
    """Carrega tabela de preços do convênio uma vez (lookup por código/nome)."""
    from servicos_medicos.models import TabelaPreco, Convenio

    conv_key = (convenio_nome or '').strip().upper()
    pool = {
        'tabelas': [],
        'por_codigo': {},
        'por_nome': {},
        'por_base': {},
    }
    if not empresa_id or not conv_key:
        return pool

    convenios = list(
        Convenio.objects.filter(empresa_id=empresa_id).filter(
            Q(nome__iexact=convenio_nome) | Q(nome__icontains=convenio_nome)
        )
    )
    if not convenios:
        convenios = [
            c for c in Convenio.objects.filter(empresa_id=empresa_id)
            if c.nome and c.nome.upper() in conv_key
        ]
    if not convenios:
        return pool

    tabelas = list(
        TabelaPreco.objects.filter(
            empresa_id=empresa_id,
            convenio__in=convenios,
        ).select_related('codigo_servico')
    )
    pool['tabelas'] = tabelas
    for tabela in tabelas:
        cod = (tabela.codigo_servico.codigo or '').strip().upper()
        if cod:
            pool['por_codigo'].setdefault(cod, tabela)
        nome = tabela.codigo_servico.servicos or ''
        nome_norm = _normalizar_nome_servico(nome)
        if nome_norm:
            pool['por_nome'].setdefault(nome_norm, tabela)
        nome_base = _nome_base_procedimento(nome)
        if nome_base:
            pool['por_base'].setdefault(nome_base, tabela)
    return pool


def _get_pool_precos(cache_precos, empresa_id, convenio_nome):
    conv_key = (convenio_nome or '').strip().upper()
    pool_key = ('__pool__', empresa_id, conv_key)
    if pool_key not in cache_precos:
        cache_precos[pool_key] = _pool_precos_convenio(empresa_id, convenio_nome)
    return cache_precos[pool_key]


def _buscar_tabela_por_nome_proximo(pool, descricao, min_ratio=FUZZY_NOME_MIN_RATIO):
    """Melhor TabelaPreco cujo nome do serviço atinge similaridade mínima."""
    if not (descricao or '').strip():
        return None, 0.0

    desc_norm = _normalizar_nome_servico(descricao)
    desc_base = _nome_base_procedimento(descricao)
    if desc_norm in pool['por_nome']:
        return pool['por_nome'][desc_norm], 1.0
    if desc_base in pool['por_base']:
        return pool['por_base'][desc_base], 1.0

    tokens = [t for t in re.split(r'\W+', desc_norm) if len(t) >= 4][:4]
    candidatos = pool['tabelas']
    if tokens:
        filtrados = []
        for tabela in candidatos:
            nome = _normalizar_nome_servico(tabela.codigo_servico.servicos or '')
            if any(tok in nome for tok in tokens):
                filtrados.append(tabela)
        if filtrados:
            candidatos = filtrados[:120]

    melhor_tabela = None
    melhor_ratio = 0.0
    for tabela in candidatos:
        nome = tabela.codigo_servico.servicos or ''
        ratio = _similaridade_nome_procedimento(descricao, nome)
        if ratio > melhor_ratio:
            melhor_ratio = ratio
            melhor_tabela = tabela
    if melhor_tabela is not None and melhor_ratio >= min_ratio:
        return melhor_tabela, melhor_ratio
    return None, melhor_ratio


def _buscar_tabela_por_nome_proximo_qs(qs, descricao, min_ratio=FUZZY_NOME_MIN_RATIO):
    """Compatibilidade: usa pool montado a partir de queryset."""
    pool = {
        'tabelas': list(qs.select_related('codigo_servico')),
        'por_codigo': {},
        'por_nome': {},
        'por_base': {},
    }
    for tabela in pool['tabelas']:
        cod = (tabela.codigo_servico.codigo or '').strip().upper()
        if cod:
            pool['por_codigo'].setdefault(cod, tabela)
        nome = tabela.codigo_servico.servicos or ''
        nome_norm = _normalizar_nome_servico(nome)
        if nome_norm:
            pool['por_nome'].setdefault(nome_norm, tabela)
        nome_base = _nome_base_procedimento(nome)
        if nome_base:
            pool['por_base'].setdefault(nome_base, tabela)
    return _buscar_tabela_por_nome_proximo(pool, descricao, min_ratio)


def _item_usa_contraste(com_contraste, descricao_servico, tipo_acomodacao=None):
    """
    Define se o preço da tabela deve ser o com contraste (preco_enfermaria).
    Prioridade: flag do item > texto do procedimento > legado apartamento/enfermaria.
    """
    if com_contraste is not None:
        return bool(com_contraste)
    desc = (descricao_servico or '').lower()
    if any(m in desc for m in ('com contraste', 'c/ contraste', 'c/contraste')):
        return True
    if any(m in desc for m in ('sem contraste', 's/ contraste', 's/contraste')):
        return False
    tipo = (tipo_acomodacao or '').strip().lower()
    if tipo == 'apartamento':
        return False
    if tipo == 'enfermaria':
        return True
    return False


def _preco_tabela_para_item(tabela, com_contraste, descricao_servico, tipo_acomodacao=None):
    usa_contraste = _item_usa_contraste(com_contraste, descricao_servico, tipo_acomodacao)
    return tabela.preco_enfermaria if usa_contraste else tabela.preco_apartamento


def _resolver_preco_tabela(
    empresa_id,
    convenio_nome,
    codigo_servico,
    descricao_servico,
    tipo_acomodacao,
    cache_precos,
    *,
    com_contraste=None,
):
    """
    Resolve preço da TabelaPreco para um item.
    preco_apartamento = sem contraste; preco_enfermaria = com contraste.
    Retorna (preco Decimal|None, codigo_encontrado, descricao_encontrada).
    """
    from servicos_medicos.models import TabelaPreco

    conv_key = (convenio_nome or '').strip().upper()
    cod = (codigo_servico or '').strip().upper()
    desc = (descricao_servico or '').strip().upper()
    usa_contraste = _item_usa_contraste(com_contraste, descricao_servico, tipo_acomodacao)
    cache_key = (conv_key, cod, desc, usa_contraste)
    if cache_key in cache_precos:
        return cache_precos[cache_key]

    resultado = (None, '', '')
    if not empresa_id or not conv_key:
        cache_precos[cache_key] = resultado
        return resultado

    pool = _get_pool_precos(cache_precos, empresa_id, convenio_nome)
    if not pool['tabelas']:
        cache_precos[cache_key] = resultado
        return resultado

    tabela = None
    if cod:
        tabela = pool['por_codigo'].get(cod)
    if tabela is None and desc:
        desc_norm = _normalizar_nome_servico(descricao_servico)
        tabela = pool['por_nome'].get(desc_norm) or pool['por_base'].get(_nome_base_procedimento(descricao_servico))
    if tabela is None and desc:
        tabela, _ratio = _buscar_tabela_por_nome_proximo(pool, descricao_servico)

    if tabela is None:
        cache_precos[cache_key] = resultado
        return resultado

    preco = _preco_tabela_para_item(tabela, com_contraste, descricao_servico, tipo_acomodacao)
    resultado = (
        Decimal(str(preco or 0)),
        (tabela.codigo_servico.codigo or ''),
        (tabela.codigo_servico.servicos or ''),
    )
    cache_precos[cache_key] = resultado
    return resultado


def verificar_corrigir_precos(request):
    """
    Filtra procedimentos por período e convênio, compara valor atual x tabela
    e permite aplicar o preço correto nos itens selecionados.
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    try:
        empresa_id = int(empresa_id)
    except (TypeError, ValueError):
        messages.error(request, 'Empresa inválida na sessão.')
        return redirect('faturamento_medico:ftlistar')

    from servicos_medicos.models import Convenio

    convenios_disponiveis = list(
        Convenio.objects.filter(empresa_id=empresa_id).order_by('nome')
    )

    hoje = date.today()
    data_inicio = request.GET.get('data_inicio') or request.POST.get('data_inicio')
    data_fim = request.GET.get('data_fim') or request.POST.get('data_fim')
    convenio = (request.GET.get('convenio') or request.POST.get('convenio') or '').strip()
    procedimento = (request.GET.get('procedimento') or request.POST.get('procedimento') or '').strip()
    status_conferencia = (request.GET.get('status_conferencia') or request.POST.get('status_conferencia') or '').strip()
    so_divergentes = (request.GET.get('so_divergentes') or request.POST.get('so_divergentes') or '') == '1'

    if not data_inicio:
        data_inicio = hoje.replace(day=1).strftime('%Y-%m-%d')
    if not data_fim:
        proximo_mes = hoje.replace(day=28) + timedelta(days=4)
        data_fim = (proximo_mes - timedelta(days=proximo_mes.day)).strftime('%Y-%m-%d')

    # Aplicar preços selecionados
    if request.method == 'POST' and request.POST.get('acao') == 'aplicar':
        ids = request.POST.getlist('itens_selecionados')
        if not ids:
            messages.warning(request, 'Selecione ao menos um procedimento para corrigir.')
        else:
            cache_precos = {}
            corrigidos = 0
            sem_preco = 0
            faturamentos_atualizar = set()
            itens = (
                ItemServico.objects
                .select_related('faturamento')
                .filter(pk__in=ids, faturamento__empresa_id=empresa_id)
            )
            for item in itens:
                fat = item.faturamento
                preco, cod_tab, _desc = _resolver_preco_tabela(
                    empresa_id,
                    fat.convenio,
                    item.codigo_servico,
                    item.servico,
                    fat.apartamento_enfermaria,
                    cache_precos,
                    com_contraste=item.com_contraste,
                )
                if preco is None:
                    sem_preco += 1
                    continue
                item.valor = preco
                if cod_tab and not (item.codigo_servico or '').strip():
                    item.codigo_servico = cod_tab
                item.save()
                faturamentos_atualizar.add(fat.id)
                corrigidos += 1

            for fat_id in faturamentos_atualizar:
                try:
                    FaturamentoMedico.objects.get(pk=fat_id).atualizar_total()
                except FaturamentoMedico.DoesNotExist:
                    pass

            if corrigidos:
                messages.success(request, f'{corrigidos} procedimento(s) atualizado(s) com o preço da tabela.')
            if sem_preco:
                messages.warning(request, f'{sem_preco} item(ns) sem preço correspondente na tabela.')

        qs = urlencode({
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'convenio': convenio,
            'procedimento': procedimento,
            'status_conferencia': status_conferencia,
            'so_divergentes': '1' if so_divergentes else '',
        })
        return redirect(f"{reverse('faturamento_medico:verificar_corrigir_precos')}?{qs}")

    linhas = []
    resumo = {
        'total': 0,
        'ok': 0,
        'divergente': 0,
        'sem_preco': 0,
        'valor_atual': Decimal('0'),
        'valor_tabela': Decimal('0'),
    }

    if convenio:
        faturamentos = (
            FaturamentoMedico.objects
            .filter(empresa_id=empresa_id)
            .exclude(_q_status_agendamento_cancelados())
            .filter(data__gte=data_inicio, data__lte=data_fim)
            .filter(_q_convenio_filtro(convenio))
            .prefetch_related('itens_servico')
            .order_by('data', 'nome')
        )

        cache_precos = {}
        for fat in faturamentos:
            for item in fat.itens_servico.all():
                if procedimento and procedimento.lower() not in (item.servico or '').lower():
                    continue

                status_label, _status_css = item.status_conferencia_badge()
                if status_conferencia and status_label != status_conferencia:
                    continue

                valor_atual = Decimal(str(item.valor or 0))
                preco, cod_tab, desc_tab = _resolver_preco_tabela(
                    empresa_id,
                    fat.convenio,
                    item.codigo_servico,
                    item.servico,
                    fat.apartamento_enfermaria,
                    cache_precos,
                    com_contraste=item.com_contraste,
                )
                if preco is None:
                    situacao = 'SEM PRECO'
                    css = 'warning'
                    diferenca = None
                    valor_esperado = None
                    pode_corrigir = False
                else:
                    valor_esperado = preco
                    diferenca = valor_atual - preco
                    if abs(diferenca) < Decimal('0.01'):
                        situacao = 'OK'
                        css = 'success'
                        pode_corrigir = False
                    else:
                        situacao = 'DIVERGENTE'
                        css = 'danger'
                        pode_corrigir = True

                if so_divergentes and situacao == 'OK':
                    continue

                if situacao == 'SEM PRECO':
                    resumo['sem_preco'] += 1
                elif situacao == 'OK':
                    resumo['ok'] += 1
                else:
                    resumo['divergente'] += 1

                if valor_esperado is not None:
                    resumo['valor_tabela'] += valor_esperado

                resumo['total'] += 1
                resumo['valor_atual'] += valor_atual
                linhas.append({
                    'item': item,
                    'faturamento': fat,
                    'valor_atual': valor_atual,
                    'valor_esperado': valor_esperado,
                    'diferenca': diferenca,
                    'situacao': situacao,
                    'situacao_css': css,
                    'status_conferencia': status_label,
                    'status_conferencia_css': _status_css,
                    'pode_corrigir': pode_corrigir,
                    'codigo_tabela': cod_tab,
                    'descricao_tabela': desc_tab,
                })

    context = {
        'linhas': linhas,
        'resumo': resumo,
        'convenios_disponiveis': convenios_disponiveis,
        'filtros': {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'convenio': convenio,
            'procedimento': procedimento,
            'status_conferencia': status_conferencia,
            'so_divergentes': so_divergentes,
        },
        'status_conferencia_choices': ItemServico.STATUS_CONFERENCIA_CHOICES,
    }
    return render(request, 'faturamento_medico/verificar_corrigir_precos.html', context)


def criar_faturamento(request):
    """Cria um novo faturamento médico"""
    logger.info("Iniciando criar_faturamento")
    empresa_id = request.session.get('empresa_id')
    logger.info(f"Empresa ID da sessão: {empresa_id}")
    if not empresa_id:
        logger.warning("Empresa não encontrada na sessão")
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    if request.method == 'POST':
        logger.info("Método POST detectado")
        form = FaturamentoMedicoForm(request.POST, request.FILES, empresa_id=empresa_id)
        logger.info(f"Form criado: {form}")
        if form.is_valid():
            logger.info("Form é válido")

            # Processar arquivos com Gemini se foram enviados
            documentos_gemini = request.FILES.getlist('documentos_gemini')
            documento_upload = request.FILES.get('documento_upload')
            dados_gemini = {}

            # Processar documento_upload se enviado
            if documento_upload:
                logger.info("Processando documento_upload com Gemini")
                dados_gemini = processar_arquivos_com_gemini([documento_upload])
                logger.info(f"Dados extraídos do documento_upload: {dados_gemini}")

            # Processar documentos_gemini adicionais se enviados
            elif documentos_gemini:
                logger.info(f"Processando {len(documentos_gemini)} arquivos com Gemini")
                dados_gemini = processar_arquivos_com_gemini(documentos_gemini)
                logger.info(f"Dados extraídos do Gemini: {dados_gemini}")

            faturamento = form.save(commit=False)
            faturamento.empresa_id = empresa_id

            # Preencher campos com dados do Gemini se disponíveis
            if dados_gemini.get('nome'):
                faturamento.nome = dados_gemini['nome']
            if dados_gemini.get('carteirinha'):
                faturamento.carteirinha = dados_gemini['carteirinha']
            if dados_gemini.get('guia'):
                faturamento.guia = dados_gemini['guia']
            if dados_gemini.get('numero_guia_lancada'):
                faturamento.numero_guia_lancada = dados_gemini['numero_guia_lancada']
            if dados_gemini.get('data_autorizacao'):
                # Tentar converter data se possível
                try:
                    from datetime import datetime
                    faturamento.data_autorizacao = datetime.strptime(dados_gemini['data_autorizacao'], '%d/%m/%Y').date()
                except:
                    pass
            if dados_gemini.get('data_internacao_cirurgia'):
                # Tentar converter data se possível
                try:
                    from datetime import datetime
                    faturamento.data = datetime.strptime(dados_gemini['data_internacao_cirurgia'], '%d/%m/%Y').date()
                except:
                    pass
            if dados_gemini.get('local'):
                faturamento.local = dados_gemini['local']
            if dados_gemini.get('medico'):
                faturamento.medico = dados_gemini['medico']
            if dados_gemini.get('anestesista'):
                faturamento.anestesista = dados_gemini['anestesista']
            if dados_gemini.get('convenio'):
                faturamento.convenio = dados_gemini['convenio']
            if dados_gemini.get('apartamento_enfermaria'):
                faturamento.apartamento_enfermaria = dados_gemini['apartamento_enfermaria']
            if dados_gemini.get('urgencia'):
                faturamento.urgencia = dados_gemini['urgencia']

            faturamento.save()
            logger.info(f"Faturamento salvo: {faturamento.id}")

            # Criar itens de serviço baseados nos dados do Gemini
            if dados_gemini.get('servicos'):
                for servico in dados_gemini['servicos']:
                    ItemServico.objects.create(
                        faturamento=faturamento,
                        servico=servico.get('descricao', ''),
                        codigo_servico=servico.get('codigo', ''),
                        valor=servico.get('valor_unitario', 0),
                        qt=servico.get('quantidade', 1)
                    )

            # Anexar documentos processados
            documentos_para_anexar = []
            if documento_upload:
                documentos_para_anexar.append(documento_upload)
            documentos_para_anexar.extend(documentos_gemini)

            for documento in documentos_para_anexar:
                DocumentoAnexado.objects.create(
                    faturamento=faturamento,
                    arquivo=documento,
                    nome=f"Documento Gemini - {documento.name}",
                    descricao="Documento processado com Gemini para extração de dados"
                )

            # Adicionar mensagem de sucesso específica se Gemini foi usado
            if documentos_gemini and dados_gemini:
                messages.success(request, f'Faturamento médico criado com sucesso! {len(documentos_gemini)} documento(s) processado(s) com Gemini.')
            else:
                messages.success(request, 'Faturamento médico criado com sucesso!')
            return _redirect_ftlistar_com_filtros_sessao(request)
        else:
            logger.warning(f"Form inválido: {form.errors}")
    else:
        logger.info("Método GET detectado")
        initial_data = {
            'data_autorizacao': timezone.now().date(),
            'data': timezone.now().date(),
        }
        form = FaturamentoMedicoForm(empresa_id=empresa_id, initial=initial_data)

    context = {
        'form': form,
        'titulo': 'Criar Faturamento Médico'
    }

    logger.info("Renderizando template form.html")
    return render(request, 'faturamento_medico/form.html', context)


def editar_faturamento(request, pk):
    """Edita um faturamento médico existente"""
    faturamento = get_object_or_404(FaturamentoMedico, pk=pk)
    empresa_id = request.session.get('empresa_id')

    if request.method == 'POST':
        form = FaturamentoMedicoForm(request.POST, instance=faturamento, empresa_id=empresa_id)
        if form.is_valid():
            form.save()
            messages.success(request, 'Faturamento médico atualizado com sucesso!')
            return _redirect_ftlistar_com_filtros_sessao(request)
    else:
        form = FaturamentoMedicoForm(instance=faturamento, empresa_id=empresa_id)

    context = {
        'form': form,
        'faturamento': faturamento,
        'titulo': 'Editar Faturamento Médico'
    }

    return render(request, 'faturamento_medico/form.html', context)


def editar_documentacao_faturamento(request, pk):
    """Altera apenas protocolo, lote, guia lançada e nota fiscal."""
    empresa_id = request.session.get('empresa_id')
    qs = FaturamentoMedico.objects.all()
    if empresa_id:
        qs = qs.filter(empresa_id=empresa_id)
    faturamento = get_object_or_404(qs, pk=pk)

    voltar = (request.GET.get('next') or request.POST.get('voltar') or '').strip()
    if not voltar:
        voltar = request.META.get('HTTP_REFERER') or reverse('faturamento_medico:ftlistar')

    if request.method == 'POST':
        form = FaturamentoDocumentacaoForm(request.POST, instance=faturamento)
        if form.is_valid():
            form.save()
            messages.success(request, 'Documentação atualizada com sucesso!')
            if voltar.startswith('/'):
                return redirect(voltar)
            return _redirect_ftlistar_com_filtros_sessao(request)
    else:
        form = FaturamentoDocumentacaoForm(instance=faturamento)

    context = {
        'form': form,
        'faturamento': faturamento,
        'titulo': 'Documentação e fechamento',
        'voltar': voltar,
    }
    return render(request, 'faturamento_medico/editar_documentacao.html', context)


def excluir_faturamento(request, pk):
    """Exclui um faturamento médico e seus itens vinculados."""
    empresa_id = request.session.get('empresa_id')
    qs = FaturamentoMedico.objects.all()
    if empresa_id:
        qs = qs.filter(empresa_id=empresa_id)
    faturamento = get_object_or_404(qs, pk=pk)

    if request.method == 'POST':
        nome = faturamento.nome or f'#{faturamento.pk}'
        # Itens e documentos são removidos por CASCADE
        faturamento.delete()
        messages.success(request, f'Faturamento médico "{nome}" excluído com sucesso!')
        return _redirect_ftlistar_com_filtros_sessao(request)

    context = {
        'faturamento': faturamento,
        'titulo': 'Confirmar Exclusão',
    }
    return render(request, 'faturamento_medico/confirmar_exclusao.html', context)


def detalhes_faturamento(request, pk):
    """Exibe detalhes de um faturamento médico"""
    faturamento = get_object_or_404(FaturamentoMedico, pk=pk)

    context = {
        'faturamento': faturamento,
    }

    return render(request, 'faturamento_medico/detalhes.html', context)


def exportar_excel(request):
    """Exporta faturamentos filtrados para Excel (conferência: 1 linha por procedimento)."""
    empresa_id = request.session.get('empresa_id')
    if empresa_id:
        faturamentos = (
            FaturamentoMedico.objects
            .filter(empresa_id=empresa_id)
            .prefetch_related('itens_servico')
            .order_by('-data', 'nome')
        )
    else:
        faturamentos = FaturamentoMedico.objects.none()

    filtros = _filtros_listagem_faturamento(request, use_session_fallback=True)
    faturamentos = _aplicar_filtros_faturamento_qs(faturamentos, filtros)
    nome = filtros['nome']
    guia = filtros['guia']
    anestesista = filtros['anestesista']
    status = filtros['status']
    status_conferencia = filtros['status_conferencia']
    lote = filtros['lote']
    data_inicio = filtros['data_inicio']
    data_fim = filtros['data_fim']
    convenios = filtros['convenios']
    codigo_relatorio = filtros['codigo_relatorio']

    # Cache preços para status de conferência
    precos_por_codigo = set()
    precos_por_descricao = set()
    if empresa_id:
        from servicos_medicos.models import TabelaPreco
        for t in (
            TabelaPreco.objects
            .filter(empresa_id=empresa_id)
            .select_related('codigo_servico', 'convenio')
        ):
            conv = (t.convenio.nome or '').strip().upper()
            cod = (t.codigo_servico.codigo or '').strip().upper()
            desc = (t.codigo_servico.servicos or '').strip().upper()
            if conv and cod:
                precos_por_codigo.add((conv, cod))
            if conv and desc:
                precos_por_descricao.add((conv, desc))

    def _tem_preco(faturamento, item):
        if not item:
            return False
        valor = item.total if item.total is not None else item.valor
        if valor is None or valor == 0:
            return False
        conv = (faturamento.convenio or '').strip().upper()
        cod = (item.codigo_servico or '').strip().upper()
        desc = (item.servico or '').strip().upper()
        if cod and (conv, cod) in precos_por_codigo:
            return True
        if desc and (conv, desc) in precos_por_descricao:
            return True
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Conferencia"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B4B7C", end_color="2B4B7C", fill_type="solid")
    empresa_font = Font(bold=True, size=14)
    empresa_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

    if empresa_id:
        try:
            empresa = Empresa.objects.get(id=empresa_id)
            empresa_info = f"{empresa.razao} - CNPJ: {empresa.cnpj}"
        except Empresa.DoesNotExist:
            empresa_info = f"Empresa ID: {empresa_id}"
    else:
        empresa_info = "Empresa não identificada"

    ws.cell(row=1, column=1).value = empresa_info
    ws.cell(row=1, column=1).font = empresa_font
    ws.cell(row=1, column=1).fill = empresa_fill

    ws.cell(row=3, column=1).value = "RELATÓRIO DE CONFERÊNCIA - FATURAMENTO MÉDICO"
    ws.cell(row=3, column=1).font = Font(bold=True, size=16)
    ws.cell(row=4, column=1).value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws.cell(row=4, column=1).font = Font(italic=True)

    filtros_excel = []
    if nome:
        filtros_excel.append(f"Nome: {nome}")
    if guia:
        filtros_excel.append(f"Guia: {guia}")
    if codigo_relatorio:
        filtros_excel.append(f"Código Relatório: {codigo_relatorio}")
    if anestesista:
        filtros_excel.append(f"Anestesista: {anestesista}")
    if data_inicio:
        filtros_excel.append(f"Data início: {data_inicio}")
    if data_fim:
        filtros_excel.append(f"Data fim: {data_fim}")
    if convenios:
        filtros_excel.append(f"Convênios: {', '.join(convenios)}")
    if status:
        filtros_excel.append(f"Status: {status}")
    if status_conferencia:
        filtros_excel.append(f"Status Conferência: {status_conferencia}")
    if lote == '__sem__':
        filtros_excel.append("Lote: Sem lote")
    elif lote:
        filtros_excel.append(f"Lote: {lote}")
    ws.cell(row=5, column=1).value = (
        "Filtros: " + ("; ".join(filtros_excel) if filtros_excel else "nenhum")
    )
    ws.cell(row=5, column=1).font = Font(italic=True)

    # Colunas alinhadas ao modelo RIS/agenda + campos do banco para conferência
    headers = [
        # Conferência (grid)
        'ID Faturamento',
        'ID Item',
        'Data',
        'Paciente',
        'Nome Associado',
        'Procedimento',
        'Modalidade',
        'Com Contraste',
        'Valor Item',
        'QT',
        'Percentual',
        'Total Item',
        'Conferido',
        'Status Conferencia',
        # Origem agenda / faturamento
        'Unidade/Local',
        'Carteirinha (CNS)',
        'CPF',
        'Prioridade',
        'Horario Inicio',
        'Horario Fim',
        'Horario',
        'Status Agendamento',
        'Motivo Cancelamento/Desistencia/Delecao',
        'Convenio (Viabilidade)',
        'Tag',
        'Agendado Via',
        'Lote',
        'Guia',
        'Codigo Servico Item',
        'Porte Item',
        'Medico',
        'Medico Solicitante',
        'Tecnico',
        'Check-in Por',
        'Agendado Por',
        'Anestesista',
        'Receber Por',
        'Apartamento/Enfermaria',
        'Urgencia',
        'Indicacao Clinica',
        'Descricao',
        'Data Autorizacao',
        'Guia Lancada',
        'Numero Guia Lancada',
        'Nota Fiscal',
        'Codigo Relatorio',
        'Status Faturamento',
        'Codigo Fechamento',
        'Data Fechamento',
        'Percentual Imposto',
        'Valor Imposto',
        'Percentual Comissao',
        'Valor Comissao',
        'Total Faturamento',
        'Observacao',
        'Data Criacao',
        'Data Atualizacao',
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    row_num = 8
    for faturamento in faturamentos:
        itens = list(faturamento.itens_servico.all())
        if not itens:
            status_label, _ = _status_linha_faturamento(faturamento)
            if status_conferencia and status_label != status_conferencia:
                continue
            itens = [None]
        else:
            itens_filtrados = []
            for item in itens:
                status_label, _ = _status_linha_faturamento(faturamento, item)
                if status_conferencia and status_label != status_conferencia:
                    continue
                itens_filtrados.append(item)
            if not itens_filtrados:
                continue
            itens = itens_filtrados

        for item in itens:
            if item is not None:
                status_label, _ = item.status_conferencia_badge()
                conferido = 'Sim' if item.conferido else 'Não'
                procedimento = item.servico or ''
                modalidade = item.modalidade or ''
                com_contraste = 'Sim' if item.com_contraste else 'Não'
                valor_item = float(item.valor) if item.valor is not None else 0
                qt = item.qt or 1
                percentual = float(item.percentual) if item.percentual is not None else 1
                total_item = float(item.total) if item.total is not None else 0
                codigo_item = item.codigo_servico or ''
                porte_item = item.porte or ''
                item_id = item.id
            else:
                if not (faturamento.guia or '').strip():
                    status_label = 'FALTA DE GUIA'
                elif not faturamento.total:
                    status_label = 'FALTA DE VALOR NA TABELA'
                else:
                    status_label = 'PENDENTE'
                conferido = 'Não'
                procedimento = faturamento.servico or ''
                modalidade = ''
                com_contraste = 'Sim' if 'contraste' in (procedimento or '').lower() else 'Não'
                valor_item = float(faturamento.valor) if faturamento.valor else 0
                qt = faturamento.qt or 1
                percentual = 1
                total_item = float(faturamento.total) if faturamento.total else 0
                codigo_item = faturamento.codigo_servico or ''
                porte_item = faturamento.porte or ''
                item_id = ''

            valores = [
                faturamento.id,
                item_id,
                faturamento.data.strftime('%d/%m/%Y') if faturamento.data else '',
                faturamento.nome or '',
                faturamento.nome_associado or faturamento.nome or '',
                procedimento,
                modalidade,
                com_contraste,
                valor_item,
                qt,
                percentual,
                total_item,
                conferido,
                status_label,
                faturamento.local or '',
                faturamento.carteirinha or '',
                faturamento.cpf or '',
                faturamento.prioridade or '',
                faturamento.horario_inicio or '',
                faturamento.horario_fim or '',
                faturamento.horario or '',
                faturamento.status_agendamento or '',
                faturamento.motivo_cancelamento or '',
                faturamento.convenio or '',
                faturamento.tag or '',
                faturamento.agendado_via or '',
                faturamento.lote or '',
                faturamento.guia or '',
                codigo_item,
                porte_item,
                faturamento.medico or '',
                faturamento.medico_solicitante or '',
                faturamento.tecnico or '',
                faturamento.checkin_por or '',
                faturamento.agendado_por or '',
                faturamento.anestesista or '',
                faturamento.receber_por or '',
                faturamento.apartamento_enfermaria or '',
                faturamento.urgencia or '',
                faturamento.indicacao_clinica or '',
                faturamento.descricao or '',
                faturamento.data_autorizacao.strftime('%d/%m/%Y') if faturamento.data_autorizacao else '',
                faturamento.guia_lancada or '',
                faturamento.numero_guia_lancada or '',
                faturamento.nota_fiscal or '',
                faturamento.codigo_relatorio or '',
                faturamento.status or '',
                faturamento.codigo_fechamento or '',
                faturamento.data_fechamento.strftime('%d/%m/%Y') if faturamento.data_fechamento else '',
                float(faturamento.percentual_imposto or 0),
                float(faturamento.valor_imposto or 0),
                float(faturamento.percentual_comissao or 0),
                float(faturamento.valor_comissao or 0),
                float(faturamento.total) if faturamento.total else 0,
                faturamento.observacao or '',
                faturamento.data_criacao.strftime('%d/%m/%Y %H:%M') if faturamento.data_criacao else '',
                faturamento.data_atualizacao.strftime('%d/%m/%Y %H:%M') if faturamento.data_atualizacao else '',
            ]

            for col_num, valor in enumerate(valores, 1):
                ws.cell(row=row_num, column=col_num).value = valor
            row_num += 1

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename=conferencia_faturamento_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    wb.save(response)
    return response


def anexar_documento(request, pk):
    """View para anexar documentos a um faturamento"""
    faturamento = get_object_or_404(FaturamentoMedico, pk=pk)

    if request.method == 'POST':
        form = DocumentoAnexadoForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.save(commit=False)
            documento.faturamento = faturamento
            documento.save()
            messages.success(request, 'Documento anexado com sucesso!')
            return redirect('faturamento_medico:detalhes', pk=faturamento.pk)
    else:
        form = DocumentoAnexadoForm()

    context = {
        'form': form,
        'faturamento': faturamento,
        'titulo': f'Anexar Documento - {faturamento}'
    }

    return render(request, 'faturamento_medico/anexar_documento.html', context)


def download_documento(request, pk):
    """View para fazer download/visualizar de um documento anexado"""
    documento = get_object_or_404(DocumentoAnexado, pk=pk)

    try:
        with open(documento.arquivo.path, 'rb') as f:
            # Determinar o content_type baseado na extensão
            extensao = documento.arquivo.name.split('.')[-1].lower()
            if extensao == 'pdf':
                content_type = 'application/pdf'
            elif extensao in ['jpg', 'jpeg']:
                content_type = 'image/jpeg'
            elif extensao == 'png':
                content_type = 'image/png'
            elif extensao == 'gif':
                content_type = 'image/gif'
            else:
                content_type = 'application/octet-stream'

            response = HttpResponse(f.read(), content_type=content_type)
            # Se for para visualização inline (ex: PDFs no modal), usar inline
            if request.GET.get('inline') == 'true':
                response['Content-Disposition'] = f'inline; filename="{documento.arquivo.name.split("/")[-1]}"'
            else:
                response['Content-Disposition'] = f'attachment; filename="{documento.arquivo.name.split("/")[-1]}"'
            return response
    except FileNotFoundError:
        raise Http404("Arquivo não encontrado")


def excluir_documento(request, pk):
    """View para excluir um documento anexado"""
    documento = get_object_or_404(DocumentoAnexado, pk=pk)
    faturamento_pk = documento.faturamento.pk

    if request.method == 'POST':
        # Remove o arquivo do sistema de arquivos
        if documento.arquivo:
            documento.arquivo.delete(save=False)
        # Remove o registro do banco
        documento.delete()
        messages.success(request, 'Documento excluído com sucesso!')
        return redirect('faturamento_medico:detalhes', pk=faturamento_pk)

    context = {
        'documento': documento,
    }

    return render(request, 'faturamento_medico/confirmar_exclusao_documento.html', context)


def adicionar_item_servico(request, pk):
    """View para adicionar item de serviço a um faturamento"""
    faturamento = get_object_or_404(FaturamentoMedico, pk=pk)

    if request.method == 'POST':
        form = ItemServicoForm(request.POST, faturamento=faturamento)
        if form.is_valid():
            # Pegar os dados do POST
            cabecalho_id = request.POST.get('cabecalho')
            codigo_servico = request.POST.get('codigo_servico')
            qt = form.cleaned_data.get('qt', 1)
            valor = form.cleaned_data.get('valor', 0)

            if cabecalho_id and codigo_servico:
                from servicos_medicos.models import Cabecalho, ServicosMedicos
                try:
                    cabecalho = Cabecalho.objects.get(id=cabecalho_id)
                    servico = ServicosMedicos.objects.get(codigo=codigo_servico)

                    ItemServico.objects.create(
                        faturamento=faturamento,
                        codigo_servico=servico.codigo,
                        servico=servico.servicos,
                        porte=servico.porte_anestesico,
                        valor=valor,
                        qt=qt
                    )
                    # Atualiza o total do faturamento
                    faturamento.atualizar_total()
                    messages.success(request, 'Item de serviço adicionado com sucesso!')
                    return redirect('faturamento_medico:detalhes', pk=faturamento.pk)
                except (Cabecalho.DoesNotExist, ServicosMedicos.DoesNotExist):
                    messages.error(request, 'Cabeçalho ou serviço não encontrado.')
            else:
                messages.error(request, 'Selecione um cabeçalho e digite um código de serviço.')
    else:
        form = ItemServicoForm(faturamento=faturamento)

    context = {
        'form': form,
        'faturamento': faturamento,
        'titulo': f'Adicionar Item de Serviço - {faturamento}'
    }

    return render(request, 'faturamento_medico/adicionar_item_servico.html', context)


def editar_item_servico(request, pk):
    """View para editar item de serviço"""
    item = get_object_or_404(ItemServico, pk=pk)

    if request.method == 'POST':
        logger.info(f"Editando item {pk}, POST data: {request.POST}")
        # Para edição, cabecalho não é necessário
        post_data = request.POST.copy()
        form = ItemServicoForm(post_data, instance=item, faturamento=item.faturamento)
        logger.info(f"Form is_valid: {form.is_valid()}")
        if form.is_valid():
            logger.info("Salvando form")
            saved_item = form.save()
            logger.info(f"Item salvo: {saved_item.id}, valor: {saved_item.valor}, qt: {saved_item.qt}, total: {saved_item.total}")
            # Atualiza o total do faturamento
            item.faturamento.atualizar_total()
            logger.info(f"Total do faturamento atualizado: {item.faturamento.total}")
            messages.success(request, 'Item de serviço atualizado com sucesso!')
            return redirect('faturamento_medico:detalhes', pk=item.faturamento.pk)
        else:
            logger.error(f"Form errors: {form.errors}")
    else:
        form = ItemServicoForm(instance=item, faturamento=item.faturamento)

    context = {
        'form': form,
        'item': item,
        'faturamento': item.faturamento,
        'titulo': f'Editar Item de Serviço - {item.faturamento}'
    }

    return render(request, 'faturamento_medico/editar_item_servico.html', context)


def excluir_item_servico(request, pk):
    """View para excluir item de serviço"""
    item = get_object_or_404(ItemServico, pk=pk)
    faturamento_pk = item.faturamento.pk

    if request.method == 'POST':
        item.delete()
        # Atualiza o total do faturamento
        item.faturamento.atualizar_total()
        messages.success(request, 'Item de serviço excluído com sucesso!')
        return redirect('faturamento_medico:detalhes', pk=faturamento_pk)

    context = {
        'item': item,
    }

    return render(request, 'faturamento_medico/confirmar_exclusao_item.html', context)


def fechamento_repasse(request):
    """View para fechamento de repasse para anestesista"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:listar')

    # Filtros para seleção
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    convenios = request.GET.getlist('convenio')
    data_fechamento = request.GET.get('data_fechamento')
    anestesista = request.GET.get('anestesista')
    mostrar_fechados = request.GET.get('mostrar_fechados', 'false').lower() == 'true'

    # Buscar convênios disponíveis para a empresa
    convenios_disponiveis = []
    if empresa_id:
        from servicos_medicos.models import Convenio
        convenios_disponiveis = list(Convenio.objects.filter(empresa_id=empresa_id).order_by('nome'))
        if not convenios_disponiveis:
            # Convênios padrão se nenhum for encontrado para a empresa
            convenios_padrao = [
                {'nome': 'CBSAUDE'},
                {'nome': 'PM'},
                {'nome': 'UNIMED'},
                {'nome': 'BRADESCO'},
                {'nome': 'GEAP'},
                {'nome': 'SAUDE CAIXA'},
                {'nome': 'POSTAL SAUDE'},
                {'nome': 'FUSEX'},
                {'nome': 'LIFE EMPRESARIAL'},
                {'nome': 'CASSI'},
                {'nome': 'GCARD'},
                {'nome': 'PERSONAL NET'},
            ]
            convenios_disponiveis = convenios_padrao

    # Query base
    faturamentos = FaturamentoMedico.objects.filter(empresa_id=empresa_id)

    # Aplicar filtros
    if data_inicio:
        faturamentos = faturamentos.filter(data_fechamento__gte=data_inicio)
    if data_fim:
        faturamentos = faturamentos.filter(data_fechamento__lte=data_fim)
    if convenios:
        q_objects = Q()
        for conv in convenios:
            if conv:
                q_objects |= _q_convenio_filtro(conv)
        faturamentos = faturamentos.filter(q_objects)
    if anestesista:
        faturamentos = faturamentos.filter(anestesista__icontains=anestesista)

    # Filtrar apenas faturamentos com anestesista
    faturamentos = faturamentos.exclude(anestesista__isnull=True).exclude(anestesista='')

    # Processar fechamento se for POST
    if request.method == 'POST':
        faturamentos_ids = request.POST.getlist('faturamentos_selecionados')

        if request.POST.get('aplicar_comissao') and faturamentos_ids:
            # Aplicar comissão aos faturamentos selecionados
            percentual_imposto = float(request.POST.get('percentual_imposto', 0))
            percentual_comissao = float(request.POST.get('percentual_comissao', 0))

            # Usar getlist para obter todos os valores do campo
            faturamentos_ids = request.POST.getlist('faturamentos_selecionados')

            # Debug: verificar o que está sendo recebido
            logger.info(f"faturamentos_ids após getlist: {faturamentos_ids} (tipo: {type(faturamentos_ids)})")

            # Converter IDs para inteiros para evitar problemas de tipo
            try:
                faturamentos_ids = [int(id.strip()) for id in faturamentos_ids if id.strip()]
            except (ValueError, TypeError) as e:
                logger.error(f"Erro ao converter IDs: {e}. faturamentos_ids: {faturamentos_ids}")
                messages.error(request, 'IDs de faturamentos inválidos.')
                return redirect('faturamento_medico:fechamento_repasse')

            faturamentos = FaturamentoMedico.objects.filter(
                id__in=faturamentos_ids,
                empresa_id=empresa_id
            )

            for faturamento in faturamentos:
                # Calcular valores com precisão decimal (manter como Decimal)
                from decimal import Decimal
                total_decimal = Decimal(str(faturamento.total))
                percentual_imposto_decimal = Decimal(str(percentual_imposto))
                percentual_comissao_decimal = Decimal(str(percentual_comissao))

                valor_imposto = (total_decimal * percentual_imposto_decimal / 100).quantize(Decimal('0.01'))
                base_comissao = total_decimal - valor_imposto
                valor_comissao = (base_comissao * percentual_comissao_decimal / 100).quantize(Decimal('0.01'))

                # Atualizar campos
                faturamento.percentual_imposto = percentual_imposto
                faturamento.percentual_comissao = percentual_comissao
                faturamento.valor_imposto = valor_imposto
                faturamento.valor_comissao = valor_comissao
                faturamento.save()

            # Verificar se é uma requisição AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({
                    'success': True,
                    'message': f'Comissão aplicada com sucesso para {len(faturamentos_ids)} faturamento(s)!'
                })
            else:
                messages.success(request, f'Comissão aplicada com sucesso para {len(faturamentos_ids)} faturamento(s)!')
                return redirect('faturamento_medico:fechamento_repasse')

        elif faturamentos_ids:
            # Usar data atual se não foi fornecida
            if not data_fechamento:
                data_fechamento = timezone.now().date()

            # Gerar código único para este fechamento com bloqueio de transação
            from django.db import transaction
            max_attempts = 10
            attempt = 0
            codigo_fechamento = None

            while attempt < max_attempts:
                codigo_fechamento = str(uuid.uuid4())[:8].upper()
                logger.info(f"Código de fechamento gerado (tentativa {attempt + 1}): {codigo_fechamento}")

                # Verificar se o código já existe
                with transaction.atomic():
                    existing_with_code = FaturamentoMedico.objects.select_for_update().filter(codigo_fechamento=codigo_fechamento)
                    if not existing_with_code.exists():
                        logger.info(f"Código de fechamento único encontrado: {codigo_fechamento}")
                        break
                    else:
                        logger.warning(f"Código de fechamento {codigo_fechamento} já existe em {existing_with_code.count()} registros")
                        attempt += 1

            if attempt >= max_attempts:
                logger.error("Não foi possível gerar um código único após várias tentativas")
                messages.error(request, 'Erro interno: não foi possível gerar código único de fechamento.')
                return redirect('faturamento_medico:fechamento_repasse')

            # Buscar faturamentos selecionados para verificar status
            faturamentos_selecionados = FaturamentoMedico.objects.filter(
                id__in=faturamentos_ids,
                empresa_id=empresa_id
            )
            logger.info(f"Faturamentos selecionados: {len(faturamentos_selecionados)}")

            # Verificar se algum já está fechado
            faturamentos_ja_fechados = faturamentos_selecionados.filter(data_fechamento__isnull=False)
            if faturamentos_ja_fechados.exists():
                logger.warning(f"Encontrados {faturamentos_ja_fechados.count()} faturamentos já fechados:")
                for fat in faturamentos_ja_fechados:
                    logger.warning(f"  ID {fat.id}: data_fechamento={fat.data_fechamento}, codigo_fechamento={fat.codigo_fechamento}")

            # Verificar códigos de fechamento existentes nos faturamentos selecionados
            faturamentos_com_codigo = faturamentos_selecionados.filter(codigo_fechamento__isnull=False)
            if faturamentos_com_codigo.exists():
                logger.info(f"Faturamentos selecionados que já têm código de fechamento:")
                for fat in faturamentos_com_codigo:
                    logger.info(f"  ID {fat.id}: codigo_fechamento={fat.codigo_fechamento}")

            # Atualizar data de fechamento, status e código para os faturamentos selecionados
            # Primeiro, verificar se algum faturamento já tem um código de fechamento
            faturamentos_com_codigo_existente = FaturamentoMedico.objects.filter(
                id__in=faturamentos_ids,
                empresa_id=empresa_id,
                codigo_fechamento__isnull=False
            )

            if faturamentos_com_codigo_existente.exists():
                logger.warning(f"Encontrados {faturamentos_com_codigo_existente.count()} faturamentos que já têm código de fechamento:")
                for fat in faturamentos_com_codigo_existente:
                    logger.warning(f"  ID {fat.id}: codigo_fechamento={fat.codigo_fechamento}")
                # Para estes, não sobrescrever o código existente
                faturamentos_ids_para_atualizar = [id for id in faturamentos_ids if id not in [fat.id for fat in faturamentos_com_codigo_existente]]
                logger.info(f"Atualizando apenas {len(faturamentos_ids_para_atualizar)} faturamentos sem código existente")
            else:
                faturamentos_ids_para_atualizar = faturamentos_ids

            if faturamentos_ids_para_atualizar:
                # SOLUÇÃO DEFINITIVA: Usar códigos únicos por faturamento
                updated_count = 0
                for faturamento_id in faturamentos_ids_para_atualizar:
                    # Gerar código único para cada faturamento
                    faturamento_codigo = str(uuid.uuid4())[:8].upper()

                    try:
                        # Tentar atualizar este faturamento específico com código único
                        count = FaturamentoMedico.objects.filter(
                            id=faturamento_id,
                            empresa_id=empresa_id,
                            codigo_fechamento__isnull=True,  # Só atualizar se não tiver código
                            status__in=['pendente', 'enviado', 'aguardando_pagamento']
                        ).update(
                            data_fechamento=data_fechamento,
                            status='finalizado',
                            codigo_fechamento=faturamento_codigo
                        )
                        if count > 0:
                            updated_count += count
                            logger.info(f"Faturamento {faturamento_id} atualizado com código único {faturamento_codigo}")
                        else:
                            logger.warning(f"Faturamento {faturamento_id} não foi atualizado (já processado ou não encontrado)")
                    except Exception as e:
                        logger.error(f"Erro ao atualizar faturamento {faturamento_id}: {e}")
                        # Mesmo com códigos únicos, pode haver race condition, mas é muito improvável
                        raise e

                logger.info(f"Total de faturamentos atualizados com códigos únicos: {updated_count}")
                if updated_count > 0:
                    messages.success(request, f'Fechamento realizado com sucesso! {updated_count} faturamento(s) finalizado(s) com códigos únicos.')
                else:
                    messages.warning(request, 'Nenhum faturamento foi atualizado. Todos podem já ter sido processados.')
            else:
                logger.info("Nenhum faturamento para atualizar (todos já têm código de fechamento)")
                messages.info(request, 'Todos os faturamentos selecionados já possuem código de fechamento.')

            messages.success(request, f'Fechamento realizado com sucesso para {len(faturamentos_ids)} faturamento(s)! Código: {codigo_fechamento}')
            return redirect('faturamento_medico:fechamento_repasse')

    # Filtrar baseado na opção selecionada
    if mostrar_fechados:
        # Mostrar apenas faturamentos já fechados
        faturamentos_filtrados = faturamentos.filter(data_fechamento__isnull=False)
    else:
        # Mostrar apenas faturamentos não fechados (padrão)
        faturamentos_filtrados = faturamentos.filter(data_fechamento__isnull=True)

    # Estatísticas
    total_faturamentos = faturamentos_filtrados.count()
    valor_total = sum(f.total for f in faturamentos_filtrados if f.total)

    context = {
        'faturamentos': faturamentos_filtrados,
        'total_faturamentos': total_faturamentos,
        'valor_total': valor_total,
        'convenios_disponiveis': convenios_disponiveis,
        'filtros': {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'convenio': convenios,
            'data_fechamento': data_fechamento,
            'anestesista': anestesista,
        },
        'mostrar_fechados': mostrar_fechados,
    }

    return render(request, 'faturamento_medico/fechamento_repasse.html', context)


def reabrir_fechamento(request, pk):
    """View para reabrir um fechamento de repasse"""
    faturamento = get_object_or_404(FaturamentoMedico, pk=pk)

    if request.method == 'POST':
        # Limpar campos de fechamento
        faturamento.data_fechamento = None
        faturamento.status = 'pendente'
        faturamento.codigo_fechamento = None
        # Manter os valores de comissão e imposto calculados
        faturamento.save()

        messages.success(request, f'Fechamento reaberto com sucesso para {faturamento.nome}!')
        return redirect('faturamento_medico:fechamento_repasse')

    context = {
        'faturamento': faturamento,
    }

    return render(request, 'faturamento_medico/confirmar_reabertura.html', context)


def exportar_excel_fechados(request):
    """Exporta os repasses fechados para Excel com cabeçalho da empresa e ordenação por convênio"""
    empresa_id = request.session.get('empresa_id')
    if empresa_id:
        faturamentos = FaturamentoMedico.objects.filter(empresa_id=empresa_id)
    else:
        faturamentos = FaturamentoMedico.objects.none()

    # Aplicar os mesmos filtros da view de fechamento_repasse
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    convenios = request.GET.getlist('convenio')
    data_fechamento = request.GET.get('data_fechamento')
    anestesista = request.GET.get('anestesista')

    # Buscar convênios disponíveis para a empresa (para compatibilidade)
    convenios_disponiveis = []
    if empresa_id:
        from servicos_medicos.models import Convenio
        convenios_disponiveis = Convenio.objects.filter(empresa_id=empresa_id).order_by('nome')

    if data_inicio:
        faturamentos = faturamentos.filter(data__gte=data_inicio)
    if data_fim:
        faturamentos = faturamentos.filter(data__lte=data_fim)
    if convenios:
        q_objects = Q()
        for conv in convenios:
            if conv:
                q_objects |= _q_convenio_filtro(conv)
        faturamentos = faturamentos.filter(q_objects)
    if anestesista:
        faturamentos = faturamentos.filter(anestesista__icontains=anestesista)

    # Filtrar apenas faturamentos com anestesista e fechados
    faturamentos = faturamentos.exclude(anestesista__isnull=True).exclude(anestesista='')
    faturamentos = faturamentos.filter(data_fechamento__isnull=False)

    # Ordenar por convênio
    faturamentos = faturamentos.order_by('convenio', 'data')

    # Criar workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Repasses Fechados"

    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Estilo para cabeçalho da empresa
    empresa_font = Font(bold=True, size=14)
    empresa_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

    # Informações básicas da empresa
    if empresa_id:
        try:
            empresa = Empresa.objects.get(id=empresa_id)
            empresa_info = f"{empresa.razao} - CNPJ: {empresa.cnpj}"
        except Empresa.DoesNotExist:
            empresa_info = f"Empresa ID: {empresa_id} - Dados não encontrados"
    else:
        empresa_info = "Empresa não identificada"

    # Adicionar cabeçalho da empresa
    ws.cell(row=1, column=1).value = empresa_info
    ws.cell(row=1, column=1).font = empresa_font
    ws.cell(row=1, column=1).fill = empresa_fill

    # Título do relatório
    ws.cell(row=3, column=1).value = "RELATÓRIO DE REPASSES FECHADOS"
    ws.cell(row=3, column=1).font = Font(bold=True, size=16)

    # Data de geração
    from datetime import datetime
    ws.cell(row=4, column=1).value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws.cell(row=4, column=1).font = Font(italic=True)

    # Filtros aplicados
    filtros_texto = "Filtros aplicados:"
    if data_inicio:
        filtros_texto += f" Data início: {data_inicio}"
    if data_fim:
        filtros_texto += f" Data fim: {data_fim}"
    if data_fechamento:
        filtros_texto += f" Data fechamento: {data_fechamento}"
    if convenios:
        filtros_texto += f" Convênios: {', '.join(convenios)}"
    if anestesista:
        filtros_texto += f" Anestesista: {anestesista}"

    ws.cell(row=5, column=1).value = filtros_texto
    ws.cell(row=5, column=1).font = Font(italic=True)

    # Cabeçalhos dos dados (linha 7)
    headers = [
        'Data', 'Nome', 'Guia', 'Anestesista', 'Convênio','Codigo Relatorio' ,'Código Serviço', 'Serviço', 'QT', 'Valor Unitário',
        'Valor Total Item','Valor Total', 'Valor do Imposto', 'Valor da Comissão', 'Valor Líquido',
        'Data de Fechamento', 'Status'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    # Dados (a partir da linha 8)
    current_row = 8
    for faturamento in faturamentos:
        # Buscar itens de serviço para este faturamento
        valor_liquido = float(faturamento.total or 0) - float(faturamento.valor_imposto or 0) - float(faturamento.valor_comissao or 0)
        ws.cell(row=current_row, column=1).value = faturamento.data.strftime('%d/%m/%Y') if faturamento.data else ''
        ws.cell(row=current_row, column=2).value = faturamento.nome or ''
        ws.cell(row=current_row, column=3).value = faturamento.guia or ''
        ws.cell(row=current_row, column=4).value = faturamento.anestesista or ''
        ws.cell(row=current_row, column=5).value = faturamento.convenio or ''
        ws.cell(row=current_row, column=6).value = faturamento.codigo_relatorio or ''
        itens_servico = faturamento.itens_servico.filter(faturamento_id = faturamento.id)

        if itens_servico.exists():
            # Para cada item de serviço, criar uma linha
            cont = 0;
            for item in itens_servico:
                # Calcular valor líquido
                

                
                ws.cell(row=current_row, column=7).value = item.codigo_servico or ''
                ws.cell(row=current_row, column=8).value = item.servico or ''
                ws.cell(row=current_row, column=9).value = item.qt or 0
                ws.cell(row=current_row, column=10).value = float(item.valor) if item.valor else 0
                ws.cell(row=current_row, column=11).value = float(item.total) if item.total else 0
                
                if cont == 0:
                   ws.cell(row=current_row, column=12).value = float(faturamento.total) if faturamento.total else 0   
                   ws.cell(row=current_row, column=13).value = float(faturamento.valor_imposto) if faturamento.valor_imposto else 0
                   ws.cell(row=current_row, column=14).value = float(faturamento.valor_comissao) if faturamento.valor_comissao else 0
                   ws.cell(row=current_row, column=15).value = valor_liquido
                   ws.cell(row=current_row, column=16).value = faturamento.data_fechamento.strftime('%d/%m/%Y') if faturamento.data_fechamento else ''
                   
                   ws.cell(row=current_row, column=17).value = faturamento.status or ''    
                cont += 1
                current_row += 1
        
        cont = 0 
        
        current_row += 1

    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=repasses_fechados.xlsx'

    wb.save(response)
    return response


# Views para Serviços Disponíveis
def listar_servicos(request):
    """Lista todos os serviços disponíveis"""
    servicos = ServicoDisponivel.objects.all().order_by('codigo')

    # Filtros
    categoria = request.GET.get('categoria')
    ativo = request.GET.get('ativo')

    if categoria:
        servicos = servicos.filter(categoria__icontains=categoria)
    if ativo:
        if ativo == '1':
            servicos = servicos.filter(ativo=True)
        elif ativo == '0':
            servicos = servicos.filter(ativo=False)

    context = {
        'servicos': servicos,
        'filtros': {
            'categoria': categoria,
            'ativo': ativo,
        }
    }

    return render(request, 'faturamento_medico/listar_servicos.html', context)


def criar_servico(request):
    """Cria um novo serviço disponível"""
    if request.method == 'POST':
        form = ServicoDisponivelForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Serviço criado com sucesso!')
            return redirect('faturamento_medico:listar_servicos')
    else:
        form = ServicoDisponivelForm()

    context = {
        'form': form,
        'titulo': 'Criar Serviço Disponível'
    }

    return render(request, 'faturamento_medico/form_servico.html', context)


def editar_servico(request, pk):
    """Edita um serviço disponível existente"""
    servico = get_object_or_404(ServicoDisponivel, pk=pk)

    if request.method == 'POST':
        form = ServicoDisponivelForm(request.POST, instance=servico)
        if form.is_valid():
            form.save()
            messages.success(request, 'Serviço atualizado com sucesso!')
            return redirect('faturamento_medico:listar_servicos')
    else:
        form = ServicoDisponivelForm(instance=servico)

    context = {
        'form': form,
        'servico': servico,
        'titulo': 'Editar Serviço Disponível'
    }

    return render(request, 'faturamento_medico/form_servico.html', context)
def extrair_dados_documento(request):
    """View para extrair dados de documento usando Gemini via AJAX"""
    if request.method == 'POST':
        documento = request.FILES.get('documento')
        if documento:
            dados = processar_arquivos_com_gemini([documento])
            logger.info(f"Dados extraídos para debug: {dados}")
            return JsonResponse(dados)
    return JsonResponse({'error': 'Invalid request'}, status=400)


def extrair_dados_documento_ocr(request):
    """View para extrair dados de documento usando OCR via AJAX"""
    if request.method == 'POST':
        documento = request.FILES.get('documento')
        if documento:
            dados = processar_arquivos_com_ocr([documento])
            logger.info(f"Dados extraídos via OCR para debug: {dados}")
            return JsonResponse(dados)
    return JsonResponse({'error': 'Invalid request'}, status=400)


def excluir_servico(request, pk):
    """Exclui um serviço disponível"""
    servico = get_object_or_404(ServicoDisponivel, pk=pk)

    if request.method == 'POST':
        servico.delete()
        messages.success(request, 'Serviço excluído com sucesso!')
        return redirect('faturamento_medico:listar_servicos')

    context = {
        'servico': servico,
    }

    return render(request, 'faturamento_medico/confirmar_exclusao_servico.html', context)


def carregar_tabelas_por_cabecalho(request, cabecalho_id):
    """View AJAX para carregar tabelas por cabeçalho"""
    try:
        from servicos_medicos.models import Cabecalho, TabelaPreco
        cabecalho = Cabecalho.objects.get(id=cabecalho_id)
        tabelas = TabelaPreco.objects.filter(cabecalho=cabecalho).select_related('codigo_servico')
        data = []
        for tabela in tabelas:
            data.append({
                'id': tabela.id,
                'codigo': tabela.codigo_servico.codigo,
                'servico': tabela.codigo_servico.servicos,
                'porte': tabela.codigo_servico.porte_anestesico,
                'preco_apartamento': str(tabela.preco_apartamento),
                'preco_enfermaria': str(tabela.preco_enfermaria),
                'display': f"{tabela.codigo_servico} - {tabela.preco_apartamento}/{tabela.preco_enfermaria}"
            })
        return JsonResponse({'success': True, 'tabelas': data})
    except Cabecalho.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cabeçalho não encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def buscar_servicos(request):
    """View AJAX para buscar serviços por código"""
    query = request.GET.get('q', '')
    if len(query) >= 3:  # Buscar a partir de 3 caracteres
        from servicos_medicos.models import ServicosMedicos
        servicos = ServicosMedicos.objects.filter(codigo__icontains=query)[:10]  # Limitar a 10 resultados
        data = []
        for servico in servicos:
            data.append({
                'codigo': servico.codigo,
                'servico': servico.servicos,
                'porte': servico.porte_anestesico
            })
        return JsonResponse({'success': True, 'servicos': data})
    return JsonResponse({'success': True, 'servicos': []})


def buscar_servicos_por_descricao(request):
    """View AJAX para buscar serviços por descrição"""
    query = request.GET.get('q', '')
    if len(query) >= 3:  # Buscar a partir de 3 caracteres
        from servicos_medicos.models import ServicosMedicos
        servicos = ServicosMedicos.objects.filter(servicos__icontains=query)[:10]  # Limitar a 10 resultados
        data = []
        for servico in servicos:
            data.append({
                'codigo': servico.codigo,
                'servico': servico.servicos,
                'porte': servico.porte_anestesico
            })
        return JsonResponse({'success': True, 'servicos': data})
    return JsonResponse({'success': True, 'servicos': []})


def buscar_precos_servico(request, cabecalho_id, codigo_servico):
    """View AJAX para buscar preços de um serviço em um cabeçalho"""
    try:
        from servicos_medicos.models import Cabecalho, TabelaPreco, ServicosMedicos
        cabecalho = Cabecalho.objects.get(id=cabecalho_id)
        servico = ServicosMedicos.objects.get(codigo=codigo_servico)
        tabela = TabelaPreco.objects.filter(
            cabecalho=cabecalho,
            codigo_servico=servico
        ).first()
        if tabela:
            return JsonResponse({
                'success': True,
                'preco_apartamento': str(tabela.preco_apartamento),
                'preco_enfermaria': str(tabela.preco_enfermaria)
            })
        else:
            return JsonResponse({
                'success': True,
                'preco_apartamento': '0.00',
                'preco_enfermaria': '0.00'
            })
    except (Cabecalho.DoesNotExist, ServicosMedicos.DoesNotExist):
        return JsonResponse({'success': False, 'error': 'Cabeçalho ou serviço não encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def selecionar_lote_imprimir(request):
    """View para selecionar lote para imprimir relatório (apenas lotes não baixados)."""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    if request.method == 'POST' and request.POST.get('acao') == 'marcar_baixado':
        lote_id_raw = (request.POST.get('lote_baixado_id') or '').strip()
        if not lote_id_raw:
            messages.error(request, 'Informe o número do lote.')
            return redirect('faturamento_medico:selecionar_lote_imprimir')
        try:
            lote = Lote.objects.get(pk=int(lote_id_raw), empresa_id=empresa_id)
        except (TypeError, ValueError, Lote.DoesNotExist):
            messages.error(request, 'Lote não encontrado.')
            return redirect('faturamento_medico:selecionar_lote_imprimir')
        if lote.baixado:
            messages.info(request, f'Lote {lote.id} já estava marcado como baixado.')
        else:
            lote.baixado = True
            lote.save(update_fields=['baixado', 'data_atualizacao'])
            messages.success(request, f'Lote {lote.id} marcado como baixado e removido da lista de impressão.')
        return redirect('faturamento_medico:selecionar_lote_imprimir')

    lotes = (
        Lote.objects.filter(empresa_id=empresa_id, baixado=False)
        .prefetch_related('linhas_extrato_pagamento')
        .order_by('-id')
    )
    context = {'lotes': lotes}
    return render(request, 'faturamento_medico/selecionar_lote_imprimir.html', context)


def imprimir_lote(request, lote_id):
    """View para imprimir relatório de lote em HTML (layout padrão)."""
    if lote_id == 0:
        lote_id = request.GET.get('lote_id')
        if not lote_id:
            return HttpResponse('Lote não selecionado')

    empresa_id = request.GET.get('empresa_id') or request.session.get('empresa_id')
    if not empresa_id:
        return HttpResponse('Sessão expirada. Faça login novamente.')

    from .lote_relatorio import montar_contexto_relatorio_lote

    try:
        context = montar_contexto_relatorio_lote(lote_id, empresa_id, layout='padrao')
    except Lote.DoesNotExist:
        return HttpResponse('Lote não encontrado')
    except PermissionError:
        return HttpResponse('Acesso negado')

    return render(request, 'faturamento_medico/imprimir_lote.html', context)


def imprimir_lote_convenio_publico(request, lote_id):
    """Relatório de lote — layout para FUSEX, PM, Bombeiro e PP Saúde."""
    if lote_id == 0:
        lote_id = request.GET.get('lote_id')
        if not lote_id:
            return HttpResponse('Lote não selecionado')

    empresa_id = request.GET.get('empresa_id') or request.session.get('empresa_id')
    if not empresa_id:
        return HttpResponse('Sessão expirada. Faça login novamente.')

    from .lote_relatorio import montar_contexto_relatorio_lote

    try:
        context = montar_contexto_relatorio_lote(lote_id, empresa_id, layout='publico')
    except Lote.DoesNotExist:
        return HttpResponse('Lote não encontrado')
    except PermissionError:
        return HttpResponse('Acesso negado')

    return render(request, 'faturamento_medico/imprimir_lote_convenio_publico.html', context)


def imprimir_repasses_fechados(request):
    """View para imprimir relatório de repasses fechados em HTML"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return HttpResponse('Sessão expirada. Faça login novamente.')

    # Aplicar os mesmos filtros da view de fechamento_repasse
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    convenios = request.GET.getlist('convenio')
    data_fechamento = request.GET.get('data_fechamento')
    anestesista = request.GET.get('anestesista')

    # Query base - apenas faturamentos fechados com anestesista
    faturamentos = FaturamentoMedico.objects.filter(
        empresa_id=empresa_id,
        data_fechamento__isnull=False,
        anestesista__isnull=False
    ).exclude(anestesista='').order_by('anestesista', 'data_fechamento', 'nome')

    # Aplicar filtros
    if data_inicio:
        faturamentos = faturamentos.filter(data_fechamento__gte=data_inicio)
    if data_fim:
        faturamentos = faturamentos.filter(data_fechamento__lte=data_fim)
    if convenios:
        q_objects = Q()
        for conv in convenios:
            if conv:
                q_objects |= _q_convenio_filtro(conv)
        faturamentos = faturamentos.filter(q_objects)
    if anestesista:
        faturamentos = faturamentos.filter(anestesista__icontains=anestesista)

    # Agrupar por anestesista
    repasses_por_anestesista = {}
    total_geral = 0
    total_imposto_geral = 0
    total_comissao_geral = 0
    total_liquido_geral = 0

    for faturamento in faturamentos:
        anestesista_nome = faturamento.anestesista or 'Sem Anestesista'
        if anestesista_nome not in repasses_por_anestesista:
            repasses_por_anestesista[anestesista_nome] = {
                'repasses': [],
                'total_valor_total': 0,
                'total_valor_imposto': 0,
                'total_valor_comissao': 0,
                'total_valor_liquido': 0,
            }

        # Calcular valores
        valor_total = float(faturamento.total or 0)
        valor_imposto = float(faturamento.valor_imposto or 0)
        valor_comissao = float(faturamento.valor_comissao or 0)
        valor_liquido = valor_total - valor_imposto - valor_comissao

        repasse_info = {
            'faturamento': faturamento,
            'valor_total': valor_total,
            'valor_imposto': valor_imposto,
            'valor_comissao': valor_comissao,
            'valor_liquido': valor_liquido,
        }

        repasses_por_anestesista[anestesista_nome]['repasses'].append(repasse_info)
        repasses_por_anestesista[anestesista_nome]['total_valor_total'] += valor_total
        repasses_por_anestesista[anestesista_nome]['total_valor_imposto'] += valor_imposto
        repasses_por_anestesista[anestesista_nome]['total_valor_comissao'] += valor_comissao
        repasses_por_anestesista[anestesista_nome]['total_valor_liquido'] += valor_liquido

        # Acumuladores gerais
        total_geral += valor_total
        total_imposto_geral += valor_imposto
        total_comissao_geral += valor_comissao
        total_liquido_geral += valor_liquido

    empresa = Empresa.objects.get(id=empresa_id)

    # Calcular período
    from django.db.models import Min, Max
    periodo_inicio = faturamentos.aggregate(min_data=Min('data_fechamento'))['min_data']
    periodo_fim = faturamentos.aggregate(max_data=Max('data_fechamento'))['max_data']

    context = {
        'empresa': empresa,
        'periodo_inicio': periodo_inicio,
        'periodo_fim': periodo_fim,
        'repasses_por_anestesista': repasses_por_anestesista,
        'total_geral': total_geral,
        'total_imposto_geral': total_imposto_geral,
        'total_comissao_geral': total_comissao_geral,
        'total_liquido_geral': total_liquido_geral,
        'filtros': {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'convenio': convenios,
            'data_fechamento': data_fechamento,
            'anestesista': anestesista,
        }
    }

    return render(request, 'faturamento_medico/imprimir_repasses_fechados.html', context)


def carregar_precos_por_cabecalho(request, cabecalho_id):
    """View AJAX para carregar preços por cabeçalho"""
    try:
        from servicos_medicos.models import Cabecalho, TabelaPreco
        cabecalho = Cabecalho.objects.get(id=cabecalho_id)
        # Pegar os preços do primeiro serviço ou calcular médias, mas como são por serviço, talvez mostrar uma mensagem
        tabelas = TabelaPreco.objects.filter(cabecalho=cabecalho)
        if tabelas.exists():
            # Como os preços variam por serviço, talvez mostrar uma mensagem ou os preços do primeiro
            preco_apartamento = tabelas.first().preco_apartamento
            preco_enfermaria = tabelas.first().preco_enfermaria
            return JsonResponse({
                'success': True,
                'preco_apartamento': str(preco_apartamento),
                'preco_enfermaria': str(preco_enfermaria)
            })
        else:
            return JsonResponse({'success': False, 'error': 'Nenhuma tabela encontrada'})
    except Cabecalho.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cabeçalho não encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def gerar_lote(request):
    """View para gerar lote a partir dos faturamentos selecionados"""
    logger.info("Iniciando gerar_lote")
    empresa_id = request.session.get('empresa_id')
    logger.info(f"Empresa ID da sessão: {empresa_id}")
    if not empresa_id:
        logger.warning("Empresa não encontrada na sessão")
        messages.error(request, 'Empresa não encontrada na sessão.')
        return _redirect_ftlistar_com_filtros_sessao(request)

    if request.method == 'POST':
        logger.info("Método POST detectado")
        faturamento_ids = request.POST.getlist('faturamentos_selecionados')
        lote_existente_id = request.POST.get('lote_existente')
        logger.info(f"Faturamento IDs selecionados: {faturamento_ids}")
        logger.info(f"Lote existente: {lote_existente_id}")

        if not faturamento_ids:
            logger.warning("Nenhum faturamento selecionado")
            messages.error(request, 'Selecione pelo menos um faturamento para gerar o lote.')
            return _redirect_ftlistar_com_filtros_sessao(request)

        # Buscar faturamentos selecionados
        faturamentos = FaturamentoMedico.objects.filter(
            id__in=faturamento_ids,
            empresa_id=empresa_id
        )
        logger.info(f"Faturamentos encontrados: {faturamentos.count()}")

        if not faturamentos.exists():
            logger.warning("Nenhum faturamento encontrado para os IDs")
            messages.error(request, 'Nenhum faturamento encontrado.')
            return _redirect_ftlistar_com_filtros_sessao(request)

        if lote_existente_id:
            # Adicionar a lote existente
            try:
                lote_existente = Lote.objects.get(id=lote_existente_id, empresa_id=empresa_id)
                logger.info(f"Adicionando a lote existente: {lote_existente.id}")
            except Lote.DoesNotExist:
                logger.error(f"Lote existente não encontrado: {lote_existente_id}")
                messages.error(request, 'Lote selecionado não encontrado.')
                return _redirect_ftlistar_com_filtros_sessao(request)

            if not lote_existente.aberto_para_adicionar():
                messages.error(request, 'Este lote já foi finalizado/baixado e não aceita novos faturamentos.')
                return _redirect_ftlistar_com_filtros_sessao(request)

            # Verificar se os faturamentos têm o mesmo convênio do lote
            faturamentos_diferente_convenio = faturamentos.exclude(convenio=lote_existente.convenio)
            if faturamentos_diferente_convenio.exists():
                logger.warning(f"Faturamentos com convênio diferente: {[f.id for f in faturamentos_diferente_convenio]}")
                messages.error(request, 'Todos os faturamentos devem ter o mesmo convênio do lote selecionado.')
                return _redirect_ftlistar_com_filtros_sessao(request)

            # Filtrar faturamentos elegíveis (pendente, sem lote interno)
            ids_internos = ids_lotes_internos(empresa_id)
            faturamentos_validos = [
                f for f in faturamentos
                if faturamento_elegivel_lote(f, ids_internos=ids_internos)
            ]
            faturamentos_invalidos = [
                f for f in faturamentos if f not in faturamentos_validos
            ]

            if faturamentos_invalidos:
                logger.warning(f"Faturamentos inelegíveis: {[f.id for f in faturamentos_invalidos]}")
                messages.warning(
                    request,
                    'Somente faturamentos pendentes, conferidos e sem lote interno podem entrar em lotes.',
                )

            if not faturamentos_validos:
                logger.warning("Nenhum faturamento válido para adicionar")
                messages.error(request, 'Nenhum faturamento válido para adicionar ao lote.')
                return _redirect_ftlistar_com_filtros_sessao(request)

            # Atualizar os faturamentos com o ID do lote e status aguardando pagamento
            fat_ids = [f.id for f in faturamentos_validos]
            try:
                updated = FaturamentoMedico.objects.filter(id__in=fat_ids).update(
                    lote=str(lote_existente.id), status='aguardando_pagamento'
                )
                logger.info(f"Faturamentos adicionados ao lote {lote_existente.id}: {updated}")
            except Exception as e:
                logger.error(f"Erro ao adicionar faturamentos ao lote {lote_existente.id}: {e}")
                messages.error(request, f'Erro ao adicionar faturamentos ao lote: {e}')
                return _redirect_ftlistar_com_filtros_sessao(request)

            # Atualizar o total do lote
            try:
                lote_existente.atualizar_total()
                lote_existente.sincronizar_extrato_pagamento()
                logger.info(f"Total do lote {lote_existente.id} atualizado: {lote_existente.total_lote}")
            except Exception as e:
                logger.error(f"Erro ao atualizar total do lote {lote_existente.id}: {e}")
                messages.error(request, f'Erro ao atualizar total do lote: {e}')
                return _redirect_ftlistar_com_filtros_sessao(request)

            url = _url_ftlistar_com_filtros_sessao(request)
            return HttpResponse(
                f'<script>alert("Faturamentos adicionados ao lote {lote_existente.id} com sucesso!"); '
                f'window.location.href = "{url}";</script>'
            )
        else:
            # Criar novo lote
            ids_internos = ids_lotes_internos(empresa_id)
            faturamentos_sem_lote = [
                f for f in faturamentos
                if faturamento_elegivel_lote(f, ids_internos=ids_internos)
            ]
            faturamentos_com_lote = [
                f for f in faturamentos if f not in faturamentos_sem_lote
            ]

            if faturamentos_com_lote:
                logger.warning(f"Alguns faturamentos inelegíveis: {[f.id for f in faturamentos_com_lote]}")
                messages.warning(
                    request,
                    'Alguns faturamentos selecionados já estão em lote interno ou não estão conferidos.',
                )

            if not faturamentos_sem_lote:
                logger.warning("Nenhum faturamento elegível encontrado")
                messages.error(request, 'Nenhum faturamento elegível para gerar lote.')
                return _redirect_ftlistar_com_filtros_sessao(request)

            faturamentos = faturamentos_sem_lote
            logger.info(f"Faturamentos elegíveis: {len(faturamentos)}")

            # Agrupar faturamentos por convênio
            faturamentos_por_convenio = {}
            for fat in faturamentos:
                convenio = fat.convenio or 'Sem Convênio'
                if convenio not in faturamentos_por_convenio:
                    faturamentos_por_convenio[convenio] = []
                faturamentos_por_convenio[convenio].append(fat)

            logger.info(f"Faturamentos agrupados por convênio: { {k: len(v) for k, v in faturamentos_por_convenio.items()} }")

            lotes_criados = []
            for convenio, fats in faturamentos_por_convenio.items():
                logger.info(f"Criando lote para convênio: {convenio}")

                # Criar o lote
                try:
                    lote = Lote.objects.create(
                        empresa_id=empresa_id,
                        convenio=convenio if convenio != 'Sem Convênio' else None
                    )
                    logger.info(f"Lote criado: {lote.id} para convênio {convenio}")
                except Exception as e:
                    logger.error(f"Erro ao criar lote para convênio {convenio}: {e}")
                    messages.error(request, f'Erro ao criar lote para convênio {convenio}: {e}')
                    continue

                # Atualizar os faturamentos com o ID do lote e status aguardando pagamento
                fat_ids = [f.id for f in fats]
                try:
                    updated = FaturamentoMedico.objects.filter(id__in=fat_ids).update(
                        lote=str(lote.id), status='aguardando_pagamento'
                    )
                    logger.info(f"Faturamentos atualizados para lote {lote.id}: {updated}")
                except Exception as e:
                    logger.error(f"Erro ao atualizar faturamentos para lote {lote.id}: {e}")
                    messages.error(request, f'Erro ao atualizar faturamentos para lote {lote.id}: {e}')
                    continue

                # Atualizar o total do lote
                try:
                    lote.atualizar_total()
                    lote.sincronizar_extrato_pagamento()
                    logger.info(f"Total do lote {lote.id} atualizado: {lote.total_lote}")
                except Exception as e:
                    logger.error(f"Erro ao atualizar total do lote {lote.id}: {e}")
                    messages.error(request, f'Erro ao atualizar total do lote {lote.id}: {e}')
                    continue

                lotes_criados.append(lote.id)

            if lotes_criados:
                url = _url_ftlistar_com_filtros_sessao(request)
                logger.info(f"Lotes criados: {lotes_criados}")
                lotes_str = ', '.join(map(str, lotes_criados))
                return HttpResponse(
                    f'<script>alert("Lotes gerados com sucesso: {lotes_str}"); '
                    f'window.location.href = "{url}";</script>'
                )
            else:
                logger.warning("Nenhum lote foi criado")
                messages.error(request, 'Nenhum lote foi criado.')
                return _redirect_ftlistar_com_filtros_sessao(request)

    logger.info("Método não é POST, redirecionando")
    return _redirect_ftlistar_com_filtros_sessao(request)


def vincular_lote_protocolo(request):
    """Vincula guias selecionadas usando lote e protocolo já informados no faturamento (convênio externo)."""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return _redirect_ftlistar_com_filtros_sessao(request)

    if request.method != 'POST':
        return _redirect_ftlistar_com_filtros_sessao(request)

    faturamento_ids = request.POST.getlist('faturamentos_selecionados')
    if not faturamento_ids:
        messages.error(request, 'Selecione pelo menos um faturamento para vincular lote e protocolo.')
        return _redirect_ftlistar_com_filtros_sessao(request)

    try:
        ids = [int(x) for x in faturamento_ids]
    except (TypeError, ValueError):
        messages.error(request, 'Seleção inválida.')
        return _redirect_ftlistar_com_filtros_sessao(request)

    from faturamento_medico.services.gerar_lotes_geap import (
        preencher_lote_protocolo_faturamentos,
        vincular_lote_protocolo_selecionados,
    )

    lote_convenio = (request.POST.get('lote_convenio') or '').strip()
    protocolo = (request.POST.get('protocolo') or '').strip()
    guia = (request.POST.get('guia') or '').strip()
    if lote_convenio or protocolo or guia:
        preenchidos = preencher_lote_protocolo_faturamentos(
            empresa_id=empresa_id,
            faturamento_ids=ids,
            lote_convenio=lote_convenio,
            protocolo=protocolo,
            guia=guia,
        )
        if preenchidos:
            messages.info(
                request,
                f'Dados informados em {preenchidos} faturamento(s) antes da vinculação.',
            )

    stats = vincular_lote_protocolo_selecionados(empresa_id=empresa_id, faturamento_ids=ids)
    if stats['lotes_criados']:
        lotes_str = ', '.join(str(x) for x in stats['lotes_criados'])
        messages.success(
            request,
            f"Vinculados {stats['faturamentos']} guia(s) em {stats['grupos']} lote(s)/protocolo(s): {lotes_str}.",
        )
    else:
        messages.warning(request, 'Nenhum lote vinculado. Verifique conferência, lote e protocolo nos faturamentos.')

    for err in stats.get('erros') or []:
        messages.info(request, err)

    return _redirect_ftlistar_com_filtros_sessao(request)


def importar_unimed(request):
    """View para importar relatório UNIMED"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    if request.method == 'POST':
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            messages.error(request, 'Selecione um arquivo para importar.')
            return redirect('faturamento_medico:importar_unimed')

        try:
            # Ler o arquivo
            content = arquivo.read().decode('utf-8')
            lines = content.split('\n')

            # Pular cabeçalho
            data_lines = lines[1:]

            # Agrupar por lote e guia
            grupos = {}
            servicos_unicos = set()

            for line in data_lines:
                if not line.strip():
                    continue
                parts = line.split(';')
                if len(parts) < 13:
                    continue

                lote = parts[0].strip()
                guia = parts[1].strip()
                cod_usuario = parts[2].strip()
                nome_usuario = parts[3].strip()
                plano = parts[4].strip()
                cod_servico = parts[5].strip()
                desc_servico = parts[6].strip()
                tp_grau = parts[7].strip()
                data_str = parts[8].strip()
                qtde_via = parts[9].strip()
                percentual = parts[10].strip().replace(',', '.')
                valor_unit = parts[11].strip().replace(',', '.')
                valor_total = parts[12].strip().replace(',', '.')
                cod_rel = parts[13].strip()
                observacao = parts[14].strip() if len(parts) > 12 else ''

                # Converter data
                try:
                    data = datetime.strptime(data_str, '%d/%m/%Y').date()
                except:
                    data = timezone.now().date()

                chave = f"{lote}_{guia}"

                if chave not in grupos:
                    grupos[chave] = {
                        'lote': lote,
                        'guia': guia,
                        'carteirinha': cod_usuario,
                        'nome': nome_usuario,
                        'plano': plano,
                        'data': data,
                        'cod_rel': cod_rel,
                        'servicos': []
                    }

                grupos[chave]['servicos'].append({
                    'codigo': cod_servico,
                    'descricao': desc_servico,
                    'porte': tp_grau,
                    'qt': int(float(qtde_via)) if qtde_via else 1,
                    'percentual': float(percentual) if percentual else 0,
                    'valor': float(valor_unit) if valor_unit else 0,
                    'total': float(valor_total) if valor_total else 0,
                    'observacao': observacao
                })

                servicos_unicos.add((cod_servico, desc_servico))

            # Verificar e criar serviços não cadastrados
            from servicos_medicos.models import ServicosMedicos
            servicos_criados = 0
            for cod, desc in servicos_unicos:
                if not ServicosMedicos.objects.filter(codigo=cod).exists():
                    ServicosMedicos.objects.create(
                        codigo=cod,
                        servicos=desc,
                        porte_anestesico=None  # Será definido depois se necessário
                    )
                    servicos_criados += 1

            # Criar faturamentos
            faturamentos_criados = 0
            itens_criados = 0

            for chave, dados in grupos.items():
                # Criar faturamento
                faturamento = FaturamentoMedico.objects.create(
                    empresa_id=empresa_id,
                    lote=dados['lote'],
                    guia=dados['guia'],
                    carteirinha=dados['carteirinha'],
                    nome=dados['nome'],
                    data=dados['data'],
                    convenio='UNIMED',
                    codigo_relatorio=dados['cod_rel'],
                    status='pendente'
                )

                # Criar itens de serviço
                for servico in dados['servicos']:
                    ItemServico.objects.create(
                        faturamento=faturamento,
                        codigo_servico=servico['codigo'],
                        servico=servico['descricao'],
                        porte=servico['porte'],
                        percentual = servico['percentual'],
                        qt=servico['qt'],
                        valor=servico['valor'],
                        total=servico['total']
                    )
                    itens_criados += 1

                # Atualizar total do faturamento
                faturamento.atualizar_total()
                faturamentos_criados += 1

            messages.success(request, f'Importação concluída! {servicos_criados} serviços criados, {faturamentos_criados} faturamentos criados, {itens_criados} itens de serviço criados.')

        except Exception as e:
            messages.error(request, f'Erro durante a importação: {str(e)}')
            return redirect('faturamento_medico:importar_unimed')

        return redirect('faturamento_medico:ftlistar')

    context = {
        'titulo': 'Importar Relatório UNIMED'
    }

    return render(request, 'faturamento_medico/importar_unimed.html', context)


def importar_xml(request):
    """View para importar XML de NFSe"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    if request.method == 'POST':
        arquivos = request.FILES.getlist('arquivos')
        if not arquivos:
            messages.error(request, 'Selecione pelo menos um arquivo XML para importar.')
            return redirect('faturamento_medico:importar_xml')

        try:
            import xml.etree.ElementTree as ET
            from servicos_medicos.models import ServicosMedicos

            faturamentos_criados = 0
            itens_criados = 0
            servicos_criados = 0
            servicos_unicos = set()

            for arquivo in arquivos:
                # Parse do XML
                content = arquivo.read().decode('utf-8')
                root = ET.fromstring(content)

                # Namespace do XML
                ns = {'nfse': 'http://www.abrasf.org.br/nfse.xsd'}

                # Encontrar o elemento Nfse
                nfse = root.find('.//nfse:Nfse', ns)
                if nfse is None:
                    continue

                inf_nfse = nfse.find('nfse:InfNfse', ns)
                if inf_nfse is None:
                    continue

                # Extrair dados básicos
                numero = inf_nfse.findtext('nfse:Numero', default='')
                data_emissao_str = inf_nfse.findtext('nfse:DataEmissao', default='')
                outras_info = inf_nfse.findtext('nfse:OutrasInformacoes', default='')

                # Valores NFSe
                valores_nfse = inf_nfse.find('nfse:ValoresNfse', ns)
                valor_liquido = 0.0
                if valores_nfse is not None:
                    valor_liquido_str = valores_nfse.findtext('nfse:ValorLiquidoNfse', default='0')
                    try:
                        valor_liquido = float(valor_liquido_str)
                    except:
                        valor_liquido = 0.0

                # Prestador
                prestador_servico = inf_nfse.find('nfse:PrestadorServico', ns)
                prestador_cnpj = ''
                prestador_razao = ''
                if prestador_servico is not None:
                    identificacao = prestador_servico.find('nfse:IdentificacaoPrestador', ns)
                    if identificacao is not None:
                        cpf_cnpj = identificacao.find('nfse:CpfCnpj', ns)
                        if cpf_cnpj is not None:
                            prestador_cnpj = cpf_cnpj.findtext('nfse:Cnpj', default='')
                    prestador_razao = prestador_servico.findtext('nfse:RazaoSocial', default='')

                # Tomador
                declaracao = inf_nfse.find('nfse:DeclaracaoPrestacaoServico', ns)
                tomador_nome = ''
                tomador_cpf = ''
                if declaracao is not None:
                    inf_declaracao = declaracao.find('nfse:InfDeclaracaoPrestacaoServico', ns)
                    if inf_declaracao is not None:
                        tomador = inf_declaracao.find('nfse:Tomador', ns)
                        if tomador is not None:
                            tomador_nome = tomador.findtext('nfse:RazaoSocial', default='')
                            identificacao_tomador = tomador.find('nfse:IdentificacaoTomador', ns)
                            if identificacao_tomador is not None:
                                cpf_cnpj_tomador = identificacao_tomador.find('nfse:CpfCnpj', ns)
                                if cpf_cnpj_tomador is not None:
                                    tomador_cpf = cpf_cnpj_tomador.findtext('nfse:Cpf', default='')

                        # Serviço
                        servico = inf_declaracao.find('nfse:Servico', ns)
                        if servico is not None:
                            discriminacao = servico.findtext('nfse:Discriminacao', default='')
                            item_lista = servico.findtext('nfse:ItemListaServico', default='')
                            codigo_cnae = servico.findtext('nfse:CodigoCnae', default='')
                            competencia_str = inf_declaracao.findtext('nfse:Competencia', default='')

                            # Valores do serviço
                            valores_servico = servico.find('nfse:Valores', ns)
                            valor_servicos = 0.0
                            if valores_servico is not None:
                                valor_servicos_str = valores_servico.findtext('nfse:ValorServicos', default='0')
                                try:
                                    valor_servicos = float(valor_servicos_str)
                                except:
                                    valor_servicos = 0.0

                            # Converter datas
                            data_emissao = None
                            try:
                                if data_emissao_str:
                                    data_emissao = datetime.fromisoformat(data_emissao_str.replace('Z', '+00:00'))
                            except:
                                data_emissao = timezone.now()

                            competencia = None
                            try:
                                if competencia_str:
                                    competencia = datetime.strptime(competencia_str, '%Y-%m-%d').date()
                            except:
                                competencia = data_emissao.date() if data_emissao else timezone.now().date()

                            # Criar faturamento
                            faturamento = FaturamentoMedico.objects.create(
                                empresa_id=empresa_id,
                                guia=numero,
                                nome=tomador_nome,
                                carteirinha=tomador_cpf,
                                data=competencia or timezone.now().date(),
                                data_autorizacao=data_emissao.date() if data_emissao else None,
                                total=valor_liquido,
                                convenio='NFSE',
                                codigo_relatorio='1',
                                status='pendente',
                                observacao=f"{discriminacao}\n{outras_info}".strip()
                            )

                            # Criar item de serviço
                            ItemServico.objects.create(
                                faturamento=faturamento,
                                codigo_servico=item_lista,
                                servico=discriminacao,
                                porte='',  # NFSe não tem porte anestésico
                                qt=1,
                                valor=valor_servicos,
                                total=valor_liquido
                            )
                            itens_criados += 1

                            # Adicionar serviço único para possível criação
                            servicos_unicos.add((item_lista, discriminacao))

                            faturamentos_criados += 1

            # Verificar e criar serviços não cadastrados
            for cod, desc in servicos_unicos:
                if cod and not ServicosMedicos.objects.filter(codigo=cod).exists():
                    ServicosMedicos.objects.create(
                        codigo=cod,
                        servicos=desc,
                        porte_anestesico=None
                    )
                    servicos_criados += 1

            messages.success(request, f'Importação XML concluída! {servicos_criados} serviços criados, {faturamentos_criados} faturamentos criados, {itens_criados} itens de serviço criados.')

        except Exception as e:
            messages.error(request, f'Erro durante a importação XML: {str(e)}')
            return redirect('faturamento_medico:importar_xml')

        return redirect('faturamento_medico:ftlistar')

    context = {
        'titulo': 'Importar XML NFSe'
    }

    return render(request, 'faturamento_medico/importar_xml.html', context)


RIS_HEADERS = [
    'Unidade',
    'Data',
    'Paciente',
    'Cartão Nacional de Saúde',
    'CPF',
    'Cor/Raça',
    'Idade',
    'E-mail',
    'Telefone',
    'Número do lote',
    'Procedimento',
    'Prioridade',
    'Horário de início',
    'Horário de fim',
    'Modalidade',
    'Valor',
    'Agendado via',
    'Acréscimo/Desconto',
    'Valor pago',
    'Observações de Pagamento',
    'Status do Agendamento',
    'Motivo Cancelamento/Desistência/Deleção',
    'Médico',
    'Médico solicitante',
    'Técnico',
    'Check-in por',
    'Agendado por',
    'Viabilidade',
    'Tag',
    'Indicação clínica',
    'Descrição',
]


def _parse_data_ris(valor):
    """Converte data do Excel RIS para date."""
    if valor is None or valor == '':
        return timezone.now().date()
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return timezone.now().date()


def _parse_valor_ris(valor):
    """Converte valor monetário do Excel RIS para Decimal."""
    if valor is None or valor == '':
        return Decimal('0')
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    texto = str(valor).strip().replace('R$', '').replace(' ', '')
    if ',' in texto and '.' in texto:
        texto = texto.replace('.', '').replace(',', '.')
    elif ',' in texto:
        texto = texto.replace(',', '.')
    try:
        return Decimal(texto)
    except (InvalidOperation, ValueError):
        return Decimal('0')


def _celula_texto(valor, max_len=None):
    if valor is None:
        return ''
    texto = str(valor).strip()
    if max_len:
        return texto[:max_len]
    return texto


def baixar_modelo_ris(request):
    """Disponibiliza o modelo próprio RIS (.xlsx) para download."""
    from django.conf import settings
    import os

    caminho = os.path.join(
        settings.BASE_DIR,
        'faturamento_medico',
        'static',
        'faturamento_medico',
        'modelo_ris.xlsx',
    )
    if os.path.exists(caminho):
        with open(caminho, 'rb') as f:
            response = HttpResponse(
                f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="modelo_ris.xlsx"'
            return response

    # Fallback: gera o modelo em memória
    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatório'
    for i, header in enumerate(RIS_HEADERS, 1):
        ws.cell(1, i, header)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="modelo_ris.xlsx"'
    return response


def renomear_guias_geap(request):
    """Renomeia PDFs de guias (DATA - CONVENIO - NOME) e devolve JSON para download."""
    import base64

    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'erro': 'Empresa não encontrada na sessão.'}, status=403)
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    if request.method == 'POST':
        arquivos = request.FILES.getlist('arquivos')
        if not arquivos:
            return JsonResponse({'erro': 'Selecione pelo menos um arquivo PDF.'}, status=400)

        convenio_padrao = (request.POST.get('convenio_padrao') or '').strip()
        anexar_no_sistema = request.POST.get('anexar_no_sistema', '1') != '0'
        from faturamento_medico.services.renomear_guias_geap import renomear_guias_geap_arquivos

        resultados = renomear_guias_geap_arquivos(
            arquivos,
            convenio_padrao=convenio_padrao,
            empresa_id=empresa_id,
            anexar_no_sistema=anexar_no_sistema,
        )
        ok_count = sum(1 for r in resultados if r.ok)
        anexo_count = sum(1 for r in resultados if r.anexo_ok)
        payload = {
            'ok_count': ok_count,
            'erro_count': len(resultados) - ok_count,
            'anexo_count': anexo_count,
            'resultados': [
                {
                    'arquivo_original': r.arquivo_original,
                    'arquivo_novo': r.arquivo_novo,
                    'data_autorizacao': r.data_autorizacao,
                    'nome_beneficiario': r.nome_beneficiario,
                    'convenio': r.convenio,
                    'tipo_guia': r.tipo_guia,
                    'faturamento_id': r.faturamento_id,
                    'anexo_ok': r.anexo_ok,
                    'anexo_tentado': r.anexo_tentado,
                    'anexo_mensagem': r.anexo_mensagem,
                    'anexo_erro': r.anexo_erro,
                    'anexo_sugestoes': r.anexo_sugestoes,
                    'ok': r.ok,
                    'erro': r.erro,
                    'pdf_base64': base64.b64encode(r.pdf_bytes).decode('ascii') if r.ok else '',
                }
                for r in resultados
            ],
        }
        return JsonResponse(payload)

    return render(
        request,
        'faturamento_medico/renomear_guias_geap.html',
        {'titulo': 'Renomear Guias'},
    )


def buscar_faturamentos_guia(request):
    """Busca manual de faturamentos para vincular guia não encontrada."""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'erro': 'Empresa não encontrada na sessão.'}, status=403)

    from faturamento_medico.services.vincular_guia_anexo import (
        buscar_faturamentos_manual,
        sugestao_para_dict,
    )

    termo = (request.GET.get('termo') or request.POST.get('termo') or '').strip()
    convenio = (request.GET.get('convenio') or request.POST.get('convenio') or '').strip()
    data_guia = (request.GET.get('data_guia') or request.POST.get('data_guia') or '').strip()

    if not termo and not convenio and not data_guia:
        return JsonResponse({'erro': 'Informe paciente, convênio ou data para buscar.'}, status=400)

    resultados = buscar_faturamentos_manual(
        empresa_id=empresa_id,
        termo=termo,
        convenio=convenio,
        data_guia_str=data_guia,
    )
    return JsonResponse({
        'resultados': [sugestao_para_dict(s) for s in resultados],
    })


def anexar_guia_manual(request):
    """Anexa PDF de guia a um faturamento escolhido manualmente."""
    import base64
    import json

    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'erro': 'Empresa não encontrada na sessão.'}, status=403)

    if request.content_type and 'application/json' in request.content_type:
        try:
            payload = json.loads(request.body.decode('utf-8'))
        except (json.JSONDecodeError, UnicodeDecodeError):
            return JsonResponse({'erro': 'JSON inválido.'}, status=400)
    else:
        payload = request.POST

    faturamento_id = payload.get('faturamento_id')
    nome_arquivo = (payload.get('nome_arquivo') or '').strip()
    pdf_base64 = (payload.get('pdf_base64') or '').strip()

    if not faturamento_id or not nome_arquivo or not pdf_base64:
        return JsonResponse({'erro': 'Dados incompletos para anexar a guia.'}, status=400)

    try:
        faturamento_id = int(faturamento_id)
        pdf_bytes = base64.b64decode(pdf_base64)
    except (ValueError, TypeError):
        return JsonResponse({'erro': 'Faturamento ou PDF inválido.'}, status=400)

    from faturamento_medico.services.vincular_guia_anexo import anexar_guia_por_faturamento_id

    resultado = anexar_guia_por_faturamento_id(
        empresa_id=empresa_id,
        faturamento_id=faturamento_id,
        pdf_bytes=pdf_bytes,
        nome_arquivo=nome_arquivo,
    )
    if not resultado.ok:
        return JsonResponse({'ok': False, 'erro': resultado.erro}, status=400)

    return JsonResponse({
        'ok': True,
        'faturamento_id': resultado.faturamento_id,
        'mensagem': resultado.mensagem,
        'paciente': resultado.paciente,
        'data_procedimento': resultado.data_procedimento,
    })


def importar_ris(request):
    """View para importar relatório no modelo próprio RIS (.xlsx)."""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    if request.method == 'POST':
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            messages.error(request, 'Selecione um arquivo para importar.')
            return redirect('faturamento_medico:importar_ris')

        nome_arquivo = (arquivo.name or '').lower()
        if not nome_arquivo.endswith(('.xlsx', '.xlsm')):
            messages.error(request, 'Envie um arquivo Excel no modelo RIS (.xlsx).')
            return redirect('faturamento_medico:importar_ris')

        try:
            wb = load_workbook(filename=BytesIO(arquivo.read()), data_only=True)
            ws = wb.active

            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                headers.append(_celula_texto(cell.value).lower())

            def idx(*nomes):
                for nome in nomes:
                    nome_l = nome.lower()
                    if nome_l in headers:
                        return headers.index(nome_l)
                return None

            col = {
                'unidade': idx('Unidade'),
                'data': idx('Data'),
                'paciente': idx('Paciente'),
                'cns': idx('Cartão Nacional de Saúde', 'Cartao Nacional de Saude'),
                'cpf': idx('CPF'),
                'lote': idx('Número do lote', 'Numero do lote'),
                'procedimento': idx('Procedimento'),
                'prioridade': idx('Prioridade'),
                'horario_inicio': idx('Horário de início', 'Horario de inicio'),
                'horario_fim': idx('Horário de fim', 'Horario de fim'),
                'modalidade': idx('Modalidade'),
                'valor': idx('Valor'),
                'agendado_via': idx('Agendado via', 'Agendado Via'),
                'status': idx('Status do Agendamento'),
                'motivo_cancelamento': idx(
                    'Motivo Cancelamento/Desistência/Deleção',
                    'Motivo Cancelamento/Desistencia/Delecao',
                ),
                'medico': idx('Médico', 'Medico'),
                'medico_solicitante': idx('Médico solicitante', 'Medico solicitante'),
                'tecnico': idx('Técnico', 'Tecnico'),
                'checkin_por': idx('Check-in por', 'Check-in Por'),
                'agendado_por': idx('Agendado por', 'Agendado Por'),
                'convenio': idx('Viabilidade'),
                'tag': idx('Tag'),
                'indicacao': idx('Indicação clínica', 'Indicacao clinica'),
                'descricao': idx('Descrição', 'Descricao'),
                'obs_pagamento': idx('Observações de Pagamento', 'Observacoes de Pagamento'),
            }

            obrigatorias = ['data', 'paciente', 'procedimento', 'valor']
            faltando = [c for c in obrigatorias if col[c] is None]
            if faltando:
                messages.error(
                    request,
                    'Arquivo fora do modelo RIS. Colunas obrigatórias não encontradas: '
                    + ', '.join(faltando)
                    + '. Baixe o modelo próprio RIS e use esse layout.'
                )
                return redirect('faturamento_medico:importar_ris')

            status_ignorar = set()  # Cancelado/Desistência/Deletado passam a ser importados
            grupos = {}
            linhas_ignoradas = 0
            linhas_canceladas = 0

            def get(row, chave):
                i = col.get(chave)
                if i is None or i >= len(row):
                    return None
                return row[i]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue

                status_raw = _celula_texto(get(row, 'status'), 50)
                status = status_raw.lower()
                if status in status_ignorar:
                    linhas_ignoradas += 1
                    continue

                paciente = _celula_texto(get(row, 'paciente'), 200)
                procedimento = _celula_texto(get(row, 'procedimento'), 200)
                if not paciente or not procedimento:
                    linhas_ignoradas += 1
                    continue

                data_fat = _parse_data_ris(get(row, 'data'))
                cpf = _celula_texto(get(row, 'cpf'), 50)
                cns = _celula_texto(get(row, 'cns'), 50)
                convenio = _celula_texto(get(row, 'convenio'), 100) or 'Particular'
                # Carteirinha = apenas CNS (não usar CPF)
                carteirinha = cns or None
                valor = _parse_valor_ris(get(row, 'valor'))
                eh_cancelado = _eh_status_agendamento_cancelado(status_raw)
                if eh_cancelado:
                    linhas_canceladas += 1

                # Separa cancelados da lista principal no agrupamento
                chave = f"{paciente}|{data_fat.isoformat()}|{cpf}|{convenio}|{status_raw or 'ok'}"
                modalidade = _celula_texto(get(row, 'modalidade'), 20)
                agendado_via = _celula_texto(get(row, 'agendado_via'), 50) or 'RIS'
                if chave not in grupos:
                    horario_inicio = _celula_texto(get(row, 'horario_inicio'), 20) or None
                    horario_fim = _celula_texto(get(row, 'horario_fim'), 20) or None
                    horario = ''
                    if horario_inicio or horario_fim:
                        horario = f"{(horario_inicio or '')} - {(horario_fim or '')}".strip(' -')
                    status_agendamento = status_raw or None
                    indicacao = _celula_texto(get(row, 'indicacao')) or None
                    descricao = _celula_texto(get(row, 'descricao')) or None
                    obs_pag = _celula_texto(get(row, 'obs_pagamento'))
                    observacao = f"Pagamento: {obs_pag}" if obs_pag else None

                    prioridade = _celula_texto(get(row, 'prioridade'), 50) or None
                    urgencia = 'Não'
                    if prioridade and prioridade.lower() not in ('eletivo', ''):
                        urgencia = 'Sim'

                    grupos[chave] = {
                        'lote': _celula_texto(get(row, 'lote'), 50) or None,
                        'carteirinha': carteirinha,
                        'cpf': cpf or None,
                        'horario': horario or None,
                        'horario_inicio': horario_inicio,
                        'horario_fim': horario_fim,
                        'prioridade': prioridade,
                        'status_agendamento': status_agendamento,
                        'motivo_cancelamento': _celula_texto(get(row, 'motivo_cancelamento'), 255) or None,
                        'nome': paciente,
                        'nome_associado': paciente,
                        'data': data_fat,
                        'local': _celula_texto(get(row, 'unidade'), 200) or None,
                        'medico': _celula_texto(get(row, 'medico'), 200) or None,
                        'medico_solicitante': _celula_texto(get(row, 'medico_solicitante'), 200) or None,
                        'tecnico': _celula_texto(get(row, 'tecnico'), 200) or None,
                        'checkin_por': _celula_texto(get(row, 'checkin_por'), 200) or None,
                        'agendado_por': _celula_texto(get(row, 'agendado_por'), 200) or None,
                        'convenio': convenio,
                        'tag': _celula_texto(get(row, 'tag'), 100) or None,
                        'indicacao_clinica': indicacao,
                        'descricao': descricao,
                        'agendado_via': agendado_via,
                        'urgencia': urgencia,
                        'observacao': observacao,
                        'servicos': [],
                    }

                grupos[chave]['servicos'].append({
                    'descricao': procedimento,
                    'modalidade': modalidade,
                    'com_contraste': 'contraste' in procedimento.lower(),
                    'valor': valor,
                    'total': valor,
                })

            faturamentos_criados = 0
            itens_criados = 0

            for dados in grupos.values():
                faturamento = FaturamentoMedico.objects.create(
                    empresa_id=empresa_id,
                    lote=dados['lote'],
                    carteirinha=dados['carteirinha'],
                    cpf=dados['cpf'],
                    horario=dados['horario'],
                    horario_inicio=dados['horario_inicio'],
                    horario_fim=dados['horario_fim'],
                    prioridade=dados['prioridade'],
                    status_agendamento=dados['status_agendamento'],
                    motivo_cancelamento=dados['motivo_cancelamento'],
                    nome=dados['nome'],
                    nome_associado=dados.get('nome_associado') or dados['nome'],
                    data=dados['data'],
                    local=dados['local'],
                    medico=dados['medico'],
                    medico_solicitante=dados['medico_solicitante'],
                    tecnico=dados['tecnico'],
                    checkin_por=dados['checkin_por'],
                    agendado_por=dados['agendado_por'],
                    convenio=dados['convenio'],
                    tag=dados['tag'],
                    indicacao_clinica=dados['indicacao_clinica'],
                    descricao=dados['descricao'],
                    agendado_via=dados['agendado_via'],
                    urgencia=dados['urgencia'],
                    observacao=dados['observacao'],
                    codigo_relatorio=None,
                    status='pendente',
                )

                for servico in dados['servicos']:
                    ItemServico.objects.create(
                        faturamento=faturamento,
                        codigo_servico='',
                        servico=servico['descricao'],
                        modalidade=servico['modalidade'] or None,
                        com_contraste=servico['com_contraste'],
                        porte='',
                        qt=1,
                        valor=servico['valor'],
                        total=servico['total'],
                    )
                    itens_criados += 1

                faturamento.atualizar_total()
                faturamentos_criados += 1

            msg = (
                f'Importação RIS concluída! '
                f'{faturamentos_criados} faturamentos criados, '
                f'{itens_criados} itens de serviço criados.'
            )
            if linhas_canceladas:
                msg += (
                    f' {linhas_canceladas} linhas Cancelado/Desistência/Deletado '
                    f'(disponíveis em Procedimentos Cancelados).'
                )
            if linhas_ignoradas:
                msg += f' {linhas_ignoradas} linhas ignoradas (vazias).'
            messages.success(request, msg)

        except Exception as e:
            logger.exception('Erro ao importar relatório RIS')
            messages.error(request, f'Erro durante a importação RIS: {str(e)}')
            return redirect('faturamento_medico:importar_ris')

        return redirect('faturamento_medico:ftlistar')

    context = {
        'titulo': 'Importar Relatório RIS'
    }

    return render(request, 'faturamento_medico/importar_ris.html', context)


def sincronizar_medcloud(request):
    """Importa agendas concluídas e/ou links de laudo via API MedCloud."""
    empresa_id = request.session.get('empresa_id')
    try:
        empresa_id = int(empresa_id) if empresa_id is not None else None
    except (TypeError, ValueError):
        empresa_id = None

    if not empresa_id:
        messages.error(request, 'Selecione uma empresa antes de sincronizar com a MedCloud.')
        return redirect('faturamento_medico:ftlistar')

    empresa = get_object_or_404(Empresa, pk=empresa_id)
    hoje = timezone.localdate()
    convenios = Convenio.objects.filter(empresa_id=empresa_id).order_by('nome')

    if request.method == 'POST':
        from faturamento_medico.medcloud.client import MedcloudAPIError
        from faturamento_medico.medcloud.sync import (
            sincronizar_agendas_concluidas,
            sincronizar_links_laudos,
        )

        acao = (request.POST.get('acao') or 'ambos').strip()
        convenio = (request.POST.get('convenio') or '').strip() or None

        def _parse_data_campo(nome, padrao):
            raw = (request.POST.get(nome) or '').strip()
            if not raw:
                return padrao
            try:
                return datetime.strptime(raw, '%Y-%m-%d').date()
            except ValueError:
                return padrao

        data_inicio = _parse_data_campo('data_inicio', hoje)
        data_fim = _parse_data_campo('data_fim', data_inicio)
        if data_fim < data_inicio:
            data_inicio, data_fim = data_fim, data_inicio

        partes_msg = []
        try:
            if acao in ('agendas', 'ambos'):
                stats = sincronizar_agendas_concluidas(
                    empresa,
                    data_inicio,
                    data_fim,
                    convenio_nome=convenio,
                )
                partes_msg.append(
                    f'Agendas: {stats["listadas"]} listadas, '
                    f'{stats.get("importadas", stats.get("concluidas", 0))} importadas, '
                    f'{stats["criadas"]} criadas, {stats["atualizadas"]} atualizadas, '
                    f'{stats["ignoradas"]} ignoradas.'
                )
            if acao in ('laudos', 'ambos'):
                stats = sincronizar_links_laudos(
                    empresa,
                    data_inicio,
                    data_fim,
                    convenio_nome=convenio,
                )
                partes_msg.append(
                    f'Laudos: {stats["atualizados"]} links gravados, '
                    f'{stats["sem_laudo"]} ainda sem laudo, '
                    f'{stats["pulados_link_valido"]} já com link válido.'
                )
            messages.success(request, 'Sincronização MedCloud concluída. ' + ' '.join(partes_msg))
        except MedcloudAPIError as exc:
            messages.error(request, f'MedCloud: {exc}')
        except Exception as exc:
            logger.exception('Erro na sincronização MedCloud')
            messages.error(request, f'Erro na sincronização MedCloud: {exc}')

        return redirect('faturamento_medico:sincronizar_medcloud')

    context = {
        'titulo': 'Sincronizar MedCloud (API)',
        'convenios': convenios,
        'data_hoje': hoje.isoformat(),
    }
    return render(request, 'faturamento_medico/sincronizar_medcloud.html', context)


def toggle_conferencia_item(request, pk):
    """Marca/desmarca conferência de um item de serviço (AJAX)."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'erro': 'Método não permitido'}, status=405)

    empresa_id = request.session.get('empresa_id')
    try:
        empresa_id = int(empresa_id) if empresa_id is not None else None
    except (TypeError, ValueError):
        empresa_id = None

    item = get_object_or_404(ItemServico.objects.select_related('faturamento'), pk=pk)
    if empresa_id is not None and item.faturamento.empresa_id != empresa_id:
        return JsonResponse({'ok': False, 'erro': 'Sem permissão'}, status=403)

    # Aceita estado explícito do checkbox; senão faz toggle
    conferido_param = request.POST.get('conferido')
    if conferido_param is not None:
        item.conferido = str(conferido_param).lower() in ('1', 'true', 'sim', 'yes')
    else:
        item.conferido = not item.conferido

    if item.conferido:
        item.status_conferencia = 'CONFERIDO'
    elif item.status_conferencia == 'CONFERIDO':
        item.status_conferencia = 'PENDENTE'

    item.save(update_fields=['conferido', 'status_conferencia', 'total'])
    status_label, status_css = item.status_conferencia_badge()
    return JsonResponse({
        'ok': True,
        'conferido': item.conferido,
        'status': status_label,
        'status_css': status_css,
    })


def alterar_status_conferencia_item(request, pk):
    """Altera manualmente o status de conferência (AJAX)."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'erro': 'Método não permitido'}, status=405)

    empresa_id = request.session.get('empresa_id')
    try:
        empresa_id = int(empresa_id) if empresa_id is not None else None
    except (TypeError, ValueError):
        empresa_id = None

    item = get_object_or_404(ItemServico.objects.select_related('faturamento'), pk=pk)
    if empresa_id is not None and item.faturamento.empresa_id != empresa_id:
        return JsonResponse({'ok': False, 'erro': 'Sem permissão'}, status=403)

    status = request.POST.get('status') or ''
    status_label, status_css = item.aplicar_status_conferencia(status)
    return JsonResponse({
        'ok': True,
        'conferido': item.conferido,
        'status': status_label,
        'status_css': status_css,
    })


def lancar_glosa_item(request, pk):
    """Registra glosa no procedimento e atualiza o total glosado no extrato do lote."""
    from emprestimos.sicoob_pdf import _dec

    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    item = get_object_or_404(
        ItemServico.objects.select_related('faturamento'),
        pk=pk,
        faturamento__empresa_id=empresa_id,
    )
    faturamento = item.faturamento
    lote_ref = (faturamento.lote or '').strip()
    lote = None
    if lote_ref:
        try:
            lote = Lote.objects.filter(pk=int(lote_ref), empresa_id=empresa_id).first()
        except (TypeError, ValueError):
            lote = None

    if request.method == 'POST':
        valor_raw = (request.POST.get('valor_glosa') or '').strip()
        if not valor_raw:
            messages.error(request, 'Informe o valor da glosa.')
            return redirect('faturamento_medico:lancar_glosa_item', pk=pk)

        item.valor_glosa = _dec(valor_raw)

        data_rec_raw = (request.POST.get('data_recorrencia') or '').strip()
        if data_rec_raw:
            try:
                item.data_recorrencia = date.fromisoformat(data_rec_raw)
            except ValueError:
                messages.error(request, 'Data de recorrência inválida.')
                return redirect('faturamento_medico:lancar_glosa_item', pk=pk)
        else:
            item.data_recorrencia = None

        item.save(update_fields=['valor_glosa', 'data_recorrencia'])

        extrato = None
        if lote:
            extrato = lote.recalcular_glosa_extrato()
            messages.success(
                request,
                f'Glosa registrada. Total glosado do lote {lote.id} atualizado para R$ {_moeda_br(extrato.valor_glosado if extrato else item.valor_glosa)}.',
            )
        else:
            messages.warning(
                request,
                'Glosa registrada no procedimento, mas o faturamento não está vinculado a um lote — o extrato não foi atualizado.',
            )

        voltar = (request.POST.get('voltar') or '').strip()
        if voltar.startswith('/'):
            return redirect(voltar)
        return redirect('faturamento_medico:ftlistar')

    valor_item = item.total if item.total is not None else (item.valor or 0)
    context = {
        'titulo': 'Lançar Glosa — Procedimento',
        'item': item,
        'faturamento': faturamento,
        'lote': lote,
        'valor_item_fmt': _moeda_br(valor_item),
        'valor_glosa_fmt': _moeda_br(item.valor_glosa),
        'voltar': request.GET.get('next') or request.META.get('HTTP_REFERER') or '',
    }
    return render(request, 'faturamento_medico/lancar_glosa_item.html', context)


def listar_extrato_pagamento(request):
    """Lista extrato de pagamento importado por convênio."""
    from django.db.models import Sum

    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    qs = ExtratoPagamentoConvenio.objects.filter(empresa_id=empresa_id)

    competencia = (request.GET.get('competencia') or '').strip()
    convenio = (request.GET.get('convenio') or '').strip()
    if competencia:
        qs = qs.filter(competencia=competencia)
    if convenio:
        qs = qs.filter(_q_convenio_filtro(convenio))

    totais = qs.aggregate(
        total_valor=Sum('valor'),
        total_glosado=Sum('valor_glosado'),
        total_liberado=Sum('valor_liberado'),
        total_retencoes=Sum('retencoes'),
        total_recebido=Sum('valor_recebido'),
    )

    competencias = (
        ExtratoPagamentoConvenio.objects.filter(empresa_id=empresa_id)
        .exclude(competencia='')
        .values_list('competencia', flat=True)
        .distinct()
        .order_by('-competencia')
    )

    context = {
        'titulo': 'Extrato de Pagamento — Convênio',
        'linhas': qs.order_by('-data_recebimento', '-competencia', 'lote'),
        'filtros': {'competencia': competencia, 'convenio': convenio},
        'competencias': competencias,
        'totais': totais,
    }
    return render(request, 'faturamento_medico/listar_extrato_pagamento.html', context)


def importar_extrato_pagamento_bradesco(request):
    """Importa Demonstrativo de Pagamento TISS Bradesco Saúde (PDF)."""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    def _parse_data_sessao(val):
        if not val:
            return None
        if isinstance(val, date):
            return val
        try:
            return date.fromisoformat(str(val)[:10])
        except (TypeError, ValueError):
            return None

    preview = None
    if request.method == 'POST':
        acao = request.POST.get('acao', 'preview')
        competencia = (request.POST.get('competencia') or '').strip()

        if acao == 'confirmar':
            dados_json = request.session.pop('extrato_pagamento_preview', None)
            if not dados_json:
                messages.error(request, 'Prévia expirada. Importe o PDF novamente.')
                return redirect('faturamento_medico:importar_extrato_pagamento_bradesco')
            competencia_sessao = (dados_json.get('competencia') or competencia or '').strip()
            if not competencia_sessao:
                messages.error(request, 'Competência não informada. Refaça a importação informando MM/AAAA.')
                return redirect('faturamento_medico:importar_extrato_pagamento_bradesco')
            criados = 0
            ignorados = 0
            for row in dados_json.get('linhas', []):
                try:
                    ExtratoPagamentoConvenio.objects.create(
                        empresa_id=empresa_id,
                        competencia=competencia_sessao,
                        convenio=row.get('convenio') or 'BRADESCO SAUDE',
                        data_lote=_parse_data_sessao(row.get('data_lote')),
                        lote=row.get('lote') or '',
                        protocolo=row.get('protocolo') or '',
                        qt_guias=row.get('qt_guias'),
                        valor=Decimal(str(row.get('valor') or 0)),
                        valor_processado=Decimal(str(row.get('valor_processado') or 0)),
                        valor_glosado=Decimal(str(row.get('valor_glosado') or 0)),
                        valor_liberado=Decimal(str(row.get('valor_liberado') or 0)),
                        retencoes=Decimal(str(row.get('retencoes') or 0)),
                        liquido=Decimal(str(row.get('liquido') or 0)),
                        data_previsao=_parse_data_sessao(row.get('data_previsao')),
                        numero_demonstrativo=row.get('numero_demonstrativo') or '',
                        nome_arquivo=dados_json.get('nome_arquivo') or '',
                    )
                    criados += 1
                except Exception:
                    ignorados += 1
            messages.success(
                request,
                f'Importação concluída: {criados} protocolo(s) gravado(s) na competência {competencia_sessao}'
                + (f', {ignorados} duplicado(s) ignorado(s).' if ignorados else '.'),
            )
            return redirect('faturamento_medico:listar_extrato_pagamento')

        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            messages.error(request, 'Selecione o PDF do demonstrativo.')
            return redirect('faturamento_medico:importar_extrato_pagamento_bradesco')

        try:
            from .bradesco_tiss_pdf import parse_extrato_pagamento_bradesco, _parse_competencia

            parsed = parse_extrato_pagamento_bradesco(
                arquivo,
                competencia=competencia,
                nome_arquivo=arquivo.name,
            )
            competencia_final = (
                _parse_competencia(competencia)
                or parsed['cabecalho'].get('competencia')
                or _parse_competencia(arquivo.name)
                or ''
            )
            linhas_sessao = []
            for row in parsed['linhas']:
                row_copy = dict(row)
                row_copy['competencia'] = competencia_final
                linhas_sessao.append({
                    k: (
                        v.isoformat() if hasattr(v, 'isoformat')
                        else float(v) if isinstance(v, Decimal)
                        else v
                    )
                    for k, v in row_copy.items()
                })
            cab_serializado = {
                k: (v.isoformat() if hasattr(v, 'isoformat') else v)
                for k, v in parsed['cabecalho'].items()
            }
            request.session['extrato_pagamento_preview'] = {
                'cabecalho': cab_serializado,
                'linhas': linhas_sessao,
                'nome_arquivo': arquivo.name,
                'competencia': competencia_final,
            }
            linhas_preview = []
            for row in parsed['linhas']:
                r = dict(row)
                r['competencia'] = competencia_final
                linhas_preview.append(r)
            preview = {
                'cabecalho': parsed['cabecalho'],
                'linhas': linhas_preview,
                'nome_arquivo': arquivo.name,
                'competencia': competencia_final,
                'qtd_protocolos': len(parsed['linhas']),
            }
        except Exception as e:
            logger.exception('Erro ao importar extrato pagamento Bradesco')
            messages.error(request, f'Erro ao ler PDF: {e}')
            return redirect('faturamento_medico:importar_extrato_pagamento_bradesco')

    context = {
        'titulo': 'Importar Extrato de Pagamento — Bradesco Saúde',
        'preview': preview,
        'competencia_padrao': (request.POST.get('competencia') or request.GET.get('competencia') or '').strip(),
    }
    return render(request, 'faturamento_medico/importar_extrato_pagamento.html', context)


def baixar_extrato_pagamento(request, pk):
    """Baixa manual — concilia recebimento com extrato bancário."""
    from extrato.models import Banco

    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    extrato = get_object_or_404(ExtratoPagamentoConvenio, pk=pk, empresa_id=empresa_id)
    bancos = Banco.objects.order_by('nome')

    if request.method == 'POST':
        data_receb = request.POST.get('data_recebimento')
        valor_raw = (request.POST.get('valor_recebido') or '').strip()
        banco = (request.POST.get('banco') or '').strip()

        if not data_receb or not valor_raw or not banco:
            messages.error(request, 'Informe data de recebimento, valor recebido e banco.')
            return redirect('faturamento_medico:baixar_extrato_pagamento', pk=pk)

        try:
            from emprestimos.sicoob_pdf import _dec
            extrato.data_recebimento = date.fromisoformat(data_receb)
            extrato.valor_recebido = _dec(valor_raw)
            extrato.banco = banco
            extrato.save(update_fields=['data_recebimento', 'valor_recebido', 'banco', 'data_atualizacao'])
            extrato.sincronizar_baixado_lote()
            n_final = FaturamentoMedico.objects.filter(
                empresa_id=empresa_id,
                lote=str(extrato.lote_faturamento_id),
                status='finalizado',
            ).count() if extrato.lote_faturamento_id else 0
            msg = 'Baixa registrada com sucesso.'
            if n_final:
                msg += f' {n_final} faturamento(s) marcado(s) como Finalizado.'
            messages.success(request, msg)
            return redirect('faturamento_medico:listar_extrato_pagamento')
        except (ValueError, InvalidOperation):
            messages.error(request, 'Data ou valor inválido.')
            return redirect('faturamento_medico:baixar_extrato_pagamento', pk=pk)

    sugestao_valor = extrato.liquido or extrato.valor_liberado or Decimal('0')
    context = {
        'titulo': 'Baixar recebimento — Extrato convênio',
        'extrato': extrato,
        'bancos': bancos,
        'sugestao_valor': sugestao_valor,
        'data_hoje': timezone.now().date().isoformat(),
    }
    return render(request, 'faturamento_medico/baixar_extrato_pagamento.html', context)


def estornar_baixa_extrato_pagamento(request, pk):
    """Remove baixa para permitir nova conciliação."""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    if request.method != 'POST':
        return redirect('faturamento_medico:listar_extrato_pagamento')

    extrato = get_object_or_404(ExtratoPagamentoConvenio, pk=pk, empresa_id=empresa_id)
    extrato.data_recebimento = None
    extrato.valor_recebido = Decimal('0')
    extrato.banco = ''
    extrato.save(update_fields=['data_recebimento', 'valor_recebido', 'banco', 'data_atualizacao'])
    extrato.sincronizar_baixado_lote()
    messages.success(request, 'Baixa estornada. Faturamentos do lote voltaram para Aguardando pagamento.')
    return redirect('faturamento_medico:listar_extrato_pagamento')


def _redirect_listar_extrato_pagamento(request):
    from urllib.parse import urlencode

    params = {}
    competencia = (request.GET.get('competencia') or request.POST.get('competencia') or '').strip()
    convenio = (request.GET.get('convenio') or request.POST.get('convenio') or '').strip()
    if competencia:
        params['competencia'] = competencia
    if convenio:
        params['convenio'] = convenio
    url = reverse('faturamento_medico:listar_extrato_pagamento')
    if params:
        url += '?' + urlencode(params)
    return redirect(url)


def editar_extrato_pagamento(request, pk):
    """Edita protocolo, nota fiscal e valores complementares do extrato."""
    from emprestimos.sicoob_pdf import _dec

    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    extrato = get_object_or_404(ExtratoPagamentoConvenio, pk=pk, empresa_id=empresa_id)

    if request.method == 'POST':
        extrato.protocolo = (request.POST.get('protocolo') or '').strip()
        extrato.nota = (request.POST.get('nota') or '').strip()

        valor_nota_raw = (request.POST.get('valor_nota') or '').strip()
        if valor_nota_raw:
            extrato.valor_nota = _dec(valor_nota_raw)
        else:
            extrato.valor_nota = None

        extrato.retencoes = _dec(request.POST.get('retencoes'))
        extrato.liquido = _dec(request.POST.get('liquido'))

        data_previsao_raw = (request.POST.get('data_previsao') or '').strip()
        if data_previsao_raw:
            try:
                extrato.data_previsao = date.fromisoformat(data_previsao_raw)
            except ValueError:
                messages.error(request, 'Data de previsão inválida.')
                return redirect('faturamento_medico:editar_extrato_pagamento', pk=pk)
        else:
            extrato.data_previsao = None

        extrato.save(update_fields=[
            'protocolo', 'nota', 'valor_nota', 'retencoes', 'liquido',
            'data_previsao', 'data_atualizacao',
        ])
        messages.success(request, 'Extrato atualizado com sucesso.')
        return _redirect_listar_extrato_pagamento(request)

    context = {
        'titulo': 'Editar Extrato de Pagamento — Convênio',
        'extrato': extrato,
        'filtros': {
            'competencia': (request.GET.get('competencia') or '').strip(),
            'convenio': (request.GET.get('convenio') or '').strip(),
        },
    }
    return render(request, 'faturamento_medico/editar_extrato_pagamento.html', context)

