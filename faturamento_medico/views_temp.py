from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from django.http import HttpResponse, Http404, JsonResponse
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import uuid
import logging
from datetime import datetime, date, timedelta
from urllib.parse import urlencode
from .models import FaturamentoMedico, DocumentoAnexado, ItemServico, ServicoDisponivel, Lote
from servicos_medicos.models import Convenio
from empresa.models import Empresa
from .forms import FaturamentoMedicoForm, DocumentoAnexadoForm, ItemServicoForm, ItemServicoFormSet, ServicoDisponivelForm
from .utils import processar_arquivos_com_gemini, processar_arquivos_com_ocr

logger = logging.getLogger(__name__)


def verificar_empresa_selecionada(request):
    """
    Verifica se há uma empresa selecionada na sessão e retorna informações úteis.
    Retorna: (empresa_id, empresa_nome, error_message)
    """
    empresa_id = request.session.get('empresa_id')
    empresa_nome = request.session.get('empresa_nome')
    
    if not empresa_id:
        # Verificar se existem empresas disponíveis para o usuário
        from empresa.models import Empresa, UsuarioEmpresa
        try:
            usuario_empresas = UsuarioEmpresa.objects.filter(
                usuario=request.user,
                ativo=True,
                empresa__status='Ativa'
            ).select_related('empresa')
            
            if not usuario_empresas.exists():
                return None, None, 'Nenhuma empresa disponível. Entre em contato com o administrador para obter acesso a uma empresa.'
            elif usuario_empresas.count() == 1:
                # Se há apenas uma empresa disponível, seleciona automaticamente
                ue = usuario_empresas.first()
                request.session['empresa_id'] = ue.empresa.id
                request.session['empresa_nome'] = ue.empresa.razao
                return ue.empresa.id, ue.empresa.razao, None
            else:
                # Múltiplas empresas disponíveis, precisa selecionar
                return None, None, 'Selecione uma empresa para continuar. <a href="/empresa/lista/" class="alert-link">Clique aqui para selecionar</a>.'
        except:
            return None, None, 'Erro ao verificar empresas disponíveis. <a href="/empresa/lista/" class="alert-link">Verificar empresas</a>.'
    
    return empresa_id, empresa_nome, None


def listar_faturamentos(request):
    """Lista todos os faturamentos médicos com filtros"""
    empresa_id = request.session.get('empresa_id')
    if empresa_id:
        faturamentos = FaturamentoMedico.objects.filter(empresa_id=empresa_id).order_by('-data')
    else:
        faturamentos = FaturamentoMedico.objects.all().order_by('-data')

    # Filtros
    nome = request.GET.get('nome')
    guia = request.GET.get('guia')
    anestesista = request.GET.get('anestesista')
    status = request.GET.get('status')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    convenios = request.GET.getlist('convenio')
    codigo_relatorio =  request.GET.get('codigo_relatorio')

    # Definir valores padrão para datas se não fornecidos
    hoje = date.today()
    if not data_inicio:
        data_inicio = hoje.replace(day=1).strftime('%Y-%m-%d')  # Primeiro dia do mês
    if not data_fim:
        # Último dia do mês
        proximo_mes = hoje.replace(day=28) + timedelta(days=4)  # Garante que vai para o próximo mês
        data_fim = (proximo_mes - timedelta(days=proximo_mes.day)).strftime('%Y-%m-%d')

    if nome:
        faturamentos = faturamentos.filter(Q(nome__icontains=nome))
    if guia:
        faturamentos = faturamentos.filter(guia__icontains=guia)
    if codigo_relatorio:
       faturamentos = faturamentos.filter(codigo_relatorio__icontains=codigo_relatorio  )     
    if anestesista:
        faturamentos = faturamentos.filter(anestesista__icontains=anestesista)
    if status:
        faturamentos = faturamentos.filter(status=status)
    if data_inicio:
        faturamentos = faturamentos.filter(data__gte=data_inicio)
    if data_fim:
        faturamentos = faturamentos.filter(data__lte=data_fim)
    if convenios:
        # Filtrar por múltiplos convênios
        q_objects = Q()
        for conv in convenios:
            if conv:  # Ignorar valores vazios
                q_objects |= Q(convenio__icontains=conv)
        faturamentos = faturamentos.filter(q_objects)

    # Estatísticas
    total_faturamentos = faturamentos.count()
    valor_total = sum(f.total for f in faturamentos if f.total)

    # Estatísticas por convênio
    from django.db.models import Sum, Count
    stats_convenio = faturamentos.values('convenio').annotate(
        total_valor=Sum('total'),
        quantidade=Count('id')
    ).order_by('-total_valor')

    # Estatísticas por anestesista
    stats_anestesista = faturamentos.values('anestesista').annotate(
        total_valor=Sum('total'),
        quantidade=Count('id')
    ).exclude(anestesista__isnull=True).exclude(anestesista='').order_by('-total_valor')

    # Buscar convênios disponíveis para a empresa
    convenios_disponiveis = []
    if empresa_id:
        from servicos_medicos.models import Convenio
        convenios_disponiveis = list(Convenio.objects.filter(empresa_id=empresa_id).order_by('nome'))
        if not convenios_disponiveis:
            # Convênios padrão se nenhum for encontrado para a empresa
            convenios_padrao = [
                {'nome': 'CBSAUDE'},
                {'nome': 'PM'},
                {'nome': 'UNIMED'},
                {'nome': 'BRADESCO'},
                {'nome': 'GEAP'},
                {'nome': 'SAUDE CAIXA'},
                {'nome': 'POSTAL SAUDE'},
                {'nome': 'FUSEX'},
                {'nome': 'LIFE EMPRESARIAL'},
                {'nome': 'CASSI'},
                {'nome': 'GCARD'},
                {'nome': 'PERSONAL NET'},
            ]
            convenios_disponiveis = convenios_padrao

    # Buscar lotes disponíveis para a empresa
    lotes_disponiveis = []
    if empresa_id:
        lotes_disponiveis = Lote.objects.filter(empresa_id=empresa_id).order_by('-id')

    # Armazenar filtros na sessão para preservar ao voltar de edição
    request.session['faturamento_filters'] = {
        'nome': nome or '',
        'guia': guia or '',
        'anestesista': anestesista or '',
        'status': status or '',
        'data_inicio': data_inicio or '',
        'data_fim': data_fim or '',
        'convenio': convenios or [],
    }

    context = {
        'faturamentos': faturamentos,
        'total_faturamentos': total_faturamentos,
        'valor_total': valor_total,
        'stats_convenio': stats_convenio,
        'stats_anestesista': stats_anestesista,
        'convenios_disponiveis': convenios_disponiveis,
        'lotes_disponiveis': lotes_disponiveis,
        'filtros': {
            'nome': nome,
            'guia': guia,
            'anestesista': anestesista,
            'status': status,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'convenio': convenios,
        }
    }

    return render(request, 'faturamento_medico/listar.html', context)


def criar_faturamento(request):
    """Cria um novo faturamento médico"""
    logger.info("Iniciando criar_faturamento")
    empresa_id = request.session.get('empresa_id')
    logger.info(f"Empresa ID da sessão: {empresa_id}")
    if not empresa_id:
        logger.warning("Empresa não encontrada na sessão")
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    if request.method == 'POST':
        logger.info("Método POST detectado")
        form = FaturamentoMedicoForm(request.POST, request.FILES, empresa_id=empresa_id)
        logger.info(f"Form criado: {form}")
        if form.is_valid():
            logger.info("Form é válido")

            # Processar arquivos com Gemini se foram enviados
            documentos_gemini = request.FILES.getlist('documentos_gemini')
            documento_upload = request.FILES.get('documento_upload')
            dados_gemini = {}

            # Processar documento_upload se enviado
            if documento_upload:
                logger.info("Processando documento_upload com Gemini")
                dados_gemini = processar_arquivos_com_gemini([documento_upload])
                logger.info(f"Dados extraídos do documento_upload: {dados_gemini}")

            # Processar documentos_gemini adicionais se enviados
            elif documentos_gemini:
                logger.info(f"Processando {len(documentos_gemini)} arquivos com Gemini")
                dados_gemini = processar_arquivos_com_gemini(documentos_gemini)
                logger.info(f"Dados extraídos do Gemini: {dados_gemini}")

            faturamento = form.save(commit=False)
            faturamento.empresa_id = empresa_id

            # Preencher campos com dados do Gemini se disponíveis
            if dados_gemini.get('nome'):
                faturamento.nome = dados_gemini['nome']
            if dados_gemini.get('carteirinha'):
                faturamento.carteirinha = dados_gemini['carteirinha']
            if dados_gemini.get('guia'):
                faturamento.guia = dados_gemini['guia']
            if dados_gemini.get('numero_guia_lancada'):
                faturamento.numero_guia_lancada = dados_gemini['numero_guia_lancada']
            if dados_gemini.get('data_autorizacao'):
                # Tentar converter data se possível
                try:
                    from datetime import datetime
                    faturamento.data_autorizacao = datetime.strptime(dados_gemini['data_autorizacao'], '%d/%m/%Y').date()
                except:
                    pass
            if dados_gemini.get('data_internacao_cirurgia'):
                # Tentar converter data se possível
                try:
                    from datetime import datetime
                    faturamento.data = datetime.strptime(dados_gemini['data_internacao_cirurgia'], '%d/%m/%Y').date()
                except:
                    pass
            if dados_gemini.get('local'):
                faturamento.local = dados_gemini['local']
            if dados_gemini.get('medico'):
                faturamento.medico = dados_gemini['medico']
            if dados_gemini.get('anestesista'):
                faturamento.anestesista = dados_gemini['anestesista']
            if dados_gemini.get('convenio'):
                faturamento.convenio = dados_gemini['convenio']
            if dados_gemini.get('apartamento_enfermaria'):
                faturamento.apartamento_enfermaria = dados_gemini['apartamento_enfermaria']
            if dados_gemini.get('urgencia'):
                faturamento.urgencia = dados_gemini['urgencia']

            faturamento.save()
            logger.info(f"Faturamento salvo: {faturamento.id}")

            # Criar itens de serviço baseados nos dados do Gemini
            if dados_gemini.get('servicos'):
                for servico in dados_gemini['servicos']:
                    ItemServico.objects.create(
                        faturamento=faturamento,
                        servico=servico.get('descricao', ''),
                        codigo_servico=servico.get('codigo', ''),
                        valor=servico.get('valor_unitario', 0),
                        qt=servico.get('quantidade', 1)
                    )

            # Anexar documentos processados
            documentos_para_anexar = []
            if documento_upload:
                documentos_para_anexar.append(documento_upload)
            documentos_para_anexar.extend(documentos_gemini)

            for documento in documentos_para_anexar:
                DocumentoAnexado.objects.create(
                    faturamento=faturamento,
                    arquivo=documento,
                    nome=f"Documento Gemini - {documento.name}",
                    descricao="Documento processado com Gemini para extração de dados"
                )

            # Adicionar mensagem de sucesso específica se Gemini foi usado
            if documentos_gemini and dados_gemini:
                messages.success(request, f'Faturamento médico criado com sucesso! {len(documentos_gemini)} documento(s) processado(s) com Gemini.')
            else:
                messages.success(request, 'Faturamento médico criado com sucesso!')
    
                messages.success(request, 'Faturamento médico criado com sucesso!')
                # Redirecionar com filtros preservados
                filters = request.session.get('faturamento_filters', {})
                url = reverse('faturamento_medico:ftlistar')
                if filters:
                    url += '?' + urlencode(filters, doseq=True)
                return redirect(url)
        else:
            logger.warning(f"Form inválido: {form.errors}")
    else:
        logger.info("Método GET detectado")
        initial_data = {
            'data_autorizacao': timezone.now().date(),
            'data': timezone.now().date(),
        }
        form = FaturamentoMedicoForm(empresa_id=empresa_id, initial=initial_data)

    context = {
        'form': form,
        'titulo': 'Criar Faturamento Médico'
    }

    logger.info("Renderizando template form.html")
    return render(request, 'faturamento_medico/form.html', context)


def editar_faturamento(request, pk):
    """Edita um faturamento médico existente"""
    faturamento = get_object_or_404(FaturamentoMedico, pk=pk)
    empresa_id = request.session.get('empresa_id')

    if request.method == 'POST':
        form = FaturamentoMedicoForm(request.POST, instance=faturamento, empresa_id=empresa_id)
        if form.is_valid():
            form.save()
            messages.success(request, 'Faturamento médico atualizado com sucesso!')
            # Redirecionar com filtros preservados
            filters = request.session.get('faturamento_filters', {})
            url = reverse('faturamento_medico:ftlistar')
            if filters:
                url += '?' + urlencode(filters, doseq=True)
            return redirect(url)
    else:
        form = FaturamentoMedicoForm(instance=faturamento, empresa_id=empresa_id)

    context = {
        'form': form,
        'faturamento': faturamento,
        'titulo': 'Editar Faturamento Médico'
    }

    return render(request, 'faturamento_medico/form.html', context)


def excluir_faturamento(request, pk):
    """Exclui um faturamento médico"""
    faturamento = get_object_or_404(FaturamentoMedico, pk=pk)

    if request.method == 'POST':
        faturamento.delete()
        messages.success(request, 'Faturamento médico excluído com sucesso!')
        # Redirecionar com filtros preservados
        filters = request.session.get('faturamento_filters', {})
        url = reverse('faturamento_medico:ftlistar')
        if filters:
            url += '?' + urlencode(filters, doseq=True)
        return redirect(url)

    context = {
        'faturamento': faturamento,
    }

    return render(request, 'faturamento_medico/confirmar_exclusao.html', context)


def detalhes_faturamento(request, pk):
    """Exibe detalhes de um faturamento médico"""
    faturamento = get_object_or_404(FaturamentoMedico, pk=pk)

    context = {
        'faturamento': faturamento,
    }

    return render(request, 'faturamento_medico/detalhes.html', context)


def exportar_excel(request):
    """Exporta os faturamentos filtrados para Excel"""
    empresa_id = request.session.get('empresa_id')
    if empresa_id:
        faturamentos = FaturamentoMedico.objects.filter(empresa_id=empresa_id).order_by('-data')
    else:
        faturamentos = FaturamentoMedico.objects.none()

    # Aplicar os mesmos filtros da view de listagem
    nome = request.GET.get('nome')
    guia = request.GET.get('guia')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    convenios = request.GET.getlist('convenio')
    anestesista = request.GET.get('anestesista')
    codigo_relatorio =  request.GET.get('codigo_relatorio')
    

    # Buscar convênios disponíveis para a empresa (para compatibilidade)
    convenios_disponiveis = []
    if empresa_id:
        from servicos_medicos.models import Convenio
        convenios_disponiveis = Convenio.objects.filter(empresa_id=empresa_id).order_by('nome')

    if nome:
        faturamentos = faturamentos.filter(Q(nome__icontains=nome))
    if guia:
        faturamentos = faturamentos.filter(guia__icontains=guia)
    if codigo_relatorio:
        faturamentos = faturamentos.filter(codigo_relatorio__icontains=codigo_relatorio)
    if anestesista:
        faturamentos = faturamentos.filter(anestesista__icontains=anestesista)
    if data_inicio:
        faturamentos = faturamentos.filter(data__gte=data_inicio)
    if data_fim:
        faturamentos = faturamentos.filter(data__lte=data_fim)
    if convenios:
        q_objects = Q()
        for conv in convenios:
            if conv:
                q_objects |= Q(convenio__icontains=conv)
        faturamentos = faturamentos.filter(q_objects)

    # Criar workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamentos Médicos"

    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Estilo para cabeçalho da empresa
    empresa_font = Font(bold=True, size=14)
    empresa_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

    # Informações básicas da empresa
    if empresa_id:
        try:
            empresa = Empresa.objects.get(id=empresa_id)
            empresa_info = f"{empresa.razao} - CNPJ: {empresa.cnpj}"
        except Empresa.DoesNotExist:
            empresa_info = f"Empresa ID: {empresa_id} - Dados não encontrados"
    else:
        empresa_info = "Empresa não identificada"

    # Adicionar cabeçalho da empresa
    ws.cell(row=1, column=1).value = empresa_info
    ws.cell(row=1, column=1).font = empresa_font
    ws.cell(row=1, column=1).fill = empresa_fill

    # Título do relatório
    ws.cell(row=3, column=1).value = "RELATÓRIO DE FATURAMENTOS MÉDICOS"
    ws.cell(row=3, column=1).font = Font(bold=True, size=16)

    # Data de geração
    from datetime import datetime
    ws.cell(row=4, column=1).value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws.cell(row=4, column=1).font = Font(italic=True)

    # Filtros aplicados
    filtros_texto = "Filtros aplicados:"
    if nome:
        filtros_texto += f" Nome: {nome}"
    if guia:
        filtros_texto += f" Guia: {guia}"
    if codigo_relatorio:
        filtros_texto += f" Guia: {codigo_relatorio}"
        
    if anestesista:
        filtros_texto += f" Anestesista: {anestesista}"
    if data_inicio:
        filtros_texto += f" Data início: {data_inicio}"
    if data_fim:
        filtros_texto += f" Data fim: {data_fim}"
    if convenios:
        filtros_texto += f" Convênios: {', '.join(convenios)}"

    ws.cell(row=5, column=1).value = filtros_texto
    ws.cell(row=5, column=1).font = Font(italic=True)

    # Cabeçalhos dos dados (linha 7)
    headers = [
        'Guia', 'Carteirinha', 'Nome', 'Anestesista', 'Código', 'Serviço', 'Data', 'Porte', 'QT', 'Valor', 'Valor Total', 'Apartamento ou Enfermaria', 'Urgência', 'Guia Lançada','codigo_relatorio', 'Lote'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    # Dados (a partir da linha 8)
    for row_num, faturamento in enumerate(faturamentos, 8):
        ws.cell(row=row_num, column=1).value = faturamento.guia or ''
        ws.cell(row=row_num, column=2).value = faturamento.carteirinha or ''
        ws.cell(row=row_num, column=3).value = faturamento.nome or ''
        ws.cell(row=row_num, column=4).value = faturamento.anestesista or ''
        ws.cell(row=row_num, column=5).value = faturamento.codigo_servico or ''
        ws.cell(row=row_num, column=6).value = faturamento.servico or ''
        ws.cell(row=row_num, column=7).value = faturamento.data.strftime('%d/%m/%Y') if faturamento.data else ''
        ws.cell(row=row_num, column=8).value = faturamento.porte or ''
        ws.cell(row=row_num, column=9).value = faturamento.qt or 0
        ws.cell(row=row_num, column=10).value = float(faturamento.valor) if faturamento.valor else 0
        ws.cell(row=row_num, column=11).value = float(faturamento.total) if faturamento.total else 0
        ws.cell(row=row_num, column=12).value = faturamento.apartamento_enfermaria or ''
        ws.cell(row=row_num, column=13).value = faturamento.urgencia or ''
        ws.cell(row=row_num, column=14).value = faturamento.guia_lancada or 0
        ws.cell(row=row_num, column=15).value = faturamento.codigo_relatorio or ''
        ws.cell(row=row_num, column=16).value = faturamento.lote or ''

    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=faturamentos_medicos.xlsx'

    wb.save(response)
    return response


def anexar_documento(request, pk):
    """View para anexar documentos a um faturamento"""
    faturamento = get_object_or_404(FaturamentoMedico, pk=pk)

    if request.method == 'POST':
        form = DocumentoAnexadoForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.save(commit=False)
            documento.faturamento = faturamento
            documento.save()
            messages.success(request, 'Documento anexado com sucesso!')
            return redirect('faturamento_medico:detalhes', pk=faturamento.pk)
    else:
        form = DocumentoAnexadoForm()

    context = {
        'form': form,
        'faturamento': faturamento,
        'titulo': f'Anexar Documento - {faturamento}'
    }

    return render(request, 'faturamento_medico/anexar_documento.html', context)


def download_documento(request, pk):
    """View para fazer download/visualizar de um documento anexado"""
    documento = get_object_or_404(DocumentoAnexado, pk=pk)

    try:
        with open(documento.arquivo.path, 'rb') as f:
            # Determinar o content_type baseado na extensão
            extensao = documento.arquivo.name.split('.')[-1].lower()
            if extensao == 'pdf':
                content_type = 'application/pdf'
            elif extensao in ['jpg', 'jpeg']:
                content_type = 'image/jpeg'
            elif extensao == 'png':
                content_type = 'image/png'
            elif extensao == 'gif':
                content_type = 'image/gif'
            else:
                content_type = 'application/octet-stream'

            response = HttpResponse(f.read(), content_type=content_type)
            # Se for para visualização inline (ex: PDFs no modal), usar inline
            if request.GET.get('inline') == 'true':
                response['Content-Disposition'] = f'inline; filename="{documento.arquivo.name.split("/")[-1]}"'
            else:
                response['Content-Disposition'] = f'attachment; filename="{documento.arquivo.name.split("/")[-1]}"'
            return response
    except FileNotFoundError:
        raise Http404("Arquivo não encontrado")


def excluir_documento(request, pk):
    """View para excluir um documento anexado"""
    documento = get_object_or_404(DocumentoAnexado, pk=pk)
    faturamento_pk = documento.faturamento.pk

    if request.method == 'POST':
        # Remove o arquivo do sistema de arquivos
        if documento.arquivo:
            documento.arquivo.delete(save=False)
        # Remove o registro do banco
        documento.delete()
        messages.success(request, 'Documento excluído com sucesso!')
        return redirect('faturamento_medico:detalhes', pk=faturamento_pk)

    context = {
        'documento': documento,
    }

    return render(request, 'faturamento_medico/confirmar_exclusao_documento.html', context)


def adicionar_item_servico(request, pk):
    """View para adicionar item de serviço a um faturamento"""
    faturamento = get_object_or_404(FaturamentoMedico, pk=pk)

    if request.method == 'POST':
        form = ItemServicoForm(request.POST, faturamento=faturamento)
        if form.is_valid():
            # Pegar os dados do POST
            cabecalho_id = request.POST.get('cabecalho')
            codigo_servico = request.POST.get('codigo_servico')
            qt = form.cleaned_data.get('qt', 1)
            valor = form.cleaned_data.get('valor', 0)

            if cabecalho_id and codigo_servico:
                from servicos_medicos.models import Cabecalho, ServicosMedicos
                try:
                    cabecalho = Cabecalho.objects.get(id=cabecalho_id)
                    servico = ServicosMedicos.objects.get(codigo=codigo_servico)

                    ItemServico.objects.create(
                        faturamento=faturamento,
                        codigo_servico=servico.codigo,
                        servico=servico.servicos,
                        porte=servico.porte_anestesico,
                        valor=valor,
                        qt=qt
                    )
                    # Atualiza o total do faturamento
                    faturamento.atualizar_total()
                    messages.success(request, 'Item de serviço adicionado com sucesso!')
                    return redirect('faturamento_medico:detalhes', pk=faturamento.pk)
                except (Cabecalho.DoesNotExist, ServicosMedicos.DoesNotExist):
                    messages.error(request, 'Cabeçalho ou serviço não encontrado.')
            else:
                messages.error(request, 'Selecione um cabeçalho e digite um código de serviço.')
    else:
        form = ItemServicoForm(faturamento=faturamento)

    context = {
        'form': form,
        'faturamento': faturamento,
        'titulo': f'Adicionar Item de Serviço - {faturamento}'
    }

    return render(request, 'faturamento_medico/adicionar_item_servico.html', context)


def editar_item_servico(request, pk):
    """View para editar item de serviço"""
    item = get_object_or_404(ItemServico, pk=pk)

    if request.method == 'POST':
        logger.info(f"Editando item {pk}, POST data: {request.POST}")
        # Para edição, cabecalho não é necessário
        post_data = request.POST.copy()
        form = ItemServicoForm(post_data, instance=item, faturamento=item.faturamento)
        logger.info(f"Form is_valid: {form.is_valid()}")
        if form.is_valid():
            logger.info("Salvando form")
            saved_item = form.save()
            logger.info(f"Item salvo: {saved_item.id}, valor: {saved_item.valor}, qt: {saved_item.qt}, total: {saved_item.total}")
            # Atualiza o total do faturamento
            item.faturamento.atualizar_total()
            logger.info(f"Total do faturamento atualizado: {item.faturamento.total}")
            messages.success(request, 'Item de serviço atualizado com sucesso!')
            return redirect('faturamento_medico:detalhes', pk=item.faturamento.pk)
        else:
            logger.error(f"Form errors: {form.errors}")
    else:
        form = ItemServicoForm(instance=item, faturamento=item.faturamento)

    context = {
        'form': form,
        'item': item,
        'faturamento': item.faturamento,
        'titulo': f'Editar Item de Serviço - {item.faturamento}'
    }

    return render(request, 'faturamento_medico/editar_item_servico.html', context)


def excluir_item_servico(request, pk):
    """View para excluir item de serviço"""
    item = get_object_or_404(ItemServico, pk=pk)
    faturamento_pk = item.faturamento.pk

    if request.method == 'POST':
        item.delete()
        # Atualiza o total do faturamento
        item.faturamento.atualizar_total()
        messages.success(request, 'Item de serviço excluído com sucesso!')
        return redirect('faturamento_medico:detalhes', pk=faturamento_pk)

    context = {
        'item': item,
    }

    return render(request, 'faturamento_medico/confirmar_exclusao_item.html', context)


def fechamento_repasse(request):
    """View para fechamento de repasse para anestesista"""
    logger.info(f"Iniciando fechamento_repasse. Método: {request.method}")
    logger.info(f"Session keys: {list(request.session.keys())}")
    
    # Verificar e potencialmente selecionar empresa automaticamente
    empresa_id, empresa_nome, error_msg = verificar_empresa_selecionada(request)
    
    if error_msg:
        logger.warning(f"Erro relacionado à empresa: {error_msg}")
        messages.error(request, error_msg)
        # Usar safe para permitir HTML na mensagem
        from django.utils.safestring import mark_safe
        messages.error(request, mark_safe(error_msg))
        return redirect('faturamento_medico:ftlistar')
    
    logger.info(f"Empresa selecionada: {empresa_id} - {empresa_nome}")

    # Filtros para seleção
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    convenios = request.GET.getlist('convenio')
    data_fechamento = request.GET.get('data_fechamento')
    anestesista = request.GET.get('anestesista')
    mostrar_fechados = request.GET.get('mostrar_fechados', 'false').lower() == 'true'

    # Buscar convênios disponíveis para a empresa
    convenios_disponiveis = []
    if empresa_id:
        from servicos_medicos.models import Convenio
        convenios_disponiveis = list(Convenio.objects.filter(empresa_id=empresa_id).order_by('nome'))
        if not convenios_disponiveis:
            # Convênios padrão se nenhum for encontrado para a empresa
            convenios_padrao = [
                {'nome': 'CBSAUDE'},
                {'nome': 'PM'},
                {'nome': 'UNIMED'},
                {'nome': 'BRADESCO'},
                {'nome': 'GEAP'},
                {'nome': 'SAUDE CAIXA'},
                {'nome': 'POSTAL SAUDE'},
                {'nome': 'FUSEX'},
                {'nome': 'LIFE EMPRESARIAL'},
                {'nome': 'CASSI'},
                {'nome': 'GCARD'},
                {'nome': 'PERSONAL NET'},
            ]
            convenios_disponiveis = convenios_padrao

    # Query base
    faturamentos = FaturamentoMedico.objects.filter(empresa_id=empresa_id)















    # Aplicar filtros
    if data_inicio:
        faturamentos = faturamentos.filter(data_fechamento__gte=data_inicio)
    if data_fim:
        faturamentos = faturamentos.filter(data_fechamento__lte=data_fim)
    if data_fechamento_inicio:
        faturamentos = faturamentos.filter(data_fechamento__gte=data_fechamento_inicio)
    if data_fechamento_fim:
        faturamentos = faturamentos.filter(data_fechamento__lte=data_fechamento_fim)
    if convenios:
        q_objects = Q()
        for conv in convenios:
            if conv:
                q_objects |= Q(convenio__icontains=conv)
        faturamentos = faturamentos.filter(q_objects)
    if anestesista:
        faturamentos = faturamentos.filter(anestesista__icontains=anestesista)

    # Filtrar apenas faturamentos com anestesista
    faturamentos = faturamentos.exclude(anestesista__isnull=True).exclude(anestesista='')

    # Processar fechamento se for POST
    if request.method == 'POST':
        faturamentos_ids = request.POST.getlist('faturamentos_selecionados')

        if request.POST.get('aplicar_comissao') and faturamentos_ids:
            # Aplicar comissão aos faturamentos selecionados
            percentual_imposto = float(request.POST.get('percentual_imposto', 0))
            percentual_comissao = float(request.POST.get('percentual_comissao', 0))

            # Usar getlist para obter todos os valores do campo
            faturamentos_ids = request.POST.getlist('faturamentos_selecionados')

            # Debug: verificar o que está sendo recebido
            logger.info(f"faturamentos_ids após getlist: {faturamentos_ids} (tipo: {type(faturamentos_ids)})")

            # Converter IDs para inteiros para evitar problemas de tipo
            try:
                faturamentos_ids = [int(id.strip()) for id in faturamentos_ids if id.strip()]
            except (ValueError, TypeError) as e:
                logger.error(f"Erro ao converter IDs: {e}. faturamentos_ids: {faturamentos_ids}")
                messages.error(request, 'IDs de faturamentos inválidos.')
                return redirect('faturamento_medico:fechamento_repasse')

            faturamentos = FaturamentoMedico.objects.filter(
                id__in=faturamentos_ids,
                empresa_id=empresa_id
            )

            for faturamento in faturamentos:
                # Calcular valores com precisão decimal (manter como Decimal)
                from decimal import Decimal
                total_decimal = Decimal(str(faturamento.total))
                percentual_imposto_decimal = Decimal(str(percentual_imposto))
                percentual_comissao_decimal = Decimal(str(percentual_comissao))

                valor_imposto = (total_decimal * percentual_imposto_decimal / 100).quantize(Decimal('0.01'))
                base_comissao = total_decimal - valor_imposto
                valor_comissao = (base_comissao * percentual_comissao_decimal / 100).quantize(Decimal('0.01'))

                # Atualizar campos
                faturamento.percentual_imposto = percentual_imposto
                faturamento.percentual_comissao = percentual_comissao
                faturamento.valor_imposto = valor_imposto
                faturamento.valor_comissao = valor_comissao
                faturamento.save()

            # Verificar se é uma requisição AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({
                    'success': True,
                    'message': f'Comissão aplicada com sucesso para {len(faturamentos_ids)} faturamento(s)!'
                })
            else:
                messages.success(request, f'Comissão aplicada com sucesso para {len(faturamentos_ids)} faturamento(s)!')
                return redirect('faturamento_medico:fechamento_repasse')

        elif faturamentos_ids:
            # Usar data atual se não foi fornecida
            if not data_fechamento:
                data_fechamento = timezone.now().date()

            # Gerar código único para este fechamento com bloqueio de transação
            from django.db import transaction
            max_attempts = 10
            attempt = 0
            codigo_fechamento = None

            while attempt < max_attempts:
                codigo_fechamento = str(uuid.uuid4())[:8].upper()
                logger.info(f"Código de fechamento gerado (tentativa {attempt + 1}): {codigo_fechamento}")

                # Verificar se o código já existe
                with transaction.atomic():
                    existing_with_code = FaturamentoMedico.objects.select_for_update().filter(codigo_fechamento=codigo_fechamento)
                    if not existing_with_code.exists():
                        logger.info(f"Código de fechamento único encontrado: {codigo_fechamento}")
                        break
                    else:
                        logger.warning(f"Código de fechamento {codigo_fechamento} já existe em {existing_with_code.count()} registros")
                        attempt += 1

            if attempt >= max_attempts:
                logger.error("Não foi possível gerar um código único após várias tentativas")
                messages.error(request, 'Erro interno: não foi possível gerar código único de fechamento.')
                return redirect('faturamento_medico:fechamento_repasse')

            # Buscar faturamentos selecionados para verificar status
            faturamentos_selecionados = FaturamentoMedico.objects.filter(
                id__in=faturamentos_ids,
                empresa_id=empresa_id
            )
            logger.info(f"Faturamentos selecionados: {len(faturamentos_selecionados)}")

            # Verificar se algum já está fechado
            faturamentos_ja_fechados = faturamentos_selecionados.filter(data_fechamento__isnull=False)
            if faturamentos_ja_fechados.exists():
                logger.warning(f"Encontrados {faturamentos_ja_fechados.count()} faturamentos já fechados:")
                for fat in faturamentos_ja_fechados:
                    logger.warning(f"  ID {fat.id}: data_fechamento={fat.data_fechamento}, codigo_fechamento={fat.codigo_fechamento}")

            # Verificar códigos de fechamento existentes nos faturamentos selecionados
            faturamentos_com_codigo = faturamentos_selecionados.filter(codigo_fechamento__isnull=False)
            if faturamentos_com_codigo.exists():
                logger.info(f"Faturamentos selecionados que já têm código de fechamento:")
                for fat in faturamentos_com_codigo:
                    logger.info(f"  ID {fat.id}: codigo_fechamento={fat.codigo_fechamento}")

            # Atualizar data de fechamento, status e código para os faturamentos selecionados
            # Primeiro, verificar se algum faturamento já tem um código de fechamento
            faturamentos_com_codigo_existente = FaturamentoMedico.objects.filter(
                id__in=faturamentos_ids,
                empresa_id=empresa_id,
                codigo_fechamento__isnull=False
            )

            if faturamentos_com_codigo_existente.exists():
                logger.warning(f"Encontrados {faturamentos_com_codigo_existente.count()} faturamentos que já têm código de fechamento:")
                for fat in faturamentos_com_codigo_existente:
                    logger.warning(f"  ID {fat.id}: codigo_fechamento={fat.codigo_fechamento}")
                # Para estes, não sobrescrever o código existente
                faturamentos_ids_para_atualizar = [id for id in faturamentos_ids if id not in [fat.id for fat in faturamentos_com_codigo_existente]]
                logger.info(f"Atualizando apenas {len(faturamentos_ids_para_atualizar)} faturamentos sem código existente")
            else:
                faturamentos_ids_para_atualizar = faturamentos_ids

            if faturamentos_ids_para_atualizar:
                # SOLUÇÃO DEFINITIVA: Usar códigos únicos por faturamento
                updated_count = 0
                for faturamento_id in faturamentos_ids_para_atualizar:
                    # Gerar código único para cada faturamento
                    faturamento_codigo = str(uuid.uuid4())[:8].upper()

                    try:
                        # Tentar atualizar este faturamento específico com código único
                        count = FaturamentoMedico.objects.filter(
                            id=faturamento_id,
                            empresa_id=empresa_id,
                            codigo_fechamento__isnull=True,  # Só atualizar se não tiver código
                            status__in=['pendente', 'enviado']  # Só atualizar se não estiver finalizado
                        ).update(
                            data_fechamento=data_fechamento,
                            status='finalizado',
                            codigo_fechamento=faturamento_codigo
                        )
                        if count > 0:
                            updated_count += count
                            logger.info(f"Faturamento {faturamento_id} atualizado com código único {faturamento_codigo}")
                        else:
                            logger.warning(f"Faturamento {faturamento_id} não foi atualizado (já processado ou não encontrado)")
                    except Exception as e:
                        logger.error(f"Erro ao atualizar faturamento {faturamento_id}: {e}")
                        # Mesmo com códigos únicos, pode haver race condition, mas é muito improvável
                        raise e

                logger.info(f"Total de faturamentos atualizados com códigos únicos: {updated_count}")
                if updated_count > 0:
                    messages.success(request, f'Fechamento realizado com sucesso! {updated_count} faturamento(s) finalizado(s) com códigos únicos.')
                else:
                    messages.warning(request, 'Nenhum faturamento foi atualizado. Todos podem já ter sido processados.')
            else:
                logger.info("Nenhum faturamento para atualizar (todos já têm código de fechamento)")
                messages.info(request, 'Todos os faturamentos selecionados já possuem código de fechamento.')

            messages.success(request, f'Fechamento realizado com sucesso para {len(faturamentos_ids)} faturamento(s)! Código: {codigo_fechamento}')
            return redirect('faturamento_medico:fechamento_repasse')

    # Filtrar baseado na opção selecionada
    if mostrar_fechados:
        # Mostrar apenas faturamentos já fechados
        faturamentos_filtrados = faturamentos.filter(data_fechamento__isnull=False)
    else:
        # Mostrar apenas faturamentos não fechados (padrão)
        faturamentos_filtrados = faturamentos.filter(data_fechamento__isnull=True)

    # Estatísticas
    total_faturamentos = faturamentos_filtrados.count()
    valor_total = sum(f.total for f in faturamentos_filtrados if f.total)

    context = {
        'faturamentos': faturamentos_filtrados,
        'total_faturamentos': total_faturamentos,
        'valor_total': valor_total,
        'convenios_disponiveis': convenios_disponiveis,
        'filtros': {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'convenio': convenios,
            'data_fechamento': data_fechamento,
            'anestesista': anestesista,
        },
        'mostrar_fechados': mostrar_fechados,
    }

    return render(request, 'faturamento_medico/fechamento_repasse.html', context)


def reabrir_fechamento(request, pk):
    """View para reabrir um fechamento de repasse"""
    faturamento = get_object_or_404(FaturamentoMedico, pk=pk)

    if request.method == 'POST':
        # Limpar campos de fechamento
        faturamento.data_fechamento = None
        faturamento.status = 'pendente'
        faturamento.codigo_fechamento = None
        # Manter os valores de comissão e imposto calculados
        faturamento.save()

        messages.success(request, f'Fechamento reaberto com sucesso para {faturamento.nome}!')
        return redirect('faturamento_medico:fechamento_repasse')

    context = {
        'faturamento': faturamento,
    }

    return render(request, 'faturamento_medico/confirmar_reabertura.html', context)


def exportar_excel_fechados(request):
    """Exporta os repasses fechados para Excel com cabeçalho da empresa e ordenação por convênio"""
    empresa_id = request.session.get('empresa_id')
    if empresa_id:
        faturamentos = FaturamentoMedico.objects.filter(empresa_id=empresa_id)
    else:
        faturamentos = FaturamentoMedico.objects.none()

    # Aplicar os mesmos filtros da view de fechamento_repasse
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    convenios = request.GET.getlist('convenio')
    data_fechamento = request.GET.get('data_fechamento')
    data_fechamento_inicio = request.GET.get('data_fechamento_inicio')
    data_fechamento_fim = request.GET.get('data_fechamento_fim')
    anestesista = request.GET.get('anestesista')

    # Buscar convênios disponíveis para a empresa (para compatibilidade)
    convenios_disponiveis = []
    if empresa_id:
        from servicos_medicos.models import Convenio
        convenios_disponiveis = Convenio.objects.filter(empresa_id=empresa_id).order_by('nome')

    if data_inicio:
        faturamentos = faturamentos.filter(data__gte=data_inicio)
    if data_fim:
        faturamentos = faturamentos.filter(data__lte=data_fim)
    if convenios:
        q_objects = Q()
        for conv in convenios:
            if conv:
                q_objects |= Q(convenio__icontains=conv)
        faturamentos = faturamentos.filter(q_objects)
    if anestesista:
        faturamentos = faturamentos.filter(anestesista__icontains=anestesista)

    # Filtrar apenas faturamentos com anestesista e fechados
    faturamentos = faturamentos.exclude(anestesista__isnull=True).exclude(anestesista='')
    faturamentos = faturamentos.filter(data_fechamento__isnull=False)

    # Ordenar por convênio
    faturamentos = faturamentos.order_by('convenio', 'data')

    # Criar workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Repasses Fechados"

    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Estilo para cabeçalho da empresa
    empresa_font = Font(bold=True, size=14)
    empresa_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

    # Informações básicas da empresa
    if empresa_id:
        try:
            empresa = Empresa.objects.get(id=empresa_id)
            empresa_info = f"{empresa.razao} - CNPJ: {empresa.cnpj}"
        except Empresa.DoesNotExist:
            empresa_info = f"Empresa ID: {empresa_id} - Dados não encontrados"
    else:
        empresa_info = "Empresa não identificada"

    # Adicionar cabeçalho da empresa
    ws.cell(row=1, column=1).value = empresa_info
    ws.cell(row=1, column=1).font = empresa_font
    ws.cell(row=1, column=1).fill = empresa_fill

    # Título do relatório
    ws.cell(row=3, column=1).value = "RELATÓRIO DE REPASSES FECHADOS"
    ws.cell(row=3, column=1).font = Font(bold=True, size=16)

    # Data de geração
    from datetime import datetime
    ws.cell(row=4, column=1).value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws.cell(row=4, column=1).font = Font(italic=True)

    # Filtros aplicados
    filtros_texto = "Filtros aplicados:"
    if data_inicio:
        filtros_texto += f" Data início: {data_inicio}"
    if data_fim:
        filtros_texto += f" Data fim: {data_fim}"
    if data_fechamento:
        filtros_texto += f" Data fechamento: {data_fechamento}"
    if convenios:
        filtros_texto += f" Convênios: {', '.join(convenios)}"
    if anestesista:
        filtros_texto += f" Anestesista: {anestesista}"

    ws.cell(row=5, column=1).value = filtros_texto
    ws.cell(row=5, column=1).font = Font(italic=True)

    # Cabeçalhos dos dados (linha 7)
    headers = [
        'Data', 'Nome', 'Guia', 'Anestesista', 'Convênio','Codigo Relatorio' ,'Código Serviço', 'Serviço', 'QT', 'Valor Unitário',
        'Valor Total Item','Valor Total', 'Valor do Imposto', 'Valor da Comissão', 'Valor Líquido',
        'Data de Fechamento', 'Status'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    # Dados (a partir da linha 8)
    current_row = 8
    for faturamento in faturamentos:
        # Buscar itens de serviço para este faturamento
        valor_liquido = float(faturamento.total or 0) - float(faturamento.valor_imposto or 0) - float(faturamento.valor_comissao or 0)
        ws.cell(row=current_row, column=1).value = faturamento.data.strftime('%d/%m/%Y') if faturamento.data else ''
        ws.cell(row=current_row, column=2).value = faturamento.nome or ''
        ws.cell(row=current_row, column=3).value = faturamento.guia or ''
        ws.cell(row=current_row, column=4).value = faturamento.anestesista or ''
        ws.cell(row=current_row, column=5).value = faturamento.convenio or ''
        ws.cell(row=current_row, column=6).value = faturamento.codigo_relatorio or ''
        itens_servico = faturamento.itens_servico.filter(faturamento_id = faturamento.id)

        if itens_servico.exists():
            # Para cada item de serviço, criar uma linha
            cont = 0;
            for item in itens_servico:
                # Calcular valor líquido
                

                
                ws.cell(row=current_row, column=7).value = item.codigo_servico or ''
                ws.cell(row=current_row, column=8).value = item.servico or ''
                ws.cell(row=current_row, column=9).value = item.qt or 0
                ws.cell(row=current_row, column=10).value = float(item.valor) if item.valor else 0
                ws.cell(row=current_row, column=11).value = float(item.total) if item.total else 0
                
                if cont == 0:
                   ws.cell(row=current_row, column=12).value = float(faturamento.total) if faturamento.total else 0   
                   ws.cell(row=current_row, column=13).value = float(faturamento.valor_imposto) if faturamento.valor_imposto else 0
                   ws.cell(row=current_row, column=14).value = float(faturamento.valor_comissao) if faturamento.valor_comissao else 0
                   ws.cell(row=current_row, column=15).value = valor_liquido
                   ws.cell(row=current_row, column=16).value = faturamento.data_fechamento.strftime('%d/%m/%Y') if faturamento.data_fechamento else ''
                   
                   ws.cell(row=current_row, column=17).value = faturamento.status or ''    
                cont += 1
                current_row += 1
        
        cont = 0 
        
        current_row += 1

    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=repasses_fechados.xlsx'

    wb.save(response)
    return response


# Views para Serviços Disponíveis
def listar_servicos(request):
    """Lista todos os serviços disponíveis"""
    servicos = ServicoDisponivel.objects.all().order_by('codigo')

    # Filtros
    categoria = request.GET.get('categoria')
    ativo = request.GET.get('ativo')

    if categoria:
        servicos = servicos.filter(categoria__icontains=categoria)
    if ativo:
        if ativo == '1':
            servicos = servicos.filter(ativo=True)
        elif ativo == '0':
            servicos = servicos.filter(ativo=False)

    context = {
        'servicos': servicos,
        'filtros': {
            'categoria': categoria,
            'ativo': ativo,
        }
    }

    return render(request, 'faturamento_medico/listar_servicos.html', context)


def criar_servico(request):
    """Cria um novo serviço disponível"""
    if request.method == 'POST':
        form = ServicoDisponivelForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Serviço criado com sucesso!')
            return redirect('faturamento_medico:listar_servicos')
    else:
        form = ServicoDisponivelForm()

    context = {
        'form': form,
        'titulo': 'Criar Serviço Disponível'
    }

    return render(request, 'faturamento_medico/form_servico.html', context)


def editar_servico(request, pk):
    """Edita um serviço disponível existente"""
    servico = get_object_or_404(ServicoDisponivel, pk=pk)

    if request.method == 'POST':
        form = ServicoDisponivelForm(request.POST, instance=servico)
        if form.is_valid():
            form.save()
            messages.success(request, 'Serviço atualizado com sucesso!')
            return redirect('faturamento_medico:listar_servicos')
    else:
        form = ServicoDisponivelForm(instance=servico)

    context = {
        'form': form,
        'servico': servico,
        'titulo': 'Editar Serviço Disponível'
    }

    return render(request, 'faturamento_medico/form_servico.html', context)
def extrair_dados_documento(request):
    """View para extrair dados de documento usando Gemini via AJAX"""
    if request.method == 'POST':
        documento = request.FILES.get('documento')
        if documento:
            dados = processar_arquivos_com_gemini([documento])
            logger.info(f"Dados extraídos para debug: {dados}")
            return JsonResponse(dados)
    return JsonResponse({'error': 'Invalid request'}, status=400)


def extrair_dados_documento_ocr(request):
    """View para extrair dados de documento usando OCR via AJAX"""
    if request.method == 'POST':
        documento = request.FILES.get('documento')
        if documento:
            dados = processar_arquivos_com_ocr([documento])
            logger.info(f"Dados extraídos via OCR para debug: {dados}")
            return JsonResponse(dados)
    return JsonResponse({'error': 'Invalid request'}, status=400)


def excluir_servico(request, pk):
    """Exclui um serviço disponível"""
    servico = get_object_or_404(ServicoDisponivel, pk=pk)

    if request.method == 'POST':
        servico.delete()
        messages.success(request, 'Serviço excluído com sucesso!')
        return redirect('faturamento_medico:listar_servicos')

    context = {
        'servico': servico,
    }

    return render(request, 'faturamento_medico/confirmar_exclusao_servico.html', context)


def carregar_tabelas_por_cabecalho(request, cabecalho_id):
    """View AJAX para carregar tabelas por cabeçalho"""
    try:
        from servicos_medicos.models import Cabecalho, TabelaPreco
        cabecalho = Cabecalho.objects.get(id=cabecalho_id)
        tabelas = TabelaPreco.objects.filter(cabecalho=cabecalho).select_related('codigo_servico')
        data = []
        for tabela in tabelas:
            data.append({
                'id': tabela.id,
                'codigo': tabela.codigo_servico.codigo,
                'servico': tabela.codigo_servico.servicos,
                'porte': tabela.codigo_servico.porte_anestesico,
                'preco_apartamento': str(tabela.preco_apartamento),
                'preco_enfermaria': str(tabela.preco_enfermaria),
                'display': f"{tabela.codigo_servico} - {tabela.preco_apartamento}/{tabela.preco_enfermaria}"
            })
        return JsonResponse({'success': True, 'tabelas': data})
    except Cabecalho.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cabeçalho não encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def buscar_servicos(request):
    """View AJAX para buscar serviços por código"""
    query = request.GET.get('q', '')
    if len(query) >= 3:  # Buscar a partir de 3 caracteres
        from servicos_medicos.models import ServicosMedicos
        servicos = ServicosMedicos.objects.filter(codigo__icontains=query)[:10]  # Limitar a 10 resultados
        data = []
        for servico in servicos:
            data.append({
                'codigo': servico.codigo,
                'servico': servico.servicos,
                'porte': servico.porte_anestesico
            })
        return JsonResponse({'success': True, 'servicos': data})
    return JsonResponse({'success': True, 'servicos': []})


def buscar_servicos_por_descricao(request):
    """View AJAX para buscar serviços por descrição"""
    query = request.GET.get('q', '')
    if len(query) >= 3:  # Buscar a partir de 3 caracteres
        from servicos_medicos.models import ServicosMedicos
        servicos = ServicosMedicos.objects.filter(servicos__icontains=query)[:10]  # Limitar a 10 resultados
        data = []
        for servico in servicos:
            data.append({
                'codigo': servico.codigo,
                'servico': servico.servicos,
                'porte': servico.porte_anestesico
            })
        return JsonResponse({'success': True, 'servicos': data})
    return JsonResponse({'success': True, 'servicos': []})


def buscar_precos_servico(request, cabecalho_id, codigo_servico):
    """View AJAX para buscar preços de um serviço em um cabeçalho"""
    try:
        from servicos_medicos.models import Cabecalho, TabelaPreco, ServicosMedicos
        cabecalho = Cabecalho.objects.get(id=cabecalho_id)
        servico = ServicosMedicos.objects.get(codigo=codigo_servico)
        tabela = TabelaPreco.objects.filter(
            cabecalho=cabecalho,
            codigo_servico=servico
        ).first()
        if tabela:
            return JsonResponse({
                'success': True,
                'preco_apartamento': str(tabela.preco_apartamento),
                'preco_enfermaria': str(tabela.preco_enfermaria)
            })
        else:
            return JsonResponse({
                'success': True,
                'preco_apartamento': '0.00',
                'preco_enfermaria': '0.00'
            })
    except (Cabecalho.DoesNotExist, ServicosMedicos.DoesNotExist):
        return JsonResponse({'success': False, 'error': 'Cabeçalho ou serviço não encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def selecionar_lote_imprimir(request):
    """View para selecionar lote para imprimir relatório"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    lotes = Lote.objects.filter(empresa_id=empresa_id).order_by('-id')
    context = {'lotes': lotes}
    return render(request, 'faturamento_medico/selecionar_lote_imprimir.html', context)


def imprimir_lote(request, lote_id):
    """View para imprimir relatório de lote em HTML"""
    if lote_id == 0:
        lote_id = request.GET.get('lote_id')
        if not lote_id:
            return HttpResponse('Lote não selecionado')

    lote = get_object_or_404(Lote, id=lote_id)
    empresa_id = request.GET.get('empresa_id') or request.session.get('empresa_id')
    if not empresa_id:
        return HttpResponse('Sessão expirada. Faça login novamente.')
    # Verificar se o lote pertence à empresa (já filtrado na seleção)
    if lote.empresa_id != int(empresa_id):
        return HttpResponse('Acesso negado')

    faturamentos = FaturamentoMedico.objects.filter(lote=str(lote.id)).order_by('data')
    items = ItemServico.objects.filter(faturamento__in=faturamentos).select_related('faturamento').order_by('faturamento__nome', 'faturamento__data', 'faturamento__guia')

    # Agrupar itens por beneficiário
    grouped_items = {}
    total_geral = 0
    for item in items:
        beneficiario = item.faturamento.nome or 'Sem Nome'
        if beneficiario not in grouped_items:
            grouped_items[beneficiario] = []
        grouped_items[beneficiario].append(item)
        total_geral += item.total or 0

    empresa = Empresa.objects.get(id=empresa_id)

    from django.db.models import Min, Max
    periodo_inicio = faturamentos.aggregate(min_data=Min('data'))['min_data']
    periodo_fim = faturamentos.aggregate(max_data=Max('data'))['max_data']

    context = {
        'lote': lote,
        'empresa': empresa,
        'periodo_inicio': periodo_inicio,
        'periodo_fim': periodo_fim,
        'grouped_items': grouped_items,
        'total_geral': total_geral,
    }
    return render(request, 'faturamento_medico/imprimir_lote.html', context)


def imprimir_repasses_fechados(request):
    """View para imprimir relatório de repasses fechados em HTML"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return HttpResponse('Sessão expirada. Faça login novamente.')

    # Aplicar os mesmos filtros da view de fechamento_repasse
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    convenios = request.GET.getlist('convenio')
    data_fechamento = request.GET.get('data_fechamento')
    anestesista = request.GET.get('anestesista')

    # Query base - apenas faturamentos fechados com anestesista
    faturamentos = FaturamentoMedico.objects.filter(
        empresa_id=empresa_id,
        data_fechamento__isnull=False,
        anestesista__isnull=False
    ).exclude(anestesista='').order_by('anestesista', 'data_fechamento', 'nome')

    # Aplicar filtros
    if data_inicio:
        faturamentos = faturamentos.filter(data_fechamento__gte=data_inicio)
    if data_fim:
        faturamentos = faturamentos.filter(data_fechamento__lte=data_fim)
    if convenios:
        q_objects = Q()
        for conv in convenios:
            if conv:
                q_objects |= Q(convenio__icontains=conv)
        faturamentos = faturamentos.filter(q_objects)
    if anestesista:
        faturamentos = faturamentos.filter(anestesista__icontains=anestesista)

    # Agrupar por anestesista
    repasses_por_anestesista = {}
    total_geral = 0
    total_imposto_geral = 0
    total_comissao_geral = 0
    total_liquido_geral = 0

    for faturamento in faturamentos:
        anestesista_nome = faturamento.anestesista or 'Sem Anestesista'
        if anestesista_nome not in repasses_por_anestesista:
            repasses_por_anestesista[anestesista_nome] = {
                'repasses': [],
                'total_valor_total': 0,
                'total_valor_imposto': 0,
                'total_valor_comissao': 0,
                'total_valor_liquido': 0,
            }

        # Calcular valores
        valor_total = float(faturamento.total or 0)
        valor_imposto = float(faturamento.valor_imposto or 0)
        valor_comissao = float(faturamento.valor_comissao or 0)
        valor_liquido = valor_total - valor_imposto - valor_comissao

        repasse_info = {
            'faturamento': faturamento,
            'valor_total': valor_total,
            'valor_imposto': valor_imposto,
            'valor_comissao': valor_comissao,
            'valor_liquido': valor_liquido,
        }

        repasses_por_anestesista[anestesista_nome]['repasses'].append(repasse_info)
        repasses_por_anestesista[anestesista_nome]['total_valor_total'] += valor_total
        repasses_por_anestesista[anestesista_nome]['total_valor_imposto'] += valor_imposto
        repasses_por_anestesista[anestesista_nome]['total_valor_comissao'] += valor_comissao
        repasses_por_anestesista[anestesista_nome]['total_valor_liquido'] += valor_liquido

        # Acumuladores gerais
        total_geral += valor_total
        total_imposto_geral += valor_imposto
        total_comissao_geral += valor_comissao
        total_liquido_geral += valor_liquido

    empresa = Empresa.objects.get(id=empresa_id)

    # Calcular período
    from django.db.models import Min, Max
    periodo_inicio = faturamentos.aggregate(min_data=Min('data_fechamento'))['min_data']
    periodo_fim = faturamentos.aggregate(max_data=Max('data_fechamento'))['max_data']

    context = {
        'empresa': empresa,
        'periodo_inicio': periodo_inicio,
        'periodo_fim': periodo_fim,
        'repasses_por_anestesista': repasses_por_anestesista,
        'total_geral': total_geral,
        'total_imposto_geral': total_imposto_geral,
        'total_comissao_geral': total_comissao_geral,
        'total_liquido_geral': total_liquido_geral,
        'filtros': {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'convenio': convenios,
            'data_fechamento': data_fechamento,
            'anestesista': anestesista,
        }
    }

    return render(request, 'faturamento_medico/imprimir_repasses_fechados.html', context)


def carregar_precos_por_cabecalho(request, cabecalho_id):
    """View AJAX para carregar preços por cabeçalho"""
    try:
        from servicos_medicos.models import Cabecalho, TabelaPreco
        cabecalho = Cabecalho.objects.get(id=cabecalho_id)
        # Pegar os preços do primeiro serviço ou calcular médias, mas como são por serviço, talvez mostrar uma mensagem
        tabelas = TabelaPreco.objects.filter(cabecalho=cabecalho)
        if tabelas.exists():
            # Como os preços variam por serviço, talvez mostrar uma mensagem ou os preços do primeiro
            preco_apartamento = tabelas.first().preco_apartamento
            preco_enfermaria = tabelas.first().preco_enfermaria
            return JsonResponse({
                'success': True,
                'preco_apartamento': str(preco_apartamento),
                'preco_enfermaria': str(preco_enfermaria)
            })
        else:
            return JsonResponse({'success': False, 'error': 'Nenhuma tabela encontrada'})
    except Cabecalho.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cabeçalho não encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def gerar_lote(request):
    """View para gerar lote a partir dos faturamentos selecionados"""
    logger.info("Iniciando gerar_lote")
    empresa_id = request.session.get('empresa_id')
    logger.info(f"Empresa ID da sessão: {empresa_id}")
    if not empresa_id:
        logger.warning("Empresa não encontrada na sessão")
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    if request.method == 'POST':
        logger.info("Método POST detectado")
        faturamento_ids = request.POST.getlist('faturamentos_selecionados')
        lote_existente_id = request.POST.get('lote_existente')
        logger.info(f"Faturamento IDs selecionados: {faturamento_ids}")
        logger.info(f"Lote existente: {lote_existente_id}")

        if not faturamento_ids:
            logger.warning("Nenhum faturamento selecionado")
            messages.error(request, 'Selecione pelo menos um faturamento para gerar o lote.')
            return redirect('faturamento_medico:ftlistar')

        # Buscar faturamentos selecionados
        faturamentos = FaturamentoMedico.objects.filter(
            id__in=faturamento_ids,
            empresa_id=empresa_id
        )
        logger.info(f"Faturamentos encontrados: {faturamentos.count()}")

        if not faturamentos.exists():
            logger.warning("Nenhum faturamento encontrado para os IDs")
            messages.error(request, 'Nenhum faturamento encontrado.')
            return redirect('faturamento_medico:ftlistar')

        if lote_existente_id:
            # Adicionar a lote existente
            try:
                lote_existente = Lote.objects.get(id=lote_existente_id, empresa_id=empresa_id)
                logger.info(f"Adicionando a lote existente: {lote_existente.id}")
            except Lote.DoesNotExist:
                logger.error(f"Lote existente não encontrado: {lote_existente_id}")
                messages.error(request, 'Lote selecionado não encontrado.')
                return redirect('faturamento_medico:ftlistar')

            # Verificar se os faturamentos têm o mesmo convênio do lote
            faturamentos_diferente_convenio = faturamentos.exclude(convenio=lote_existente.convenio)
            if faturamentos_diferente_convenio.exists():
                logger.warning(f"Faturamentos com convênio diferente: {[f.id for f in faturamentos_diferente_convenio]}")
                messages.error(request, 'Todos os faturamentos devem ter o mesmo convênio do lote selecionado.')
                return redirect('faturamento_medico:ftlistar')

            # Filtrar faturamentos sem lote ou com status != finalizado
            faturamentos_validos = faturamentos.filter(status__in=['pendente', 'enviado'])
            faturamentos_invalidos = faturamentos.exclude(status__in=['pendente', 'enviado'])

            if faturamentos_invalidos.exists():
                logger.warning(f"Faturamentos com status finalizado: {[f.id for f in faturamentos_invalidos]}")
                messages.warning(request, 'Faturamentos finalizados não podem ser adicionados a lotes.')

            if not faturamentos_validos.exists():
                logger.warning("Nenhum faturamento válido para adicionar")
                messages.error(request, 'Nenhum faturamento válido para adicionar ao lote.')
                return redirect('faturamento_medico:ftlistar')

            # Atualizar os faturamentos com o ID do lote e status para enviado
            fat_ids = [f.id for f in faturamentos_validos]
            try:
                updated = FaturamentoMedico.objects.filter(id__in=fat_ids).update(lote=str(lote_existente.id), status='enviado')
                logger.info(f"Faturamentos adicionados ao lote {lote_existente.id}: {updated}")
            except Exception as e:
                logger.error(f"Erro ao adicionar faturamentos ao lote {lote_existente.id}: {e}")
                messages.error(request, f'Erro ao adicionar faturamentos ao lote: {e}')
                return redirect('faturamento_medico:ftlistar')

            # Atualizar o total do lote
            try:
                lote_existente.atualizar_total()
                logger.info(f"Total do lote {lote_existente.id} atualizado: {lote_existente.total_lote}")
            except Exception as e:
                logger.error(f"Erro ao atualizar total do lote {lote_existente.id}: {e}")
                messages.error(request, f'Erro ao atualizar total do lote: {e}')
                return redirect('faturamento_medico:ftlistar')

            url = reverse('faturamento_medico:ftlistar')
            return HttpResponse(f'<script>alert("Faturamentos adicionados ao lote {lote_existente.id} com sucesso!"); window.location.href = "{url}";</script>')
        else:
            # Criar novo lote
            # Filtrar apenas faturamentos sem lote
            faturamentos_sem_lote = faturamentos.filter(lote__isnull=True) | faturamentos.filter(lote='')
            faturamentos_com_lote = faturamentos.exclude(lote__isnull=True).exclude(lote='')

            if faturamentos_com_lote.exists():
                logger.warning(f"Alguns faturamentos já têm lote: {[f.id for f in faturamentos_com_lote]}")
                messages.warning(request, f'Alguns faturamentos selecionados já estão incluídos em outro lote e foram ignorados.')

            if not faturamentos_sem_lote.exists():
                logger.warning("Nenhum faturamento sem lote encontrado")
                messages.error(request, 'Todos os faturamentos selecionados já estão incluídos em lotes.')
                return redirect('faturamento_medico:ftlistar')

            faturamentos = faturamentos_sem_lote
            logger.info(f"Faturamentos sem lote: {faturamentos.count()}")

            # Agrupar faturamentos por convênio
            faturamentos_por_convenio = {}
            for fat in faturamentos:
                convenio = fat.convenio or 'Sem Convênio'
                if convenio not in faturamentos_por_convenio:
                    faturamentos_por_convenio[convenio] = []
                faturamentos_por_convenio[convenio].append(fat)

            logger.info(f"Faturamentos agrupados por convênio: { {k: len(v) for k, v in faturamentos_por_convenio.items()} }")

            lotes_criados = []
            for convenio, fats in faturamentos_por_convenio.items():
                logger.info(f"Criando lote para convênio: {convenio}")

                # Criar o lote
                try:
                    lote = Lote.objects.create(
                        empresa_id=empresa_id,
                        convenio=convenio if convenio != 'Sem Convênio' else None
                    )
                    logger.info(f"Lote criado: {lote.id} para convênio {convenio}")
                except Exception as e:
                    logger.error(f"Erro ao criar lote para convênio {convenio}: {e}")
                    messages.error(request, f'Erro ao criar lote para convênio {convenio}: {e}')
                    continue

                # Atualizar os faturamentos com o ID do lote e status para enviado
                fat_ids = [f.id for f in fats]
                try:
                    updated = FaturamentoMedico.objects.filter(id__in=fat_ids).update(lote=str(lote.id), status='enviado')
                    logger.info(f"Faturamentos atualizados para lote {lote.id}: {updated}")
                except Exception as e:
                    logger.error(f"Erro ao atualizar faturamentos para lote {lote.id}: {e}")
                    messages.error(request, f'Erro ao atualizar faturamentos para lote {lote.id}: {e}')
                    continue

                # Atualizar o total do lote
                try:
                    lote.atualizar_total()
                    logger.info(f"Total do lote {lote.id} atualizado: {lote.total_lote}")
                except Exception as e:
                    logger.error(f"Erro ao atualizar total do lote {lote.id}: {e}")
                    messages.error(request, f'Erro ao atualizar total do lote {lote.id}: {e}')
                    continue

                lotes_criados.append(lote.id)

            if lotes_criados:
                url = reverse('faturamento_medico:ftlistar')
                logger.info(f"Lotes criados: {lotes_criados}")
                lotes_str = ', '.join(map(str, lotes_criados))
                return HttpResponse(f'<script>alert("Lotes gerados com sucesso: {lotes_str}"); window.location.href = "{url}";</script>')
            else:
                logger.warning("Nenhum lote foi criado")
                messages.error(request, 'Nenhum lote foi criado.')
                return redirect('faturamento_medico:ftlistar')

    logger.info("Método não é POST, redirecionando")
    return redirect('faturamento_medico:ftlistar')


def importar_unimed(request):
    """View para importar relatório UNIMED"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    if request.method == 'POST':
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            messages.error(request, 'Selecione um arquivo para importar.')
            return redirect('faturamento_medico:importar_unimed')

        try:
            # Ler o arquivo
            content = arquivo.read().decode('utf-8')
            lines = content.split('\n')

            # Pular cabeçalho
            data_lines = lines[1:]

            # Agrupar por lote e guia
            grupos = {}
            servicos_unicos = set()

            for line in data_lines:
                if not line.strip():
                    continue
                parts = line.split(';')
                if len(parts) < 13:
                    continue

                lote = parts[0].strip()
                guia = parts[1].strip()
                cod_usuario = parts[2].strip()
                nome_usuario = parts[3].strip()
                plano = parts[4].strip()
                cod_servico = parts[5].strip()
                desc_servico = parts[6].strip()
                tp_grau = parts[7].strip()
                data_str = parts[8].strip()
                qtde_via = parts[9].strip()
                percentual = parts[10].strip().replace(',', '.')
                valor_unit = parts[11].strip().replace(',', '.')
                valor_total = parts[12].strip().replace(',', '.')
                cod_rel = parts[13].strip()
                observacao = parts[14].strip() if len(parts) > 12 else ''

                # Converter data
                try:
                    data = datetime.strptime(data_str, '%d/%m/%Y').date()
                except:
                    data = timezone.now().date()

                chave = f"{lote}_{guia}"

                if chave not in grupos:
                    grupos[chave] = {
                        'lote': lote,
                        'guia': guia,
                        'carteirinha': cod_usuario,
                        'nome': nome_usuario,
                        'plano': plano,
                        'data': data,
                        'cod_rel': cod_rel,
                        'servicos': []
                    }

                grupos[chave]['servicos'].append({
                    'codigo': cod_servico,
                    'descricao': desc_servico,
                    'porte': tp_grau,
                    'qt': int(float(qtde_via)) if qtde_via else 1,
                    'percentual': float(percentual) if percentual else 0,
                    'valor': float(valor_unit) if valor_unit else 0,
                    'total': float(valor_total) if valor_total else 0,
                    'observacao': observacao
                })

                servicos_unicos.add((cod_servico, desc_servico))

            # Verificar e criar serviços não cadastrados
            from servicos_medicos.models import ServicosMedicos
            servicos_criados = 0
            for cod, desc in servicos_unicos:
                if not ServicosMedicos.objects.filter(codigo=cod).exists():
                    ServicosMedicos.objects.create(
                        codigo=cod,
                        servicos=desc,
                        porte_anestesico=None  # Será definido depois se necessário
                    )
                    servicos_criados += 1

            # Criar faturamentos
            faturamentos_criados = 0
            itens_criados = 0

            for chave, dados in grupos.items():
                # Criar faturamento
                faturamento = FaturamentoMedico.objects.create(
                    empresa_id=empresa_id,
                    lote=dados['lote'],
                    guia=dados['guia'],
                    carteirinha=dados['carteirinha'],
                    nome=dados['nome'],
                    data=dados['data'],
                    convenio='UNIMED',
                    codigo_relatorio=dados['cod_rel'],
                    status='pendente'
                )

                # Criar itens de serviço
                for servico in dados['servicos']:
                    ItemServico.objects.create(
                        faturamento=faturamento,
                        codigo_servico=servico['codigo'],
                        servico=servico['descricao'],
                        porte=servico['porte'],
                        percentual = servico['percentual'],
                        qt=servico['qt'],
                        valor=servico['valor'],
                        total=servico['total']
                    )
                    itens_criados += 1

                # Atualizar total do faturamento
                faturamento.atualizar_total()
                faturamentos_criados += 1

            messages.success(request, f'Importação concluída! {servicos_criados} serviços criados, {faturamentos_criados} faturamentos criados, {itens_criados} itens de serviço criados.')

        except Exception as e:
            messages.error(request, f'Erro durante a importação: {str(e)}')
            return redirect('faturamento_medico:importar_unimed')

        return redirect('faturamento_medico:ftlistar')

    context = {
        'titulo': 'Importar Relatório UNIMED'
    }

    return render(request, 'faturamento_medico/importar_unimed.html', context)


def importar_xml(request):
    """View para importar XML de NFSe"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Empresa não encontrada na sessão.')
        return redirect('faturamento_medico:ftlistar')

    if request.method == 'POST':
        arquivos = request.FILES.getlist('arquivos')
        if not arquivos:
            messages.error(request, 'Selecione pelo menos um arquivo XML para importar.')
            return redirect('faturamento_medico:importar_xml')

        try:
            import xml.etree.ElementTree as ET
            from servicos_medicos.models import ServicosMedicos

            faturamentos_criados = 0
            itens_criados = 0
            servicos_criados = 0
            servicos_unicos = set()

            for arquivo in arquivos:
                # Parse do XML
                content = arquivo.read().decode('utf-8')
                root = ET.fromstring(content)

                # Namespace do XML
                ns = {'nfse': 'http://www.abrasf.org.br/nfse.xsd'}

                # Encontrar o elemento Nfse
                nfse = root.find('.//nfse:Nfse', ns)
                if nfse is None:
                    continue

                inf_nfse = nfse.find('nfse:InfNfse', ns)
                if inf_nfse is None:
                    continue

                # Extrair dados básicos
                numero = inf_nfse.findtext('nfse:Numero', default='')
                data_emissao_str = inf_nfse.findtext('nfse:DataEmissao', default='')
                outras_info = inf_nfse.findtext('nfse:OutrasInformacoes', default='')

                # Valores NFSe
                valores_nfse = inf_nfse.find('nfse:ValoresNfse', ns)
                valor_liquido = 0.0
                if valores_nfse is not None:
                    valor_liquido_str = valores_nfse.findtext('nfse:ValorLiquidoNfse', default='0')
                    try:
                        valor_liquido = float(valor_liquido_str)
                    except:
                        valor_liquido = 0.0

                # Prestador
                prestador_servico = inf_nfse.find('nfse:PrestadorServico', ns)
                prestador_cnpj = ''
                prestador_razao = ''
                if prestador_servico is not None:
                    identificacao = prestador_servico.find('nfse:IdentificacaoPrestador', ns)
                    if identificacao is not None:
                        cpf_cnpj = identificacao.find('nfse:CpfCnpj', ns)
                        if cpf_cnpj is not None:
                            prestador_cnpj = cpf_cnpj.findtext('nfse:Cnpj', default='')
                    prestador_razao = prestador_servico.findtext('nfse:RazaoSocial', default='')

                # Tomador
                declaracao = inf_nfse.find('nfse:DeclaracaoPrestacaoServico', ns)
                tomador_nome = ''
                tomador_cpf = ''
                if declaracao is not None:
                    inf_declaracao = declaracao.find('nfse:InfDeclaracaoPrestacaoServico', ns)
                    if inf_declaracao is not None:
                        tomador = inf_declaracao.find('nfse:Tomador', ns)
                        if tomador is not None:
                            tomador_nome = tomador.findtext('nfse:RazaoSocial', default='')
                            identificacao_tomador = tomador.find('nfse:IdentificacaoTomador', ns)
                            if identificacao_tomador is not None:
                                cpf_cnpj_tomador = identificacao_tomador.find('nfse:CpfCnpj', ns)
                                if cpf_cnpj_tomador is not None:
                                    tomador_cpf = cpf_cnpj_tomador.findtext('nfse:Cpf', default='')

                        # Serviço
                        servico = inf_declaracao.find('nfse:Servico', ns)
                        if servico is not None:
                            discriminacao = servico.findtext('nfse:Discriminacao', default='')
                            item_lista = servico.findtext('nfse:ItemListaServico', default='')
                            codigo_cnae = servico.findtext('nfse:CodigoCnae', default='')
                            competencia_str = inf_declaracao.findtext('nfse:Competencia', default='')

                            # Valores do serviço
                            valores_servico = servico.find('nfse:Valores', ns)
                            valor_servicos = 0.0
                            if valores_servico is not None:
                                valor_servicos_str = valores_servico.findtext('nfse:ValorServicos', default='0')
                                try:
                                    valor_servicos = float(valor_servicos_str)
                                except:
                                    valor_servicos = 0.0

                            # Converter datas
                            data_emissao = None
                            try:
                                if data_emissao_str:
                                    data_emissao = datetime.fromisoformat(data_emissao_str.replace('Z', '+00:00'))
                            except:
                                data_emissao = timezone.now()

                            competencia = None
                            try:
                                if competencia_str:
                                    competencia = datetime.strptime(competencia_str, '%Y-%m-%d').date()
                            except:
                                competencia = data_emissao.date() if data_emissao else timezone.now().date()

                            # Criar faturamento
                            faturamento = FaturamentoMedico.objects.create(
                                empresa_id=empresa_id,
                                guia=numero,
                                nome=tomador_nome,
                                carteirinha=tomador_cpf,
                                data=competencia or timezone.now().date(),
                                data_autorizacao=data_emissao.date() if data_emissao else None,
                                total=valor_liquido,
                                convenio='NFSE',
                                codigo_relatorio='1',
                                status='pendente',
                                observacao=f"{discriminacao}\n{outras_info}".strip()
                            )

                            # Criar item de serviço
                            ItemServico.objects.create(
                                faturamento=faturamento,
                                codigo_servico=item_lista,
                                servico=discriminacao,
                                porte='',  # NFSe não tem porte anestésico
                                qt=1,
                                valor=valor_servicos,
                                total=valor_liquido
                            )
                            itens_criados += 1

                            # Adicionar serviço único para possível criação
                            servicos_unicos.add((item_lista, discriminacao))

                            faturamentos_criados += 1

            # Verificar e criar serviços não cadastrados
            for cod, desc in servicos_unicos:
                if cod and not ServicosMedicos.objects.filter(codigo=cod).exists():
                    ServicosMedicos.objects.create(
                        codigo=cod,
                        servicos=desc,
                        porte_anestesico=None
                    )
                    servicos_criados += 1

            messages.success(request, f'Importação XML concluída! {servicos_criados} serviços criados, {faturamentos_criados} faturamentos criados, {itens_criados} itens de serviço criados.')

        except Exception as e:
            messages.error(request, f'Erro durante a importação XML: {str(e)}')
            return redirect('faturamento_medico:importar_xml')

        return redirect('faturamento_medico:ftlistar')

    context = {
        'titulo': 'Importar XML NFSe'
    }

    return render(request, 'faturamento_medico/importar_xml.html', context)
