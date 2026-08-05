from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.http import HttpResponseRedirect
from django.views.decorators.http import require_POST
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Prefetch
from django.db import transaction
from django.utils import timezone
from django.http import JsonResponse
from django.core.paginator import Paginator
from datetime import timedelta, date
from decimal import Decimal
from calendar import monthrange
from dateutil.relativedelta import relativedelta
from .models import ContaAReceber, BaixaContaAReceber
from .forms import BaixaContaIndividualForm, EscolhaContaBaixaForm
from cliente.models import Cliente
from cobranca.models import Cobranca
from empresa.models import Empresa
from notasfiscais.models import NotaFiscalServico
from socio.models import Socio
from categoria.models import Categoria
from extrato.models import ContaBancaria
import re
import json
# import Levenshtein  # Temporariamente desabilitado - módulo não instalado


def _redirect_sem_empresa(request):
    """Evita loop infinito (crlistar → crlistar) quando a sessão perde empresa_id."""
    messages.error(request, 'Empresa não encontrada na sessão.')
    return redirect('empresa:lista')


def _parcela_para_titulo_diferenca(conta):
    """Parcela do novo título gerado pela diferença (tenta sequência da NF; senão rótulo padrão)."""
    p = (conta.parcela or '').strip().replace(' ', '')
    if p and '/' in p:
        try:
            n_s, tot_s = p.split('/', 1)
            n_i, tot_i = int(n_s), int(tot_s)
            if tot_i > 0 and 1 <= n_i < tot_i:
                return f'{n_i + 1}/{tot_i}'
        except ValueError:
            pass
    return 'Parc. dif. 1/1'


def _parcela_primeira_apos_split(conta):
    """Ao dividir 1 título em 2 parcelas: primeira parcela (ex.: 1/1 → 1/2)."""
    p = (conta.parcela or '').strip().replace(' ', '')
    if p and '/' in p:
        try:
            n_s, tot_s = p.split('/', 1)
            n_i, tot_i = int(n_s), int(tot_s)
            if tot_i > 0 and n_i >= 1:
                return f'{n_i}/{tot_i + 1}'
        except ValueError:
            pass
    return '1/2'


def _decimal_para_js_attr(valor):
    """Decimal com ponto fixo para value em HTML/parseFloat (evita locale 400,00)."""
    if valor is None:
        return '0.00'
    d = Decimal(str(valor))
    return f'{d:.2f}'


def _context_valores_js_baixa(conta):
    """Saldo nominal = valor parcela − já recebido (quitação do título: líquido nominal ≈ valor recebido informado na baixa)."""
    v = conta.valor_a_receber or Decimal('0')
    r = conta.valor_recebido or Decimal('0')
    saldo = max(v - r, Decimal('0'))
    return {
        'js_saldo_nominal': _decimal_para_js_attr(saldo),
        'js_valor_parcela': _decimal_para_js_attr(v),
        'js_valor_recebido_conta': _decimal_para_js_attr(r),
    }


def _total_recebido_efetivo_conta(conta):
    """
    Valor efetivamente creditado no banco (soma das baixas): VR + juros − desconto por baixa
    (tarifa não reduz o líquido do extrato; ela entra na quitação do título em separado).
    Se não houver baixa gravada, usa o campo na parcela; se status pago/cartão com zero, usa o face (alinhado ao resumo fechamento).
    """
    s = Decimal('0')
    for bx in conta.baixas.all():
        s += bx.valor_total_com_ajustes()
    if s != 0:
        return s
    vr = conta.valor_recebido
    if vr is not None and vr != 0:
        return vr if isinstance(vr, Decimal) else Decimal(str(vr))
    if conta.status in ('pago', 'cartao') and (vr is None or vr == 0):
        return conta.valor_a_receber or Decimal('0')
    return Decimal('0')


def _par_datas_iso(ini, fim):
    """Garante ini <= fim para strings YYYY-MM-DD."""
    a = (ini or '').strip()
    b = (fim or '').strip()
    if len(a) >= 10 and len(b) >= 10 and a > b:
        return b, a
    return a, b


def extrair_filtros_contas_receber(request):
    """Extrai filtros da tela de contas a receber do request (validação: strip em textos)."""
    fp_list = [x.strip() for x in request.GET.getlist('forma_pagamento') if x and str(x).strip()]
    if not fp_list:
        fp_single = (request.GET.get('forma_pagamento') or '').strip()
        if fp_single:
            fp_list = [fp_single]
    return {
        'search': (request.GET.get('search') or '').strip(),
        'status': (request.GET.get('status') or '').strip(),
        'forma_pagamento': fp_list,
        'socio': (request.GET.get('socio') or '').strip(),
        'data_inicio': (request.GET.get('data_inicio') or '').strip(),
        'data_fim': (request.GET.get('data_fim') or '').strip(),
        'vencimento_inicio': (request.GET.get('vencimento_inicio') or '').strip(),
        'vencimento_fim': (request.GET.get('vencimento_fim') or '').strip(),
        'recebimento_inicio': (request.GET.get('recebimento_inicio') or '').strip(),
        'recebimento_fim': (request.GET.get('recebimento_fim') or '').strip(),
        'per_page': (request.GET.get('per_page') or '25').strip(),
    }


def _aplicar_filtro_socio_car(queryset, socio_val):
    """Filtra por sócio na conta ou na NF (compatível com registros antigos)."""
    if not socio_val or not str(socio_val).strip():
        return queryset
    s = str(socio_val).strip()
    if s in ('sem', 'none'):
        q = (
            Q(socio__isnull=True, nota__isnull=True)
            | Q(socio__isnull=True, nota__socio__isnull=True)
        )
        return queryset.filter(q).distinct()
    if s.isdigit():
        sid = int(s)
        return queryset.filter(Q(socio_id=sid) | Q(nota__socio_id=sid)).distinct()
    return queryset


def construir_url_crlistar_com_filtros(filtros):
    """Constrói URL para crlistar com filtros preservados (listas → vários parâmetros GET)."""
    from urllib.parse import urlencode

    pairs = []
    for k, v in filtros.items():
        if v is None or v == '' or v == []:
            continue
        if isinstance(v, (list, tuple)):
            for item in v:
                if item is not None and str(item).strip():
                    pairs.append((k, str(item).strip()))
        else:
            pairs.append((k, v))
    if pairs:
        return f"{reverse('contasareceber:crlistar')}?{urlencode(pairs)}"
    return reverse('contasareceber:crlistar')


CAR_DEFAULT_SORT = 'data_vencimento'
CAR_DEFAULT_SORT_DIR = 'desc'

CAR_SORT_FIELDS = {
    'cliente': 'cliente',
    'cnpj_cpf': 'cnpj_cpf',
    'socio': 'socio',
    'categoria': 'categoria',
    'data_emissao': 'data_emissao',
    'data_recebimento': 'data_recebimento',
    'documento': 'documento',
    'autorizacao': 'autorizacao',
    'cobranca': 'cobranca',
    'data_vencimento': 'data_vencimento',
    'valor': 'valor',
    'saldo_nominal': 'saldo_nominal',
    'total_liquido': 'total_liquido',
    'desconto': 'desconto',
    'juros': 'juros',
    'tarifas': 'tarifas',
    'status': 'status',
    'dias_atraso': 'dias_atraso',
}


def _get_car_sort_from_request(request, filtros):
    sort_get = (request.GET.get('sort') or '').strip()
    dir_get = (request.GET.get('dir') or '').strip().lower()
    if sort_get in CAR_SORT_FIELDS:
        filtros['sort'] = sort_get
        filtros['dir'] = dir_get if dir_get in ('asc', 'desc') else 'asc'

    sort_col = filtros.get('sort', CAR_DEFAULT_SORT)
    sort_dir = filtros.get('dir', CAR_DEFAULT_SORT_DIR)
    if sort_col not in CAR_SORT_FIELDS:
        sort_col = CAR_DEFAULT_SORT
    if sort_dir not in ('asc', 'desc'):
        sort_dir = CAR_DEFAULT_SORT_DIR
    return sort_col, sort_dir


def _car_sort_key(conta, sort_col):
    if sort_col == 'cliente':
        return (conta.cliente or '').lower()
    if sort_col == 'cnpj_cpf':
        return (conta.cnpj_cpf or '').lower()
    if sort_col == 'socio':
        if conta.socio:
            return str(conta.socio).lower()
        if conta.nota and conta.nota.socio:
            return str(conta.nota.socio).lower()
        return ''
    if sort_col == 'categoria':
        return (conta.categoria.nome if conta.categoria else '').lower()
    if sort_col == 'data_emissao':
        return conta.data_emissao or date.min
    if sort_col == 'data_recebimento':
        return conta.data_recebimento or date.min
    if sort_col == 'documento':
        doc = conta.doc or (conta.nota.numero_nota if conta.nota else '') or ''
        return str(doc).lower()
    if sort_col == 'autorizacao':
        return (conta.autorizacao or '').lower()
    if sort_col == 'cobranca':
        return (conta.forma_pagamento.descricao if conta.forma_pagamento else '').lower()
    if sort_col == 'data_vencimento':
        return conta.data_vencimento or date.min
    if sort_col == 'valor':
        return conta.valor_a_receber or Decimal('0')
    if sort_col == 'saldo_nominal':
        return conta.get_saldo_nominal_pendente()
    if sort_col == 'total_liquido':
        return conta.get_total_liquido_listagem()
    if sort_col == 'desconto':
        return conta.desconto or Decimal('0')
    if sort_col == 'juros':
        return conta.juros or Decimal('0')
    if sort_col == 'tarifas':
        return conta.tarifas or Decimal('0')
    if sort_col == 'status':
        return conta.status or ''
    if sort_col == 'dias_atraso':
        return conta.dias_atraso
    return conta.data_vencimento or date.min


def _apply_car_sort(contas_list, sort_col, sort_dir):
    reverse = sort_dir == 'desc'
    return sorted(contas_list, key=lambda c: _car_sort_key(c, sort_col), reverse=reverse)


def extrair_filtros_categorizar_baixados(request):
    """Filtros da tela de categorização de contas já recebidas (pago/cartão)."""
    return {
        'search': (request.GET.get('search') or '').strip(),
        'forma_pagamento': request.GET.getlist('forma_pagamento'),
        'data_inicio': (request.GET.get('data_inicio') or '').strip(),
        'data_fim': (request.GET.get('data_fim') or '').strip(),
        'per_page': (request.GET.get('per_page') or '50').strip(),
    }


def categorizar_recebidos_baixados(request):
    """Lista contas já baixadas (pago/cartão) para aplicar categoria em lote, com filtros por cobrança e data de recebimento."""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    filtros = extrair_filtros_categorizar_baixados(request)
    per_page = filtros['per_page']
    try:
        per_page = int(per_page)
        if per_page not in [25, 50, 100, 150, 500, 900]:
            per_page = 50
    except ValueError:
        per_page = 50

    hoje = timezone.now().date()
    data_inicio = filtros['data_inicio'] or (hoje - timedelta(days=365)).strftime('%Y-%m-%d')
    data_fim = filtros['data_fim'] or hoje.strftime('%Y-%m-%d')
    if data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio

    contas = (
        ContaAReceber.objects.filter(empresa_id=empresa_id, status__in=['pago', 'cartao'])
        .filter(data_recebimento__isnull=False)
        .filter(data_recebimento__gte=data_inicio, data_recebimento__lte=data_fim)
        .select_related('categoria', 'forma_pagamento', 'nota', 'nota__socio')
        .order_by('-data_recebimento', '-id')
    )

    if filtros['search']:
        contas = contas.filter(
            Q(nota__numero_nota__icontains=filtros['search'], nota__isnull=False)
            | Q(cliente__icontains=filtros['search'])
            | Q(cnpj_cpf__icontains=filtros['search'])
            | Q(doc__icontains=filtros['search'])
        )

    fp_ids = []
    for fp in filtros['forma_pagamento']:
        if fp and str(fp).strip():
            try:
                fp_ids.append(int(fp))
            except ValueError:
                continue
    if fp_ids:
        contas = contas.filter(forma_pagamento_id__in=fp_ids)

    categorias = Categoria.objects.filter(empresa_id=empresa_id).order_by('classificacao', 'nome')
    formas_pagamento = Cobranca.objects.all().order_by('descricao')

    paginator = Paginator(contas, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    q_nav = request.GET.copy()
    if 'page' in q_nav:
        del q_nav['page']
    filtros_query_sem_page = q_nav.urlencode()

    context = {
        'contas': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'categorias': categorias,
        'formas_pagamento': formas_pagamento,
        'filtros_query_sem_page': filtros_query_sem_page,
        'filtros': {
            'search': filtros['search'],
            'forma_pagamento': filtros['forma_pagamento'],
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'per_page': str(per_page),
        },
    }
    return render(request, 'contasareceber/categorizar_baixados.html', context)


def criar_conta_a_receber(request):
    """Cria uma nova conta a receber"""
    if request.method == 'POST':
        empresa_id = request.session.get('empresa_id')
        if not empresa_id:
            return _redirect_sem_empresa(request)

        try:
            empresa = Empresa.objects.get(id=empresa_id)

            # Obter dados do formulário
            cliente_nome = request.POST.get('cliente')
            cnpj_cpf = request.POST.get('cnpj_cpf')
            valor_str = request.POST.get('valor')
            # Convert Brazilian decimal format (comma) to Python decimal format (dot)
            if valor_str:
                valor = float(valor_str.replace(',', '.'))
            else:
                valor = 0
            data_emissao = request.POST.get('data_emissao')
            data_vencimento = request.POST.get('data_vencimento')
            numero_nota = request.POST.get('nota')
            doc = request.POST.get('doc')
            autorizacao = request.POST.get('autorizacao')
            forma_pagamento_id = request.POST.get('forma_pagamento')
            conta_bancaria_id = request.POST.get('conta_bancaria')

            # Criar ou obter nota fiscal apenas se numero_nota for fornecido
            nota = None
            if numero_nota:
                nota, created = NotaFiscalServico.objects.get_or_create(
                    empresa=empresa,
                    numero_nota=numero_nota,
                    defaults={
                        'cliente': cliente_nome or 'CLIENTE DIVERSOS',
                        'valor_bruto': valor or 0,
                        'valor_liquido': valor or 0,
                        'data_emissao': data_emissao or timezone.now().date(),
                    }
                )

            socio_id = request.POST.get('socio')
            socio_obj = None
            if socio_id and str(socio_id).strip().isdigit():
                try:
                    socio_obj = Socio.objects.get(id=int(socio_id), empresa_id=empresa_id)
                except Socio.DoesNotExist:
                    pass

            if nota and socio_obj:
                nota.socio = socio_obj
                nota.save(update_fields=['socio'])

            if valor <= 0:
                messages.error(request, 'O valor a receber deve ser maior que zero.')
                filtros = extrair_filtros_contas_receber(request)
                return redirect(construir_url_crlistar_com_filtros(filtros))

            if nota and nota.is_cancelada():
                messages.error(request, 'Não é possível criar conta a receber para nota cancelada.')
                filtros = extrair_filtros_contas_receber(request)
                return redirect(construir_url_crlistar_com_filtros(filtros))

            # DEBUG: Log valores atuais antes da criação da conta
            data_emissao_nota = nota.data_emissao if nota else (data_emissao or timezone.now().date())
            forma_pagamento_desc = None
            if forma_pagamento_id:
                try:
                    forma_pgto_obj = Cobranca.objects.get(id=forma_pagamento_id)
                    forma_pagamento_desc = forma_pgto_obj.descricao
                except Cobranca.DoesNotExist:
                    forma_pagamento_desc = "FORMA_PAGAMENTO_NAO_ENCONTRADA"

            print(f"DEBUG CRIAR_CONTA_RECEBER: data_emissao_nota={data_emissao_nota}, data_vencimento_atual={data_vencimento}, forma_pagamento_id={forma_pagamento_id}, forma_pagamento_desc={forma_pagamento_desc}")

            # Criar conta a receber
            conta = ContaAReceber.objects.create(
                empresa=empresa,
                nota=nota,
                socio=socio_obj,
                cliente=cliente_nome or (nota.cliente if nota else 'CLIENTE DIVERSOS'),
                cnpj_cpf=cnpj_cpf or (nota.cnpj_cpf if nota else None),
                data_emissao=data_emissao or timezone.now().date(),
                data_vencimento=data_vencimento,
                valor_a_receber=valor,
                doc=doc,
                autorizacao=autorizacao,
                forma_pagamento_id=forma_pagamento_id if forma_pagamento_id else None,
            )

            # Associar conta bancária se fornecida
            # Temporariamente desabilitado para evitar erro de foreign key
            # if conta_bancaria_id:
            #     try:
            #         conta_bancaria = ContaBancaria.objects.get(id=conta_bancaria_id, empresa=empresa)
            #         conta.conta_banco = conta_bancaria
            #         conta.save()
            #     except ContaBancaria.DoesNotExist:
            #         pass

            # Categoria (sócio já aplicado em nota + conta na criação)
            categoria_id = request.POST.get('categoria')

            if categoria_id:
                try:
                    categoria = Categoria.objects.get(id=categoria_id, empresa_id=empresa_id)
                    conta.categoria = categoria
                    conta.save()
                except Categoria.DoesNotExist:
                    pass
            else:
                # Se não foi passado categoria_id, tentar buscar pelo nome da categoria
                categoria_nome = request.POST.get('categoria')
                if categoria_nome:
                    try:
                        # Tentar encontrar categoria pelo nome completo (classificacao + nome)
                        categoria = Categoria.objects.filter(
                            empresa_id=empresa_id
                        ).filter(
                            Q(nome__iexact=categoria_nome.split(' ', 1)[-1]) |
                            Q(classificacao__iexact=categoria_nome.split(' ')[0])
                        ).first()
                        if categoria:
                            conta.categoria = categoria
                            conta.save()
                    except Exception:
                        pass

            messages.success(request, 'Conta a receber criada com sucesso!')
            filtros = extrair_filtros_contas_receber(request)
            return redirect(construir_url_crlistar_com_filtros(filtros))

        except Exception as e:
            messages.error(request, f'Erro ao criar conta: {str(e)}')
            filtros = extrair_filtros_contas_receber(request)
            return redirect(construir_url_crlistar_com_filtros(filtros))

    empresa_id = request.session.get('empresa_id')
    clientes = Cliente.objects.filter(empresa_id=empresa_id) if empresa_id else Cliente.objects.none()
    contas_bancarias = ContaBancaria.objects.filter(empresa_id=empresa_id, status='A') if empresa_id else ContaBancaria.objects.none()
    formas_pagamento = Cobranca.objects.all()
    socios = Socio.objects.filter(empresa_id=empresa_id) if empresa_id else Socio.objects.none()
    categorias = Categoria.objects.filter(empresa_id=empresa_id) if empresa_id else Categoria.objects.none()
    hoje = timezone.now().date()
    context = {
        'clientes': clientes,
        'contas_bancarias': contas_bancarias,
        'formas_pagamento': formas_pagamento,
        'socios': socios,
        'categorias': categorias,
        'hoje': hoje,
        'conta_banco_selecionado_id': None,  # Para compatibilidade com o template
    }
    return render(request, 'contasareceber/criar.html', context)

def listar_contas_a_receber(request):
    """Lista todas as contas a receber com filtros"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    contas = (
        ContaAReceber.objects.filter(empresa_id=empresa_id)
        .select_related('nota', 'nota__socio', 'socio', 'categoria', 'forma_pagamento')
        .order_by('-data_vencimento')
    )

    # Extrair filtros usando a função helper
    filtros = extrair_filtros_contas_receber(request)
    search = filtros['search']
    status = filtros['status']
    forma_pagamento = filtros['forma_pagamento']
    socio_filtro = filtros['socio']
    data_inicio = filtros['data_inicio']
    data_fim = filtros['data_fim']
    per_page = filtros['per_page']

    # Paginação
    try:
        per_page = int(per_page)
        if per_page not in [25, 50, 100, 150, 500, 900]:
            per_page = 25
    except ValueError:
        per_page = 25

    # Buscar formas de pagamento
    formas_pagamento = Cobranca.objects.all()

    # Buscar categorias
    categorias = Categoria.objects.filter(empresa_id=empresa_id)

    # Definir datas padrão (últimos 12 meses)
    hoje = timezone.now().date()
    data_inicio_padrao = hoje - timedelta(days=365)
    data_fim_padrao = hoje

    # Se não há datas selecionadas, usar as datas padrão
    if not data_inicio:
        data_inicio = data_inicio_padrao.strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = data_fim_padrao.strftime('%Y-%m-%d')
    # Garantir data_inicio <= data_fim (trocar se invertidas)
    if data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio

    # Refletir no template as datas realmente usadas na query (evita intervalo vazio na tela vs. últimos 12 meses no backend)
    filtros = dict(filtros)
    filtros['data_inicio'] = data_inicio
    filtros['data_fim'] = data_fim
    filtros['per_page'] = str(per_page)

    venc_ini, venc_fim = _par_datas_iso(
        filtros.get('vencimento_inicio', ''),
        filtros.get('vencimento_fim', ''),
    )
    filtros['vencimento_inicio'] = venc_ini
    filtros['vencimento_fim'] = venc_fim

    rec_ini, rec_fim = _par_datas_iso(
        filtros.get('recebimento_inicio', ''),
        filtros.get('recebimento_fim', ''),
    )
    filtros['recebimento_inicio'] = rec_ini
    filtros['recebimento_fim'] = rec_fim

    # Aplicar filtros
    if search:
        contas = contas.filter(
            Q(nota__numero_nota__icontains=search, nota__isnull=False)
            | Q(cliente__icontains=search)
            | Q(cnpj_cpf__icontains=search)
            | Q(doc__icontains=search)
            | Q(autorizacao__icontains=search)
        )

    # Drill-down a partir do Resumo fechamento (coluna «Val. a receber»): mesma base do relatório (pendente/vencido)
    if (request.GET.get('resumo_val_receber') or '').strip() == '1':
        contas = contas.filter(status__in=['pendente', 'vencido'])
    elif status:
        contas = contas.filter(status=status)

    # Filtrar e converter forma_pagamento para inteiros válidos
    if forma_pagamento:
        forma_pagamento_ids = []
        for fp in forma_pagamento:
            if fp.strip():  # Verificar se não está vazio
                try:
                    fp_id = int(fp)
                    forma_pagamento_ids.append(fp_id)
                except ValueError:
                    continue  # Ignorar valores inválidos
        if forma_pagamento_ids:
            contas = contas.filter(forma_pagamento_id__in=forma_pagamento_ids)

    contas = _aplicar_filtro_socio_car(contas, socio_filtro)

    # Emissão: intervalo (padrão últimos 12 meses se vazio no GET)
    contas = contas.filter(data_emissao__gte=data_inicio)
    contas = contas.filter(data_emissao__lte=data_fim)

    # Vencimento (opcional)
    if venc_ini:
        contas = contas.filter(data_vencimento__gte=venc_ini)
    if venc_fim:
        contas = contas.filter(data_vencimento__lte=venc_fim)

    # Recebimento (opcional): apenas títulos com data de recebimento preenchida
    if rec_ini or rec_fim:
        contas = contas.filter(data_recebimento__isnull=False)
        if rec_ini:
            contas = contas.filter(data_recebimento__gte=rec_ini)
        if rec_fim:
            contas = contas.filter(data_recebimento__lte=rec_fim)

    contas = contas.prefetch_related('baixas')

    sort_col, sort_dir = _get_car_sort_from_request(request, filtros)
    filtros['sort'] = sort_col
    filtros['dir'] = sort_dir

    # Uma única materialização: cards e tabela usam o mesmo conjunto (evita divergência queryset vs página).
    contas_filtradas = list(contas)
    contas_filtradas = _apply_car_sort(contas_filtradas, sort_col, sort_dir)
    total_pendente = sum(
        conta.get_valor_pendente() for conta in contas_filtradas if conta.status not in ('pago', 'cartao')
    )
    total_vencido = sum(
        conta.get_valor_pendente()
        for conta in contas_filtradas
        if conta.is_vencida() and conta.status not in ('pago', 'cartao')
    )
    total_recebido = sum(_total_recebido_efetivo_conta(conta) for conta in contas_filtradas)
    total_contas_receber = sum(conta.valor_a_receber or Decimal('0') for conta in contas_filtradas)

    paginator = Paginator(contas_filtradas, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'contas': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'total_pendente': total_pendente,
        'total_vencido': total_vencido,
        'total_recebido': total_recebido,
        'total_contas_receber': total_contas_receber,
        'formas_pagamento': formas_pagamento,
        'categorias': categorias,
        'search': search,
        'status_filter': status,
        'forma_pagamento_filter': forma_pagamento,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'per_page': per_page,
        'filtros': filtros,
        'socios': Socio.objects.filter(empresa_id=empresa_id).order_by('socio', 'lastname'),
        'resumo_val_receber': (request.GET.get('resumo_val_receber') or '').strip() == '1',
        'sort_col': sort_col,
        'sort_dir': sort_dir,
    }

    return render(request, 'contasareceber/listar.html', context)


@login_required
@require_POST
def alterar_socio_lote(request):
    """Define o sócio na NF vinculada e replica o mesmo sócio em todas as parcelas (contas a receber) da nota."""
    from .socio_sync import propagar_socio_nota_para_contas_receber

    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    contas_ids = request.POST.getlist('contas_ids')
    socio_id_raw = (request.POST.get('socio_id') or '').strip()
    return_path = (request.POST.get('return_path') or '').strip()

    if socio_id_raw and not socio_id_raw.isdigit():
        messages.error(request, 'Seleção de sócio inválida.')
        return _redirect_car_com_return(request, return_path)

    if not contas_ids:
        messages.error(request, 'Nenhuma conta selecionada.')
        return _redirect_car_com_return(request, return_path)

    socio_obj = None
    if socio_id_raw.isdigit():
        socio_obj = Socio.objects.filter(pk=int(socio_id_raw), empresa_id=empresa_id).first()
        if not socio_obj:
            messages.error(request, 'Sócio inválido ou de outra empresa.')
            return _redirect_car_com_return(request, return_path)

    contas = ContaAReceber.objects.filter(id__in=contas_ids, empresa_id=empresa_id).select_related('nota')
    atualizadas = 0
    sem_nota = 0
    notas_ja_gravadas = set()
    for conta in contas:
        if not conta.nota_id:
            sem_nota += 1
            continue
        nf = conta.nota
        if nf.pk in notas_ja_gravadas:
            continue
        notas_ja_gravadas.add(nf.pk)
        nf.socio = socio_obj
        nf.save(update_fields=['socio'])
        propagar_socio_nota_para_contas_receber(nf)
        atualizadas += 1

    if atualizadas:
        messages.success(
            request,
            f'Sócio atualizado em {atualizadas} nota(s) fiscal(is) e nas contas a receber vinculadas.',
        )
    else:
        messages.warning(
            request,
            'Nenhuma nota fiscal vinculada às contas selecionadas; o sócio é definido na NF.',
        )
    if sem_nota:
        messages.info(
            request,
            f'{sem_nota} conta(s) sem NF vinculada foram ignoradas (cadastre NF ou edite a conta).',
        )

    return _redirect_car_com_return(request, return_path)


def _redirect_car_com_return(request, return_path):
    if return_path.startswith('/contasareceber'):
        return redirect(return_path)
    return redirect('contasareceber:crlistar')


@login_required
def detalhes_conta_a_receber(request, pk):
    """Exibe detalhes de uma conta a receber"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    conta = get_object_or_404(
        ContaAReceber.objects.prefetch_related(
            Prefetch(
                'baixas',
                queryset=BaixaContaAReceber.objects.prefetch_related('extrato_movimentos'),
            )
        ),
        pk=pk,
        empresa_id=empresa_id,
    )

    # Capturar parâmetros GET para preservar filtros na navegação
    filtros_query = request.GET.urlencode()

    context = {
        'conta': conta,
        'filtros_query': filtros_query,
    }

    return render(request, 'contasareceber/detalhes.html', context)

def baixar_conta_a_receber(request, pk):
    """Permite baixar (receber) uma conta a receber"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    conta = get_object_or_404(ContaAReceber, pk=pk, empresa_id=empresa_id)

    # Capturar parâmetros GET para preservar filtros na navegação
    filtros_query = request.GET.urlencode()

    # Verificar se a conta já foi totalmente paga
    if conta.status == 'pago':
        messages.warning(request, 'Esta conta já foi totalmente paga. Use a opção de estorno para reverter o pagamento.')
        return redirect('contasareceber:estornar', pk=pk)

    if request.method == 'POST':
        print(f"DEBUG: POST request received for conta {pk}")
        # Obter parâmetros de período do POST
        data_inicio = request.POST.get('data_inicio')
        data_fim = request.POST.get('data_fim')
        print(f"DEBUG: data_inicio={data_inicio}, data_fim={data_fim}")

        form = BaixaContaIndividualForm(request.POST, empresa_id=empresa_id, conta=conta, data_inicio=data_inicio, data_fim=data_fim)
        print(f"DEBUG: Form is_valid: {form.is_valid()}")

        if not form.is_valid():
            print(f"DEBUG: Form errors: {form.errors}")
            print(f"DEBUG: Form non_field_errors: {form.non_field_errors()}")

        if form.is_valid():
            print("DEBUG: Form is valid, processing baixa...")
            # Verifica se há lançamentos do extrato selecionados para conciliação
            lancamentos_ids_str = form.cleaned_data.get('lancamentos_extrato_ids', '')
            lancamentos_selecionados = []
            if lancamentos_ids_str:
                from extrato.models import Lancamento
                raw_ids = [int(x.strip()) for x in lancamentos_ids_str.split(',') if x.strip()]
                lancamentos_ids = list(dict.fromkeys(raw_ids))  # sem duplicar IDs
                lancamentos_selecionados = Lancamento.objects.filter(id__in=lancamentos_ids)

            # Verifica se a conta bancária selecionada permite baixa sem lançamentos
            conta_banco = form.cleaned_data.get('conta_banco')
            tipos_permitidos_sem_lancamentos = ['CAIXA', 'INVESTIMENTO', 'EMPRESTIMO', 'FATURA_CARTAO']

            # Se não há lançamentos selecionados, verifica se o tipo da conta permite baixa
            if not lancamentos_selecionados and conta_banco:
                if conta_banco.tipo not in tipos_permitidos_sem_lancamentos:
                    # Em vez de erro, mostrar modal com contas que permitem baixa direta
                    from extrato.models import ContaBancaria
                    contas_para_baixa_direta = ContaBancaria.objects.filter(
                        empresa_id=empresa_id,
                        status='A',
                        tipo__in=tipos_permitidos_sem_lancamentos
                    ).order_by('banco__nome')

                    context = {
                        'conta': conta,
                        'form': form,
                        'hoje': timezone.now().date(),
                        'data_inicio': data_inicio,
                        'data_fim': data_fim,
                        'mostrar_modal_contas': True,
                        'contas_para_baixa_direta': contas_para_baixa_direta,
                        'form_data': form.cleaned_data,  # Para repopular o form
                        **_context_valores_js_baixa(conta),
                    }
                    return render(request, 'contasareceber/baixar.html', context)

            # Calcula o valor total dos lançamentos selecionados (extrato)
            valor_lancamentos = sum(lancamento.valor for lancamento in lancamentos_selecionados)
            print(f"DEBUG: valor_lancamentos: {valor_lancamentos}")

            # Com lançamentos do extrato, o valor recebido é fixado na gravação pela soma das linhas
            # (VR = extrato − juros + desconto); não bloquear por divergência do campo digitado.
            resolucao = (form.cleaned_data.get('resolucao_diferenca_extrato') or 'igual').strip()

            # Validar conta bancária antes de processar
            if conta_banco:
                try:
                    # Verificar se a conta bancária existe e pertence à empresa
                    from extrato.models import ContaBancaria
                    ContaBancaria.objects.get(id=conta_banco.id, empresa_id=empresa_id)
                except ContaBancaria.DoesNotExist:
                    messages.error(request, f'Conta bancária selecionada não existe ou não pertence à empresa atual.')
                    return redirect('contasareceber:baixar', pk=pk)
                except Exception as e:
                    messages.error(request, f'Erro ao validar conta bancária: {str(e)}')
                    return redirect('contasareceber:baixar', pk=pk)

            # Processa a baixa normalmente (com ou sem conciliação)
            return processar_baixa_com_ajustes(
                conta,
                form,
                lancamentos_selecionados,
                request.user,
                request,
                filtros_query,
                resolucao_diferenca=resolucao,
            )
    hoje = timezone.now().date()

    if request.method == 'POST':
        # Obter parâmetros de período do POST
        data_inicio = request.POST.get('data_inicio')
        data_fim = request.POST.get('data_fim')
        form = BaixaContaIndividualForm(request.POST, empresa_id=empresa_id, conta=conta, data_inicio=data_inicio, data_fim=data_fim)
    else:
        # Obter parâmetros de período do GET (para caso de reload da página)
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')

        # Definir período padrão baseado na data de vencimento da conta
        if not data_inicio:
            if conta.data_vencimento:
                data_inicio = (conta.data_vencimento - timedelta(days=7)).strftime('%Y-%m-%d')
            else:
                data_inicio = hoje.strftime('%Y-%m-%d')
        if not data_fim:
            if conta.data_vencimento:
                data_fim = (conta.data_vencimento + timedelta(days=7)).strftime('%Y-%m-%d')
            else:
                data_fim = hoje.strftime('%Y-%m-%d')

        form = BaixaContaIndividualForm(empresa_id=empresa_id, conta=conta, data_inicio=data_inicio, data_fim=data_fim)

    context = {
        'conta': conta,
        'form': form,
        'hoje': hoje,
        'data_inicio': data_inicio or hoje.strftime('%Y-%m-%d'),
        'data_fim': data_fim or hoje.strftime('%Y-%m-%d'),
        'filtros_query': filtros_query,
        **_context_valores_js_baixa(conta),
    }

    return render(request, 'contasareceber/baixar.html', context)


def autocomplete_cliente(request):
    """Autocomplete para clientes"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'results': []})

    term = request.GET.get('term', '')
    clientes = Cliente.objects.filter(
        empresa_id=empresa_id
    ).filter(
        Q(razao__icontains=term) | Q(cnpj__icontains=term)
    )[:10]  # Limita a 10 resultados

    results = []
    for cliente in clientes:
        results.append({
            'id': cliente.id,
            'text': f"{cliente.razao} - {cliente.cnpj}"
        })

    return JsonResponse({'results': results})


def buscar_lancamentos_extrato(request):
    """AJAX: Busca lançamentos do extrato para uma conta bancária"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'error': 'Empresa não encontrada'}, status=400)

    conta_banco_id = request.GET.get('conta_banco_id')
    if not conta_banco_id:
        return JsonResponse({'error': 'Conta bancária não informada'}, status=400)

    try:
        from extrato.models import Lancamento

        # Buscar ContaBancaria para obter agencia e conta
        conta_banco = ContaBancaria.objects.get(id=conta_banco_id, empresa_id=empresa_id)

        # Verificar conta bancária

        # Apenas lançamentos ainda não conciliados (exclude cobre NULL no MySQL).
        # Contas a receber: só créditos (valor > 0); débitos não servem à conciliação de recebimento.
        lancamentos_query = Lancamento.objects.filter(
            empresa_id=empresa_id,
            conta_id=conta_banco.id,
            valor__gt=0,
        ).exclude(conciliado=True).filter(idconciliacao__isnull=True)

        # Aplicar filtros de período se fornecidos
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')

        if data_inicio:
            lancamentos_query = lancamentos_query.filter(data__gte=data_inicio)
        if data_fim:
            lancamentos_query = lancamentos_query.filter(data__lte=data_fim)

        # Paginação
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 20))
        total = lancamentos_query.count()
        start = (page - 1) * limit
        end = start + limit

        # Ordenar por data decrescente e paginar
        lancamentos = lancamentos_query.order_by('-data')[start:end]

        # Processar lançamentos

        data = []
        for lancamento in lancamentos:
            data.append({
                'id': lancamento.id,
                'data': lancamento.data.strftime('%d/%m/%Y'),
                'valor': float(lancamento.valor),
                'historico': lancamento.historico,
                'documento': lancamento.documento or '',
                'fitid': lancamento.fitid or '',
                'conciliado': lancamento.conciliado
            })

        return JsonResponse({
            'lancamentos': data,
            'total': total,
            'page': page,
            'limit': limit
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def buscar_lancamentos_selecionados(request):
    """AJAX: Busca detalhes dos lançamentos selecionados para exibir na conferência"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'error': 'Empresa não encontrada'}, status=400)

    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)

    try:
        import json
        data = json.loads(request.body.decode('utf-8'))
        lancamentos_ids = data.get('lancamentos_ids', [])

        if not lancamentos_ids:
            return JsonResponse({'lancamentos': []})

        from extrato.models import Lancamento

        # Buscar lançamentos pelos IDs (apenas não conciliados)
        lancamentos = Lancamento.objects.filter(
            id__in=lancamentos_ids,
            empresa_id=empresa_id,
        ).exclude(conciliado=True).filter(idconciliacao__isnull=True).order_by('data')

        data_response = []
        for lancamento in lancamentos:
            data_response.append({
                'id': lancamento.id,
                'data': lancamento.data.strftime('%d/%m/%Y'),
                'valor': float(lancamento.valor),
                'historico': lancamento.historico,
                'documento': lancamento.documento or '',
                'fitid': lancamento.fitid or ''
            })

        return JsonResponse({'lancamentos': data_response})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def baixar_contas_a_receber(request):
    """Escolhe uma conta a receber e redireciona para a baixa com extrato (vários lançamentos)."""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    if request.method == 'POST':
        form = EscolhaContaBaixaForm(request.POST, empresa_id=empresa_id)
        if form.is_valid():
            conta = form.cleaned_data['conta']
            return redirect('contasareceber:baixar', pk=conta.pk)
    else:
        form = EscolhaContaBaixaForm(empresa_id=empresa_id)

    return render(
        request,
        'contasareceber/baixar_multiplas.html',
        {'form': form},
    )


def editar_conta_a_receber(request, pk):
    """Edita uma conta a receber existente"""
    empresa_id = request.session.get('empresa_id')

    if not empresa_id:
        return _redirect_sem_empresa(request)

    conta = get_object_or_404(ContaAReceber, pk=pk, empresa_id=empresa_id)

    # Verificar se a conta está paga
    if conta.status == 'pago':
        messages.warning(request, 'Esta conta já foi totalmente paga. Para editar, primeiro estorne o pagamento.')
        return redirect('contasareceber:detalhes', pk=pk)

    if request.method == 'POST':
        try:
            # Obter dados do formulário
            cliente_nome = request.POST.get('cliente')
            cnpj_cpf = request.POST.get('cnpj_cpf')
            valor_str = request.POST.get('valor')
            # Convert Brazilian decimal format (comma) to Python decimal format (dot)
            if valor_str:
                valor = float(valor_str.replace(',', '.'))
            else:
                valor = 0
            data_emissao = request.POST.get('data_emissao')
            data_vencimento = request.POST.get('data_vencimento')
            numero_nota = request.POST.get('nota')
            doc = request.POST.get('doc')
            autorizacao = request.POST.get('autorizacao')
            forma_pagamento_id = request.POST.get('forma_pagamento')
            conta_bancaria_id = request.POST.get('conta_bancaria')
            raw_socio = (request.POST.get('socio') or '').strip()
            categoria_id = request.POST.get('categoria')

            # Criar ou obter nota fiscal apenas se numero_nota for fornecido
            nota = None
            if numero_nota:
                nota, created = NotaFiscalServico.objects.get_or_create(
                    empresa=conta.empresa,
                    numero_nota=numero_nota,
                    defaults={
                        'cliente': cliente_nome or 'Cliente não informado',
                        'valor_bruto': float(str(valor).replace(',', '.')) if valor else 0,
                        'valor_liquido': float(str(valor).replace(',', '.')) if valor else 0,
                        'data_emissao': data_emissao or timezone.now().date(),
                    }
                )

            # DEBUG: Log valores atuais antes da atualização da conta
            data_emissao_nota = nota.data_emissao if nota else (data_emissao or timezone.now().date())
            forma_pagamento_desc_edit = None
            if forma_pagamento_id:
                try:
                    forma_pgto_obj_edit = Cobranca.objects.get(id=forma_pagamento_id)
                    forma_pagamento_desc_edit = forma_pgto_obj_edit.descricao
                except Cobranca.DoesNotExist:
                    forma_pagamento_desc_edit = "FORMA_PAGAMENTO_NAO_ENCONTRADA"

            print(f"DEBUG EDITAR_CONTA_RECEBER: data_emissao_nota={data_emissao_nota}, data_vencimento_atual={data_vencimento}, forma_pagamento_id={forma_pagamento_id}, forma_pagamento_desc={forma_pagamento_desc_edit}")

            # Atualizar conta a receber
            conta.nota = nota
            conta.cliente = cliente_nome or (nota.cliente if nota else 'Cliente não informado')
            conta.cnpj_cpf = cnpj_cpf or (nota.cnpj_cpf if nota else None)
            conta.data_emissao = data_emissao or timezone.now().date()
            conta.data_vencimento = data_vencimento
            conta.valor_a_receber = float(str(valor).replace(',', '.'))
            conta.doc = doc
            conta.autorizacao = autorizacao
            conta.forma_pagamento_id = forma_pagamento_id if forma_pagamento_id else None

            socio_obj = None
            if raw_socio.isdigit():
                try:
                    socio_obj = Socio.objects.get(id=int(raw_socio), empresa_id=empresa_id)
                except Socio.DoesNotExist:
                    pass

            conta.socio = socio_obj
            if nota:
                nota.socio = socio_obj
                nota.save(update_fields=['socio'])

            if categoria_id:
                try:
                    categoria = Categoria.objects.get(id=categoria_id, empresa_id=empresa_id)
                    conta.categoria = categoria
                except Categoria.DoesNotExist:
                    pass

            # Associar conta bancária se fornecida
            # Temporariamente desabilitado para evitar erro de foreign key
            # if conta_bancaria_id:
            #     try:
            #         conta_banco_selecionado = ContaBancaria.objects.get(id=conta_bancaria_id, empresa=conta.empresa)
            #         # Buscar ContaBancaria correspondente baseado em agencia e conta
            #         # from extrato.models import ContaBancaria
            #         conta_bancaria_correspondente = ContaBancaria.objects.filter(
            #             empresa=conta.empresa,
            #             agencia=conta_banco_selecionado.agencia,
            #             conta=conta_banco_selecionado.conta
            #         ).first()
            #         if conta_bancaria_correspondente:
            #             conta.conta_banco = conta_bancaria_correspondente
            #     except ContaBancaria.DoesNotExist:
            #         pass

            conta.save()

            messages.success(request, 'Conta a receber atualizada com sucesso!')
            filtros = extrair_filtros_contas_receber(request)
            return redirect(construir_url_crlistar_com_filtros(filtros))

        except Exception as e:
            messages.error(request, f'Erro ao atualizar conta: {str(e)}')
            return redirect('contasareceber:editar', pk=pk)

    # Dados para o template
    clientes = Cliente.objects.filter(empresa_id=empresa_id) if empresa_id else Cliente.objects.none()
    contas_bancarias = ContaBancaria.objects.filter(empresa_id=empresa_id, status='A') if empresa_id else ContaBancaria.objects.none()
    formas_pagamento = Cobranca.objects.all()
    socios = Socio.objects.filter(empresa_id=empresa_id) if empresa_id else Socio.objects.none()
    categorias = Categoria.objects.filter(empresa_id=empresa_id) if empresa_id else Categoria.objects.none()
    hoje = timezone.now().date()

    # Encontrar ContaBancaria correspondente ao ContaBancaria da conta
    conta_banco_selecionado_id = None
    if conta.conta_banco:
        conta_banco_correspondente = ContaBancaria.objects.filter(
            empresa_id=empresa_id,
            agencia=conta.conta_banco.agencia,
            conta=conta.conta_banco.conta
        ).first()
        if conta_banco_correspondente:
            conta_banco_selecionado_id = conta_banco_correspondente.id

    context = {
        'conta': conta,
        'clientes': clientes,
        'contas_bancarias': contas_bancarias,
        'formas_pagamento': formas_pagamento,
        'socios': socios,
        'categorias': categorias,
        'hoje': hoje,
        'conta_banco_selecionado_id': conta_banco_selecionado_id,
    }

    return render(request, 'contasareceber/editar.html', context)


def estornar_conta_a_receber(request, pk):
    """Permite estornar (reverter) pagamentos de uma conta a receber"""
    empresa_id = request.session.get('empresa_id')

    if not empresa_id:
        return _redirect_sem_empresa(request)

    conta = get_object_or_404(ContaAReceber, pk=pk, empresa_id=empresa_id)

    # Verificar se a conta tem pagamentos para estornar
    baixas = (
        BaixaContaAReceber.objects.filter(conta_a_receber=conta)
        .prefetch_related('extrato_movimentos')
        .order_by('-data_recebimento')
    )

    if not baixas.exists():
        messages.warning(request, 'Esta conta não possui pagamentos para estornar.')
        return redirect('contasareceber:detalhes', pk=pk)

    if request.method == 'POST':
        baixa_id = request.POST.get('baixa_id')
        if baixa_id:
            try:
                baixa = BaixaContaAReceber.objects.get(id=baixa_id, conta_a_receber=conta)

                valor_revertido = baixa.get_valor_credito_extrato()

                # Remover movimento do extrato relacionado a esta baixa
                from extrato.models import ExtratoMovimento
                ExtratoMovimento.objects.filter(
                    conta_receber=conta,
                    baixa_receber=baixa
                ).delete()

                # Excluir baixa; BaixaContaAReceber.delete() recalcula totais na conta (sem soma duplicada)
                baixa.delete()
                conta.refresh_from_db()

                # Marcar nota fiscal com status apropriado
                if conta.nota:
                    if not conta.valor_recebido or conta.valor_recebido <= 0:
                        conta.nota.status_conciliacao = 'nao_conciliado'
                    else:
                        conta.nota.status_conciliacao = 'parcialmente_conciliado'
                    conta.nota.save()

                messages.success(request, f'Pagamento estornado com sucesso! Valor revertido: R$ {valor_revertido}')
                return redirect('contasareceber:detalhes', pk=pk)

            except BaixaContaAReceber.DoesNotExist:
                messages.error(request, 'Baixa não encontrada.')
        else:
            messages.error(request, 'Selecione um pagamento para estornar.')

    context = {
        'conta': conta,
        'baixas': baixas,
    }

    return render(request, 'contasareceber/estornar.html', context)


def listar_baixas(request):
    """Lista todas as baixas de contas a receber"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    baixas = (
        BaixaContaAReceber.objects.filter(empresa_id=empresa_id)
        .prefetch_related('extrato_movimentos')
        .order_by('-data_recebimento')
    )

    # Filtros
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    conta_banco = request.GET.get('conta_banco')

    if data_inicio:
        baixas = baixas.filter(data_recebimento__gte=data_inicio)
    if data_fim:
        baixas = baixas.filter(data_recebimento__lte=data_fim)
    if conta_banco:
        baixas = baixas.filter(conta_banco_id=conta_banco)

    # Estatísticas (coluna alinhada ao extrato quando houver movimentos vinculados)
    total_recebido = sum((b.get_valor_credito_extrato() for b in baixas), Decimal('0'))
    total_descontos = sum(baixa.desconto for baixa in baixas)
    total_juros = sum(baixa.juros for baixa in baixas)
    total_tarifas = sum(baixa.tarifas for baixa in baixas)

    context = {
        'baixas': baixas,
        'total_recebido': total_recebido,
        'total_descontos': total_descontos,
        'total_juros': total_juros,
        'total_tarifas': total_tarifas,
        'filtros': {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'conta_banco': conta_banco,
        }
    }

    return render(request, 'contasareceber/listar_baixas.html', context)

def extrair_nome_pagador(observacao):
    match = re.search(r'PAGADOR:\s+([A-Za-zÀ-ÿ\s]+)', observacao, re.IGNORECASE)
    
    if match:
        nome = match.group(1).strip()
        # Remover caracteres indesejados no final, como apóstrofo ou aspas
        nome = re.sub(r"[\"']+$", "", nome)
        print(f"DEBUG EXTRACAO NOME PAGADOR: observacao='{nome}', match={match}")
        return nome
    return None


def _prefixo_razao_conciliacao(conta, n=15):
    """Primeiros n caracteres do nome/razão do tomador (conta a receber ou NF vinculada)."""
    texto = (getattr(conta, 'cliente', None) or '').strip()
    if not texto:
        try:
            if getattr(conta, 'nota_id', None) and conta.nota:
                texto = (conta.nota.cliente or '').strip()
        except Exception:
            texto = ''
    if not texto:
        return ''
    return texto[:n]


def encontrar_lancamento_aproximado(conta_receber, lancamentos_extrato, valor_conta):
    cliente = conta_receber.cliente.upper()
    
    observacao = getattr(conta_receber, 'observacao', '') or ""
    data_vencimento = conta_receber.data_vencimento

    # Filtrar lançamentos por data (±30 dias)
    lancamentos_filtrados = [
        l for l in lancamentos_extrato
        if abs((l.data - data_vencimento).days) <= 30
    ]

    # 1. Comparar cliente com histórico
 
    for lanc in lancamentos_filtrados:
        if cliente.upper() in lanc.historico.upper() and abs(lanc.valor - valor_conta) <= 0.05:
          print(f"DEBUG ENCONTRAR LANÇAMENTO: Lançamento encontrado: {lanc}")
          return lanc

    # 2. Extrair nome do pagador da observação e comparar
        nome_pagador = extrair_nome_pagador(observacao)
        if nome_pagador:
          for lanc in lancamentos_filtrados:
            if nome_pagador.upper() in lanc.historico.upper() and abs(lanc.valor - valor_conta) <= 0.05:
              print(f"DEBUG ENCONTRAR LANÇAMENTO: Lançamento encontrado: {lanc}")
            return lanc

    return None

def conciliar_contas_a_receber(request):
    """Conciliar múltiplas contas a receber com lançamentos bancários aplicando regras específicas"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    if request.method == 'POST':
        contas_ids = request.POST.getlist('contas')
        if not contas_ids:
            messages.error(request, 'Selecione pelo menos uma conta a receber.')
            filtros = extrair_filtros_contas_receber(request)
            return redirect(construir_url_crlistar_com_filtros(filtros))

        contas = ContaAReceber.objects.filter(
            id__in=contas_ids, empresa_id=empresa_id, status='pendente'
        ).select_related('nota')
        conciliadas = 0
        erros_conciliacao = []

        for conta in contas:
            try:
                if _conciliar_uma_conta(request, conta, empresa_id):
                    conciliadas += 1
            except Exception as e:
                import traceback
                if settings.DEBUG:
                    traceback.print_exc()
                erros_conciliacao.append(f"Conta {conta.id} ({getattr(conta, 'cliente', '')}): {str(e)}")
                continue

        if erros_conciliacao:
            messages.error(
                request,
                "Conciliação concluída com erros: " + "; ".join(erros_conciliacao[:5])
                + (" ..." if len(erros_conciliacao) > 5 else "")
            )
        if conciliadas:
            messages.success(request, f'{conciliadas} conta(s) conciliada(s) automaticamente.')
        elif not erros_conciliacao:
            messages.info(request, 'Nenhuma conta foi conciliada (nenhuma correspondência encontrada).')

        filtros = extrair_filtros_contas_receber(request)
        return redirect(construir_url_crlistar_com_filtros(filtros))

    # GET: mostrar contas pendentes para seleção
    contas_pendentes = ContaAReceber.objects.filter(
        empresa_id=empresa_id,
        status='pendente'
    ).order_by('data_vencimento')

    context = {
        'contas_pendentes': contas_pendentes,
    }

    return render(request, 'contasareceber/conciliar.html', context)


def _conciliar_uma_conta(request, conta, empresa_id):
    """Processa conciliação de uma conta a receber. Retorna True se conciliou, False se sem correspondência. Levanta exceção em caso de erro."""
    if settings.DEBUG:
        print(f"DEBUG CONCILIAR: Processando conta {conta.id} - Cliente: {conta.cliente}, CNPJ_CPF: {conta.cnpj_cpf}")
    # Regra 1: Pegar CPF/CNPJ, data de vencimento e valor
    cnpj_cpf = conta.cnpj_cpf
    nome_cliente = conta.cliente
    data_vencimento = conta.data_vencimento
    valor_conta = conta.get_valor_pendente()
    if settings.DEBUG:
        print(f"DEBUG CONCILIAR: Data vencimento: {data_vencimento}, Valor pendente: {valor_conta}")

    # Verificar se data_vencimento existe
    if not data_vencimento:
        if settings.DEBUG:
            print(f"DEBUG CONCILIAR: Conta {conta.id} não tem data de vencimento, pulando...")
        return False  # não é erro, apenas sem data para buscar

    # Filtrar lançamentos bancários dentro do mês da data de vencimento
    from extrato.models import Lancamento, Conciliacao, ExtratoMovimento
    from contasareceber.models import BaixaContaAReceber
    from datetime import timedelta

    # Mês de referência da data de vencimento
    mes_referencia = data_vencimento.replace(day=1)
    # Calcular mês anterior e próximo mês
    mes_anterior = (mes_referencia - relativedelta(months=1))
    proximo_mes = (mes_referencia + relativedelta(months=1))
    if settings.DEBUG:
        print(f"DEBUG CONCILIAR: Buscando lançamentos entre {mes_anterior} e {proximo_mes}")

    # Buscar lançamentos do mês anterior até o próximo mês
    lancamentos = Lancamento.objects.filter(
        empresa_id=empresa_id,
        data__gte=mes_anterior,
        data__lt=proximo_mes,
        conciliado=False
    )
    if settings.DEBUG:
        print(f"DEBUG CONCILIAR: Encontrados {lancamentos.count()} lançamentos no mês")

    # Verificar se encontra correspondência
    lancamento_encontrado = None

    # Regra 1: Buscar por CPF/CNPJ no histórico
    if cnpj_cpf:
        if settings.DEBUG:
            print(f"DEBUG CONCILIAR: Buscando por CPF/CNPJ: '{cnpj_cpf}' (tipo: {type(cnpj_cpf)}, len: {len(cnpj_cpf) if cnpj_cpf else 0})")
        # Verificar se cpf_cnpj é válido
        if not isinstance(cnpj_cpf, str):
            if settings.DEBUG:
                print(f"DEBUG CONCILIAR: AVISO - cnpj_cpf não é string: {cnpj_cpf}")
            cnpj_cpf = str(cnpj_cpf) if cnpj_cpf is not None else ""
        if not cnpj_cpf.strip():
            if settings.DEBUG:
                print(f"DEBUG CONCILIAR: AVISO - cnpj_cpf está vazio após strip")
            return False
        for lancamento in lancamentos:
            # Logs detalhados para identificar a causa do OSError
            try:
                lancamento_id = lancamento.id
                if settings.DEBUG:
                    print(f"DEBUG CONCILIAR: Verificando lançamento ID: {lancamento_id}")
            except Exception as e:
                if settings.DEBUG:
                    print(f"DEBUG CONCILIAR: ERRO ao acessar lancamento.id: {str(e)}")
                continue

            try:
                banco_codigo = lancamento.conta.banco.codigo if lancamento.conta and lancamento.conta.banco else None
                if settings.DEBUG:
                    print(f"DEBUG CONCILIAR: Banco código: {banco_codigo}")
            except Exception as e:
                if settings.DEBUG:
                    print(f"DEBUG CONCILIAR: ERRO ao acessar lancamento.conta.banco.codigo: {str(e)}")
                banco_codigo = None

            try:
                historico = lancamento.historico
                if settings.DEBUG:
                    print(f"DEBUG CONCILIAR: Histórico: '{historico}'")
            except Exception as e:
                if settings.DEBUG:
                    print(f"DEBUG CONCILIAR: ERRO ao acessar lancamento.historico: {str(e)}")
                historico = ""

            try:
                valor = lancamento.valor
                if settings.DEBUG:
                    print(f"DEBUG CONCILIAR: Valor: {valor}")
            except Exception as e:
                if settings.DEBUG:
                    print(f"DEBUG CONCILIAR: ERRO ao acessar lancamento.valor: {str(e)}")
                continue

            # Formatar CPF/CNPJ conforme banco
            cpf_cnpj_formatado = cnpj_cpf
            if banco_codigo == '001':  # Banco do Brasil
                if len(cnpj_cpf) == 11:  # CPF
                    cpf_cnpj_formatado = cnpj_cpf
                elif len(cnpj_cpf) == 14:  # CNPJ
                    cpf_cnpj_formatado = cnpj_cpf
                else:
                    if settings.DEBUG:
                        print(f"DEBUG CONCILIAR: CPF/CNPJ inválido para BB: {cnpj_cpf}")
                    continue
            elif banco_codigo == '756':  # SICOOB
                if len(cnpj_cpf) == 11:  # CPF
                    cpf_cnpj_formatado = f"***.{cnpj_cpf[3:6]}.{cnpj_cpf[6:9]}-**"
                elif len(cnpj_cpf) == 14:  # CNPJ
                    cpf_cnpj_formatado = f"{cnpj_cpf[:2]}.{cnpj_cpf[2:5]}.{cnpj_cpf[5:8]} {cnpj_cpf[8:12]}-{cnpj_cpf[12:]}"
                else:
                    if settings.DEBUG:
                        print(f"DEBUG CONCILIAR: CPF/CNPJ inválido para SICOOB: {cnpj_cpf}")
                    continue

            # Verificar se CPF/CNPJ está no histórico
            try:
                historico_lancamento = lancamento.historico or ""
                if settings.DEBUG:
                    print(f"DEBUG CONCILIAR: Verificando se '{cpf_cnpj_formatado}' está em '{historico_lancamento}'")
                if cpf_cnpj_formatado in historico_lancamento:
                    if settings.DEBUG:
                        print(f"DEBUG CONCILIAR: CPF/CNPJ encontrado no histórico!")
                    # Verificar valor com tolerância de 0.05
                    valor_lancamento = lancamento.valor
                    if abs(valor_lancamento - valor_conta) <= 0.05:
                        if settings.DEBUG:
                            print(f"DEBUG CONCILIAR: Valor compatível! Lançamento: {valor_lancamento}, Conta: {valor_conta}")
                        lancamento_encontrado = lancamento
                        break
                    elif settings.DEBUG:
                        print(f"DEBUG CONCILIAR: Valor não compatível. Lançamento: {valor_lancamento}, Conta: {valor_conta}")
                elif settings.DEBUG:
                    print(f"DEBUG CONCILIAR: CPF/CNPJ não encontrado no histórico")
            except Exception as e:
                if settings.DEBUG:
                    print(f"DEBUG CONCILIAR: ERRO ao verificar CPF/CNPJ no histórico: {str(e)}")
                continue

    # Regra 1b: CNPJ no extrato não fechou a conciliação — primeiros 15 caracteres da razão no histórico,
    # mesmo valor (tolerância R$ 0,05), data do lançamento entre vencimento −10 e +10 dias.
    if not lancamento_encontrado and cnpj_cpf and str(cnpj_cpf).strip():
        prefixo = _prefixo_razao_conciliacao(conta, 15)
        if len(prefixo) >= 3:
            d_min = data_vencimento - timedelta(days=10)
            d_max = data_vencimento + timedelta(days=10)
            lancamentos_janela = Lancamento.objects.filter(
                empresa_id=empresa_id,
                data__gte=d_min,
                data__lte=d_max,
                conciliado=False,
            )
            pref_u = prefixo.upper()
            if settings.DEBUG:
                print(
                    f"DEBUG CONCILIAR: Fallback razão (15 chars) após CNPJ sem match: "
                    f"'{prefixo}' | janela {d_min} a {d_max}"
                )
            for lancamento in lancamentos_janela:
                try:
                    historico_u = (lancamento.historico or '').upper()
                    if pref_u not in historico_u:
                        continue
                    valor_lancamento = lancamento.valor
                    if abs(valor_lancamento - valor_conta) <= 0.05:
                        lancamento_encontrado = lancamento
                        if settings.DEBUG:
                            print(
                                f"DEBUG CONCILIAR: Lançamento {lancamento.id} por prefixo razão + valor + ±10 dias."
                            )
                        break
                except Exception as e:
                    if settings.DEBUG:
                        print(
                            f"DEBUG CONCILIAR: ERRO fallback razão no lançamento {getattr(lancamento, 'id', '?')}: {e}"
                        )
                    continue

    # Regra 1c: cadastro Cliente com o mesmo CNPJ/CPF — texto do extrato no cadastro contido no histórico;
    # valor ±0,05; data do lançamento entre vencimento −10 e +10 dias.
    if not lancamento_encontrado and cnpj_cpf and str(cnpj_cpf).strip():
        from fornecedor.cnpj_utils import limpar_cnpj
        from cliente.models import Cliente

        cnpj_limpo = limpar_cnpj(str(cnpj_cpf))
        if len(cnpj_limpo) in (11, 14):
            cli = Cliente.objects.filter(empresa_id=empresa_id, cnpj=cnpj_limpo).first()
            trecho = (cli.descricao_extrato_bancario or '').strip() if cli else ''
            if trecho and len(trecho) >= 2:
                d_min = data_vencimento - timedelta(days=10)
                d_max = data_vencimento + timedelta(days=10)
                lancamentos_janela_cli = Lancamento.objects.filter(
                    empresa_id=empresa_id,
                    data__gte=d_min,
                    data__lte=d_max,
                    conciliado=False,
                )
                trecho_u = trecho.upper()
                if settings.DEBUG:
                    print(
                        f"DEBUG CONCILIAR: Fallback cadastro Cliente (CNPJ {cnpj_limpo}) texto extrato: "
                        f"'{trecho}' | janela {d_min} a {d_max}"
                    )
                for lancamento in lancamentos_janela_cli:
                    try:
                        historico_u = (lancamento.historico or '').upper()
                        if trecho_u not in historico_u:
                            continue
                        valor_lancamento = lancamento.valor
                        if abs(valor_lancamento - valor_conta) <= 0.05:
                            lancamento_encontrado = lancamento
                            if settings.DEBUG:
                                print(
                                    f"DEBUG CONCILIAR: Lançamento {lancamento.id} por Cliente.descricao_extrato_bancario."
                                )
                            break
                    except Exception as e:
                        if settings.DEBUG:
                            print(
                                f"DEBUG CONCILIAR: ERRO fallback Cliente extrato no lançamento "
                                f"{getattr(lancamento, 'id', '?')}: {e}"
                            )
                        continue

    # Regra 2: Se não encontrou por CPF/CNPJ, buscar pelo nome do cliente
    if not lancamento_encontrado and nome_cliente:
        if settings.DEBUG:
            print(f"DEBUG CONCILIAR: Buscando por nome do cliente: {nome_cliente}")
        for lancamento in lancamentos:
            try:
                historico = lancamento.historico or ""
                if nome_cliente.upper() in historico.upper():
                    if settings.DEBUG:
                        print(f"DEBUG CONCILIAR: Nome do cliente '{nome_cliente}' encontrado no histórico!")
                    # Verificar valor com tolerância de 0.05
                    valor_lancamento = lancamento.valor
                    if abs(valor_lancamento - valor_conta) <= 0.05:
                        lancamento_encontrado = lancamento
                        if settings.DEBUG:
                            print(f"DEBUG CONCILIAR: Lançamento encontrado por nome do cliente!")
                        break
                    elif settings.DEBUG:
                        print(f"DEBUG CONCILIAR: Valor não compatível - Lançamento: {valor_lancamento}, Conta: {valor_conta}")
            except Exception as e:
                if settings.DEBUG:
                    print(f"DEBUG CONCILIAR: ERRO ao verificar nome do cliente no lançamento {lancamento.id}: {str(e)}")
                continue

        if not lancamento_encontrado:
            lancamento_aproximado = encontrar_lancamento_aproximado(conta, lancamentos, valor_conta)
            if lancamento_aproximado:
                lancamento_encontrado = lancamento_aproximado

    # Regra 3: Se não encontrou por nome do cliente, buscar pela máquina ou bandeira (case insensitive)
    if not lancamento_encontrado:
        # Pegar a máquina e bandeira da conta através do relatório de recebíveis
        maquina = None
        bandeira = None
        from relatoriorecebiveis.models import RelatorioRecebiveisMaquinaCartao
        from .cartao_aproximacao import buscar_relatorios_por_autorizacao
        relatorio_cartao = buscar_relatorios_por_autorizacao(
            empresa_id, conta.autorizacao, conciliado=None
        ).first()
        if relatorio_cartao:
            maquina = relatorio_cartao.maquinha
            bandeira = relatorio_cartao.bandeira

        # Primeiro tentar pela máquina
        if maquina:
            if settings.DEBUG:
                print(f"DEBUG CONCILIAR: Buscando por máquina: {maquina}")
            for lancamento in lancamentos:
                try:
                    historico = lancamento.historico or ""
                    if maquina.lower() in historico.lower():
                        if settings.DEBUG:
                            print(f"DEBUG CONCILIAR: Máquina '{maquina}' encontrada no histórico!")
                        # Verificar valor com tolerância de 0.05
                        valor_lancamento = lancamento.valor
                        if abs(valor_lancamento - valor_conta) <= 0.05:
                            lancamento_encontrado = lancamento
                            if settings.DEBUG:
                                print(f"DEBUG CONCILIAR: Lançamento encontrado por máquina!")
                            break
                        elif settings.DEBUG:
                            print(f"DEBUG CONCILIAR: Valor não compatível - Lançamento: {valor_lancamento}, Conta: {valor_conta}")
                except Exception as e:
                    if settings.DEBUG:
                        print(f"DEBUG CONCILIAR: ERRO ao verificar máquina no lançamento {lancamento.id}: {str(e)}")
                    continue

        # Se não encontrou pela máquina, tentar pela bandeira
        if not lancamento_encontrado and bandeira:
            if settings.DEBUG:
                print(f"DEBUG CONCILIAR: Buscando por bandeira: {bandeira}")
            for lancamento in lancamentos:
                try:
                    historico = lancamento.historico or ""
                    if bandeira.lower() in historico.lower():
                        if settings.DEBUG:
                            print(f"DEBUG CONCILIAR: Bandeira '{bandeira}' encontrada no histórico!")
                        # Verificar valor com tolerância de 0.05
                        valor_lancamento = lancamento.valor
                        if abs(valor_lancamento - valor_conta) <= 0.05:
                            lancamento_encontrado = lancamento
                            if settings.DEBUG:
                                print(f"DEBUG CONCILIAR: Lançamento encontrado por bandeira!")
                            break
                        elif settings.DEBUG:
                            print(f"DEBUG CONCILIAR: Valor não compatível - Lançamento: {valor_lancamento}, Conta: {valor_conta}")
                except Exception as e:
                    if settings.DEBUG:
                        print(f"DEBUG CONCILIAR: ERRO ao verificar bandeira no lançamento {lancamento.id}: {str(e)}")
                    continue

    if lancamento_encontrado:
        if settings.DEBUG:
            print(f"DEBUG CONCILIAR: Correspondência encontrada para conta {conta.id} com lançamento {lancamento_encontrado.id}")
        # Encontrou correspondência - aplicar regras de conciliação

        # Criar idconciliacao
        conciliacao = Conciliacao.objects.create(
            criado_por=request.user if request.user.is_authenticated else None,
            observacao=f'Conciliação automática - Conta {conta.id}'
        )

        # Atualizar lançamento bancário
        lancamento_encontrado.conciliado = True
        lancamento_encontrado.idconciliacao = conciliacao
        lancamento_encontrado.save()

        # Encontrar ContaBancaria correspondente
        conta_bancoB = None
        if not getattr(lancamento_encontrado, 'conta', None):
            if settings.DEBUG:
                print(f"DEBUG CONCILIAR: Lançamento {lancamento_encontrado.id} sem conta bancária, pulando conta {conta.id}")
            return False
        try:
            if settings.DEBUG:
                print(f"DEBUG CONCILIAR: Tentando encontrar ContaBancaria para lancamento {lancamento_encontrado.id}")
            conta_bancoB = ContaBancaria.objects.get(
                empresa=conta.empresa,
                id=lancamento_encontrado.conta.id
            )
        except ContaBancaria.DoesNotExist:
            if settings.DEBUG:
                print(f"DEBUG CONCILIAR: ContaBancaria não encontrada para empresa {conta.empresa.id} e conta {lancamento_encontrado.conta.id}")
            conta_bancoB = None
        except Exception as e:
            if settings.DEBUG:
                print(f"DEBUG CONCILIAR: ERRO ao buscar ContaBancaria: {str(e)}")
            conta_bancoB = None

        # Verificar se conta_banco é válido
        if conta_bancoB:
            try:
                ContaBancaria.objects.get(id=conta_bancoB.id, empresa=conta.empresa)
            except ContaBancaria.DoesNotExist:
                conta_bancoB = None
            except Exception as e:
                if settings.DEBUG:
                    print(f"DEBUG CONCILIAR: ERRO ao validar ContaBancaria: {str(e)}")
                conta_bancoB = None

        # Criar baixa para conta a receber
        baixa = BaixaContaAReceber.objects.create(
            conta_a_receber=conta,
            empresa=conta.empresa,
            data_recebimento=lancamento_encontrado.data,
            valor_recebido=lancamento_encontrado.valor,
            conta_banco=conta_bancoB if conta_bancoB else None,
            tipo_baixa='total'
        )

        # BaixaContaAReceber.save() já agregou valor_recebido, juros, tarifas, desconto e status na conta.
        # Só persistir conta bancária na conta (não sobrescrever totais nem forçar status).
        try:
            if lancamento_encontrado and hasattr(lancamento_encontrado, 'data') and lancamento_encontrado.data:
                conta.refresh_from_db()
                conta.conta_banco = conta_bancoB
                conta.save(update_fields=['conta_banco'])
            else:
                raise ValueError("lançamento encontrado inválido (sem data)")
        except AttributeError as e:
            raise AttributeError(f"ao atualizar conta {conta.id}: {str(e)}")

        # Atualizar nota fiscal se existir
        if conta.nota:
            conta.nota.status_conciliacao = 'conciliado'
            conta.nota.save()

        # Criar ExtratoMovimento com dados do relatório de recebíveis
        from relatoriorecebiveis.models import RelatorioRecebiveisMaquinaCartao
        from .cartao_aproximacao import buscar_relatorios_por_autorizacao

        relatorio = None
        if conta.autorizacao:
            relatorio = buscar_relatorios_por_autorizacao(
                empresa_id, conta.autorizacao, conciliado=None
            ).first()

        if relatorio:
            relatorio.conta_a_receber = conta
            relatorio.save()

        # Construir descrição conforme especificação (evitar None para não quebrar em produção)
        if relatorio:
            nota_fiscal = (relatorio.nota_fiscal or "Sem Nota")
            razao = (relatorio.razao or conta.cliente or "")
        else:
            nota_fiscal = (conta.nota.numero_nota if conta.nota and getattr(conta.nota, 'numero_nota', None) else "Sem Nota")
            razao = (conta.cliente or "")
        parcela_info = (conta.parcela or "1/1")
        if relatorio and relatorio.parcelas and relatorio.total_parcelas:
            parcela_info = f"{relatorio.parcelas}/{relatorio.total_parcelas}"

        descricao_movimento = f"{lancamento_encontrado.historico} - {nota_fiscal} {parcela_info} - {razao}"

        valor_liquido = getattr(relatorio, 'valor_liquido', None) if relatorio else None
        valor_movimento = valor_liquido if valor_liquido is not None else lancamento_encontrado.valor

        ExtratoMovimento.objects.create(
            empresa=conta.empresa,
            data_baixa=lancamento_encontrado.data,
            descricao=descricao_movimento,
            situacao='recebido',
            valor=valor_movimento,
            conta_receber=conta,
            baixa_receber=baixa,
            lancamento=lancamento_encontrado,
            conta_banco=conta_bancoB,
            categoria=conta.categoria
        )
        return True  # conciliado com sucesso

    # Nenhum lançamento bancário correspondente encontrado para esta conta
    if settings.DEBUG:
        print(f"DEBUG CONCILIAR: Nenhuma correspondência para conta {conta.id}; baixa não criada.")
    return False  # não é erro, apenas sem correspondência


def nao_conciliados(request):
    """Exibe relatório de itens não conciliados: contas a receber, cartão e extrato bancário"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    # Filtros de período
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    # Contas a receber não conciliadas (status pendente ou vencido)
    contas_nao_conciliadas = ContaAReceber.objects.filter(
        empresa_id=empresa_id,
        status__in=['pendente', 'vencido']
    )
    if data_inicio:
        contas_nao_conciliadas = contas_nao_conciliadas.filter(data_vencimento__gte=data_inicio)
    if data_fim:
        contas_nao_conciliadas = contas_nao_conciliadas.filter(data_vencimento__lte=data_fim)
    contas_nao_conciliadas = contas_nao_conciliadas.order_by('data_vencimento')[:50]

    # Relatório de cartão não conciliado (sem conta_a_receber vinculada)
    from relatoriorecebiveis.models import RelatorioRecebiveisMaquinaCartao
    cartao_nao_conciliado = RelatorioRecebiveisMaquinaCartao.objects.filter(
        empresa_id=empresa_id,
        conta_a_receber__isnull=True
    )
    if data_inicio:
        cartao_nao_conciliado = cartao_nao_conciliado.filter(data_pagamento__gte=data_inicio)
    if data_fim:
        cartao_nao_conciliado = cartao_nao_conciliado.filter(data_pagamento__lte=data_fim)
    cartao_nao_conciliado = cartao_nao_conciliado.order_by('-data_pagamento')[:50]

    # Extrato bancário não conciliado
    from extrato.models import Lancamento
    extrato_nao_conciliado = Lancamento.objects.filter(
        empresa_id=empresa_id,
        conciliado=False
    )
    if data_inicio:
        extrato_nao_conciliado = extrato_nao_conciliado.filter(data__gte=data_inicio)
    if data_fim:
        extrato_nao_conciliado = extrato_nao_conciliado.filter(data__lte=data_fim)
    extrato_nao_conciliado = extrato_nao_conciliado.order_by('-data')[:50]

    context = {
        'contas_nao_conciliadas': contas_nao_conciliadas,
        'cartao_nao_conciliado': cartao_nao_conciliado,
        'extrato_nao_conciliado': extrato_nao_conciliado,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
    }

    return render(request, 'contasareceber/nao_conciliados.html', context)


def nao_conciliados_excel(request):
    """Exporta os itens não conciliados para Excel"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    # Obter empresa
    from empresa.models import Empresa
    empresa = Empresa.objects.get(id=empresa_id)

    # Filtros de período
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    # Mesmo código para obter os dados
    contas_nao_conciliadas = ContaAReceber.objects.filter(
        empresa_id=empresa_id,
        status__in=['pendente', 'vencido']
    )
    if data_inicio:
        contas_nao_conciliadas = contas_nao_conciliadas.filter(data_vencimento__gte=data_inicio)
    if data_fim:
        contas_nao_conciliadas = contas_nao_conciliadas.filter(data_vencimento__lte=data_fim)
    contas_nao_conciliadas = contas_nao_conciliadas.order_by('data_vencimento')

    from relatoriorecebiveis.models import RelatorioRecebiveisMaquinaCartao
    cartao_nao_conciliado = RelatorioRecebiveisMaquinaCartao.objects.filter(
        empresa_id=empresa_id,
        conta_a_receber__isnull=True
    )
    if data_inicio:
        cartao_nao_conciliado = cartao_nao_conciliado.filter(data_pagamento__gte=data_inicio)
    if data_fim:
        cartao_nao_conciliado = cartao_nao_conciliado.filter(data_pagamento__lte=data_fim)
    cartao_nao_conciliado = cartao_nao_conciliado.order_by('-data_pagamento')

    from extrato.models import Lancamento
    extrato_nao_conciliado = Lancamento.objects.filter(
        empresa_id=empresa_id,
        conciliado=False
    )
    if data_inicio:
        extrato_nao_conciliado = extrato_nao_conciliado.filter(data__gte=data_inicio)
    if data_fim:
        extrato_nao_conciliado = extrato_nao_conciliado.filter(data__lte=data_fim)
    extrato_nao_conciliado = extrato_nao_conciliado.order_by('-data')

    # Criar workbook
    from openpyxl import Workbook
    wb = Workbook()

    # Função para adicionar cabeçalho
    def add_header(ws, titulo):
        ws['A1'] = 'Razão:'
        ws['B1'] = empresa.razao
        ws['A2'] = 'CNPJ:'
        ws['B2'] = empresa.cnpj
        ws['A3'] = 'Período:'
        periodo = f"{data_inicio or 'N/A'} a {data_fim or 'N/A'}"
        ws['B3'] = periodo
        ws['A4'] = 'Nome do Relatório:'
        ws['B4'] = titulo
        # Deixar linha 5 em branco
        return 6  # Próxima linha

    # Sheet 1: Contas a Receber Não Conciliadas
    ws1 = wb.active
    ws1.title = "Contas Não Conciliadas"
    next_row = add_header(ws1, "Contas a Receber Não Conciliadas")
    headers1 = ['Cliente', 'Data Vencimento', 'Valor Pendente', 'Status', 'Forma de Pgto', 'Doc']
    for col_num, header in enumerate(headers1, start=1):
        ws1.cell(row=next_row, column=col_num, value=header)
    for conta in contas_nao_conciliadas:
        ws1.append([
            conta.cliente,
            conta.data_vencimento.strftime('%d/%m/%Y') if conta.data_vencimento else '',
            float(conta.get_valor_pendente()),
            conta.get_status_display(),
            conta.forma_pagamento.descricao if conta.forma_pagamento else '',
            conta.doc or ''
        ])

    # Sheet 2: Relatório de Cartão Não Conciliado
    ws2 = wb.create_sheet("Cartão Não Conciliado")
    next_row = add_header(ws2, "Relatório de Cartão Não Conciliado")
    headers2 = ['Data da Venda', 'Data do Pagamento', 'Valor Bruto', 'Valor Líquido', 'Valor da Taxa', 'Autorização', 'Razão', 'Bandeira', 'Forma de Pgto']
    for col_num, header in enumerate(headers2, start=1):
        ws2.cell(row=next_row, column=col_num, value=header)
    for rel in cartao_nao_conciliado:
        ws2.append([
            rel.data_venda.strftime('%d/%m/%Y') if rel.data_venda and hasattr(rel.data_venda, 'strftime') else str(rel.data_venda) if rel.data_venda else '',
            rel.data_pagamento.strftime('%d/%m/%Y') if rel.data_pagamento and hasattr(rel.data_pagamento, 'strftime') else str(rel.data_pagamento) if rel.data_pagamento else '',
            float(rel.valor_bruto),
            float(rel.valor_liquido),
            float(rel.taxa_maquinha),
            rel.numero_autorizacao,
            rel.razao,
            rel.bandeira,
            rel.forma_pagamento
        ])

    # Sheet 3: Extrato Bancário Não Conciliado
    ws3 = wb.create_sheet("Extrato Não Conciliado")
    next_row = add_header(ws3, "Extrato Bancário Não Conciliado")
    headers3 = ['Data', 'Valor', 'Histórico', 'Documento', 'Banco']
    for col_num, header in enumerate(headers3, start=1):
        ws3.cell(row=next_row, column=col_num, value=header)
    for lanc in extrato_nao_conciliado:
        ws3.append([
            lanc.data.strftime('%d/%m/%Y'),
            float(lanc.valor),
            lanc.historico,
            lanc.documento or '',
            lanc.conta.banco.nome
        ])

    # Resposta HTTP
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=nao_conciliados.xlsx'
    wb.save(response)
    return response


def lancar_contas_selecionadas(request):
    """Cria lançamentos para contas a receber selecionadas"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    if request.method == 'POST':
        contas_ids = request.POST.getlist('contas')
        if not contas_ids:
            messages.error(request, 'Selecione pelo menos uma conta a receber.')
            filtros = extrair_filtros_contas_receber(request)
            return redirect(construir_url_crlistar_com_filtros(filtros))

        contas = ContaAReceber.objects.filter(id__in=contas_ids, empresa_id=empresa_id, status='pendente')
        lancamentos_criados = 0

        for conta in contas:
            # Criar movimento no extrato
            from extrato.models import ExtratoMovimento

            # Usar conta bancária da conta a receber se disponível, senão primeira conta ativa
            conta_banco = conta.conta_banco
            if not conta_banco:
                try:
                    conta_banco = ContaBancaria.objects.filter(empresa_id=empresa_id, status='A').first()
                except:
                    conta_banco = None

            # Gerar descrição
            numero_nota = conta.nota.numero_nota if conta.nota else "Sem Nota"
            parcela = conta.parcela if conta.parcela else "1/1"
            descricao = f'Conta a Receber - NF {numero_nota} - {parcela} - {conta.cliente}'

            # Criar movimento
            ExtratoMovimento.objects.create(
                empresa=conta.empresa,
                data_baixa=conta.data_vencimento,
                descricao=descricao,
                situacao='pendente',
                valor=conta.get_valor_pendente(),
                conta_receber=conta,
                conta_banco=conta_banco,
                categoria=conta.categoria
            )

            lancamentos_criados += 1

        messages.success(request, f'{lancamentos_criados} lançamento(s) criado(s) para as contas selecionadas.')
        filtros = extrair_filtros_contas_receber(request)
        return redirect(construir_url_crlistar_com_filtros(filtros))

    # Se não for POST, redirecionar para a lista
    return redirect('contasareceber:crlistar')


@login_required
def detalhes_modal(request, tipo, id):
    """
    View para fornecer dados ao modal de detalhes
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'error': 'Empresa não selecionada'}, status=400)

    try:
        # Buscar o objeto baseado no tipo
        if tipo == 'nf':
            from notasfiscais.models import NotaFiscalServico
            obj = NotaFiscalServico.objects.get(id=id, empresa_id=empresa_id)
            nf = obj
        elif tipo == 'conta':
            obj = ContaAReceber.objects.get(id=id, empresa_id=empresa_id)
            nf = obj.nota if obj.nota else None
        elif tipo == 'movimento':
            from extrato.models import ExtratoMovimento
            obj = ExtratoMovimento.objects.select_related(
                'baixa_receber', 'conta_receber', 'conta_receber__nota'
            ).get(id=id, empresa_id=empresa_id)
            nf = obj.conta_receber.nota if obj.conta_receber and obj.conta_receber.nota else None
        elif tipo == 'lancamento':
            from extrato.models import Lancamento
            obj = Lancamento.objects.get(id=id, empresa_id=empresa_id)
            # Para lançamento, tentar encontrar NF através de ExtratoMovimento
            movimento = ExtratoMovimento.objects.filter(lancamento=obj).first()
            nf = movimento.conta_receber.nota if movimento and movimento.conta_receber and movimento.conta_receber.nota else None
        else:
            return JsonResponse({'error': 'Tipo inválido'}, status=400)

        # Buscar dados relacionados
        conta_receber = None
        movimentos = []
        lancamento_conciliado = None

        if tipo == 'conta':
            # Para tipo conta, usar diretamente a conta encontrada
            conta_receber = obj
        elif nf:
            # Para outros tipos, buscar conta a receber da NF
            conta_receber = ContaAReceber.objects.filter(nota=nf, empresa_id=empresa_id).first()

        if conta_receber:
            # Buscar movimentos da conta a receber
            from extrato.models import ExtratoMovimento
            movimentos = (
                ExtratoMovimento.objects.filter(conta_receber=conta_receber)
                .order_by('-data_baixa')
                .select_related('baixa_receber')
            )

        if movimentos:
            # Buscar lançamentos conciliados dos movimentos
            from extrato.models import Lancamento
            lancamentos_ids = [m.lancamento_id for m in movimentos if m.lancamento_id]
            if lancamentos_ids:
                lancamento_conciliado = Lancamento.objects.filter(id__in=lancamentos_ids).first()

        movimentos_html_cr = '<p>Nenhum movimento encontrado</p>'
        if movimentos:
            partes_m = []
            for m in movimentos[:3]:
                br = m.baixa_receber
                d = float(br.desconto or 0) if br else 0.0
                j = float(br.juros or 0) if br else 0.0
                t = float(br.tarifas or 0) if br else 0.0
                partes_m.append(
                    f'<p><strong>{m.data_baixa}:</strong> R$ {m.valor:.2f} — '
                    f'Desconto R$ {d:.2f} | Juros R$ {j:.2f} | Tarifa R$ {t:.2f} — {m.descricao}</p>'
                )
            movimentos_html_cr = ''.join(partes_m)

        # Construir HTML do modal
        html = f"""
        <div class="row">
            <div class="col-md-6">
                <h5>Nota Fiscal</h5>
                {f'<p><strong>Número:</strong> {nf.numero_nota}</p>' if nf else '<p>Nenhuma NF encontrada</p>'}
                {f'<p><strong>Cliente:</strong> {nf.cliente}</p>' if nf else ''}
                {f'<p><strong>Valor:</strong> R$ {nf.valor_liquido:.2f}</p>' if nf else ''}
            </div>
            <div class="col-md-6">
                <h5>Conta a Receber</h5>
                {f'<p><strong>Cliente:</strong> {conta_receber.cliente}</p>' if conta_receber else '<p>Nenhuma conta encontrada</p>'}
                {f'<p><strong>Valor:</strong> R$ {conta_receber.valor_a_receber:.2f}</p>' if conta_receber else ''}
                {f'<p><strong>Status:</strong> {conta_receber.get_status_display()}</p>' if conta_receber else ''}
            </div>
        </div>
        <div class="row mt-3">
            <div class="col-md-6">
                <h5>Movimentos do Extrato</h5>
                {movimentos_html_cr}
            </div>
            <div class="col-md-6">
                <h5>Lançamento Conciliado</h5>
                {f'<p><strong>Data:</strong> {lancamento_conciliado.data}</p>' if lancamento_conciliado else '<p>Nenhum lançamento conciliado</p>'}
                {f'<p><strong>Valor:</strong> R$ {lancamento_conciliado.valor:.2f}</p>' if lancamento_conciliado else ''}
                {f'<p><strong>Histórico:</strong> {lancamento_conciliado.historico}</p>' if lancamento_conciliado else ''}
            </div>
        </div>
        """

        return JsonResponse({'html': html})

    except Exception as e:
        return JsonResponse({'error': f'Erro interno: {str(e)}'}, status=500)


def processar_baixa_com_ajustes(
    conta, form, lancamentos_selecionados, user, request, filtros_query='', resolucao_diferenca='igual'
):
    """Processa a baixa com possíveis ajustes e conciliação."""
    try:
        print(f"DEBUG: Iniciando processamento da baixa para conta {conta.id}")
        from extrato.models import Conciliacao, ExtratoMovimento

        # Verificar se a conta já foi totalmente paga
        if conta.status == 'pago':
            print("DEBUG: ERRO - Conta já foi totalmente paga")
            messages.error(request, 'Esta conta já foi totalmente paga.')
            return redirect('contasareceber:baixar', pk=conta.pk)

        # Cria a baixa
        print("DEBUG: Criando baixa...")
        baixa = form.save(commit=False)
        baixa.conta_a_receber = conta
        baixa.empresa = conta.empresa
        print(f"DEBUG: Baixa criada - conta_a_receber: {baixa.conta_a_receber}, empresa: {baixa.empresa}")

        lista_lanc = []
        if lancamentos_selecionados is not None:
            qs = (
                lancamentos_selecionados.order_by('data', 'id')
                if hasattr(lancamentos_selecionados, 'order_by')
                else lancamentos_selecionados
            )
            lista_lanc = list(qs)

        if lista_lanc:
            valor_lancamentos_dec = sum(Decimal(str(l.valor)) for l in lista_lanc)
            baixa.data_recebimento = lista_lanc[0].data
            # VR + juros − desconto = soma do extrato (crédito no banco); tarifa só no título.
            baixa.valor_recebido = (
                valor_lancamentos_dec
                - Decimal(str(baixa.juros or 0))
                + Decimal(str(baixa.desconto or 0))
            )
            print(
                f"DEBUG: Conciliação — {len(lista_lanc)} lançamento(s); soma extrato R$ {valor_lancamentos_dec}; "
                f"valor_recebido R$ {baixa.valor_recebido}"
            )

        vr_final = Decimal(str(baixa.valor_recebido or 0))
        if vr_final <= 0:
            print("DEBUG: ERRO - Valor recebido deve ser maior que zero")
            messages.error(request, 'O valor recebido deve ser maior que zero.')
            url = reverse('contasareceber:baixar', kwargs={'pk': conta.pk})
            if filtros_query:
                url += '?' + filtros_query
            return HttpResponseRedirect(url)

        # Verificar se já existe uma baixa similar para evitar duplicatas
        baixas_existentes = BaixaContaAReceber.objects.filter(
            conta_a_receber=conta,
            data_recebimento=baixa.data_recebimento,
            valor_recebido=baixa.valor_recebido,
            conta_banco=baixa.conta_banco
        )
        if baixas_existentes.exists():
            print(f"DEBUG: AVISO - Já existe uma baixa similar! ID existente: {baixas_existentes.first().id}")
            messages.warning(request, 'Já existe um recebimento similar registrado para esta conta.')
            return redirect('contasareceber:baixar', pk=conta.pk)

        # PRIMEIRO: Atualizar a conta principal com a conta bancária selecionada
        conta_banco_selecionada = form.cleaned_data.get('conta_banco')
        print(f"DEBUG: Conta bancária selecionada: {conta_banco_selecionada}")

        if conta_banco_selecionada:
            print(f"DEBUG: Conta bancária ID: {conta_banco_selecionada.id}, Tipo: {conta_banco_selecionada.tipo}")

            # Verificar se o ID está no queryset do form
            ids_disponiveis = list(form.fields['conta_banco'].queryset.values_list('id', flat=True))
            if conta_banco_selecionada.id not in ids_disponiveis:
                print(f"DEBUG: AVISO - Conta ID {conta_banco_selecionada.id} não está no queryset do form! IDs disponíveis: {ids_disponiveis}")
                messages.error(request, f'Conta bancária selecionada não está disponível para esta empresa. Selecione outra conta bancária.')
                return redirect('contasareceber:baixar', pk=conta.pk)

            # Validação adicional: verificar se a conta bancária existe no banco
            try:
                from extrato.models import ContaBancaria
                conta_banco_validada = ContaBancaria.objects.get(id=conta_banco_selecionada.id)
                print(f"DEBUG: Conta bancária validada no banco: {conta_banco_validada} (empresa: {conta_banco_validada.empresa}, status: {conta_banco_validada.status})")

                # Verificar se pertence à empresa correta
                if conta_banco_validada.empresa != conta.empresa:
                    print(f"DEBUG: ERRO - Conta bancária pertence a empresa {conta_banco_validada.empresa}, mas conta é da empresa {conta.empresa}")
                    messages.error(request, f'Conta bancária selecionada pertence a outra empresa.')
                    return redirect('contasareceber:baixar', pk=conta.pk)

            except ContaBancaria.DoesNotExist:
                print(f"DEBUG: ERRO - Conta bancária ID {conta_banco_selecionada.id} não existe no banco!")
                messages.error(request, f'Conta bancária selecionada não existe mais. Selecione outra conta bancária.')
                return redirect('contasareceber:baixar', pk=conta.pk)
            except Exception as e:
                print(f"DEBUG: ERRO ao validar conta bancária: {str(e)}")
                messages.error(request, f'Erro ao validar conta bancária: {str(e)}')
                return redirect('contasareceber:baixar', pk=conta.pk)

            # Atualizar a conta principal com a conta bancária selecionada
            print(f"DEBUG: Conta atual tem conta_banco: {conta.conta_banco}")
            if not conta.conta_banco or conta.conta_banco.id != conta_banco_selecionada.id:
                print(f"DEBUG: Atualizando conta principal com nova conta bancária...")
                try:
                    conta.conta_banco = conta_banco_selecionada
                    conta.save()
                    print(f"DEBUG: Conta principal atualizada com conta bancária: {conta.conta_banco}")
                except Exception as e:
                    print(f"DEBUG: ERRO ao salvar conta principal: {str(e)}")
                    messages.error(request, f'Erro ao atualizar conta com conta bancária: {str(e)}')
                    return redirect('contasareceber:baixar', pk=conta.pk)
            else:
                print(f"DEBUG: Conta principal já tem a conta bancária correta")

        # Agora definir conta_banco na baixa (usando da conta principal)
        baixa.conta_banco = conta.conta_banco
        print(f"DEBUG: Conta bancária definida na baixa: {baixa.conta_banco}")

        print("DEBUG: Salvando baixa...")
        try:
            baixa.save()
            print(f"DEBUG: Baixa salva com ID: {baixa.id}")
        except Exception as save_error:
            print(f"DEBUG: ERRO ao salvar baixa: {str(save_error)}")
            # Verificar se é erro de foreign key
            if "foreign key constraint" in str(save_error).lower() or "cannot add or update a child row" in str(save_error).lower():
                print(f"DEBUG: ERRO DE FOREIGN KEY - conta_banco_id: {baixa.conta_banco.id if baixa.conta_banco else None}")
                messages.error(request, 'Erro de integridade: A conta bancária selecionada não existe ou não está disponível. Selecione outra conta bancária.')
            else:
                messages.error(request, f'Erro ao salvar baixa: {str(save_error)}')
            return redirect('contasareceber:baixar', pk=conta.pk)

        # Verificar se a conta_banco é válida para a empresa
        # Para contas do tipo CAIXA, permitir uso entre empresas (caixa compartilhado)
        if baixa.conta_banco:
            try:
                from extrato.models import ContaBancaria
                conta_banco_obj = ContaBancaria.objects.get(id=baixa.conta_banco.id)
                if conta_banco_obj.tipo != 'CAIXA' and conta_banco_obj.empresa != conta.empresa:
                    print(f"DEBUG: ERRO - Conta bancária {baixa.conta_banco.id} não pertence à empresa {conta.empresa.id}")
                    messages.error(request, f'Conta bancária selecionada não pertence à empresa atual.')
                    return redirect('contasareceber:baixar', pk=conta.pk)
                print("DEBUG: Conta bancária validada com sucesso")
            except ContaBancaria.DoesNotExist:
                print(f"DEBUG: ERRO - Conta bancária {baixa.conta_banco.id} não existe")
                messages.error(request, f'Conta bancária selecionada não existe.')
                return redirect('contasareceber:baixar', pk=conta.pk)

        # Verifica se deve fazer conciliação automática
        conciliacao_realizada = False
        if lista_lanc:
            print("DEBUG: Verificando conciliação automática...")
            valor_lancamentos = sum(Decimal(str(l.valor)) for l in lista_lanc)
            liquido_banco_baixa = (
                Decimal(str(baixa.valor_recebido))
                + Decimal(str(baixa.juros or 0))
                - Decimal(str(baixa.desconto or 0))
            )
            print(f"DEBUG: Líquido bancário da baixa (vs extrato): R$ {liquido_banco_baixa}")

            # Conciliação: soma do extrato = líquido creditado (VR + juros − desconto).
            if abs(valor_lancamentos - liquido_banco_baixa) <= Decimal('0.01'):
                print("DEBUG: Criando conciliação...")
                # Criar idconciliacao
                conciliacao = Conciliacao.objects.create(
                    criado_por=user if user.is_authenticated else None,
                    observacao=f'Conciliação automática - Baixa conta {conta.id}'
                )
                print(f"DEBUG: Conciliação criada com ID: {conciliacao.id}")

                # Marcar lançamentos como conciliados (mesmo que já estejam)
                for lancamento in lista_lanc:
                    lancamento.conciliado = True
                    lancamento.idconciliacao = conciliacao
                    lancamento.save()
                    print(f"DEBUG: Lançamento {lancamento.id} marcado como conciliado")

                conciliacao_realizada = True
                print("DEBUG: Conciliação realizada com sucesso")
            else:
                print(
                    f"DEBUG: Valores não batem - Lançamentos: R$ {valor_lancamentos}, líquido baixa: R$ {liquido_banco_baixa}"
                )

        # Registrar movimento no extrato
        print("DEBUG: Registrando movimento no extrato...")

        # Buscar relatório de recebíveis relacionado para usar dados corretos
        from .cartao_aproximacao import buscar_relatorios_por_autorizacao
        relatorio = buscar_relatorios_por_autorizacao(
            conta.empresa_id, conta.autorizacao, conciliado=None
        ).first()

        # Usa a descrição fornecida pelo usuário ou gera uma conforme especificação
        descricao_movimento = form.cleaned_data.get('descricao')
        if not descricao_movimento:
            if conciliacao_realizada and lista_lanc:
                # Para conciliação, usar formato específico com dados do relatório
                lancamento = lista_lanc[0]  # Usar primeiro lançamento para histórico
                nota_fiscal = relatorio.nota_fiscal if relatorio else (conta.nota.numero_nota if conta.nota else "Sem Nota")
                parcela_info = ""
                if relatorio and relatorio.parcelas and relatorio.total_parcelas:
                    parcela_info = f"{relatorio.parcelas}/{relatorio.total_parcelas}"
                else:
                    parcela_info = conta.parcela if conta.parcela else "1/1"
                razao = relatorio.razao if relatorio else conta.cliente
                descricao_movimento = f"{lancamento.historico} - {nota_fiscal} {parcela_info} - {razao}"
            else:
                # Movimento normal sem conciliação
                numero_nota = conta.nota.numero_nota if conta.nota else "Sem Nota"
                parcela = conta.parcela if conta.parcela else "1/1"
                descricao_movimento = f"VLR REF A NF {numero_nota} - {parcela} - {conta.cliente}"
        print(f"DEBUG: Descrição do movimento: {descricao_movimento}")

        # Se houve conciliação, um ExtratoMovimento por lançamento (valor = linha do extrato)
        if conciliacao_realizada:
            print("DEBUG: Criando movimentos conciliados...")
            for lancamento in lista_lanc:
                print(f"DEBUG: Criando movimento para lançamento {lancamento.id}")
                valor_movimento = lancamento.valor
                try:
                    movimento = ExtratoMovimento.objects.create(
                        empresa=conta.empresa,
                        data_baixa=lancamento.data,
                        descricao=descricao_movimento,
                        situacao='recebido',
                        valor=valor_movimento,
                        conta_receber=conta,
                        baixa_receber=baixa,
                        lancamento=lancamento,
                        conta_banco=baixa.conta_banco,
                        categoria=conta.categoria
                    )
                    print(f"DEBUG: Movimento criado com ID: {movimento.id}")
                except Exception as e:
                    print(f"DEBUG: ERRO ao criar movimento conciliado: {str(e)}")
                    messages.error(request, f'Erro ao registrar movimento no extrato: {str(e)}')
                    return redirect('contasareceber:baixar', pk=conta.pk)
        else:
            print("DEBUG: Criando movimento normal...")
            # Movimento normal sem conciliação
            if lista_lanc:
                valor_movimento = baixa.valor_recebido
            else:
                valor_movimento = relatorio.valor_liquido if relatorio else baixa.valor_recebido
            try:
                movimento = ExtratoMovimento.objects.create(
                    empresa=conta.empresa,
                    data_baixa=baixa.data_recebimento,
                    descricao=descricao_movimento,
                    situacao='recebido',
                    valor=valor_movimento,
                    conta_receber=conta,
                    baixa_receber=baixa,
                    conta_banco=baixa.conta_banco,
                    categoria=conta.categoria
                )
                print(f"DEBUG: Movimento criado com ID: {movimento.id}")
            except Exception as e:
                print(f"DEBUG: ERRO ao criar movimento normal: {str(e)}")
                messages.error(request, f'Erro ao registrar movimento no extrato: {str(e)}')
                return redirect('contasareceber:baixar', pk=conta.pk)

        # CAIXA sem linhas de extrato: um lançamento manual espelhando a baixa.
        # Com extrato selecionado, os lançamentos bancários já existem (conciliação).
        if baixa.conta_banco and baixa.conta_banco.tipo == 'CAIXA' and not lista_lanc:
            print("DEBUG: Conta tipo CAIXA detectada - criando lançamento no extrato bancário...")
            try:
                from extrato.models import Lancamento

                # Criar hash único para evitar duplicatas
                import hashlib
                hash_unico = hashlib.md5(f"{conta.empresa.id}_{baixa.conta_banco.id}_{baixa.data_recebimento}_{baixa.valor_recebido}_{descricao_movimento}".encode()).hexdigest()

                # Verificar se já existe lançamento com mesmo hash
                lancamento_existente = Lancamento.objects.filter(
                    empresa=conta.empresa,
                    conta=baixa.conta_banco,
                    data=baixa.data_recebimento,
                    valor=baixa.valor_recebido,
                    hash_unico=hash_unico
                ).first()

                if not lancamento_existente:
                    # Criar lançamento no extrato bancário
                    novo_lancamento = Lancamento.objects.create(
                        empresa=conta.empresa,
                        conta=baixa.conta_banco,
                        banco=baixa.conta_banco.banco,
                        data=baixa.data_recebimento,
                        documento=conta.doc or '',
                        historico=descricao_movimento,
                        valor=baixa.valor_recebido,
                        conciliado=True,  # Já conciliado pois é entrada direta
                        origem="MANUAL_CAIXA"
                    )
                    print(f"DEBUG: Lançamento bancário criado com ID: {novo_lancamento.id}")

                    # Vincular o movimento do extrato ao lançamento bancário
                    movimento.lancamento = novo_lancamento
                    movimento.save()
                    print("DEBUG: Movimento vinculado ao lançamento bancário")
                else:
                    print(f"DEBUG: Lançamento bancário já existe (ID: {lancamento_existente.id})")

            except Exception as e:
                print(f"DEBUG: ERRO ao criar lançamento bancário para conta CAIXA: {str(e)}")
                messages.error(request, f'Erro ao registrar lançamento no extrato bancário: {str(e)}')
                return redirect('contasareceber:baixar', pk=conta.pk)

        # BaixaContaAReceber.save() já recalculou valor_recebido, tarifas, status na conta (fonte única).
        conta.refresh_from_db()

        # Gerar conta da diferença (parcela seguinte): título atual = valor recebido (1ª parcela); saldo = próxima parcela
        if lista_lanc and resolucao_diferenca == 'nova_conta':
            vl_nc = sum(Decimal(str(l.valor)) for l in lista_lanc)
            vr_nc = Decimal(str(baixa.valor_recebido))
            orig_valor = conta.valor_a_receber
            remainder = orig_valor - vr_nc

            if remainder > Decimal('0.01'):
                parcela_1 = _parcela_primeira_apos_split(conta)
                conta.parcela = parcela_1
                conta.valor_a_receber = vr_nc
                conta.save()
                BaixaContaAReceber.atualizar_totais_na_conta(conta)
                conta.refresh_from_db()

                parcela_nova = _parcela_para_titulo_diferenca(conta)
                obs = (
                    f'Saldo após baixa #{baixa.id} (conta #{conta.id}). '
                    f'Valor original R$ {orig_valor}; valor recebido na baixa R$ {vr_nc}; extrato R$ {vl_nc}.'
                )
                ContaAReceber.objects.create(
                    empresa=conta.empresa,
                    nota=conta.nota,
                    cliente=conta.cliente,
                    cnpj_cpf=conta.cnpj_cpf or '',
                    data_emissao=baixa.data_recebimento,
                    data_vencimento=baixa.data_recebimento,
                    valor_a_receber=remainder,
                    parcela=parcela_nova,
                    doc='',
                    forma_pagamento=conta.forma_pagamento,
                    autorizacao='',
                    observacao=obs,
                    categoria=conta.categoria,
                    conta_banco=None,
                    status='pendente',
                )
                messages.info(
                    request,
                    f'Parcela 1 ajustada ao valor recebido (R$ {vr_nc:.2f}). '
                    f'Foi criada a parcela {parcela_nova} no valor de R$ {remainder:.2f}.',
                )
            else:
                diff_nc = abs(vl_nc - vr_nc)
                if diff_nc > Decimal('0.01'):
                    parcela_nova = _parcela_para_titulo_diferenca(conta)
                    obs = (
                        f'Diferença na baixa #{baixa.id} (conta a receber #{conta.id}). '
                        f'Total extrato R$ {vl_nc} vs valor recebido R$ {vr_nc}.'
                    )
                    ContaAReceber.objects.create(
                        empresa=conta.empresa,
                        nota=conta.nota,
                        cliente=conta.cliente,
                        cnpj_cpf=conta.cnpj_cpf or '',
                        data_emissao=baixa.data_recebimento,
                        data_vencimento=baixa.data_recebimento,
                        valor_a_receber=diff_nc,
                        parcela=parcela_nova,
                        doc='',
                        forma_pagamento=conta.forma_pagamento,
                        autorizacao='',
                        observacao=obs,
                        categoria=conta.categoria,
                        conta_banco=None,
                        status='pendente',
                    )
                    messages.info(
                        request,
                        f'Foi criada uma conta a receber em aberto (parcela {parcela_nova}) no valor de R$ {diff_nc:.2f} referente à diferença entre extrato e valor recebido.',
                    )

        conta.refresh_from_db()

        # Marcar nota fiscal com status de conciliação apropriado
        print("DEBUG: Atualizando status da nota fiscal...")
        if conta.nota:
            try:
                if conciliacao_realizada:
                    # Houve conciliação - marcar como conciliada
                    conta.nota.status_conciliacao = 'conciliado'
                elif conta.status == 'pago':
                    # Conta totalmente paga sem conciliação - marcar como conciliada
                    conta.nota.status_conciliacao = 'conciliado'
                elif conta.valor_recebido > 0:
                    # Conta com pagamentos parciais - marcar como parcialmente conciliada
                    conta.nota.status_conciliacao = 'parcialmente_conciliado'
                else:
                    # Conta sem pagamentos - manter como não conciliada
                    conta.nota.status_conciliacao = 'nao_conciliado'
                conta.nota.save()
                print(f"DEBUG: Status da nota fiscal atualizado para: {conta.nota.status_conciliacao}")
            except Exception as e:
                print(f"DEBUG: ERRO ao atualizar nota fiscal: {str(e)}")
                messages.error(request, f'Erro ao atualizar status da nota fiscal: {str(e)}')
                return redirect('contasareceber:baixar', pk=conta.pk)

        print("DEBUG: Processamento concluído com sucesso")
        print(f"DEBUG: conciliacao_realizada: {conciliacao_realizada}")
        print(f"DEBUG: baixa.valor_recebido: {baixa.valor_recebido}")
        if conciliacao_realizada:
            success_msg = f'Conta baixada e conciliada com sucesso! Valor: R$ {baixa.valor_recebido:.2f}'
            print(f"DEBUG: Success message: {success_msg}")
            messages.success(request, success_msg)
        else:
            success_msg = f'Conta baixada com sucesso! Valor recebido: R$ {baixa.valor_recebido:.2f}'
            print(f"DEBUG: Success message: {success_msg}")
            messages.success(request, success_msg)
        print("DEBUG: Redirecionando para lista de contas...")
        url = reverse('contasareceber:crlistar')
        if filtros_query:
            url += '?' + filtros_query
        return HttpResponseRedirect(url)

    except Exception as e:
        # Verificar se é erro de foreign key
        error_message = str(e)
        if "foreign key constraint" in error_message.lower() or "cannot add or update a child row" in error_message.lower():
            messages.error(request, 'Erro de integridade: A conta bancária selecionada não existe ou não está disponível. Selecione outra conta bancária.')
        else:
            messages.error(request, f'Erro ao processar baixa: {str(e)}')

        # Tentar reverter a baixa se ela foi criada
        try:
            if 'baixa' in locals():
                baixa.delete()
        except:
            pass
        return redirect('contasareceber:baixar', pk=conta.pk)


def validar_contas_dinheiro(request):
    """AJAX: Validar se contas selecionadas têm cobrança DINHEIRO"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'error': 'Empresa não encontrada'}, status=400)

    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)

    try:
        import json
        data = json.loads(request.body.decode('utf-8'))
        contas_ids = data.get('contas_ids', [])

        if not contas_ids:
            return JsonResponse({'contas': [], 'contas_invalidas': []})

        # Buscar contas
        contas = ContaAReceber.objects.filter(id__in=contas_ids, empresa_id=empresa_id)

        contas_validas = []
        contas_invalidas = []

        for conta in contas:
            # Verificar se cobrança é DINHEIRO
            forma_pagamento_desc = conta.forma_pagamento.descricao if conta.forma_pagamento else ''
            if forma_pagamento_desc.upper() == 'DINHEIRO':
                contas_validas.append({
                    'id': conta.id,
                    'cliente': conta.cliente,
                    'numero_nota': conta.nota.numero_nota if conta.nota else None,
                    'valor_pendente': float(conta.get_valor_pendente()),
                    'data_vencimento': conta.data_vencimento.strftime('%d/%m/%Y') if conta.data_vencimento else None,
                    'forma_pagamento': forma_pagamento_desc
                })
            else:
                contas_invalidas.append({
                    'cliente': conta.cliente,
                    'forma_pagamento': forma_pagamento_desc
                })

        return JsonResponse({
            'contas': contas_validas,
            'contas_invalidas': contas_invalidas
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def contas_caixa(request):
    """AJAX: Buscar contas bancárias do tipo CAIXA"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'error': 'Empresa não encontrada'}, status=400)

    try:
        from extrato.models import ContaBancaria
        contas_caixa = ContaBancaria.objects.filter(
            empresa_id=empresa_id,
            tipo='CAIXA',
            status='A'
        ).order_by('banco__nome', 'descricao')

        data = []
        for conta in contas_caixa:
            data.append({
                'id': conta.id,
                'descricao': str(conta),
                'banco': conta.banco.nome if conta.banco else '',
                'agencia': conta.agencia,
                'conta': conta.conta
            })

        return JsonResponse({'contas_caixa': data})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def aplicar_categoria(request):
    """Aplicar categoria a múltiplas contas a receber selecionadas"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    if request.method != 'POST':
        return redirect('contasareceber:crlistar')

    redirect_to = (request.POST.get('redirect_to') or '').strip()
    filtros_query = request.POST.get('filtros_query') or ''

    def _redirect_apos_categoria():
        if redirect_to == 'categorizar_baixados':
            base = reverse('contasareceber:categorizar_baixados')
            if filtros_query:
                return redirect(f'{base}?{filtros_query}')
            return redirect(base)
        filtros = extrair_filtros_contas_receber(request)
        return redirect(construir_url_crlistar_com_filtros(filtros))

    try:
        empresa = Empresa.objects.get(id=empresa_id)
    except Empresa.DoesNotExist:
        messages.error(request, 'Empresa não encontrada.')
        return redirect('empresa:lista')

    # Obter dados do formulário
    contas_ids = request.POST.getlist('contas_ids')
    categoria_id = request.POST.get('categoria_id')

    if not contas_ids:
        messages.error(request, 'Nenhuma conta selecionada.')
        return _redirect_apos_categoria()

    if not categoria_id:
        messages.error(request, 'Categoria não selecionada.')
        return _redirect_apos_categoria()

    try:
        # Buscar categoria
        categoria = Categoria.objects.get(id=categoria_id, empresa_id=empresa_id)
    except Categoria.DoesNotExist:
        messages.error(request, 'Categoria não encontrada ou não pertence à empresa.')
        return _redirect_apos_categoria()

    # Buscar contas selecionadas
    contas = ContaAReceber.objects.filter(id__in=contas_ids, empresa=empresa)

    aplicadas = 0
    notas_atualizadas = 0

    for conta in contas:
        try:
            # Aplicar categoria à conta
            conta.categoria = categoria
            conta.save()

            # Se houver nota fiscal vinculada, aplicar categoria também
            if conta.nota:
                conta.nota.categoria = categoria
                conta.nota.save()
                notas_atualizadas += 1

            aplicadas += 1

        except Exception as e:
            print(f"Erro ao aplicar categoria à conta {conta.id}: {str(e)}")
            continue

    if aplicadas > 0:
        messages.success(request, f'Categoria "{categoria.nome}" aplicada a {aplicadas} conta(s) com sucesso.')
        if notas_atualizadas > 0:
            messages.info(request, f'Categoria também aplicada a {notas_atualizadas} nota(s) fiscal(is) vinculada(s).')
    else:
        messages.warning(request, 'Nenhuma conta foi atualizada.')

    return _redirect_apos_categoria()


def baixar_contas_dinheiro(request):
    """Processar baixa de múltiplas contas a receber em dinheiro"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    # Capturar parâmetros GET para preservar filtros na navegação
    filtros_query = request.GET.urlencode()

    if request.method != 'POST':
        return redirect('contasareceber:crlistar')

    try:
        empresa = Empresa.objects.get(id=empresa_id)
    except Empresa.DoesNotExist:
        messages.error(request, 'Empresa não encontrada.')
        return redirect('empresa:lista')

    # Obter dados do formulário
    contas_ids = request.POST.getlist('contas_ids')
    conta_banco_id = request.POST.get('conta_banco')
    data_recebimento = request.POST.get('data_recebimento')

    if not contas_ids:
        messages.error(request, 'Nenhuma conta selecionada.')
        filtros = extrair_filtros_contas_receber(request)
        return redirect(construir_url_crlistar_com_filtros(filtros))

    if not conta_banco_id:
        messages.error(request, 'Conta caixa não selecionada.')
        filtros = extrair_filtros_contas_receber(request)
        return redirect(construir_url_crlistar_com_filtros(filtros))

    if not data_recebimento:
        messages.error(request, 'Data de recebimento não informada.')
        filtros = extrair_filtros_contas_receber(request)
        return redirect(construir_url_crlistar_com_filtros(filtros))

    try:
        # Buscar conta bancária
        from extrato.models import ContaBancaria
        conta_banco = ContaBancaria.objects.get(id=conta_banco_id, empresa=empresa, tipo='CAIXA')
    except ContaBancaria.DoesNotExist:
        messages.error(request, 'Conta caixa não encontrada ou não pertence à empresa.')
        filtros = extrair_filtros_contas_receber(request)
        return redirect(construir_url_crlistar_com_filtros(filtros))

    # Buscar contas selecionadas
    contas = ContaAReceber.objects.filter(id__in=contas_ids, empresa=empresa, status='pendente')

    baixas_criadas = 0
    erros = []

    for conta in contas:
        try:
            # Validar novamente se é DINHEIRO
            forma_pagamento_desc = conta.forma_pagamento.descricao if conta.forma_pagamento else ''
            if forma_pagamento_desc.upper() != 'DINHEIRO':
                erros.append(f'Conta {conta.cliente} não tem cobrança DINHEIRO')
                continue

            # DEBUG: Log valores antes da baixa
            print(f"DEBUG BAIXA_DINHEIRO: Conta {conta.id} - Cliente: {conta.cliente}")
            print(f"DEBUG BAIXA_DINHEIRO: Status atual: {conta.status}")
            print(f"DEBUG BAIXA_DINHEIRO: Valor a receber: {conta.valor_a_receber}")
            print(f"DEBUG BAIXA_DINHEIRO: Valor já recebido: {conta.valor_recebido or 0}")
            print(f"DEBUG BAIXA_DINHEIRO: Desconto: {conta.desconto or 0}")
            print(f"DEBUG BAIXA_DINHEIRO: Juros: {conta.juros or 0}")
            print(f"DEBUG BAIXA_DINHEIRO: Tarifas: {conta.tarifas or 0}")

            # Calcular valor pendente ANTES de marcar conta como paga
            valor_pendente = conta.get_valor_pendente()
            print(f"DEBUG BAIXA_DINHEIRO: Valor pendente calculado: {valor_pendente}")

            # Criar baixa
            baixa = BaixaContaAReceber.objects.create(
                conta_a_receber=conta,
                empresa=empresa,
                data_recebimento=data_recebimento,  # Usar data informada pelo usuário
                valor_recebido=valor_pendente,  # Valor a receber
                conta_banco=conta_banco,
                tipo_baixa='total',
                observacao=f'Baixa em dinheiro - NF {conta.nota.numero_nota if conta.nota else "N/A"}'
            )
            print(f"DEBUG BAIXA_DINHEIRO: Baixa criada com ID {baixa.id} e valor {baixa.valor_recebido}")

            # BaixaContaAReceber.save() já agrega valor_recebido (e demais totais) na conta.
            # Não chamar conta.save() aqui: a instância em memória estaria defasada e sobrescreveria o valor gravado.
            conta.refresh_from_db()
            print(f"DEBUG BAIXA_DINHEIRO: Conta após baixa - Status: {conta.status}, Valor recebido: {conta.valor_recebido}")

            # Criar movimento no extrato
            numero_nota = conta.nota.numero_nota if conta.nota else "Sem Nota"
            historico = f"VLR REF A RECEB COM NF {numero_nota} – {conta.cliente}"

            from extrato.models import ExtratoMovimento
            movimento = ExtratoMovimento.objects.create(
                empresa=empresa,
                data_baixa=data_recebimento,
                descricao=historico,
                situacao='recebido',
                valor=valor_pendente,
                conta_receber=conta,
                baixa_receber=baixa,
                conta_banco=conta_banco,
                categoria=conta.categoria
            )
            print(f"DEBUG BAIXA_DINHEIRO: ExtratoMovimento criado com ID {movimento.id} e valor {movimento.valor}")

            # Criar lançamento no extrato bancário para conta CAIXA
            from extrato.models import Lancamento
            import hashlib
            hash_unico = hashlib.md5(f"{empresa.id}_{conta_banco.id}_{data_recebimento}_{valor_pendente}_{historico}".encode()).hexdigest()

            lancamento_existente = Lancamento.objects.filter(
                empresa=empresa,
                conta=conta_banco,
                data=data_recebimento,
                valor=valor_pendente,
                hash_unico=hash_unico
            ).first()

            if not lancamento_existente:
                lancamento = Lancamento.objects.create(
                    empresa=empresa,
                    conta=conta_banco,
                    banco=conta_banco.banco,
                    data=data_recebimento,
                    documento=conta.doc or '',
                    historico=historico,
                    valor=valor_pendente,
                    conciliado=True,
                    origem="MANUAL_CAIXA"
                )
                print(f"DEBUG BAIXA_DINHEIRO: Lancamento criado com ID {lancamento.id} e valor {lancamento.valor}")
            else:
                print(f"DEBUG BAIXA_DINHEIRO: Lancamento já existe com ID {lancamento_existente.id}")

            # Atualizar status de conciliação da nota fiscal
            if conta.nota:
                try:
                    conta.nota.status_conciliacao = 'conciliado'
                    conta.nota.save()
                    print(f"DEBUG BAIXA_DINHEIRO: Status da nota fiscal {conta.nota.numero_nota} atualizado para 'conciliado'")
                except Exception as e:
                    print(f"DEBUG BAIXA_DINHEIRO: ERRO ao atualizar nota fiscal {conta.nota.numero_nota}: {str(e)}")
                    erros.append(f'Erro ao atualizar status da nota fiscal {conta.nota.numero_nota}: {str(e)}')

            baixas_criadas += 1

        except Exception as e:
            print(f"DEBUG BAIXA_DINHEIRO: ERRO ao processar conta {conta.id}: {str(e)}")
            erros.append(f'Erro ao processar conta {conta.cliente}: {str(e)}')

    if baixas_criadas > 0:
        messages.success(request, f'{baixas_criadas} conta(s) baixada(s) com sucesso em dinheiro.')

    if erros:
        for erro in erros:
            messages.warning(request, erro)

    # Redirecionar mantendo os filtros
    url = reverse('contasareceber:crlistar')
    if filtros_query:
        url += '?' + filtros_query
    return HttpResponseRedirect(url)


def conciliar_cartao_por_autorizacao(request):
    """Conciliar contas a receber com relatório de recebíveis baseado no número de autorização.

    Vincula recebíveis da maquininha ao título e mantém status «cartão».
    O status só passa a «pago» quando o extrato bancário conciliar com esses recebíveis.
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return _redirect_sem_empresa(request)

    if request.method == 'POST':
        contas_ids = request.POST.getlist('contas')
        if not contas_ids:
            messages.error(request, 'Selecione pelo menos uma conta a receber.')
            return redirect('contasareceber:crlistar')

        contas = ContaAReceber.objects.filter(
            id__in=contas_ids,
            empresa_id=empresa_id,
            status__in=['pendente', 'cartao'],
        )
        conciliadas = 0
        from .cartao_aproximacao import (
            buscar_relatorios_por_autorizacao,
            _atualizar_conta_totais_cartao,
            _vincular_relatorio_na_conta,
        )
        from relatoriorecebiveis.models import RelatorioRecebiveisMaquinaCartao

        for conta in contas:
            autorizacao = conta.autorizacao
            if not autorizacao:
                continue

            # Já vinculados a este título
            ja_vinculados = list(
                RelatorioRecebiveisMaquinaCartao.objects.filter(
                    empresa_id=empresa_id,
                    conta_a_receber=conta,
                )
            )
            ids_vinculados = {r.pk for r in ja_vinculados}

            # Novos matches por autorização (ainda sem vínculo, ou só desta conta)
            candidatos = buscar_relatorios_por_autorizacao(
                empresa_id, autorizacao, conciliado=False
            )
            novos = []
            for rel in candidatos:
                if rel.pk in ids_vinculados:
                    continue
                if rel.conta_a_receber_id and rel.conta_a_receber_id != conta.pk:
                    continue
                novos.append(rel)

            if not novos and not ja_vinculados:
                continue

            for relatorio in novos:
                _vincular_relatorio_na_conta(conta, relatorio)
                conciliadas += 1

            # Recalcula soma de TODOS os recebíveis do título (parcelas acumuladas)
            todos = list(
                RelatorioRecebiveisMaquinaCartao.objects.filter(
                    empresa_id=empresa_id,
                    conta_a_receber=conta,
                )
            )
            if todos:
                _atualizar_conta_totais_cartao(conta, todos)

        if conciliadas > 0:
            messages.success(
                request,
                f'{conciliadas} lançamento(s) do relatório de recebíveis incluído(s) na baixa '
                f'(parcelas acumuladas por autorização).',
            )
        else:
            messages.warning(
                request,
                'Nenhum recebível novo encontrado para incluir na baixa das contas selecionadas.',
            )
        filtros = extrair_filtros_contas_receber(request)
        return redirect(construir_url_crlistar_com_filtros(filtros))

    # GET: mostrar contas pendentes para seleção
    contas_pendentes = ContaAReceber.objects.filter(
        empresa_id=empresa_id,
        status='pendente'
    ).order_by('data_vencimento')

    context = {
        'contas_pendentes': contas_pendentes,
    }

    return render(request, 'contasareceber/conciliar.html', context)

def _excluir_conta_a_receber_obj(conta, empresa_id):
    """
    Remove vínculos e exclui a conta.
    Retorna (ok: bool, mensagem: str). Não permite pago/cartão.
    """
    if conta.status == 'pago':
        return False, f'Conta #{conta.pk} ({conta.cliente}): está paga — estorne antes de excluir.'
    if conta.status == 'cartao':
        return False, f'Conta #{conta.pk} ({conta.cliente}): conciliada com cartão — desconciliar antes de excluir.'

    from relatoriorecebiveis.models import RelatorioRecebiveisMaquinaCartao
    from extrato.models import ExtratoMovimento

    RelatorioRecebiveisMaquinaCartao.objects.filter(
        conta_a_receber=conta,
        empresa_id=empresa_id,
    ).update(conta_a_receber=None)

    ExtratoMovimento.objects.filter(conta_receber=conta).delete()
    BaixaContaAReceber.objects.filter(conta_a_receber=conta).delete()

    if conta.nota:
        conta.nota.status_conciliacao = 'nao_conciliado'
        conta.nota.save(update_fields=['status_conciliacao'])

    conta.delete()
    return True, f'Conta #{conta.pk} excluída.'


def deletar_conta_a_receber(request, pk):
    """Deletar uma conta a receber"""
    empresa_id = request.session.get('empresa_id')

    if not empresa_id:
        return _redirect_sem_empresa(request)

    conta = get_object_or_404(ContaAReceber, pk=pk, empresa_id=empresa_id)

    # Verificar se a conta está paga
    if conta.status == 'pago':
        messages.warning(request, 'Esta conta já foi totalmente paga. Para excluir, primeiro estorne o pagamento.')
        return redirect('contasareceber:detalhes', pk=pk)

    # Verificar se a conta está conciliada com cartão
    if conta.status == 'cartao':
        messages.warning(request, 'Esta conta está conciliada com cartão. Para excluir, primeiro desconciliar.')
        return redirect('contasareceber:detalhes', pk=pk)

    if request.method == 'POST':
        try:
            ok, msg = _excluir_conta_a_receber_obj(conta, empresa_id)
            if ok:
                messages.success(request, 'Conta a receber excluída com sucesso!')
                filtros = extrair_filtros_contas_receber(request)
                return redirect(construir_url_crlistar_com_filtros(filtros))
            messages.warning(request, msg)
            return redirect('contasareceber:detalhes', pk=pk)
        except Exception as e:
            messages.error(request, f'Erro ao excluir conta: {str(e)}')
            return redirect('contasareceber:detalhes', pk=pk)

    # GET: mostrar confirmação
    context = {
        'conta': conta,
    }

    return render(request, 'contasareceber/deletar.html', context)


@login_required
def excluir_contas_selecionadas(request):
    """Exclui em lote as contas a receber selecionadas (POST)."""
    empresa_id = request.session.get('empresa_id')
    filtros = extrair_filtros_contas_receber(request)

    if not empresa_id:
        return _redirect_sem_empresa(request)

    if request.method != 'POST':
        messages.warning(request, 'Use a seleção na listagem para excluir contas.')
        return redirect(construir_url_crlistar_com_filtros(filtros))

    raw_ids = request.POST.getlist('contas')
    if not raw_ids:
        messages.warning(request, 'Nenhuma conta selecionada para exclusão.')
        return redirect(construir_url_crlistar_com_filtros(filtros))

    try:
        contas_ids = [int(x) for x in raw_ids]
    except (TypeError, ValueError):
        messages.error(request, 'IDs de conta inválidos.')
        return redirect(construir_url_crlistar_com_filtros(filtros))

    contas = ContaAReceber.objects.filter(pk__in=contas_ids, empresa_id=empresa_id)
    sucesso = 0
    avisos = []

    with transaction.atomic():
        for conta in contas:
            try:
                ok, msg = _excluir_conta_a_receber_obj(conta, empresa_id)
                if ok:
                    sucesso += 1
                else:
                    avisos.append(msg)
            except Exception as e:
                avisos.append(f'Conta #{conta.pk}: {e}')

    if sucesso:
        messages.success(
            request,
            f'{sucesso} conta(s) a receber excluída(s) com sucesso.',
        )
    if avisos:
        for a in avisos[:8]:
            messages.warning(request, a)
        if len(avisos) > 8:
            messages.warning(request, f'... e mais {len(avisos) - 8} aviso(s).')
    if not sucesso and not avisos:
        messages.warning(request, 'Nenhuma conta válida encontrada para exclusão.')

    return redirect(construir_url_crlistar_com_filtros(filtros))


@login_required
def buscar_categorias(request):
    """Busca categorias para autocomplete"""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'error': 'Empresa não encontrada na sessão.'}, status=400)

    try:
        empresa = Empresa.objects.get(id=empresa_id)
    except Empresa.DoesNotExist:
        return JsonResponse({'error': 'Empresa não encontrada.'}, status=404)

    termo = request.GET.get('q', '').strip()
    if len(termo) < 4:
        return JsonResponse({'categorias': []})

    # Buscar categorias que contenham o termo no nome ou classificação
    categorias = Categoria.objects.filter(
        empresa=empresa
    ).filter(
        Q(nome__icontains=termo) | Q(classificacao__icontains=termo)
    ).order_by('tipo', 'nome')[:20]  # Limitar a 20 resultados

    # Preparar dados para resposta
    categorias_data = []
    for categoria in categorias:
        tipo_display = {
            'R': 'Receita',
            'D': 'Despesa',
            'I': 'Investimento',
            'L': 'Distribuição de Lucro'
        }.get(categoria.tipo, categoria.tipo)

        categorias_data.append({
            'id': categoria.id,
            'nome': categoria.nome,
            'classificacao': categoria.classificacao,
            'nome_completo': f"{categoria.classificacao} {categoria.nome}",
            'tipo': categoria.tipo,
            'tipo_display': tipo_display,
            'grupo': categoria.grupo
        })

    return JsonResponse({'categorias': categorias_data})


def sugerir_conciliacao_cartao_aproximacao(request):
    """AJAX: sugere pares título ↔ recebível (±2 dias, valor aproximado)."""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'error': 'Empresa não encontrada na sessão.'}, status=400)
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)
    raw_ids = data.get('contas_ids') or []
    if not isinstance(raw_ids, list) or not raw_ids:
        return JsonResponse({'error': 'Informe contas_ids (lista de IDs).'}, status=400)
    try:
        conta_ids = [int(x) for x in raw_ids]
    except (TypeError, ValueError):
        return JsonResponse({'error': 'contas_ids deve ser uma lista de inteiros.'}, status=400)

    from .cartao_aproximacao import build_suggestions

    out = build_suggestions(int(empresa_id), conta_ids)
    if out.get('erro'):
        return JsonResponse({**out, 'ok': False}, status=400)
    return JsonResponse({**out, 'ok': True})


def confirmar_conciliacao_cartao_aproximacao(request):
    """AJAX: aplica vínculos (grupos com vários recebíveis por título ou legado pares)."""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'error': 'Empresa não encontrada na sessão.'}, status=400)
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    from .cartao_aproximacao import aplicar_grupos, aplicar_pares

    grupos = data.get('grupos')
    if isinstance(grupos, list) and grupos:
        aplicados, erros = aplicar_grupos(int(empresa_id), grupos)
    else:
        pares = data.get('pares')
        if not isinstance(pares, list):
            return JsonResponse(
                {'error': 'Informe grupos [{conta_id, relatorio_ids}] ou pares legados.'},
                status=400,
            )
        aplicados, erros = aplicar_pares(int(empresa_id), pares)
    return JsonResponse(
        {
            'ok': True,
            'aplicados': aplicados,
            'erros': erros,
        }
    )
