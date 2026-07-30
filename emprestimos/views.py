from decimal import Decimal, ROUND_HALF_UP
import json
import re
from datetime import date, datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q, Sum
from django.db.models.functions import TruncMonth
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_GET, require_POST

from empresa.models import Empresa
from extrato.models import Banco

from .bradesco_pdf import detectar_e_parsear_pdf_emprestimo
from .forms import EmprestimoForm
from .models import Emprestimo, IndicadorCalculoSicoob, ParcelaEmprestimo, SimulacaoQuitacaoEmprestimo
from .sicoob_pdf import parse_extrato_sicoob


def _empresa_sessao(request):
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return None
    return Empresa.objects.filter(pk=empresa_id).first()


_MESES_PT = (
    '',
    'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro',
)


def _rotulo_mes_ano(valor: str) -> str:
    try:
        dt = datetime.strptime(valor, '%Y-%m')
    except (TypeError, ValueError):
        return valor
    nome = _MESES_PT[dt.month] if 1 <= dt.month <= 12 else str(dt.month)
    return f'{nome}/{dt.year}'


def _meses_abertos_opcoes(qs) -> list[dict[str, str]]:
    meses = (
        qs.annotate(mes=TruncMonth('data_vencimento'))
        .values_list('mes', flat=True)
        .distinct()
        .order_by('mes')
    )
    opcoes: list[dict[str, str]] = []
    for mes in meses:
        if not mes:
            continue
        valor = mes.strftime('%Y-%m')
        opcoes.append({'valor': valor, 'rotulo': _rotulo_mes_ano(valor)})
    return opcoes


def _banco_sicoob():
    """Resolve o cadastro Sicoob (com logo) usado nos extratos importados."""
    banco = (
        Banco.objects.filter(codigo='756').first()
        or Banco.objects.filter(nome__iexact='SICOOB').first()
        or Banco.objects.filter(nome__icontains='sicoob').first()
    )
    if banco:
        return banco
    return Banco.objects.create(nome='SICOOB', codigo='756')


def _banco_bradesco():
    """Resolve o cadastro Bradesco usado nos documentos de evolução de dívida."""
    banco = (
        Banco.objects.filter(codigo__in=('237', '0237')).first()
        or Banco.objects.filter(codigo='237').first()
        or Banco.objects.filter(nome__iexact='BRADESCO').first()
        or Banco.objects.filter(nome__icontains='bradesco').first()
    )
    if banco:
        return banco
    return Banco.objects.create(nome='BRADESCO', codigo='237')


def _banco_caixa():
    """Resolve o cadastro Caixa usado nos demonstrativos de evolução contratual."""
    banco = (
        Banco.objects.filter(codigo__in=('104', '0104')).first()
        or Banco.objects.filter(nome__iexact='CAIXA').first()
        or Banco.objects.filter(nome__icontains='caixa econ').first()
        or Banco.objects.filter(nome__icontains='caixa').first()
    )
    if banco:
        return banco
    return Banco.objects.create(nome='CAIXA', codigo='104')


def _banco_daycoval():
    """Resolve o cadastro Daycoval usado nos fluxos financeiros de leasing."""
    banco = (
        Banco.objects.filter(codigo__in=('707', '0707')).first()
        or Banco.objects.filter(nome__iexact='DAYCOVAL').first()
        or Banco.objects.filter(nome__icontains='daycoval').first()
    )
    if banco:
        return banco
    return Banco.objects.create(nome='DAYCOVAL', codigo='707')


def _banco_do_pdf(dados: dict):
    banco_key = (dados.get('banco') or '').lower()
    if banco_key == 'bradesco':
        return _banco_bradesco()
    if banco_key == 'caixa':
        return _banco_caixa()
    if banco_key == 'daycoval':
        return _banco_daycoval()
    return _banco_sicoob()


def _parse_taxa_post(raw) -> Decimal | None:
    """Converte taxa do formulário (BR ou ponto). None se vazio/inválido."""
    if raw is None:
        return None
    t = str(raw).strip()
    if not t:
        return None
    t = t.replace('%', '').strip()
    if ',' in t:
        t = t.replace('.', '').replace(',', '.')
    try:
        return Decimal(t)
    except Exception:
        return None


def _taxa_juros_am_efetiva(emp, request=None) -> Decimal:
    """
    Taxa de juros a.m. usada nos cálculos (sem mora).
    No POST da simulação, aceita override de taxa_juros_am.
    """
    juros = emp.taxa_juros_am or Decimal('0')
    if request is not None and request.method == 'POST':
        j = _parse_taxa_post(request.POST.get('taxa_juros_am'))
        if j is not None:
            juros = j
    return juros.quantize(Decimal('0.0001'))


def _taxa_calculo_am(emp, request=None) -> Decimal:
    """Alias legado — cálculos usam só taxa de juros a.m."""
    return _taxa_juros_am_efetiva(emp, request)


def _emprestimo_filtros_request(request):
    """Lê filtros GET da listagem de empréstimos."""
    banco_id = (request.GET.get('banco') or '').strip()
    clientes = [c.strip() for c in request.GET.getlist('cliente') if c and str(c).strip()]
    try:
        banco_id = int(banco_id) if banco_id else None
    except (TypeError, ValueError):
        banco_id = None
    return {'banco_id': banco_id, 'clientes': clientes}


def _emprestimo_filtros_opcoes(empresa):
    """Bancos e clientes disponíveis nos empréstimos da empresa."""
    qs = Emprestimo.objects.filter(empresa=empresa)
    bancos = (
        Banco.objects.filter(emprestimos__empresa=empresa)
        .distinct()
        .order_by('nome')
    )
    clientes = sorted({
        c.strip() for c in qs.values_list('cliente', flat=True) if c and str(c).strip()
    })
    return bancos, clientes


def _emprestimo_list_rows(empresa, *, banco_id=None, clientes=None):
    """Monta linhas e totais da listagem de empréstimos (tela e Excel)."""
    qs = (
        Emprestimo.objects.filter(empresa=empresa)
        .select_related('indicador', 'banco')
        .order_by('-data_operacao', '-id')
    )
    if banco_id:
        qs = qs.filter(banco_id=banco_id)
    if clientes:
        q_clientes = Q()
        for cliente in clientes:
            q_clientes |= Q(cliente__iexact=cliente)
        qs = qs.filter(q_clientes)
    rows = []
    tot_valor_contrato = Decimal('0')
    tot_saldo_principal = Decimal('0')
    tot_parcelas_abertas = Decimal('0')
    tot_ja_pago = Decimal('0')
    tot_juros_pago = Decimal('0')
    tot_correcao = Decimal('0')
    tot_mora = Decimal('0')
    tot_qtd_abertas = 0
    tot_qtd_parcelas = 0
    for emp in qs:
        abertas = emp.parcelas.filter(status='aberta')
        pagas = emp.parcelas.filter(status='paga')
        total_aberto = abertas.aggregate(t=Sum('valor_parcela'))['t'] or Decimal('0')
        saldo_principal = abertas.aggregate(t=Sum('amortizacao'))['t'] or Decimal('0')
        ja_pago = (
            pagas.aggregate(t=Sum(Coalesce('valor_pago', 'valor_parcela')))['t']
            or Decimal('0')
        )
        juros_pago = pagas.aggregate(t=Sum('juros'))['t'] or Decimal('0')
        total_correcao = emp.parcelas.aggregate(t=Sum('correcao'))['t'] or Decimal('0')
        total_mora = emp.parcelas.aggregate(t=Sum('mora'))['t'] or Decimal('0')
        qtd_parcelas = emp.parcelas.count()
        qtd_abertas = abertas.count()
        rows.append({
            'emprestimo': emp,
            'qtd_parcelas': qtd_parcelas,
            'qtd_abertas': qtd_abertas,
            'total_aberto': total_aberto,
            'saldo_principal': saldo_principal,
            'ja_pago': ja_pago,
            'juros_pago': juros_pago,
            'total_correcao': total_correcao,
            'total_mora': total_mora,
        })
        tot_valor_contrato += emp.valor_contrato or Decimal('0')
        tot_saldo_principal += saldo_principal
        tot_parcelas_abertas += total_aberto
        tot_ja_pago += ja_pago
        tot_juros_pago += juros_pago
        tot_correcao += total_correcao
        tot_mora += total_mora
        tot_qtd_abertas += qtd_abertas
        tot_qtd_parcelas += qtd_parcelas

    totais = {
        'valor_contrato': tot_valor_contrato,
        'saldo_principal': tot_saldo_principal,
        'total_aberto': tot_parcelas_abertas,
        'ja_pago': tot_ja_pago,
        'juros_pago': tot_juros_pago,
        'total_correcao': tot_correcao,
        'total_mora': tot_mora,
        'qtd_abertas': tot_qtd_abertas,
        'qtd_parcelas': tot_qtd_parcelas,
        'qtd_contratos': len(rows),
    }
    return rows, totais


@login_required
def emprestimo_list(request):
    from urllib.parse import urlencode

    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    filtros = _emprestimo_filtros_request(request)
    bancos, clientes = _emprestimo_filtros_opcoes(empresa)
    rows, totais = _emprestimo_list_rows(
        empresa,
        banco_id=filtros['banco_id'],
        clientes=filtros['clientes'],
    )
    export_params = []
    if filtros['banco_id']:
        export_params.append(('banco', str(filtros['banco_id'])))
    for cliente in filtros['clientes']:
        export_params.append(('cliente', cliente))

    return render(request, 'emprestimos/listar.html', {
        'title': 'Empréstimos bancários',
        'rows': rows,
        'empresa': empresa,
        'totais': totais,
        'bancos_filtro': bancos,
        'clientes_filtro': clientes,
        'filtros': filtros,
        'export_query_string': urlencode(export_params),
    })


@login_required
@require_GET
def emprestimo_list_excel(request):
    """Exporta a listagem de empréstimos para Excel."""
    from io import BytesIO

    from django.http import HttpResponse
    from django.utils import timezone as dj_tz
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    filtros = _emprestimo_filtros_request(request)
    rows, totais = _emprestimo_list_rows(
        empresa,
        banco_id=filtros['banco_id'],
        clientes=filtros['clientes'],
    )
    hoje = dj_tz.localdate()
    money_format = '#,##0.00'
    header_fill = PatternFill('solid', fgColor='198754')
    header_font = Font(bold=True, color='FFFFFF')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Empréstimos'

    ws['A1'] = 'Empréstimos bancários'
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:O1')
    ws['A2'] = str(empresa)
    ws['A3'] = f'Exportado em: {hoje.strftime("%d/%m/%Y")}'
    ws['A4'] = f'Contratos: {totais["qtd_contratos"]}'

    headers = [
        'Banco',
        'Contrato',
        'Cliente',
        'Valor contrato',
        'Início',
        'Indicador de cálculo',
        'Taxa juros a.m.',
        'Parcelas',
        'Em aberto',
        'Saldo principal',
        'Total parcelas abertas',
        'Total já pago',
        'Total juros pago',
        'Total correção',
        'Total mora',
    ]
    row0 = 6
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row0, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for i, row in enumerate(rows):
        r = row0 + 1 + i
        emp = row['emprestimo']
        banco_nome = emp.banco.nome if emp.banco_id else ''
        if emp.taxa_juros_am:
            taxa_txt = float(emp.taxa_juros_am)
        elif emp.taxa_juros_aa:
            taxa_txt = f'{emp.taxa_juros_aa} a.a.'
        else:
            taxa_txt = ''

        ws.cell(row=r, column=1, value=banco_nome)
        ws.cell(row=r, column=2, value=emp.numero_contrato)
        ws.cell(row=r, column=3, value=emp.cliente or '')
        cell = ws.cell(row=r, column=4, value=float(emp.valor_contrato or 0))
        cell.number_format = money_format
        ws.cell(
            row=r, column=5,
            value=emp.data_operacao.strftime('%d/%m/%Y') if emp.data_operacao else '',
        )
        ws.cell(row=r, column=6, value=emp.indicador_display if emp.indicador_display != '—' else '')
        taxa_cell = ws.cell(row=r, column=7, value=taxa_txt)
        if isinstance(taxa_txt, float):
            taxa_cell.number_format = '0.0000'
        ws.cell(row=r, column=8, value=row['qtd_parcelas'])
        ws.cell(row=r, column=9, value=row['qtd_abertas'])
        for col, key in enumerate(
            ('saldo_principal', 'total_aberto', 'ja_pago', 'juros_pago', 'total_correcao', 'total_mora'),
            start=10,
        ):
            cell = ws.cell(row=r, column=col, value=float(row[key] or 0))
            cell.number_format = money_format

    total_row = row0 + 1 + len(rows)
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    cell = ws.cell(row=total_row, column=4, value=float(totais['valor_contrato'] or 0))
    cell.number_format = money_format
    cell.font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=totais['qtd_parcelas']).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=totais['qtd_abertas']).font = Font(bold=True)
    for col, key in enumerate(
        ('saldo_principal', 'total_aberto', 'ja_pago', 'juros_pago', 'total_correcao', 'total_mora'),
        start=10,
    ):
        cell = ws.cell(row=total_row, column=col, value=float(totais[key] or 0))
        cell.number_format = money_format
        cell.font = Font(bold=True)

    widths = [14, 18, 28, 16, 12, 22, 14, 10, 10, 16, 20, 16, 16, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    empresa_slug = re.sub(r'[^\w.-]+', '_', str(empresa)[:40])
    nome = f'emprestimos_{empresa_slug}_{hoje.strftime("%Y%m%d")}.xlsx'

    response = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nome}"'
    return response


@login_required
def emprestimo_parcelas_abertas(request):
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    parcelas_qs = (
        ParcelaEmprestimo.objects.filter(
            emprestimo__empresa=empresa,
            status='aberta',
        )
        .select_related('emprestimo', 'emprestimo__banco', 'emprestimo__indicador')
    )
    meses_opcoes = _meses_abertos_opcoes(parcelas_qs)

    mes_filtro = (request.GET.get('mes') or '').strip()
    mes_rotulo = ''
    if mes_filtro:
        try:
            ano, mes = mes_filtro.split('-', 1)
            parcelas = parcelas_qs.filter(
                data_vencimento__year=int(ano),
                data_vencimento__month=int(mes),
            )
            mes_rotulo = _rotulo_mes_ano(mes_filtro)
        except (ValueError, TypeError):
            mes_filtro = ''
            parcelas = parcelas_qs
    else:
        parcelas = parcelas_qs

    parcelas = parcelas.order_by('data_vencimento', 'emprestimo__numero_contrato', 'numero')

    totais = parcelas.aggregate(
        valor_parcela=Sum('valor_parcela'),
        amortizacao=Sum('amortizacao'),
        juros=Sum('juros'),
        mora=Sum('mora'),
        correcao=Sum('correcao'),
    )

    return render(request, 'emprestimos/parcelas_abertas.html', {
        'title': 'Parcelas em aberto',
        'empresa': empresa,
        'parcelas': parcelas,
        'mes_filtro': mes_filtro,
        'mes_rotulo': mes_rotulo,
        'meses_opcoes': meses_opcoes,
        'totais': {
            'qtd': parcelas.count(),
            'valor_parcela': totais['valor_parcela'] or Decimal('0'),
            'amortizacao': totais['amortizacao'] or Decimal('0'),
            'juros': totais['juros'] or Decimal('0'),
            'mora': totais['mora'] or Decimal('0'),
            'correcao': totais['correcao'] or Decimal('0'),
        },
    })


@login_required
def emprestimo_cadastrar(request):
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    if request.method == 'POST':
        form = EmprestimoForm(request.POST, empresa=empresa)
        if form.is_valid():
            emp = form.save(commit=False)
            emp.empresa = empresa
            emp.save()
            messages.success(
                request,
                f'Contrato {emp.numero_contrato} cadastrado. '
                'Gere as parcelas (Price/SAC) ou importe o PDF na tela de detalhe.',
            )
            return redirect('emprestimos:detalhe', pk=emp.pk)
    else:
        form = EmprestimoForm(empresa=empresa)

    return render(request, 'emprestimos/cadastrar.html', {
        'title': 'Cadastrar empréstimo',
        'empresa': empresa,
        'form': form,
    })


@login_required
def emprestimo_importar(request):
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    if request.method == 'POST':
        pdf = request.FILES.get('pdf_file')
        if not pdf:
            messages.error(request, 'Selecione um arquivo PDF.')
            return redirect('emprestimos:importar')
        if not pdf.name.lower().endswith('.pdf'):
            messages.error(request, 'Envie um arquivo PDF.')
            return redirect('emprestimos:importar')

        try:
            dados = detectar_e_parsear_pdf_emprestimo(pdf)
        except Exception as exc:
            messages.error(request, f'Erro ao ler PDF: {exc}')
            return redirect('emprestimos:importar')

        with transaction.atomic():
            indicador_obj = IndicadorCalculoSicoob.from_texto_pdf(dados.get('indicador_calculo') or '')
            emp, created = Emprestimo.objects.update_or_create(
                empresa=empresa,
                numero_contrato=dados['numero_contrato'],
                defaults={
                    'banco': _banco_do_pdf(dados),
                    'cooperativa': dados['cooperativa'],
                    'cliente': dados['cliente'],
                    'modalidade': dados['modalidade'],
                    'data_operacao': dados['data_operacao'],
                    'data_vencimento': dados['data_vencimento'],
                    'prazo_dias': dados['prazo_dias'],
                    'valor_contrato': dados['valor_contrato'],
                    'valor_tributos': dados.get('valor_tributos') or Decimal('0'),
                    'valor_tarifas': dados.get('valor_tarifas') or Decimal('0'),
                    'valor_registros': dados.get('valor_registros') or Decimal('0'),
                    'valor_servicos_terceiros': dados.get('valor_servicos_terceiros') or Decimal('0'),
                    'saldo_devedor_atualizado': dados.get('saldo_devedor_atualizado') or Decimal('0'),
                    'taxa_juros_am': dados['taxa_juros_am'],
                    'taxa_juros_aa': dados.get('taxa_juros_aa') or Decimal('0'),
                    'taxa_multa_am': dados['taxa_multa_am'],
                    'taxa_mora_am': dados['taxa_mora_am'],
                    'indice_correcao': dados.get('indice_correcao') or '',
                    'indice_correcao_atraso': dados.get('indice_correcao_atraso') or '',
                    'pct_correcao_am': dados.get('pct_correcao_am') or Decimal('0'),
                    'pct_correcao_atraso_am': dados.get('pct_correcao_atraso_am') or Decimal('0'),
                    'indicador': indicador_obj,
                    'indicador_calculo': dados['indicador_calculo'],
                    'data_extrato': dados['data_extrato'],
                    'arquivo_origem': pdf.name[:255],
                },
            )
            # Substitui parcelas pelo extrato importado
            emp.parcelas.all().delete()
            ParcelaEmprestimo.objects.bulk_create([
                ParcelaEmprestimo(
                    emprestimo=emp,
                    numero=p['numero'],
                    data_vencimento=p['data_vencimento'],
                    valor_parcela=p['valor_parcela'],
                    amortizacao=p['amortizacao'],
                    juros=p['juros'],
                    data_pagamento=p['data_pagamento'],
                    historico=p['historico'],
                    valor_pago=p['valor_pago'],
                    mora=p['mora'],
                    multa=p.get('multa') or Decimal('0'),
                    iof=p['iof'],
                    correcao=p.get('correcao') or Decimal('0'),
                    status=p['status'],
                )
                for p in dados['parcelas']
            ])

        abertas = emp.parcelas.filter(status='aberta').count()
        acao = 'importado' if created else 'atualizado'
        messages.success(
            request,
            f'Contrato {emp.numero_contrato} {acao}: {len(dados["parcelas"])} parcelas '
            f'({abertas} em aberto).',
        )
        if dados.get('aviso'):
            messages.warning(request, dados['aviso'])
        return redirect('emprestimos:detalhe', pk=emp.pk)

    return render(request, 'emprestimos/importar.html', {
        'title': 'Importar PDF de empréstimo',
        'empresa': empresa,
    })


def _vals_parcela_pdf(row: dict) -> dict:
    valor_parcela = row.get('valor_parcela') or Decimal('0')
    juros = row.get('juros') or Decimal('0')
    amortizacao = row.get('amortizacao') or Decimal('0')
    correcao = row.get('correcao') or Decimal('0')
    if valor_parcela > 0 and amortizacao <= 0:
        amortizacao = max(
            Decimal('0'),
            (valor_parcela - juros - correcao).quantize(Decimal('0.01')),
        )
    return {
        'data_vencimento': row.get('data_vencimento'),
        'valor_parcela': valor_parcela,
        'amortizacao': amortizacao,
        'juros': juros,
        'data_pagamento': None,
        'historico': row.get('historico') or '',
        'valor_pago': None,
        'mora': row.get('mora') or Decimal('0'),
        'iof': row.get('iof') or Decimal('0'),
        'correcao': correcao,
        'status': 'aberta',
    }


def _reimportar_parcelas_daycoval(emp, parcelas_pdf: list[dict]) -> tuple[int, int, int, int, date | None]:
    """
    Daycoval: PARCELA + RESIDUAL com vencimentos distintos — casa por data de vencimento.
    """
    ultima_paga_venc = (
        emp.parcelas.filter(status='paga')
        .order_by('-data_vencimento')
        .values_list('data_vencimento', flat=True)
        .first()
    )
    por_vencimento = {p.data_vencimento: p for p in emp.parcelas.all()}
    max_num = emp.parcelas.order_by('-numero').values_list('numero', flat=True).first() or 0

    atualizadas = criadas = ignoradas = restauradas = 0

    for row in parcelas_pdf:
        venc = row.get('data_vencimento')
        if not venc:
            continue

        if ultima_paga_venc and venc <= ultima_paga_venc:
            existente = por_vencimento.get(venc)
            if existente and existente.status != 'paga':
                existente.status = 'paga'
                existente.save(update_fields=['status'])
                restauradas += 1
            ignoradas += 1
            continue

        vals = _vals_parcela_pdf(row)
        existente = por_vencimento.get(venc)
        if existente:
            if existente.status == 'paga':
                ignoradas += 1
                continue
            for k, v in vals.items():
                setattr(existente, k, v)
            existente.save()
            atualizadas += 1
        else:
            max_num += 1
            nova = ParcelaEmprestimo.objects.create(emprestimo=emp, numero=max_num, **vals)
            por_vencimento[venc] = nova
            criadas += 1

    return atualizadas, criadas, ignoradas, restauradas, ultima_paga_venc


@login_required
@require_POST
def emprestimo_importar_parcelas_pdf(request, pk):
    """
    Importa/corrige SOMENTE parcelas em aberto a partir do PDF.
    - Não altera parcelas já baixadas (status=paga) no sistema
    - Começa após a última parcela paga do sistema (ex.: última paga=31 → a partir da 32)
    - Grava valor da parcela, juros e amortização (provisão) conforme o PDF
    """
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    emp = get_object_or_404(Emprestimo, pk=pk, empresa=empresa)
    pdf = request.FILES.get('pdf_file')
    if not pdf:
        messages.error(request, 'Selecione o arquivo PDF do extrato.')
        return redirect('emprestimos:detalhe', pk=pk)
    if not pdf.name.lower().endswith('.pdf'):
        messages.error(request, 'Envie um arquivo PDF.')
        return redirect('emprestimos:detalhe', pk=pk)

    try:
        dados = detectar_e_parsear_pdf_emprestimo(pdf)
    except Exception as exc:
        messages.error(request, f'Erro ao ler PDF: {exc}')
        return redirect('emprestimos:detalhe', pk=pk)

    contrato_pdf = str(dados.get('numero_contrato') or '').strip()
    contrato_emp = str(emp.numero_contrato or '').strip()
    if contrato_pdf and contrato_emp and contrato_pdf != contrato_emp:
        messages.error(
            request,
            f'O PDF é do contrato {contrato_pdf}, mas esta tela é do contrato {contrato_emp}.',
        )
        return redirect('emprestimos:detalhe', pk=pk)

    parcelas_pdf = dados.get('parcelas') or []
    if not parcelas_pdf:
        messages.error(request, 'Nenhuma parcela encontrada no PDF.')
        return redirect('emprestimos:detalhe', pk=pk)

    # Âncora: última baixada no sistema — não mexer nela nem nas anteriores
    ultima_paga = (
        emp.parcelas.filter(status='paga')
        .order_by('-numero')
        .values_list('numero', flat=True)
        .first()
    )
    inicio_aberto = int(ultima_paga or 0) + 1

    atualizadas = 0
    criadas = 0
    ignoradas_baixadas = 0
    restauradas_status = 0
    ultima_paga_venc = None

    with transaction.atomic():
        indicador_obj = IndicadorCalculoSicoob.from_texto_pdf(dados.get('indicador_calculo') or '')
        emp.banco = emp.banco or _banco_do_pdf(dados)
        if dados.get('cooperativa'):
            emp.cooperativa = dados['cooperativa']
        if dados.get('cliente'):
            emp.cliente = dados['cliente']
        if dados.get('modalidade'):
            emp.modalidade = dados['modalidade']
        if dados.get('data_operacao'):
            emp.data_operacao = dados['data_operacao']
        if dados.get('data_vencimento'):
            emp.data_vencimento = dados['data_vencimento']
        if dados.get('prazo_dias') is not None:
            emp.prazo_dias = dados['prazo_dias']
        if dados.get('valor_contrato') is not None:
            emp.valor_contrato = dados['valor_contrato']
        if dados.get('valor_tributos') is not None:
            emp.valor_tributos = dados.get('valor_tributos') or Decimal('0')
        if dados.get('valor_tarifas') is not None:
            emp.valor_tarifas = dados.get('valor_tarifas') or Decimal('0')
        if dados.get('valor_registros') is not None:
            emp.valor_registros = dados.get('valor_registros') or Decimal('0')
        if dados.get('valor_servicos_terceiros') is not None:
            emp.valor_servicos_terceiros = dados.get('valor_servicos_terceiros') or Decimal('0')
        if dados.get('saldo_devedor_atualizado') is not None:
            emp.saldo_devedor_atualizado = dados.get('saldo_devedor_atualizado') or Decimal('0')
        if dados.get('taxa_juros_am') is not None:
            emp.taxa_juros_am = dados['taxa_juros_am']
        if dados.get('taxa_juros_aa') is not None:
            emp.taxa_juros_aa = dados['taxa_juros_aa'] or Decimal('0')
        if dados.get('taxa_multa_am') is not None:
            emp.taxa_multa_am = dados['taxa_multa_am']
        if dados.get('taxa_mora_am') is not None:
            emp.taxa_mora_am = dados['taxa_mora_am']
        emp.indice_correcao = dados.get('indice_correcao') or emp.indice_correcao or ''
        emp.indice_correcao_atraso = (
            dados.get('indice_correcao_atraso') or emp.indice_correcao_atraso or ''
        )
        if dados.get('pct_correcao_am') is not None:
            emp.pct_correcao_am = dados.get('pct_correcao_am') or Decimal('0')
        if dados.get('pct_correcao_atraso_am') is not None:
            emp.pct_correcao_atraso_am = dados.get('pct_correcao_atraso_am') or Decimal('0')
        if indicador_obj:
            emp.indicador = indicador_obj
        if dados.get('indicador_calculo'):
            emp.indicador_calculo = dados['indicador_calculo']
        if dados.get('data_extrato'):
            emp.data_extrato = dados['data_extrato']
        emp.arquivo_origem = pdf.name[:255]
        emp.save()

        if (dados.get('banco') or '').lower() == 'daycoval':
            atualizadas, criadas, ignoradas_baixadas, restauradas_status, ultima_paga_venc = (
                _reimportar_parcelas_daycoval(emp, parcelas_pdf)
            )
            ultima_paga = (
                emp.parcelas.filter(status='paga')
                .order_by('-numero')
                .values_list('numero', flat=True)
                .first()
            )
            inicio_aberto = None
        else:
            por_numero = {p.numero: p for p in emp.parcelas.all()}

            # Garante que baixadas acidentais (reabertas) voltem a paga se nº <= última paga
            for num, p in por_numero.items():
                if num < inicio_aberto and p.status != 'paga':
                    p.status = 'paga'
                    p.save(update_fields=['status'])
                    restauradas_status += 1

            for row in parcelas_pdf:
                numero = int(row.get('numero') or 0)
                if numero < inicio_aberto:
                    # Nunca altera baixadas / anteriores à âncora
                    ignoradas_baixadas += 1
                    continue

                vals = _vals_parcela_pdf(row)
                existente = por_numero.get(numero)
                if existente:
                    if existente.status == 'paga':
                        # Segurança: não reabre baixada mesmo se número >= início
                        ignoradas_baixadas += 1
                        continue
                    for k, v in vals.items():
                        setattr(existente, k, v)
                    existente.save()
                    atualizadas += 1
                else:
                    ParcelaEmprestimo.objects.create(emprestimo=emp, numero=numero, **vals)
                    criadas += 1

    if (dados.get('banco') or '').lower() == 'daycoval':
        ancora = (
            ultima_paga_venc.strftime('%d/%m/%Y')
            if ultima_paga_venc
            else 'nenhuma'
        )
        messages.success(
            request,
            f'Parcelas Daycoval atualizadas após vencimento {ancora} '
            f'(inclui PARCELA e RESIDUAL VRG). '
            f'{atualizadas} atualizada(s), {criadas} criada(s). '
            f'Baixadas preservadas: {ignoradas_baixadas}.',
        )
    else:
        messages.success(
            request,
            f'Parcelas em aberto atualizadas a partir da nº {inicio_aberto} '
            f'(última baixada no sistema: {ultima_paga or "nenhuma"}). '
            f'{atualizadas} atualizada(s), {criadas} criada(s). '
            f'Baixadas preservadas: {ignoradas_baixadas}.',
        )
    if restauradas_status:
        if (dados.get('banco') or '').lower() == 'daycoval':
            messages.warning(
                request,
                f'{restauradas_status} parcela(s) com vencimento já quitado '
                f'voltaram para status paga.',
            )
        else:
            messages.warning(
                request,
                f'{restauradas_status} parcela(s) que tinham sido reabertas por engano '
                f'voltaram para status paga (nº < {inicio_aberto}).',
            )
    if dados.get('aviso'):
        messages.warning(request, dados['aviso'])
    return redirect('emprestimos:detalhe', pk=emp.pk)


def _metodo_flags(emp):
    is_sac = bool(
        (emp.indicador_id and emp.indicador.tipo == 'sac')
        or 'sac' in (emp.indicador_calculo or '').lower()
        or (emp.indice_correcao or '').upper() == 'SELIC'
    )
    if (
        (emp.indicador_id and emp.indicador.tipo == 'sac')
        or 'sac' in (emp.indicador_calculo or '').lower()
    ):
        metodo = 'sac'
    elif (
        (emp.indicador_id and emp.indicador.tipo == 'price')
        or 'price' in (emp.indicador_calculo or '').lower()
        or 'pric' in (emp.indicador_calculo or '').lower()
    ):
        metodo = 'price'
    else:
        metodo = 'outro'
    return is_sac, metodo


def _data_ref_quitacao(emp, pagas, hoje):
    """Início do CDI/SELIC na quitação: último vencimento (ou pagamento) das pagas."""
    if pagas:
        datas_venc = [p.data_vencimento for p in pagas if p.data_vencimento]
        if datas_venc:
            return max(datas_venc)
        datas_pag = [p.data_pagamento for p in pagas if p.data_pagamento]
        if datas_pag:
            return max(datas_pag)
    if emp.data_extrato:
        return emp.data_extrato
    if emp.data_operacao:
        return emp.data_operacao
    return hoje


def _juros_price_pro_rata(saldo: Decimal, taxa_am: Decimal, dias: int) -> Decimal:
    """
    Juros Price entre datas (extrato Sicoob):
      J = saldo × ((1 + i)^(dias/30) − 1)
    i = taxa juros a.m. (sem mora).
    """
    if not saldo or saldo <= 0 or dias <= 0:
        return Decimal('0.00')
    i = (taxa_am or Decimal('0')) / Decimal('100')
    if i <= 0:
        return Decimal('0.00')
    juros = saldo * ((Decimal('1') + i) ** (Decimal(dias) / Decimal('30')) - Decimal('1'))
    return juros.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def _saldo_principal_abertas(parcelas):
    """Soma das amortizações das parcelas em aberto (saldo principal do extrato)."""
    if hasattr(parcelas, 'filter'):
        qs = parcelas.filter(status='aberta')
    else:
        qs = [p for p in parcelas if getattr(p, 'status', None) == 'aberta']
    return sum(
        (p.amortizacao or Decimal('0') for p in qs),
        Decimal('0'),
    ).quantize(Decimal('0.01'))


def _resumo_quitacao_contrato(emp, abertas, pagas, data_quitacao, is_sac: bool, taxa_juros_am: Decimal):
    """
    Valores da quitação TOTAL do contrato (todas as abertas) na data informada.
    Retorna dict com principal, face, quitação e diferença.
    """
    from datetime import date as date_cls

    from .sac_calculo import saldo_quitacao_sac_taxa_fixa

    parcela_carencia = next((p for p in abertas if p.numero == 0), None)
    pago_amort = sum((p.amortizacao or Decimal('0') for p in pagas), Decimal('0'))
    taxa = taxa_juros_am or Decimal('0')

    if (
        is_sac
        and taxa > 0
        and parcela_carencia
        and emp.data_operacao
        and parcela_carencia.data_vencimento
        and data_quitacao
    ):
        saldo, juros, quitacao = saldo_quitacao_sac_taxa_fixa(
            valor_contrato=emp.valor_contrato or Decimal('0'),
            taxa_juros_am=taxa,
            data_operacao=emp.data_operacao,
            data_quitacao=data_quitacao,
            vencimento_p0=parcela_carencia.data_vencimento,
            pago_amort=pago_amort,
        )
        face = sum((p.valor_parcela or Decimal('0') for p in abertas), Decimal('0'))
        dias = max(0, (data_quitacao - emp.data_operacao).days)
        return {
            'valor_principal': saldo.quantize(Decimal('0.01')),
            'valor_parcela_original': face.quantize(Decimal('0.01')),
            'valor_quitacao': quitacao,
            'diferenca': (face - quitacao).quantize(Decimal('0.01')),
            'juros': juros,
            'dias': dias,
        }

    saldo = _saldo_principal_abertas(abertas)
    if saldo <= 0:
        saldo = max(
            Decimal('0'),
            (emp.valor_contrato or Decimal('0'))
            - sum((p.amortizacao or Decimal('0') for p in pagas), Decimal('0')),
        )
    face = sum((p.valor_parcela or Decimal('0') for p in abertas), Decimal('0'))
    data_ref = _data_ref_quitacao(emp, pagas, data_quitacao or date_cls.today())
    dias = max(0, (data_quitacao - data_ref).days) if data_quitacao and data_ref else 0
    juros = Decimal('0.00')

    if is_sac and saldo > 0 and data_quitacao and data_ref:
        pct_corr = emp.pct_correcao_am or Decimal('100')
        if pct_corr <= 0:
            pct_corr = Decimal('100')
        indice = (emp.indice_correcao or '').upper()
        usa_cdi = (
            'CDI' in indice
            or (pct_corr != Decimal('100') and 'SELIC' not in indice)
            or 'SELIC' not in indice
        )
        if usa_cdi:
            from .cdi import juros_cdi_sobre_saldo
            juros, _info = juros_cdi_sobre_saldo(
                saldo=saldo,
                data_inicio=data_ref,
                data_fim=data_quitacao,
                pct_indice=pct_corr,
                incluir_data_inicio=True,
                taxa_prefixada_am=taxa_juros_am or Decimal('0'),
            )
        else:
            from .selic import juros_selic_sobre_saldo
            juros, _info = juros_selic_sobre_saldo(
                saldo=saldo,
                data_inicio=data_ref,
                data_fim=data_quitacao,
                pct_correcao=pct_corr,
            )
        juros = (juros or Decimal('0')).quantize(Decimal('0.01'))
    else:
        juros = _juros_price_pro_rata(saldo, taxa_juros_am or Decimal('0'), dias)

    quitacao = (saldo + juros).quantize(Decimal('0.01'))
    return {
        'valor_principal': saldo.quantize(Decimal('0.01')),
        'valor_parcela_original': face.quantize(Decimal('0.01')),
        'valor_quitacao': quitacao,
        'diferenca': (face - quitacao).quantize(Decimal('0.01')),
        'juros': juros,
        'dias': dias,
    }


def _valor_presente_price(parcelas, data_quitacao, taxa_am: Decimal) -> Decimal:
    """
    Saldo p/ quitação Tabela Price (Sicoob):
      VP = Σ parcela_k / (1 + i)^(dias_k / 30)
    dias_k = data_vencimento − data_quitação (dias corridos).
    Usa só a taxa de juros a.m. (não mora). Cada termo arredondado a 0,01.
    """
    if not data_quitacao or not parcelas:
        return Decimal('0.00')
    i = (taxa_am or Decimal('0')) / Decimal('100')
    if i < 0:
        i = Decimal('0')
    total = Decimal('0')
    for p in parcelas:
        venc = getattr(p, 'data_vencimento', None)
        if not venc:
            continue
        dias = (venc - data_quitacao).days
        if dias < 0:
            continue
        parc = getattr(p, 'valor_parcela', None) or Decimal('0')
        if parc <= 0:
            continue
        if i == 0 or dias == 0:
            termo = parc
        else:
            fator = (Decimal('1') + i) ** (Decimal(dias) / Decimal('30'))
            termo = parc / fator
        total += termo.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return total.quantize(Decimal('0.01'))


def _processar_simulacao_quitacao(request, emp, abertas, pagas, data_quitacao):
    """Processa POST acao=simular. Retorna (sim_dict|None, selecionadas_list)."""
    from datetime import datetime

    ids = request.POST.getlist('parcelas')
    selecionadas = list(
        emp.parcelas.filter(id__in=ids, status='aberta').order_by('numero')
    )
    if not selecionadas:
        messages.warning(request, 'Selecione ao menos uma parcela em aberto.')
        return None, selecionadas

    from .selic import juros_selic_sobre_saldo

    dados_price = None
    raw_price = (request.POST.get('dados_price_json') or '').strip()
    if raw_price:
        try:
            dados_price = json.loads(raw_price)
        except (json.JSONDecodeError, TypeError, ValueError):
            dados_price = None
    linhas_price = (dados_price or {}).get('linhas') or {}
    origem_raw = ((dados_price or {}).get('origem') or '').strip().lower()
    tem_linhas_sim = bool(linhas_price)
    is_sac, _metodo = _metodo_flags(emp)
    origem_sac_sim = origem_raw == 'sac' or (is_sac and tem_linhas_sim and origem_raw != 'price')
    origem_price = origem_raw == 'price' or (tem_linhas_sim and not is_sac and origem_raw != 'sac')
    if tem_linhas_sim and not origem_raw and not is_sac:
        origem_price = True
    origem_simulacao = bool(tem_linhas_sim or origem_raw in ('price', 'sac'))

    def _dec_price(parcela, campo, fallback):
        item = linhas_price.get(str(parcela.id)) or linhas_price.get(parcela.id)
        if not item:
            return fallback
        try:
            return Decimal(str(item.get(campo))).quantize(Decimal('0.01'))
        except Exception:
            return fallback

    def amort_p(p):
        return _dec_price(p, 'amort', p.amortizacao)

    def juros_p(p):
        return _dec_price(p, 'juros', p.juros)

    def parcela_p(p):
        return _dec_price(p, 'parcela', p.valor_parcela)

    total_parcela = sum((parcela_p(p) for p in selecionadas), Decimal('0'))
    total_amort = sum((amort_p(p) for p in selecionadas), Decimal('0'))
    total_juros = sum((juros_p(p) for p in selecionadas), Decimal('0'))
    restantes = [
        p for p in abertas
        if p.id not in {s.id for s in selecionadas}
    ]
    saldo_restante_amort = sum((amort_p(p) for p in restantes), Decimal('0'))
    saldo_restante_parcela = sum((parcela_p(p) for p in restantes), Decimal('0'))

    parcela_ref = selecionadas[0]
    data_ref_parcela = parcela_ref.data_vencimento

    pagas_antes = [
        p for p in pagas
        if p.data_pagamento and p.data_pagamento <= data_ref_parcela
    ]
    if pagas_antes:
        data_ref = max(p.data_pagamento for p in pagas_antes)
    elif emp.data_extrato:
        data_ref = emp.data_extrato
    elif emp.data_operacao:
        data_ref = emp.data_operacao
    else:
        from django.utils import timezone as dj_tz
        data_ref = dj_tz.localdate()

    data_fim_juros = data_quitacao

    raw_selic = (request.POST.get('selic_periodo_pct') or '').strip().replace(',', '.')
    taxa_manual = None
    if raw_selic:
        try:
            taxa_manual = Decimal(raw_selic)
        except Exception:
            messages.warning(request, 'Taxa SELIC manual inválida — usando consulta BCB.')

    dias_juros = max(0, (data_fim_juros - data_ref).days)
    taxa = _taxa_juros_am_efetiva(emp, request)
    juros_pro_rata = Decimal('0.00')
    juros_selic = Decimal('0.00')
    juros_aa = Decimal('0.00')
    selic_info = {}
    juros_parcela_ref = Decimal('0.00')

    if is_sac:
        # Saldo p/ quitação: CDI/SELIC até a DATA DE QUITAÇÃO (não até o venc. da 1ª parcela)
        from datetime import date as date_cls

        data_fim_selic = data_quitacao
        pct_corr = emp.pct_correcao_am or Decimal('100')
        if pct_corr <= 0:
            pct_corr = Decimal('100')
        indice = (emp.indice_correcao or '').upper()
        usa_cdi = (
            'CDI' in indice
            or (pct_corr != Decimal('100') and 'SELIC' not in indice)
            or (is_sac and 'SELIC' not in indice)
        )
        # Início: último vencimento das pagas (banco inclui esse dia no CDI da quitação)
        if pagas:
            ultima_paga = max(
                pagas,
                key=lambda p: (
                    p.data_vencimento or p.data_pagamento or date_cls.min,
                    p.numero,
                ),
            )
            data_ref = (
                ultima_paga.data_vencimento
                or ultima_paga.data_pagamento
                or data_ref
            )
        if usa_cdi:
            from .cdi import juros_cdi_sobre_saldo
            juros_selic, selic_info = juros_cdi_sobre_saldo(
                saldo=total_amort,
                data_inicio=data_ref,
                data_fim=data_fim_selic,
                pct_indice=pct_corr,
                taxa_manual_periodo_pct=taxa_manual,
                incluir_data_inicio=True,
                taxa_prefixada_am=emp.taxa_juros_am or Decimal('0'),
            )
            selic_info['indice'] = 'CDI'
        else:
            juros_selic, selic_info = juros_selic_sobre_saldo(
                saldo=total_amort,
                data_inicio=data_ref,
                data_fim=data_fim_selic,
                pct_correcao=pct_corr,
                taxa_manual_periodo_pct=taxa_manual,
            )
            selic_info['indice'] = 'SELIC'
        # Taxa prefixada a.m. já entra no fator (SACD). Não somar taxa a.a. de novo.
        juros_aa = Decimal('0.00')
        juros_parcela_ref = (juros_selic or Decimal('0')).quantize(Decimal('0.01'))
        juros_pro_rata = juros_parcela_ref
        dias_juros = max(0, (data_fim_selic - data_ref).days)
        data_fim_juros = data_fim_selic
    elif origem_price or not is_sac:
        # Tabela Price: saldo p/ quitação = principal + juros compostos até a data
        # (Saldo p/ Quitação do extrato Sicoob / demonstrativo de saldo devedor).
        taxa_juros_only = emp.taxa_juros_am or Decimal('0')
        if request is not None and request.method == 'POST':
            j_ov = _parse_taxa_post(request.POST.get('taxa_juros_am'))
            if j_ov is not None:
                taxa_juros_only = j_ov
        saldo_price = max(
            Decimal('0'),
            (emp.valor_contrato or Decimal('0'))
            - sum((p.amortizacao or Decimal('0') for p in pagas), Decimal('0')),
        )
        # Seleção parcial: base = soma das amortizações marcadas
        todas_abertas = len(selecionadas) == len(abertas) and len(abertas) > 0
        if not todas_abertas:
            saldo_price = total_amort
        if pagas:
            from datetime import date as date_cls
            ultima_paga = max(
                pagas,
                key=lambda p: (
                    p.data_vencimento or p.data_pagamento or date_cls.min,
                    p.numero,
                ),
            )
            data_ref = (
                ultima_paga.data_vencimento
                or ultima_paga.data_pagamento
                or data_ref
            )
        dias_juros = max(0, (data_quitacao - data_ref).days) if data_ref else 0
        data_fim_juros = data_quitacao
        juros_pro_rata = _juros_price_pro_rata(saldo_price, taxa_juros_only, dias_juros)
        total_amort = saldo_price
        taxa = taxa_juros_only
    else:
        juros_pro_rata = (
            total_amort * (taxa / Decimal('100')) * (Decimal(dias_juros) / Decimal('30'))
        ).quantize(Decimal('0.01'))

    valor_quitacao = (total_amort + juros_pro_rata).quantize(Decimal('0.01'))
    valor_parcela_ref = (
        (amort_p(parcela_ref) + juros_parcela_ref).quantize(Decimal('0.01'))
        if is_sac else None
    )

    sim = {
        'qtd': len(selecionadas),
        'total_parcela': total_parcela,
        'total_amortizacao': total_amort,
        'total_juros': total_juros,
        'data_quitacao': data_quitacao,
        'data_referencia': data_ref,
        'data_fim_juros': data_fim_juros,
        'dias_juros': dias_juros,
        'juros_pro_rata': juros_pro_rata,
        'juros_selic': juros_selic,
        'juros_aa': juros_aa,
        'juros_parcela_ref': juros_parcela_ref,
        'parcela_ref_numero': parcela_ref.numero,
        'valor_parcela_ref': valor_parcela_ref,
        'is_sac': is_sac,
        'origem_price': origem_price,
        'origem_sac': origem_sac_sim,
        'origem_simulacao': origem_simulacao,
        'selic_info': selic_info,
        'taxa_juros_am': taxa,
        'taxa_juros_aa': emp.taxa_juros_aa or Decimal('0'),
        'valor_quitacao_estimado': valor_quitacao,
        'parcelas_restantes': len(restantes),
        'saldo_restante_amort': saldo_restante_amort,
        'saldo_restante_parcela': saldo_restante_parcela,
        'numeros': [p.numero for p in selecionadas],
        'selic_periodo_pct': raw_selic,
    }

    if request.POST.get('salvar_simulacao'):
        titulo = (request.POST.get('titulo_simulacao') or '').strip()[:200]
        metodo = 'sac' if is_sac else 'price'
        detalhes = {
            'numeros': sim['numeros'],
            'is_sac': is_sac,
            'origem_price': origem_price,
            'origem_sac': origem_sac_sim,
            'origem_simulacao': 'sac' if origem_sac_sim else ('price' if origem_price else 'extrato'),
            'juros_selic': str(juros_selic),
            'juros_aa': str(juros_aa),
            'selic_info': {
                k: (str(v) if not isinstance(v, (str, int, float, bool, type(None))) else v)
                for k, v in (selic_info or {}).items()
            },
            'taxa_juros_am': str(taxa),
            'valor_parcela_ref': str(valor_parcela_ref) if valor_parcela_ref is not None else None,
            'valor_parcela_extrato': str(emp.valor_parcela_do_extrato()),
            'total_parcela_original': str(total_parcela),
            'diferenca': str((total_parcela - valor_quitacao).quantize(Decimal('0.01'))),
        }
        if origem_simulacao:
            detalhes['linhas_price'] = {
                str(p.id): linhas_price.get(str(p.id)) or linhas_price.get(p.id)
                for p in selecionadas
            }

        # Quitação TOTAL do contrato (todas as abertas) — para o relatório sintético
        taxa_tot = emp.taxa_juros_am or Decimal('0')
        if request is not None and request.method == 'POST':
            j_ov = _parse_taxa_post(request.POST.get('taxa_juros_am'))
            if j_ov is not None:
                taxa_tot = j_ov
        resumo_tot = _resumo_quitacao_contrato(
            emp, abertas, pagas, data_quitacao, is_sac, taxa_tot,
        )
        detalhes['quitacao_contrato'] = {
            'valor_principal': str(resumo_tot['valor_principal']),
            'valor_parcela_original': str(resumo_tot['valor_parcela_original']),
            'valor_quitacao': str(resumo_tot['valor_quitacao']),
            'diferenca': str(resumo_tot['diferenca']),
            'juros': str(resumo_tot.get('juros') or Decimal('0')),
        }

        # Simulação parcial (novas parcelas / novo saldo devedor)
        sim_parcial = None
        raw_sp = (request.POST.get('sim_parcial_json') or '').strip()
        if raw_sp:
            try:
                sim_parcial = json.loads(raw_sp)
            except (json.JSONDecodeError, TypeError, ValueError):
                sim_parcial = None
        parcelas_restantes_salvar = len(restantes)
        if isinstance(sim_parcial, dict) and sim_parcial:
            detalhes['sim_parcial'] = sim_parcial
            try:
                ns = Decimal(str(sim_parcial.get('novo_saldo') or 0)).quantize(Decimal('0.01'))
                if ns > 0:
                    saldo_restante_amort = ns
            except Exception:
                pass
            try:
                nr = int(sim_parcial.get('n_restante') or 0)
                if nr > 0:
                    parcelas_restantes_salvar = nr
            except Exception:
                pass

        SimulacaoQuitacaoEmprestimo.objects.create(
            emprestimo=emp,
            titulo=titulo or (
                f'Quitação {data_quitacao.strftime("%d/%m/%Y")} — '
                f'parcelas {",".join(str(n) for n in sim["numeros"])}'
                + (' (SAC)' if origem_sac_sim or is_sac else (' (Price)' if origem_price else ''))
            ),
            data_quitacao=data_quitacao,
            metodo=metodo,
            indicador_rotulo=emp.indicador_display[:100],
            parcelas_numeros=','.join(str(n) for n in sim['numeros']),
            qtd_parcelas=sim['qtd'],
            total_amortizacao=total_amort,
            total_parcela_original=total_parcela,
            total_juros_extrato=total_juros,
            juros_calculado=juros_pro_rata,
            valor_quitacao=valor_quitacao,
            dias_juros=dias_juros,
            data_referencia=data_ref,
            data_fim_juros=data_fim_juros,
            parcelas_restantes=parcelas_restantes_salvar,
            saldo_restante_amort=saldo_restante_amort,
            detalhes_json=json.dumps(detalhes, ensure_ascii=False),
            criado_por=request.user if request.user.is_authenticated else None,
        )
        messages.success(
            request,
            'Simulação salva. Gere o relatório sintético na lista de simulações.',
        )

    return sim, selecionadas


@login_required
def emprestimo_detalhe(request, pk):
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    from django.utils import timezone as dj_tz

    emp = get_object_or_404(
        Emprestimo.objects.select_related('indicador', 'banco'),
        pk=pk,
        empresa=empresa,
    )
    parcelas = list(emp.parcelas.all())
    abertas = [p for p in parcelas if p.status == 'aberta']
    pagas = [p for p in parcelas if p.status == 'paga']
    hoje = dj_tz.localdate()
    is_sac, metodo = _metodo_flags(emp)

    simulacoes_recentes = emp.simulacoes_quitacao.all()[:8]
    data_ref_quitacao = _data_ref_quitacao(emp, pagas, hoje)

    ultima_parcela = parcelas[-1] if parcelas else None
    sugestao_total_parcelas = None
    sugestao_dia_vencimento = None
    if ultima_parcela and ultima_parcela.data_vencimento:
        sugestao_dia_vencimento = ultima_parcela.data_vencimento.day
        if emp.data_vencimento:
            base = ultima_parcela.data_vencimento
            fim = emp.data_vencimento
            meses = (fim.year - base.year) * 12 + (fim.month - base.month)
            sugestao_total_parcelas = max(ultima_parcela.numero, ultima_parcela.numero + max(0, meses))
        else:
            sugestao_total_parcelas = ultima_parcela.numero

    # Parcela fixa Price = valor atual da lista (última paga / 1ª aberta), não a 1ª histórica
    parcela_fixa_price = Decimal('0')
    for p in reversed(pagas):
        if (p.valor_parcela or Decimal('0')) > 0:
            parcela_fixa_price = p.valor_parcela
            break
    if parcela_fixa_price <= 0:
        for p in abertas:
            if (p.valor_parcela or Decimal('0')) > 0:
                parcela_fixa_price = p.valor_parcela
                break
    if parcela_fixa_price <= 0 and parcelas:
        parcela_fixa_price = parcelas[0].valor_parcela or Decimal('0')

    parcela_carencia = next((p for p in parcelas if p.numero == 0), None)
    primeira_parcela_pagamento = next(
        (p for p in abertas if (p.amortizacao or Decimal('0')) > 0),
        None,
    )
    juros_carencia_p0 = (parcela_carencia.juros or Decimal('0')) if parcela_carencia else Decimal('0')
    juros_carencia_total = sum(
        (p.juros or Decimal('0')) for p in parcelas
        if p.numero >= 1 and 'encargos mensais' in (p.historico or '').lower()
    )
    juros_carencia_mensais = sum(
        (p.juros or Decimal('0')) for p in abertas
        if p.numero >= 1 and 'encargos mensais' in (p.historico or '').lower()
    )
    carencia_juros_mensais = juros_carencia_total > 0
    juros_carencia_incorporados = juros_carencia_total
    if (
        parcela_carencia
        and juros_carencia_p0 <= 0
        and juros_carencia_mensais > 0
        and 'incorporados' in (parcela_carencia.historico or '').lower()
    ):
        juros_carencia_p0 = juros_carencia_mensais
    total_juros_aberto = sum((p.juros for p in abertas), Decimal('0'))
    total_juros_futuros = sum(
        (p.juros or Decimal('0')) for p in abertas if (p.amortizacao or Decimal('0')) > 0
    )
    if total_juros_futuros < 0:
        total_juros_futuros = Decimal('0')

    saldo_principal = sum((p.amortizacao for p in abertas), Decimal('0'))
    saldo_principal_total = saldo_principal
    if carencia_juros_mensais and juros_carencia_total > 0:
        saldo_principal_total = (saldo_principal + juros_carencia_total).quantize(Decimal('0.01'))

    return render(request, 'emprestimos/detalhe.html', {
        'title': f'Empréstimo {emp.numero_contrato}',
        'emprestimo': emp,
        'parcelas': parcelas,
        'abertas': abertas,
        'pagas': pagas,
        'total_aberto': sum((p.valor_parcela for p in abertas), Decimal('0')),
        'amortizacao_paga': sum((p.amortizacao for p in pagas), Decimal('0')),
        'saldo_principal': saldo_principal,
        'saldo_principal_total': saldo_principal_total,
        'juros_carencia_total': juros_carencia_total,
        'total_juros_aberto': total_juros_aberto,
        'juros_carencia_p0': juros_carencia_p0,
        'juros_carencia_mensais': juros_carencia_mensais,
        'juros_carencia_incorporados': juros_carencia_incorporados,
        'carencia_juros_mensais': carencia_juros_mensais,
        'total_juros_futuros': total_juros_futuros,
        'simulacao': None,
        'selecionadas_ids': set(),
        'data_quitacao': hoje,
        'data_ref_quitacao': data_ref_quitacao,
        'taxa_quitacao_am': _taxa_juros_am_efetiva(emp),
        'juros_por_parcela': {},
        'selic_periodo_pct': '',
        'is_sac_emprestimo': is_sac,
        'metodo_calculo': metodo,
        'metodo_calculo_rotulo': emp.indicador_display,
        'simulacoes_recentes': simulacoes_recentes,
        'ultima_parcela': ultima_parcela,
        'sugestao_total_parcelas': sugestao_total_parcelas,
        'sugestao_dia_vencimento': sugestao_dia_vencimento,
        'parcela_fixa_price': parcela_fixa_price,
        'parcela_carencia': parcela_carencia,
        'primeira_parcela_pagamento': primeira_parcela_pagamento,
        'data_ref_cobranca': hoje,
    })


@login_required
def emprestimo_quitacao(request, pk):
    """Página de simulação de quitação (substitui o modal)."""
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    from datetime import datetime
    from django.utils import timezone as dj_tz

    emp = get_object_or_404(
        Emprestimo.objects.select_related('indicador', 'banco'),
        pk=pk,
        empresa=empresa,
    )
    parcelas = list(emp.parcelas.all())
    abertas = [p for p in parcelas if p.status == 'aberta']
    pagas = [p for p in parcelas if p.status == 'paga']
    hoje = dj_tz.localdate()
    is_sac, metodo = _metodo_flags(emp)
    data_ref_quitacao = _data_ref_quitacao(emp, pagas, hoje)
    parcela_carencia = next((p for p in parcelas if p.numero == 0), None)
    juros_carencia_total = sum(
        (p.juros or Decimal('0')) for p in parcelas
        if p.numero >= 1 and 'encargos mensais' in (p.historico or '').lower()
    )
    carencia_juros_mensais = juros_carencia_total > 0
    if not parcela_carencia and carencia_juros_mensais:
        parcela_carencia = next(
            (p for p in reversed(parcelas) if 'encargos mensais' in (p.historico or '').lower()),
            None,
        )
    saldo_pos_carencia = None
    if (
        parcela_carencia
        and emp.data_operacao
        and parcela_carencia.data_vencimento
        and (emp.taxa_juros_am or Decimal('0')) > 0
    ):
        from .sac_calculo import _juros_carencia_divididos

        pago_amort = sum((p.amortizacao or Decimal('0') for p in pagas), Decimal('0'))
        principal_base = max(
            Decimal('0'),
            (emp.valor_contrato or Decimal('0')) - pago_amort,
        )
        _jp0, _jc, saldo_pos_carencia = _juros_carencia_divididos(
            saldo_inicial=principal_base,
            taxa_juros_am=emp.taxa_juros_am,
            data_inicio=emp.data_operacao,
            vencimento_p0=parcela_carencia.data_vencimento,
        )

    session_key = f'quitacao_sim_{pk}'
    session_key_legacy = f'quitacao_price_{pk}'
    data_quitacao = hoje
    sim = None
    selecionadas = []
    dados_price_json = ''
    origem_simulacao = ''
    origem_price = False
    origem_sac = False
    selecionadas_ids = set()

    # Entrada vinda da simulação (Price ou SAC)
    if request.method == 'POST' and request.POST.get('acao') in ('entrar_simulacao', 'entrar_price'):
        raw_price = (request.POST.get('dados_price_json') or '').strip()
        if raw_price:
            try:
                dados = json.loads(raw_price)
                origem = (dados or {}).get('origem') or 'price'
                if dados and (origem in ('price', 'sac') or dados.get('linhas')):
                    if origem not in ('price', 'sac'):
                        dados['origem'] = 'price'
                        raw_price = json.dumps(dados, ensure_ascii=False)
                    request.session[session_key] = raw_price
                    request.session.pop(session_key_legacy, None)
            except (json.JSONDecodeError, TypeError, ValueError):
                messages.warning(request, 'Não foi possível carregar os dados da simulação.')
        return redirect('emprestimos:quitacao', pk=pk)

    # Abrir com valores do extrato (limpa simulação da sessão)
    if request.method == 'GET' and request.GET.get('origem') == 'extrato':
        request.session.pop(session_key, None)
        request.session.pop(session_key_legacy, None)

    if request.method == 'POST' and request.POST.get('acao') == 'simular':
        raw_data = (request.POST.get('data_quitacao') or '').strip()
        if raw_data:
            try:
                data_quitacao = datetime.strptime(raw_data, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, 'Data de quitação inválida.')
                data_quitacao = hoje
        raw_price = (request.POST.get('dados_price_json') or '').strip()
        if raw_price:
            request.session[session_key] = raw_price
            dados_price_json = raw_price
        else:
            request.session.pop(session_key, None)
            request.session.pop(session_key_legacy, None)
        sim, selecionadas = _processar_simulacao_quitacao(
            request, emp, abertas, pagas, data_quitacao
        )
        selecionadas_ids = {p.id for p in selecionadas}
    else:
        dados_price_json = (
            (request.session.get(session_key) or '').strip()
            or (request.session.get(session_key_legacy) or '').strip()
        )

    if dados_price_json:
        try:
            d = json.loads(dados_price_json)
            origem_simulacao = (d or {}).get('origem') or ''
            if origem_simulacao not in ('price', 'sac') and d.get('linhas'):
                origem_simulacao = 'sac' if is_sac else 'price'
            origem_price = origem_simulacao == 'price'
            origem_sac = origem_simulacao == 'sac'
            if (origem_price or origem_sac) and not selecionadas_ids:
                selecionadas_ids = {
                    int(x) for x in (d.get('selecionadas') or []) if str(x).isdigit()
                }
        except (json.JSONDecodeError, TypeError, ValueError):
            dados_price_json = ''
            origem_simulacao = ''
            origem_price = False
            origem_sac = False

    return render(request, 'emprestimos/quitacao.html', {
        'title': f'Quitação — {emp.numero_contrato}',
        'emprestimo': emp,
        'parcelas': parcelas,
        'abertas': abertas,
        'pagas': pagas,
        'total_aberto': sum((p.valor_parcela for p in abertas), Decimal('0')),
        'amortizacao_paga': sum((p.amortizacao for p in pagas), Decimal('0')),
        'saldo_principal': sum((p.amortizacao or Decimal('0') for p in abertas), Decimal('0')),
        'saldo_principal_total': (
            sum((p.amortizacao or Decimal('0') for p in abertas), Decimal('0'))
            + (juros_carencia_total if carencia_juros_mensais else Decimal('0'))
        ).quantize(Decimal('0.01')),
        'juros_carencia_total': juros_carencia_total,
        'carencia_juros_mensais': carencia_juros_mensais,
        'total_juros_aberto': sum((p.juros for p in abertas), Decimal('0')),
        'simulacao': sim,
        'selecionadas_ids': selecionadas_ids,
        'data_quitacao': data_quitacao,
        'data_ref_quitacao': data_ref_quitacao,
        'taxa_quitacao_am': _taxa_juros_am_efetiva(emp, request if request.method == 'POST' else None),
        'taxa_quitacao_aa': emp.taxa_juros_aa or Decimal('0'),
        'selic_periodo_pct': (request.POST.get('selic_periodo_pct') if request.method == 'POST' else '') or '',
        'is_sac_emprestimo': is_sac,
        'metodo_calculo': metodo,
        'metodo_calculo_rotulo': emp.indicador_display,
        'dados_price_json': dados_price_json,
        'origem_simulacao': origem_simulacao,
        'origem_price': origem_price,
        'origem_sac': origem_sac,
        'parcela_carencia': parcela_carencia,
        'saldo_pos_carencia': saldo_pos_carencia,
    })


@login_required
def emprestimo_simulacoes_list(request):
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    q = (request.GET.get('q') or '').strip()
    qs = (
        SimulacaoQuitacaoEmprestimo.objects.filter(emprestimo__empresa=empresa)
        .select_related('emprestimo', 'emprestimo__banco', 'criado_por')
        .prefetch_related('emprestimo__parcelas')
        .order_by('-criado_em')
    )
    if q:
        qs = qs.filter(
            Q(titulo__icontains=q)
            | Q(emprestimo__numero_contrato__icontains=q)
            | Q(emprestimo__cliente__icontains=q)
            | Q(parcelas_numeros__icontains=q)
            | Q(indicador_rotulo__icontains=q)
        )

    simulacoes = list(qs[:200])
    tot_principal = Decimal('0')
    tot_juros = Decimal('0')
    tot_quitacao = Decimal('0')
    tot_diferenca = Decimal('0')
    tot_principal_contrato = Decimal('0')
    tot_juros_contrato = Decimal('0')
    tot_quitacao_contrato = Decimal('0')
    tot_diferenca_contrato = Decimal('0')
    for s in simulacoes:
        tot_principal += s.total_amortizacao or Decimal('0')
        tot_juros += s.juros_quitacao or Decimal('0')
        tot_quitacao += s.valor_quitacao or Decimal('0')
        tot_diferenca += s.diferenca or Decimal('0')
        tot_principal_contrato += s.principal_contrato or Decimal('0')
        tot_juros_contrato += s.juros_contrato or Decimal('0')
        tot_quitacao_contrato += s.quitacao_contrato or Decimal('0')
        tot_diferenca_contrato += s.diferenca_contrato or Decimal('0')

    return render(request, 'emprestimos/simulacoes_listar.html', {
        'title': 'Simulações de quitação',
        'empresa': empresa,
        'simulacoes': simulacoes,
        'q': q,
        'tot_principal': tot_principal,
        'tot_juros': tot_juros,
        'tot_quitacao': tot_quitacao,
        'tot_diferenca': tot_diferenca,
        'tot_principal_contrato': tot_principal_contrato,
        'tot_juros_contrato': tot_juros_contrato,
        'tot_quitacao_contrato': tot_quitacao_contrato,
        'tot_diferenca_contrato': tot_diferenca_contrato,
    })


@login_required
@require_POST
def emprestimo_simulacao_excluir(request, pk):
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')
    sim = get_object_or_404(
        SimulacaoQuitacaoEmprestimo,
        pk=pk,
        emprestimo__empresa=empresa,
    )
    emp_pk = sim.emprestimo_id
    sim.delete()
    messages.success(request, 'Simulação excluída.')
    nxt = request.POST.get('next') or ''
    if nxt == 'detalhe':
        return redirect('emprestimos:detalhe', pk=emp_pk)
    return redirect('emprestimos:simulacoes')


@login_required
@require_GET
def emprestimo_simulacoes_relatorio_sintetico(request):
    """
    Relatório sintético Excel das simulações salvas:
    Banco | Contrato | Valor principal | Parcela original | Quitação total | Diferença
    """
    from io import BytesIO

    from django.http import HttpResponse
    from django.utils import timezone as dj_tz
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    ids_raw = (request.GET.get('ids') or '').strip()
    exportar_todas = (request.GET.get('todas') or '').strip() in ('1', 'true', 'yes')
    qs = (
        SimulacaoQuitacaoEmprestimo.objects.filter(emprestimo__empresa=empresa)
        .select_related('emprestimo', 'emprestimo__banco')
        .order_by('-criado_em')
    )
    try:
        ids = [int(x) for x in ids_raw.split(',') if x.strip().isdigit()]
    except ValueError:
        ids = []

    if ids:
        qs = qs.filter(pk__in=ids)
        by_id = {s.pk: s for s in qs}
        sims = [by_id[i] for i in ids if i in by_id]
    elif exportar_todas:
        sims = list(qs[:500])
    else:
        messages.warning(
            request,
            'Selecione ao menos uma simulação na lista para gerar o relatório sintético.',
        )
        return redirect('emprestimos:simulacoes')

    if not sims:
        messages.warning(request, 'Nenhuma simulação encontrada para o relatório.')
        return redirect('emprestimos:simulacoes')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sintético'

    header_fill = PatternFill('solid', fgColor='198754')
    header_font = Font(bold=True, color='FFFFFF')
    group_fill = PatternFill('solid', fgColor='198754')
    group_font = Font(bold=True, color='FFFFFF')
    money_format = '#,##0.00'
    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    center = Alignment(horizontal='center', vertical='center')

    ws['A1'] = 'Relatório sintético — Simulações de quitação'
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:L1')
    ws['A2'] = f'Empresa: {getattr(empresa, "razao", None) or getattr(empresa, "nome_fantasia", None) or empresa}'
    ws['A3'] = f'Gerado em: {dj_tz.localtime().strftime("%d/%m/%Y %H:%M")}'
    ws['A4'] = f'Simulações: {len(sims)}'

    # Linha 5 — agrupadores
    row_grp = 5
    ws.merge_cells(start_row=row_grp, start_column=4, end_row=row_grp, end_column=7)
    c_parc = ws.cell(row=row_grp, column=4, value='parcial')
    c_parc.fill = group_fill
    c_parc.font = group_font
    c_parc.alignment = center
    for col in range(4, 8):
        ws.cell(row=row_grp, column=col).fill = group_fill
        ws.cell(row=row_grp, column=col).border = thin

    ws.merge_cells(start_row=row_grp, start_column=9, end_row=row_grp, end_column=12)
    c_tot = ws.cell(row=row_grp, column=9, value='total')
    c_tot.fill = group_fill
    c_tot.font = group_font
    c_tot.alignment = center
    for col in range(9, 13):
        ws.cell(row=row_grp, column=col).fill = group_fill
        ws.cell(row=row_grp, column=col).border = thin

    ws.merge_cells(start_row=row_grp, start_column=14, end_row=row_grp, end_column=16)
    c_qp = ws.cell(row=row_grp, column=14, value='quando tiver parcial')
    c_qp.fill = group_fill
    c_qp.font = group_font
    c_qp.alignment = center
    for col in (14, 15, 16):
        ws.cell(row=row_grp, column=col).fill = group_fill
        ws.cell(row=row_grp, column=col).border = thin

    # Linha 6 — cabeçalhos
    row0 = 6
    headers = {
        1: 'Banco',
        2: 'Contrato',
        3: 'Parcela atual',
        4: 'Valor principal',
        5: 'Valor parcela original',
        6: 'Quitação total',
        7: 'Diferença',
        8: 'Data quitação',
        9: 'Valor principal',
        10: 'Valor parcela original',
        11: 'Quitação total',
        12: 'Diferença',
        14: 'Valor restante',
        15: 'Valor parcelas',
        16: 'Remanescentes',
    }
    for col, h in headers.items():
        cell = ws.cell(row=row0, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    sums = {
        'p_prin': Decimal('0'),
        'p_face': Decimal('0'),
        'p_quit': Decimal('0'),
        'p_diff': Decimal('0'),
        't_prin': Decimal('0'),
        't_face': Decimal('0'),
        't_quit': Decimal('0'),
        't_diff': Decimal('0'),
        'r_saldo': Decimal('0'),
        'r_parc': Decimal('0'),
    }

    def _money(cell, val, green=False):
        if val is None:
            return
        cell.value = float(val)
        cell.number_format = money_format
        if green:
            cell.font = Font(bold=True, color='198754')

    for i, s in enumerate(sims):
        r = row0 + 1 + i
        emp = s.emprestimo
        banco = ''
        if emp.banco_id and emp.banco:
            banco = emp.banco.nome or ''

        # parcial = valores da simulação (parcelas marcadas)
        p_prin = s.total_amortizacao or Decimal('0')
        p_face = s.parcela_original_efetiva or Decimal('0')
        p_quit = s.valor_quitacao or Decimal('0')
        p_diff = (p_face - p_quit).quantize(Decimal('0.01'))

        # total = quitação do contrato inteiro
        t_prin = s.principal_contrato
        t_face = s.face_contrato
        t_quit = s.quitacao_contrato
        t_diff = s.diferenca_contrato
        # Fallback: recalcula se a simulação foi salva antes deste campo
        if t_prin <= 0 and t_quit <= 0:
            try:
                abertas_emp = list(emp.parcelas.filter(status='aberta'))
                pagas_emp = list(emp.parcelas.filter(status='paga'))
                is_sac_emp = (s.metodo or '') == 'sac'
                taxa_emp = emp.taxa_juros_am or Decimal('0')
                resumo = _resumo_quitacao_contrato(
                    emp, abertas_emp, pagas_emp, s.data_quitacao, is_sac_emp, taxa_emp,
                )
                t_prin = resumo['valor_principal']
                t_face = resumo['valor_parcela_original']
                t_quit = resumo['valor_quitacao']
                t_diff = resumo['diferenca']
            except Exception:
                if not s.sim_parcial:
                    t_prin = p_prin
                    t_face = p_face
                    t_quit = p_quit
                    t_diff = p_diff

        # quando tiver parcial
        r_saldo = s.novo_saldo_devedor if s.sim_parcial else Decimal('0')
        r_parc = s.nova_parcela if s.sim_parcial else Decimal('0')
        r_rem = s.remanescentes_rotulo if s.sim_parcial else '—'

        ws.cell(row=r, column=1, value=banco or '—').border = thin
        ws.cell(row=r, column=2, value=emp.numero_contrato or '').border = thin

        parcela_atual = s.valor_parcela_extrato or emp.valor_parcela_do_extrato()
        _money(ws.cell(row=r, column=3), parcela_atual if parcela_atual else None)
        ws.cell(row=r, column=3).border = thin

        _money(ws.cell(row=r, column=4), p_prin)
        _money(ws.cell(row=r, column=5), p_face)
        _money(ws.cell(row=r, column=6), p_quit, green=True)
        cdiff = ws.cell(row=r, column=7)
        _money(cdiff, p_diff, green=(p_diff >= 0))
        if p_diff < 0:
            cdiff.font = Font(bold=True, color='DC3545')

        ws.cell(
            row=r, column=8,
            value=s.data_quitacao.strftime('%d/%m/%Y') if s.data_quitacao else '',
        ).border = thin

        _money(ws.cell(row=r, column=9), t_prin if t_prin else None)
        _money(ws.cell(row=r, column=10), t_face if t_face else None)
        _money(ws.cell(row=r, column=11), t_quit if t_quit else None, green=bool(t_quit))
        cdiff2 = ws.cell(row=r, column=12)
        if t_face or t_quit:
            _money(cdiff2, t_diff, green=(t_diff >= 0))
            if t_diff < 0:
                cdiff2.font = Font(bold=True, color='DC3545')

        if r_saldo:
            _money(ws.cell(row=r, column=14), r_saldo)
        if r_parc:
            _money(ws.cell(row=r, column=15), r_parc, green=True)
        crem = ws.cell(row=r, column=16, value=r_rem)
        crem.border = thin

        for col in (3, 4, 5, 6, 7, 9, 10, 11, 12, 14, 15):
            ws.cell(row=r, column=col).border = thin

        sums['p_prin'] += p_prin
        sums['p_face'] += p_face
        sums['p_quit'] += p_quit
        sums['p_diff'] += p_diff
        sums['t_prin'] += t_prin
        sums['t_face'] += t_face
        sums['t_quit'] += t_quit
        sums['t_diff'] += t_diff
        sums['r_saldo'] += r_saldo
        sums['r_parc'] += r_parc

    rtot = row0 + 1 + len(sims)
    ws.cell(row=rtot, column=1, value='TOTAL').font = Font(bold=True)
    for col, key, green in (
        (4, 'p_prin', False),
        (5, 'p_face', False),
        (6, 'p_quit', True),
        (7, 'p_diff', True),
        (9, 't_prin', False),
        (10, 't_face', False),
        (11, 't_quit', True),
        (12, 't_diff', True),
        (14, 'r_saldo', False),
        (15, 'r_parc', True),
    ):
        cell = ws.cell(row=rtot, column=col, value=float(sums[key]))
        cell.number_format = money_format
        cell.font = Font(bold=True, color='198754') if green else Font(bold=True)
        cell.border = thin

    widths = {
        'A': 18, 'B': 12, 'C': 14, 'D': 15, 'E': 18, 'F': 14, 'G': 12, 'H': 13,
        'I': 15, 'J': 18, 'K': 14, 'L': 12, 'M': 3, 'N': 14, 'O': 14, 'P': 28,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    nome = f'relatorio_sintetico_quitacao_{dj_tz.localdate().strftime("%Y%m%d")}.xlsx'
    resp = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{nome}"'
    return resp


@login_required
@require_POST
def emprestimo_quitacao_excel(request, pk):
    """Exporta para Excel as parcelas da simulação de quitação (valores da tela/Price)."""
    from io import BytesIO

    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    emp = get_object_or_404(
        Emprestimo.objects.select_related('indicador', 'banco'),
        pk=pk,
        empresa=empresa,
    )

    raw = (request.POST.get('payload') or '').strip()
    try:
        payload = json.loads(raw) if raw else {}
    except (json.JSONDecodeError, TypeError, ValueError):
        messages.error(request, 'Dados inválidos para exportação.')
        return redirect('emprestimos:detalhe', pk=pk)

    linhas = payload.get('linhas') or []
    if not linhas:
        messages.warning(request, 'Nenhuma parcela selecionada para exportar.')
        return redirect('emprestimos:detalhe', pk=pk)

    data_quitacao = (payload.get('data_quitacao') or '').strip()
    titulo = (payload.get('titulo') or '').strip()
    origem = (payload.get('origem') or 'extrato').strip()
    totais = payload.get('totais') or {}
    soma_amort = Decimal('0')
    soma_parcela = Decimal('0')
    soma_juros = Decimal('0')

    def _dec(v, default='0'):
        try:
            if v in (None, '', '—'):
                return Decimal(default)
            return Decimal(str(v)).quantize(Decimal('0.01'))
        except Exception:
            return Decimal(default)

    taxa_am = _taxa_juros_am_efetiva(emp)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Quitação'

    header_fill = PatternFill('solid', fgColor='198754')
    header_font = Font(bold=True, color='FFFFFF')
    label_fill = PatternFill('solid', fgColor='E8F5E9')
    money_format = '#,##0.00'

    ws['A1'] = f'Simulação de quitação — Contrato {emp.numero_contrato}'
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:F1')
    ws['A2'] = emp.cliente or ''
    ws['A3'] = f'Indicador: {emp.indicador_display}'
    ws['A4'] = f'Data pretendida de quitação: {data_quitacao or "—"}'
    ws['A5'] = f'Título: {titulo or "—"}'
    ws['A6'] = (
        'Origem dos valores: Tabela Price (simulação)'
        if origem == 'price'
        else 'Origem dos valores: extrato / lista principal'
    )
    ws['A7'] = 'Taxa de juros a.m.:'
    ws['A7'].font = Font(bold=True)
    ws['B7'] = float(taxa_am)
    ws['B7'].number_format = '0.000000'
    ws['C7'] = '%'
    if emp.taxa_juros_aa:
        ws['D7'] = f'(taxa a.a.: {emp.taxa_juros_aa}%)'

    headers = [
        'Nº',
        'Vencimento',
        'Saldo devedor (início)',
        'Parcela',
        'Amortização',
        'Juros',
    ]
    row0 = 9
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row0, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for i, item in enumerate(linhas):
        r = row0 + 1 + i
        amort = _dec(item.get('amortizacao'))
        parcela = _dec(item.get('parcela'))
        juros = _dec(item.get('juros'))
        try:
            saldo = item.get('saldo_inicio')
            saldo_dec = _dec(saldo) if saldo not in (None, '', '—') else None
        except Exception:
            saldo_dec = None

        soma_amort += amort
        soma_parcela += parcela
        soma_juros += juros

        ws.cell(row=r, column=1, value=item.get('numero'))
        ws.cell(row=r, column=2, value=item.get('vencimento') or '')
        c_saldo = ws.cell(
            row=r, column=3, value=float(saldo_dec) if saldo_dec is not None else None
        )
        if saldo_dec is not None:
            c_saldo.number_format = money_format
        c_parc = ws.cell(row=r, column=4, value=float(parcela))
        c_parc.number_format = money_format
        c_amort = ws.cell(row=r, column=5, value=float(amort))
        c_amort.number_format = money_format
        c_juros = ws.cell(row=r, column=6, value=float(juros))
        c_juros.number_format = money_format

    total_row = row0 + 1 + len(linhas)
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=float(soma_parcela)).number_format = money_format
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=float(soma_amort)).number_format = money_format
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=float(soma_juros)).number_format = money_format
    ws.cell(row=total_row, column=6).font = Font(bold=True)

    # Totais: total parcelas | principal (amort) | juros s/ saldo DEVEDOR | liquidar | diferença
    saldo_principal = _dec(totais.get('saldo_principal'), str(soma_amort))
    saldo_devedor = _dec(totais.get('saldo_devedor'), str(saldo_principal))
    for item in linhas:
        try:
            si = item.get('saldo_inicio')
            if si not in (None, '', '—'):
                saldo_devedor = _dec(si)
                break
        except Exception:
            pass
    juros_quitacao = _dec(totais.get('juros_quitacao'))
    valor_quitar = _dec(totais.get('valor_quitar'))
    if valor_quitar <= 0:
        valor_quitar = (saldo_principal + juros_quitacao).quantize(Decimal('0.01'))
    total_parcelas = _dec(totais.get('total_parcelas'), str(soma_parcela))
    # Diferença = TOTAL DAS PARCELAS − TOTAL DE LIQUIDAR
    diferenca = (total_parcelas - valor_quitar).quantize(Decimal('0.01'))
    dias_juros = int(totais.get('dias_juros') or 0)

    label_row = total_row + 2
    value_row = total_row + 3

    ws.cell(row=label_row, column=1, value='Quitação parcial (parcelas marcadas)')
    ws.cell(row=label_row, column=1).font = Font(bold=True)

    headers_tot = [
        (2, 'total parcelas'),
        (3, 'principal (amort.)'),
        (4, 'juros s/ saldo DEVEDOR'),
        (5, 'Quitação total'),
        (6, 'diferença'),
    ]
    for col, label in headers_tot:
        cell = ws.cell(row=label_row, column=col, value=label)
        cell.font = Font(bold=True)
        cell.fill = label_fill
        cell.alignment = Alignment(horizontal='center')

    c_tot_parc = ws.cell(row=value_row, column=2, value=float(total_parcelas))
    c_tot_parc.number_format = money_format
    c_tot_parc.font = Font(bold=True)

    c_saldo = ws.cell(row=value_row, column=3, value=float(saldo_principal))
    c_saldo.number_format = money_format
    c_saldo.font = Font(bold=True, color='0D6EFD')

    c_juros_q = ws.cell(row=value_row, column=4, value=float(juros_quitacao))
    c_juros_q.number_format = money_format
    c_juros_q.font = Font(bold=True, color='DC3545')

    c_liq = ws.cell(row=value_row, column=5, value=float(valor_quitar))
    c_liq.number_format = money_format
    c_liq.font = Font(bold=True, color='198754')

    c_diff = ws.cell(row=value_row, column=6, value=float(diferenca))
    c_diff.number_format = money_format
    if diferenca >= 0:
        c_diff.font = Font(bold=True, color='198754')
    else:
        c_diff.font = Font(bold=True, color='DC3545')

    ws.cell(row=value_row + 1, column=2, value='Soma das parcelas marcadas (originais)')
    ws.cell(row=value_row + 1, column=2).font = Font(italic=True, size=9, color='666666')

    ws.cell(row=value_row + 1, column=3, value='Soma da amortização marcada')
    ws.cell(row=value_row + 1, column=3).font = Font(italic=True, size=9, color='666666')

    saldo_fmt = f'{saldo_devedor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    ws.cell(
        row=value_row + 1,
        column=4,
        value=(
            f'Juros s/ saldo DEVEDOR R$ {saldo_fmt} até {data_quitacao or "—"}'
            + (f' ({dias_juros} dia{"s" if dias_juros != 1 else ""})' if dias_juros else '')
        ),
    )
    ws.cell(row=value_row + 1, column=4).font = Font(italic=True, size=9, color='666666')

    ws.cell(row=value_row + 1, column=5, value='Quitação total = principal + juros')
    ws.cell(row=value_row + 1, column=5).font = Font(italic=True, size=9, color='666666')

    rotulo_diff = (
        'Economia: TOTAL PARCELAS − QUITAÇÃO TOTAL'
        if diferenca >= 0
        else 'Prejuízo: TOTAL PARCELAS − QUITAÇÃO TOTAL'
    )
    ws.cell(row=value_row + 1, column=6, value=rotulo_diff)
    ws.cell(row=value_row + 1, column=6).font = Font(
        italic=True,
        size=9,
        color='198754' if diferenca >= 0 else 'DC3545',
    )

    # --- Simulação parcial: tabela do restante do saldo devedor ---
    sim_parcial = payload.get('sim_parcial') or {}
    linhas_rest = sim_parcial.get('linhas') or []
    if linhas_rest:
        header_parcial_fill = PatternFill('solid', fgColor='0D6EFD')
        r0 = value_row + 3
        ws.cell(row=r0, column=1, value='Simulação parcial — restante do saldo devedor')
        ws.cell(row=r0, column=1).font = Font(bold=True, size=12, color='0D6EFD')
        ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=6)

        ws.cell(
            row=r0 + 1,
            column=1,
            value=(
                f'Modo: {sim_parcial.get("modo_rotulo") or sim_parcial.get("modo") or "—"}'
                f' · {sim_parcial.get("consolida") or ""}'
                f' · Remanescentes: {sim_parcial.get("n_restante_rotulo") or sim_parcial.get("n_restante") or "—"}'
            ),
        )
        ws.merge_cells(start_row=r0 + 1, start_column=1, end_row=r0 + 1, end_column=6)

        ws.cell(row=r0 + 2, column=1, value='Novo saldo devedor:')
        c_ns = ws.cell(row=r0 + 2, column=2, value=float(_dec(sim_parcial.get('novo_saldo'))))
        c_ns.number_format = money_format
        c_ns.font = Font(bold=True)
        ws.cell(row=r0 + 2, column=3, value=(
            'Nova parcela (SAC):' if (sim_parcial.get('metodo') == 'sac' or payload.get('origem') == 'sac')
            else 'Nova parcela (Price):'
        ))
        c_np = ws.cell(row=r0 + 2, column=4, value=float(_dec(sim_parcial.get('nova_parcela'))))
        c_np.number_format = money_format
        c_np.font = Font(bold=True, color='198754')

        hint = (sim_parcial.get('hint') or '').strip()
        if hint:
            ws.cell(row=r0 + 3, column=1, value=hint)
            ws.cell(row=r0 + 3, column=1).font = Font(italic=True, size=9, color='666666')
            ws.merge_cells(start_row=r0 + 3, start_column=1, end_row=r0 + 3, end_column=6)
            header_rest_row = r0 + 4
        else:
            header_rest_row = r0 + 3

        headers_rest = [
            'Nº',
            'Saldo início',
            'Juros',
            'Amortização',
            'Parcela',
            'Saldo final',
        ]
        for col, h in enumerate(headers_rest, start=1):
            cell = ws.cell(row=header_rest_row, column=col, value=h)
            cell.fill = header_parcial_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        soma_parc_r = Decimal('0')
        soma_amort_r = Decimal('0')
        soma_juros_r = Decimal('0')
        for i, item in enumerate(linhas_rest):
            r = header_rest_row + 1 + i
            amort = _dec(item.get('amortizacao'))
            parcela = _dec(item.get('parcela'))
            juros = _dec(item.get('juros'))
            saldo_i = _dec(item.get('saldo_inicio'))
            saldo_f = _dec(item.get('saldo_fim'))
            soma_parc_r += parcela
            soma_amort_r += amort
            soma_juros_r += juros
            ws.cell(row=r, column=1, value=item.get('numero'))
            for col, val in enumerate(
                (saldo_i, juros, amort, parcela, saldo_f), start=2
            ):
                cell = ws.cell(row=r, column=col, value=float(val))
                cell.number_format = money_format

        tot_r = header_rest_row + 1 + len(linhas_rest)
        ws.cell(row=tot_r, column=1, value='TOTAL').font = Font(bold=True)
        for col, val in enumerate((soma_juros_r, soma_amort_r, soma_parc_r), start=3):
            cell = ws.cell(row=tot_r, column=col, value=float(val))
            cell.number_format = money_format
            cell.font = Font(bold=True)

    widths = [36, 14, 22, 14, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    nome = f'quitacao_{emp.numero_contrato}'
    if data_quitacao:
        nome += f'_{data_quitacao.replace("-", "")}'
    nome += '.xlsx'

    response = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nome}"'
    return response


def _rotulo_status_parcela(parcela, data_ref):
    if parcela.status == 'paga':
        return 'Paga'
    if parcela.status == 'aberta':
        if parcela.situacao_cobranca(data_ref) == 'atrasada':
            return 'Atrasada'
        return 'A vencer'
    return parcela.get_status_display()


@login_required
@require_GET
def emprestimo_parcelas_excel(request, pk):
    """Exporta a tabela de parcelas do contrato para Excel."""
    from io import BytesIO

    from django.http import HttpResponse
    from django.utils import timezone as dj_tz
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    emp = get_object_or_404(
        Emprestimo.objects.select_related('indicador', 'banco'),
        pk=pk,
        empresa=empresa,
    )
    hoje = dj_tz.localdate()
    parcelas = list(emp.parcelas.order_by('numero'))

    ids_raw = (request.GET.get('ids') or '').strip()
    if ids_raw:
        try:
            ids = {int(x) for x in ids_raw.split(',') if x.strip().isdigit()}
        except ValueError:
            ids = set()
        if ids:
            parcelas = [p for p in parcelas if p.id in ids]

    if not parcelas:
        messages.warning(request, 'Nenhuma parcela para exportar.')
        return redirect('emprestimos:detalhe', pk=pk)

    money_format = '#,##0.00'
    header_fill = PatternFill('solid', fgColor='198754')
    header_font = Font(bold=True, color='FFFFFF')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Parcelas'

    ws['A1'] = f'Parcelas — Contrato {emp.numero_contrato}'
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:M1')
    ws['A2'] = emp.cliente or ''
    ws['A3'] = f'Indicador: {emp.indicador_display}'
    ws['A4'] = f'Valor contrato: R$ {emp.valor_contrato or 0}'
    ws['A5'] = f'Taxa juros a.m.: {emp.taxa_juros_am or 0}%'
    if emp.data_operacao:
        ws['A6'] = f'Data operação: {emp.data_operacao.strftime("%d/%m/%Y")}'
    ws['A7'] = f'Exportado em: {hoje.strftime("%d/%m/%Y")}'

    headers = [
        'Nº',
        'Vencimento',
        'Parcela',
        'Amortização',
        'Juros',
        'Taxa juros a.m.',
        'Pagamento',
        'Valor pago',
        'Multa',
        'Mora',
        'Correção',
        'Status',
        'Histórico',
    ]
    row0 = 9
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row0, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    tot_parcela = Decimal('0')
    tot_amort = Decimal('0')
    tot_juros = Decimal('0')
    tot_pago = Decimal('0')
    tot_multa = Decimal('0')
    tot_mora = Decimal('0')
    tot_correcao = Decimal('0')

    for i, p in enumerate(parcelas):
        r = row0 + 1 + i
        multa = p.multa if (p.multa or Decimal('0')) > 0 else p.multa_atraso_calculada(hoje)
        taxa = p.taxa_juros_efetiva_am(hoje) if p.status == 'aberta' else (emp.taxa_juros_am or Decimal('0'))
        valor_pago = p.valor_pago if p.valor_pago is not None else None

        tot_parcela += p.valor_parcela or Decimal('0')
        tot_amort += p.amortizacao or Decimal('0')
        tot_juros += p.juros or Decimal('0')
        if valor_pago is not None:
            tot_pago += valor_pago
        tot_multa += multa or Decimal('0')
        tot_mora += p.mora or Decimal('0')
        tot_correcao += p.correcao or Decimal('0')

        ws.cell(row=r, column=1, value=p.numero)
        ws.cell(row=r, column=2, value=p.data_vencimento.strftime('%d/%m/%Y') if p.data_vencimento else '')
        for col, val in enumerate(
            (p.valor_parcela, p.amortizacao, p.juros), start=3
        ):
            cell = ws.cell(row=r, column=col, value=float(val or 0))
            cell.number_format = money_format
        ws.cell(row=r, column=6, value=float(taxa or 0))
        ws.cell(row=r, column=6).number_format = '0.0000'
        ws.cell(
            row=r, column=7,
            value=p.data_pagamento.strftime('%d/%m/%Y') if p.data_pagamento else '',
        )
        if valor_pago is not None:
            cell = ws.cell(row=r, column=8, value=float(valor_pago))
            cell.number_format = money_format
        for col, val in enumerate((multa, p.mora, p.correcao), start=9):
            cell = ws.cell(row=r, column=col, value=float(val or 0))
            cell.number_format = money_format
        ws.cell(row=r, column=12, value=_rotulo_status_parcela(p, hoje))
        ws.cell(row=r, column=13, value=(p.historico or '').strip())

    total_row = row0 + 1 + len(parcelas)
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    for col, val in enumerate(
        (tot_parcela, tot_amort, tot_juros, tot_pago, tot_multa, tot_mora, tot_correcao),
        start=3,
    ):
        cell = ws.cell(row=total_row, column=col, value=float(val))
        cell.number_format = money_format
        cell.font = Font(bold=True)

    widths = [6, 12, 14, 14, 14, 14, 12, 14, 12, 12, 12, 12, 28]
    for i, w in enumerate(widths, start=1):
        col_letter = chr(64 + i) if i <= 26 else None
        if col_letter:
            ws.column_dimensions[col_letter].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    contrato_slug = re.sub(r'[^\w.-]+', '_', emp.numero_contrato or str(pk))
    nome = f'parcelas_{contrato_slug}_{hoje.strftime("%Y%m%d")}.xlsx'

    response = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nome}"'
    return response


@login_required
@require_GET
def emprestimo_quitacao_juros_preview(request, pk):
    """
    Prévia do juros de quitação SAC/SACD até a data informada.
    CDI × % índice, e se houver taxa juros a.m., multiplica o fator:
      J = saldo × (fator_CDI × (1+taxa_am)^ (dias/30) − 1)
    Query: saldo, data_inicio, data_fim, pct_indice, taxa_juros_am (opc.)
    """
    from datetime import datetime

    from .cdi import juros_cdi_sobre_saldo
    from .selic import juros_selic_sobre_saldo

    empresa = _empresa_sessao(request)
    if not empresa:
        return JsonResponse({'ok': False, 'erro': 'Selecione uma empresa.'}, status=403)
    emp = get_object_or_404(Emprestimo, pk=pk, empresa=empresa)

    def _dec_q(raw, default='0'):
        t = str(raw or default).strip().replace('%', '')
        if ',' in t:
            t = t.replace('.', '').replace(',', '.')
        try:
            return Decimal(t)
        except Exception:
            return Decimal(default)

    def _parse_iso(s):
        try:
            return datetime.strptime(str(s)[:10], '%Y-%m-%d').date()
        except Exception:
            return None

    saldo = _dec_q(request.GET.get('saldo'), '0')
    data_inicio = _parse_iso(request.GET.get('data_inicio'))
    data_fim = _parse_iso(request.GET.get('data_fim'))
    pct = _dec_q(request.GET.get('pct_indice'), str(emp.pct_correcao_am or '100'))
    if pct <= 0:
        pct = Decimal('100')
    if request.GET.get('taxa_juros_am') not in (None, ''):
        taxa_am = _dec_q(request.GET.get('taxa_juros_am'), '0')
    else:
        taxa_am = emp.taxa_juros_am or Decimal('0')
    if not data_inicio or not data_fim or saldo <= 0:
        return JsonResponse({'ok': False, 'erro': 'Parâmetros inválidos.'}, status=400)

    parcela_carencia = emp.parcelas.filter(numero=0).first()
    if (
        taxa_am > 0
        and parcela_carencia
        and emp.data_operacao
        and parcela_carencia.data_vencimento
        and _metodo_flags(emp)[0]
    ):
        from .sac_calculo import saldo_quitacao_sac_taxa_fixa

        pago_amort = sum(
            (p.amortizacao or Decimal('0') for p in emp.parcelas.filter(status='paga')),
            Decimal('0'),
        )
        principal, juros, total = saldo_quitacao_sac_taxa_fixa(
            valor_contrato=emp.valor_contrato or saldo,
            taxa_juros_am=taxa_am,
            data_operacao=emp.data_operacao,
            data_quitacao=data_fim,
            vencimento_p0=parcela_carencia.data_vencimento,
            pago_amort=pago_amort,
        )
        em_carencia = data_fim < parcela_carencia.data_vencimento
        dias = max(0, (data_fim - emp.data_operacao).days)
        return JsonResponse({
            'ok': True,
            'juros': str(juros),
            'total': str(total),
            'saldo': str(principal),
            'principal': str(principal),
            'dias_corridos': dias,
            'dias_uteis': 0,
            'taxa_juros_am': str(taxa_am),
            'pct_indice': str(pct),
            'indice': 'Taxa fixa',
            'em_carencia': em_carencia,
            'modo': 'carencia_parcial' if em_carencia else 'pos_carencia',
        })

    indice = (emp.indice_correcao or '').upper()
    usa_cdi = 'CDI' in indice or (pct != Decimal('100') and 'SELIC' not in indice) or (
        'SELIC' not in indice and _metodo_flags(emp)[0]
    )
    if usa_cdi:
        juros, info = juros_cdi_sobre_saldo(
            saldo=saldo,
            data_inicio=data_inicio,
            data_fim=data_fim,
            pct_indice=pct,
            incluir_data_inicio=True,
            taxa_prefixada_am=taxa_am,
        )
        indice_nome = 'CDI'
    else:
        juros, info = juros_selic_sobre_saldo(
            saldo=saldo,
            data_inicio=data_inicio,
            data_fim=data_fim,
            pct_correcao=pct,
        )
        indice_nome = 'SELIC'

    return JsonResponse({
        'ok': True,
        'juros': str(juros),
        'total': str((saldo + juros).quantize(Decimal('0.01'))),
        'saldo': str(saldo),
        'dias_uteis': info.get('dias_uteis') or 0,
        'dias_corridos': info.get('dias_corridos') or 0,
        'taxa_juros_am': str(taxa_am),
        'pct_indice': str(pct),
        'indice': indice_nome,
        'mensagem': info.get('mensagem') or '',
        'data_inicio': data_inicio.isoformat(),
        'data_fim': data_fim.isoformat(),
    })


@login_required
@require_GET
def emprestimo_sac_tabela_cdi(request, pk):
    """
    Gera linhas da tabela SAC:
      - Parcelas pagas: amortização, juros e valor do extrato (já quitados)
      - Em aberto: J = saldo × (fator_CDI − 1) × (pct/100)
        Parcela = amortização + juros; Saldo = anterior − amortização
    """
    from datetime import date, datetime

    from .cdi import carregar_cdi_diario, juros_cdi_sobre_saldo

    empresa = _empresa_sessao(request)
    if not empresa:
        return JsonResponse({'ok': False, 'erro': 'Selecione uma empresa.'}, status=403)

    emp = get_object_or_404(
        Emprestimo.objects.select_related('indicador').prefetch_related('parcelas'),
        pk=pk,
        empresa=empresa,
    )

    def _dec_q(raw, default='0'):
        t = str(raw or default).strip().replace('%', '')
        if ',' in t:
            t = t.replace('.', '').replace(',', '.')
        try:
            return Decimal(t)
        except Exception:
            return Decimal(default)

    def _parse_iso(s):
        if not s:
            return None
        try:
            return datetime.strptime(str(s)[:10], '%Y-%m-%d').date()
        except ValueError:
            return None

    pv = _dec_q(request.GET.get('pv'), str(emp.valor_contrato or '0'))
    try:
        n = int(request.GET.get('n') or emp.parcelas.count() or 0)
    except ValueError:
        n = emp.parcelas.count()
    pct = _dec_q(
        request.GET.get('pct_indice'),
        str(emp.pct_correcao_am or '100'),
    )
    if pct <= 0:
        pct = Decimal('100')

    data_inicio = _parse_iso(request.GET.get('data_inicio')) or emp.data_operacao
    if not data_inicio:
        return JsonResponse({'ok': False, 'erro': 'Informe a data de operação.'}, status=400)
    if pv <= 0 or n < 1:
        return JsonResponse({'ok': False, 'erro': 'PV e nº de parcelas inválidos.'}, status=400)

    recalc_parcial = request.GET.get('recalc_parcial') in ('1', 'true', 'True')
    try:
        num_inicio = int(request.GET.get('num_inicio') or 1)
    except ValueError:
        num_inicio = 1
    if num_inicio < 1:
        num_inicio = 1

    parcelas = list(emp.parcelas.order_by('numero'))
    venc_por_n = {p.numero: p for p in parcelas}

    if recalc_parcial:
        import calendar as cal

        ultimo_venc = None
        for k in range(num_inicio, num_inicio + n):
            p = venc_por_n.get(k)
            if p and p.data_vencimento:
                ultimo_venc = p.data_vencimento
        if ultimo_venc is None:
            y, m, d = data_inicio.year, data_inicio.month, data_inicio.day
            m += n
            while m > 12:
                m -= 12
                y += 1
            ultimo_venc = date(y, m, min(d, cal.monthrange(y, m)[1]))

        series = []
        avisos = []
        try:
            series = carregar_cdi_diario(data_inicio, max(ultimo_venc, date.today()))
        except Exception as exc:
            avisos.append(
                f'CDI BCB indisponível: {exc}. Usando estimativa com última taxa quando possível.'
            )

        amort_fixa = (pv / Decimal(n)).quantize(Decimal('0.01'))
        saldo = pv
        linhas = []
        data_ant = data_inicio

        for idx in range(n):
            k = num_inicio + idx
            p = venc_por_n.get(k)
            if p and p.data_vencimento:
                data_venc = p.data_vencimento
            else:
                y, m = data_ant.year, data_ant.month + 1
                if m > 12:
                    m = 1
                    y += 1
                data_venc = date(y, m, min(data_ant.day, cal.monthrange(y, m)[1]))

            saldo_inicio = saldo.quantize(Decimal('0.01'))
            amort = amort_fixa
            if idx == n - 1 or amort > saldo_inicio:
                amort = saldo_inicio
            amort = amort.quantize(Decimal('0.01'))
            juros, det = juros_cdi_sobre_saldo(
                saldo=saldo_inicio,
                data_inicio=data_ant,
                data_fim=data_venc,
                pct_indice=pct,
                series=series if series else None,
            )
            parcela = (amort + juros).quantize(Decimal('0.01'))
            saldo_fim = (saldo_inicio - amort).quantize(Decimal('0.01'))
            if saldo_fim < 0:
                saldo_fim = Decimal('0.00')
            if idx == n - 1:
                saldo_fim = Decimal('0.00')

            linhas.append({
                'numero': k,
                'id': p.id if p else None,
                'status': p.status if p else 'aberta',
                'origem': 'calculado',
                'vencimento': data_venc.isoformat(),
                'dias_uteis': det.get('dias_uteis') or 0,
                'saldo_inicio': str(saldo_inicio),
                'juros': str(juros),
                'amortizacao': str(amort),
                'parcela': str(parcela),
                'saldo_fim': str(saldo_fim),
                'amort_extrato': None,
                'juros_extrato': None,
                'parcela_extrato': None,
                'mensagem': det.get('mensagem') or '',
                'fator': str(det.get('fator') or '1'),
            })
            saldo = saldo_fim
            data_ant = data_venc

        return JsonResponse({
            'ok': True,
            'pv': str(pv),
            'n': n,
            'num_inicio': num_inicio,
            'recalc_parcial': True,
            'pct_indice': str(pct),
            'amort_fixa': str(amort_fixa),
            'data_inicio': data_inicio.isoformat(),
            'indice': emp.indice_correcao or 'CDI',
            'qtd_extrato': 0,
            'avisos': avisos,
            'linhas': linhas,
            'formula': (
                'Simulação parcial: juros = saldo × (fator CDI × pct/100 − 1) '
                'por período entre vencimentos; parcela = amort + juros.'
            ),
        })

    # Carrega CDI uma vez cobrindo todo o horizonte (com margem futura estimada)
    ultimo_venc = None
    for k in range(1, n + 1):
        p = venc_por_n.get(k)
        if p and p.data_vencimento:
            ultimo_venc = p.data_vencimento
    if ultimo_venc is None:
        # estima n meses à frente a partir da operação
        import calendar as cal
        y, m, d = data_inicio.year, data_inicio.month, data_inicio.day
        m += n
        while m > 12:
            m -= 12
            y += 1
        ultimo_venc = date(y, m, min(d, cal.monthrange(y, m)[1]))

    series = []
    avisos = []
    try:
        series = carregar_cdi_diario(data_inicio, max(ultimo_venc, date.today()))
    except Exception as exc:
        avisos.append(f'CDI BCB indisponível: {exc}. Usando estimativa com última taxa quando possível.')

    amort_fixa = (pv / Decimal(n)).quantize(Decimal('0.01'))
    # Prefer amortização constante das parcelas já pagas no extrato
    for p in parcelas:
        if p.status == 'paga' and (p.amortizacao or 0) > 0:
            amort_fixa = (p.amortizacao or Decimal('0')).quantize(Decimal('0.01'))
            break

    saldo = pv
    saldo_principal_abertas = _saldo_principal_abertas(parcelas)
    anchor_aplicado = False
    linhas = []
    data_ant = data_inicio
    qtd_extrato = 0

    for k in range(1, n + 1):
        p = venc_por_n.get(k)
        if p and p.data_vencimento:
            data_venc = p.data_vencimento
        else:
            y, m = data_ant.year, data_ant.month + 1
            if m > 12:
                m = 1
                y += 1
            import calendar as cal
            data_venc = date(y, m, min(data_ant.day, cal.monthrange(y, m)[1]))

        if p and p.status != 'paga' and not anchor_aplicado and saldo_principal_abertas > 0:
            saldo = saldo_principal_abertas
            anchor_aplicado = True

        saldo_inicio = saldo.quantize(Decimal('0.01'))
        origem = 'calculado'
        det = {'dias_uteis': 0, 'mensagem': '', 'fator': Decimal('1')}

        # Parcelas pagas: traz valores já quitados do extrato
        if p and p.status == 'paga':
            amort = (p.amortizacao or Decimal('0')).quantize(Decimal('0.01'))
            juros = (p.juros or Decimal('0')).quantize(Decimal('0.01'))
            if p.valor_pago is not None and p.valor_pago > 0:
                parcela = p.valor_pago.quantize(Decimal('0.01'))
            elif p.valor_parcela is not None and p.valor_parcela > 0:
                parcela = p.valor_parcela.quantize(Decimal('0.01'))
            else:
                parcela = (amort + juros).quantize(Decimal('0.01'))
            # Se amort do extrato for 0 mas parcela/juros existem, deriva amort
            if amort <= 0 and parcela > juros:
                amort = (parcela - juros).quantize(Decimal('0.01'))
            origem = 'extrato'
            qtd_extrato += 1
            det = {
                'dias_uteis': 0,
                'mensagem': 'Valores do extrato (parcela paga/quitada)',
                'fator': Decimal('1'),
            }
        else:
            amort = amort_fixa
            if k == n or amort > saldo_inicio:
                amort = saldo_inicio
            amort = amort.quantize(Decimal('0.01'))
            juros, det = juros_cdi_sobre_saldo(
                saldo=saldo_inicio,
                data_inicio=data_ant,
                data_fim=data_venc,
                pct_indice=pct,
                series=series if series else None,
            )
            parcela = (amort + juros).quantize(Decimal('0.01'))

        saldo_fim = (saldo_inicio - amort).quantize(Decimal('0.01'))
        if saldo_fim < 0:
            saldo_fim = Decimal('0.00')
        if k == n:
            saldo_fim = Decimal('0.00')

        linhas.append({
            'numero': k,
            'id': p.id if p else None,
            'status': p.status if p else 'aberta',
            'origem': origem,
            'vencimento': data_venc.isoformat(),
            'dias_uteis': det.get('dias_uteis') or 0,
            'saldo_inicio': str(saldo_inicio),
            'juros': str(juros),
            'amortizacao': str(amort),
            'parcela': str(parcela),
            'saldo_fim': str(saldo_fim),
            'amort_extrato': str(p.amortizacao) if p else None,
            'juros_extrato': str(p.juros) if p else None,
            'parcela_extrato': str(p.valor_parcela) if p else None,
            'mensagem': det.get('mensagem') or '',
            'fator': str(det.get('fator') or '1'),
        })
        saldo = saldo_fim
        data_ant = data_venc

    if qtd_extrato:
        avisos.insert(
            0,
            f'{qtd_extrato} parcela(s) paga(s) com valores do extrato (já quitadas).',
        )

    return JsonResponse({
        'ok': True,
        'pv': str(pv),
        'n': n,
        'pct_indice': str(pct),
        'amort_fixa': str(amort_fixa),
        'data_inicio': data_inicio.isoformat(),
        'indice': emp.indice_correcao or 'CDI',
        'qtd_extrato': qtd_extrato,
        'avisos': avisos,
        'linhas': linhas,
        'formula': (
            'Pagas: valores do extrato. '
            'Abertas: J = saldo × (fator_CDI − 1) × (pct/100); '
            'parcela = amortização + juros; saldo = anterior − amortização'
        ),
    })


@login_required
@require_POST
def emprestimo_gerar_parcelas_sac(request, pk):
    """
    Gera parcelas em aberto no SAC a partir do último lançamento:
    amortização constante (= último), vencimentos mensais no dia informado.
    """
    import calendar
    from datetime import date

    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    emp = get_object_or_404(
        Emprestimo.objects.select_related('indicador'),
        pk=pk,
        empresa=empresa,
    )
    is_sac, _metodo = _metodo_flags(emp)
    if not is_sac:
        messages.error(request, 'Esta função é apenas para contratos SAC.')
        return redirect('emprestimos:detalhe', pk=pk)

    ultima = emp.parcelas.order_by('-numero').first()
    if not ultima:
        messages.error(request, 'Não há parcelas importadas para usar como referência.')
        return redirect('emprestimos:detalhe', pk=pk)

    try:
        total = int((request.POST.get('total_parcelas') or '').strip())
    except ValueError:
        total = 0
    try:
        dia_venc = int((request.POST.get('dia_vencimento') or '').strip())
    except ValueError:
        dia_venc = ultima.data_vencimento.day if ultima.data_vencimento else 0

    if total <= ultima.numero:
        messages.error(
            request,
            f'Informe o total de parcelas maior que {ultima.numero} (última existente).',
        )
        return redirect('emprestimos:detalhe', pk=pk)
    if dia_venc < 1 or dia_venc > 31:
        messages.error(request, 'Dia de vencimento inválido (1–31).')
        return redirect('emprestimos:detalhe', pk=pk)

    amort = (ultima.amortizacao or Decimal('0')).quantize(Decimal('0.01'))
    if amort <= 0:
        messages.error(request, 'A última parcela não tem amortização válida.')
        return redirect('emprestimos:detalhe', pk=pk)

    # Taxa mensal para prévia de juros (somente juros contratuais)
    taxa_calc = _taxa_juros_am_efetiva(emp)
    if taxa_calc > 0:
        i_am = taxa_calc / Decimal('100')
    elif emp.taxa_juros_aa and emp.taxa_juros_aa > 0:
        i_am = (emp.taxa_juros_aa / Decimal('100')) / Decimal('12')
    else:
        i_am = Decimal('0')

    pago_amort = emp.parcelas.filter(numero__lte=ultima.numero).aggregate(
        t=Sum('amortizacao')
    )['t'] or Decimal('0')
    saldo = (emp.valor_contrato or Decimal('0')) - pago_amort
    if saldo < 0:
        saldo = Decimal('0')

    def _venc_mes(base: date, meses_a_frente: int, dia: int) -> date:
        y = base.year
        m = base.month + meses_a_frente
        while m > 12:
            m -= 12
            y += 1
        while m < 1:
            m += 12
            y -= 1
        ultimo_dia = calendar.monthrange(y, m)[1]
        return date(y, m, min(dia, ultimo_dia))

    existentes = set(emp.parcelas.values_list('numero', flat=True))
    criadas = []
    base_venc = ultima.data_vencimento

    with transaction.atomic():
        for n in range(ultima.numero + 1, total + 1):
            if n in existentes:
                continue
            meses = n - ultima.numero
            venc = _venc_mes(base_venc, meses, dia_venc)
            saldo_inicio = saldo
            juros = (saldo_inicio * i_am).quantize(Decimal('0.01')) if i_am > 0 else Decimal('0.00')
            amort_n = amort
            if amort_n > saldo_inicio and saldo_inicio > 0:
                amort_n = saldo_inicio.quantize(Decimal('0.01'))
            if n == total and saldo_inicio > 0:
                # última: zera residual
                amort_n = saldo_inicio.quantize(Decimal('0.01'))
            valor = (amort_n + juros).quantize(Decimal('0.01'))
            saldo_fim = (saldo_inicio - amort_n).quantize(Decimal('0.01'))
            if saldo_fim < 0:
                saldo_fim = Decimal('0.00')

            p = ParcelaEmprestimo.objects.create(
                emprestimo=emp,
                numero=n,
                data_vencimento=venc,
                valor_parcela=valor,
                amortizacao=amort_n,
                juros=juros,
                data_pagamento=None,
                historico='Gerada SAC (parcelas em aberto)',
                valor_pago=None,
                mora=Decimal('0'),
                iof=Decimal('0'),
                correcao=Decimal('0'),
                status='aberta',
            )
            criadas.append(p)
            saldo = saldo_fim

    if not criadas:
        messages.warning(
            request,
            'Nenhuma parcela nova foi criada (números já existiam até o total informado).',
        )
    else:
        messages.success(
            request,
            f'{len(criadas)} parcela(s) em aberto gerada(s) '
            f'(nº {criadas[0].numero}–{criadas[-1].numero}) com amortização '
            f'constante R$ {amort:,.2f}.'.replace(',', 'X').replace('.', ',').replace('X', '.'),
        )
    return redirect('emprestimos:detalhe', pk=pk)


@login_required
@require_POST
def emprestimo_gerar_cronograma_inicial(request, pk):
    """Gera cronograma inicial (Price + carência) quando o contrato ainda não tem parcelas."""
    from .price_calculo import gerar_cronograma_price_com_carencia

    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    emp = get_object_or_404(Emprestimo.objects.select_related('indicador'), pk=pk, empresa=empresa)
    if emp.parcelas.exists():
        messages.error(request, 'Este contrato já possui parcelas. Use importar PDF ou atualizar valores.')
        return redirect('emprestimos:detalhe', pk=pk)

    is_sac, _metodo = _metodo_flags(emp)
    if is_sac:
        messages.error(request, 'Geração automática disponível apenas para Tabela Price.')
        return redirect('emprestimos:detalhe', pk=pk)

    pv = (emp.valor_contrato or Decimal('0')).quantize(Decimal('0.01'))
    if pv <= 0:
        messages.error(request, 'Informe o valor do contrato antes de gerar o cronograma.')
        return redirect('emprestimos:detalhe', pk=pk)

    if not emp.data_operacao:
        messages.error(request, 'Informe a data da operação antes de gerar o cronograma.')
        return redirect('emprestimos:detalhe', pk=pk)

    taxa = _taxa_juros_am_efetiva(emp)
    if taxa <= 0:
        messages.error(request, 'Informe a taxa de juros (% a.m.) antes de gerar o cronograma.')
        return redirect('emprestimos:detalhe', pk=pk)

    try:
        n_parcelas = int((request.POST.get('n_parcelas') or '').strip())
    except ValueError:
        n_parcelas = 0
    try:
        meses_carencia = int((request.POST.get('meses_carencia') or '0').strip())
    except ValueError:
        meses_carencia = 0
    try:
        dia_venc = int((request.POST.get('dia_vencimento') or '').strip())
    except ValueError:
        dia_venc = emp.data_operacao.day

    if n_parcelas < 1 or n_parcelas > 600:
        messages.error(request, 'Nº de parcelas inválido (1–600).')
        return redirect('emprestimos:detalhe', pk=pk)
    if meses_carencia < 0 or meses_carencia > 120:
        messages.error(request, 'Meses de carência inválido (0–120).')
        return redirect('emprestimos:detalhe', pk=pk)
    if dia_venc < 1 or dia_venc > 31:
        messages.error(request, 'Dia de vencimento inválido (1–31).')
        return redirect('emprestimos:detalhe', pk=pk)

    tipo_carencia = (request.POST.get('tipo_carencia') or 'juros_mensais').strip().lower()
    if tipo_carencia not in ('juros_mensais', 'capitalizar', 'sem'):
        tipo_carencia = 'juros_mensais'

    linhas = gerar_cronograma_price_com_carencia(
        valor_contrato=pv,
        taxa_juros_am=taxa,
        data_operacao=emp.data_operacao,
        n_parcelas=n_parcelas,
        meses_carencia=meses_carencia,
        dia_vencimento=dia_venc,
        tipo_carencia=tipo_carencia,
    )
    if not linhas:
        messages.error(request, 'Não foi possível gerar o cronograma.')
        return redirect('emprestimos:detalhe', pk=pk)

    ultima_venc = linhas[-1]['data_vencimento']
    prazo_dias = max(1, (ultima_venc - emp.data_operacao).days)
    if tipo_carencia == 'juros_mensais' and meses_carencia > 0:
        juros_carencia = sum(
            (p['juros'] for p in linhas if p['numero'] >= 1 and 'encargos mensais' in (p.get('historico') or '')),
            Decimal('0'),
        )
        saldo_devedor = (pv + juros_carencia).quantize(Decimal('0.01'))
    else:
        juros_incorporados = sum(
            (p['juros'] for p in linhas if p['numero'] >= 1 and 'encargos mensais' in (p.get('historico') or '')),
            Decimal('0'),
        )
        saldo_devedor = (pv + juros_incorporados).quantize(Decimal('0.01')) if juros_incorporados > 0 else pv

    with transaction.atomic():
        ParcelaEmprestimo.objects.bulk_create([
            ParcelaEmprestimo(
                emprestimo=emp,
                numero=p['numero'],
                data_vencimento=p['data_vencimento'],
                valor_parcela=p['valor_parcela'],
                amortizacao=p['amortizacao'],
                juros=p['juros'],
                data_pagamento=p['data_pagamento'],
                historico=p['historico'],
                valor_pago=p['valor_pago'],
                mora=p['mora'],
                iof=p['iof'],
                correcao=p['correcao'],
                status=p['status'],
            )
            for p in linhas
        ])
        emp.data_vencimento = ultima_venc
        emp.prazo_dias = prazo_dias
        emp.saldo_devedor_atualizado = saldo_devedor
        emp.save(update_fields=['data_vencimento', 'prazo_dias', 'saldo_devedor_atualizado', 'atualizado_em'])

    qtd_carencia = (
        meses_carencia
        if tipo_carencia == 'juros_mensais' and meses_carencia > 0
        else (1 if tipo_carencia == 'capitalizar' and meses_carencia > 0 else 0)
    )
    rotulo_carencia = {
        'juros_mensais': 'juros mensais',
        'capitalizar': 'capitalizada (parc. 0)',
        'sem': 'sem carência',
    }.get(tipo_carencia, tipo_carencia)
    messages.success(
        request,
        f'Cronograma gerado: {qtd_carencia} mês(es) carência ({rotulo_carencia}) + '
        f'{n_parcelas} parcela(s) Price (taxa {taxa}% a.m.).',
    )
    return redirect('emprestimos:detalhe', pk=pk)


@login_required
@require_POST
def emprestimo_atualizar_parcelas_sac(request, pk):
    """
    Recalcula parcelas EM ABERTO pela regra SAC:
      Amortização = constante (da última paga ou PV÷n)
      Juros = saldo × (fator_CDI − 1) × (% índice / 100)
      Parcela = amortização + juros
      Saldo = anterior − amortização
    Parcelas pagas não são alteradas.
    """
    from datetime import date

    from .cdi import carregar_cdi_diario, juros_cdi_sobre_saldo
    from .sac_calculo import recalcular_sac_taxa_fixa_modelos

    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    emp = get_object_or_404(
        Emprestimo.objects.select_related('indicador').prefetch_related('parcelas'),
        pk=pk,
        empresa=empresa,
    )
    is_sac, _metodo = _metodo_flags(emp)
    if not is_sac:
        messages.error(request, 'Esta função é apenas para contratos SAC.')
        return redirect('emprestimos:detalhe', pk=pk)

    parcelas = list(emp.parcelas.order_by('numero'))
    if not parcelas:
        messages.error(request, 'Não há parcelas neste contrato.')
        return redirect('emprestimos:detalhe', pk=pk)

    pagas = [p for p in parcelas if p.status == 'paga']
    abertas = [p for p in parcelas if p.status == 'aberta']
    if not abertas:
        messages.warning(request, 'Não há parcelas em aberto para atualizar.')
        return redirect('emprestimos:detalhe', pk=pk)

    pv = (emp.valor_contrato or Decimal('0')).quantize(Decimal('0.01'))
    n_total = len(parcelas)
    amort_fixa = (pv / Decimal(n_total)).quantize(Decimal('0.01')) if n_total else Decimal('0')
    for p in reversed(pagas):
        if (p.amortizacao or Decimal('0')) > 0:
            amort_fixa = (p.amortizacao or Decimal('0')).quantize(Decimal('0.01'))
            break

    pago_amort = sum((p.amortizacao or Decimal('0') for p in pagas), Decimal('0'))
    saldo = (pv - pago_amort).quantize(Decimal('0.01'))
    if saldo < 0:
        saldo = Decimal('0')

    if pagas:
        ultima_paga = max(
            pagas,
            key=lambda p: (p.data_vencimento or p.data_pagamento or date.min, p.numero),
        )
        data_ant = ultima_paga.data_vencimento or ultima_paga.data_pagamento or emp.data_operacao
    else:
        data_ant = emp.data_operacao
    if not data_ant:
        messages.error(request, 'Informe a data de operação do contrato.')
        return redirect('emprestimos:detalhe', pk=pk)

    pct = emp.pct_correcao_am or Decimal('192')
    if pct <= 0:
        pct = Decimal('192')

    taxa_am = emp.taxa_juros_am or Decimal('0')
    taxa_mora = emp.taxa_mora_am or Decimal('0')
    if taxa_am > 0:
        from django.utils import timezone as dj_tz

        data_ref = dj_tz.localdate()
        atualizadas = recalcular_sac_taxa_fixa_modelos(
            parcelas=parcelas,
            abertas=abertas,
            pagas=pagas,
            valor_contrato=pv,
            taxa_juros_am=taxa_am,
            taxa_mora_am=taxa_mora,
            data_operacao=emp.data_operacao,
            data_ref=data_ref,
        )
        messages.success(
            request,
            f'{atualizadas} parcela(s) em aberto recalculada(s) pela regra SAC '
            f'(a vencer: juros {taxa_am}% a.m.; atrasada: juros+mora; multa 2% se vencida).',
        )
        return redirect('emprestimos:detalhe', pk=pk)

    ultimo_venc = abertas[-1].data_vencimento or data_ant
    series = []
    try:
        series = carregar_cdi_diario(data_ant, max(ultimo_venc, date.today()))
    except Exception as exc:
        messages.warning(
            request,
            f'CDI BCB indisponível ({exc}). Usando estimativa com a última taxa conhecida.',
        )

    atualizadas = 0
    with transaction.atomic():
        for idx, p in enumerate(abertas):
            if not p.data_vencimento:
                continue
            saldo_inicio = saldo.quantize(Decimal('0.01'))
            amort = amort_fixa
            if amort > saldo_inicio:
                amort = saldo_inicio
            # última aberta (maior número): zera residual
            if idx == len(abertas) - 1 and saldo_inicio > 0:
                amort = saldo_inicio
            amort = amort.quantize(Decimal('0.01'))

            juros, _det = juros_cdi_sobre_saldo(
                saldo=saldo_inicio,
                data_inicio=data_ant,
                data_fim=p.data_vencimento,
                pct_indice=pct,
                series=series if series else None,
                incluir_data_inicio=False,
            )
            valor = (amort + juros).quantize(Decimal('0.01'))
            saldo_fim = (saldo_inicio - amort).quantize(Decimal('0.01'))
            if saldo_fim < 0:
                saldo_fim = Decimal('0.00')

            p.amortizacao = amort
            p.juros = juros
            p.valor_parcela = valor
            hist = (p.historico or '').strip()
            tag = 'Atualizada SAC/CDI'
            if tag not in hist:
                p.historico = f'{hist} | {tag}'.strip(' |') if hist else tag
            p.save(update_fields=[
                'amortizacao', 'juros', 'valor_parcela', 'historico',
            ])
            atualizadas += 1
            saldo = saldo_fim
            data_ant = p.data_vencimento

    messages.success(
        request,
        f'{atualizadas} parcela(s) em aberto atualizada(s) pela regra SAC '
        f'(amort. R$ {amort_fixa} + juros CDI × {pct}% do índice).',
    )
    return redirect('emprestimos:detalhe', pk=pk)


def _parcela_fixa_price(parcelas) -> Decimal:
    """PMT de referência: valor de parcela mais frequente na lista."""
    from collections import Counter

    vals = [
        (p.valor_parcela or Decimal('0')).quantize(Decimal('0.01'))
        for p in parcelas
        if (p.valor_parcela or Decimal('0')) > 0
    ]
    if vals:
        return Counter(vals).most_common(1)[0][0]
    return Decimal('0')


def _recalcular_juros_amort_price(
    emp,
    parcelas,
    *,
    data_ref=None,
) -> int:
    """
    Recalcula juros e amortização pela Tabela Price em todo o cronograma.

    Percorre pagas + abertas desde o valor do contrato para manter saldo coerente.
    """
    from datetime import date

    from django.utils import timezone as dj_tz

    from .taxas_parcela import multa_atraso_parcela, taxa_juros_am_parcela

    if not parcelas:
        return 0

    taxa_juros = emp.taxa_juros_am or Decimal('0')
    taxa_mora = emp.taxa_mora_am or Decimal('0')
    if taxa_juros <= 0:
        return 0

    data_ref = data_ref or dj_tz.localdate()
    parcela_fixa = _parcela_fixa_price(parcelas)
    if parcela_fixa <= 0:
        return 0

    saldo = (emp.valor_contrato or Decimal('0')).quantize(Decimal('0.01'))
    data_ant = emp.data_operacao
    atualizadas = 0

    for idx, p in enumerate(parcelas):
        if not p.data_vencimento:
            continue

        saldo_inicio = saldo.quantize(Decimal('0.01'))
        if p.status == 'paga':
            taxa_p = taxa_juros
        else:
            taxa_p = taxa_juros_am_parcela(
                p,
                taxa_juros_am=taxa_juros,
                taxa_mora_am=taxa_mora,
                data_ref=data_ref,
            )

        dias = max(0, (p.data_vencimento - data_ant).days) if data_ant else 30
        if dias <= 0:
            dias = 30
        i = taxa_p / Decimal('100')
        if i > 0 and dias > 0 and saldo_inicio > 0:
            fator = (Decimal('1') + i) ** (Decimal(dias) / Decimal('30'))
            juros = (saldo_inicio * (fator - Decimal('1'))).quantize(
                Decimal('0.01'), rounding=ROUND_HALF_UP,
            )
        else:
            juros = Decimal('0.00')

        parcela = (p.valor_parcela or Decimal('0')).quantize(Decimal('0.01'))
        if parcela <= 0:
            parcela = parcela_fixa
        elif p.status == 'aberta' and parcela_fixa > 0:
            parcela = parcela_fixa

        amort = (parcela - juros).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if amort < 0:
            amort = Decimal('0.00')
        if amort > saldo_inicio:
            amort = saldo_inicio.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        saldo_fim = (saldo_inicio - amort).quantize(Decimal('0.01'))
        if saldo_fim < 0:
            saldo_fim = Decimal('0.00')

        mudou = (
            p.amortizacao != amort
            or p.juros != juros
            or (p.valor_parcela or Decimal('0')) != parcela
        )
        if mudou:
            p.amortizacao = amort
            p.juros = juros
            if (p.valor_parcela or Decimal('0')) <= 0 or p.status == 'aberta':
                p.valor_parcela = parcela
            if p.status == 'aberta':
                p.multa = multa_atraso_parcela(p, data_ref)
            hist = (p.historico or '').strip()
            tag = 'Atualizada Price'
            if tag not in hist:
                p.historico = f'{hist} | {tag}'.strip(' |') if hist else tag
            campos = ['amortizacao', 'juros', 'historico']
            if (p.valor_parcela or Decimal('0')) == parcela:
                campos.append('valor_parcela')
            if p.status == 'aberta':
                campos.append('multa')
            p.save(update_fields=campos)
            atualizadas += 1

        saldo = saldo_fim
        data_ant = p.data_vencimento

    return atualizadas


@login_required
@require_POST
def emprestimo_atualizar_parcelas_price(request, pk):
    """
    Recalcula juros e amortização pela Tabela Price em todo o cronograma:
      Parcela = valor do extrato (ou PMT fixa se zerada)
      Juros = saldo × ((1 + i)^(dias/30) − 1)
      Amortização = parcela − juros
    """
    from django.utils import timezone as dj_tz

    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    emp = get_object_or_404(
        Emprestimo.objects.select_related('indicador').prefetch_related('parcelas'),
        pk=pk,
        empresa=empresa,
    )
    is_sac, _metodo = _metodo_flags(emp)
    if is_sac:
        messages.error(request, 'Esta função é apenas para contratos Tabela Price.')
        return redirect('emprestimos:detalhe', pk=pk)

    parcelas = list(emp.parcelas.order_by('numero'))
    if not parcelas:
        messages.error(request, 'Não há parcelas neste contrato.')
        return redirect('emprestimos:detalhe', pk=pk)

    if not (emp.taxa_juros_am or Decimal('0')) > 0:
        messages.error(request, 'Informe a Taxa de Juros (% a.m.) do contrato antes de recalcular.')
        return redirect('emprestimos:detalhe', pk=pk)

    with transaction.atomic():
        atualizadas = _recalcular_juros_amort_price(
            emp, parcelas, data_ref=dj_tz.localdate(),
        )

    if atualizadas <= 0:
        messages.warning(request, 'Nenhuma parcela foi atualizada.')
    else:
        messages.success(
            request,
            f'{atualizadas} parcela(s) atualizada(s) pela Tabela Price '
            f'(juros {emp.taxa_juros_am}% a.m.; atrasada: juros+mora).',
        )
    return redirect('emprestimos:detalhe', pk=pk)


@login_required
@require_POST
def emprestimo_atualizar_taxas(request, pk):
    """Salva taxas do contrato (juros, mora, multa, % índice). Juros a.m. no cálculo; mora só se atrasada."""
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')
    emp = get_object_or_404(Emprestimo, pk=pk, empresa=empresa)

    juros = _parse_taxa_post(request.POST.get('taxa_juros_am'))
    mora = _parse_taxa_post(request.POST.get('taxa_mora_am'))
    if juros is None or mora is None:
        messages.error(request, 'Informe Taxa de Juros e Taxa de Mora válidas.')
        return redirect('emprestimos:detalhe', pk=pk)
    if juros < 0 or mora < 0:
        messages.error(request, 'As taxas não podem ser negativas.')
        return redirect('emprestimos:detalhe', pk=pk)

    juros_aa = _parse_taxa_post(request.POST.get('taxa_juros_aa'))
    multa = _parse_taxa_post(request.POST.get('taxa_multa_am'))
    pct_idx = _parse_taxa_post(request.POST.get('pct_correcao_am'))
    pct_atr = _parse_taxa_post(request.POST.get('pct_correcao_atraso_am'))
    for rotulo, val in (
        ('Taxa Juros a.a.', juros_aa),
        ('Taxa Multa', multa),
        ('% do Índice', pct_idx),
        ('% Correção Atraso', pct_atr),
    ):
        if val is None:
            messages.error(request, f'{rotulo}: valor inválido.')
            return redirect('emprestimos:detalhe', pk=pk)
        if val < 0:
            messages.error(request, f'{rotulo}: não pode ser negativo.')
            return redirect('emprestimos:detalhe', pk=pk)

    indice = (request.POST.get('indice_correcao') or '').strip()[:40]
    indice_atraso = (request.POST.get('indice_correcao_atraso') or '').strip()[:40]
    # Se colaram "CDI % Índice: 192,00" no índice, separa nome e percentual
    m_idx = re.search(
        r'^([A-Za-zÀ-ú]+)\s*%\s*[IÍ]ndice\s*:?\s*([\d.,]+)',
        indice,
        flags=re.IGNORECASE,
    )
    if m_idx:
        indice = m_idx.group(1).strip()
        parsed = _parse_taxa_post(m_idx.group(2))
        if parsed is not None and pct_idx == 0:
            pct_idx = parsed
    m_atr = re.search(
        r'^([A-Za-zÀ-ú]+)\s*%\s*Corre[cç][aã]o\s*Atraso\s*:?\s*([\d.,]+)',
        indice_atraso,
        flags=re.IGNORECASE,
    )
    if m_atr:
        indice_atraso = m_atr.group(1).strip()
        parsed = _parse_taxa_post(m_atr.group(2))
        if parsed is not None and pct_atr == 0:
            pct_atr = parsed

    emp.taxa_juros_am = juros.quantize(Decimal('0.0001'))
    emp.taxa_mora_am = mora.quantize(Decimal('0.0001'))
    emp.taxa_juros_aa = juros_aa.quantize(Decimal('0.0001'))
    emp.taxa_multa_am = multa.quantize(Decimal('0.0001'))
    emp.pct_correcao_am = pct_idx.quantize(Decimal('0.0001'))
    emp.pct_correcao_atraso_am = pct_atr.quantize(Decimal('0.0001'))
    emp.indice_correcao = indice
    emp.indice_correcao_atraso = indice_atraso
    emp.save(update_fields=[
        'taxa_juros_am', 'taxa_mora_am', 'taxa_juros_aa', 'taxa_multa_am',
        'pct_correcao_am', 'pct_correcao_atraso_am',
        'indice_correcao', 'indice_correcao_atraso', 'atualizado_em',
    ])
    messages.success(
        request,
        f'Taxas salvas: juros {emp.taxa_juros_am}% a.m.; '
        f'mora {emp.taxa_mora_am}% (só parcelas atrasadas); '
        f'índice {emp.indice_correcao or "—"} {emp.pct_correcao_am}%.',
    )
    return redirect('emprestimos:detalhe', pk=pk)


@login_required
@require_POST
def emprestimo_excluir(request, pk):
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')
    emp = get_object_or_404(Emprestimo, pk=pk, empresa=empresa)
    num = emp.numero_contrato
    emp.delete()
    messages.success(request, f'Empréstimo {num} excluído.')
    return redirect('emprestimos:listar')
