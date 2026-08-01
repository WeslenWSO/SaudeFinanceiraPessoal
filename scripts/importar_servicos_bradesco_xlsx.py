#!/usr/bin/env python
"""
Importa ServicosMedicos a partir da planilha Bradesco (codigo + procedimento).
Importa somente codigos que ainda nao existem no banco.

  set DATABASE_URL=postgresql://...
  python scripts/importar_servicos_bradesco_xlsx.py "C:\\Users\\...\\Bradesco itens.xlsx"
  python scripts/importar_servicos_bradesco_xlsx.py --dry-run
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "SaudeFinanceira.settings")

MAX_CODIGO = 20
MAX_SERVICO = 200
DEFAULT_XLSX = Path(r"c:\Users\wesle\OneDrive\Desktop\Bradesco itens.xlsx")


def _codigo_str(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        n = int(val)
        if n <= 0:
            return None
        return str(n).zfill(8)
    s = re.sub(r"\D", "", str(val).strip())
    if not s or int(s) <= 0:
        return None
    return s.zfill(8) if len(s) <= 8 else s[:8]


def parse_xlsx(caminho: Path) -> dict[str, str]:
    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    registros: dict[str, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or len(row) < 3:
            continue
        codigo = _codigo_str(row[0])
        if not codigo:
            continue
        nome = (row[2] or "").strip()
        if not nome:
            continue
        nome = re.sub(r"\s+", " ", nome)[:MAX_SERVICO]
        registros[codigo] = nome
    wb.close()
    return registros


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("arquivo", nargs="?", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.arquivo.is_file():
        print(f"Arquivo nao encontrado: {args.arquivo}", file=sys.stderr)
        return 1

    servicos = parse_xlsx(args.arquivo)
    print(f"Codigos na planilha: {len(servicos)}")

    if args.dry_run:
        for codigo in sorted(servicos)[:15]:
            print(f"  {codigo}\t{servicos[codigo][:60]}")
        print("  ...")
        return 0

    if not os.environ.get("DATABASE_URL"):
        url = ROOT / "render_db.url"
        if url.is_file():
            os.environ["DATABASE_URL"] = url.read_text(encoding="utf-8").strip()
    if not os.environ.get("DATABASE_URL"):
        print("Defina DATABASE_URL.", file=sys.stderr)
        return 1

    import django

    django.setup()
    from servicos_medicos.models import ServicosMedicos

    existentes = set(ServicosMedicos.objects.values_list("codigo", flat=True))
    criados = pulados = 0

    for codigo in sorted(servicos):
        if codigo in existentes:
            pulados += 1
            continue
        ServicosMedicos.objects.create(codigo=codigo, servicos=servicos[codigo])
        criados += 1

    total = ServicosMedicos.objects.count()
    print(f"Novos: {criados} | ja cadastrados (ignorados): {pulados} | total no banco: {total}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
