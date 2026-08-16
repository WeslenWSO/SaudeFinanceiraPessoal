#!/usr/bin/env python
"""
Cadastra serviços ausentes e importa tabela de preços CASSI (Medicinarte).

Planilha (.xlsx): Código | Descrição | Grupo | Total (R$)

  python scripts/importar_tabela_preco_cassi.py --arquivo "C:\\...\\Medicinarte_Tabela_Dotacoes.xlsx"
  python scripts/importar_tabela_preco_cassi.py --dry-run
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "SaudeFinanceira.settings")

CONVENIO_NOME = "CASSI"
CABECALHO_NOME = "TABELA - CASSI"
EMPRESA_ID_DEFAULT = 16
XLSX_DEFAULT = Path.home() / "OneDrive" / "Desktop" / "Medicinarte_Tabela_Dotacoes.xlsx"

MAX_CODIGO = 20
MAX_SERVICO = 200


def _tuss_para_cbhpm(codigo: str) -> str | None:
    c = re.sub(r"\D", "", codigo or "")
    if len(c) != 8:
        return None
    return f"{c[0]}.{c[1:3]}.{c[3:5]}.{c[5:7]}-{c[7]}"


def _codigo_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return str(int(val)).zfill(8)
    s = re.sub(r"\D", "", str(val).strip())
    return s.zfill(8) if s else ""


def _parse_valor(val) -> Decimal:
    if val is None or val == "":
        raise ValueError("valor vazio")
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    s = str(val).strip().replace("R$", "").strip()
    if not s:
        raise ValueError("valor vazio")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif s.count(".") > 1:
        parts = s.split(".")
        s = "".join(parts[:-1]) + "." + parts[-1]
    return Decimal(s).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _parse_xlsx(caminho: Path) -> list[tuple[str, str, Decimal]]:
    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    registros: list[tuple[str, str, Decimal]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or row[0] is None:
            continue
        codigo = _codigo_str(row[0])
        if len(codigo) != 8:
            continue
        nome = str(row[1] or "").strip()[:MAX_SERVICO] or f"Servico {codigo}"
        valor_col = row[3] if len(row) > 3 else row[2]
        valor = _parse_valor(valor_col)
        registros.append((codigo[:MAX_CODIGO], nome, valor))
    wb.close()
    return registros


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--arquivo", type=Path, default=XLSX_DEFAULT)
    parser.add_argument("--empresa-id", type=int, default=EMPRESA_ID_DEFAULT)
    parser.add_argument("--somente-novos", action="store_true")
    args = parser.parse_args()

    if not args.arquivo.is_file():
        print(f"Arquivo não encontrado: {args.arquivo}", file=sys.stderr)
        return 1

    linhas = _parse_xlsx(args.arquivo)
    if not linhas:
        print("Nenhuma linha válida na planilha.", file=sys.stderr)
        return 1

    print(f"Linhas: {len(linhas)} | Empresa: {args.empresa_id} | Convênio: {CONVENIO_NOME}")

    if args.dry_run:
        for codigo, nome, valor in linhas[:10]:
            print(f"  {codigo}\t{nome[:50]}\t{valor}")
        print("  ...")
        return 0

    if not os.environ.get("DATABASE_URL"):
        url = ROOT / "render_db.url"
        if url.is_file():
            os.environ["DATABASE_URL"] = url.read_text(encoding="utf-8").strip()
    if not os.environ.get("DATABASE_URL"):
        print("Defina DATABASE_URL.", file=sys.stderr)
        return 1

    import django

    django.setup()
    from empresa.models import Empresa
    from servicos_medicos.models import Cabecalho, Convenio, ServicosMedicos, TabelaPreco

    empresa = Empresa.objects.filter(pk=args.empresa_id).first()
    if not empresa:
        print(f"Empresa id={args.empresa_id} não encontrada.", file=sys.stderr)
        return 1

    convenio = Convenio.objects.filter(empresa=empresa, nome=CONVENIO_NOME).first()
    if not convenio:
        alt = Convenio.objects.filter(empresa=empresa, nome__icontains="CASSI").first()
        convenio = alt
    if not convenio:
        convenio, _ = Convenio.objects.get_or_create(empresa=empresa, nome=CONVENIO_NOME)

    cabecalho = Cabecalho.objects.filter(
        empresa=empresa, convenio=convenio, nome_tabela=CABECALHO_NOME
    ).first()
    if not cabecalho:
        cabecalho = Cabecalho.objects.create(
            empresa=empresa, convenio=convenio, nome_tabela=CABECALHO_NOME
        )

    print(f"Convênio: {convenio.nome} (id={convenio.pk})")
    print(f"Cabeçalho: {cabecalho.nome_tabela} (id={cabecalho.pk})")

    servicos_por_codigo = {s.codigo: s for s in ServicosMedicos.objects.all()}
    cbhpm_map = {_tuss_para_cbhpm(c): c for c in servicos_por_codigo if _tuss_para_cbhpm(c)}

    existentes: set[int] = set()
    if args.somente_novos:
        existentes = set(
            TabelaPreco.objects.filter(
                empresa=empresa, convenio=convenio, cabecalho=cabecalho
            ).values_list("codigo_servico_id", flat=True)
        )

    servicos_criados = criados = atualizados = pulados = erros = 0

    for codigo, nome, valor in linhas:
        if codigo not in servicos_por_codigo:
            cbhpm = _tuss_para_cbhpm(codigo)
            alt = cbhpm_map.get(cbhpm) if cbhpm else None
            if not alt:
                obj = ServicosMedicos.objects.create(codigo=codigo, servicos=nome)
                servicos_por_codigo[codigo] = obj
                if cbhpm:
                    cbhpm_map[cbhpm] = codigo
                servicos_criados += 1
            else:
                servicos_por_codigo.setdefault(alt, servicos_por_codigo[alt])

        servico = servicos_por_codigo.get(codigo)
        if not servico:
            cbhpm = _tuss_para_cbhpm(codigo)
            alt = cbhpm_map.get(cbhpm) if cbhpm else None
            if alt:
                servico = servicos_por_codigo[alt]
        if not servico:
            print(f"AVISO: não foi possível resolver serviço {codigo}", file=sys.stderr)
            erros += 1
            continue

        if args.somente_novos and servico.pk in existentes:
            pulados += 1
            continue

        _, created = TabelaPreco.objects.update_or_create(
            empresa=empresa,
            convenio=convenio,
            cabecalho=cabecalho,
            codigo_servico=servico,
            defaults={
                "preco_apartamento": valor,
                "preco_enfermaria": valor,
            },
        )
        if created:
            criados += 1
        else:
            atualizados += 1

    total = TabelaPreco.objects.filter(
        empresa=empresa, convenio=convenio, cabecalho=cabecalho
    ).count()
    print(
        f"ServicosMedicos novos: {servicos_criados} | TabelaPreco criados: {criados} | "
        f"atualizados: {atualizados} | pulados: {pulados} | erros: {erros} | total cabeçalho: {total}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
