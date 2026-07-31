#!/usr/bin/env python
"""Gera tabela_preco_geap.tsv a partir da planilha GEAP (servicos_contrato_*.xlsx)."""
from __future__ import annotations

import re
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

TSV_DEFAULT = Path(__file__).resolve().parent / "dados" / "tabela_preco_geap.tsv"


def _codigo_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return str(int(val)).zfill(8)
    s = re.sub(r"\D", "", str(val).strip())
    return s.zfill(8) if s else ""


def parse_xlsx(caminho: Path) -> list[tuple[str, str, Decimal]]:
    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    registros: list[tuple[str, str, Decimal]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or row[0] is None:
            continue
        codigo = _codigo_str(row[0])
        if len(codigo) != 8:
            continue
        nome = (row[1] or "").strip()[:200]
        if row[2] is None:
            continue
        valor = Decimal(str(row[2])).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        registros.append((codigo, nome, valor))
    wb.close()
    return registros


def main() -> int:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    tsv_path = Path(sys.argv[2]) if len(sys.argv) > 2 else TSV_DEFAULT

    if not xlsx_path or not xlsx_path.is_file():
        print("Uso: python scripts/gerar_tsv_geap_xlsx.py <arquivo.xlsx> [saida.tsv]", file=sys.stderr)
        return 1

    registros = parse_xlsx(xlsx_path)
    lines = ["codigo\tnome\tvalor"]
    for codigo, nome, valor in registros:
        lines.append(f"{codigo}\t{nome}\t{valor}")

    tsv_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Registros: {len(registros)} -> {tsv_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
