#!/usr/bin/env python
"""Gera tabela_preco_bradesco.tsv a partir de TabelaProcedimentosTUSS.xlsx."""
from __future__ import annotations

import re
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

TSV_DEFAULT = Path(__file__).resolve().parent / "dados" / "tabela_preco_bradesco.tsv"
DEFAULT_XLSX = Path(r"c:\Users\wesle\Downloads\TabelaProcedimentosTUSS.xlsx")

MAX_SERVICO = 200


def _codigo_str(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        n = int(val)
        if n <= 0:
            return None
        return str(n).zfill(8)
    s = re.sub(r"\D", "", str(val).strip())
    if not s or int(s) <= 0:
        return None
    return s.zfill(8) if len(s) <= 8 else s[:8]


def parse_xlsx(caminho: Path) -> list[tuple[str, str, str, str]]:
    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    registros: dict[str, tuple[str, str, str]] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or len(row) < 3:
            continue
        codigo = _codigo_str(row[0])
        if not codigo:
            continue
        nome = re.sub(r"\s+", " ", (row[1] or "").strip())[:MAX_SERVICO]
        if row[2] is None:
            continue
        valor = Decimal(str(row[2])).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        vi = ve = str(valor)
        # coluna empresarial opcional (coluna 4, índice 3)
        if len(row) > 3 and row[3] is not None:
            ve = str(Decimal(str(row[3])).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
        registros[codigo] = (nome, vi, ve)
    wb.close()
    return [(c, registros[c][0], registros[c][1], registros[c][2]) for c in sorted(registros)]


def main() -> int:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    tsv = Path(sys.argv[2]) if len(sys.argv) > 2 else TSV_DEFAULT

    if not xlsx.is_file():
        print(f"Arquivo não encontrado: {xlsx}", file=sys.stderr)
        return 1

    rows = parse_xlsx(xlsx)
    lines = ["codigo\tnome\tvalor_individual\tvalor_empresarial"]
    for codigo, nome, vi, ve in rows:
        lines.append(f"{codigo}\t{nome}\t{vi}\t{ve}")

    tsv.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Registros: {len(rows)} -> {tsv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
