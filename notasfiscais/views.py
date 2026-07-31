from typing import Any
import os
import subprocess
import sys
import tempfile
from django.db.models.query import QuerySet
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import UpdateView, CreateView, DeleteView
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum, Count, F, IntegerField, Max
from django.db.models.functions import Cast
from django.views import View
from django.views.generic import FormView
from django.http import JsonResponse
import logging
try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

from .models import NotaFiscalServico, AnexoSimplesNacional, FolhaSalario
from .forms import (
    XMLUploadForm,
    NFSeForm,
    NFSeUpdateForm,
    NFSeRecebimentoForm,
    NFSeSegmentForm,
    PortalNacionalNfseForm,
    PortalExtensaoNfseForm,
)
from .utils import import_nfse_from_xml, extrair_forma_pagamento, extrair_socio, _nome_socio_completo
from typing import Optional


from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from decimal import Decimal
import json
from datetime import datetime, date, timedelta
from calendar import monthrange
from dateutil.relativedelta import relativedelta
from empresa.models import Empresa
from cobranca.models import Cobranca
# views.py
import re
from django.db import transaction
from django.http import JsonResponse, HttpResponseBadRequest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.conf import settings as django_settings
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.utils.dateparse import parse_date

from socio.models import Socio


def serialize_resultado(resultado):
    """Converte o resultado da importação para formato serializável"""
    def convert_decimal(obj):
        """Converte valores Decimal para float recursivamente"""
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, dict):
            return {key: convert_decimal(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [convert_decimal(item) for item in obj]
        else:
            return obj
    
    # Converter todos os valores Decimal para float
    resultado_serializado = convert_decimal(resultado)
    
    # Garantir que os tipos básicos sejam serializáveis
    return {
        'notas_importadas': resultado_serializado.get('notas_importadas', []),
        'notas_ignoradas': resultado_serializado.get('notas_ignoradas', []),
        'notas_canceladas': resultado_serializado.get('notas_canceladas', []),
        'total_processadas': int(resultado_serializado.get('total_processadas', 0)),
        'total_importadas': int(resultado_serializado.get('total_importadas', 0)),
        'total_canceladas': int(resultado_serializado.get('total_canceladas', 0)),
        'total_ignoradas': int(resultado_serializado.get('total_ignoradas', 0))
    }

# Chave de session para persistir filtros da listagem NFSe (limpa ao sair do módulo via middleware)
NFS_FILTROS_SESSION_KEY = 'nfs_filtros'

NFSE_DEFAULT_SORT = 'numero'
NFSE_DEFAULT_SORT_DIR = 'asc'

# Colunas ordenáveis da listagem (chave data-col -> campo ORM)
NFSE_SORT_FIELDS = {
    'numero': 'numero_nota',
    'data_emissao': 'data_emissao',
    'cnpj': 'cnpj_cpf',
    'cliente': 'cliente',
    'socio': 'socio__socio',
    'base_servico': 'base_servico',
    'regra_imposto': 'codigo_da_regra_do_imposto__DescricaoRegraImposto',
    'status_nota': 'data_cancelamento',
    'valor_bruto': 'valor_bruto',
    'iss_retido': 'iss_retido',
    'aliquota_iss': 'aliquota',
    'valor_iss_retido': 'valor_iss_retido',
    'valor_pis': 'valor_pis',
    'valor_cofins': 'valor_cofins',
    'valor_csll': 'valor_csll',
    'valor_irpj': 'valor_ir',
    'valor_outras_ret': 'outras_retencoes',
    'valor_inss': 'valor_inss',
    'valor_liquido': 'valor_liquido',
    'discriminacao': 'discriminacao',
    'forma_pag': 'forma_pagamento__descricao',
    'autorizacao': 'nsu',
    'conciliacao': 'status_conciliacao',
}


def _get_default_dates():
    """Retorna (data_inicio, data_fim) padrão: últimos 12 meses."""
    hoje = timezone.now().date()
    return (hoje - timedelta(days=365), hoje)


def _parse_date(s):
    """Converte string para date. Aceita YYYY-MM-DD. Retorna None se inválido."""
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return None


def _apply_filtro_socio_nfse(queryset, socio_val):
    """
    Filtra NFSe por sócio responsável.
    - vazio: sem filtro
    - 'sem' / 'none': apenas notas sem sócio cadastrado
    - ID numérico: apenas notas daquele sócio (empresa.Socio / socio.Socio)
    """
    if socio_val is None:
        return queryset
    s = str(socio_val).strip()
    if not s:
        return queryset
    if s in ('sem', 'none'):
        return queryset.filter(socio__isnull=True)
    if s.isdigit():
        return queryset.filter(socio_id=int(s))
    return queryset


def _parse_valor_filtro(valor_str):
    """Converte texto de valor (pt-BR ou US) em Decimal, ou None se inválido."""
    if valor_str is None:
        return None
    s = str(valor_str).strip()
    if not s:
        return None
    s = s.replace('R$', '').replace(' ', '').strip()
    if not s:
        return None
    # 1.234,56 (pt-BR) ou 1234,56
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return Decimal(s).quantize(Decimal('0.01'))
    except Exception:
        return None


def _apply_filtro_valor_nfse(queryset, valor_str):
    """Filtra por valor bruto ou líquido igual ao informado."""
    valor = _parse_valor_filtro(valor_str)
    if valor is None:
        return queryset
    return queryset.filter(Q(valor_bruto=valor) | Q(valor_liquido=valor))


def _apply_filtro_search_nfse(queryset, search):
    """Busca em número, cliente, CPF/CNPJ, autorização, série e discriminação."""
    termo = (search or '').strip()
    if not termo:
        return queryset
    return queryset.filter(
        Q(numero_nota__icontains=termo)
        | Q(cliente__icontains=termo)
        | Q(cnpj_cpf__icontains=termo)
        | Q(nsu__icontains=termo)
        | Q(serie__icontains=termo)
        | Q(discriminacao__icontains=termo)
    )


def _get_nfse_sort_from_request(request, filters):
    """
    Lê sort/dir do GET (clique no cabeçalho), persiste na session e retorna valores válidos.
    """
    sort_get = (request.GET.get('sort') or '').strip()
    dir_get = (request.GET.get('dir') or '').strip().lower()
    if sort_get in NFSE_SORT_FIELDS:
        filters['sort'] = sort_get
        filters['dir'] = dir_get if dir_get in ('asc', 'desc') else 'asc'
        request.session[NFS_FILTROS_SESSION_KEY] = filters

    sort_col = filters.get('sort', NFSE_DEFAULT_SORT)
    sort_dir = filters.get('dir', NFSE_DEFAULT_SORT_DIR)
    if sort_col not in NFSE_SORT_FIELDS:
        sort_col = NFSE_DEFAULT_SORT
    if sort_dir not in ('asc', 'desc'):
        sort_dir = NFSE_DEFAULT_SORT_DIR
    return sort_col, sort_dir


def _apply_nfse_sort(queryset, sort_col, sort_dir):
    """Aplica order_by conforme coluna e direção; desempate por número da nota (texto)."""
    field = NFSE_SORT_FIELDS[sort_col]
    prefix = '-' if sort_dir == 'desc' else ''
    order_fields = [f'{prefix}{field}']
    if sort_col != 'numero':
        # numero_nota pode ser "4729-1" (NF segmentada) — não usar Cast para int no Postgres
        order_fields.append('numero_nota')
    return queryset.order_by(*order_fields)


def _normalize_filters(raw_dict, default_paginate_by=20):
    """
    Normaliza filtros: strip em search/status/forma_pagamento; datas em ISO; inverte se início > fim.
    raw_dict: dict com chaves search, status, forma_pagamento, data_inicio, data_fim, paginate_by (opcional).
    Retorna dict com data_inicio e data_fim em ISO 'YYYY-MM-DD'.
    """
    get = lambda k: (raw_dict.get(k) or '').strip() if raw_dict else ''
    search = get('search')
    valor = get('valor')
    status = get('status')
    status_nota = get('status_nota')
    socio = get('socio')
    forma_pagamento = get('forma_pagamento')
    data_inicio_str = get('data_inicio')
    data_fim_str = get('data_fim')
    paginate_by_str = get('paginate_by')

    d_ini, d_fim = _get_default_dates()
    if data_inicio_str:
        parsed = _parse_date(data_inicio_str)
        if parsed:
            d_ini = parsed
    if data_fim_str:
        parsed = _parse_date(data_fim_str)
        if parsed:
            d_fim = parsed
    if d_ini > d_fim:
        d_ini, d_fim = d_fim, d_ini

    paginate_by = default_paginate_by
    if paginate_by_str:
        try:
            paginate_by = int(paginate_by_str)
            if paginate_by <= 0:
                paginate_by = default_paginate_by
        except ValueError:
            pass

    return {
        'search': search,
        'valor': valor,
        'status': status,
        'status_nota': status_nota,
        'socio': socio,
        'forma_pagamento': forma_pagamento,
        'data_inicio': d_ini.isoformat(),
        'data_fim': d_fim.isoformat(),
        'paginate_by': paginate_by,
    }


class NFSeListView(LoginRequiredMixin, ListView):
    model = NotaFiscalServico
    template_name = 'notasfiscais/nfse_list.html'
    context_object_name = 'notas'
    paginate_by = 20

    def _raw_get_filters(self):
        """Parâmetros de filtro vindos do GET (sem normalizar)."""
        g = self.request.GET
        return {
            'search': g.get('search', ''),
            'valor': g.get('valor', ''),
            'status': g.get('status', ''),
            'status_nota': g.get('status_nota', ''),
            'socio': g.get('socio', ''),
            'forma_pagamento': g.get('forma_pagamento', ''),
            'data_inicio': g.get('data_inicio', ''),
            'data_fim': g.get('data_fim', ''),
            'paginate_by': g.get('paginate_by', ''),
        }

    def _has_filter_params_in_get(self, raw):
        """True se alguma chave de filtro veio preenchida no GET."""
        return bool(
            raw.get('search') or raw.get('valor') or raw.get('status') or raw.get('status_nota')
            or raw.get('socio')
            or raw.get('forma_pagamento')
            or raw.get('data_inicio') or raw.get('data_fim') or raw.get('paginate_by')
        )

    def _get_current_filters(self):
        """
        Filtros atuais: se GET trouxer filtros, normaliza e salva na session;
        senão carrega da session; se não houver session, usa default e salva.
        Retorna dict normalizado (com data_inicio/data_fim em ISO).
        """
        raw = self._raw_get_filters()
        if self._has_filter_params_in_get(raw):
            filters = _normalize_filters(raw, default_paginate_by=self.paginate_by)
            self.request.session[NFS_FILTROS_SESSION_KEY] = filters
            return filters
        filters = self.request.session.get(NFS_FILTROS_SESSION_KEY)
        if not filters:
            d_ini, d_fim = _get_default_dates()
            filters = {
                'search': '',
                'valor': '',
                'status': '',
                'status_nota': '',
                'socio': '',
                'forma_pagamento': '',
                'data_inicio': d_ini.isoformat(),
                'data_fim': d_fim.isoformat(),
                'paginate_by': self.paginate_by,
                'sort': NFSE_DEFAULT_SORT,
                'dir': NFSE_DEFAULT_SORT_DIR,
            }
            self.request.session[NFS_FILTROS_SESSION_KEY] = filters
        return filters

    def get_paginate_by(self, queryset):
        """Usa paginate_by dos filtros persistidos (session)."""
        filters = self.request.session.get(NFS_FILTROS_SESSION_KEY) or {}
        return filters.get('paginate_by', self.paginate_by)

    def get_queryset(self):
        empresa_id = self.request.session.get('empresa_id')
        if not empresa_id:
            messages.warning(self.request, 'Selecione uma empresa para continuar.')
            return NotaFiscalServico.objects.none()

        filters = self._get_current_filters()
        sort_col, sort_dir = _get_nfse_sort_from_request(self.request, filters)
        self._sort_col = sort_col
        self._sort_dir = sort_dir

        queryset = NotaFiscalServico.objects.filter(empresa_id=empresa_id).select_related(
            'codigo_da_regra_do_imposto', 'socio', 'forma_pagamento'
        )

        queryset = _apply_filtro_search_nfse(queryset, filters.get('search'))

        queryset = _apply_filtro_valor_nfse(queryset, filters.get('valor'))

        status = filters.get('status', '')
        if status:
            queryset = queryset.filter(status_conciliacao=status)

        status_nota = filters.get('status_nota', '')
        if status_nota == 'ativa':
            queryset = queryset.filter(data_cancelamento__isnull=True)
        elif status_nota == 'cancelada':
            queryset = queryset.filter(data_cancelamento__isnull=False)

        forma_pagamento = filters.get('forma_pagamento', '')
        if forma_pagamento:
            if forma_pagamento == 'none':
                queryset = queryset.filter(forma_pagamento__isnull=True)
            else:
                try:
                    queryset = queryset.filter(forma_pagamento_id=int(forma_pagamento))
                except (ValueError, TypeError):
                    pass

        queryset = _apply_filtro_socio_nfse(queryset, filters.get('socio'))

        d_ini = _parse_date(filters.get('data_inicio'))
        d_fim = _parse_date(filters.get('data_fim'))
        if not d_ini or not d_fim:
            d_ini, d_fim = _get_default_dates()
        if d_ini > d_fim:
            d_ini, d_fim = d_fim, d_ini
        queryset = queryset.filter(data_emissao__range=(d_ini, d_fim))

        return _apply_nfse_sort(queryset, sort_col, sort_dir)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa_id = self.request.session.get('empresa_id')

        # Filtros atuais vêm da session (já definida em get_queryset)
        filters = self.request.session.get(NFS_FILTROS_SESSION_KEY) or {}
        search = filters.get('search', '')
        valor_filter = filters.get('valor', '')
        status_filter = filters.get('status', '')
        status_nota_filter = filters.get('status_nota', '')
        socio_filter = filters.get('socio', '')
        forma_pagamento_filter = filters.get('forma_pagamento', '')
        data_inicio = filters.get('data_inicio', '')
        data_fim = filters.get('data_fim', '')
        paginate_by = str(filters.get('paginate_by', self.paginate_by))

        # Verificar quais notas já têm contas a receber
        notas_ids = [nfse.id for nfse in context['notas']]
        if notas_ids:
            from contasareceber.models import ContaAReceber
            contas_existentes = ContaAReceber.objects.filter(
                empresa_id=empresa_id,
                nota_id__in=notas_ids
            ).values_list('nota_id', flat=True)
            context['notas_com_conta'] = set(contas_existentes)

        # Período fechado: comparar com date (ApuracaoPeriodo usa DateField)
        periodo_fechado = False
        from .models import ApuracaoPeriodo
        d_ini = _parse_date(data_inicio) or _get_default_dates()[0]
        d_fim = _parse_date(data_fim) or _get_default_dates()[1]
        if d_ini and d_fim and empresa_id:
            try:
                periodo = ApuracaoPeriodo.objects.get(
                    empresa_id=empresa_id,
                    data_inicio=d_ini,
                    data_fim=d_fim
                )
                periodo_fechado = periodo.status == 'fechado'
            except ApuracaoPeriodo.DoesNotExist:
                pass

        from regraImposto.models import RegraImposto
        forma_pagamento_choices = list(Cobranca.objects.all().order_by('descricao'))
        forma_pagamento_choices.insert(0, {'id': 'none', 'descricao': 'Sem forma definida'})

        # data_inicio/data_fim em ISO para input type=date; versão dd/mm/aaaa para exibição
        data_inicio_display = ''
        data_fim_display = ''
        if d_ini:
            data_inicio_display = d_ini.strftime('%d/%m/%Y')
        if d_fim:
            data_fim_display = d_fim.strftime('%d/%m/%Y')

        # Sócios da empresa selecionada
        socios_list = Socio.objects.filter(empresa_id=empresa_id).order_by('socio') if empresa_id else []
        # URL para aplicar sócio (evita NoReverseMatch com namespace duplicado)
        path_prefix = (self.request.path.strip('/').split('/') or ['notasfiscais'])[0] or 'notasfiscais'
        aplicar_socio_url = f'/{path_prefix}/nfse/aplicar-socio/'
        aplicar_cobranca_url = f'/{path_prefix}/nfse/aplicar-cobranca/'
        context.update({
            'search': search,
            'valor_filter': valor_filter,
            'status_filter': status_filter,
            'status_nota_filter': status_nota_filter,
            'socio_filter': socio_filter,
            'forma_pagamento_filter': forma_pagamento_filter,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'data_inicio_display': data_inicio_display,
            'data_fim_display': data_fim_display,
            'paginate_by': paginate_by,
            'forma_pagamento_choices': forma_pagamento_choices,
            'regras_imposto': RegraImposto.objects.all().order_by('DescricaoRegraImposto'),
            'status_choices': NotaFiscalServico.STATUS_CONCILIACAO_CHOICES,
            'periodo_fechado': periodo_fechado,
            'socios': socios_list,
            'aplicar_socio_url': aplicar_socio_url,
            'aplicar_cobranca_url': aplicar_cobranca_url,
            'sort_col': getattr(self, '_sort_col', filters.get('sort', NFSE_DEFAULT_SORT)),
            'sort_dir': getattr(self, '_sort_dir', filters.get('dir', NFSE_DEFAULT_SORT_DIR)),
        })

        if empresa_id:
            # Queryset com os mesmos filtros da listagem (para os indicadores refletirem a tela)
            qs = self.get_queryset()
            total = qs.count()
            cancelado = qs.filter(data_cancelamento__isnull=False).count()
            pago = qs.filter(status_conciliacao='conciliado').count()
            pendente = total - cancelado - pago
            if pendente < 0:
                pendente = 0

            # Estatísticas para os indicadores (Total, Pendentes, Pagas, Canceladas)
            context['stats'] = {
                'total': total,
                'pendente': pendente,
                'pago': pago,
                'cancelado': cancelado,
            }

            # Valores (mantido para compatibilidade)
            total_notas = NotaFiscalServico.objects.filter(empresa_id=empresa_id).count()
            valor_total = NotaFiscalServico.objects.filter(empresa_id=empresa_id).aggregate(
                total=Sum('valor_liquido')
            )['total'] or 0
            valor_pendente = NotaFiscalServico.objects.filter(
                empresa_id=empresa_id
            ).aggregate(total=Sum('valor_liquido'))['total'] or 0

            context.update({
                'total_notas': total_notas,
                'valor_total': valor_total,
                'valor_pendente': valor_pendente,
            })
        else:
            context['stats'] = {'total': 0, 'pendente': 0, 'pago': 0, 'cancelado': 0}

        return context

class NFSeCreateView(LoginRequiredMixin, CreateView):
    model = NotaFiscalServico
    form_class = NFSeForm
    template_name = 'notasfiscais/nfse_form.html'
    success_url = reverse_lazy('notasfiscais:list')
    
    def form_valid(self, form):
        empresa_id = self.request.session.get('empresa_id')
        if not empresa_id:
            messages.error(self.request, 'Selecione uma empresa para continuar.')
            return redirect('empresa:lista')
        
        form.instance.empresa_id = empresa_id
        messages.success(self.request, 'NFSe criada com sucesso!')
        response = super().form_valid(form)
        from contasareceber.socio_sync import propagar_socio_nota_para_contas_receber

        propagar_socio_nota_para_contas_receber(self.object)
        return response
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        empresa_id = self.request.session.get('empresa_id')
        if 'forma_pagamento' in form.fields and hasattr(form.fields['forma_pagamento'], 'queryset'):
            form.fields['forma_pagamento'].queryset = Cobranca.objects.all().order_by('descricao')  # ajuste o campo de ordenação se quiser
        if empresa_id:
            # Filtra formas de pagamento pela empresa
            #form.fields['forma_pagamento'].queryset = form.fields['forma_pagamento'].queryset.filter(empresa_id=empresa_id)
            form.fields['socio'].queryset = form.fields['socio'].queryset.filter(empresa_id=empresa_id)
        return form

class NFSeUpdateView(LoginRequiredMixin, UpdateView):
    model = NotaFiscalServico
    form_class = NFSeUpdateForm
    template_name = 'notasfiscais/nfse_form.html'

    def get_queryset(self):
        empresa_id = self.request.session.get('empresa_id')
        if empresa_id:
            return NotaFiscalServico.objects.filter(empresa_id=empresa_id)
        return NotaFiscalServico.objects.none()

    def dispatch(self, request, *args, **kwargs):
        # Verificar se o usuário tem empresa selecionada
        empresa_id = request.session.get('empresa_id')
        print(f"DEBUG NFSeUpdateView.dispatch: empresa_id da sessão = {empresa_id}")
        if not empresa_id:
            print("DEBUG NFSeUpdateView.dispatch: Empresa não selecionada, redirecionando para lista de empresas")
            messages.warning(request, 'Selecione uma empresa para continuar.')
            return redirect('empresa:lista')

        # Converter empresa_id para inteiro para evitar problemas de tipo
        try:
            empresa_id = int(empresa_id)
        except (ValueError, TypeError):
            print("DEBUG NFSeUpdateView.dispatch: empresa_id inválido, redirecionando para lista de empresas")
            messages.warning(request, 'Empresa selecionada inválida. Selecione novamente.')
            return redirect('empresa:lista')

        # Verificar se a nota pertence à empresa selecionada
        nota = self.get_object()
        print(f"DEBUG NFSeUpdateView.dispatch: Nota ID {nota.id}, empresa_id da nota = {nota.empresa_id}, empresa_id da sessão = {empresa_id}")
        if nota.empresa_id != empresa_id:
            print("DEBUG NFSeUpdateView.dispatch: Nota não pertence à empresa selecionada, redirecionando para lista")
            messages.error(request, 'Você não tem permissão para editar esta nota fiscal.')
            return redirect('notasfiscais:list')

        # Verificar se o período está fechado
        from .models import ApuracaoPeriodo
        periodos_fechados = ApuracaoPeriodo.objects.filter(
            empresa_id=int(empresa_id),
            data_inicio__lte=nota.data_emissao,
            data_fim__gte=nota.data_emissao,
            status='fechado'
        )
        print(f"DEBUG NFSeUpdateView.dispatch: Data emissão da nota = {nota.data_emissao}, períodos fechados encontrados = {periodos_fechados.count()}")
        if periodos_fechados.exists():
            print("DEBUG NFSeUpdateView.dispatch: Período fechado, redirecionando para lista")
            messages.error(request, 'Período fechado - vá para Apuração de Impostos e reabra o período para editar notas.')
            return redirect('notasfiscais:list')

        print("DEBUG NFSeUpdateView.dispatch: Todas as validações passaram, prosseguindo")
        return super().dispatch(request, *args, **kwargs)
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        empresa_id = self.request.session.get('empresa_id')
    #    form.fields['forma_pagamento'].queryset = form.fields['forma_pagamento'].cobranca()
        if 'forma_pagamento' in form.fields and hasattr(form.fields['forma_pagamento'], 'queryset'):
            form.fields['forma_pagamento'].queryset = Cobranca.objects.all().order_by('descricao')  # ajuste o campo de ordenação se quiser

        if empresa_id:
            # Filtra formas de pagamento pela empresa
            #form.fields['forma_pagamento'].queryset = form.fields['forma_pagamento'].queryset.filter(empresa_id=empresa_id)
            
            form.fields['socio'].queryset = form.fields['socio'].queryset.filter(empresa_id=empresa_id)
        return form
    def form_invalid(self, form):
        # Adiciona todas as mensagens de erro no messages framework
        for field, errors in form.errors.items():
            for error in errors:
                if field == '__all__':
                    messages.error(self.request, error)
                else:
                    messages.error(self.request, f"{form.fields[field].label}: {error}")
        return super().form_invalid(form)
    def form_valid(self, form):
        messages.success(self.request, 'NFSe atualizada com sucesso!')
        response = super().form_valid(form)
        from contasareceber.socio_sync import propagar_socio_nota_para_contas_receber

        propagar_socio_nota_para_contas_receber(self.object)
        return response

    def get_success_url(self):
        # Manter os filtros da URL atual na URL de sucesso
        current_url = self.request.META.get('HTTP_REFERER', '')
        if '?' in current_url:
            query_string = current_url.split('?', 1)[1]
            return f"{reverse_lazy('notasfiscais:list')}?{query_string}"
        return reverse_lazy('notasfiscais:list')

class NFSeDetailView(LoginRequiredMixin, DetailView):
    model = NotaFiscalServico
    template_name = 'notasfiscais/nfse_detail.html'
    context_object_name = 'nfse'
    
    def get_queryset(self):
        empresa_id = self.request.session.get('empresa_id')
        if empresa_id:
            return NotaFiscalServico.objects.filter(empresa_id=empresa_id)
        return NotaFiscalServico.objects.none()

class NFSeDeleteView(LoginRequiredMixin, DeleteView):
    model = NotaFiscalServico
    template_name = 'notasfiscais/nfse_confirm_delete.html'
    
    def get_queryset(self):
        empresa_id = self.request.session.get('empresa_id')
        pk = self.kwargs.get('pk')  # pega o PK da URL
        if empresa_id:
            return NotaFiscalServico.objects.filter(empresa_id=empresa_id, pk=pk )
        return NotaFiscalServico.objects.none()
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'NFSe excluída com sucesso!')
        return super().delete(request, *args, **kwargs)

    def get_success_url(self):
        # Manter os filtros da URL atual na URL de sucesso
        current_url = self.request.META.get('HTTP_REFERER', '')
        if '?' in current_url:
            query_string = current_url.split('?', 1)[1]
            return f"{reverse_lazy('notasfiscais:list')}?{query_string}"
        return reverse_lazy('notasfiscais:list')
    def dispatch(self, request, *args, **kwargs):
        empresa_id = request.session.get('empresa_id')
        self.nfse = get_object_or_404(
            NotaFiscalServico,
            pk=self.kwargs['pk'],
            empresa_id=empresa_id
        )

        # Verificar se o período está fechado
        if empresa_id:
            from .models import ApuracaoPeriodo
            periodos_fechados = ApuracaoPeriodo.objects.filter(
                empresa_id=int(empresa_id),
                data_inicio__lte=self.nfse.data_emissao,
                data_fim__gte=self.nfse.data_emissao,
                status='fechado'
            )
            if periodos_fechados.exists():
                messages.error(request, 'Período fechado - vá para Apuração de Impostos e reabra o período para excluir notas.')
                return redirect('notasfiscais:list')

        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['nfse'] = self.nfse
        return ctx


# Sessão guarda cópias em disco dos XML após "Visualizar notas", para "Confirmar importação"
# (o browser não reenvia arquivos num segundo POST só com nome do arquivo).
NFSE_SESSION_PENDING = 'nfse_xml_import_pending'


def _nfse_safe_unlink(path: str) -> None:
    if path and os.path.isfile(path):
        try:
            os.unlink(path)
        except OSError:
            pass


def _nfse_clear_pending_session(request) -> None:
    old = request.session.pop(NFSE_SESSION_PENDING, None)
    if not old:
        return
    for item in old:
        _nfse_safe_unlink(item.get('path', ''))


def _nfse_persist_upload_to_temp(uploaded_file, user_id: int) -> dict[str, str]:
    uploaded_file.seek(0)
    ext = os.path.splitext(uploaded_file.name)[1] or '.xml'
    fd, path = tempfile.mkstemp(prefix=f'nfse_u{user_id}_', suffix=ext)
    try:
        with os.fdopen(fd, 'wb') as dest:
            for chunk in uploaded_file.chunks():
                dest.write(chunk)
    except Exception:
        _nfse_safe_unlink(path)
        raise
    return {'path': path, 'name': os.path.basename(uploaded_file.name) or 'nota.xml'}


class XMLImportView(LoginRequiredMixin, FormView):
    template_name = 'notasfiscais/xml_import.html'
    form_class = XMLUploadForm
    success_url = reverse_lazy('notasfiscais:import')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Recuperar resultado da importação da sessão
        resultado_importacao = self.request.session.pop('resultado_importacao', None)
        if resultado_importacao:
            context['resultado_importacao'] = resultado_importacao
        
        return context
    
    def post(self, request, *args, **kwargs):
        # Se for preview, mostra as notas sem salvar
        if request.POST.get('action') == 'preview':
            return self.preview_xml(request)
        # Confirma importação usando XMLs salvos na sessão (após preview)
        if request.POST.get('action') == 'confirm_import':
            return self.confirm_import_from_session(request)

        return super().post(request, *args, **kwargs)
    
    def preview_xml(self, request):
        """Mostra preview das notas dos XMLs antes de importar"""
        try:
            xml_files = request.FILES.getlist('xml_file')
            empresa_id = request.session.get('empresa_id')
            importar_canceladas = bool(request.POST.get('importar_canceladas'))

            if not empresa_id:
                messages.error(request, 'Selecione uma empresa para continuar.')
                return redirect('empresa:lista')

            if not xml_files:
                messages.error(request, 'Arquivos XML não fornecidos.')
                return redirect('notasfiscais:import')

            empresa = Empresa.objects.get(id=empresa_id)

            # Remove cópias antigas e grava novos XMLs em temp para o passo "Confirmar importação"
            _nfse_clear_pending_session(request)
            pending_paths: list[dict[str, str]] = []
            try:
                for xml_file in xml_files:
                    pending_paths.append(_nfse_persist_upload_to_temp(xml_file, request.user.id))
            except Exception as e:
                for item in pending_paths:
                    _nfse_safe_unlink(item.get('path', ''))
                raise e
            request.session[NFSE_SESSION_PENDING] = pending_paths
            request.session['nfse_import_importar_canceladas'] = importar_canceladas
            request.session.modified = True

            # Extrai as notas de todos os XMLs para preview (reposiciona leitura após persistir)
            notas_preview = []
            file_names = []
            for xml_file in xml_files:
                xml_file.seek(0)
                notas_arquivo = self.extract_notas_preview(xml_file, empresa)
                notas_preview.extend(notas_arquivo)
                file_names.append(xml_file.name)

            valid_count = sum(1 for n in notas_preview if n.get('status') == 'valido')
            invalid_count = len(notas_preview) - valid_count
            context = {
                'form': XMLUploadForm(),
                'empresa': empresa,
                'notas_preview': notas_preview,
                'xml_file_name': ', '.join(file_names),
                'total_notas': len(notas_preview),
                'valid_count': valid_count,
                'invalid_count': invalid_count,
                'show_preview': True,
                'importar_canceladas': importar_canceladas,
            }

            return render(request, 'notasfiscais/xml_import.html', context)

        except Exception as e:
            _nfse_clear_pending_session(request)
            print(f"ERRO em preview_xml: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Erro ao processar XML: {str(e)}')
            return redirect('notasfiscais:import')

    def confirm_import_from_session(self, request):
        """Importa os XMLs gravados em disco no preview (confirmar após listar notas)."""
        empresa_id = request.session.get('empresa_id')
        if not empresa_id:
            messages.error(request, 'Selecione uma empresa para continuar.')
            return redirect('empresa:lista')

        pending = request.session.pop(NFSE_SESSION_PENDING, None)
        # Checkbox só vem no POST se marcado; senão usa o que foi guardado no preview
        if request.POST.get('importar_canceladas') == 'on':
            importar_canceladas = True
            request.session.pop('nfse_import_importar_canceladas', None)
        else:
            importar_canceladas = bool(request.session.pop('nfse_import_importar_canceladas', False))
        request.session.modified = True

        if not pending:
            messages.error(
                request,
                'Não há XML pendente para importar. Clique em "Visualizar notas" e depois em "Confirmar importação".',
            )
            return redirect('notasfiscais:import')

        try:
            empresa = Empresa.objects.get(id=empresa_id)
        except Empresa.DoesNotExist:
            for item in pending:
                _nfse_safe_unlink(item.get('path', ''))
            messages.error(request, 'Empresa não encontrada.')
            return redirect('notasfiscais:import')

        resultado_total = {
            'notas_importadas': [],
            'notas_ignoradas': [],
            'notas_canceladas': [],
            'total_processadas': 0,
            'total_importadas': 0,
            'total_canceladas': 0,
            'total_ignoradas': 0,
        }

        try:
            for item in pending:
                path = item.get('path')
                name = item.get('name') or 'nota.xml'
                try:
                    with open(path, 'rb') as f:
                        raw = f.read()
                    uploaded = SimpleUploadedFile(name, raw, content_type='application/xml')
                    resultado = import_nfse_from_xml(
                        uploaded,
                        request.user,
                        empresa,
                        importar_canceladas=importar_canceladas,
                    )
                    resultado_total['notas_importadas'].extend(resultado.get('notas_importadas', []))
                    resultado_total['notas_ignoradas'].extend(resultado.get('notas_ignoradas', []))
                    resultado_total['notas_canceladas'].extend(resultado.get('notas_canceladas', []))
                    resultado_total['total_processadas'] += resultado.get('total_processadas', 0)
                    resultado_total['total_importadas'] += resultado.get('total_importadas', 0)
                    resultado_total['total_canceladas'] += resultado.get('total_canceladas', 0)
                    resultado_total['total_ignoradas'] += resultado.get('total_ignoradas', 0)
                finally:
                    _nfse_safe_unlink(path)

            if resultado_total.get('total_canceladas', 0) > 0:
                if resultado_total['total_canceladas'] == 1:
                    n = resultado_total['notas_canceladas'][0]
                    messages.success(
                        request,
                        f"NFSe {n['numero_nota']} cancelada com sucesso. Motivo: {n.get('motivo', '—')}",
                    )
                else:
                    messages.success(
                        request,
                        f"{resultado_total['total_canceladas']} NFSe canceladas com sucesso via evento.",
                    )
            elif resultado_total['total_importadas'] > 0:
                if resultado_total['total_importadas'] == 1:
                    messages.success(
                        request,
                        f'NFSe {resultado_total["notas_importadas"][0]["numero_nota"]} importada com sucesso!',
                    )
                else:
                    messages.success(
                        request,
                        f'{resultado_total["total_importadas"]} NFSe importadas com sucesso!',
                    )

            if resultado_total['total_ignoradas'] > 0:
                if resultado_total['total_ignoradas'] == 1:
                    motivo = resultado_total['notas_ignoradas'][0]['motivo']
                    messages.warning(request, f'1 NFSe ignorada: {motivo}')
                else:
                    messages.warning(
                        request,
                        f'{resultado_total["total_ignoradas"]} NFSe ignoradas (duplicatas ou com erro)',
                    )

            if resultado_total['total_importadas'] == 0:
                if resultado_total['total_ignoradas'] > 0:
                    messages.warning(
                        request,
                        'Nenhuma nova NFSe foi importada. Todas as notas já existem no sistema.',
                    )
                else:
                    messages.error(request, 'Nenhuma NFSe válida encontrada no XML.')

            request.session['resultado_importacao'] = serialize_resultado(resultado_total)
            return redirect('notasfiscais:import')

        except ValueError as e:
            messages.error(request, f'Erro de validação: {str(e)}')
            return redirect('notasfiscais:import')
        except Exception as e:
            messages.error(request, f'Erro ao importar XML: {str(e)}')
            return redirect('notasfiscais:import')
    
    def extract_notas_preview(self, xml_file, empresa):
        """Extrai dados das notas para preview sem salvar no banco"""
        try:
            from .utils import extract_xml_data_preview
            return extract_xml_data_preview(xml_file, empresa)
        except Exception as e:
            print(f"Erro ao extrair preview: {str(e)}")
            return []
    
    def form_valid(self, form):
        empresa_id = self.request.session.get('empresa_id')

        if not empresa_id:
            messages.error(self.request, 'Selecione uma empresa para continuar.')
            return redirect('empresa:lista')

        try:
            xml_files = form.files.getlist('xml_file')
            importar_canceladas = form.cleaned_data.get('importar_canceladas', False)

            if not xml_files:
                messages.error(self.request, 'Selecione pelo menos um arquivo XML para importar.')
                return self.form_invalid(form)

            # Importação direta (sem passar pelo preview): remove XMLs temporários de um preview antigo
            _nfse_clear_pending_session(self.request)

            # Busca a empresa
            empresa = Empresa.objects.get(id=empresa_id)

            # Tenta importar todos os XMLs
            resultado_total = {
                'notas_importadas': [],
                'notas_ignoradas': [],
                'notas_canceladas': [],
                'total_processadas': 0,
                'total_importadas': 0,
                'total_canceladas': 0,
                'total_ignoradas': 0
            }

            for xml_file in xml_files:
                resultado = import_nfse_from_xml(
                    xml_file,
                    self.request.user,
                    empresa,
                    importar_canceladas=importar_canceladas,
                )

                # Agrega os resultados
                resultado_total['notas_importadas'].extend(resultado.get('notas_importadas', []))
                resultado_total['notas_ignoradas'].extend(resultado.get('notas_ignoradas', []))
                resultado_total['notas_canceladas'].extend(resultado.get('notas_canceladas', []))
                resultado_total['total_processadas'] += resultado.get('total_processadas', 0)
                resultado_total['total_importadas'] += resultado.get('total_importadas', 0)
                resultado_total['total_canceladas'] += resultado.get('total_canceladas', 0)
                resultado_total['total_ignoradas'] += resultado.get('total_ignoradas', 0)
            
            # Preparar mensagens baseadas no resultado total (todos os arquivos)
            if resultado_total.get('total_canceladas', 0) > 0:
                if resultado_total['total_canceladas'] == 1:
                    n = resultado_total['notas_canceladas'][0]
                    messages.success(
                        self.request,
                        f"NFSe {n['numero_nota']} cancelada com sucesso. Motivo: {n.get('motivo', '—')}",
                    )
                else:
                    messages.success(
                        self.request,
                        f"{resultado_total['total_canceladas']} NFSe canceladas com sucesso via evento.",
                    )
            elif resultado_total['total_importadas'] > 0:
                if resultado_total['total_importadas'] == 1:
                    messages.success(self.request, f'NFSe {resultado_total["notas_importadas"][0]["numero_nota"]} importada com sucesso!')
                else:
                    messages.success(self.request, f'{resultado_total["total_importadas"]} NFSe importadas com sucesso!')
            
            if resultado_total['total_ignoradas'] > 0:
                if resultado_total['total_ignoradas'] == 1:
                    motivo = resultado_total['notas_ignoradas'][0]['motivo']
                    messages.warning(self.request, f'1 NFSe ignorada: {motivo}')
                else:
                    messages.warning(self.request, f'{resultado_total["total_ignoradas"]} NFSe ignoradas (duplicatas ou com erro)')
            
            # Se não importou nenhuma nota
            if resultado_total['total_importadas'] == 0:
                if resultado_total['total_ignoradas'] > 0:
                    messages.warning(self.request, 'Nenhuma nova NFSe foi importada. Todas as notas já existem no sistema.')
                else:
                    messages.error(self.request, 'Nenhuma NFSe válida encontrada no XML.')
            
            # Armazenar resultado total na sessão para mostrar no template (sem objetos Django)
            self.request.session['resultado_importacao'] = serialize_resultado(resultado_total)
            
            return super().form_valid(form)
            
        except ValueError as e:
            messages.error(self.request, f'Erro de validação: {str(e)}')
            return self.form_invalid(form)

        except Exception as e:
            messages.error(self.request, f'Erro ao importar XML: {str(e)}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        print("=== DEBUG XMLImportView.form_invalid ===")
        print(f"Form errors: {form.errors}")
        print(f"Form data: {form.data}")
        print(f"Form files: {form.files}")
        return super().form_invalid(form)


class NfseEventoCancelamentoImportView(XMLImportView):
    """Importação de XML de eventos de cancelamento NFS-e (pasta Eventos do portal nacional)."""

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Importar eventos de cancelamento NFS-e'
        ctx['page_intro'] = (
            'Selecione os XML da pasta <strong>Eventos</strong> '
            '(ex.: <code>4642_Cancelamento de NFS-e_15559.xml</code>). '
            'O sistema localiza a NFSe, marca como cancelada, zera os valores e grava o motivo.'
        )
        return ctx


class NfsePortalNacionalImportView(LoginRequiredMixin, FormView):
    """
    Baixa XML (e tenta PDF) na API nacional SEFIN por identificador DPS (mTLS com PFX).
    Uma requisição = uma DPS; opcionalmente um intervalo de números com a mesma série (até 100 por envio).
    Período no formulário só filtra após o download.
    """

    template_name = "notasfiscais/nfse_portal_nacional.html"
    form_class = PortalNacionalNfseForm
    success_url = reverse_lazy("notasfiscais:list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        eid = self.request.session.get("empresa_id")
        if eid:
            kwargs["empresa"] = Empresa.objects.filter(pk=eid).first()
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        eid = self.request.session.get("empresa_id")
        empresa = Empresa.objects.filter(pk=eid).first() if eid else None
        if not empresa:
            return initial
        di = parse_date((self.request.GET.get("data_inicio") or "").strip())
        df = parse_date((self.request.GET.get("data_fim") or "").strip())
        if di:
            initial["data_periodo_inicio"] = di
        if df:
            initial["data_periodo_fim"] = df

        qs = NotaFiscalServico.objects.filter(empresa=empresa)
        if di and df:
            qs = qs.filter(data_emissao__gte=di, data_emissao__lte=df)
        ult = qs.order_by("-data_emissao", "-pk").first()

        pad = re.sub(r"\D", "", (getattr(empresa, "nfse_nacional_dps_serie_padrao", None) or "").strip())
        if pad:
            initial["serie_dps"] = pad.zfill(5)[-5:]
        elif ult and (ult.serie or "").strip():
            s = re.sub(r"\D", "", (ult.serie or "").strip())
            if s:
                initial["serie_dps"] = s.zfill(5)[-5:]

        nd = (getattr(ult, "numero_dps", None) or "").strip() if ult else ""
        if nd and re.sub(r"\D", "", nd):
            initial["numero_dps"] = str(int(re.sub(r"\D", "", nd)))
        else:
            prox = getattr(empresa, "nfse_nacional_dps_proximo_numero", None)
            if prox is not None and int(prox) > 0:
                initial["numero_dps"] = str(int(prox))
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from empresa.nfse_nacional_config import nfse_nacional_resolvido_para_empresa

        d = getattr(django_settings, "NFSE_NACIONAL", {}) or {}
        empresa = None
        eid = self.request.session.get("empresa_id")
        if eid:
            empresa = Empresa.objects.filter(pk=eid).first()
        nf = nfse_nacional_resolvido_para_empresa(empresa) if empresa else {}
        pfx = (nf.get("pfx_path") or (d.get("pfx_path") or "")).strip()
        ctx["nfse_nacional_configurado"] = bool(pfx) and os.path.isfile(pfx)
        ctx["nfse_nacional_base_url"] = (nf.get("base_url") or d.get("base_url") or "").strip()
        ibge_ok = False
        if empresa:
            ib = re.sub(r"\D", "", (getattr(empresa, "nfse_nacional_codigo_ibge_municipio", None) or "").strip())
            ibge_ok = len(ib) == 7
        ctx["nfse_portal_ibge_ok"] = ibge_ok
        return ctx

    def form_valid(self, form):
        empresa_id = self.request.session.get("empresa_id")
        if not empresa_id:
            messages.error(self.request, "Selecione uma empresa para continuar.")
            return redirect("empresa:lista")

        from empresa.nfse_nacional_config import nfse_nacional_resolvido_para_empresa

        empresa = Empresa.objects.get(id=empresa_id)
        d = getattr(django_settings, "NFSE_NACIONAL", {}) or {}
        cfg = nfse_nacional_resolvido_para_empresa(empresa)
        pfx_path = (cfg.get("pfx_path") or "").strip()
        pfx_password = cfg.get("pfx_password") or ""
        base_url = (cfg.get("base_url") or d.get("base_url") or "https://sefin.nfse.gov.br").strip()
        verify_ssl = bool(cfg.get("verify_ssl", d.get("verify_ssl", True)))

        from .portal_nacional_client import (
            baixar_nfse_pdf_por_identificador_dps,
            baixar_nfse_xml_por_identificador_dps,
            montar_identificador_dps,
        )
        from .nfse_xml_copia import salvar_baixados_portal_nacional_files, validar_periodo_xml_nfse

        importar_canceladas = bool(form.cleaned_data.get("importar_canceladas"))

        ibge_d = form.cleaned_data["_ibge_dps"]
        tipo_d = form.cleaned_data.get("_tipo_inscricao_dps") or "2"
        insc_d = form.cleaned_data["_inscricao_dps"]
        serie_d = form.cleaned_data["_serie_dps"]
        rango = form.cleaned_data.get("_dps_numero_range")
        if rango:
            n_ini, n_fim = rango
            numeros_dps = list(range(n_ini, n_fim + 1))
        else:
            try:
                numeros_dps = [int(form.cleaned_data["_numero_dps"])]
            except (TypeError, ValueError):
                numeros_dps = []
        if not numeros_dps:
            messages.error(self.request, "Número da DPS inválido; confira o formulário.")
            return redirect("notasfiscais:list")

        di = form.cleaned_data.get("data_periodo_inicio")
        df = form.cleaned_data.get("data_periodo_fim")

        total_imp = 0
        total_ign = 0
        avisos_pdf: list[str] = []
        erros_baixa: list[str] = []

        for n_int in numeros_dps:
            num_pad = str(n_int).zfill(15)[-15:]
            try:
                id_dps = montar_identificador_dps(ibge_d, tipo_d, insc_d, serie_d, num_pad)
            except ValueError as e:
                erros_baixa.append(f"nº {n_int}: {e}")
                continue

            xml_bytes, err = baixar_nfse_xml_por_identificador_dps(
                id_dps,
                pfx_path,
                pfx_password,
                base_url,
                verify_ssl=verify_ssl,
            )
            nome_arquivo = f"nfse_nacional_dps_{id_dps}.xml"

            if err or not xml_bytes:
                erros_baixa.append(f"Série {serie_d} nº {n_int}: {err or 'sem XML'}")
                continue

            msg_periodo = validar_periodo_xml_nfse(xml_bytes, di, df)
            if msg_periodo:
                erros_baixa.append(f"Série {serie_d} nº {n_int}: {msg_periodo}")
                continue

            pdf_bytes, pdf_err = baixar_nfse_pdf_por_identificador_dps(
                id_dps, pfx_path, pfx_password, base_url, verify_ssl=verify_ssl
            )
            if not pdf_bytes and len(numeros_dps) == 1:
                messages.warning(
                    self.request,
                    "XML baixado. O PDF não foi obtido pela API com as rotas padrão (o manual da SEFIN pode usar outro caminho). "
                    + (pdf_err or ""),
                )
            elif not pdf_bytes and pdf_err:
                avisos_pdf.append(f"nº {n_int}: {pdf_err}")

            stem = os.path.splitext(nome_arquivo)[0] or "nfse_nacional"
            salvar_baixados_portal_nacional_files(
                xml_bytes,
                pdf_bytes,
                stem,
                empresa,
                importar_canceladas=importar_canceladas,
            )

            nome = nome_arquivo
            uploaded = SimpleUploadedFile(nome, xml_bytes, content_type="application/xml")

            try:
                resultado = import_nfse_from_xml(
                    uploaded,
                    self.request.user,
                    empresa,
                    importar_canceladas=importar_canceladas,
                )
            except ValueError as e:
                erros_baixa.append(f"Série {serie_d} nº {n_int}: {e}")
                continue
            except Exception as e:
                erros_baixa.append(f"Série {serie_d} nº {n_int}: {e}")
                continue

            total_imp += int(resultado.get("total_importadas") or 0)
            total_ign += int(resultado.get("total_ignoradas") or 0)

        if rango and form.cleaned_data.get("_sequencia_numero_automatico") and not form.cleaned_data.get(
            "_numero_dps_informado_pelo_usuario"
        ):
            _, n_fim = rango
            Empresa.objects.filter(pk=empresa_id).update(nfse_nacional_dps_proximo_numero=n_fim + 1)
        elif not rango and form.cleaned_data.get("_sequencia_numero_automatico"):
            if total_imp > 0 or total_ign > 0:
                try:
                    n = int(form.cleaned_data["_numero_dps"])
                except (TypeError, ValueError):
                    n = 0
                if n > 0:
                    Empresa.objects.filter(pk=empresa_id).update(nfse_nacional_dps_proximo_numero=n + 1)

        for a in avisos_pdf[:5]:
            messages.warning(self.request, f"PDF: {a}")
        for e in erros_baixa[:12]:
            messages.warning(self.request, e)
        if len(erros_baixa) > 12:
            messages.warning(self.request, f"... e mais {len(erros_baixa) - 12} erro(s) no lote.")

        if total_imp > 0:
            messages.success(
                self.request,
                f"{total_imp} NFSe importada(s) a partir do Portal Nacional (série DPS {serie_d}).",
            )
            return redirect("notasfiscais:list")
        if total_ign > 0:
            messages.warning(
                self.request,
                f"Nenhuma nota nova; {total_ign} já existente(s) ou ignorada(s) no lote. Verifique duplicidade ou números sem NFS-e na SEFIN.",
            )
            return redirect("notasfiscais:list")
        if erros_baixa:
            messages.error(
                self.request,
                "Nenhuma NFSe foi importada neste envio. Confira série/número(s), certificado e mensagens acima.",
            )
            return redirect("notasfiscais:list")
        messages.error(self.request, "Nenhuma NFSe válida foi importada a partir do XML retornado.")
        return redirect("notasfiscais:list")


def _extrair_chave_acesso_nfse_do_xml(xml_bytes: bytes) -> Optional[str]:
    try:
        txt = xml_bytes.decode("utf-8", errors="ignore")
    except Exception:
        return None
    # Ex.: infNFSe Id="NFS1200401221980..."
    m = re.search(r'Id\s*=\s*"NFS(\d{40,60})"', txt, re.IGNORECASE)
    if m:
        return re.sub(r"\D", "", m.group(1))
    # fallback: qualquer sequência longa associada a "chave"
    m2 = re.search(r"chave[^0-9]*(\d{40,60})", txt, re.IGNORECASE)
    if m2:
        return re.sub(r"\D", "", m2.group(1))
    return None


class NfseAdnImportView(LoginRequiredMixin, View):
    """
    Sincronização ADN pela interface web desativada.
    Em servidor com certificado: ``python manage.py sync_adn_until`` (ou integração própria).
    """

    def get(self, request, *args, **kwargs):
        messages.info(
            request,
            "A sincronização ADN por esta página foi desativada. Use «Portal (extensão)» ou "
            "«Importar XML», ou o comando ``python manage.py sync_adn_until`` onde o PFX estiver configurado.",
        )
        return redirect("notasfiscais:list")

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)


class NfsePortalExtensaoView(LoginRequiredMixin, FormView):
    """
    Tela de integração com o portal nacional: período do filtro, Selenium e importação a partir
    da pasta do mês. O envio manual de XML/PDF por formulário nesta página foi removido da UI;
    ``form_valid`` permanece para compatibilidade se houver POST com ficheiros.
    """

    template_name = "notasfiscais/nfse_portal_extensao.html"
    form_class = PortalExtensaoNfseForm
    success_url = reverse_lazy("notasfiscais:list")

    def get_initial(self):
        initial = super().get_initial()
        di_q = parse_date((self.request.GET.get("data_inicio") or "").strip())
        df_q = parse_date((self.request.GET.get("data_fim") or "").strip())

        ontem = timezone.localdate() - timedelta(days=1)
        month_start = date(ontem.year, ontem.month, 1)
        eid = self.request.session.get("empresa_id")
        ultima_importacao: date | None = None
        if eid:
            ultimo_dt = (
                NotaFiscalServico.objects.filter(empresa_id=eid)
                .aggregate(m=Max("data_criacao"))
                .get("m")
            )
            if ultimo_dt is not None:
                ultima_importacao = timezone.localtime(ultimo_dt).date()

        if ultima_importacao is not None:
            # Última NFSe gravada no sistema: mesmo dia em início e fim (filtro típico no portal).
            d = min(ultima_importacao, ontem)
            default_di = default_df = d
        else:
            default_di = month_start
            default_df = ontem
            if default_di > default_df:
                default_di = default_df

        initial["data_inicio"] = di_q if di_q else default_di
        initial["data_fim"] = df_q if df_q else default_df
        if initial["data_inicio"] > initial["data_fim"]:
            initial["data_inicio"] = initial["data_fim"]
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .nfse_xml_copia import _nfse_xml_base_prestador, _nfse_xml_base_tomador

        eid = self.request.session.get("empresa_id")
        empresa = Empresa.objects.filter(pk=eid).first() if eid else None
        ctx["empresa"] = empresa
        ctx["pasta_xml_prestador"] = _nfse_xml_base_prestador(empresa) if empresa else ""
        ctx["pasta_xml_tomador"] = _nfse_xml_base_tomador(empresa) if empresa else ""
        # Tela de login do Portal Contribuinte / Emissor Nacional (mesmo fluxo do navegador e extensões).
        ctx["portal_nfse_url"] = "https://www.nfse.gov.br/EmissorNacional/Login"
        ctx["portal_emitidas_url"] = "https://www.nfse.gov.br/EmissorNacional/Notas/Emitidas"
        if empresa:
            ctx["portal_site_login_cadastrado"] = bool(
                (getattr(empresa, "nfse_portal_nacional_login", None) or "").strip()
            )
            ctx["portal_site_senha_gravada"] = bool(
                (getattr(empresa, "nfse_portal_nacional_senha_cifrada", None) or "").strip()
            )
        else:
            ctx["portal_site_login_cadastrado"] = False
            ctx["portal_site_senha_gravada"] = False
        ctx["portal_selenium_via_http"] = getattr(django_settings, "NFSE_PORTAL_PLAYWRIGHT_HTTP", True)
        ctx["nfse_baixar_webstore_url"] = getattr(
            django_settings,
            "NFSE_PORTAL_NFSE_BAIXAR_WEBSTORE_URL",
            "https://chromewebstore.google.com/detail/enehmclajcndmgefbmjhecccoegbdgea",
        )
        # Período para o POST do Selenium (YYYY-MM-DD), alinhado a get_initial / query string
        ini = self.get_initial()
        di0 = ini.get("data_inicio")
        df0 = ini.get("data_fim")
        ctx["portal_pw_data_inicio_iso"] = di0.isoformat() if isinstance(di0, date) else ""
        ctx["portal_pw_data_fim_iso"] = df0.isoformat() if isinstance(df0, date) else ""
        from .portal_emitidas_selenium import diagnostico_selenium_carregavel

        _sel_ok, _sel_err = diagnostico_selenium_carregavel()
        ctx["portal_selenium_carregavel_ok"] = _sel_ok
        ctx["portal_selenium_carregavel_erro"] = _sel_err
        return ctx

    def form_valid(self, form):
        empresa_id = self.request.session.get("empresa_id")
        if not empresa_id:
            messages.error(self.request, "Selecione uma empresa para continuar.")
            return redirect("empresa:lista")

        empresa = Empresa.objects.filter(pk=empresa_id).first()
        if not empresa:
            messages.error(self.request, "Empresa não encontrada na sessão.")
            return redirect("empresa:lista")

        arquivos = self.request.FILES.getlist("arquivos")
        if not arquivos:
            form.add_error(
                None,
                "Selecione ao menos um arquivo XML (e os PDFs correspondentes, se a extensão os gerou).",
            )
            return self.form_invalid(form)

        from .nfse_xml_copia import (
            extrair_chave_acesso_nfse_html,
            extrair_chave_acesso_nfse_xml,
            html_extensao_portal_indica_nfse_cancelada,
            salvar_baixados_portal_nacional_files,
            validar_periodo_xml_nfse,
            xml_nfse_portal_indica_cancelada,
        )

        di = form.cleaned_data["data_inicio"]
        df = form.cleaned_data["data_fim"]

        pdf_por_stem: dict[str, bytes] = {}
        html_por_stem: dict[str, bytes] = {}
        xml_files: list = []
        for f in arquivos:
            nome = (getattr(f, "name", "") or "").strip()
            ext = os.path.splitext(nome)[1].lower()
            stem = os.path.splitext(os.path.basename(nome))[0].lower()
            if ext == ".xml":
                xml_files.append(f)
            elif ext == ".pdf":
                try:
                    pdf_por_stem[stem] = f.read()
                except Exception:
                    pass
            elif ext in (".html", ".htm"):
                try:
                    html_por_stem[stem] = f.read()
                except Exception:
                    pass

        html_por_chave: dict[str, bytes] = {}
        for _st, hb in html_por_stem.items():
            ch_h = extrair_chave_acesso_nfse_html(hb)
            if ch_h:
                html_por_chave[ch_h] = hb

        if not xml_files:
            form.add_error(
                None,
                "Nenhum XML foi enviado. Inclua os arquivos .xml baixados (formato XML+PDF da extensão).",
            )
            return self.form_invalid(form)

        total_imp = 0
        total_ign = 0
        total_erros = 0
        erros_periodo: list[str] = []

        for xf in xml_files:
            try:
                xml_bytes = xf.read()
            except Exception as e:
                total_erros += 1
                messages.warning(self.request, f"Erro ao ler {getattr(xf, 'name', 'arquivo')}: {e}")
                continue

            msg_periodo = validar_periodo_xml_nfse(xml_bytes, di, df)
            if msg_periodo:
                erros_periodo.append(f"{getattr(xf, 'name', 'XML')}: {msg_periodo}")
                total_erros += 1
                continue

            stem = os.path.splitext(os.path.basename(getattr(xf, "name", "nfse.xml")))[0] or "nfse"
            stem_l = stem.lower()
            pdf_bytes = pdf_por_stem.get(stem_l)
            chave_xml = extrair_chave_acesso_nfse_xml(xml_bytes)
            html_bytes = html_por_chave.get(chave_xml) if chave_xml else None
            if not html_bytes:
                html_bytes = html_por_stem.get(stem_l)
            cancel_ef = bool(
                xml_nfse_portal_indica_cancelada(xml_bytes)
                or (html_bytes and html_extensao_portal_indica_nfse_cancelada(html_bytes))
            )

            salvar_baixados_portal_nacional_files(
                xml_bytes,
                pdf_bytes,
                stem,
                empresa,
                importar_canceladas=cancel_ef,
                html_bytes=html_bytes,
            )

            nome_xml = os.path.basename(getattr(xf, "name", "nfse.xml")) or "nfse.xml"
            uploaded = SimpleUploadedFile(nome_xml, xml_bytes, content_type="application/xml")
            try:
                resultado = import_nfse_from_xml(
                    uploaded,
                    self.request.user,
                    empresa,
                    importar_canceladas=cancel_ef,
                )
            except ValueError as e:
                total_erros += 1
                messages.warning(self.request, f"{nome_xml}: {e}")
                continue
            except Exception as e:
                total_erros += 1
                messages.warning(self.request, f"{nome_xml}: erro na importação — {e}")
                continue

            total_imp += int(resultado.get("total_importadas") or 0)
            total_ign += int(resultado.get("total_ignoradas") or 0)

        for ep in erros_periodo[:8]:
            messages.warning(self.request, ep)
        if len(erros_periodo) > 8:
            messages.warning(self.request, f"... e mais {len(erros_periodo) - 8} arquivo(s) fora do período.")

        if total_imp > 0:
            messages.success(self.request, f"{total_imp} NFSe importada(s) a partir dos XML enviados.")
        elif total_ign > 0:
            messages.warning(
                self.request,
                f"Nenhuma nota nova; {total_ign} já existente(s) ou ignorada(s). Verifique duplicidade ou período.",
            )
        elif total_imp == 0 and total_erros > 0:
            messages.error(
                self.request,
                "Nenhuma NFSe foi importada. Corrija o período, os arquivos ou o cadastro da empresa.",
            )
        elif total_imp == 0:
            messages.info(
                self.request,
                "Nenhuma NFSe nova importada. Confira se os XML correspondem à empresa e ao período.",
            )

        return redirect("notasfiscais:list")


@login_required
@require_POST
def portal_extensao_credenciais(request):
    """
    Devolve login e senha do portal nacional (cadastro da empresa) em JSON, para o navegador
    copiar na área de transferência. Não é possível preencher o site nfse.gov.br automaticamente
    a partir desta aplicação (origem diferente — política de segurança do navegador).
    """
    empresa_id = request.session.get("empresa_id")
    if not empresa_id:
        return JsonResponse({"ok": False, "error": "Selecione uma empresa."}, status=400)

    empresa = Empresa.objects.filter(pk=empresa_id).first()
    if not empresa:
        return JsonResponse({"ok": False, "error": "Empresa não encontrada."}, status=400)

    from empresa.nfse_nacional_config import portal_nacional_site_credenciais_para_empresa

    cred = portal_nacional_site_credenciais_para_empresa(empresa)
    login = (cred.get("login") or "").strip()
    senha = (cred.get("senha") or "").strip()
    if not login or not senha:
        return JsonResponse(
            {
                "ok": False,
                "error": "Cadastre o login e a senha do portal em Configuração de integração da empresa.",
            },
            status=400,
        )
    return JsonResponse({"ok": True, "login": login, "senha": senha})


@login_required
@require_POST
def portal_extensao_executar_selenium(request):
    """
    Inicia na máquina onde o Django roda o comando ``nfse_portal_emitidas_automacao``
    (Selenium: login, período, tentativa de download e importação).

    No Windows abre um novo console; em Linux usa nova sessão. Não funciona se o servidor
    for remoto sem ambiente gráfico — nesse caso use o comando manualmente no seu PC.
    """
    if not getattr(django_settings, "NFSE_PORTAL_PLAYWRIGHT_HTTP", True):
        messages.error(
            request,
            "Automação Selenium via navegador está desativada (NFSE_PORTAL_PLAYWRIGHT_HTTP).",
        )
        return redirect("notasfiscais:portal_extensao_import")

    empresa_id = request.session.get("empresa_id")
    if not empresa_id:
        messages.error(request, "Selecione uma empresa para continuar.")
        return redirect("empresa:lista")

    empresa = Empresa.objects.filter(pk=empresa_id).first()
    if not empresa:
        messages.error(request, "Empresa não encontrada na sessão.")
        return redirect("empresa:lista")

    di = parse_date((request.POST.get("data_inicio") or "").strip())
    df = parse_date((request.POST.get("data_fim") or "").strip())
    if not di or not df:
        messages.error(request, "Informe data inicial e final válidas antes de executar a automação.")
        return redirect("notasfiscais:portal_extensao_import")
    if di > df:
        messages.error(request, "A data inicial não pode ser maior que a data final.")
        return redirect("notasfiscais:portal_extensao_import")
    if (di.year, di.month) != (df.year, df.month):
        messages.error(
            request,
            "Para a automação, use um único mês civil na data inicial e final (ex.: 01/04/2026 a 30/04/2026). "
            "Os downloads gravam em «código externo-razão/MMAAAA» na pasta prestador.",
        )
        return redirect("notasfiscais:portal_extensao_import")

    from empresa.nfse_nacional_config import portal_nacional_site_credenciais_para_empresa
    from notasfiscais.nfse_xml_copia import PastaNfseInacessivelError, pasta_inbox_downloads_portal_nacional
    from notasfiscais.portal_emitidas_selenium import diagnostico_selenium_carregavel

    ok_sel, err_sel = diagnostico_selenium_carregavel()
    if not ok_sel:
        messages.error(request, err_sel)
        return redirect("notasfiscais:portal_extensao_import")

    cred = portal_nacional_site_credenciais_para_empresa(empresa)
    if not (cred.get("login") or "").strip() or not (cred.get("senha") or "").strip():
        messages.error(
            request,
            "Cadastre na empresa o login e a senha do Portal nacional (site) em Configuração de integração.",
        )
        return redirect("notasfiscais:portal_extensao_import")

    try:
        pasta_inbox_downloads_portal_nacional(empresa, di, df)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("notasfiscais:portal_extensao_import")
    except PastaNfseInacessivelError as exc:
        messages.error(request, str(exc))
        return redirect("notasfiscais:portal_extensao_import")

    base = django_settings.BASE_DIR
    manage_py = os.path.join(base, "manage.py")
    if not os.path.isfile(manage_py):
        messages.error(request, "manage.py não encontrado; não é possível disparar a automação.")
        return redirect("notasfiscais:portal_extensao_import")

    cmd = [
        sys.executable,
        manage_py,
        "nfse_portal_emitidas_automacao",
        "--empresa-id",
        str(empresa_id),
        "--inicio",
        di.isoformat(),
        "--fim",
        df.isoformat(),
        "--modo",
        "emitidas",
        "--usuario",
        request.user.username,
    ]
    if request.POST.get("selenium_pausa"):
        cmd.append("--pausa")
    # Perfil persistente (user-data-dir dedicado) só se o utilizador marcar a opção.
    if request.POST.get("selenium_perfil") or request.POST.get("playwright_perfil"):
        cmd.append("--perfil")
    else:
        cmd.append("--sem-perfil")

    popen_kw: dict = {"cwd": base}
    try:
        if sys.platform == "win32":
            popen_kw["creationflags"] = subprocess.CREATE_NEW_CONSOLE  # type: ignore[attr-defined]
        else:
            popen_kw["start_new_session"] = True
        subprocess.Popen(cmd, **popen_kw)
    except OSError as e:
        messages.error(request, f"Não foi possível iniciar a automação: {e}")
        return redirect("notasfiscais:portal_extensao_import")

    messages.success(
        request,
        "Automação iniciada: abra a nova janela do terminal (Selenium com Chrome ou Edge; pip install selenium). "
        "Se marcou perfil em disco, feche outras janelas que usem a mesma pasta de perfil. "
        "Por omissão, após a grade o Chrome fecha quando os downloads terminarem e os XML da pasta do mês são importados no mesmo comando; "
        "a janela do terminal pode fechar sozinha quando o processo terminar. "
        "Com «Pausar no fim», prima Enter no terminal para fechar o Chrome e só então corre a importação. "
        "Se interromper o comando, use «Importar XML da pasta do mês» com o mesmo período.",
    )
    return redirect("notasfiscais:portal_extensao_import")


@login_required
@require_POST
def portal_extensao_importar_pasta_mes(request):
    """
    Importa NFSe a partir dos .xml (e Cancelada/) já gravados na pasta do mês do cadastro
    (a mesma usada pelo Selenium: prestador → código-razão/MMAAAA).
    """
    empresa_id = request.session.get("empresa_id")
    if not empresa_id:
        messages.error(request, "Selecione uma empresa para continuar.")
        return redirect("empresa:lista")

    empresa = Empresa.objects.filter(pk=empresa_id).first()
    if not empresa:
        messages.error(request, "Empresa não encontrada na sessão.")
        return redirect("empresa:lista")

    di = parse_date((request.POST.get("data_inicio") or "").strip())
    df = parse_date((request.POST.get("data_fim") or "").strip())
    if not di or not df:
        messages.error(request, "Informe data inicial e final válidas (igual ao período dos downloads).")
        return redirect("notasfiscais:portal_extensao_import")
    if di > df:
        messages.error(request, "A data inicial não pode ser maior que a data final.")
        return redirect("notasfiscais:portal_extensao_import")
    if (di.year, di.month) != (df.year, df.month):
        messages.error(
            request,
            "Use um único mês civil na data inicial e final (ex.: 01/04/2026 a 30/04/2026), alinhado à pasta MMAAAA.",
        )
        return redirect("notasfiscais:portal_extensao_import")

    from notasfiscais.nfse_xml_copia import PastaNfseInacessivelError, pasta_inbox_downloads_portal_nacional
    from notasfiscais.portal_extensao_service import (
        coletar_xml_pdf_de_diretorio,
        organizar_arquivos_cancelados_na_inbox_portal,
        processar_portal_extensao_arquivos,
    )

    try:
        work_dir = pasta_inbox_downloads_portal_nacional(empresa, di, df)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("notasfiscais:portal_extensao_import")
    except PastaNfseInacessivelError as exc:
        messages.error(request, str(exc))
        return redirect("notasfiscais:portal_extensao_import")

    if work_dir is None:
        messages.error(
            request,
            "Configure «Pasta cópias XML NFSe (prestador)» na empresa (integração) para existir uma pasta do mês "
            "onde os downloads do portal são gravados; depois use este botão para importar sem reenviar ficheiros.",
        )
        return redirect("notasfiscais:portal_extensao_import")

    n_org = organizar_arquivos_cancelados_na_inbox_portal(work_dir)
    if n_org:
        messages.info(
            request,
            f"{n_org} ficheiro(s) movido(s) para «Cancelada/» na pasta do mês (manifesto ou conteúdo XML/HTML).",
        )

    itens = coletar_xml_pdf_de_diretorio(work_dir)
    if not itens:
        messages.warning(
            request,
            f"Nenhum ficheiro .xml encontrado em {work_dir} (nem em Cancelada/). "
            "Confirme que o Selenium ou a extensão gravou os XML nessa pasta e que o período acima coincide com o mês da pasta.",
        )
        return redirect("notasfiscais:portal_extensao_import")

    n_ca = sum(1 for it in itens if len(it) == 4 and it[3])
    if n_ca:
        messages.info(
            request,
            f"{n_ca} XML(s) em «Cancelada/» serão importados como notas canceladas (valores zerados).",
        )

    res = processar_portal_extensao_arquivos(
        empresa,
        request.user,
        di,
        df,
        itens,
        False,
        on_warning=lambda m: messages.warning(request, m),
        pasta_manifest=work_dir,
    )

    if res["total_importadas"] > 0:
        messages.success(
            request,
            f"{res['total_importadas']} NFSe importada(s) a partir da pasta do mês. "
            f"Ignoradas: {res['total_ignoradas']}; erros: {res['total_erros']}; fora do período: {res['erros_periodo_count']}.",
        )
        return redirect("notasfiscais:list")
    if res["total_ignoradas"] > 0:
        messages.warning(
            request,
            f"Nenhuma nota nova; {res['total_ignoradas']} já existente(s) ou ignorada(s). "
            f"Erros: {res['total_erros']}; fora do período: {res['erros_periodo_count']}.",
        )
        return redirect("notasfiscais:list")
    if res["total_erros"] > 0 or res["erros_periodo_count"] > 0:
        messages.error(
            request,
            "Nenhuma NFSe importada. Corrija o período, os XML na pasta ou as mensagens de aviso acima.",
        )
        return redirect("notasfiscais:portal_extensao_import")

    messages.info(request, "Nenhuma NFSe nova foi importada a partir da pasta (sem erros reportados).")
    return redirect("notasfiscais:portal_extensao_import")


@method_decorator(csrf_exempt, name='dispatch')
@login_required
def import_xml_ajax(request):
    """Importa XML via AJAX"""
    if request.method == 'POST':
        empresa_id = request.session.get('empresa_id')
        if not empresa_id:
            return JsonResponse({'error': 'Selecione uma empresa para continuar.'}, status=400)
        
        try:
            xml_file = request.FILES.get('xml_file')
            if not xml_file:
                return JsonResponse({'error': 'Arquivo XML não fornecido.'}, status=400)
            
            # Busca a empresa
            empresa = Empresa.objects.get(id=empresa_id)
            
            # Tenta importar o XML (não usa flag de canceladas neste fluxo)
            resultado = import_nfse_from_xml(xml_file, request.user, empresa)
            
            # Preparar resposta baseada no resultado
            if resultado['total_importadas'] > 0:
                if resultado['total_importadas'] == 1:
                    message = f'NFSe {resultado["notas_importadas"][0]["numero_nota"]} importada com sucesso!'
                else:
                    message = f'{resultado["total_importadas"]} NFSe importadas com sucesso!'
                
                if resultado['total_ignoradas'] > 0:
                    message += f' ({resultado["total_ignoradas"]} ignoradas)'
                
                return JsonResponse({
                    'success': True,
                    'message': message,
                    'resultado': serialize_resultado(resultado)
                })
            else:
                if resultado['total_ignoradas'] > 0:
                    return JsonResponse({
                        'success': False,
                        'message': 'Nenhuma nova NFSe foi importada. Todas as notas já existem no sistema.',
                        'resultado': serialize_resultado(resultado)
                    }, status=400)
                else:
                    return JsonResponse({
                        'success': False,
                        'message': 'Nenhuma NFSe válida encontrada no XML.',
                        'resultado': serialize_resultado(resultado)
                    }, status=400)
            
        except ValueError as e:
            # Erro de validação (incluindo CNPJ não correspondente)
            return JsonResponse({'error': str(e)}, status=400)
            
        except Exception as e:
            # Outros erros
            return JsonResponse({'error': f'Erro inesperado: {str(e)}'}, status=400)
    
    return JsonResponse({'error': 'Método não permitido.'}, status=405)


# ---- PARSERS ---------------------------------------------------

def _normaliza(s: str) -> str:
    return (s or "").strip()


def extrair_autorizacao(discriminacao: str) -> Optional[str]:
    from notasfiscais.utils import extrair_autorizacao as _extrair_autorizacao_util
    return _extrair_autorizacao_util(discriminacao)


def extrair_data_cancelamento(discriminacao: str)  -> Optional[str]:
    pass
#     """
#     Procura data padrão brasileiro indicando cancelamento.
#     Exemplos: 'Cancelado em 17/07/2025', 'Data de cancelamento: 12/05/2025'
#     """
#     if not discriminacao:
#         return None
#     m = re.search(r'(cancelad[oa].*?|data\s*de\s*cancelamento[: ]\s*)(\d{2}/\d{2}/\d{4})',
#                   discriminacao, flags=re.IGNORECASE)
#     return m.group(2) if m else None


def extrair_base_servico(discriminacao: str, empresa) -> str:
    """
    Extrai a base de serviço baseada na discriminação e configuração da empresa
    """
    if not discriminacao:
        return 'NORMAL'

    # Se a empresa não usa base de cálculo reduzido, tudo é NORMAL
    if not getattr(empresa, 'usa_base_calculo_reduzido', False):
        return 'NORMAL'

    # Se usa base reduzido, verificar o tipo de serviço na discriminação
    texto = discriminacao.lower()

    # Mapeamento baseado na discriminação conforme especificação
    if 'consulta' in texto:
        return 'NORMAL'
    elif 'procedimento' in texto or 'exames' in texto or 'exame' in texto or 'cirurgia' in texto:
        return 'DEMAIS'
    else:
        # Se não encontrar nenhum padrão específico, usar NORMAL
        return 'NORMAL'

# ---- VIEW BULK -------------------------------------------------

@login_required
@require_POST
def aplicar_discriminacao_bulk(request):
    """
    Recebe IDs de NFS-e selecionadas e, para cada uma:
      - Se houver nome de sócio na discriminação, troca o sócio
      - Se houver forma de pgto, vincula à Cobranca correspondente (por descricao/nome)
      - Se houver AUT: <valor>, grava em 'autorizacao'
      - Se houver data de cancelamento, seta status='Cancelado'
    """
    try:
        ids = request.POST.getlist("ids[]") or request.POST.getlist("ids")
        if not ids:
            return JsonResponse({"ok": False, "error": "IDs não enviados."}, status=400)

        # Converter IDs para inteiros
        try:
            ids = [int(id) for id in ids]
        except ValueError:
            return JsonResponse({"ok": False, "error": "IDs inválidos."}, status=400)

        empresa_id = request.session.get('empresa_id')
        if not empresa_id:
            return JsonResponse({"ok": False, "error": "Empresa não selecionada."}, status=400)

        qs = NotaFiscalServico.objects.filter(id__in=ids, empresa_id=empresa_id)

        if not qs.exists():
            return JsonResponse({"ok": False, "error": "Nenhuma NFSe encontrada com os IDs fornecidos."}, status=404)

        # cache de formas de pagamento (Cobranca) por nome normalizado
        todas_formas = list(Cobranca.objects.all())

        def encontra_cobranca(nome: str):
            if not nome:
                return None
            alvo = nome.lower().strip()
            for cb in todas_formas:
                rotulo = (getattr(cb, "descricao", None) or getattr(cb, "nome", None) or "").lower().strip()
                if rotulo and rotulo == alvo:
                    return cb
            from notasfiscais.utils import _encontra_cobranca_flexivel

            hit = _encontra_cobranca_flexivel(nome.strip().upper(), todas_formas)
            if hit:
                return hit
            return None

        atualizados = 0
        erros = []
        resultados = []

        with transaction.atomic():
            for nf in qs.select_related("socio", "forma_pagamento"):
                mudou = False
                txt = _normaliza(nf.discriminacao)

                # Sócios da empresa do lançamento
                socios_qs = Socio.objects.filter(empresa_id=nf.empresa_id)

                try:
                    # SOCIO
                    socio_detectado = extrair_socio(txt, socios_qs)
                    if socio_detectado and nf.socio_id != socio_detectado.id:
                        nf.socio = socio_detectado
                        mudou = True

                    # FORMA PGTO
                    forma_txt = extrair_forma_pagamento(txt)
                    print(f"NF {nf.numero_nota}: Forma extraida: {forma_txt}")
                    if forma_txt:
                        cobranca = encontra_cobranca(forma_txt)
                        print(f"NF {nf.numero_nota}: Cobranca encontrada: {cobranca.descricao if cobranca else 'None'}")
                        if cobranca and (not nf.forma_pagamento_id or nf.forma_pagamento_id != cobranca.id):
                            nf.forma_pagamento = cobranca
                            mudou = True

                    # AUTORIZACAO
                    aut = extrair_autorizacao(txt)
                    if aut and getattr(nf, "nsu", None) != aut:
                        nf.nsu = aut
                        mudou = True

                    # CANCELAMENTO
                    data_canc = extrair_data_cancelamento(txt)
                    if data_canc:
                        # Se existir campo data_cancelamento, grave; e ajuste status.
                        if hasattr(nf, "data_cancelamento") and not nf.data_cancelamento:
                            try:
                                from datetime import datetime
                                nf.data_cancelamento = datetime.strptime(data_canc, "%d/%m/%Y").date()
                            except Exception:
                                pass

                    # BASE SERVICO
                    base_servico_calculada = extrair_base_servico(txt, nf.empresa)
                    if hasattr(nf, "base_servico") and nf.base_servico != base_servico_calculada:
                        nf.base_servico = base_servico_calculada
                        mudou = True

                    if mudou:
                        # Salvar apenas campos que existem no modelo
                        campos_para_salvar = []
                        if hasattr(nf, 'socio') and nf.socio is not None:
                            campos_para_salvar.append('socio')
                        if hasattr(nf, 'forma_pagamento') and nf.forma_pagamento is not None:
                            campos_para_salvar.append('forma_pagamento')
                        if hasattr(nf, 'nsu') and nf.nsu:  # Verifica se nsu tem valor
                            campos_para_salvar.append('nsu')
                        if hasattr(nf, 'data_cancelamento') and nf.data_cancelamento is not None:
                            campos_para_salvar.append('data_cancelamento')
                        if hasattr(nf, 'base_servico'):
                            campos_para_salvar.append('base_servico')

                        if campos_para_salvar:
                            nf.save(update_fields=campos_para_salvar)
                            atualizados += 1

                    resultados.append({
                        "id": nf.id,
                        "numero_nota": nf.numero_nota,
                        "socio": getattr(nf.socio, "socio", None),
                        "forma_pagamento": getattr(nf.forma_pagamento, "descricao", None) or getattr(nf.forma_pagamento, "nome", None),
                        "nsu": getattr(nf, "nsu", None),
                        "atualizado": mudou
                    })

                except Exception as e:
                    erros.append(f"Erro ao processar NFSe {nf.numero_nota}: {str(e)}")

        response_data = {
            "ok": True,
            "atualizados": atualizados,
            "total_processados": len(resultados),
            "resultados": resultados,
            "message": f"Processamento concluído. {atualizados} NFSe(s) atualizada(s) de {len(resultados)} processada(s)."
        }

        if erros:
            response_data["avisos"] = erros
            if atualizados == 0:
                response_data["ok"] = False
                response_data["error"] = "Nenhuma NFSe foi atualizada. Verifique os avisos."

        return JsonResponse(response_data)

    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Erro interno: {str(e)}"}, status=500)

@login_required
@require_POST
def gerar_contas_receber_bulk(request):
    """
    Gera contas a receber para as NFSe selecionadas
    """
    try:
        ids = request.POST.getlist("ids[]") or request.POST.getlist("ids")
        if not ids:
            return JsonResponse({"ok": False, "error": "IDs não enviados."}, status=400)

        # Converter IDs para inteiros
        try:
            ids = [int(id) for id in ids]
        except ValueError:
            return JsonResponse({"ok": False, "error": "IDs inválidos."}, status=400)

        empresa_id = request.session.get('empresa_id')
        if not empresa_id:
            return JsonResponse({"ok": False, "error": "Empresa não selecionada."}, status=400)

        qs = NotaFiscalServico.objects.filter(id__in=ids, empresa_id=empresa_id)

        if not qs.exists():
            return JsonResponse({"ok": False, "error": "Nenhuma NFSe encontrada com os IDs fornecidos."}, status=404)

        from contasareceber.models import ContaAReceber
        from empresa.models import Empresa

        empresa = Empresa.objects.get(id=empresa_id)
        contas_criadas = 0
        erros = []
        resultados = []

        with transaction.atomic():
            for nf in qs:
                try:
                    # Verificar se já existe conta a receber para esta nota
                    conta_existente = ContaAReceber.objects.filter(
                        empresa=empresa,
                        nota=nf
                    ).exists()

                    if conta_existente:
                        resultados.append({
                            "id": nf.id,
                            "numero_nota": nf.numero_nota,
                            "status": "já existe",
                            "conta_id": None
                        })
                        continue

                    if nf.is_cancelada():
                        resultados.append({
                            "id": nf.id,
                            "numero_nota": nf.numero_nota,
                            "status": "ignorada",
                            "motivo": "nota cancelada",
                            "conta_id": None
                        })
                        continue

                    if nf.valor_liquido is None or nf.valor_liquido <= Decimal('0'):
                        resultados.append({
                            "id": nf.id,
                            "numero_nota": nf.numero_nota,
                            "status": "ignorada",
                            "motivo": "valor líquido zerado",
                            "conta_id": None
                        })
                        continue

                    # Definir a data de vencimento baseada no tipo da forma de pagamento
                    data_vencimento = nf.data_emissao
                    if nf.forma_pagamento:
                        if nf.forma_pagamento.formapgto == '0':  # A Vista
                            # Data de vencimento = data de emissão
                            data_vencimento = nf.data_emissao
                        elif nf.forma_pagamento.formapgto == '1':  # A Prazo
                            # Data de vencimento = data de emissão + intervalo entre parcelas
                            intervalo_dias = int(nf.forma_pagamento.intervaloparcelas or 0)
                            data_vencimento = nf.data_emissao + timezone.timedelta(days=intervalo_dias)

                    # Criar conta a receber
                    conta = ContaAReceber.objects.create(
                        empresa=empresa,
                        nota=nf,
                        cliente=nf.cliente,
                        cnpj_cpf=nf.cnpj_cpf,
                        data_emissao=nf.data_emissao,
                        data_vencimento=data_vencimento,
                        valor_a_receber=nf.valor_liquido,
                        doc=f"NF {nf.numero_nota}",
                        forma_pagamento=nf.forma_pagamento,
                        autorizacao=nf.autorizacao_para_conta_receber(),
                    )

                    contas_criadas += 1
                    resultados.append({
                        "id": nf.id,
                        "numero_nota": nf.numero_nota,
                        "status": "criada",
                        "conta_id": conta.id
                    })

                except Exception as e:
                    erros.append(f"Erro ao criar conta para NFSe {nf.numero_nota}: {str(e)}")
                    resultados.append({
                        "id": nf.id,
                        "numero_nota": nf.numero_nota,
                        "status": "erro",
                        "erro": str(e)
                    })

        response_data = {
            "ok": True,
            "contas_criadas": contas_criadas,
            "total_processados": len(resultados),
            "resultados": resultados,
            "message": f"Processamento concluído. {contas_criadas} conta(s) a receber criada(s) de {len(resultados)} NFSe(s) processada(s)."
        }

        if erros:
            response_data["avisos"] = erros

        return JsonResponse(response_data)

    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Erro interno: {str(e)}"}, status=500)

@login_required
@require_POST
def aplicar_regra_imposto_bulk(request):
    """
    Recebe IDs de NFS-e selecionadas e uma regra de imposto, aplicando-a a todas
    """
    try:
        ids = request.POST.getlist("ids[]") or request.POST.getlist("ids")
        regra_imposto_id = request.POST.get("regra_imposto_id")
        data_inicio = request.POST.get("data_inicio")
        data_fim = request.POST.get("data_fim")

        if not ids:
            return JsonResponse({"ok": False, "error": "IDs não enviados."}, status=400)

        if not data_inicio or not data_fim:
            return JsonResponse({"ok": False, "error": "Período não especificado."}, status=400)

        # Converter IDs para inteiros
        try:
            ids = [int(id) for id in ids]
        except ValueError:
            return JsonResponse({"ok": False, "error": "IDs inválidos."}, status=400)

        empresa_id = request.session.get('empresa_id')
        if not empresa_id:
            return JsonResponse({"ok": False, "error": "Empresa não selecionada."}, status=400)

        # Verificar se o período está fechado
        from .models import ApuracaoPeriodo
        try:
            periodo = ApuracaoPeriodo.objects.get(
                empresa_id=int(empresa_id),
                data_inicio=data_inicio,
                data_fim=data_fim
            )
            if periodo.status == 'fechado':
                return JsonResponse({"ok": False, "error": "Período fechado - vá para Apuração de Impostos e reabra o período para alterar regras."}, status=400)
        except ApuracaoPeriodo.DoesNotExist:
            pass

        qs = NotaFiscalServico.objects.filter(id__in=ids, empresa_id=empresa_id)
        if not qs.exists():
            return JsonResponse({"ok": False, "error": "Nenhuma NFSe encontrada com os IDs fornecidos."}, status=404)

        limpar_regra = not (regra_imposto_id and str(regra_imposto_id).strip())

        if limpar_regra:
            # "Sem Regra": remover regra e zerar apurações
            from regraImposto.models import RegraImposto
            atualizados = 0
            resultados = []
            with transaction.atomic():
                for nf in qs.select_related("codigo_da_regra_do_imposto"):
                    mudou = nf.codigo_da_regra_do_imposto_id is not None
                    if mudou:
                        nf.codigo_da_regra_do_imposto = None
                        nf.pisapuracao = nf.calcular_pis_apuracao()
                        nf.cofinsapuracao = nf.calcular_cofins_apuracao()
                        nf.csllapuracao = nf.calcular_csll_apuracao()
                        nf.irpjapuracao = nf.calcular_irpj_apuracao()
                        nf.issapuracao = nf.calcular_iss_apuracao()
                        campos_para_salvar = ['codigo_da_regra_do_imposto', 'pisapuracao', 'cofinsapuracao', 'csllapuracao', 'irpjapuracao', 'issapuracao']
                        nf.save(update_fields=campos_para_salvar)
                        atualizados += 1
                    resultados.append({
                        "id": nf.id,
                        "numero_nota": nf.numero_nota,
                        "regra_imposto": "",
                        "atualizado": mudou
                    })
            return JsonResponse({
                "ok": True,
                "atualizados": atualizados,
                "total_processados": len(resultados),
                "resultados": resultados,
                "message": f"Regra removida com sucesso! {atualizados} NFSe(s) atualizada(s)."
            })

        # Aplicar regra de imposto selecionada (apenas regras da mesma empresa)
        from regraImposto.models import RegraImposto
        try:
            regra_imposto = RegraImposto.objects.get(id=regra_imposto_id)
        except RegraImposto.DoesNotExist:
            return JsonResponse({"ok": False, "error": "Regra de imposto não encontrada ou não pertence a esta empresa."}, status=404)

        atualizados = 0
        erros = []
        resultados = []

        with transaction.atomic():
            for nf in qs.select_related("codigo_da_regra_do_imposto"):
                mudou = False

                if nf.codigo_da_regra_do_imposto_id != regra_imposto.id:
                    nf.codigo_da_regra_do_imposto = regra_imposto
                    mudou = True

                if mudou:
                    nf.pisapuracao = nf.calcular_pis_apuracao()
                    nf.cofinsapuracao = nf.calcular_cofins_apuracao()
                    nf.csllapuracao = nf.calcular_csll_apuracao()
                    nf.irpjapuracao = nf.calcular_irpj_apuracao()
                    nf.issapuracao = nf.calcular_iss_apuracao()
                    campos_para_salvar = []
                    if hasattr(nf, 'codigo_da_regra_do_imposto'):
                        campos_para_salvar.append('codigo_da_regra_do_imposto')
                    if hasattr(nf, 'pisapuracao'):
                        campos_para_salvar.append('pisapuracao')
                    if hasattr(nf, 'cofinsapuracao'):
                        campos_para_salvar.append('cofinsapuracao')
                    if hasattr(nf, 'csllapuracao'):
                        campos_para_salvar.append('csllapuracao')
                    if hasattr(nf, 'irpjapuracao'):
                        campos_para_salvar.append('irpjapuracao')
                    if hasattr(nf, 'issapuracao'):
                        campos_para_salvar.append('issapuracao')
                    if campos_para_salvar:
                        nf.save(update_fields=campos_para_salvar)
                        atualizados += 1

                    resultados.append({
                        "id": nf.id,
                        "numero_nota": nf.numero_nota,
                        "regra_imposto": regra_imposto.DescricaoRegraImposto,
                        "atualizado": True
                    })
                else:
                    resultados.append({
                        "id": nf.id,
                        "numero_nota": nf.numero_nota,
                        "regra_imposto": regra_imposto.DescricaoRegraImposto,
                        "atualizado": False
                    })

        response_data = {
            "ok": True,
            "atualizados": atualizados,
            "total_processados": len(resultados),
            "resultados": resultados,
            "message": f"Regra de imposto aplicada com sucesso! {atualizados} NFSe(s) atualizada(s) de {len(resultados)} processada(s)."
        }

        if erros:
            response_data["avisos"] = erros
            if atualizados == 0:
                response_data["ok"] = False
                response_data["error"] = "Nenhuma NFSe foi atualizada. Verifique os avisos."

        return JsonResponse(response_data)

    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Erro interno: {str(e)}"}, status=500)


@login_required
@require_POST
def aplicar_socio_bulk(request):
    """
    Recebe IDs de NFS-e selecionadas e o ID do sócio, atualizando o campo sócio
    de todas as notas para o sócio escolhido. O sócio deve pertencer à mesma
    empresa das notas.
    """
    try:
        ids = request.POST.getlist("ids[]") or request.POST.getlist("ids")
        socio_id = request.POST.get("socio_id", "").strip()

        if not ids:
            return JsonResponse({"ok": False, "error": "IDs não enviados."}, status=400)

        if not socio_id:
            return JsonResponse({"ok": False, "error": "Selecione um sócio."}, status=400)

        try:
            ids = [int(i) for i in ids]
            socio_id = int(socio_id)
        except (ValueError, TypeError):
            return JsonResponse({"ok": False, "error": "IDs ou sócio inválidos."}, status=400)

        empresa_id = request.session.get('empresa_id')
        if not empresa_id:
            return JsonResponse({"ok": False, "error": "Empresa não selecionada."}, status=400)

        qs = NotaFiscalServico.objects.filter(id__in=ids, empresa_id=empresa_id)
        if not qs.exists():
            return JsonResponse({"ok": False, "error": "Nenhuma NFSe encontrada com os IDs fornecidos."}, status=404)

        try:
            socio = Socio.objects.get(id=socio_id, empresa_id=empresa_id)
        except Socio.DoesNotExist:
            return JsonResponse({"ok": False, "error": "Sócio não encontrado ou não pertence à empresa das notas."}, status=404)

        nf_ids = list(qs.values_list('id', flat=True))
        atualizados = qs.update(socio=socio)
        from contasareceber.socio_sync import propagar_socio_para_contas_das_notas

        propagar_socio_para_contas_das_notas(nf_ids, socio_id)
        return JsonResponse({
            "ok": True,
            "atualizados": atualizados,
            "message": f"Sócio alterado com sucesso! {atualizados} NFSe(s) atualizada(s)."
        })
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Erro interno: {str(e)}"}, status=500)


@login_required
@require_POST
def aplicar_cobranca_bulk(request):
    """
    Recebe IDs de NFS-e selecionadas e o ID da cobrança (forma de pagamento),
    atualizando o campo forma_pagamento de todas as notas.
    Envie forma_pagamento_id=__clear__ para remover a cobrança nas notas.
    """
    try:
        ids = request.POST.getlist("ids[]") or request.POST.getlist("ids")
        forma_id_raw = (request.POST.get("forma_pagamento_id") or "").strip()

        if not ids:
            return JsonResponse({"ok": False, "error": "IDs não enviados."}, status=400)

        if not forma_id_raw:
            return JsonResponse({"ok": False, "error": "Selecione uma cobrança ou “Sem forma definida”."}, status=400)

        try:
            ids = [int(i) for i in ids]
        except (ValueError, TypeError):
            return JsonResponse({"ok": False, "error": "IDs inválidos."}, status=400)

        empresa_id = request.session.get("empresa_id")
        if not empresa_id:
            return JsonResponse({"ok": False, "error": "Empresa não selecionada."}, status=400)

        qs = NotaFiscalServico.objects.filter(id__in=ids, empresa_id=empresa_id)
        if not qs.exists():
            return JsonResponse({"ok": False, "error": "Nenhuma NFSe encontrada com os IDs fornecidos."}, status=404)

        from contasareceber.socio_sync import propagar_forma_pagamento_para_contas_das_notas

        nf_ids = list(qs.values_list("pk", flat=True))
        with transaction.atomic():
            if forma_id_raw == "__clear__":
                atualizados = qs.update(forma_pagamento=None)
                propagar_forma_pagamento_para_contas_das_notas(nf_ids, None)
                msg = f"Cobrança removida em {atualizados} NFSe(s)."
            else:
                try:
                    forma_id = int(forma_id_raw)
                except (ValueError, TypeError):
                    return JsonResponse({"ok": False, "error": "Cobrança inválida."}, status=400)
                try:
                    cob = Cobranca.objects.get(pk=forma_id)
                except Cobranca.DoesNotExist:
                    return JsonResponse({"ok": False, "error": "Cobrança não encontrada."}, status=404)
                atualizados = qs.update(forma_pagamento=cob)
                propagar_forma_pagamento_para_contas_das_notas(nf_ids, cob.pk)
                msg = f"Cobrança aplicada com sucesso! {atualizados} NFSe(s) atualizada(s)."

        return JsonResponse({"ok": True, "atualizados": atualizados, "message": msg})
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Erro interno: {str(e)}"}, status=500)


@login_required
def extrair_discriminacao_ajax(request):
    """
    AJAX view para extrair dados da discriminação e retornar valores para preencher o formulário
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)

    discriminacao = request.POST.get('discriminacao', '').strip()
    if not discriminacao:
        return JsonResponse({'error': 'Discriminação não fornecida'}, status=400)

    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'error': 'Empresa não selecionada'}, status=400)

    try:
        # Extrair forma de pagamento
        forma_pagamento_txt = extrair_forma_pagamento(discriminacao)

        # Extrair sócio
        socios_qs = Socio.objects.filter(empresa_id=empresa_id)
        socio = extrair_socio(discriminacao, socios_qs)

        # Extrair autorização
        autorizacao = extrair_autorizacao(discriminacao)

        # Extrair base serviço
        empresa = Empresa.objects.get(id=empresa_id)
        base_servico = extrair_base_servico(discriminacao, empresa)

        # Preparar resposta
        response_data = {}

        # Forma de pagamento - encontrar correspondente no banco
        if forma_pagamento_txt:
            cobranca = None
            alvo = forma_pagamento_txt.lower()
            todas_formas = list(Cobranca.objects.all())

            for cb in todas_formas:
                rotulo = (getattr(cb, "descricao", None) or "").lower()
                if alvo in rotulo or rotulo in alvo:
                    cobranca = cb
                    break

            if cobranca:
                response_data['forma_pagamento'] = cobranca.id
                response_data['forma_pagamento_nome'] = cobranca.descricao

        # Sócio
        if socio:
            response_data['socio'] = socio.id
            response_data['socio_nome'] = socio.socio

        # Autorização
        if autorizacao:
            response_data['nsu'] = autorizacao  # assumindo que nsu é o campo de autorização

        # Base Serviço
        response_data['base_servico'] = base_servico

        return JsonResponse({'success': True, 'data': response_data})

    except Exception as e:
        return JsonResponse({'error': f'Erro ao processar: {str(e)}'}, status=500)


@login_required
def get_filtered_ids(request):
    """
    Retorna os IDs de todas as NFSe filtradas
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'error': 'Empresa não selecionada'}, status=400)

    queryset = NotaFiscalServico.objects.filter(empresa_id=empresa_id)

    # Aplicar os mesmos filtros do NFSeListView (sessão não disponível no AJAX — usa GET)
    search = request.GET.get('search', '')
    valor = request.GET.get('valor', '')
    status = request.GET.get('status', '')
    status_nota = request.GET.get('status_nota', '')
    forma_pagamento = request.GET.get('forma_pagamento', '')
    socio = (request.GET.get('socio') or '').strip()
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')

    # Definir datas padrão
    hoje = timezone.now().date()
    primeiro_dia_mes = date(hoje.year, hoje.month, 1)
    ultimo_dia_mes = date(hoje.year, hoje.month, monthrange(hoje.year, hoje.month)[1])

    if not data_inicio:
        data_inicio = primeiro_dia_mes.strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = ultimo_dia_mes.strftime('%Y-%m-%d')

    if search:
        queryset = _apply_filtro_search_nfse(queryset, search)

    queryset = _apply_filtro_valor_nfse(queryset, valor)

    if status:
        queryset = queryset.filter(status_conciliacao=status)

    if status_nota == 'ativa':
        queryset = queryset.filter(data_cancelamento__isnull=True)
    elif status_nota == 'cancelada':
        queryset = queryset.filter(data_cancelamento__isnull=False)

    if forma_pagamento:
        if forma_pagamento == 'none':
            queryset = queryset.filter(forma_pagamento__isnull=True)
        else:
            try:
                queryset = queryset.filter(forma_pagamento_id=int(forma_pagamento))
            except (ValueError, TypeError):
                pass

    queryset = _apply_filtro_socio_nfse(queryset, socio)

    # Sempre aplicar filtro de data
    queryset = queryset.filter(data_emissao__gte=data_inicio)
    queryset = queryset.filter(data_emissao__lte=data_fim)

    ids = list(queryset.values_list('id', flat=True))
    return JsonResponse({'ids': ids})


@login_required
def export_excel(request):
    """
    Exporta os dados filtrados das NFSe para Excel
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Selecione uma empresa para continuar.')
        return redirect('empresa:lista')

    # Aplicar os mesmos filtros da NFSeListView
    queryset = NotaFiscalServico.objects.filter(empresa_id=empresa_id)

    # Filtros (alinhados à listagem NFSe)
    search = request.GET.get('search', '')
    valor = request.GET.get('valor', '')
    status = request.GET.get('status', '')
    status_nota = request.GET.get('status_nota', '')
    forma_pagamento = request.GET.get('forma_pagamento', '')
    socio = (request.GET.get('socio') or '').strip()
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')

    # Definir datas padrão
    hoje = timezone.now().date()
    primeiro_dia_mes = date(hoje.year, hoje.month, 1)
    ultimo_dia_mes = date(hoje.year, hoje.month, monthrange(hoje.year, hoje.month)[1])

    if not data_inicio:
        data_inicio = primeiro_dia_mes.strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = ultimo_dia_mes.strftime('%Y-%m-%d')

    if search:
        queryset = _apply_filtro_search_nfse(queryset, search)

    queryset = _apply_filtro_valor_nfse(queryset, valor)

    if status:
        queryset = queryset.filter(status_conciliacao=status)
    if status_nota == 'ativa':
        queryset = queryset.filter(data_cancelamento__isnull=True)
    elif status_nota == 'cancelada':
        queryset = queryset.filter(data_cancelamento__isnull=False)

    if forma_pagamento:
        if forma_pagamento == 'none':
            queryset = queryset.filter(forma_pagamento__isnull=True)
        else:
            try:
                queryset = queryset.filter(forma_pagamento_id=int(forma_pagamento))
            except (ValueError, TypeError):
                pass

    queryset = _apply_filtro_socio_nfse(queryset, socio)

    # Sempre aplicar filtro de data
    queryset = queryset.filter(data_emissao__gte=data_inicio)
    queryset = queryset.filter(data_emissao__lte=data_fim)

    # Ordenar por número da nota (texto — suporta "4729-1" segmentada)
    queryset = queryset.order_by('numero_nota')

    # Criar workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = 'NFSe'

    # Cabeçalhos
    headers = [
        'Número',
        'Série',
        'Data Emissão',
        'Cliente',
        'CNPJ/CPF',
        'Sócio',
        'Valor Bruto',
        'ISS Retido',
        'Aliquota ISS',
        'Valor ISS Retido',
        'Valor PIS',
        'Valor COFINS',
        'Valor CSLL',
        'Valor IRPJ',
        'Outras Retenções',
        'Valor INSS',
        'Valor Líquido',
        'ISS APURAÇÃO',
        'PIS APURAÇÃO',
        'COFINS APURAÇÃO',
        'CSLL APURAÇÃO',
        'IRPJ APURAÇÃO',
        'IRPJ ADICIONAL',
        'Discriminação',
        'Forma Pagamento',
        'Autorização',
        'Status',
        'Conciliação'
    ]

    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Dados
    for row_num, nfse in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=nfse.numero_nota)
        ws.cell(row=row_num, column=2, value=nfse.serie)
        ws.cell(row=row_num, column=3, value=nfse.data_emissao.strftime('%d/%m/%Y') if nfse.data_emissao else '')
        ws.cell(row=row_num, column=4, value=nfse.cliente)
        ws.cell(row=row_num, column=5, value=nfse.cnpj_cpf)
        ws.cell(row=row_num, column=6, value=nfse.socio.socio if nfse.socio else '')
        ws.cell(row=row_num, column=7, value=float(nfse.valor_bruto) if nfse.valor_bruto else 0)
        ws.cell(row=row_num, column=8, value='Sim' if nfse.iss_retido else 'Não')
        ws.cell(row=row_num, column=9, value=float(nfse.aliquota) if nfse.aliquota else 0)
        ws.cell(row=row_num, column=10, value=float(nfse.valor_iss_retido) if nfse.valor_iss_retido else 0)
        ws.cell(row=row_num, column=11, value=float(nfse.valor_pis) if nfse.valor_pis else 0)
        ws.cell(row=row_num, column=12, value=float(nfse.valor_cofins) if nfse.valor_cofins else 0)
        ws.cell(row=row_num, column=13, value=float(nfse.valor_csll) if nfse.valor_csll else 0)
        ws.cell(row=row_num, column=14, value=float(nfse.valor_ir) if nfse.valor_ir else 0)
        ws.cell(row=row_num, column=15, value=float(nfse.outras_retencoes) if nfse.outras_retencoes else 0)
        ws.cell(row=row_num, column=16, value=float(nfse.valor_inss) if nfse.valor_inss else 0)
        ws.cell(row=row_num, column=17, value=float(nfse.valor_liquido) if nfse.valor_liquido else 0)
        ws.cell(row=row_num, column=18, value=float(nfse.issapuracao) if nfse.issapuracao else 0)
        ws.cell(row=row_num, column=19, value=float(nfse.pisapuracao) if nfse.pisapuracao else 0)
        ws.cell(row=row_num, column=20, value=float(nfse.cofinsapuracao) if nfse.cofinsapuracao else 0)
        ws.cell(row=row_num, column=21, value=float(nfse.csllapuracao) if nfse.csllapuracao else 0)
        ws.cell(row=row_num, column=22, value=float(nfse.irpjapuracao) if nfse.irpjapuracao else 0)
        ws.cell(row=row_num, column=23, value=float(nfse.irpjadicional) if nfse.irpjadicional else 0)
        ws.cell(row=row_num, column=24, value=nfse.discriminacao)
        ws.cell(row=row_num, column=25, value=nfse.forma_pagamento.descricao if nfse.forma_pagamento else '')
        ws.cell(row=row_num, column=26, value=nfse.nsu)
        ws.cell(row=row_num, column=27, value='')
        ws.cell(row=row_num, column=28, value=dict(NotaFiscalServico.STATUS_CONCILIACAO_CHOICES).get(nfse.status_conciliacao, nfse.status_conciliacao))

    # Ajustar largura das colunas
    for col_num, header in enumerate(headers, 1):
        column_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[column_letter].width = max(len(header) + 2, 15)

    # Nome do arquivo
    filename = f'nfse_export_{data_inicio}_to_{data_fim}.xlsx'

    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Salvar workbook na resposta
    wb.save(response)

    return response


@login_required
def export_apuracao_excel(request):
    """
    Exporta os dados de apuração de impostos para Excel
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Selecione uma empresa para continuar.')
        return redirect('empresa:lista')

    # Filtros de período
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    # Definir datas padrão (mês atual)
    hoje = timezone.now().date()
    primeiro_dia_mes = date(hoje.year, hoje.month, 1)
    ultimo_dia_mes = date(hoje.year, hoje.month, monthrange(hoje.year, hoje.month)[1])

    if not data_inicio:
        data_inicio = primeiro_dia_mes.strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = ultimo_dia_mes.strftime('%Y-%m-%d')

    # Buscar notas fiscais do período
    notas = NotaFiscalServico.objects.filter(
        empresa_id=empresa_id,
        data_emissao__gte=data_inicio,
        data_emissao__lte=data_fim
    ).select_related('socio', 'codigo_da_regra_do_imposto').order_by('data_emissao')

    # Calcular demonstrativo por sócio
    from collections import defaultdict

    demonstrativo_por_socio = defaultdict(lambda: {
        'socio_nome': '',
        'valor_bruto': 0,
        'pis': 0,
        'cofins': 0,
        'iss': 0,
        'csll': 0,
        'irpj': 0,
        'adicional_irpj': 0,
        'total_impostos': 0,
        'valor_liquido': 0
    })

    # Totais gerais
    totais_gerais = {
        'valor_bruto': 0,
        'pis': 0,
        'cofins': 0,
        'iss': 0,
        'csll': 0,
        'irpj': 0,
        'adicional_irpj': 0,
        'total_impostos': 0,
        'valor_liquido': 0
    }

    for nota in notas:
        socio_nome = nota.socio.socio if nota.socio else 'Sem Sócio'

        # Atualizar totais por sócio
        demonstrativo_por_socio[socio_nome]['socio_nome'] = socio_nome
        demonstrativo_por_socio[socio_nome]['valor_bruto'] += nota.valor_bruto or 0
        demonstrativo_por_socio[socio_nome]['pis'] += nota.pisapuracao or 0
        demonstrativo_por_socio[socio_nome]['cofins'] += nota.cofinsapuracao or 0
        demonstrativo_por_socio[socio_nome]['iss'] += nota.issapuracao or 0
        demonstrativo_por_socio[socio_nome]['csll'] += nota.csllapuracao or 0
        demonstrativo_por_socio[socio_nome]['irpj'] += nota.irpjapuracao or 0
        demonstrativo_por_socio[socio_nome]['adicional_irpj'] += nota.irpjadicional or 0

        # Calcular total impostos e valor líquido por sócio
        total_impostos_socio = (
            demonstrativo_por_socio[socio_nome]['pis'] +
            demonstrativo_por_socio[socio_nome]['cofins'] +
            demonstrativo_por_socio[socio_nome]['iss'] +
            demonstrativo_por_socio[socio_nome]['csll'] +
            demonstrativo_por_socio[socio_nome]['irpj'] +
            demonstrativo_por_socio[socio_nome]['adicional_irpj']
        )
        demonstrativo_por_socio[socio_nome]['total_impostos'] = total_impostos_socio
        demonstrativo_por_socio[socio_nome]['valor_liquido'] = demonstrativo_por_socio[socio_nome]['valor_bruto'] - total_impostos_socio

        # Atualizar totais gerais
        totais_gerais['valor_bruto'] += nota.valor_bruto or 0
        totais_gerais['pis'] += nota.pisapuracao or 0
        totais_gerais['cofins'] += nota.cofinsapuracao or 0
        totais_gerais['iss'] += nota.issapuracao or 0
        totais_gerais['csll'] += nota.csllapuracao or 0
        totais_gerais['irpj'] += nota.irpjapuracao or 0
        totais_gerais['adicional_irpj'] += nota.irpjadicional or 0

    # Calcular totais gerais
    totais_gerais['total_impostos'] = (
        totais_gerais['pis'] +
        totais_gerais['cofins'] +
        totais_gerais['iss'] +
        totais_gerais['csll'] +
        totais_gerais['irpj'] +
        totais_gerais['adicional_irpj']
    )
    totais_gerais['valor_liquido'] = totais_gerais['valor_bruto'] - totais_gerais['total_impostos']

    # Converter para lista ordenada por nome do sócio
    demonstrativo_lista = sorted(demonstrativo_por_socio.values(), key=lambda x: x['socio_nome'])

    # Criar workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = Workbook()

    # Primeira aba: Apuração de Impostos
    ws1 = wb.active
    ws1.title = 'Apuração de Impostos'

    # Cabeçalhos
    headers = [
        'Sócio',
        'Valor Bruto',
        'PIS',
        'COFINS',
        'ISS',
        'CSLL',
        'IRPJ',
        'Adicional IRPJ',
        'Total dos Impostos',
        'Valor Líquido'
    ]

    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Dados por sócio
    for row_num, item in enumerate(demonstrativo_lista, 2):
        ws1.cell(row=row_num, column=1, value=item['socio_nome'])
        ws1.cell(row=row_num, column=2, value=float(item['valor_bruto']))
        ws1.cell(row=row_num, column=3, value=float(item['pis']))
        ws1.cell(row=row_num, column=4, value=float(item['cofins']))
        ws1.cell(row=row_num, column=5, value=float(item['iss']))
        ws1.cell(row=row_num, column=6, value=float(item['csll']))
        ws1.cell(row=row_num, column=7, value=float(item['irpj']))
        ws1.cell(row=row_num, column=8, value=float(item['adicional_irpj']))
        ws1.cell(row=row_num, column=9, value=float(item['total_impostos']))
        ws1.cell(row=row_num, column=10, value=float(item['valor_liquido']))

    # Linha de totais
    row_num = len(demonstrativo_lista) + 3
    ws1.cell(row=row_num, column=1, value='TOTAIS')
    ws1.cell(row=row_num, column=2, value=float(totais_gerais['valor_bruto']))
    ws1.cell(row=row_num, column=3, value=float(totais_gerais['pis']))
    ws1.cell(row=row_num, column=4, value=float(totais_gerais['cofins']))
    ws1.cell(row=row_num, column=5, value=float(totais_gerais['iss']))
    ws1.cell(row=row_num, column=6, value=float(totais_gerais['csll']))
    ws1.cell(row=row_num, column=7, value=float(totais_gerais['irpj']))
    ws1.cell(row=row_num, column=8, value=float(totais_gerais['adicional_irpj']))
    ws1.cell(row=row_num, column=9, value=float(totais_gerais['total_impostos']))
    ws1.cell(row=row_num, column=10, value=float(totais_gerais['valor_liquido']))

    # Aplicar estilo aos totais
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    for col_num in range(1, 11):
        cell = ws1.cell(row=row_num, column=col_num)
        cell.font = total_font
        cell.fill = total_fill

    # Ajustar largura das colunas
    for col_num, header in enumerate(headers, 1):
        column_letter = ws1.cell(row=1, column=col_num).column_letter
        ws1.column_dimensions[column_letter].width = max(len(header) + 2, 15)

    # Segunda aba: Detalhes do Cálculo do Adicional
    ws2 = wb.create_sheet(title='Detalhes do Cálculo')

    # Buscar dados de detalhes do cálculo
    from .models import ApuracaoPeriodo
    periodo, created = ApuracaoPeriodo.objects.get_or_create(
        empresa_id=empresa_id,
        data_inicio=data_inicio,
        data_fim=data_fim,
        defaults={'status': 'aberto', 'adicional_calculado': False}
    )

    detalhes_data = periodo.get_detalhes_calculo_adicional()

    # Cabeçalhos da segunda aba
    ws2.cell(row=1, column=1, value='Informações do Período')
    ws2.cell(row=2, column=1, value='Meses no período:')
    ws2.cell(row=2, column=2, value=detalhes_data.get('meses_periodo', 0))
    ws2.cell(row=3, column=1, value='Limite mensal:')
    ws2.cell(row=3, column=2, value=f"R$ {detalhes_data.get('limite_mensal', 0):.2f}")
    ws2.cell(row=4, column=1, value='Limite do período:')
    ws2.cell(row=4, column=2, value=f"R$ {detalhes_data.get('limite_periodo', 0):.2f}")
    ws2.cell(row=5, column=1, value='Total faturamento:')
    ws2.cell(row=5, column=2, value=f"R$ {detalhes_data.get('total_faturamento', 0):.2f}")
    ws2.cell(row=6, column=1, value='Total adicional calculado:')
    ws2.cell(row=6, column=2, value=f"R$ {detalhes_data.get('total_adicional_periodo', 0):.2f}")

    # Tabela de regras
    row_start = 8
    ws2.cell(row=row_start, column=1, value='Regra')
    ws2.cell(row=row_start, column=2, value='Percentual')
    ws2.cell(row=row_start, column=3, value='Faturamento')
    ws2.cell(row=row_start, column=4, value='Base Calculada')
    ws2.cell(row=row_start, column=5, value='Base Adicional')
    ws2.cell(row=row_start, column=6, value='Adicional Calculado')
    ws2.cell(row=row_start, column=7, value='Índice')

    # Aplicar estilo ao cabeçalho da segunda aba
    for col_num in range(1, 8):
        cell = ws2.cell(row=row_start, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Dados das regras
    for row_num, regra in enumerate(detalhes_data.get('regras', []), row_start + 1):
        ws2.cell(row=row_num, column=1, value=regra.get('regra_nome', ''))
        ws2.cell(row=row_num, column=2, value=f"{regra.get('percentual', 0)}%")
        ws2.cell(row=row_num, column=3, value=float(regra.get('total_faturamento', 0)))
        ws2.cell(row=row_num, column=4, value=float(regra.get('base_calculada', 0)))
        ws2.cell(row=row_num, column=5, value=float(regra.get('base_adicional_calculada', 0)))
        ws2.cell(row=row_num, column=6, value=float(regra.get('adicional_calculado', 0)))
        ws2.cell(row=row_num, column=7, value=f"{(regra.get('indice', 0) * 100):.8f}%")

    # Ajustar largura das colunas da segunda aba
    for col_num in range(1, 8):
        column_letter = ws2.cell(row=1, column=col_num).column_letter
        ws2.column_dimensions[column_letter].width = 20

    # Nome do arquivo
    filename = f'apuracao_impostos_{data_inicio}_to_{data_fim}.xlsx'

    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Salvar workbook na resposta
    wb.save(response)

    return response

class NFSeRecebimentoView(LoginRequiredMixin, UpdateView):
    model = NotaFiscalServico
    form_class = NFSeRecebimentoForm
    template_name = 'notasfiscais/nfse_recebimento.html'
    success_url = reverse_lazy('notasfiscais:detail')
    context_object_name = 'nfse'

    def get_queryset(self):
        empresa_id = self.request.session.get('empresa_id')
        if empresa_id:
            return NotaFiscalServico.objects.filter(empresa_id=empresa_id)
        return NotaFiscalServico.objects.none()

    def get_success_url(self):
        return reverse_lazy('notasfiscais:detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        # Update conciliation status based on received amount
        valor_recebido = form.cleaned_data.get('valor_recebido')
        if valor_recebido and self.object.valor_liquido:
            if valor_recebido >= self.object.valor_liquido:
                form.instance.status_conciliacao = 'conciliado'
            else:
                form.instance.status_conciliacao = 'parcialmente_conciliado'

        messages.success(self.request, 'Recebimento registrado com sucesso!')
        return super().form_valid(form)

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                if field == '__all__':
                    messages.error(self.request, error)
                else:
                    messages.error(self.request, f"{form.fields[field].label}: {error}")
        return super().form_invalid(form)


class NFSeSegmentView(LoginRequiredMixin, FormView):
    template_name = 'notasfiscais/nfse_segment.html'
    success_url = reverse_lazy('notasfiscais:list')

    def get(self, request, *args, **kwargs):
        # Get the original NFSe
        nfse_id = kwargs.get('pk')
        empresa_id = request.session.get('empresa_id')

        if not empresa_id:
            messages.error(request, 'Selecione uma empresa para continuar.')
            return redirect('empresa:lista')

        try:
            nfse = NotaFiscalServico.objects.get(id=nfse_id, empresa_id=empresa_id)
        except NotaFiscalServico.DoesNotExist:
            messages.error(request, 'NFSe não encontrada.')
            return redirect('notasfiscais:list')

        # Create form with initial data
        form = NFSeSegmentForm(request.POST or None, numero_segmentos=2, empresa_id=empresa_id)
        form.fields['numero_segmentos'].initial = 2

        # Get querysets for template
        from cobranca.models import Cobranca
        from socio.models import Socio
        from regraImposto.models import RegraImposto

        context = {
            'form': form,
            'nfse': nfse,
            'formas_pagamento': Cobranca.objects.all(),
            'socios': Socio.objects.filter(empresa_id=empresa_id),
            'regras_imposto': RegraImposto.objects.all().order_by('DescricaoRegraImposto')
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        nfse_id = kwargs.get('pk')
        empresa_id = request.session.get('empresa_id')

        if not empresa_id:
            messages.error(request, 'Selecione uma empresa para continuar.')
            return redirect('empresa:lista')

        try:
            nfse_original = NotaFiscalServico.objects.get(id=nfse_id, empresa_id=empresa_id)
        except NotaFiscalServico.DoesNotExist:
            messages.error(request, 'NFSe não encontrada.')
            return redirect('notasfiscais:list')

        numero_segmentos = int(request.POST.get('numero_segmentos', 2))
        form = NFSeSegmentForm(request.POST, numero_segmentos=numero_segmentos, empresa_id=empresa_id)
        if form.is_valid():
            return self.form_valid(form, nfse_original)
        else:
            # Get querysets for template
            from cobranca.models import Cobranca
            from socio.models import Socio
            from regraImposto.models import RegraImposto

            return render(request, self.template_name, {
                'form': form,
                'nfse': nfse_original,
                'formas_pagamento': Cobranca.objects.all(),
                'socios': Socio.objects.filter(empresa_id=empresa_id),
                'regras_imposto': RegraImposto.objects.all().order_by('DescricaoRegraImposto')
            })

    def form_valid(self, form, nfse_original):
        numero_segmentos = form.cleaned_data['numero_segmentos']
        segmentos_data = []

        # Collect data for each segment
        for i in range(1, numero_segmentos + 1):
            segmento = {
                'segmento': form.cleaned_data.get(f'segmento_{i}'),
                'valor_bruto': form.cleaned_data[f'valor_bruto_{i}'],
                'forma_pagamento': form.cleaned_data[f'forma_pagamento_{i}'],
                'socio': form.cleaned_data[f'socio_{i}'],
                'codigo_regra_imposto': form.cleaned_data.get(f'codigo_regra_imposto_{i}'),
                'base_servico': form.cleaned_data.get(f'base_servico_{i}'),
            }
            segmentos_data.append(segmento)

        # Validate total value
        total_segmentos = sum(s['valor_bruto'] for s in segmentos_data)
        if abs(total_segmentos - nfse_original.valor_bruto) > 0.05:
            # Get querysets for template
            from cobranca.models import Cobranca
            from socio.models import Socio

            messages.error(self.request, f'O total dos segmentos (R$ {total_segmentos:.2f}) deve ser igual ao valor bruto da NFSe original (R$ {nfse_original.valor_bruto:.2f}).')
            context = self.get_context_data(form=form, nfse=nfse_original)
            context.update({
                'formas_pagamento': Cobranca.objects.all(),
                'socios': Socio.objects.filter(empresa_id=self.request.session.get('empresa_id')),
            })
            return self.render_to_response(context)

        # Create segments
        segmentos_criados = []
        with transaction.atomic():
            # Primeiro, criar log da nota original antes de qualquer modificação
            from .models import LogNotaFiscal
            log_nota = LogNotaFiscal.objects.create(
                empresa=nfse_original.empresa,
                numero_nota=nfse_original.numero_nota,
                serie=nfse_original.serie,
                data_emissao=nfse_original.data_emissao,
                cnpj_cpf=nfse_original.cnpj_cpf,
                cliente=nfse_original.cliente,
                valor_bruto=nfse_original.valor_bruto,
                valor_liquido=nfse_original.valor_liquido,
                valor_deducoes=nfse_original.valor_deducoes,
                valor_pis=nfse_original.valor_pis,
                valor_cofins=nfse_original.valor_cofins,
                valor_inss=nfse_original.valor_inss,
                valor_ir=nfse_original.valor_ir,
                valor_csll=nfse_original.valor_csll,
                iss_retido=nfse_original.iss_retido,
                valor_iss_retido=nfse_original.valor_iss_retido,
                outras_retencoes=nfse_original.outras_retencoes,
                aliquota=nfse_original.aliquota,
                socio=nfse_original.socio,
                discriminacao=nfse_original.discriminacao,
                observacoes=nfse_original.observacoes,
                segmento=nfse_original.segmento,
                base_servico=nfse_original.base_servico,
                forma_pagamento=nfse_original.forma_pagamento,
                nsu=nfse_original.nsu,
                status_conciliacao=nfse_original.status_conciliacao,
                issapuracao=nfse_original.issapuracao,
                pisapuracao=nfse_original.pisapuracao,
                cofinsapuracao=nfse_original.cofinsapuracao,
                csllapuracao=nfse_original.csllapuracao,
                irpjapuracao=nfse_original.irpjapuracao,
                irpjadicional=nfse_original.irpjadicional,
                codigo_da_regra_do_imposto=nfse_original.codigo_da_regra_do_imposto,
                motivo_exclusao='segmentacao',
                usuario_segmentacao=self.request.user if hasattr(self.request, 'user') and self.request.user.is_authenticated else None
            )

            for i, segmento in enumerate(segmentos_data, 1):
                # Calculate proportional values for the segment
                proporcao = segmento['valor_bruto'] / nfse_original.valor_bruto
                valor_liquido_segmento = nfse_original.valor_liquido * proporcao

                # Create observation message
                observacao_segmento = f"nota segmentada valores {nfse_original.valor_bruto} e {nfse_original.valor_liquido}"

                # Copy original NFSe data
                novo_segmento = NotaFiscalServico.objects.create(
                    empresa=nfse_original.empresa,
                    numero_nota=f"{nfse_original.numero_nota}-{i}",
                    serie=nfse_original.serie,
                    data_emissao=nfse_original.data_emissao,
                    cnpj_cpf=nfse_original.cnpj_cpf,
                    cliente=nfse_original.cliente,
                    valor_bruto=segmento['valor_bruto'],
                    valor_liquido=valor_liquido_segmento,
                    valor_deducoes=nfse_original.valor_deducoes * proporcao,
                    valor_pis=nfse_original.valor_pis * proporcao,
                    valor_cofins=nfse_original.valor_cofins * proporcao,
                    valor_inss=nfse_original.valor_inss * proporcao,
                    valor_ir=nfse_original.valor_ir * proporcao,
                    valor_csll=nfse_original.valor_csll * proporcao,
                    iss_retido=nfse_original.iss_retido,
                    valor_iss_retido=nfse_original.valor_iss_retido * proporcao,
                    outras_retencoes=nfse_original.outras_retencoes * proporcao,
                    aliquota=nfse_original.aliquota,
                    socio=segmento['socio'],
                    discriminacao=nfse_original.discriminacao,
                    observacoes=observacao_segmento,
                    segmento=segmento['segmento'],
                    base_servico=segmento['base_servico'] or nfse_original.base_servico,
                    forma_pagamento=segmento['forma_pagamento'],
                    nsu=nfse_original.nsu,
                    codigo_da_regra_do_imposto=segmento['codigo_regra_imposto'] or nfse_original.codigo_da_regra_do_imposto,
                )
                segmentos_criados.append(novo_segmento)

            # Delete original note after segmentation
            nfse_original.delete()

        messages.success(self.request, f'NFSe segmentada com sucesso! Criados {len(segmentos_criados)} segmentos.')
        return redirect('notasfiscais:list')


@login_required
def detalhes_modal(request, tipo, id):
    """
    View para fornecer dados ao modal de detalhes
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'error': 'Empresa não selecionada'}, status=400)

    try:
        # Importar modelos necessários
        from .models import NotaFiscalServico
        from contasareceber.models import ContaAReceber
        from extrato.models import ExtratoMovimento, Lancamento

        # Buscar o objeto baseado no tipo
        if tipo == 'nf':
            obj = NotaFiscalServico.objects.get(id=id, empresa_id=empresa_id)
            nf = obj
        elif tipo == 'conta':
            obj = ContaAReceber.objects.get(id=id, empresa_id=empresa_id)
            nf = obj.nota if obj.nota else None
        elif tipo == 'movimento':
            obj = ExtratoMovimento.objects.get(id=id, empresa_id=empresa_id)
            nf = obj.conta_receber.nota if obj.conta_receber and obj.conta_receber.nota else None
        elif tipo == 'lancamento':
            obj = Lancamento.objects.get(id=id, empresa_id=empresa_id)
            # Para lançamento, tentar encontrar NF através de ExtratoMovimento
            movimento = ExtratoMovimento.objects.filter(lancamento=obj).first()
            nf = movimento.conta_receber.nota if movimento and movimento.conta_receber and movimento.conta_receber.nota else None
        else:
            return JsonResponse({'error': 'Tipo inválido'}, status=400)

        # Buscar dados relacionados
        conta_receber = None
        movimentos = []
        lancamento_conciliado = None

        if nf:
            # Buscar conta a receber da NF
            conta_receber = ContaAReceber.objects.filter(nota=nf, empresa_id=empresa_id).first()

        if conta_receber:
            # Buscar movimentos da conta a receber
            movimentos = ExtratoMovimento.objects.filter(conta_receber=conta_receber).order_by('-data_baixa')

        if movimentos:
            # Buscar lançamentos conciliados dos movimentos
            lancamentos_ids = [m.lancamento_id for m in movimentos if m.lancamento_id]
            if lancamentos_ids:
                lancamento_conciliado = Lancamento.objects.filter(id__in=lancamentos_ids).first()

        # Construir HTML do modal
        html = f"""
        <div class="row">
            <div class="col-md-6">
                <h5>Nota Fiscal</h5>
                {f'<p><strong>Número:</strong> {nf.numero_nota}</p>' if nf else '<p>Nenhuma NF encontrada</p>'}
                {f'<p><strong>Cliente:</strong> {nf.cliente}</p>' if nf else ''}
                {f'<p><strong>Valor:</strong> R$ {nf.valor_liquido:.2f}</p>' if nf else ''}
            </div>
            <div class="col-md-6">
                <h5>Conta a Receber</h5>
                {f'<p><strong>Cliente:</strong> {conta_receber.cliente}</p>' if conta_receber else '<p>Nenhuma conta encontrada</p>'}
                {f'<p><strong>Valor:</strong> R$ {conta_receber.valor_a_receber:.2f}</p>' if conta_receber else ''}
                {f'<p><strong>Status:</strong> {conta_receber.get_status_display()}</p>' if conta_receber else ''}
            </div>
        </div>
        <div class="row mt-3">
            <div class="col-md-6">
                <h5>Movimentos do Extrato</h5>
                {''.join([f'<p><strong>{m.data_baixa}:</strong> R$ {m.valor:.2f} - {m.descricao}</p>' for m in movimentos[:3]]) if movimentos else '<p>Nenhum movimento encontrado</p>'}
            </div>
            <div class="col-md-6">
                <h5>Lançamento Conciliado</h5>
                {f'<p><strong>Data:</strong> {lancamento_conciliado.data}</p>' if lancamento_conciliado else '<p>Nenhum lançamento conciliado</p>'}
                {f'<p><strong>Valor:</strong> R$ {lancamento_conciliado.valor:.2f}</p>' if lancamento_conciliado else ''}
                {f'<p><strong>Histórico:</strong> {lancamento_conciliado.historico}</p>' if lancamento_conciliado else ''}
            </div>
        </div>
        """

        return JsonResponse({'html': html})

    except Exception as e:
        return JsonResponse({'error': f'Erro interno: {str(e)}'}, status=500)

@login_required
@require_POST
def ai_segmentacao(request):
    """
    View para segmentação assistida por IA (ChatGPT)
    """
    if not OPENAI_AVAILABLE:
        return JsonResponse({"ok": False, "error": "Módulo OpenAI não instalado. Instale com: pip install openai"}, status=500)

    try:
        ids = request.POST.getlist("ids[]") or request.POST.getlist("ids")
        if not ids:
            return JsonResponse({"ok": False, "error": "IDs não enviados."}, status=400)

        # Converter IDs para inteiros
        try:
            ids = [int(id) for id in ids]
        except ValueError:
            return JsonResponse({"ok": False, "error": "IDs inválidos."}, status=400)

        empresa_id = request.session.get('empresa_id')
        if not empresa_id:
            return JsonResponse({"ok": False, "error": "Empresa não selecionada."}, status=400)

        qs = NotaFiscalServico.objects.filter(id__in=ids, empresa_id=empresa_id)

        if not qs.exists():
            return JsonResponse({"ok": False, "error": "Nenhuma NFSe encontrada com os IDs fornecidos."}, status=404)

        # Coletar descrições
        descricoes = []
        for nf in qs:
            descricoes.append(f"NF {nf.numero_nota}: {nf.discriminacao}")

        texto_completo = "\n\n".join(descricoes)

        # Prompt para ChatGPT
        prompt = f"""
Analise as seguintes descrições de notas fiscais e determine se alguma delas indica múltiplas formas de pagamento com valores distintos.

Para cada nota que tenha múltiplas formas de pagamento, sugira como segmentar a nota em partes, indicando:
- Valor para cada segmento
- Forma de pagamento para cada segmento
- Qualquer outra informação relevante

Descrições:
{texto_completo}

Responda em português, de forma clara e estruturada.
"""

        # Chamar ChatGPT (defina OPENAI_API_KEY no ambiente)
        api_key = os.environ.get("OPENAI_API_KEY", "").strip()
        if not api_key:
            return JsonResponse({"ok": False, "error": "OPENAI_API_KEY não configurada."}, status=500)
        client = openai.OpenAI(api_key=api_key)
        #assistant_id = "asst_LkIP3aAUb72Nj4IlYMFvk2W1"
        # Cria uma thread para conversar com o assistente salvo
        thread = client.beta.threads.create()
        # Adiciona a mensagem do usuário
        client.beta.threads.messages.create(
           thread_id = thread.id,
           role="user",
           content=prompt
        )
        # Roda o assistente salvo
        run = client.beta.threads.runs.create(
           thread_id=thread.id,
           assistant_id="asst_LkIP3aAUb72Nj4IlYMFvk2W1"
           )
        # Recupera a resposta
        # Código comentado - funcionalidade OpenAI desabilitada
        ai_response = "Funcionalidade de IA temporariamente indisponível."

        return JsonResponse({
            "ok": True,
            "ai_response": ai_response,
            "notas_analisadas": len(descricoes)
        })

    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Erro na análise por IA: {str(e)}"}, status=500)


@login_required
def dashboard_nfse(request):
    """
    Dashboard para visualização de notas fiscais com estatísticas e filtros
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Selecione uma empresa para continuar.')
        return redirect('empresa:lista')

    # Filtros de período
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    # Definir datas padrão (mês atual)
    hoje = timezone.now().date()
    primeiro_dia_mes = date(hoje.year, hoje.month, 1)
    ultimo_dia_mes = date(hoje.year, hoje.month, monthrange(hoje.year, hoje.month)[1])

    if not data_inicio:
        data_inicio = primeiro_dia_mes.strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = ultimo_dia_mes.strftime('%Y-%m-%d')

    # Base queryset com filtros
    queryset = NotaFiscalServico.objects.filter(
        empresa_id=empresa_id,
        data_emissao__gte=data_inicio,
        data_emissao__lte=data_fim
    ).select_related('socio', 'forma_pagamento', 'codigo_da_regra_do_imposto')

    # Estatísticas gerais
    total_notas = queryset.count()
    total_valor_bruto = queryset.aggregate(Sum('valor_bruto'))['valor_bruto__sum'] or 0
    total_valor_liquido = queryset.aggregate(Sum('valor_liquido'))['valor_liquido__sum'] or 0

    # Estatísticas por base de serviço
    base_servico_stats = queryset.values('base_servico').annotate(
        count=Count('id'),
        total_bruto=Sum('valor_bruto'),
        total_liquido=Sum('valor_liquido')
    ).order_by('base_servico')

    # Estatísticas por sócio
    socio_stats = queryset.values('socio__socio').annotate(
        count=Count('id'),
        total_bruto=Sum('valor_bruto'),
        total_liquido=Sum('valor_liquido')
    ).order_by('socio__socio')

    # Estatísticas por status de conciliação
    conciliacao_stats = queryset.values('status_conciliacao').annotate(
        count=Count('id'),
        total_bruto=Sum('valor_bruto'),
        total_liquido=Sum('valor_liquido')
    ).order_by('status_conciliacao')

    # Estatísticas por regra de imposto
    regra_imposto_stats = queryset.values('codigo_da_regra_do_imposto__DescricaoRegraImposto').annotate(
        count=Count('id'),
        total_bruto=Sum('valor_bruto'),
        total_liquido=Sum('valor_liquido')
    ).order_by('codigo_da_regra_do_imposto__DescricaoRegraImposto')

    # Estatísticas por forma de pagamento
    forma_pagamento_stats = queryset.values('forma_pagamento__descricao').annotate(
        count=Count('id'),
        total_bruto=Sum('valor_bruto'),
        total_liquido=Sum('valor_liquido')
    ).order_by('forma_pagamento__descricao')

    # Estatísticas de contas a receber
    from contasareceber.models import ContaAReceber
    contas_existentes = ContaAReceber.objects.filter(
        empresa_id=empresa_id,
        nota__data_emissao__gte=data_inicio,
        nota__data_emissao__lte=data_fim
    ).values_list('nota_id', flat=True)

    notas_com_conta = set(contas_existentes)
    total_com_conta = len(notas_com_conta)
    total_sem_conta = total_notas - total_com_conta

    # Listagem das notas (limitada para performance)
    notas = queryset.order_by('-data_emissao')[:100]  # Últimas 100 notas

    context = {
        'notas': notas,
        'data_inicio': data_inicio,
        'data_fim': data_fim,

        # Estatísticas gerais
        'total_notas': total_notas,
        'total_valor_bruto': total_valor_bruto,
        'total_valor_liquido': total_valor_liquido,

        # Estatísticas detalhadas
        'base_servico_stats': base_servico_stats,
        'socio_stats': socio_stats,
        'conciliacao_stats': conciliacao_stats,
        'regra_imposto_stats': regra_imposto_stats,
        'forma_pagamento_stats': forma_pagamento_stats,

        # Estatísticas de contas a receber
        'total_com_conta': total_com_conta,
        'total_sem_conta': total_sem_conta,
        'notas_com_conta': notas_com_conta,
    }

    return render(request, 'notasfiscais/dashboard.html', context)


@login_required
def apuracao_impostos(request):
    """
    View para apuração de impostos das notas fiscais por período
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Selecione uma empresa para continuar.')
        return redirect('empresa:lista')

    # Filtros de período
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    # Definir datas padrão (mês atual)
    hoje = timezone.now().date()
    primeiro_dia_mes = date(hoje.year, hoje.month, 1)
    ultimo_dia_mes = date(hoje.year, hoje.month, monthrange(hoje.year, hoje.month)[1])

    if not data_inicio:
        data_inicio = primeiro_dia_mes.strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = ultimo_dia_mes.strftime('%Y-%m-%d')

    # Buscar notas fiscais do período
    notas = NotaFiscalServico.objects.filter(
        empresa_id=empresa_id,
        data_emissao__gte=data_inicio,
        data_emissao__lte=data_fim
    ).select_related('socio', 'codigo_da_regra_do_imposto').order_by('data_emissao')

    # Calcular demonstrativo por sócio
    from collections import defaultdict

    demonstrativo_por_socio = defaultdict(lambda: {
        'socio_nome': '',
        'valor_bruto': 0,
        'pis': 0,
        'cofins': 0,
        'iss': 0,
        'csll': 0,
        'irpj': 0,
        'adicional_irpj': 0,
        'total_impostos': 0,
        'valor_liquido': 0
    })

    # Totais gerais
    totais_gerais = {
        'valor_bruto': 0,
        'pis': 0,
        'cofins': 0,
        'iss': 0,
        'csll': 0,
        'irpj': 0,
        'adicional_irpj': 0,
        'total_impostos': 0,
        'valor_liquido': 0
    }

    for nota in notas:
        socio_nome = nota.socio.socio if nota.socio else 'Sem Sócio'

        # Atualizar totais por sócio
        demonstrativo_por_socio[socio_nome]['socio_nome'] = socio_nome
        demonstrativo_por_socio[socio_nome]['valor_bruto'] += nota.valor_bruto or 0
        demonstrativo_por_socio[socio_nome]['pis'] += nota.pisapuracao or 0
        demonstrativo_por_socio[socio_nome]['cofins'] += nota.cofinsapuracao or 0
        demonstrativo_por_socio[socio_nome]['iss'] += nota.issapuracao or 0
        demonstrativo_por_socio[socio_nome]['csll'] += nota.csllapuracao or 0
        demonstrativo_por_socio[socio_nome]['irpj'] += nota.irpjapuracao or 0
        demonstrativo_por_socio[socio_nome]['adicional_irpj'] += nota.irpjadicional or 0

        # Calcular total impostos e valor líquido por sócio
        total_impostos_socio = (
            demonstrativo_por_socio[socio_nome]['pis'] +
            demonstrativo_por_socio[socio_nome]['cofins'] +
            demonstrativo_por_socio[socio_nome]['iss'] +
            demonstrativo_por_socio[socio_nome]['csll'] +
            demonstrativo_por_socio[socio_nome]['irpj'] +
            demonstrativo_por_socio[socio_nome]['adicional_irpj']
        )
        demonstrativo_por_socio[socio_nome]['total_impostos'] = total_impostos_socio
        demonstrativo_por_socio[socio_nome]['valor_liquido'] = demonstrativo_por_socio[socio_nome]['valor_bruto'] - total_impostos_socio

        # Atualizar totais gerais
        totais_gerais['valor_bruto'] += nota.valor_bruto or 0
        totais_gerais['pis'] += nota.pisapuracao or 0
        totais_gerais['cofins'] += nota.cofinsapuracao or 0
        totais_gerais['iss'] += nota.issapuracao or 0
        totais_gerais['csll'] += nota.csllapuracao or 0
        totais_gerais['irpj'] += nota.irpjapuracao or 0
        totais_gerais['adicional_irpj'] += nota.irpjadicional or 0

    # Calcular totais gerais
    totais_gerais['total_impostos'] = (
        totais_gerais['pis'] +
        totais_gerais['cofins'] +
        totais_gerais['iss'] +
        totais_gerais['csll'] +
        totais_gerais['irpj'] +
        totais_gerais['adicional_irpj']
    )
    totais_gerais['valor_liquido'] = totais_gerais['valor_bruto'] - totais_gerais['total_impostos']

    # Converter para lista ordenada por nome do sócio
    demonstrativo_lista = sorted(demonstrativo_por_socio.values(), key=lambda x: x['socio_nome'])

    # Verificar ou criar período de apuração
    from .models import ApuracaoPeriodo
    periodo, created = ApuracaoPeriodo.objects.get_or_create(
        empresa_id=int(empresa_id),
        data_inicio=data_inicio,
        data_fim=data_fim,
        defaults={'status': 'aberto', 'adicional_calculado': False}
    )
    periodo_fechado = periodo.status == 'fechado'
    adicional_calculado = periodo.adicional_calculado

    context = {
        'notas': notas,
        'demonstrativo_lista': demonstrativo_lista,
        'totais_gerais': totais_gerais,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'titulo': 'Demonstrativo de Apuração de Impostos',
        'periodo_fechado': periodo_fechado,
        'adicional_calculado': adicional_calculado
    }

    return render(request, 'notasfiscais/apuracao_LP.html', context)


@login_required
def apuracao_simples(request):
    """
    View para apuração do Simples Nacional - faturamento dos últimos 12 meses e anexos
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Selecione uma empresa para continuar.')
        return redirect('empresa:lista')

    # Filtro de ano
    ano = request.GET.get('ano')
    if not ano:
        ano = timezone.now().year
    else:
        try:
            ano = int(ano)
        except ValueError:
            ano = timezone.now().year

    # Buscar empresa para obter anexo
    empresa = Empresa.objects.get(id=empresa_id)

    # Verificar se a empresa está no regime Simples Nacional
    if empresa.regime_tributario != 'SIMPLES_NACIONAL':
        messages.error(request, 'A empresa deve estar no regime Simples Nacional para visualizar a apuração.')
        return redirect('empresa:lista')

    # Determinar anexo baseado nos campos booleanos da empresa
    def determinar_anexo_empresa(empresa, ano):
        """
        Determina qual anexo usar baseado na configuração da empresa
        """
        print(f"DEBUG: tem_fator_r = {empresa.tem_fator_r}")
        print(f"DEBUG: anexo_i = {empresa.anexo_i}, anexo_ii = {empresa.anexo_ii}, anexo_iii = {empresa.anexo_iii}, anexo_iv = {empresa.anexo_iv}, anexo_v = {empresa.anexo_v}")

        # Se não tem fator R, usar o anexo marcado nos campos booleanos
        if not empresa.tem_fator_r:
            print("DEBUG: Empresa não tem fator R, usando anexo marcado")
            anexos_booleanos = {
                empresa.anexo_i: 'I',
                empresa.anexo_ii: 'II',
                empresa.anexo_iii: 'III',
                empresa.anexo_iv: 'IV',
                empresa.anexo_v: 'V'
            }

            # Encontrar qual anexo está marcado como True
            for campo, anexo in anexos_booleanos.items():
                if campo:
                    print(f"DEBUG: Anexo marcado encontrado: {anexo}")
                    return anexo
            print("DEBUG: Nenhum anexo marcado encontrado")
            return None

        # Se tem fator R, verificar se faturamento > 28% no mês referência
        print("DEBUG: Empresa tem fator R, calculando...")
        hoje = timezone.now().date()
        mes_referencia = hoje.month
        ano_referencia = hoje.year

        # Calcular faturamento dos 12 meses anteriores ao mês referência
        data_inicio_12m = date(ano_referencia, mes_referencia, 1) - relativedelta(months=12)
        data_fim_12m = date(ano_referencia, mes_referencia, 1) - timedelta(days=1)

        print(f"DEBUG: Período para cálculo: {data_inicio_12m} a {data_fim_12m}")

        faturamento_12m = NotaFiscalServico.objects.filter(
            empresa_id=empresa_id,
            data_emissao__gte=data_inicio_12m,
            data_emissao__lte=data_fim_12m
        ).aggregate(total=Sum('valor_bruto'))['total'] or 0

        print(f"DEBUG: Faturamento 12m: {faturamento_12m}")

        # Calcular folha de salários dos 12 meses anteriores
        folha_salarios_12m = 0
        current_date = data_inicio_12m
        while current_date <= data_fim_12m:
            ano_folha = current_date.year
            mes_folha = current_date.month
            folha_mes = FolhaSalario.objects.filter(
                empresa_id=empresa_id,
                ano=ano_folha,
                mes=mes_folha
            ).aggregate(total=Sum('total_salario'))['total'] or 0
            print(f"DEBUG: Folha {ano_folha}/{mes_folha}: {folha_mes}")
            folha_salarios_12m += folha_mes
            current_date += relativedelta(months=1)

        print(f"DEBUG: Total folha salários 12m: {folha_salarios_12m}")

        # Calcular Fator R
        if faturamento_12m > 0:
            fator_r = (folha_salarios_12m / faturamento_12m) * 100
            print(f"DEBUG: Fator R calculado: {fator_r}%")
            if fator_r > 28:
                print("DEBUG: Fator R > 28%, retornando Anexo III")
                return 'III'  # Anexo III se Fator R > 28%
            else:
                print("DEBUG: Fator R <= 28%, retornando Anexo V")
                return 'V'    # Anexo V se Fator R <= 28%
        else:
            print("DEBUG: Faturamento 12m = 0, usando fallback")
            # Fallback: usar anexo marcado se não conseguir calcular
            anexos_booleanos = {
                empresa.anexo_i: 'I',
                empresa.anexo_ii: 'II',
                empresa.anexo_iii: 'III',
                empresa.anexo_iv: 'IV',
                empresa.anexo_v: 'V'
            }
            for campo, anexo in anexos_booleanos.items():
                if campo:
                    return anexo
            return None

    print(f"DEBUG: Chamando determinar_anexo_empresa para empresa {empresa.razao}")
    anexo_empresa = determinar_anexo_empresa(empresa, ano)
    print(f"DEBUG: Anexo determinado: {anexo_empresa}")

    # Calcular Receita Bruta dos últimos 12 meses
    meses_portugues = [
        'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
        'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
    ]

    receita_bruta_12_meses = {}
    folha_salarios_12_meses = {}
    faturamento_sem_retencao_12_meses = {}
    faturamento_com_retencao_12_meses = {}
    faturamento_total_mes = {}
    total_receita_bruta_12m = 0
    total_folha_salarios_12m = 0

    # O cálculo da receita bruta dos 12 meses será feito mensalmente
    # Para cada mês do ano, calculamos os 12 meses anteriores

    # Agora calcular os valores mensais para a tabela
    for mes in range(1, 13):
        data_inicio_mes = date(ano, mes, 1)
        if mes == 12:
            data_fim_mes = date(ano, 12, 31)
        else:
            data_fim_mes = date(ano, mes + 1, 1) - timedelta(days=1)

        # Receita Bruta (valor_bruto total do mês corrente)
        receita_bruta_mes_corrente = NotaFiscalServico.objects.filter(
            empresa_id=empresa_id,
            data_emissao__gte=data_inicio_mes,
            data_emissao__lte=data_fim_mes
        ).aggregate(total=Sum('valor_bruto'))['total'] or 0

        # Receita Bruta dos 12 meses anteriores (para alíquota efetiva)
        # Para janeiro/2025: janeiro/2024 a dezembro/2024
        data_inicio_12m_anteriores = data_inicio_mes - relativedelta(months=12)
        data_fim_12m_anteriores = data_inicio_mes - timedelta(days=1)

        receita_bruta_12m_anteriores = NotaFiscalServico.objects.filter(
            empresa_id=empresa_id,
            data_emissao__gte=data_inicio_12m_anteriores,
            data_emissao__lte=data_fim_12m_anteriores
        ).aggregate(total=Sum('valor_bruto'))['total'] or 0

        # Folha de Salários dos 12 meses anteriores (correção)
        folha_salarios_12m_anteriores = 0
        current_date = data_inicio_12m_anteriores
        while current_date <= data_fim_12m_anteriores:
            ano_folha = current_date.year
            mes_folha = current_date.month
            folha_salarios_12m_anteriores += FolhaSalario.objects.filter(
                empresa_id=empresa_id,
                ano=ano_folha,
                mes=mes_folha
            ).aggregate(total=Sum('total_salario'))['total'] or 0
            current_date += relativedelta(months=1)

        # Faturamento SEM retenção (valor_bruto = valor_liquido)
        faturamento_sem_retencao = NotaFiscalServico.objects.filter(
            empresa_id=empresa_id,
            data_emissao__gte=data_inicio_mes,
            data_emissao__lte=data_fim_mes,
            valor_bruto=F('valor_liquido')
        ).aggregate(total=Sum('valor_bruto'))['total'] or 0

        # Faturamento COM retenção (valor_bruto != valor_liquido)
        faturamento_com_retencao = NotaFiscalServico.objects.filter(
            empresa_id=empresa_id,
            data_emissao__gte=data_inicio_mes,
            data_emissao__lte=data_fim_mes
        ).exclude(valor_bruto=F('valor_liquido')).aggregate(total=Sum('valor_bruto'))['total'] or 0

        # Faturamento TOTAL do Mês
        faturamento_total = faturamento_sem_retencao + faturamento_com_retencao

        # A linha "Receita Bruta - 12 Meses" mostra o acumulado dos 12 meses anteriores
        receita_bruta_12_meses[meses_portugues[mes-1]] = receita_bruta_12m_anteriores
        folha_salarios_12_meses[meses_portugues[mes-1]] = folha_salarios_12m_anteriores
        faturamento_sem_retencao_12_meses[meses_portugues[mes-1]] = faturamento_sem_retencao
        faturamento_com_retencao_12_meses[meses_portugues[mes-1]] = faturamento_com_retencao
        faturamento_total_mes[meses_portugues[mes-1]] = faturamento_total

        # REMOVIDO: Esta linha estava somando meses correntes em vez dos 12 meses anteriores
        # total_receita_bruta_12m += receita_bruta_mes_corrente
        # total_folha_salarios_12m += folha_salarios_12m_anteriores

    # Calcular alíquota efetiva para cada mês separadamente
    aliquota_efetiva_mensal = {}
    faixa_mensal = {}

    # Calcular folha de salários dos 12 meses anteriores a janeiro (mesmo período)
    data_inicio_janeiro = date(ano, 1, 1)
    data_inicio_12m_anteriores = data_inicio_janeiro - timedelta(days=365)
    data_fim_12m_anteriores = data_inicio_janeiro - timedelta(days=1)

    total_folha_salarios_12m = 0
    current_date = data_inicio_12m_anteriores
    while current_date <= data_fim_12m_anteriores:
        ano_folha = current_date.year
        mes_folha = current_date.month
        folha_mes = FolhaSalario.objects.filter(
            empresa_id=empresa_id,
            ano=ano_folha,
            mes=mes_folha
        ).aggregate(total=Sum('total_salario'))['total'] or 0
        total_folha_salarios_12m += folha_mes
        current_date += relativedelta(months=1)


    # Calcular Fator R mensal
    fator_r_12_meses = {}
    aliquota_efetiva_12_meses = {}
    fator_r_lista = []
    aliquota_efetiva_lista = []
    faixa_lista = []
    parcela_deduzir_lista = []
    aliquota_tabela_lista = []
    percentual_cpp_lista = []
    percentual_iss_lista = []
    aliquota_cpp_lista = []
    aliquota_iss_lista = []
    aliquota_efetiva_com_iss_lista = []

    # Primeiro, calcular novamente os valores mensais para o Fator R
    receita_bruta_mensal = {}
    for mes in range(1, 13):
        data_inicio_mes = date(ano, mes, 1)
        if mes == 12:
            data_fim_mes = date(ano, 12, 31)
        else:
            data_fim_mes = date(ano, mes + 1, 1) - timedelta(days=1)

        # Receita Bruta (valor_bruto total do mês corrente)
        receita_mes_corrente = NotaFiscalServico.objects.filter(
            empresa_id=empresa_id,
            data_emissao__gte=data_inicio_mes,
            data_emissao__lte=data_fim_mes
        ).aggregate(total=Sum('valor_bruto'))['total'] or 0

        receita_bruta_mensal[meses_portugues[mes-1]] = receita_mes_corrente

    # Calcular Fator R mensal
    fator_r_12_meses = {}
    aliquota_efetiva_12_meses = {}
    fator_r_lista = []
    aliquota_efetiva_lista = []
    faixa_lista = []
    parcela_deduzir_lista = []
    aliquota_tabela_lista = []
    percentual_cpp_lista = []
    percentual_iss_lista = []
    aliquota_cpp_lista = []
    aliquota_iss_lista = []
    aliquota_efetiva_com_iss_lista = []
    anexo_mensal_lista = []

    for mes in range(1, 13):
        # Data início dos 12 meses anteriores
        data_inicio_mes = date(ano, mes, 1)
        data_inicio_12m_anteriores = data_inicio_mes - timedelta(days=365)
        # Data fim dos 12 meses anteriores (um dia antes do mês corrente)
        data_fim_12m_anteriores = data_inicio_mes - timedelta(days=1)

        # Receita bruta dos 12 meses anteriores
        receita_bruta_12m_anteriores = NotaFiscalServico.objects.filter(
            empresa_id=empresa_id,
            data_emissao__gte=data_inicio_12m_anteriores,
            data_emissao__lte=data_fim_12m_anteriores
        ).aggregate(total=Sum('valor_bruto'))['total'] or 0

        # Folha de salários dos 12 meses anteriores
        folha_mes = 0
        current_date = data_inicio_12m_anteriores
        while current_date <= data_fim_12m_anteriores:
            ano_folha = current_date.year
            mes_folha = current_date.month
            folha_mes += FolhaSalario.objects.filter(
                empresa_id=empresa_id,
                ano=ano_folha,
                mes=mes_folha
            ).aggregate(total=Sum('total_salario'))['total'] or 0
            current_date += relativedelta(months=1)

        # Receita bruta dos 12 meses anteriores (para cálculo do Fator R)
        receita_mes = receita_bruta_12m_anteriores

        # Determinar anexo mensal baseado no Fator R
        if empresa.tem_fator_r and receita_mes > 0:
            fator_r_mes = (folha_mes / receita_mes) * 100
            if fator_r_mes > 28:
                anexo_mes = 'III'
            else:
                anexo_mes = 'V'
        else:
            # Se não tem fator R, usar anexo marcado
            if empresa.anexo_i:
                anexo_mes = 'I'
            elif empresa.anexo_ii:
                anexo_mes = 'II'
            elif empresa.anexo_iii:
                anexo_mes = 'III'
            elif empresa.anexo_iv:
                anexo_mes = 'IV'
            elif empresa.anexo_v:
                anexo_mes = 'V'
            else:
                anexo_mes = None

        # Calcular Fator R mensal (Salário ÷ Faturamento)
        if empresa.tem_fator_r and receita_mes > 0:
            fator_r_mes = (folha_mes / receita_mes) * 100
        else:
            fator_r_mes = 0

        # Cálculo do Fator R mensal


        fator_r_12_meses[meses_portugues[mes-1]] = fator_r_mes
        fator_r_lista.append(fator_r_mes)
        anexo_mensal_lista.append(anexo_mes)

        # Calcular alíquota efetiva mensal baseada nos 12 meses anteriores
        # Para janeiro/2025: usar janeiro/2024 a dezembro/2024
        # Para fevereiro/2025: usar fevereiro/2024 a janeiro/2025
        # E assim por diante

        # Data início dos 12 meses anteriores
        data_inicio_12m_anteriores = data_inicio_mes - relativedelta(months=12)
        # Data fim dos 12 meses anteriores (um dia antes do mês corrente)
        data_fim_12m_anteriores = data_inicio_mes - timedelta(days=1)

        # Calcular alíquota efetiva baseada no faturamento dos 12 meses anteriores ao mês corrente
        aliquota_efetiva_mes = 0
        receita_mes = receita_bruta_12_meses[meses_portugues[mes-1]]
        # Usar anexo_mes se disponível, senão anexo_empresa
        anexo_para_faixa = anexo_mes if anexo_mes else anexo_empresa
        faixa_aplicavel_mes = None  # Initialize variable
        parcela_deduzir_mes = 0  # Initialize variable
        aliquota_tabela_mes = 0  # Initialize variable
        percentual_cpp_mes = 0  # Initialize variable
        percentual_iss_mes = 0  # Initialize variable
        if anexo_para_faixa and receita_mes > 0:
            # Find the faixa where the faturamento falls within the range
            # First try to find faixas with specific upper limits
            faixa_aplicavel_mes = AnexoSimplesNacional.objects.filter(
                anexo=anexo_para_faixa,
                ano_vigencia=ano,
                limite_inferior__lte=receita_mes,
                limite_superior__gte=receita_mes
            ).first()

            # If no faixa found, try the last faixa (with null limite_superior)
            if not faixa_aplicavel_mes:
                faixa_aplicavel_mes = AnexoSimplesNacional.objects.filter(
                    anexo=anexo_para_faixa,
                    ano_vigencia=ano,
                    limite_inferior__lte=receita_mes,
                    limite_superior__isnull=True
                ).first()

            if faixa_aplicavel_mes:
                parcela_deduzir_mes = float(faixa_aplicavel_mes.valor_deduzir)
                aliquota_tabela_mes = float(faixa_aplicavel_mes.aliquota)
                percentual_cpp_mes = float(faixa_aplicavel_mes.percentual_cpp)
                percentual_iss_mes = float(faixa_aplicavel_mes.percentual_iss)
                if aliquota_tabela_mes > 0 and receita_mes > 0:
                    # Alíquota efetiva = ((receita * aliquota_tabela/100) - parcela_deduzir) / receita * 100
                    aliquota_tabela_percent = aliquota_tabela_mes / 100.0
                    imposto_bruto = float(receita_mes) * aliquota_tabela_percent
                    imposto_liquido = imposto_bruto - parcela_deduzir_mes
                    aliquota_efetiva_mes = (imposto_liquido / float(receita_mes)) * 100
            else:
                parcela_deduzir_mes = 0
                aliquota_tabela_mes = 0
                percentual_cpp_mes = 0
                percentual_iss_mes = 0

        aliquota_efetiva_12_meses[meses_portugues[mes-1]] = aliquota_efetiva_mes
        aliquota_efetiva_lista.append(aliquota_efetiva_mes)
        faixa_mensal[meses_portugues[mes-1]] = faixa_aplicavel_mes.faixa if faixa_aplicavel_mes else "N/A"
        faixa_lista.append(faixa_aplicavel_mes.faixa if faixa_aplicavel_mes else "N/A")
        parcela_deduzir_lista.append(parcela_deduzir_mes)
        aliquota_tabela_lista.append(aliquota_tabela_mes)
        percentual_cpp_lista.append(percentual_cpp_mes)
        percentual_iss_lista.append(percentual_iss_mes)

        # Calcular alíquotas mensais
        aliquota_cpp_mes = (aliquota_efetiva_mes * percentual_cpp_mes) / 100 if aliquota_efetiva_mes > 0 else 0
        aliquota_iss_mes = (aliquota_efetiva_mes * percentual_iss_mes) / 100 if aliquota_efetiva_mes > 0 else 0
        aliquota_efetiva_com_iss_mes = aliquota_efetiva_mes - aliquota_iss_mes

        aliquota_cpp_lista.append(aliquota_cpp_mes)
        aliquota_iss_lista.append(aliquota_iss_mes)
        aliquota_efetiva_com_iss_lista.append(aliquota_efetiva_com_iss_mes)

    # Calcular Fator R total (para compatibilidade)
    if empresa.tem_fator_r and total_receita_bruta_12m > 0:
        fator_r = (total_folha_salarios_12m / total_receita_bruta_12m) * 100
    else:
        fator_r = 0

    # Encontrar a faixa aplicável baseada no faturamento total
    faixa_aplicavel = None
    parcela_deduzir = 0
    aliquota_tabela = 0
    percentual_cpp = 0
    percentual_iss = 0

    if anexo_empresa and total_receita_bruta_12m > 0:
        # Find the faixa where the faturamento falls within the range
        # First try to find faixas with specific upper limits
        faixa_aplicavel = AnexoSimplesNacional.objects.filter(
            anexo=anexo_empresa,
            ano_vigencia=ano,
            limite_inferior__lte=total_receita_bruta_12m,
            limite_superior__gte=total_receita_bruta_12m
        ).first()

        # If no faixa found, try the last faixa (with null limite_superior)
        if not faixa_aplicavel:
            faixa_aplicavel = AnexoSimplesNacional.objects.filter(
                anexo=anexo_empresa,
                ano_vigencia=ano,
                limite_inferior__lte=total_receita_bruta_12m,
                limite_superior__isnull=True
            ).first()

        if faixa_aplicavel:
            parcela_deduzir = float(faixa_aplicavel.valor_deduzir)
            aliquota_tabela = float(faixa_aplicavel.aliquota)
            percentual_cpp = float(faixa_aplicavel.percentual_cpp)
            percentual_iss = float(faixa_aplicavel.percentual_iss)

    # Calcular alíquota efetiva baseada no total dos 12 meses
    aliquota_efetiva = 0
    if aliquota_tabela > 0 and total_receita_bruta_12m > 0:
        receita_float = float(total_receita_bruta_12m)
        aliquota_tabela_percent = aliquota_tabela / 100.0

        print(f"DEBUG ALÍQUOTA EFETIVA: receita_float = {receita_float}")
        print(f"DEBUG ALÍQUOTA EFETIVA: aliquota_tabela = {aliquota_tabela}%")
        print(f"DEBUG ALÍQUOTA EFETIVA: parcela_deduzir = {parcela_deduzir}")

        # Cálculo passo a passo
        imposto_bruto = receita_float * aliquota_tabela_percent
        print(f"DEBUG ALÍQUOTA EFETIVA: imposto_bruto = {receita_float} * {aliquota_tabela_percent} = {imposto_bruto}")

        imposto_liquido = imposto_bruto - parcela_deduzir
        print(f"DEBUG ALÍQUOTA EFETIVA: imposto_liquido = {imposto_bruto} - {parcela_deduzir} = {imposto_liquido}")

        aliquota_efetiva = (imposto_liquido / receita_float) * 100
        print(f"DEBUG ALÍQUOTA EFETIVA: aliquota_efetiva = ({imposto_liquido} / {receita_float}) * 100 = {aliquota_efetiva}%")

    # Calcular Alíquotas do período
    aliquota_cpp_periodo = (aliquota_efetiva * percentual_cpp) / 100
    aliquota_iss_periodo = (aliquota_efetiva * percentual_iss) / 100
    aliquota_efetiva_com_iss_retido = aliquota_efetiva - aliquota_iss_periodo

    # Calcular totais do período
    total_faturamento_periodo = float(sum(faturamento_total_mes.values()))
    total_das_devido_periodo = total_faturamento_periodo * aliquota_efetiva / 100  # Corrigido: usar faturamento do período atual * alíquota efetiva

    # Calcular DAS devido mensalmente
    total_das_devido_mes = {}
    total_cpp_mes = {}
    for i, (mes, faturamento) in enumerate(faturamento_total_mes.items()):
        aliquota_efetiva_mes = aliquota_efetiva_lista[i] if i < len(aliquota_efetiva_lista) else 0
        aliquota_cpp_mes = aliquota_cpp_lista[i] if i < len(aliquota_cpp_lista) else 0
        total_das_devido_mes[mes] = float(faturamento) * aliquota_efetiva_mes / 100
        total_cpp_mes[mes] = float(faturamento) * aliquota_cpp_mes / 100

    # Calcular DAS sem retenção e com retenção mensalmente
    total_das_sem_retencao_mes = {}
    total_das_com_retencao_mes = {}
    total_das_pagar_mes = {}

    for i, mes in enumerate(faturamento_total_mes.keys()):
        sem_retencao = float(faturamento_sem_retencao_12_meses.get(mes, 0))
        com_retencao = float(faturamento_com_retencao_12_meses.get(mes, 0))
        aliquota_efetiva_mes = aliquota_efetiva_lista[i] if i < len(aliquota_efetiva_lista) else 0
        aliquota_efetiva_com_iss_mes = aliquota_efetiva_com_iss_lista[i] if i < len(aliquota_efetiva_com_iss_lista) else 0

        total_das_sem_retencao_mes[mes] = sem_retencao * aliquota_efetiva_mes / 100
        total_das_com_retencao_mes[mes] = com_retencao * aliquota_efetiva_com_iss_mes / 100
        total_das_pagar_mes[mes] = total_das_sem_retencao_mes[mes] + total_das_com_retencao_mes[mes]

    # Totais gerais para compatibilidade
    total_faturamento_sem_retencao = float(sum(faturamento_sem_retencao_12_meses.values()))
    total_faturamento_com_retencao = float(sum(faturamento_com_retencao_12_meses.values()))

    total_das_sem_retencao = total_faturamento_sem_retencao * aliquota_efetiva / 100
    total_das_com_retencao = total_faturamento_com_retencao * aliquota_efetiva / 100
    total_das_pagar_periodo = total_das_sem_retencao + total_das_com_retencao

    print(f"DEBUG ALÍQUOTA EFETIVA: totallddlsl = {faturamento_sem_retencao_12_meses} ")
    print(f"DEBUG ALÍQUOTA EFETIVA: fatut per = {total_faturamento_periodo} ")
    print(f"DEBUG ALÍQUOTA EFETIVA: aliquota_efetiva = {aliquota_efetiva} ")
    print(f"DEBUG ALÍQUOTA EFETIVA: total_das_sem_retencao = {total_das_sem_retencao} ")
    print(f"DEBUG ALÍQUOTA EFETIVA: total_das_com_retencao = {total_das_com_retencao} ")
    print(f"DEBUG ALÍQUOTA EFETIVA: total_das_pagar_periodo = {total_das_pagar_periodo} ")
    total_cpp_periodo = total_faturamento_periodo * aliquota_cpp_periodo / 100  # Corrigido: usar faturamento do período atual

    # Calcular totais para as colunas de total
    total_receita_bruta_acumulada = sum(receita_bruta_12_meses.values())


    # Preparar dados para o template
    context = {
        'ano': ano,
        'receita_bruta_12_meses': receita_bruta_12_meses,
        'folha_salarios_12_meses': folha_salarios_12_meses,
        'faturamento_sem_retencao_12_meses': faturamento_sem_retencao_12_meses,
        'faturamento_com_retencao_12_meses': faturamento_com_retencao_12_meses,
        'faturamento_total_mes': faturamento_total_mes,
        'fator_r_12_meses': fator_r_12_meses,
        'fator_r_lista': fator_r_lista,
        'anexo_mensal_lista': anexo_mensal_lista,
        'aliquota_efetiva_12_meses': aliquota_efetiva_12_meses,
        'aliquota_efetiva_lista': aliquota_efetiva_lista,
        'faixa_mensal': faixa_mensal,
        'faixa_lista': faixa_lista,
        'parcela_deduzir_lista': parcela_deduzir_lista,
        'aliquota_tabela_lista': aliquota_tabela_lista,
        'percentual_cpp_lista': percentual_cpp_lista,
        'percentual_iss_lista': percentual_iss_lista,
        'aliquota_cpp_lista': aliquota_cpp_lista,
        'aliquota_iss_lista': aliquota_iss_lista,
        'aliquota_efetiva_com_iss_lista': aliquota_efetiva_com_iss_lista,
        'total_receita_bruta_12m': total_receita_bruta_12m,
        'total_receita_bruta_acumulada': total_receita_bruta_acumulada,
        'total_folha_salarios_12m': total_folha_salarios_12m,
        'fator_r': fator_r,
        'anexo_empresa': anexo_empresa,
        'empresa_anexo': anexo_empresa,
        'faixa_aplicavel': faixa_aplicavel,
        'parcela_deduzir': parcela_deduzir,
        'aliquota_tabela': aliquota_tabela,
        'aliquota_efetiva': aliquota_efetiva,
        'percentual_cpp': percentual_cpp,
        'aliquota_cpp_periodo': aliquota_cpp_periodo,
        'percentual_iss': percentual_iss,
        'aliquota_iss_periodo': aliquota_iss_periodo,
        'aliquota_efetiva_com_iss_retido': aliquota_efetiva_com_iss_retido,
        'total_das_devido_periodo': total_das_devido_periodo,
        'total_das_devido_mes': total_das_devido_mes,
        'total_cpp_periodo': total_cpp_periodo,
        'total_cpp_mes': total_cpp_mes,
        'total_das_sem_retencao': total_das_sem_retencao,
        'total_das_com_retencao': total_das_com_retencao,
        'total_das_pagar_periodo': total_das_pagar_periodo,
        'total_das_sem_retencao_mes': total_das_sem_retencao_mes,
        'total_das_com_retencao_mes': total_das_com_retencao_mes,
        'total_das_pagar_mes': total_das_pagar_mes,
        'meses_portugues': meses_portugues,
    }

    if not anexo_empresa:
        messages.warning(request, 'A empresa não possui anexo do Simples Nacional configurado.')

    return render(request, 'notasfiscais/apuracao_simples.html', context)


@login_required
def anexos_simples_list(request):
    """
    View para listar anexos do Simples Nacional
    """
    # Filtros
    anexo_filter = request.GET.get('anexo', '')
    ano_vigencia_filter = request.GET.get('ano_vigencia', '')

    # Buscar anexos do banco de dados
    anexos = AnexoSimplesNacional.objects.all().order_by('anexo', 'limite_inferior')

    # Aplicar filtros
    if anexo_filter:
        anexos = anexos.filter(anexo=anexo_filter)

    if ano_vigencia_filter:
        try:
            ano_vigencia_filter = int(ano_vigencia_filter)
            anexos = anexos.filter(ano_vigencia=ano_vigencia_filter)
        except ValueError:
            pass

    context = {
        'anexos': anexos,
        'titulo': 'Anexos do Simples Nacional'
    }
    return render(request, 'notasfiscais/anexos_simples_list.html', context)


@login_required
def anexos_simples_create(request):
    """
    View para criar anexo do Simples Nacional
    """
    if request.method == 'POST':
        anexo = request.POST.get('anexo')
        faixa = request.POST.get('faixa')
        limite_inferior = request.POST.get('limite_inferior')
        limite_superior = request.POST.get('limite_superior')
        aliquota = request.POST.get('aliquota')
        valor_deduzir = request.POST.get('valor_deduzir')
        ano_vigencia = request.POST.get('ano_vigencia', 2024)

        # Novos campos de percentuais
        percentual_irpj = request.POST.get('percentual_irpj', 0)
        percentual_csll = request.POST.get('percentual_csll', 0)
        percentual_cofins = request.POST.get('percentual_cofins', 0)
        percentual_pis = request.POST.get('percentual_pis', 0)
        percentual_cpp = request.POST.get('percentual_cpp', 0)
        percentual_iss = request.POST.get('percentual_iss', 0)

        try:
            AnexoSimplesNacional.objects.create(
                anexo=anexo,
                faixa=faixa,
                limite_inferior=limite_inferior,
                limite_superior=limite_superior or None,
                aliquota=aliquota,
                valor_deduzir=valor_deduzir,
                ano_vigencia=ano_vigencia,
                percentual_irpj=percentual_irpj,
                percentual_csll=percentual_csll,
                percentual_cofins=percentual_cofins,
                percentual_pis=percentual_pis,
                percentual_cpp=percentual_cpp,
                percentual_iss=percentual_iss
            )
            messages.success(request, 'Anexo criado com sucesso!')
            return redirect('notasfiscais:anexos_simples_list')
        except Exception as e:
            messages.error(request, f'Erro ao criar anexo: {str(e)}')

    context = {
        'titulo': 'Cadastrar Anexo do Simples Nacional',
        'ANEXO_CHOICES': AnexoSimplesNacional.ANEXO_CHOICES
    }
    return render(request, 'notasfiscais/anexos_simples_form.html', context)


@login_required
def anexos_simples_update(request, pk):
    """
    View para editar anexo do Simples Nacional
    """
    try:
        anexo_obj = AnexoSimplesNacional.objects.get(pk=pk)
    except AnexoSimplesNacional.DoesNotExist:
        messages.error(request, 'Anexo não encontrado!')
        return redirect('notasfiscais:anexos_simples_list')

    if request.method == 'POST':
        anexo = request.POST.get('anexo')
        faixa = request.POST.get('faixa')
        limite_inferior = request.POST.get('limite_inferior')
        limite_superior = request.POST.get('limite_superior')
        aliquota = request.POST.get('aliquota')
        valor_deduzir = request.POST.get('valor_deduzir')
        ano_vigencia = request.POST.get('ano_vigencia', 2024)

        # Novos campos de percentuais
        percentual_irpj = request.POST.get('percentual_irpj', 0)
        percentual_csll = request.POST.get('percentual_csll', 0)
        percentual_cofins = request.POST.get('percentual_cofins', 0)
        percentual_pis = request.POST.get('percentual_pis', 0)
        percentual_cpp = request.POST.get('percentual_cpp', 0)
        percentual_iss = request.POST.get('percentual_iss', 0)

        try:
            anexo_obj.anexo = anexo
            anexo_obj.faixa = faixa
            anexo_obj.limite_inferior = limite_inferior
            anexo_obj.limite_superior = limite_superior or None
            anexo_obj.aliquota = aliquota
            anexo_obj.valor_deduzir = valor_deduzir
            anexo_obj.ano_vigencia = ano_vigencia
            anexo_obj.percentual_irpj = percentual_irpj
            anexo_obj.percentual_csll = percentual_csll
            anexo_obj.percentual_cofins = percentual_cofins
            anexo_obj.percentual_pis = percentual_pis
            anexo_obj.percentual_cpp = percentual_cpp
            anexo_obj.percentual_iss = percentual_iss
            anexo_obj.save()

            messages.success(request, 'Anexo atualizado com sucesso!')
            return redirect('notasfiscais:anexos_simples_list')
        except Exception as e:
            messages.error(request, f'Erro ao atualizar anexo: {str(e)}')

    context = {
        'titulo': 'Editar Anexo do Simples Nacional',
        'anexo': anexo_obj,
        'ANEXO_CHOICES': AnexoSimplesNacional.ANEXO_CHOICES
    }
    return render(request, 'notasfiscais/anexos_simples_form.html', context)


@login_required
def anexos_simples_delete(request, pk):
    """
    View para excluir anexo do Simples Nacional
    """
    try:
        anexo_obj = AnexoSimplesNacional.objects.get(pk=pk)
    except AnexoSimplesNacional.DoesNotExist:
        messages.error(request, 'Anexo não encontrado!')
        return redirect('notasfiscais:anexos_simples_list')

    if request.method == 'POST':
        try:
            anexo_obj.delete()
            messages.success(request, 'Anexo excluído com sucesso!')
            return redirect('notasfiscais:anexos_simples_list')
        except Exception as e:
            messages.error(request, f'Erro ao excluir anexo: {str(e)}')

    context = {
        'titulo': 'Confirmar Exclusão',
        'anexo': anexo_obj,
        'mensagem': f'Tem certeza que deseja excluir o anexo {anexo_obj.anexo} - Faixa {anexo_obj.faixa}?'
    }
    return render(request, 'notasfiscais/anexos_simples_confirm_delete.html', context)


def retencoes_nota_ajax(request, nota_id):
    """
    View AJAX para retornar dados de retenções de uma nota fiscal
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'success': False, 'error': 'Empresa não selecionada'}, status=400)

    try:
        nota = NotaFiscalServico.objects.get(id=nota_id, empresa_id=empresa_id)

        # Preparar dados de retenção
        dados_retencao = {
            'numero_nota': nota.numero_nota,
            'iss_retido': nota.iss_retido,
            'aliquota': float(nota.aliquota or 0),
            'valor_iss_retido': float(nota.valor_iss_retido or 0),
            'valor_pis': float(nota.valor_pis or 0),
            'valor_cofins': float(nota.valor_cofins or 0),
            'valor_csll': float(nota.valor_csll or 0),
            'valor_ir': float(nota.valor_ir or 0),
            'outras_retencoes': float(nota.outras_retencoes or 0),
            'valor_inss': float(nota.valor_inss or 0)
        }

        return JsonResponse({
            'success': True,
            'nota': dados_retencao
        })

    except NotaFiscalServico.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Nota fiscal não encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erro interno: {str(e)}'}, status=500)


@login_required
@require_POST
def calcular_adicional_trimestral(request):
    """
    View AJAX para calcular o adicional trimestral de IRPJ ou preview
    """
    print("=== DEBUG calcular_adicional_trimestral: INÍCIO ===")
    empresa_id = request.session.get('empresa_id')
    print(f"Empresa ID da sessão: {empresa_id}")
    if not empresa_id:
        print("ERRO: Empresa não selecionada")
        return JsonResponse({'success': False, 'error': 'Empresa não selecionada'}, status=400)

    try:
        # Obter período dos filtros
        data_inicio = request.POST.get('data_inicio')
        data_fim = request.POST.get('data_fim')
        preview = request.POST.get('preview') == 'true'
        detalhes = request.POST.get('detalhes') == 'true'
        print(f"Período solicitado: {data_inicio} a {data_fim}, Preview: {preview}, Detalhes: {detalhes}")

        if not data_inicio or not data_fim:
            print("ERRO: Período não especificado")
            return JsonResponse({'success': False, 'error': 'Período não especificado'}, status=400)

        # Buscar ou criar período de apuração
        print("Buscando ou criando período de apuração...")
        from .models import ApuracaoPeriodo
        periodo, created = ApuracaoPeriodo.objects.get_or_create(
            empresa_id=int(empresa_id),
            data_inicio=data_inicio,
            data_fim=data_fim,
            defaults={'status': 'aberto', 'adicional_calculado': False}
        )
        if created:
            print(f"Período de apuração criado: {periodo}")

        if preview:
            # Retornar dados de preview
            print("Gerando preview...")
            preview_data = periodo.get_preview_adicional_irpj()
            return JsonResponse({
                'success': True,
                'preview': True,
                'data': preview_data
            })
        elif detalhes:
            # Retornar detalhes do cálculo
            print("Buscando detalhes do cálculo...")
            detalhes_data = periodo.get_detalhes_calculo_adicional()
            return JsonResponse({
                'success': True,
                'detalhes': True,
                'data': detalhes_data
            })
        else:
            # Calcular adicional
            print("Iniciando cálculo do adicional...")
            valor_adicional, debug_messages = periodo.calcular_adicional_irpj()
            print(f"Cálculo concluído. Valor adicional: R$ {valor_adicional:.2f}")

            print("=== DEBUG calcular_adicional_trimestral: SUCESSO ===")
            return JsonResponse({
                'success': True,
                'message': f'Adicional trimestral calculado com sucesso! Valor: R$ {valor_adicional:.2f}',
                'valor_adicional': float(valor_adicional),
                'debug': '\n'.join(debug_messages)
            })

    except Exception as e:
        print(f"=== DEBUG calcular_adicional_trimestral: ERRO ===")
        print(f"Erro: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Erro ao calcular adicional: {str(e)}'}, status=500)


@login_required
@require_POST
def fechar_periodo(request):
    """
    View AJAX para fechar o período de apuração
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'success': False, 'error': 'Empresa não selecionada'}, status=400)

    try:
        # Obter período dos filtros
        data_inicio = request.POST.get('data_inicio')
        data_fim = request.POST.get('data_fim')

        if not data_inicio or not data_fim:
            return JsonResponse({'success': False, 'error': 'Período não especificado'}, status=400)

        # Buscar ou criar período de apuração
        from .models import ApuracaoPeriodo
        periodo, created = ApuracaoPeriodo.objects.get_or_create(
            empresa_id=int(empresa_id),
            data_inicio=data_inicio,
            data_fim=data_fim,
            defaults={'status': 'aberto', 'adicional_calculado': False}
        )

        # Fechar período
        periodo.fechar_periodo(request.user)

        return JsonResponse({
            'success': True,
            'message': 'Período fechado com sucesso!'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erro ao fechar período: {str(e)}'}, status=500)


@login_required
@require_POST
def reabrir_periodo(request):
    """
    View AJAX para reabrir o período de apuração
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'success': False, 'error': 'Empresa não selecionada'}, status=400)

    try:
        # Obter período dos filtros
        data_inicio = request.POST.get('data_inicio')
        data_fim = request.POST.get('data_fim')

        if not data_inicio or not data_fim:
            return JsonResponse({'success': False, 'error': 'Período não especificado'}, status=400)

        # Buscar ou criar período de apuração
        from .models import ApuracaoPeriodo
        periodo, created = ApuracaoPeriodo.objects.get_or_create(
            empresa_id=int(empresa_id),
            data_inicio=data_inicio,
            data_fim=data_fim,
            defaults={'status': 'aberto', 'adicional_calculado': False}
        )

        # Reabrir período
        periodo.reabrir_periodo()

        return JsonResponse({
            'success': True,
            'message': 'Período reaberto com sucesso!'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erro ao reabrir período: {str(e)}'}, status=500)


@login_required
def restaurar_nota_fiscal_view(request, log_id):
    """
    View para restaurar uma nota fiscal do log
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Selecione uma empresa para continuar.')
        return redirect('empresa:lista')

    try:
        # Buscar o log da nota fiscal
        from .models import LogNotaFiscal
        log_nota = LogNotaFiscal.objects.get(id=log_id, empresa_id=empresa_id)

        # Verificar se já existe uma nota com o mesmo número e série
        nota_existente = NotaFiscalServico.objects.filter(
            empresa_id=empresa_id,
            numero_nota=log_nota.numero_nota,
            serie=log_nota.serie
        ).exists()

        if nota_existente:
            messages.error(request, f'Já existe uma nota fiscal com o número {log_nota.numero_nota} e série {log_nota.serie}.')
            return redirect('notasfiscais:list')

        # Criar nova nota fiscal baseada no log
        nova_nota = NotaFiscalServico.objects.create(
            empresa_id=empresa_id,
            numero_nota=log_nota.numero_nota,
            serie=log_nota.serie,
            data_emissao=log_nota.data_emissao,
            cnpj_cpf=log_nota.cnpj_cpf,
            cliente=log_nota.cliente,
            valor_bruto=log_nota.valor_bruto,
            valor_liquido=log_nota.valor_liquido,
            valor_deducoes=log_nota.valor_deducoes,
            valor_pis=log_nota.valor_pis,
            valor_cofins=log_nota.valor_cofins,
            valor_inss=log_nota.valor_inss,
            valor_ir=log_nota.valor_ir,
            valor_csll=log_nota.valor_csll,
            iss_retido=log_nota.iss_retido,
            valor_iss_retido=log_nota.valor_iss_retido,
            outras_retencoes=log_nota.outras_retencoes,
            aliquota=log_nota.aliquota,
            socio=log_nota.socio,
            discriminacao=log_nota.discriminacao,
            observacoes=log_nota.observacoes,
            segmento=log_nota.segmento,
            base_servico=log_nota.base_servico,
            forma_pagamento=log_nota.forma_pagamento,
            nsu=log_nota.nsu,
            status_conciliacao=log_nota.status_conciliacao,
            issapuracao=log_nota.issapuracao,
            pisapuracao=log_nota.pisapuracao,
            cofinsapuracao=log_nota.cofinsapuracao,
            csllapuracao=log_nota.csllapuracao,
            irpjapuracao=log_nota.irpjapuracao,
            irpjadicional=log_nota.irpjadicional,
            codigo_da_regra_do_imposto=log_nota.codigo_da_regra_do_imposto
        )

        # Excluir o log após restaurar
        log_nota.delete()

        messages.success(request, f'Nota fiscal {nova_nota.numero_nota} restaurada com sucesso!')
        return redirect('notasfiscais:list')

    except LogNotaFiscal.DoesNotExist:
        messages.error(request, 'Log da nota fiscal não encontrado.')
        return redirect('notasfiscais:list')
    except Exception as e:
        messages.error(request, f'Erro ao restaurar nota fiscal: {str(e)}')
        return redirect('notasfiscais:list')


@login_required
def listar_logs_notas_fiscais_view(request):
    """
    View para listar logs de notas fiscais excluídas
    """
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        messages.error(request, 'Selecione uma empresa para continuar.')
        return redirect('empresa:lista')

    # Filtros
    search = request.GET.get('search', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')

    # Buscar logs da empresa
    from .models import LogNotaFiscal
    logs = LogNotaFiscal.objects.filter(empresa_id=empresa_id)

    # Aplicar filtros
    if search:
        logs = logs.filter(
            Q(numero_nota__icontains=search) |
            Q(cliente__icontains=search) |
            Q(cnpj_cpf__icontains=search)
        )

    if data_inicio:
        try:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            logs = logs.filter(data_segmentacao__date__gte=data_inicio)
        except ValueError:
            pass

    if data_fim:
        try:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
            logs = logs.filter(data_segmentacao__date__lte=data_fim)
        except ValueError:
            pass

    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(logs.order_by('-data_segmentacao'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search': search,
        'data_inicio': request.GET.get('data_inicio', ''),
        'data_fim': request.GET.get('data_fim', ''),
        'title': 'Logs de Notas Fiscais Excluídas'
    }

    return render(request, 'notasfiscais/log_nota_list.html', context)


@login_required
@require_POST
def marcar_nfse_cancelada(request, pk):
    """Marca ou reativa NFSe (define/limpa data_cancelamento). Ao cancelar, zera valores."""
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return JsonResponse({'ok': False, 'error': 'Empresa não selecionada.'}, status=400)

    nfse = get_object_or_404(NotaFiscalServico, pk=pk, empresa_id=empresa_id)
    acao = (request.POST.get('acao') or 'cancelar').strip().lower()
    wants_json = (
        request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        or 'application/json' in (request.headers.get('Accept') or '')
    )

    from .models import ApuracaoPeriodo

    periodos_fechados = ApuracaoPeriodo.objects.filter(
        empresa_id=int(empresa_id),
        data_inicio__lte=nfse.data_emissao,
        data_fim__gte=nfse.data_emissao,
        status='fechado',
    )
    if periodos_fechados.exists():
        msg = 'Período fechado — reabra o período em Apuração de Impostos para alterar o status da nota.'
        if wants_json:
            return JsonResponse({'ok': False, 'error': msg}, status=400)
        messages.error(request, msg)
        return redirect('notasfiscais:list')

    next_url = request.POST.get('next') or reverse('notasfiscais:detail', kwargs={'pk': nfse.pk})
    zero = Decimal('0')

    if acao == 'reativar':
        if not nfse.is_cancelada():
            msg = 'Esta NFSe já está ativa.'
            if wants_json:
                return JsonResponse({'ok': False, 'error': msg}, status=400)
            messages.warning(request, msg)
            return redirect(next_url)

        nfse.data_cancelamento = None
        nfse.save(update_fields=['data_cancelamento', 'data_atualizacao'])
        msg = f'NFSe {nfse.numero_nota} reativada com sucesso.'
        if wants_json:
            return JsonResponse({
                'ok': True,
                'cancelada': False,
                'message': msg,
                'data_cancelamento': None,
            })
        messages.success(request, msg)
        return redirect(next_url)

    # cancelar: marca data e zera valores (como importação de canceladas)
    if not nfse.data_cancelamento:
        nfse.data_cancelamento = timezone.localdate()
    nfse.valor_bruto = zero
    nfse.valor_liquido = zero
    nfse.valor_pis = zero
    nfse.valor_cofins = zero
    nfse.valor_inss = zero
    nfse.valor_ir = zero
    nfse.valor_csll = zero
    nfse.valor_iss_retido = zero
    nfse.outras_retencoes = zero
    nfse.aliquota = zero
    nfse.issapuracao = zero
    nfse.pisapuracao = zero
    nfse.cofinsapuracao = zero
    nfse.csllapuracao = zero
    nfse.irpjapuracao = zero
    nfse.irpjadicional = zero
    nfse.save(update_fields=[
        'data_cancelamento', 'valor_bruto', 'valor_liquido',
        'valor_pis', 'valor_cofins', 'valor_inss', 'valor_ir', 'valor_csll',
        'valor_iss_retido', 'outras_retencoes', 'aliquota',
        'issapuracao', 'pisapuracao', 'cofinsapuracao', 'csllapuracao',
        'irpjapuracao', 'irpjadicional', 'data_atualizacao',
    ])
    msg = f'NFSe {nfse.numero_nota} marcada como cancelada (valores zerados).'
    if wants_json:
        return JsonResponse({
            'ok': True,
            'cancelada': True,
            'message': msg,
            'data_cancelamento': nfse.data_cancelamento.isoformat(),
        })
    messages.success(request, msg)
    return redirect(next_url)


# Template filters
# Removido filtro get_item não utilizado
