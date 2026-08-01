from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal
from urllib.parse import urlencode
import json

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.db.models import Prefetch, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_POST

from empresa.models import Empresa

from .forms import ItemOrcamentoForm
from .models import ItemOrcamento, LancamentoOrcamento


TIPOS_VALIDOS = {c[0] for c in ItemOrcamento.TIPO_CHOICES}


def _empresa_sessao(request):
    empresa_id = request.session.get('empresa_id')
    if not empresa_id:
        return None
    try:
        return Empresa.objects.get(pk=empresa_id)
    except Empresa.DoesNotExist:
        return None


def _parse_date(val):
    if not val:
        return None
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _periodo_padrao():
    hoje = date.today()
    ini = hoje.replace(day=1)
    fim = hoje.replace(day=monthrange(hoje.year, hoje.month)[1])
    return ini, fim


def _periodo_consulta(request):
    """Lê data_ini / data_fim do GET; padrão = mês corrente."""
    padrao_ini, padrao_fim = _periodo_padrao()
    data_ini = _parse_date(request.GET.get('data_ini')) or padrao_ini
    data_fim = _parse_date(request.GET.get('data_fim')) or padrao_fim
    if data_fim < data_ini:
        data_ini, data_fim = data_fim, data_ini
    return data_ini, data_fim


def _qs_params(data_ini, data_fim):
    return urlencode({
        'data_ini': data_ini.isoformat(),
        'data_fim': data_fim.isoformat(),
    })


def _periodo_mes_vizinho(data_ini: date, delta: int = 1) -> tuple[date, date]:
    """Retorna o 1º e o último dia do mês deslocado a partir de data_ini."""
    y, m = data_ini.year, data_ini.month
    m += int(delta)
    while m > 12:
        m -= 12
        y += 1
    while m < 1:
        m += 12
        y -= 1
    ini = date(y, m, 1)
    fim = date(y, m, monthrange(y, m)[1])
    return ini, fim


def _periodo_nav_qs(data_ini: date) -> dict:
    ant_ini, ant_fim = _periodo_mes_vizinho(data_ini, -1)
    prox_ini, prox_fim = _periodo_mes_vizinho(data_ini, 1)
    return {
        'periodo_ant_qs': _qs_params(ant_ini, ant_fim),
        'periodo_prox_qs': _qs_params(prox_ini, prox_fim),
    }


def _total_receitas_periodo(empresa, data_ini, data_fim):
    return (
        LancamentoOrcamento.objects.filter(
            empresa=empresa,
            item__tipo=ItemOrcamento.TIPO_RECEITA,
            item__ativo=True,
            data_lancamento__gte=data_ini,
            data_lancamento__lte=data_fim,
        ).aggregate(t=Sum('valor'))['t']
        or Decimal('0')
    )


def _somar_tipo_periodo(empresa, tipo, data_ini, data_fim):
    qs = LancamentoOrcamento.objects.filter(
        empresa=empresa,
        item__tipo=tipo,
        item__ativo=True,
        data_lancamento__gte=data_ini,
        data_lancamento__lte=data_fim,
    )
    total = qs.aggregate(t=Sum('valor'))['t'] or Decimal('0')
    qtd_itens = qs.values('item_id').distinct().count()
    qtd_lanc = qs.count()
    return total, qtd_itens, qtd_lanc


def _total_receitas(empresa):
    """Valor mensal dos itens ativos — fallback na geração de %."""
    itens = ItemOrcamento.objects.filter(
        empresa=empresa,
        tipo=ItemOrcamento.TIPO_RECEITA,
        ativo=True,
    )
    total = Decimal('0')
    for it in itens:
        total += it.valor_estimado()
    return total


def _url_listar_item(obj):
    """Lista do tipo com período cobrindo o mês do item (para ver o valor gravado)."""
    base = reverse('planejamento_orcamentario:listar_tipo', kwargs={'tipo': obj.tipo})
    if not obj.data_inicio:
        return base
    d = obj.data_inicio
    ini = d.replace(day=1)
    fim = d.replace(day=monthrange(d.year, d.month)[1])
    return base + '?' + _qs_params(ini, fim)


def _salvar_e_gerar(obj, total_receitas=None):
    try:
        obj.save()
    except IntegrityError as exc:
        if '_pkey' in str(exc).lower() or 'duplicate key' in str(exc).lower():
            raise IntegrityError(
                'Conflito de ID no banco (sequência PostgreSQL desatualizada após importação). '
                'Execute: python scripts/corrigir_sequencias_postgres.py'
            ) from exc
        raise
    n = obj.gerar_lancamentos(total_receitas=total_receitas)
    # Receita alterada → recalcula impostos/variáveis em % sobre receitas
    if obj.tipo == ItemOrcamento.TIPO_RECEITA:
        ItemOrcamento.regenerar_percentuais(obj.empresa)
    return n


def _meses_periodo(data_ini, data_fim):
    """Lista de meses (ano, mês, rótulo) cobertos pelo período filtrado."""
    meses_lbl = (
        '', 'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
        'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez',
    )
    meses = []
    y, m = data_ini.year, data_ini.month
    y_fim, m_fim = data_fim.year, data_fim.month
    while (y, m) <= (y_fim, m_fim):
        meses.append({
            'ano': y,
            'mes': m,
            'chave': f'{y}-{m:02d}',
            'rotulo': f'{meses_lbl[m]}/{y}',
        })
        if m == 12:
            y, m = y + 1, 1
        else:
            m += 1
    return meses


def _tabela_despesas_por_mes(empresa, data_ini, data_fim):
    """
    Monta linhas: descrição × valor por mês do período.
    Inclui receitas, despesas (fixas/semi/variáveis), impostos e resultado.
    """
    meses = _meses_periodo(data_ini, data_fim)
    tipos_ordem = (
        ItemOrcamento.TIPO_RECEITA,
        ItemOrcamento.TIPO_FIXA,
        ItemOrcamento.TIPO_SEMI_FIXA,
        ItemOrcamento.TIPO_VARIAVEL,
        ItemOrcamento.TIPO_IMPOSTO,
    )
    tipos_despesa = {
        ItemOrcamento.TIPO_FIXA,
        ItemOrcamento.TIPO_SEMI_FIXA,
        ItemOrcamento.TIPO_VARIAVEL,
    }

    lancs = (
        LancamentoOrcamento.objects.filter(
            empresa=empresa,
            item__tipo__in=tipos_ordem,
            item__ativo=True,
            data_lancamento__gte=data_ini,
            data_lancamento__lte=data_fim,
        )
        .select_related('item', 'item__categoria')
        .order_by('item__tipo', 'item__ordem', 'item__nome', 'data_lancamento')
    )

    por_item = {}
    for L in lancs:
        it = L.item
        row = por_item.get(it.pk)
        if row is None:
            row = {
                'item_id': it.pk,
                'nome': it.nome,
                'tipo': it.tipo,
                'tipo_meta': ItemOrcamento.tipo_meta(it.tipo),
                'categoria': it.categoria.nome if it.categoria_id else '',
                'valores': {m['chave']: Decimal('0') for m in meses},
                'total': Decimal('0'),
            }
            por_item[it.pk] = row
        chave = f'{L.data_lancamento.year}-{L.data_lancamento.month:02d}'
        if chave in row['valores']:
            row['valores'][chave] += L.valor or Decimal('0')
            row['total'] += L.valor or Decimal('0')

    grupos = []
    receitas_mes = {m['chave']: Decimal('0') for m in meses}
    despesas_mes = {m['chave']: Decimal('0') for m in meses}
    impostos_mes = {m['chave']: Decimal('0') for m in meses}
    receitas_total = Decimal('0')
    despesas_total = Decimal('0')
    impostos_total = Decimal('0')

    for tipo in tipos_ordem:
        linhas = [r for r in por_item.values() if r['tipo'] == tipo]
        if not linhas:
            continue
        linhas.sort(key=lambda r: (r['nome'] or '').lower())
        subtotal = {m['chave']: Decimal('0') for m in meses}
        subtotal_geral = Decimal('0')
        for r in linhas:
            r['valores_lista'] = [r['valores'][m['chave']] for m in meses]
            for m in meses:
                v = r['valores'][m['chave']]
                subtotal[m['chave']] += v
                if tipo == ItemOrcamento.TIPO_RECEITA:
                    receitas_mes[m['chave']] += v
                elif tipo == ItemOrcamento.TIPO_IMPOSTO:
                    impostos_mes[m['chave']] += v
                elif tipo in tipos_despesa:
                    despesas_mes[m['chave']] += v
            subtotal_geral += r['total']
            if tipo == ItemOrcamento.TIPO_RECEITA:
                receitas_total += r['total']
            elif tipo == ItemOrcamento.TIPO_IMPOSTO:
                impostos_total += r['total']
            elif tipo in tipos_despesa:
                despesas_total += r['total']
        grupos.append({
            'tipo': tipo,
            'meta': ItemOrcamento.tipo_meta(tipo),
            'linhas': linhas,
            'subtotal_lista': [subtotal[m['chave']] for m in meses],
            'subtotal': subtotal_geral,
        })

    saidas_lista = [
        despesas_mes[m['chave']] + impostos_mes[m['chave']] for m in meses
    ]
    saidas_total = despesas_total + impostos_total
    resultado_lista = [
        receitas_mes[m['chave']] - despesas_mes[m['chave']] - impostos_mes[m['chave']]
        for m in meses
    ]
    resultado_total = receitas_total - despesas_total - impostos_total

    return {
        'meses': meses,
        'grupos': grupos,
        'receitas_lista': [receitas_mes[m['chave']] for m in meses],
        'receitas_total': receitas_total,
        'despesas_lista': [despesas_mes[m['chave']] for m in meses],
        'despesas_total': despesas_total,
        'impostos_lista': [impostos_mes[m['chave']] for m in meses],
        'impostos_total': impostos_total,
        'totais_lista': saidas_lista,
        'total_geral': saidas_total,
        'resultado_lista': resultado_lista,
        'resultado_total': resultado_total,
    }


def _serie_mensal_tipo(empresa, tipo, meses, data_ini, data_fim):
    """Soma dos lançamentos do tipo por mês (lista alinhada a `meses`)."""
    totais = {m['chave']: Decimal('0') for m in meses}
    qs = LancamentoOrcamento.objects.filter(
        empresa=empresa,
        item__tipo=tipo,
        item__ativo=True,
        data_lancamento__gte=data_ini,
        data_lancamento__lte=data_fim,
    ).values_list('data_lancamento', 'valor')
    for data_lanc, valor in qs:
        chave = f'{data_lanc.year}-{data_lanc.month:02d}'
        if chave in totais:
            totais[chave] += valor or Decimal('0')
    return [float(totais[m['chave']]) for m in meses]


def _grafico_torre(empresa, data_ini, data_fim, tabela_despesas):
    """Dados JSON para gráfico de barras (torres) por mês."""
    meses = tabela_despesas['meses'] or _meses_periodo(data_ini, data_fim)
    labels = [m['rotulo'] for m in meses]

    receitas = _serie_mensal_tipo(
        empresa, ItemOrcamento.TIPO_RECEITA, meses, data_ini, data_fim
    )
    fixas = _serie_mensal_tipo(
        empresa, ItemOrcamento.TIPO_FIXA, meses, data_ini, data_fim
    )
    semi = _serie_mensal_tipo(
        empresa, ItemOrcamento.TIPO_SEMI_FIXA, meses, data_ini, data_fim
    )
    variavel = _serie_mensal_tipo(
        empresa, ItemOrcamento.TIPO_VARIAVEL, meses, data_ini, data_fim
    )
    impostos = _serie_mensal_tipo(
        empresa, ItemOrcamento.TIPO_IMPOSTO, meses, data_ini, data_fim
    )
    despesas = [
        round(fixas[i] + semi[i] + variavel[i], 2)
        for i in range(len(meses))
    ]
    resultado = [
        round(receitas[i] - despesas[i] - impostos[i], 2)
        for i in range(len(meses))
    ]

    return {
        'labels_json': json.dumps(labels, ensure_ascii=False),
        'receitas_json': json.dumps(receitas),
        'despesas_json': json.dumps(despesas),
        'fixas_json': json.dumps(fixas),
        'semi_json': json.dumps(semi),
        'variavel_json': json.dumps(variavel),
        'impostos_json': json.dumps(impostos),
        'resultado_json': json.dumps(resultado),
        'tem_dados': any(
            v > 0 for serie in (receitas, despesas, impostos) for v in serie
        ),
    }


@login_required
def dashboard(request):
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    data_ini, data_fim = _periodo_consulta(request)
    qs = _qs_params(data_ini, data_fim)

    total_rec = _total_receitas_periodo(empresa, data_ini, data_fim)
    grupos = []
    total_despesas = Decimal('0')
    total_impostos = Decimal('0')

    for tipo, _label in ItemOrcamento.TIPO_CHOICES:
        soma, qtd_itens, qtd_lanc = _somar_tipo_periodo(empresa, tipo, data_ini, data_fim)
        meta = ItemOrcamento.tipo_meta(tipo)
        grupos.append({
            'tipo': tipo,
            'meta': meta,
            'total': soma,
            'qtd': qtd_itens,
            'qtd_lanc': qtd_lanc,
            'url': reverse('planejamento_orcamentario:listar_tipo', kwargs={'tipo': tipo}) + '?' + qs,
        })
        if tipo == ItemOrcamento.TIPO_RECEITA:
            pass
        elif tipo == ItemOrcamento.TIPO_IMPOSTO:
            total_impostos += soma
        else:
            total_despesas += soma

    resultado = total_rec - total_despesas - total_impostos
    tabela_despesas = _tabela_despesas_por_mes(empresa, data_ini, data_fim)
    grafico_torre = _grafico_torre(empresa, data_ini, data_fim, tabela_despesas)

    return render(request, 'planejamento_orcamentario/dashboard.html', {
        'title': 'Planejamento orçamentário',
        'empresa': empresa,
        'grupos': grupos,
        'total_receitas': total_rec,
        'total_despesas': total_despesas,
        'total_impostos': total_impostos,
        'resultado': resultado,
        'data_ini': data_ini,
        'data_fim': data_fim,
        'periodo_qs': qs,
        'tabela_despesas': tabela_despesas,
        'grafico_torre': grafico_torre,
        **_periodo_nav_qs(data_ini),
    })


@login_required
def exportar_excel(request):
    """Exporta o demonstrativo mensal (mesma tabela do dashboard) para .xlsx."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    data_ini, data_fim = _periodo_consulta(request)
    tabela = _tabela_despesas_por_mes(empresa, data_ini, data_fim)
    meses = tabela['meses']
    n_meses = len(meses)
    col_total = n_meses + 2  # A=descrição, B..=meses, última=Total

    wb = Workbook()
    ws = wb.active
    ws.title = 'Demonstrativo'

    header_fill = PatternFill('solid', fgColor='212529')
    header_font = Font(bold=True, color='FFFFFF')
    grupo_fill = PatternFill('solid', fgColor='DEE2E6')
    grupo_font = Font(bold=True)
    sub_fill = PatternFill('solid', fgColor='F8F9FA')
    sub_font = Font(bold=True)
    rec_fill = PatternFill('solid', fgColor='D1E7DD')
    desp_fill = PatternFill('solid', fgColor='E2E3E5')
    imp_fill = PatternFill('solid', fgColor='F8D7DA')
    res_fill = PatternFill('solid', fgColor='CFE2FF')
    money_fmt = '#,##0.00'
    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    right = Alignment(horizontal='right', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    center = Alignment(horizontal='center', vertical='center')

    empresa_nome = getattr(empresa, 'nome_fantasia', None) or getattr(empresa, 'razao', '') or str(empresa)
    ws['A1'] = 'Planejamento orçamentário — Demonstrativo mensal'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Empresa: {empresa_nome}'
    ws['A3'] = (
        f'Período: {data_ini.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}'
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_total)

    # Cabeçalho
    row = 5
    headers = ['Descrição'] + [m['rotulo'] for m in meses] + ['Total']
    for col, texto in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=texto)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center if col > 1 else left
        cell.border = thin

    def _escrever_valores(r, valores, total, fill=None, font=None, money=True):
        for col, v in enumerate(valores, start=2):
            cell = ws.cell(row=r, column=col, value=float(v or 0) if money else v)
            if money:
                cell.number_format = money_fmt
            cell.alignment = right
            cell.border = thin
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
        cell_t = ws.cell(row=r, column=col_total, value=float(total or 0) if money else total)
        if money:
            cell_t.number_format = money_fmt
        cell_t.alignment = right
        cell_t.border = thin
        if fill:
            cell_t.fill = fill
        if font:
            cell_t.font = font

    row = 6
    for grupo in tabela['grupos']:
        # Título do grupo
        cell = ws.cell(row=row, column=1, value=grupo['meta']['titulo'])
        cell.fill = grupo_fill
        cell.font = grupo_font
        cell.border = thin
        for col in range(2, col_total + 1):
            c = ws.cell(row=row, column=col, value='')
            c.fill = grupo_fill
            c.border = thin
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_total)
        row += 1

        for linha in grupo['linhas']:
            nome = linha['nome']
            if linha.get('categoria'):
                nome = f"{nome} ({linha['categoria']})"
            c0 = ws.cell(row=row, column=1, value=nome)
            c0.alignment = left
            c0.border = thin
            _escrever_valores(row, linha['valores_lista'], linha['total'])
            row += 1

        # Subtotal
        c0 = ws.cell(row=row, column=1, value=f"Subtotal — {grupo['meta']['titulo']}")
        c0.fill = sub_fill
        c0.font = sub_font
        c0.border = thin
        _escrever_valores(
            row, grupo['subtotal_lista'], grupo['subtotal'],
            fill=sub_fill, font=sub_font,
        )
        row += 1

    # Totais finais
    bold = Font(bold=True)
    for label, valores, total, fill in (
        ('Total das receitas', tabela['receitas_lista'], tabela['receitas_total'], rec_fill),
        ('Total das despesas', tabela['despesas_lista'], tabela['despesas_total'], desp_fill),
        ('Total dos impostos', tabela['impostos_lista'], tabela['impostos_total'], imp_fill),
        (
            'Resultado (Receitas − Despesas − Impostos)',
            tabela['resultado_lista'],
            tabela['resultado_total'],
            res_fill,
        ),
    ):
        c0 = ws.cell(row=row, column=1, value=label)
        c0.fill = fill
        c0.font = bold
        c0.border = thin
        _escrever_valores(row, valores, total, fill=fill, font=bold)
        row += 1

    ws.column_dimensions['A'].width = 42
    for col in range(2, col_total + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_arq = (
        f'planejamento_orcamentario_{data_ini.strftime("%Y%m%d")}_'
        f'{data_fim.strftime("%Y%m%d")}.xlsx'
    )
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{nome_arq}"'
    return resp


@login_required
def listar_tipo(request, tipo):
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')
    if tipo not in TIPOS_VALIDOS:
        messages.error(request, 'Tipo inválido.')
        return redirect('planejamento_orcamentario:dashboard')

    data_ini, data_fim = _periodo_consulta(request)
    qs = _qs_params(data_ini, data_fim)
    meta = ItemOrcamento.tipo_meta(tipo)

    lanc_periodo = LancamentoOrcamento.objects.filter(
        data_lancamento__gte=data_ini,
        data_lancamento__lte=data_fim,
    ).order_by('data_lancamento')

    itens = list(
                ItemOrcamento.objects.filter(empresa=empresa, tipo=tipo)
        .select_related('categoria')
        .prefetch_related(
            Prefetch('lancamentos', queryset=lanc_periodo, to_attr='lancamentos_periodo')
        )
        .order_by('ordem', 'nome')
    )
    total_tipo = Decimal('0')
    itens_periodo = []
    for it in itens:
        lancs = getattr(it, 'lancamentos_periodo', [])
        it.estimado = sum((l.valor for l in lancs), Decimal('0'))
        it.periodo = it.periodo_rotulo()
        it.qtd_lanc = len(lancs)
        # Só lista o que tem lançamento no período consultado
        if not lancs:
            continue
        if it.ativo:
            total_tipo += it.estimado
        itens_periodo.append(it)

    return render(request, 'planejamento_orcamentario/listar.html', {
        'title': meta['titulo'],
        'empresa': empresa,
        'tipo': tipo,
        'meta': meta,
        'itens': itens_periodo,
        'total_tipo': total_tipo,
        'total_receitas': _total_receitas_periodo(empresa, data_ini, data_fim),
        'data_ini': data_ini,
        'data_fim': data_fim,
        'periodo_qs': qs,
        **_periodo_nav_qs(data_ini),
    })


@login_required
def criar(request, tipo):
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')
    if tipo not in TIPOS_VALIDOS:
        messages.error(request, 'Tipo inválido.')
        return redirect('planejamento_orcamentario:dashboard')

    meta = ItemOrcamento.tipo_meta(tipo)
    if request.method == 'POST':
        form = ItemOrcamentoForm(request.POST, tipo=tipo, empresa=empresa)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.empresa = empresa
            obj.tipo = tipo
            if tipo == ItemOrcamento.TIPO_RECEITA:
                obj.forma_calculo = ItemOrcamento.FORMA_FIXO
            total_rec = (
                _total_receitas(empresa)
                if obj.forma_calculo in ItemOrcamento.FORMAS_SOBRE_RECEITA
                else None
            )
            try:
                n = _salvar_e_gerar(obj, total_receitas=total_rec)
            except IntegrityError as exc:
                messages.error(request, str(exc))
            else:
                messages.success(
                    request,
                    f'Item “{obj.nome}” cadastrado com {n} lançamento{"s" if n != 1 else ""}.',
                )
                return redirect(_url_listar_item(obj))
    else:
        initial = {}
        if tipo == ItemOrcamento.TIPO_IMPOSTO:
            initial['forma_calculo'] = ItemOrcamento.FORMA_PRESUMIDO_IRPJ
        form = ItemOrcamentoForm(initial=initial, tipo=tipo, empresa=empresa)

    return render(request, 'planejamento_orcamentario/form.html', {
        'title': f'Novo — {meta["titulo"]}',
        'empresa': empresa,
        'tipo': tipo,
        'meta': meta,
        'form': form,
        'acao': 'criar',
    })


@login_required
def editar(request, pk):
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    obj = get_object_or_404(ItemOrcamento, pk=pk, empresa=empresa)
    tipo = obj.tipo
    meta = ItemOrcamento.tipo_meta(tipo)

    if request.method == 'POST':
        form = ItemOrcamentoForm(request.POST, instance=obj, tipo=tipo, empresa=empresa)
        if form.is_valid():
            obj = form.save(commit=False)
            total_rec = (
                _total_receitas(empresa)
                if obj.forma_calculo in ItemOrcamento.FORMAS_SOBRE_RECEITA
                else None
            )
            try:
                n = _salvar_e_gerar(obj, total_receitas=total_rec)
            except IntegrityError as exc:
                messages.error(request, str(exc))
            else:
                messages.success(
                    request,
                    f'Item “{obj.nome}” atualizado ({n} lançamento{"s" if n != 1 else ""}).',
                )
                return redirect(_url_listar_item(obj))
    else:
        form = ItemOrcamentoForm(instance=obj, tipo=tipo, empresa=empresa)

    return render(request, 'planejamento_orcamentario/form.html', {
        'title': f'Editar — {obj.nome}',
        'empresa': empresa,
        'tipo': tipo,
        'meta': meta,
        'form': form,
        'item': obj,
        'acao': 'editar',
        'lancamentos': obj.lancamentos.order_by('data_lancamento')[:24],
    })


@login_required
@require_POST
def excluir(request, pk):
    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    obj = get_object_or_404(ItemOrcamento, pk=pk, empresa=empresa)
    tipo = obj.tipo
    nome = obj.nome
    obj.delete()
    messages.success(request, f'Item “{nome}” excluído.')
    return redirect('planejamento_orcamentario:listar_tipo', tipo=tipo)


@login_required
def visao_real_plan(request):
    """
    Visão do planejamento (só valores planejados por mês).
    Período: ano cheio, um mês, ou intervalo (ex.: 08/2026–07/2027).
    Após o RESULTADO, lista parcelas de empréstimo do período.
    """
    from fluxo_de_caixa.services.montar_fluxo_mensal import MESES_NOME
    from planejamento_orcamentario.services.montar_visao_real_plan import (
        gerar_colunas,
        montar_visao_real_plan,
        rotulos_colunas,
    )

    empresa = _empresa_sessao(request)
    if not empresa:
        messages.error(request, 'Selecione uma empresa.')
        return redirect('accounts:login')

    hoje = date.today()
    modo = (request.GET.get('modo') or 'ano').strip().lower()
    if modo not in ('ano', 'mes', 'periodo'):
        modo = 'ano'

    def _int(name, default):
        try:
            return int(request.GET.get(name) or default)
        except (TypeError, ValueError):
            return default

    ano = _int('ano', hoje.year)
    mes = max(1, min(12, _int('mes', hoje.month)))
    ano_ini = _int('ano_ini', hoje.year)
    mes_ini = max(1, min(12, _int('mes_ini', 8 if hoje.month >= 8 else 1)))
    ano_fim = _int('ano_fim', ano_ini + 1 if mes_ini > 1 else ano_ini)
    mes_fim = max(1, min(12, _int('mes_fim', 7 if mes_ini == 8 else 12)))

    # Atalho: 12 meses a partir do mês inicial
    if modo == 'periodo' and request.GET.get('atalho') == '12m':
        ano_fim, mes_fim = ano_ini, mes_ini
        for _ in range(11):
            mes_fim += 1
            if mes_fim > 12:
                mes_fim = 1
                ano_fim += 1

    colunas = gerar_colunas(
        modo,
        ano=ano,
        mes=mes,
        ano_ini=ano_ini,
        mes_ini=mes_ini,
        ano_fim=ano_fim,
        mes_fim=mes_fim,
    )
    dados, grafico_torre = montar_visao_real_plan(empresa, colunas)
    cab = rotulos_colunas(colunas)

    if modo == 'mes':
        periodo_label = f'{MESES_NOME[mes - 1]}/{ano}'
    elif modo == 'periodo':
        periodo_label = f'{cab[0]["rotulo"]} → {cab[-1]["rotulo"]} ({len(colunas)} meses)'
    else:
        periodo_label = str(ano)

    return render(request, 'planejamento_orcamentario/visao_real_plan.html', {
        'title': 'Planejamento orçamentário',
        'empresa': empresa,
        'modo': modo,
        'ano': ano,
        'mes': mes,
        'ano_ini': ano_ini,
        'mes_ini': mes_ini,
        'ano_fim': ano_fim,
        'mes_fim': mes_fim,
        'anos_disponiveis': range(2020, hoje.year + 3),
        'meses_opcoes': list(enumerate(MESES_NOME, start=1)),
        'colunas': cab,
        'periodo_label': periodo_label,
        'dados': dados,
        'grafico_torre': grafico_torre,
    })
