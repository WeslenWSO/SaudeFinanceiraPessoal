"""
Parser do relatório detalhado de recebíveis Cielo (.xlsx).
Layout esperado: planilha com cabeçalho na 1ª linha (exportação Cielo / Recebíveis).
"""
from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook


def _norm_header(value: Any) -> str:
    if value is None:
        return ''
    s = str(value).strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'\s+', ' ', s)
    return s


def _cell_str(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    if isinstance(value, date):
        return value.strftime('%d/%m/%Y')
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_date_cell(value: Any) -> date | None:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == '':
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = str(value).strip()
    s = re.sub(r'[^\d,.\-]', '', s)
    if not s or s in ('-', '.', ',', '-.', '-,'):
        return None
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _parse_int(value: Any, default: int = 1) -> int:
    if value is None or value == '':
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    s = str(value).strip()
    if s.isdigit():
        return int(s)
    m = re.search(r'\d+', s)
    return int(m.group(0)) if m else default


# Cabeçalhos normalizados → chave interna
_HEADER_ALIASES = {
    'data de pagamento': 'data_pagamento',
    'data da venda': 'data_venda',
    'forma de pagamento': 'forma_pagamento',
    'bandeira': 'bandeira',
    'valor bruto': 'valor_bruto',
    'taxa/tarifa': 'taxa_tarifa',
    'valor da taxa administrativa (mdr)': 'taxa_mdr',
    'valor liquido': 'valor_liquido',
    'codigo da autorizacao': 'codigo_autorizacao',
    'nsu/doc': 'nsu_doc',
    'numero da parcela': 'numero_parcela',
    'quantidade total de parcelas': 'total_parcelas',
    'nota fiscal': 'nota_fiscal',
    'banco': 'banco',
    'agencia': 'agencia',
    'conta': 'conta',
    'tipo de lancamento': 'tipo_lancamento',
    'numero da maquina': 'numero_maquina',
    'status de pagamento': 'status_pagamento',
}


def _map_headers(header_row: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _HEADER_ALIASES.get(_norm_header(cell))
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


def _get(row: tuple[Any, ...], mapping: dict[str, int], key: str) -> Any:
    idx = mapping.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def parse_cielo_xlsx_bytes(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """
    Lê o XLSX Cielo e devolve linhas normalizadas + avisos.
    Cada linha é um dict serializável (strings / números simples) para a sessão.
    """
    warnings: list[str] = []
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], ['Planilha vazia.']

    mapping = _map_headers(list(header or []))
    required = ['data_pagamento', 'valor_bruto']
    missing = [k for k in required if k not in mapping]
    if missing:
        wb.close()
        return [], [
            'Cabeçalho Cielo não reconhecido. '
            f'Colunas obrigatórias ausentes: {", ".join(missing)}. '
            'Exporte o relatório detalhado de recebíveis da Cielo (.xlsx).'
        ]

    if 'codigo_autorizacao' not in mapping and 'nsu_doc' not in mapping:
        warnings.append('Colunas de autorização/NSU não encontradas — campos ficarão vazios.')

    out: list[dict] = []
    for row_num, row in enumerate(rows_iter, start=2):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue

        data_pag = _parse_date_cell(_get(row, mapping, 'data_pagamento'))
        if not data_pag:
            warnings.append(f'Linha {row_num}: data de pagamento inválida/ausente — ignorada')
            continue

        valor_bruto = _parse_decimal(_get(row, mapping, 'valor_bruto'))
        if valor_bruto is None:
            warnings.append(f'Linha {row_num}: valor bruto ausente — ignorada')
            continue

        taxa = _parse_decimal(_get(row, mapping, 'taxa_tarifa'))
        if taxa is None:
            taxa = _parse_decimal(_get(row, mapping, 'taxa_mdr'))
        if taxa is None:
            taxa = Decimal('0')
        taxa_abs = abs(taxa)

        valor_liquido = _parse_decimal(_get(row, mapping, 'valor_liquido'))
        if valor_liquido is None:
            valor_liquido = valor_bruto - taxa_abs

        forma = _cell_str(_get(row, mapping, 'forma_pagamento'))
        forma_l = forma.lower()
        is_debito = 'debito' in forma_l or 'débito' in forma_l

        if is_debito:
            parcelas = 1
            total_parcelas = 1
        else:
            parcelas = _parse_int(_get(row, mapping, 'numero_parcela'), 1)
            total_parcelas = _parse_int(_get(row, mapping, 'total_parcelas'), parcelas or 1)

        data_venda = _parse_date_cell(_get(row, mapping, 'data_venda'))
        data_venda_str = data_venda.strftime('%d/%m/%Y') if data_venda else _cell_str(
            _get(row, mapping, 'data_venda')
        )

        autorizacao = _cell_str(_get(row, mapping, 'codigo_autorizacao'))
        nsu = _cell_str(_get(row, mapping, 'nsu_doc'))

        banco = _cell_str(_get(row, mapping, 'banco'))
        agencia = _cell_str(_get(row, mapping, 'agencia'))
        conta = _cell_str(_get(row, mapping, 'conta'))
        conta_bancaria_parts = [p for p in (banco, agencia, conta) if p]
        conta_bancaria = ' / '.join(conta_bancaria_parts) if conta_bancaria_parts else ''

        out.append({
            'linha': row_num - 1,
            'data_pagamento': data_pag.strftime('%d/%m/%Y'),
            'forma_pagamento': forma,
            'bandeira': _cell_str(_get(row, mapping, 'bandeira')),
            'valor_bruto': f'{valor_bruto:.2f}',
            'taxa_maquinha': f'{taxa_abs:.2f}',
            'valor_liquido': f'{valor_liquido:.2f}',
            'maquinha': 'CIELO',
            'numero_autorizacao': autorizacao,
            'data_venda': data_venda_str,
            'nsu_doc': nsu or autorizacao,
            'parcelas': str(parcelas),
            'total_parcelas': str(total_parcelas),
            'parcela_texto': f'{parcelas} / {total_parcelas}',
            'conciliado': 'Não',
            'nota_fiscal': _cell_str(_get(row, mapping, 'nota_fiscal')),
            'razao': '',
            'conta_bancaria': conta_bancaria,
            'tipo_lancamento': _cell_str(_get(row, mapping, 'tipo_lancamento')),
            'numero_maquina': _cell_str(_get(row, mapping, 'numero_maquina')),
            'status_pagamento': _cell_str(_get(row, mapping, 'status_pagamento')),
        })

    wb.close()
    if not out and not warnings:
        warnings.append('Nenhuma linha de recebível encontrada no arquivo.')
    return out, warnings
